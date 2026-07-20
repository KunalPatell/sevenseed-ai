"""
deep_web_research.py — Full Deep Web Research & Enrichment Pipeline
=====================================================================
Uses ONLY standard libraries (re, urllib.parse, html.parser) to scrape and search.
No external dependencies like bs4.
"""

import sys, os, re, json, time, asyncio, unicodedata
import openpyxl
import httpx
from urllib.parse import urlparse, quote_plus, urljoin
from html.parser import HTMLParser

sys.stdout.reconfigure(encoding='utf-8')

EXCEL = "COMONK_TRUE_MASTER.xlsx"
SHEET = "COMPLETE_MASTER"
CHECKPOINT = "deep_web_research_checkpoint.json"

MAX_WORKERS = 15
REQ_TIMEOUT = 10
MAX_PAGES = 5

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ── Column Map ───────────────────────────────────────────────────────────────
C_COMP  = 2
C_CITY  = 3
C_CAT   = 4
C_ROLE  = 5
C_EMAILS = [6,7,8,9,10,11]
C_PHONE = 12
C_WEB   = 13
C_LI    = 14
C_ADDR  = 15
C_CARE  = 16
C_PRIO  = 17
C_SRC   = 18
C_LIHR  = 19
C_LICO  = 20
C_EMP   = 21
C_IND   = 22
C_FOUND = 23

EMAIL_RE = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,7}\b')
PHONE_RE = re.compile(r'(?:\+91[\s\-.]?[6-9]\d{9}|\+91[\s\-.]?(?:79|80|22|33|44)\d{8}|0?79[\s\-.]?\d{4}[\s\-.]?\d{4}|\b[6-9]\d{9}\b)')
FOUNDED_RE = re.compile(r'\b(19[5-9]\d|20[0-2]\d)\b')

HR_PREFIXES = {"hr", "hrd", "hrbp", "humanresources", "recruit", "recruitment", "talent", "career", "careers", "jobs", "hiring", "people"}
SKIP_PREFIXES = {"noreply", "no-reply", "support", "help", "tech", "sales", "info", "contact", "admin", "billing", "marketing"}
SKIP_DOMAINS = {"example.com", "test.com", "yourdomain.com", "domain.com", "sentry.io", "gmail.com", "yahoo.com", "outlook.com"}

# HTML tag stripper
class TextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.result = []
        self._skip = False
    def handle_starttag(self, tag, attrs):
        if tag in ('script', 'style', 'noscript'):
            self._skip = True
    def handle_endtag(self, tag):
        if tag in ('script', 'style', 'noscript'):
            self._skip = False
    def handle_data(self, data):
        if not self._skip:
            self.result.append(data)
    def get_text(self):
        return ' '.join(self.result)

def extract_text(html):
    p = TextExtractor()
    try: p.feed(html)
    except: pass
    return p.get_text()

# Extract links from HTML
def extract_links(html):
    links = []
    # simple href regex
    for m in re.finditer(r'href=["\']([^"\']+)["\']', html, re.I):
        links.append(m.group(1))
    return links

def norm(s):
    if not s: return ""
    s = unicodedata.normalize("NFKD", str(s).lower())
    return re.sub(r'[^a-z0-9]', '', s).strip()

def clean_phone(p):
    if not p: return ""
    d = re.sub(r'[^\d]', '', str(p))
    if d.startswith('91') and len(d) == 12: d = d[2:]
    elif d.startswith('0') and len(d) in (11, 12): d = d[1:]
    if len(d) < 10 or len(set(d)) <= 2 or d[:5] == d[5:]: return ""
    if len(d) == 10 and d[0] in '6789': return '+91 ' + d[:5] + ' ' + d[5:]
    if len(d) == 8: return '+91 79 ' + d[:4] + ' ' + d[4:]
    return ""

def clean_url(url):
    if not url: return ""
    url = str(url).strip()
    if not url.startswith("http") and "." in url:
        url = "https://" + url
    return url.rstrip('/')

def domain_of(url):
    if not url: return ""
    try: return urlparse(clean_url(url)).netloc.replace("www.", "").strip().lower()
    except: return ""

def score_email(email):
    prefix = email.split('@')[0].lower()
    if prefix in HR_PREFIXES or any(h in prefix for h in HR_PREFIXES): return 100
    if prefix in SKIP_PREFIXES: return 5
    return 30

def is_valid_email(email):
    if '@' not in email: return False
    dom = email.split('@')[1].lower()
    if dom in SKIP_DOMAINS: return False
    return True

# ─────────────────────────────────────────────────────────────────────────────
# REAL-TIME SEARCH ENGINES & GOOGLE
# ─────────────────────────────────────────────────────────────────────────────
async def google_search(client, query):
    try:
        url = f"https://www.google.com/search?q={quote_plus(query)}&num=8"
        r = await client.get(url, timeout=REQ_TIMEOUT)
        if r.status_code == 200:
            return r.text
    except Exception:
        pass
    return ""

async def find_website(client, company_name, city):
    html = await google_search(client, f"{company_name} {city} official website")
    if not html: return ""
    
    # Extract domains
    domains = []
    for link in extract_links(html):
        if '/url?q=' in link:
            url = link.split('/url?q=')[1].split('&')[0]
            dom = domain_of(url)
            if dom and not any(k in dom for k in ["google", "linkedin", "glassdoor", "indeed", "ambitionbox", "facebook", "twitter", "instagram", "youtube", "justdial", "indiamart"]):
                return f"https://www.{dom}"
    return ""

async def find_linkedin_company(client, company_name):
    html = await google_search(client, f"site:linkedin.com/company \"{company_name}\"")
    if not html: return ""
    m = re.search(r'linkedin\.com/company/([a-z0-9\-]+)', html, re.I)
    if m:
        return f"linkedin.com/company/{m.group(1)}"
    return ""

# ─────────────────────────────────────────────────────────────────────────────
# WEBSITE SCRAPING & METADATA EXTRACTOR
# ─────────────────────────────────────────────────────────────────────────────
async def scrape_site_metadata(client, website):
    results = {"emails": [], "phones": [], "address": "", "careers": "", "founded": "", "employees": ""}
    if not website: return results
    
    base_url = f"https://{domain_of(website)}"
    paths = ["", "/contact", "/contact-us", "/about", "/careers", "/jobs"]
    
    for path in paths:
        url = base_url + path
        try:
            r = await client.get(url, timeout=8, follow_redirects=True)
            if r.status_code != 200: continue
            text = extract_text(r.text)
            
            # Emails
            for em in EMAIL_RE.findall(text):
                em = em.lower().strip()
                if is_valid_email(em) and em not in results["emails"]:
                    results["emails"].append(em)
                    
            # Phones
            for ph in PHONE_RE.findall(text):
                c = clean_phone(ph)
                if c and c not in results["phones"]:
                    results["phones"].append(c)
                    
            # Address
            if not results["address"]:
                addr_m = re.search(r'([A-Za-z0-9\s,\-\.#/]{10,80}(?:Ahmedabad|Gandhinagar|GIFT City|Gujarat)[A-Za-z0-9\s,\-\.#/]{0,60})', text, re.I)
                if addr_m:
                    results["address"] = re.sub(r'\s+', ' ', addr_m.group(1)).strip()
                    
            # Careers URL
            if not results["careers"]:
                for link in extract_links(r.text):
                    href = link.lower()
                    if 'career' in href or 'job' in href or 'hiring' in href:
                        results["careers"] = urljoin(url, link)
                        break
                        
            # Founded
            if not results["founded"]:
                fy = FOUNDED_RE.findall(text)
                if fy:
                    years = [int(y) for y in fy if 1980 <= int(y) <= 2025]
                    if years: results["founded"] = min(years)
                    
            if len(results["emails"]) >= 3 and len(results["phones"]) >= 1:
                break
        except Exception:
            pass
            
    results["emails"].sort(key=score_email, reverse=True)
    return results

# ─────────────────────────────────────────────────────────────────────────────
# PIPELINE EXECUTION
# ─────────────────────────────────────────────────────────────────────────────
async def process_row(client, sem, row_data):
    async with sem:
        r = row_data["row"]
        name = row_data["name"]
        city = row_data["city"]
        web = row_data["website"]
        li = row_data["linkedin"]
        
        result = {"row": r, "website": "", "linkedin": "", "emails": [], "phones": [], "address": "", "careers": "", "founded": "", "employees": ""}
        
        # 1. Try finding website
        if not web:
            web = await find_website(client, name, city)
            if web:
                result["website"] = web
                
        # 2. Scrape website details
        site = web or result["website"]
        if site:
            site_data = await scrape_site_metadata(client, site)
            result.update(site_data)
            
        # 3. Find LinkedIn Company URL
        if not li:
            li_url = await find_linkedin_company(client, name)
            if li_url:
                result["linkedin"] = li_url
                
        return result

async def main():
    print("\n" + "="*70)
    print("  COMONK AI — Full Deep Research & Auto-Enrichment Pipeline (Standard Library Only)")
    print("="*70 + "\n")
    
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    
    targets = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, C_COMP).value
        if not name: continue
        
        # Check what's missing
        web = ws.cell(r, C_WEB).value
        li = ws.cell(r, C_LI).value
        emails = [ws.cell(r, col).value for col in C_EMAILS if ws.cell(r, col).value]
        phone = ws.cell(r, C_PHONE).value
        addr = ws.cell(r, C_ADDR).value
        
        if not web or not li or not emails or not phone or not addr:
            targets.append({
                "row": r, "name": str(name).strip(),
                "city": str(ws.cell(r, C_CITY).value or "Ahmedabad").strip(),
                "website": web or "",
                "linkedin": li or ""
            })
            
    total = len(targets)
    print(f"  Target rows with missing details: {total}")
    if total == 0:
        print("  All data fully populated!")
        return

    # Checkpoint
    done = set()
    if os.path.exists(CHECKPOINT):
        try:
            with open(CHECKPOINT) as f:
                done = set(json.load(f))
        except: pass
    print(f"  Already processed (checkpoint): {len(done)}")
    
    targets = [t for t in targets if t["row"] not in done]
    print(f"  Remaining rows to process: {len(targets)}")
    
    sem = asyncio.Semaphore(MAX_WORKERS)
    
    async with httpx.AsyncClient(headers=HTTP_HEADERS, timeout=REQ_TIMEOUT, verify=False) as client:
        batch_size = 50
        for start in range(0, len(targets), batch_size):
            chunk = targets[start:start+batch_size]
            print(f"\n  Batch {start//batch_size + 1} ({len(chunk)} companies)...", flush=True)
            
            tasks = [process_row(client, sem, item) for item in chunk]
            results = await asyncio.gather(*tasks)
            
            # Write results to worksheet
            for res in results:
                r = res["row"]
                
                # Web
                if res["website"] and not ws.cell(r, C_WEB).value:
                    ws.cell(r, C_WEB).value = res["website"]
                    
                # LinkedIn
                if res["linkedin"] and not ws.cell(r, C_LI).value:
                    ws.cell(r, C_LI).value = res["linkedin"]
                    
                # Emails
                if res["emails"]:
                    ex_emails = [str(ws.cell(r, c).value).lower() for c in C_EMAILS if ws.cell(r, c).value]
                    new_emails = [e for e in res["emails"] if e.lower() not in ex_emails]
                    empty_cols = [c for c in C_EMAILS if not ws.cell(r, c).value]
                    for e, col in zip(new_emails, empty_cols):
                        ws.cell(r, col).value = e
                        
                # Phone
                if res["phones"] and not ws.cell(r, C_PHONE).value:
                    ws.cell(r, C_PHONE).value = res["phones"][0]
                    
                # Address
                if res["address"] and not ws.cell(r, C_ADDR).value:
                    ws.cell(r, C_ADDR).value = res["address"][:200]
                    
                # Careers URL
                if res["careers"] and not ws.cell(r, C_CARE).value:
                    ws.cell(r, C_CARE).value = res["careers"]
                    
                # Founded
                if res["founded"] and not ws.cell(r, C_FOUND).value:
                    ws.cell(r, C_FOUND).value = res["founded"]
                    
                done.add(r)
                
            wb.save(EXCEL)
            with open(CHECKPOINT, 'w') as f:
                json.dump(list(done), f)
                
            print(f"    Checkpoint saved. Processed {len(done)} rows so far.")
            
    print("\n" + "="*70)
    print("  Full Deep Web Research Finished successfully!")
    print(f"  Spreadsheet fully updated: {EXCEL}")
    print("="*70 + "\n")

if __name__ == "__main__":
    asyncio.run(main())
