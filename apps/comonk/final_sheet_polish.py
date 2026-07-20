# final_sheet_polish.py
# Unified master polisher and sorter for Ahmedabad_IT_AIML_FINAL_MASTER.xlsx
# Priorities order:
# 1. 1 - Apply Now (MNC Gujarat) ⭐⭐  (MNC + Gujarat location)
# 2. 1 - Apply Now (MNC) ⭐             (MNC outside Gujarat)
# 3. 2 - Gujarat Priority 🏆            (Gujarat non-MNC company)
# 4. 3 - General IT / Other

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

EXCEL_FILE = "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
SHEET_NAME = "All Companies"
FAKE_PHONE = "+91 73260 59369"
GUJARAT_CITIES = {"ahmedabad", "gandhinagar", "surat", "vadodara", "rajkot",
                  "anand", "bharuch", "mehsana", "morbi", "gift city",
                  "gift", "sanand", "naroda", "vatva", "koba", "infocity"}

def normalize(s):
    return str(s or "").strip().lower()

def is_gujarat(city_raw):
    city = normalize(city_raw)
    return any(g in city for g in GUJARAT_CITIES)

def safe_save(wb, filename):
    import time
    while True:
        try:
            wb.save(filename)
            return
        except PermissionError:
            print(f"  [⚠️ WARNING] Permission Denied: Cannot save '{filename}'. Please close it in Microsoft Excel! Retrying in 5 seconds...")
            time.sleep(5)
        except Exception as e:
            print(f"  [❌ ERROR] Failed to save: {e}")
            raise e

def main():
    print(f"\n========================================================")
    print(f"  MASTER SHEET POLISH & PRIORITY SORTING")
    print(f"========================================================\n")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    # Read headers from Row 3
    headers = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
    hmap = {str(h).strip().lower(): i+1 for i, h in enumerate(headers) if h}
    total_cols = len(headers)

    COL_NO   = hmap.get("no", 1)
    COL_CO   = hmap.get("company name", 2)
    COL_CAT  = hmap.get("category", 3)
    COL_ROLE = hmap.get("roles / skills", 4)
    COL_PH   = hmap.get("phone", 10)
    COL_WEB  = hmap.get("website", 11)
    COL_LI   = hmap.get("linkedin", 12)
    COL_ADDR = hmap.get("address", 13)
    COL_CITY = hmap.get("city", 14)
    COL_PRI  = hmap.get("priority", 15)
    
    email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]
    email_cols = [ec for ec in email_cols if ec and ec <= total_cols]

    rows_data = []

    # Read all data rows (starting at Row 4)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        row_vals = [cell.value for cell in row]
        if len(row_vals) < COL_CO or not row_vals[COL_CO-1]:
            continue # skip empty rows
        
        # Pad row to total columns to avoid index errors
        row_vals = (row_vals + [None] * total_cols)[:total_cols]
        
        # 1. Clean fake phone number
        phone_val = str(row_vals[COL_PH-1] or "").strip()
        if phone_val == FAKE_PHONE:
            row_vals[COL_PH-1] = None

        # 2. Count active emails
        active_emails = []
        for ec in email_cols:
            email_val = str(row_vals[ec-1] or "").strip()
            if email_val and email_val.lower() != "none" and "@" in email_val:
                active_emails.append(email_val)
            else:
                row_vals[ec-1] = None

        # 3. Clean string representations of "None"
        for idx in range(len(row_vals)):
            if str(row_vals[idx]).strip() == "None":
                row_vals[idx] = None

        # 4. Set Intelligent Priority
        cat = normalize(row_vals[COL_CAT-1])
        city = normalize(row_vals[COL_CITY-1])

        is_mnc = "mnc" in cat
        is_gj = is_gujarat(city)

        if is_mnc and is_gj:
            row_vals[COL_PRI-1] = "1 - Apply Now (MNC Gujarat) ⭐⭐"
        elif is_mnc:
            row_vals[COL_PRI-1] = "1 - Apply Now (MNC) ⭐"
        elif is_gj:
            row_vals[COL_PRI-1] = "2 - Gujarat Priority 🏆"
        else:
            row_vals[COL_PRI-1] = "3 - General IT / Other"

        rows_data.append(row_vals)

    # 5. Sort Rows
    def get_sort_key(r):
        pri = str(r[COL_PRI-1] or "").strip()
        co_name = str(r[COL_CO-1] or "").strip()
        if "⭐⭐" in pri:
            return (0, co_name)
        elif "MNC" in pri:
            return (1, co_name)
        elif "Gujarat" in pri:
            return (2, co_name)
        else:
            return (3, co_name)

    rows_data.sort(key=get_sort_key)

    # 6. Clear sheet data starting at Row 4
    ws.delete_rows(4, ws.max_row+1)

    # 7. Write back sorted & cleaned data
    for idx, r in enumerate(rows_data, start=1):
        r[COL_NO-1] = idx
        ws.append(r)

    # Styling and coloring
    thin = Side(border_style="thin", color="D3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Format header row 3
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fills
    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Top priority Gujarat MNC
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Other MNC
    blue_fill   = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Gujarat non-MNC

    # Style data rows (starting at Row 4)
    for r_idx in range(4, ws.max_row + 1):
        pri_val = str(ws.cell(r_idx, COL_PRI).value or "")
        row_fill = None
        if "⭐⭐" in pri_val:
            row_fill = green_fill
        elif "MNC" in pri_val:
            row_fill = yellow_fill
        elif "Gujarat" in pri_val:
            row_fill = blue_fill
        
        for c_idx in range(1, total_cols + 1):
            cell = ws.cell(r_idx, c_idx)
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10, bold=(c_idx == COL_CO))
            
            # Align center for No, Category, Priority, Phone, Website, City
            if c_idx in [COL_NO, COL_CAT, COL_PRI, COL_PH, COL_CITY]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    safe_save(wb, EXCEL_FILE)

    # Quick summary of priorities
    summary = {}
    for r in rows_data:
        pri = r[COL_PRI-1]
        summary[pri] = summary.get(pri, 0) + 1

    print("========================================================")
    print("  POLISH & SORT COMPLETE!")
    print("  Summary of rows by priority:")
    for pri, count in sorted(summary.items()):
        print(f"    - {pri:<40}: {count} companies")
    print(f"  Total records in database: {len(rows_data)}")
    print("========================================================\n")

if __name__ == "__main__":
    main()
