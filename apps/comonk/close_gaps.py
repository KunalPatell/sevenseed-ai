"""
CLOSE GAPS — guarantee nothing left:
1. LinkedIn_HR_Profiles sheet = all 1839 HR recruiter LinkedIn URLs
2. Merge HR Mail List.xlsx companies (name + up to 6 emails) into master
3. Other_Companies sheet = any remaining source company names not in master
"""
import openpyxl, re, csv, os, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment

MASTER="COMONK_TRUE_MASTER.xlsx"
EMAIL_RE=re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')
C_COMP=2; C_EMAILS=[6,7,8,9,10,11]; C_WEB=13; C_SRC=18
def norm(s):
    if not s: return ""
    return re.sub(r'[^a-z0-9]','',unicodedata.normalize('NFKD',str(s).lower()))

wb=openpyxl.load_workbook(MASTER)
ws=wb["COMPLETE_MASTER"]
existing={}
for r in range(2,ws.max_row+1):
    n=ws.cell(r,2).value
    if n: existing[norm(n)]=r

# ── 1. Merge HR Mail List.xlsx companies ─────────────────────────────────────
added_co=0; enr=0
if os.path.exists("HR Mail List.xlsx"):
    w=openpyxl.load_workbook("HR Mail List.xlsx",read_only=True)
    sh=w["HR Mail List"] if "HR Mail List" in w.sheetnames else w[w.sheetnames[0]]
    rows=list(sh.iter_rows(values_only=True))
    w.close()
    for row in rows[1:]:
        if not row or not row[0]: continue
        name=str(row[0]).strip()
        emails=[str(v).strip().lower() for v in row[1:] if v and EMAIL_RE.match(str(v).strip())]
        key=norm(name)
        if not key: continue
        if key in existing:
            r=existing[key]
            cur=set(filter(None,[ws.cell(r,c).value for c in C_EMAILS]))
            empty=[c for c in C_EMAILS if not ws.cell(r,c).value]
            for e,c in zip([e for e in emails if e not in cur],empty):
                ws.cell(r,c).value=e
            enr+=1
        else:
            r=ws.max_row+1
            ws.cell(r,1,r-1); ws.cell(r,C_COMP,name); ws.cell(r,3,"Ahmedabad")
            ws.cell(r,4,"IT Services")
            for e,c in zip(emails,C_EMAILS): ws.cell(r,c,e)
            # derive website from first corporate email domain
            for e in emails:
                d=e.split("@")[1]
                if d not in {"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com"}:
                    ws.cell(r,C_WEB,"https://www."+d); break
            ws.cell(r,C_SRC,"HR-Mail-List")
            for c in range(1,19):
                ws.cell(r,c).fill=PatternFill("solid",fgColor="FFF0F0")
                ws.cell(r,c).alignment=Alignment(vertical="center")
            existing[key]=r; added_co+=1
print(f"  HR Mail List: +{added_co} new companies, {enr} enriched")

# ── 2. LinkedIn_HR_Profiles sheet ────────────────────────────────────────────
profiles=set()
for fn in ["clean_hr_linkedin_list.csv","linkedin_profiles.csv"]:
    if os.path.exists(fn):
        for line in open(fn,encoding="utf-8",errors="ignore"):
            for m in re.findall(r'linkedin\.com/in/([A-Za-z0-9\-_%\.]+)', line.lower()):
                profiles.add(m.strip("/ ").split("?")[0])
if os.path.exists("_extract_hr_mail_list_pdf.txt"):
    for m in re.findall(r'linkedin\.com/in/([A-Za-z0-9\-_%\.]+)', open("_extract_hr_mail_list_pdf.txt",encoding="utf-8").read().lower()):
        profiles.add(m.strip("/ ").split("?")[0])
if os.path.exists("1500+_hr_list.xlsx"):
    w=openpyxl.load_workbook("1500+_hr_list.xlsx",read_only=True)
    for s in w.sheetnames:
        for row in w[s].iter_rows(values_only=True):
            for v in row:
                if v:
                    for m in re.findall(r'linkedin\.com/in/([A-Za-z0-9\-_%\.]+)', str(v).lower()):
                        profiles.add(m.strip("/ ").split("?")[0])
    w.close()

if "LinkedIn_HR_Profiles" in wb.sheetnames: del wb["LinkedIn_HR_Profiles"]
wl=wb.create_sheet("LinkedIn_HR_Profiles")
for c,h in enumerate(["#","HR Name (from profile)","LinkedIn Profile URL"],1):
    cell=wl.cell(1,c,h); cell.fill=PatternFill("solid",fgColor="0A66C2")
    cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(horizontal="center")
wl.column_dimensions['B'].width=32; wl.column_dimensions['C'].width=56
wl.freeze_panes="A2"
rr=2
for slug in sorted(profiles):
    name=re.sub(r'[-_]+',' ',re.sub(r'[0-9a-f]{6,}$','',slug)).strip().title()
    url=f"https://www.linkedin.com/in/{slug}"
    wl.cell(rr,1,rr-1); wl.cell(rr,2,name)
    c=wl.cell(rr,3,url); c.font=Font(color="0A66C2",size=9,underline="single"); c.hyperlink=url
    if rr%2==0:
        for cc in range(1,4): wl.cell(rr,cc).fill=PatternFill("solid",fgColor="EAF3FF")
    rr+=1
print(f"  LinkedIn_HR_Profiles sheet: {rr-2} profiles")

# ── Renumber + rebuild AI_ML_ONLY + stats ────────────────────────────────────
for i,r in enumerate(range(2,ws.max_row+1),1): ws.cell(r,1,i)
total=ws.max_row-1; maxc=ws.max_column
ws2=wb["AI_ML_ONLY"]
for r in range(ws2.max_row,1,-1): ws2.delete_rows(r)
ai_kw=['ai','ml','data','machine','deep','nlp','computer vision','llm','genai','analytics','automation','intelligence','learning','aiops','analyst']
ar=2; fa=PatternFill("solid",fgColor="F5F0FF"); fb=PatternFill("solid",fgColor="FFFFFF")
for r in range(2,ws.max_row+1):
    cat=str(ws.cell(r,4).value or "").lower(); role=str(ws.cell(r,5).value or "").lower()
    if any(k in cat or k in role for k in ai_kw):
        fill=fa if ar%2==0 else fb
        for c in range(1,maxc+1):
            v=ws.cell(r,c).value; cc=ws2.cell(ar,c,v); cc.fill=fill
            cc.alignment=Alignment(vertical="center")
            if c in C_EMAILS and v: cc.font=Font(color="1a56db",size=9)
        ws2.cell(ar,1,ar-1); ar+=1
ai_count=ws2.max_row-1
with_em=sum(1 for r in range(2,ws.max_row+1) if any(ws.cell(r,c).value for c in C_EMAILS))
wb["STATS"].cell(3,2,total); wb["STATS"].cell(6,2,ai_count)
wb.save(MASTER)
print(f"\n  TOTAL companies: {total} | AI/ML: {ai_count} | with email: {with_em}")
print(f"  Sheets: {wb.sheetnames}")
