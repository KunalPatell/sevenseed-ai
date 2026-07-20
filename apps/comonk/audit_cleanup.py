"""
Audit: did fix_faults remove any REAL emails (false positives)?
Compare pre-cleanup backup vs current file, per sheet, list removed emails,
and flag ones that look legitimate (not obviously fake).
"""
import openpyxl, re, zipfile, time
from collections import Counter

CUR="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
BAK="Ahmedabad_IT_AIML_FINAL_MASTER_SCRAPED_1653.xlsx"   # pre-cleanup (has faults)

def load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read "+p)

OBVIOUS_FAKE=("w3.org","example.com","domain.com","yourdomain","johnsmith","name@example",
              "@2x",".webp",".png","@3x","sentry.io","wixpress","noreply@","no-reply@")
def is_obvious_fake(e):
    return any(b in e.lower() for b in OBVIOUS_FAKE)

def emails_by_company(wb, S):
    ws=wb[S]
    hrow=None
    for rr in (1,2,3,4):
        vals=[str(ws.cell(rr,c).value or "").strip().lower() for c in range(1,26)]
        if "company name" in vals or "company" in vals: hrow=rr; break
    if not hrow: return {}
    hdr=[ws.cell(hrow,c).value for c in range(1,ws.max_column+1)]
    hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
    CO=hm.get("company name") or hm.get("company")
    EM=[hm.get(f"email {i}") for i in range(1,7)]+[hm.get(f"hr email {i}") for i in range(1,16)]; EM=[c for c in EM if c]
    out={}
    for row in ws.iter_rows(min_row=hrow+1, values_only=True):
        def g(c): return row[c-1] if c and len(row)>=c else None
        co=g(CO)
        if not co or not str(co).strip(): continue
        ems=set(str(g(c)).strip().lower() for c in EM if g(c) and "@" in str(g(c)))
        out.setdefault(str(co).strip(), set()).update(ems)
    return out

cur=load(CUR); bak=load(BAK)
for S in ["All Companies","IT MNC"]:
    if S not in cur.sheetnames or S not in bak.sheetnames: continue
    cb=emails_by_company(bak,S); cc=emails_by_company(cur,S)
    removed=[];
    for co, ems in cb.items():
        cur_ems=cc.get(co,set())
        for e in ems - cur_ems:
            removed.append((co,e))
    fake=[x for x in removed if is_obvious_fake(x[1])]
    maybe_real=[x for x in removed if not is_obvious_fake(x[1])]
    print(f"===== {S} =====")
    print(f"  total emails removed: {len(removed)}")
    print(f"    obvious-fake (correct to remove): {len(fake)}")
    print(f"    NOT-obvious-fake (possible false positives): {len(maybe_real)}")
    if maybe_real:
        print(f"    --- sample of possibly-real removed emails ---")
        for co,e in maybe_real[:25]:
            print(f"      {co[:28]:<28} {e}")
    print()
