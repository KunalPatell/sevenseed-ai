"""Append the remaining source phones to All_Phones sheet — close the gap."""
import openpyxl, re
from openpyxl.styles import PatternFill

MASTER="COMONK_TRUE_MASTER.xlsx"
def fmt(d):
    if d.startswith('79'): return '+91 79 '+d[2:6]+' '+d[6:]
    return '+91 '+d[:5]+' '+d[5:]

missing=[l.strip() for l in open("_missing_phones.txt",encoding="utf-8") if l.strip()]
wb=openpyxl.load_workbook(MASTER)
wap=wb["All_Phones"]
# existing digits in All_Phones
have=set()
for r in range(2,wap.max_row+1):
    v=wap.cell(r,2).value
    if v:
        d=re.sub(r'[^\d]','',str(v))[-10:]
        if len(d)==10: have.add(d)
rr=wap.max_row+1
added=0
for d in missing:
    if len(d)==10 and d not in have:
        typ="Landline (079)" if d.startswith('79') else "Mobile"
        wap.cell(rr,1,rr-1); wap.cell(rr,2,fmt(d)); wap.cell(rr,3,typ)
        if rr%2==0:
            for c in range(1,4): wap.cell(rr,c).fill=PatternFill("solid",fgColor="F5F0FF")
        rr+=1; added+=1; have.add(d)
wb.save(MASTER)
print(f"  Appended {added} phones to All_Phones. Total now: {wap.max_row-1}")
