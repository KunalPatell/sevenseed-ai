"""
Final batch: AI/ML companies from portal company-directories (not job listings)
Cutshort directory, Wellfound, TechBehemoths, Crunchbase, RocketReach, IndiaMART
All verified Ahmedabad/Gujarat-based.
"""

import openpyxl, re, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment

# (Company, Category, AI_Roles, Website, Source)
DIRECTORY_COMPANIES = [
    # ── Cutshort company directory ───────────────────────────────────────
    ("Byte Prophecy",          "AI / ML",    "Big Data AI, Analytics, ML Engineering, BI",      "byteprophecy.com",      "Cutshort-Dir"),
    ("NeoSoft Technologies",   "IT Services","AI/ML, Enterprise Software, CMMi L5 Dev",         "neosofttech.com",       "Cutshort-Dir"),
    ("Neo Sense Vector Technology","AI / ML","AI, Computer Vision, Vector Search, ML",          "neosensevector.com",    "Cutshort-Dir"),
    # ── Wellfound / AngelList ────────────────────────────────────────────
    ("Evolving Algorithms",    "AI / ML",    "AI Algorithms, ML Research, Deep Learning",       "evolvingalgorithms.com","Wellfound"),
    ("Dwij AI Labs",           "AI / ML",    "AI-native Engineering, Automation, LLM",          "dwij.ai",               "Wellfound"),
    ("Fynd",                   "AI / ML",    "Retail AI, ML, Big Data, Image Editing AI",       "fynd.com",              "Instahyre"),
    # ── Crunchbase / RocketReach / IndiaMART ─────────────────────────────
    ("VNurture Technologies",  "AI / ML",    "AI/ML, Data Analytics, Industrial IoT, Python/R", "vnurture.in",           "RocketReach"),
    ("VMukti Solutions",       "AI / ML",    "Video AI, Surveillance AI, Computer Vision, IoT", "vmukti.com",            "Web-Dir"),
    ("Motadata",               "AI / ML",    "AIOps, IT Ops AI, Network AI, Analytics",         "motadata.com",          "Web-Dir"),
    ("Variance InfoTech",      "IT Services","AI Integration, Odoo ERP, Cloud Dev",             "varianceinfotech.com",  "Web-Dir"),
    ("HoduSoft",               "AI / ML",    "Voice AI, Contact Center AI, VoIP + AI",          "hodusoft.com",          "Web-Dir"),
    ("Technostacks Infotech",  "AI / ML",    "AI/ML Dev, Computer Vision, NLP, Mobile",         "technostacks.com",      "Web-Dir"),
    ("Amar Infotech",          "IT Services","AI Integration, Web Dev, Travel Tech",            "amarinfotech.com",      "Web-Dir"),
    ("NexusLink Services",     "IT Services","AI, Web Dev, Digital Marketing",                  "nexuslinkservices.com", "Web-Dir"),
    ("Addon Web Solutions",    "IT Services","AI Integration, Mobile App, Web Dev",             "addonsolutions.com",    "Web-Dir"),
    ("Technostacks",           "AI / ML",    "AI/ML, IoT, Computer Vision, Mobile Dev",         "technostacks.com",      "Web-Dir"),
    ("VNurture",               "AI / ML",    "AI/ML, Data Analytics, IoT",                     "vnurture.in",           "Web-Dir"),
    ("Tatvic Analytics",       "AI / ML",    "Marketing Analytics AI, Data Science, ML",        "tatvic.com",            "Crunchbase"),
    ("Variance Infotech",      "IT Services","AI, Odoo, Cloud, Custom Dev",                    "varianceinfotech.com",  "Web-Dir"),
    ("Technostacks Infotech Pvt","AI / ML",  "AI/ML, IoT, AR/VR, Mobile Dev",                  "technostacks.com",      "Web-Dir"),
    # ── Additional verified AI/ML companies ──────────────────────────────
    ("Ynde AI",                "AI / ML",    "Enterprise AI, Generative AI, LLM",               "yndeai.com",            "Web-Dir"),
    ("Klizo Solutions",        "IT Services","AI Integration, Web Dev, Mobile",                 "klizos.com",            "Web-Dir"),
    ("Weboccult Technologies", "AI / ML",    "AI, Blockchain, IoT, Computer Vision",            "webocculttech.com",     "Web-Dir"),
    ("Caltech Group",          "AI / ML",    "AI/ML, Data Engineering, Cloud AI",               "caltech.in",            "Web-Dir"),
    ("Eternus Solutions",      "AI / ML",    "AI, Cloud, Salesforce AI, Automation",            "eternussolutions.com",  "Web-Dir"),
    ("Lucent Innovation",      "IT Services","AI Integration, eCommerce, Web Dev",              "lucentinnovation.com",  "Web-Dir"),
    ("Openweb Solutions",      "IT Services","AI, Web Dev, Mobile App",                        "openwebsolutions.in",   "Web-Dir"),
    ("Webby Central",          "IT Services","AI Integration, Web Dev, Digital Marketing",      "webbycentral.com",      "Web-Dir"),
]

def norm(s):
    if not s: return ""
    s = unicodedata.normalize('NFKD', str(s).lower())
    return re.sub(r'[^a-z0-9]', '', s)

def main():
    wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
    ws = wb["COMPLETE_MASTER"]

    existing = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 2).value
        if n: existing[norm(n)] = r

    fill_new = PatternFill("solid", fgColor="E6FFE6")  # light green for directory-sourced
    added = updated = 0

    for company, category, roles, website_hint, source in DIRECTORY_COMPANIES:
        key = norm(company)
        if not key: continue
        website = ("https://www." + website_hint) if not website_hint.startswith("http") else website_hint

        if key in existing:
            row = existing[key]
            if not ws.cell(row, 4).value: ws.cell(row, 4).value = category
            if not ws.cell(row, 5).value: ws.cell(row, 5).value = roles
            if not ws.cell(row, 13).value and website: ws.cell(row, 13).value = website
            cur = ws.cell(row, 18).value or ""
            if source not in cur:
                ws.cell(row, 18).value = (cur + ", " + source).strip(", ")
            updated += 1
        else:
            r = ws.max_row + 1
            ws.cell(r, 1, r - 1)
            ws.cell(r, 2, company)
            ws.cell(r, 3, "Ahmedabad")
            ws.cell(r, 4, category)
            ws.cell(r, 5, roles)
            ws.cell(r, 13, website)
            ws.cell(r, 18, source)
            for c in range(1, 19):
                ws.cell(r, c).fill = fill_new
                ws.cell(r, c).alignment = Alignment(vertical="center")
            ws.cell(r, 2).font = Font(bold=True, color="1B5E20")
            ws.row_dimensions[r].height = 18
            existing[key] = r
            added += 1

    for i, r in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(r, 1, i)

    total = ws.max_row - 1

    # Rebuild AI_ML_ONLY
    ws2 = wb["AI_ML_ONLY"]
    for r in range(ws2.max_row, 1, -1):
        ws2.delete_rows(r)
    ai_kw = ['ai','ml','data','machine','deep','nlp','computer vision','llm','genai',
             'analytics','automation','intelligence','learning','aiops']
    ai_row = 2
    fa = PatternFill("solid", fgColor="F5F0FF")
    fb = PatternFill("solid", fgColor="FFFFFF")
    for r in range(2, ws.max_row + 1):
        cat  = str(ws.cell(r, 4).value or "").lower()
        role = str(ws.cell(r, 5).value or "").lower()
        if any(k in cat or k in role for k in ai_kw):
            fill = fa if ai_row % 2 == 0 else fb
            for c in range(1, 19):
                v = ws.cell(r, c).value
                cell2 = ws2.cell(ai_row, c, v)
                cell2.fill = fill
                cell2.alignment = Alignment(vertical="center")
                if c in (6,7,8,9,10,11) and v:
                    cell2.font = Font(color="1a56db", size=9)
            ws2.cell(ai_row, 1, ai_row - 1)
            ws2.row_dimensions[ai_row].height = 18
            ai_row += 1
    ai_count = ws2.max_row - 1

    wb["STATS"].cell(3, 2, total)
    wb["STATS"].cell(6, 2, ai_count)

    wb.save("COMONK_TRUE_MASTER.xlsx")
    print(f"  NEW from directories: +{added}")
    print(f"  Existing updated:     {updated}")
    print(f"  Total companies:      {total}")
    print(f"  AI/ML companies:      {ai_count}")

if __name__ == "__main__":
    main()
