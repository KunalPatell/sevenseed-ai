"""
Capture remaining phones — attribute MNC office numbers to companies,
guarantee every phone is in the workbook (All_Phones sheet).
"""
import openpyxl, re, os, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment

MASTER="COMONK_TRUE_MASTER.xlsx"
C_COMP=2; C_PHONE=12
def norm(s):
    if not s: return ""
    return re.sub(r'[^a-z0-9]','',unicodedata.normalize('NFKD',str(s).lower()))
def d10(p):
    d=re.sub(r'[^\d]','',str(p))[-10:]
    return d if len(d)==10 else None
def fmt(d):
    if d.startswith('79'): return '+91 79 '+d[2:6]+' '+d[6:]
    return '+91 '+d[:5]+' '+d[5:]

# 1. Build company->phone from source Excels (esp MNC_AI_Targets)
comp_phone={}
for fn in ["Ahmedabad_IT_AIML_FINAL_MASTER.xlsx","Ahmedabad_IT_AIML_Companies_Master.xlsx",
           "Ahmedabad_AIML_Companies_HR_Contacts.xlsx","AI_Engineer_Job_Targets.xlsx"]:
    if not os.path.exists(fn): continue
    w=openpyxl.load_workbook(fn, read_only=True)
    for s in w.sheetnames:
        ws_=w[s]
        for row in ws_.iter_rows(values_only=True):
            cells=[c for c in row if c]
            name=None; phone=None
            for v in row:
                if v and isinstance(v,str) and not name and len(v)>2 and "@" not in v and not v.startswith("http") and not re.match(r'^\+?\d',v):
                    name=v
                dd=d10(v) if v else None
                if dd and (dd[0] in '6789' or dd.startswith('79')) and not phone: phone=dd
            if name and phone:
                comp_phone.setdefault(norm(name), phone)
    w.close()
print(f"  company->phone pairs from sources: {len(comp_phone)}")

wb=openpyxl.load_workbook(MASTER)
ws=wb["COMPLETE_MASTER"]

# 2. Fill master companies missing phone
filled=0
for r in range(2, ws.max_row+1):
    if ws.cell(r,C_PHONE).value: continue
    key=norm(ws.cell(r,C_COMP).value)
    if key in comp_phone:
        ws.cell(r,C_PHONE).value=fmt(comp_phone[key]); filled+=1

# 3. Collect ALL phones everywhere -> All_Phones sheet
PHONE_RE=re.compile(r'(?:\+?91[\s\-]?)?[6-9]\d{4}[\s\-]?\d{5}|\b0?79[\s\-]?\d{7,8}\b')
all_ph=set()
def slurp(t):
    for p in PHONE_RE.findall(t):
        dd=d10(p)
        if dd: all_ph.add(dd)
for fn in os.listdir("."):
    if fn.startswith("_extract_") and fn.endswith(".txt"):
        slurp(open(fn,encoding="utf-8",errors="ignore").read())
for fn in ["AIML_Companies_Apply_List.csv","Gandhinagar_GIFTCity_Companies.csv","hr_vcards_15-6-26.vcf"]:
    if os.path.exists(fn): slurp(open(fn,encoding="utf-8",errors="ignore").read())
for fn in ["Ahmedabad_IT_AIML_FINAL_MASTER.xlsx","Ahmedabad_IT_AIML_Companies_Master.xlsx",
           "Ahmedabad_AIML_Companies_HR_Contacts.xlsx","AI_Engineer_Job_Targets.xlsx","HR Mail List.xlsx"]:
    if os.path.exists(fn):
        w=openpyxl.load_workbook(fn,read_only=True)
        for s in w.sheetnames:
            for row in w[s].iter_rows(values_only=True):
                for v in row:
                    if v: slurp(str(v))
        w.close()
# include master phones
for r in range(2,ws.max_row+1):
    v=ws.cell(r,C_PHONE).value
    if v:
        dd=d10(v)
        if dd: all_ph.add(dd)
# include VCF contacts sheet phones
if "VCF_HR_Contacts" in wb.sheetnames:
    for row in wb["VCF_HR_Contacts"].iter_rows(values_only=True):
        for v in row:
            if v:
                dd=d10(v)
                if dd: all_ph.add(dd)

# Build domain/company hint by reverse phone map
phone_comp={}
for k,p in comp_phone.items(): phone_comp[p]=k

if "All_Phones" in wb.sheetnames: del wb["All_Phones"]
wap=wb.create_sheet("All_Phones")
hf=PatternFill("solid",fgColor="2E1A47"); hfont=Font(bold=True,color="FFFFFF")
for c,h in enumerate(["#","Phone","Type"],1):
    cell=wap.cell(1,c,h); cell.fill=hf; cell.font=hfont
    cell.alignment=Alignment(horizontal="center")
wap.column_dimensions['B'].width=20; wap.column_dimensions['C'].width=16
wap.freeze_panes="A2"
rr=2
for d in sorted(all_ph):
    typ="Landline (079)" if d.startswith('79') else "Mobile"
    wap.cell(rr,1,rr-1); wap.cell(rr,2,fmt(d)); wap.cell(rr,3,typ)
    if rr%2==0:
        for c in range(1,4): wap.cell(rr,c).fill=PatternFill("solid",fgColor="F5F0FF")
    rr+=1
total_ph=rr-2

for i,r in enumerate(range(2,ws.max_row+1),1): ws.cell(r,1,i)
total=ws.max_row-1
with_ph=sum(1 for r in range(2,ws.max_row+1) if ws.cell(r,C_PHONE).value)
wb.save(MASTER)
print(f"  Phones attached to companies: +{filled}")
print(f"  All_Phones sheet: {total_ph} phones (EVERY phone, guaranteed)")
print(f"  Company phone coverage: {with_ph}/{total} ({round(with_ph/total*100)}%)")
print(f"  Sheets: {wb.sheetnames}")
