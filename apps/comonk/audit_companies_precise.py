"""
PRECISE company audit — read the ACTUAL Company column from each structured file
(not a heuristic), compare to master. Report truly-missing real companies.
"""
import openpyxl, re, csv, os, unicodedata
def norm(s):
    if not s: return ""
    return re.sub(r'[^a-z0-9]','',unicodedata.normalize('NFKD',str(s).lower()))

wb=openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
ws=wb["COMPLETE_MASTER"]
master={norm(ws.cell(r,2).value) for r in range(2,ws.max_row+1) if ws.cell(r,2).value}

src={}   # norm -> display
def addn(n):
    if n and isinstance(n,str):
        n=n.strip()
        k=norm(n)
        if k and len(k)>=4 and not n.startswith("http") and "@" not in n:
            src.setdefault(k,n)

# CSVs with explicit Company column
for fn in ["AIML_Companies_Apply_List.csv","Gandhinagar_GIFTCity_Companies.csv","NEW_COMPANIES_DISCOVERED.csv"]:
    if os.path.exists(fn):
        for row in csv.DictReader(open(fn,encoding="utf-8-sig",errors="ignore")):
            for k in row:
                if k and k.strip().lower() in ("company","company name"): addn(row[k])

# Excels with known company column (header row detection)
def excel_companies(fn, col, header_row=1):
    if not os.path.exists(fn): return
    w=openpyxl.load_workbook(fn,read_only=True)
    for s in w.sheetnames:
        ws_=w[s]
        for i,row in enumerate(ws_.iter_rows(values_only=True),1):
            if i<=header_row: continue
            if len(row)>=col and row[col-1]: addn(row[col-1])
    w.close()

excel_companies("HR Mail List.xlsx",1,1)
excel_companies("Ahmedabad_AIML_Companies_HR_Contacts.xlsx",2,1)
excel_companies("Ahmedabad_IT_AIML_Companies_Master.xlsx",2,3)
excel_companies("Ahmedabad_IT_AIML_FINAL_MASTER.xlsx",2,3)
excel_companies("AI_Engineer_Job_Targets.xlsx",2,1)

# DataScience PDF numbered list
if os.path.exists("_extract_datascience_compeny_list_pdf.txt"):
    for line in open("_extract_datascience_compeny_list_pdf.txt",encoding="utf-8"):
        m=re.match(r'^\d+\.\s+(.+)',line.strip())
        if m: addn(m.group(1))

missing={k:v for k,v in src.items() if k not in master}
print(f"Structured source companies: {len(src)}")
print(f"In master: {len(src)-len(missing)}")
print(f"TRULY missing real companies: {len(missing)}")
print("\nMissing list:")
for k,v in sorted(missing.items(), key=lambda x:x[1].lower()):
    print(f"  {v}")
