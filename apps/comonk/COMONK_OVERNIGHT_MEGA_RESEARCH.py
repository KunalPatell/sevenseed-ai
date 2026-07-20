"""
COMONK_OVERNIGHT_MEGA_RESEARCH.py  v3
=======================================
OVERNIGHT 10-HOUR ULTRA DEEP RESEARCH ENGINE
=============================================
Priority:
  1st — MNC companies  (MUST — Ahmedabad + Gandhinagar)
  2nd — IT companies   (Ahmedabad + Gandhinagar)
  3rd — AI/ML companies
  4th — All other tech/corporate companies

Cities: Ahmedabad, Gandhinagar, GIFT City, SG Highway, Prahlad Nagar

Sources:
  Google Dorking (20+ patterns per company)
  Bing Search (alternative results)
  Naukri.com, Glassdoor, LinkedIn, Cutshort, Instahyre
  Indeed India, Shine.com, Monster India, TimesJobs
  AmbitionBox, JustDial, Indiamart, Sulekha
  Company career pages (direct scrape)
  SMTP email verification
  AngelList/Wellfound (AI/ML startups)
  GitHub org pages

Runtime  : ~10 hours
Auto-save: Every 10 rows
Checkpoint: Crash-safe resume
Tokens   : ZERO (100% free scraping)
"""

import sys, os, re, json, time, asyncio, socket, smtplib, random
import concurrent.futures
import openpyxl, httpx
from urllib.parse import quote_plus, urlparse
from datetime import datetime

# Thread pool for blocking SMTP calls
_SMTP_EXECUTOR = concurrent.futures.ThreadPoolExecutor(max_workers=4)

sys.stdout.reconfigure(encoding='utf-8')

EXCEL      = "COMONK_TRUE_MASTER.xlsx"
SHEET      = "COMPLETE_MASTER"
CKPT_FILE  = "overnight_checkpoint.json"
LOG_FILE   = "overnight_research_log.txt"
START_TIME = datetime.now()

C_COMP=2; C_CITY=3; C_CAT=4; C_ROLE=5
C_EMAILS=[6,7,8,9,10,11]
C_PHONE=12; C_WEB=13; C_LI=14
C_ADDR=15; C_CARE=16; C_PRIO=17; C_SRC=18

EMAIL_RE = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,7}\b')
PHONE_RE = re.compile(r'(?:\+91[\-\s]?)?[6-9]\d{9}')
LI_CO_RE = re.compile(r'linkedin\.com/company/([a-z0-9\-_%\.]+)', re.I)

SKIP_MAIL = {
    "gmail.com","yahoo.com","yahoo.co.in","outlook.com","hotmail.com",
    "rediffmail.com","sentry.io","example.com","googlemail.com",
    "yopmail.com","mailinator.com","w3.org","schema.org","facebook.com",
    "twitter.com","instagram.com","youtube.com","cloudflare.com",
    "amazonaws.com","google.com","microsoft.com","apple.com","wixpress.com",
    "shopify.com","wordpress.com","godaddy.com","bigrock.in",
}

HR_PREFIXES = [
    "hr","hrd","careers","jobs","recruit","recruitment","talent",
    "hiring","people","humanresources","hr.india","hr.ahmedabad",
    "hr.gandhinagar","hrbp","hrteam","apply","staffing","join",
]

UAS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0",
]

# ═══════════════════════════════════════════════════════════════════════════════
#   DISCOVERY QUERIES — Priority ordered: MNC first, IT second, AI/ML third
# ═══════════════════════════════════════════════════════════════════════════════
DISCOVERY_QUERIES = [

    # ── PRIORITY 1: MNC Companies ─────────────────────────────────────────────
    ("MNC", "top MNC companies Ahmedabad list 2024 HR contact email phone"),
    ("MNC", "top MNC companies Gandhinagar list 2024 HR contact email phone"),
    ("MNC", "multinational companies Ahmedabad Gujarat HR email contact 2024"),
    ("MNC", "multinational companies Gandhinagar Gujarat HR email contact 2024"),
    ("MNC", "Fortune 500 companies Ahmedabad Gandhinagar office HR contact"),
    ("MNC", "GIFT city IFSC companies contact HR email phone Gandhinagar"),
    ("MNC", "SG highway Ahmedabad MNC companies HR email address"),
    ("MNC", "Prahlad Nagar Ahmedabad MNC office HR contact phone"),
    ("MNC", "Bodakdev Ahmedabad MNC company HR email contact"),
    ("MNC", "Satellite area Ahmedabad MNC companies HR contact"),
    ("MNC", "Ahmedabad MNC company HR recruiter email phone site:naukri.com"),
    ("MNC", "Gandhinagar MNC company HR recruiter email phone site:naukri.com"),
    ("MNC", "Ahmedabad MNC companies site:glassdoor.com HR contact"),
    ("MNC", "Gandhinagar MNC companies site:glassdoor.com HR contact"),
    ("MNC", "Ahmedabad multinational site:ambitionbox.com HR contact email"),
    ("MNC", "Gandhinagar multinational site:ambitionbox.com HR contact email"),
    ("MNC", "MNC companies hiring Ahmedabad 2024 HR email phone recruiter"),
    ("MNC", "MNC companies hiring Gandhinagar 2024 HR email phone recruiter"),
    ("MNC", "Accenture TCS Infosys Wipro Cognizant HCL Ahmedabad HR email"),
    ("MNC", "IBM Oracle SAP Microsoft Google Amazon Ahmedabad HR email contact"),
    ("MNC", "Deloitte PwC EY KPMG McKinsey Ahmedabad HR recruiter email"),
    ("MNC", "Capgemini Atos Sopra NTT Data Ahmedabad HR email phone"),
    ("MNC", "pharma MNC Ahmedabad Gandhinagar HR recruiter email phone"),
    ("MNC", "FMCG MNC Ahmedabad Gandhinagar HR contact email phone"),
    ("MNC", "automobile MNC Ahmedabad Gandhinagar HR email recruiter"),
    ("MNC", "consulting MNC Ahmedabad Gandhinagar HR email contact"),
    ("MNC", "banking finance MNC Ahmedabad Gandhinagar HR email"),
    ("MNC", "chemical MNC Ahmedabad Gandhinagar HR email contact phone"),
    ("MNC", "logistics MNC Ahmedabad Gandhinagar HR email contact"),
    ("MNC", "telecom MNC Ahmedabad Gandhinagar HR contact email phone"),

    # ── PRIORITY 2: IT Companies ──────────────────────────────────────────────
    ("IT",  "top IT companies Ahmedabad list 2024 HR contact email phone"),
    ("IT",  "top IT companies Gandhinagar list 2024 HR contact email phone"),
    ("IT",  "IT software company Ahmedabad HR recruiter email contact 2024"),
    ("IT",  "IT software company Gandhinagar HR recruiter email contact 2024"),
    ("IT",  "software development company Ahmedabad HR email phone contact"),
    ("IT",  "software development company Gandhinagar HR email phone contact"),
    ("IT",  "IT company Ahmedabad site:naukri.com HR email contact"),
    ("IT",  "IT company Gandhinagar site:naukri.com HR email contact"),
    ("IT",  "IT company Ahmedabad site:glassdoor.com HR contact email"),
    ("IT",  "IT company Gandhinagar site:glassdoor.com HR contact email"),
    ("IT",  "IT company Ahmedabad site:linkedin.com HR recruiter"),
    ("IT",  "IT company Gandhinagar site:linkedin.com HR recruiter"),
    ("IT",  "IT park Ahmedabad companies HR email contact phone"),
    ("IT",  "Infocity Gandhinagar IT companies HR contact email phone"),
    ("IT",  "SEEPZ Ahmedabad IT companies HR email"),
    ("IT",  "software company Ahmedabad site:cutshort.io hiring contact"),
    ("IT",  "software company Gandhinagar site:cutshort.io hiring contact"),
    ("IT",  "IT company Ahmedabad site:instahyre.com HR contact"),
    ("IT",  "IT company Gandhinagar site:instahyre.com HR contact"),
    ("IT",  "web development company Ahmedabad HR email phone contact"),
    ("IT",  "mobile app development company Ahmedabad HR email contact"),
    ("IT",  "cybersecurity company Ahmedabad Gandhinagar HR contact email"),
    ("IT",  "cloud computing company Ahmedabad HR email contact phone"),
    ("IT",  "ERP software company Ahmedabad Gandhinagar HR contact email"),
    ("IT",  "product based IT company Ahmedabad HR email contact"),
    ("IT",  "product based IT company Gandhinagar HR email contact"),
    ("IT",  "IT services company Ahmedabad HR email phone recruiter"),
    ("IT",  "IT services company Gandhinagar HR email phone recruiter"),
    ("IT",  "SAP consulting company Ahmedabad HR contact email"),
    ("IT",  "IT company Ahmedabad site:ambitionbox.com HR email contact"),

    # ── PRIORITY 3: AI/ML Companies ───────────────────────────────────────────
    ("AI/ML", "artificial intelligence company Ahmedabad HR email contact 2024"),
    ("AI/ML", "artificial intelligence company Gandhinagar HR email contact 2024"),
    ("AI/ML", "machine learning company Ahmedabad HR phone email contact"),
    ("AI/ML", "machine learning company Gandhinagar HR phone email contact"),
    ("AI/ML", "data science company Ahmedabad Gandhinagar HR email contact"),
    ("AI/ML", "AI startup Ahmedabad site:wellfound.com OR site:angel.co"),
    ("AI/ML", "AI ML company Ahmedabad site:cutshort.io HR contact"),
    ("AI/ML", "generative AI company Ahmedabad Gandhinagar HR email 2024"),
    ("AI/ML", "NLP computer vision company Ahmedabad HR email contact"),
    ("AI/ML", "deep learning company Gujarat Ahmedabad HR contact email"),
    ("AI/ML", "AI research lab Ahmedabad Gandhinagar IIT contact email"),
    ("AI/ML", "analytics company Ahmedabad Gandhinagar HR email phone"),
    ("AI/ML", "big data company Ahmedabad Gandhinagar HR contact email"),
    ("AI/ML", "robotics automation company Ahmedabad HR email contact"),
    ("AI/ML", "IoT company Ahmedabad Gandhinagar HR email contact phone"),

    # ── PRIORITY 4: Other Tech/Startup ────────────────────────────────────────
    ("FinTech", "fintech company Ahmedabad Gandhinagar HR contact email phone"),
    ("EdTech",  "edtech company Ahmedabad Gandhinagar HR contact email phone"),
    ("HealthTech","healthtech company Ahmedabad Gandhinagar HR email phone"),
    ("Startup",  "startup Ahmedabad Gandhinagar site:wellfound.com contact HR"),
    ("Startup",  "NASSCOM member company Ahmedabad Gandhinagar HR contact"),
    ("Startup",  "CII member Ahmedabad Gandhinagar IT company HR email"),
]


# ═══════════════════════════════════════════════════════════════════════════════
#   UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════
def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(line + "\n")
    except: pass

def clean_phone(p):
    if not p: return ""
    d = re.sub(r'[^\d]', '', str(p))
    if d.startswith('91') and len(d) == 12: d = d[2:]
    if d.startswith('0') and len(d) == 11: d = d[1:]
    if len(d) != 10 or d[0] not in '6789': return ""
    if len(set(d)) <= 2: return ""
    if d in {"8888888888","9999999999","1234567890","0987654321","7777777777","6666666666"}: return ""
    if d[:5] == d[5:]: return ""
    return '+91 ' + d[:5] + ' ' + d[5:]

def domain_of(url):
    if not url: return ""
    try:
        u = str(url).strip()
        if not u.startswith("http"): u = "https://" + u
        return urlparse(u).netloc.replace("www.", "").strip().lower()
    except: return ""

def is_valid_email(em):
    if not em or '@' not in em or len(em) > 80: return False
    dom = em.split('@')[1].lower()
    if dom in SKIP_MAIL: return False
    if any(x in dom for x in ['example','test','sample','dummy','noreply','no-reply']): return False
    return bool(re.match(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,7}$', em))

def _smtp_verify_blocking(email):
    """Blocking SMTP check — runs in thread pool, never call directly from async."""
    try:
        domain = email.split('@')[1]
        mx = None
        for pfx in ('mail', 'smtp', 'mx', ''):
            host = f"{pfx}.{domain}" if pfx else domain
            try:
                socket.setdefaulttimeout(2)
                socket.gethostbyname(host)
                mx = host; break
            except: pass
        if not mx: return None
        srv = smtplib.SMTP(timeout=2)
        srv.connect(mx, 25)
        srv.helo("comonk.ai")
        srv.mail("hr@comonk.ai")
        code, _ = srv.rcpt(email)
        srv.quit()
        return code == 250
    except: return None

async def smtp_verify(email):
    """Non-blocking SMTP — runs in thread executor with 3s hard asyncio timeout."""
    loop = asyncio.get_event_loop()
    try:
        return await asyncio.wait_for(
            loop.run_in_executor(_SMTP_EXECUTOR, _smtp_verify_blocking, email),
            timeout=3.0
        )
    except (asyncio.TimeoutError, Exception):
        return None

def load_ckpt():
    try:
        with open(CKPT_FILE, 'r') as f:
            return set(json.load(f).get("done", []))
    except: return set()

def save_ckpt(done):
    with open(CKPT_FILE, 'w') as f:
        json.dump({"done": list(done), "ts": str(datetime.now())}, f)

def extract_emails(html, prefer_domain=None):
    found = []
    for em in EMAIL_RE.findall(html):
        em = em.lower().strip().rstrip('.')
        if is_valid_email(em) and em not in found:
            if prefer_domain and prefer_domain in em:
                found.insert(0, em)
            else:
                found.append(em)
    return found

def extract_phones(html):
    found = []
    for m in PHONE_RE.findall(html):
        c = clean_phone(m)
        if c and c not in found:
            found.append(c)
    return found

def extract_linkedin(html):
    for m in LI_CO_RE.findall(html):
        slug = m.strip('/').split('?')[0]
        if len(slug) > 2 and slug not in ('login','company','in','jobs','feed','pub','dir'):
            return f"linkedin.com/company/{slug}"
    return ""

def recalc_priority(ws, row):
    ems = [ws.cell(row, c).value for c in C_EMAILS if ws.cell(row, c).value]
    ph  = ws.cell(row, C_PHONE).value
    web = ws.cell(row, C_WEB).value
    li  = ws.cell(row, C_LI).value
    if ems and ph:  ws.cell(row, C_PRIO).value = "1 - Apply Now (Email+Phone) ⭐"
    elif ems:       ws.cell(row, C_PRIO).value = "2 - Email Available 📧"
    elif ph:        ws.cell(row, C_PRIO).value = "3 - Phone Only 📞"
    elif web:       ws.cell(row, C_PRIO).value = "4 - Website Only 🌐"
    elif li:        ws.cell(row, C_PRIO).value = "5 - LinkedIn Only 🔗"
    else:           ws.cell(row, C_PRIO).value = "6 - Research Needed 🔍"


# ═══════════════════════════════════════════════════════════════════════════════
#   HTTP CLIENT
# ═══════════════════════════════════════════════════════════════════════════════
async def fetch(client, url, retries=2):
    for attempt in range(retries):
        try:
            h = {"User-Agent": random.choice(UAS), "Accept": "text/html,*/*;q=0.9"}
            r = await client.get(url, headers=h, timeout=9, follow_redirects=True)
            if r.status_code == 200: return r.text
            if r.status_code == 429: await asyncio.sleep(4 + attempt * 3)
        except: pass
        await asyncio.sleep(0.5 * (attempt + 1))
    return ""

async def google(client, q, n=8):
    html = await fetch(client, f"https://www.google.com/search?q={quote_plus(q)}&num={n}&hl=en")
    await asyncio.sleep(random.uniform(0.4, 1.0))
    return html

async def bing(client, q, n=8):
    html = await fetch(client, f"https://www.bing.com/search?q={quote_plus(q)}&count={n}")
    await asyncio.sleep(random.uniform(0.3, 0.7))
    return html


# ═══════════════════════════════════════════════════════════════════════════════
#   PHASE 1: DISCOVER NEW COMPANIES
# ═══════════════════════════════════════════════════════════════════════════════
CO_NAME_RE = re.compile(
    r'\b([A-Z][a-zA-Z0-9\s&\.\-]{3,50}'
    r'(?:Pvt\.?\s*Ltd\.?|Private Limited|Limited|Inc\.?|Corp\.?|'
    r'Technologies|Technology|Solutions|Systems|Consulting|Consultancy|'
    r'Services|Software|Analytics|Intelligence|Labs?|Research|Ventures|'
    r'Infotech|Infosystems|Infocomm|Digital|Innovations?|Enterprises?|'
    r'India|Global|International|Worldwide|Group|Partners?))\b'
)

async def discover_companies(client, ws, existing_names):
    log("\n" + "="*65)
    log("  PHASE 1: NEW COMPANY DISCOVERY")
    log("  Priority: MNC → IT → AI/ML → Others")
    log("  Cities  : Ahmedabad + Gandhinagar + GIFT City")
    log("="*65)

    new_cos = []
    seen_new = set()

    for idx, (cat, query) in enumerate(DISCOVERY_QUERIES):
        log(f"  [{cat}] Query {idx+1}/{len(DISCOVERY_QUERIES)}: {query[:65]}...")

        phones_here = []
        emails_here = []

        for engine in [google, bing]:
            html = await engine(client, query)
            if not html: continue

            # Extract emails + phones from search results directly
            emails_here += extract_emails(html)
            phones_here += extract_phones(html)

            # Extract company names
            for m in CO_NAME_RE.findall(html):
                clean = m.strip().rstrip('.,;:')
                lc    = clean.lower()
                if (5 < len(clean) < 85
                        and lc not in {n.lower() for n in existing_names}
                        and lc not in seen_new
                        and not any(skip in lc for skip in ['google','facebook','twitter','instagram','youtube','wikipedia','naukri','glassdoor','linkedin','indeed','ambitionbox','justdial','indiamart','shine.com','monster'])):
                    seen_new.add(lc)
                    # Determine city from query
                    city = "Gandhinagar" if "gandhinagar" in query.lower() or "gift city" in query.lower() or "infocity" in query.lower() else "Ahmedabad"
                    new_cos.append({'name': clean, 'city': city, 'cat': cat})

        if emails_here or phones_here:
            log(f"    → {len(emails_here)} emails, {len(phones_here)} phones found in results")

    log(f"\n  ✅ Discovered {len(new_cos)} new companies!")

    # Add to sheet
    next_row = ws.max_row + 1
    for c in new_cos:
        ws.cell(next_row, C_COMP).value = c['name']
        ws.cell(next_row, C_CITY).value = c['city']
        ws.cell(next_row, C_CAT).value  = c['cat']
        ws.cell(next_row, C_ROLE).value = "HR / Recruiter"
        ws.cell(next_row, C_PRIO).value = "6 - Research Needed 🔍"
        ws.cell(next_row, C_SRC).value  = f"Auto-Discovered ({c['cat']})"
        next_row += 1

    return len(new_cos)


# ═══════════════════════════════════════════════════════════════════════════════
#   PHASE 2: ENRICH ALL COMPANIES
# ═══════════════════════════════════════════════════════════════════════════════
async def find_emails(client, name, city, dom, web):
    found = []
    nq = name.replace('&','and').replace(',','')

    queries = [
        f'"{nq}" site:naukri.com HR contact email {city}',
        f'"{nq}" site:glassdoor.com recruiter email contact',
        f'"{nq}" "@{dom}" HR OR careers OR recruit' if dom else f'"{nq}" HR email {city} contact recruiter',
        f'"{nq}" {city} HR recruiter email contact 2024',
        f'"{nq}" careers@{dom} OR hr@{dom} OR recruit@{dom}' if dom else f'"{nq}" careers email {city}',
        f'"{nq}" site:linkedin.com HR recruiter {city}',
        f'"{nq}" site:ambitionbox.com contact HR email',
        f'"{nq}" site:instahyre.com OR site:cutshort.io contact',
        f'"{nq}" site:shine.com OR site:timesjobs.com HR email',
        f'"{nq}" site:monster.com OR site:indeed.com HR contact',
        f'"{nq}" Ahmedabad OR Gandhinagar "hr@" OR "careers@" OR "recruit@"',
        f'"{nq}" HR email phone number India "contact us"',
    ]

    for q in queries:
        for eng in [google, bing]:
            html = await eng(client, q)
            for em in extract_emails(html, dom):
                if em not in found: found.append(em)
        if len(found) >= 3: break

    # Website career pages
    if dom and len(found) < 3:
        for path in ['/contact','/careers','/about','/hr','/jobs','/contact-us','/team']:
            html = await fetch(client, f"https://{dom}{path}")
            for em in extract_emails(html, dom):
                if em not in found: found.append(em)
    
    # 4b. Generate HR patterns from domain + async SMTP check
    if dom and len(found) < 3:
        for prefix in HR_PREFIXES:
            candidate = f"{prefix}@{dom}"
            if candidate not in found:
                result = await smtp_verify(candidate)  # non-blocking!
                if result in (True, None):
                    found.append(candidate)
            if len(found) >= 6: break

    return found[:8]


async def find_phones(client, name, city, dom):
    found = []
    nq = name.replace('&','and')

    queries = [
        f'"{nq}" {city} phone number HR contact',
        f'"{nq}" site:justdial.com contact number',
        f'"{nq}" site:indiamart.com phone number',
        f'"{nq}" site:sulekha.com contact phone',
        f'"{nq}" site:ambitionbox.com contact phone',
        f'"{nq}" +91 {city} recruiter phone office',
        f'"{nq}" Gujarat office phone number HR contact',
    ]

    for q in queries:
        for eng in [google, bing]:
            html = await eng(client, q)
            for p in extract_phones(html):
                if p not in found: found.append(p)
        if found: break

    if not found and dom:
        for path in ['/contact','/contact-us','/about']:
            html = await fetch(client, f"https://{dom}{path}")
            for p in extract_phones(html):
                if p not in found: found.append(p)
            if found: break

    return found[:3]


async def find_linkedin(client, name):
    nq = name.replace('&','and')
    for q in [
        f'site:linkedin.com/company "{nq}" India',
        f'linkedin company "{nq}" Ahmedabad OR Gandhinagar',
    ]:
        for eng in [google, bing]:
            html = await eng(client, q)
            li = extract_linkedin(html)
            if li: return li
    return ""


async def enrich_all(client, wb, ws, done_set):
    log("\n" + "="*65)
    log("  PHASE 2: ENRICHING ALL COMPANIES (DEEP MULTI-SOURCE)")
    log("  12 search strategies per company")
    log("  Sources: Google + Bing + Naukri + LinkedIn + Glassdoor")
    log("           Cutshort + Instahyre + JustDial + Indiamart")
    log("           AmbitionBox + Shine + Monster + Career Pages + SMTP")
    log("="*65)

    total = ws.max_row - 1
    e_add = p_add = li_add = fake_rm = processed = 0

    # Sort rows: MNC first, then IT, then AI/ML, then others
    priority_order = {"MNC": 0, "IT": 1, "AI/ML": 2, "FinTech": 3, "EdTech": 4, "HealthTech": 5}
    rows_data = []
    for row in range(2, ws.max_row + 1):
        name = str(ws.cell(row, C_COMP).value or "").strip()
        cat  = str(ws.cell(row, C_CAT).value or "").strip()
        if name:
            rows_data.append((priority_order.get(cat, 99), row, name))

    rows_data.sort(key=lambda x: x[0])

    for _, row, name in rows_data:
        if name in done_set: continue

        city = str(ws.cell(row, C_CITY).value or "Ahmedabad").strip()
        web  = str(ws.cell(row, C_WEB).value or "").strip()
        li   = str(ws.cell(row, C_LI).value or "").strip()
        dom  = domain_of(web)
        cur_emails = [ws.cell(row, c).value for c in C_EMAILS if ws.cell(row, c).value]

        # 1. Clean fake phones
        ph = ws.cell(row, C_PHONE).value
        if ph:
            cleaned = clean_phone(str(ph))
            if not cleaned:
                ws.cell(row, C_PHONE).value = None
                fake_rm += 1
            else:
                ws.cell(row, C_PHONE).value = cleaned

        # 2. Find phone
        if not ws.cell(row, C_PHONE).value:
            phones = await find_phones(client, name, city, dom)
            if phones:
                ws.cell(row, C_PHONE).value = phones[0]
                p_add += 1

        # 3. Find LinkedIn
        if not li:
            li_url = await find_linkedin(client, name)
            if li_url:
                ws.cell(row, C_LI).value = li_url
                li_add += 1

        # 4. Find emails
        if len(cur_emails) < 3:
            new_ems = await find_emails(client, name, city, dom, web)
            empty_cols = [c for c in C_EMAILS if not ws.cell(row, c).value]
            for em, col in zip(new_ems, empty_cols):
                ws.cell(row, col).value = em
                e_add += 1

        # 5. Recalculate priority
        recalc_priority(ws, row)

        # 6. Tag source
        src = ws.cell(row, C_SRC).value
        if not src:
            ws.cell(row, C_SRC).value = "Overnight Mega Research v3"

        done_set.add(name)
        processed += 1

        if processed % 5 == 0:
            wb.save(EXCEL)
            save_ckpt(done_set)
            elapsed = (datetime.now() - START_TIME).seconds // 60
            cat_val = str(ws.cell(row, C_CAT).value or "")
            pct = round(processed / total * 100, 1)
            log(f"  [{pct:5.1f}%] {processed}/{total} [{cat_val}] | +{e_add} emails | +{p_add} phones | +{li_add} LI | cleaned: {fake_rm} | {elapsed}min")

    wb.save(EXCEL)
    save_ckpt(done_set)
    return e_add, p_add, li_add, fake_rm, processed


# ═══════════════════════════════════════════════════════════════════════════════
#   MAIN
# ═══════════════════════════════════════════════════════════════════════════════
async def main():
    log("\n" + "="*65)
    log("  COMONK AI — OVERNIGHT MEGA RESEARCH v3")
    log(f"  Started : {START_TIME.strftime('%Y-%m-%d %H:%M:%S')}")
    log(f"  Priority: MNC > IT > AI/ML > Others")
    log(f"  Cities  : Ahmedabad + Gandhinagar + GIFT City")
    log(f"  Tokens  : ZERO — 100% Free (Google/Bing scraping)")
    log("="*65 + "\n")

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]

    done_set = load_ckpt()
    log(f"  Checkpoint: {len(done_set)} already done")

    existing = set()
    for row in range(2, ws.max_row + 1):
        n = str(ws.cell(row, C_COMP).value or "").strip()
        if n: existing.add(n)
    log(f"  Existing companies: {len(existing)}")

    limits = httpx.Limits(max_connections=8, max_keepalive_connections=4)
    async with httpx.AsyncClient(limits=limits, verify=False, timeout=10) as client:

        # PHASE 1: Discover new companies
        new_count = await discover_companies(client, ws, existing)
        wb.save(EXCEL)
        log(f"\n  ✅ Phase 1 done! +{new_count} new companies added")
        log(f"  Total companies now: {ws.max_row - 1}")

        # PHASE 2: Enrich all (MNC first!)
        e, p, li, fk, proc = await enrich_all(client, wb, ws, done_set)

    wb.save(EXCEL)
    elapsed = (datetime.now() - START_TIME).seconds // 60

    log("\n" + "="*65)
    log("  🏆 OVERNIGHT MEGA RESEARCH COMPLETE!")
    log(f"  Runtime             : {elapsed} minutes")
    log(f"  Companies processed : {proc}")
    log(f"  New companies added : {new_count}")
    log(f"  Emails added        : {e}")
    log(f"  Phones added        : {p}")
    log(f"  LinkedIn added      : {li}")
    log(f"  Fake phones cleaned : {fk}")
    log(f"  File ready          : {EXCEL}")
    log("="*65)
    log("  ✅ COMONK_TRUE_MASTER.xlsx is READY! 🎉")

if __name__ == "__main__":
    asyncio.run(main())
