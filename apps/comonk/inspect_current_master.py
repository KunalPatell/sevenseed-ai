import openpyxl

wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
ws = wb["COMPLETE_MASTER"]
print("Sheets:", wb.sheetnames)
print("Rows:", ws.max_row - 1)

headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i+1 for i, h in enumerate(headers) if h}
print("Headers:", headers)

COL_CO = hmap.get("company name", 2)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]
email_cols = [c for c in email_cols if c]
COL_PH = hmap.get("phone", 12)
COL_CAT = hmap.get("category", 4)

MNC_NAMES = ["tcs", "tata consultancy", "infosys", "wipro", "hcl", "tech mahindra",
             "capgemini", "accenture", "ibm", "cognizant", "deloitte", "ey ",
             "ernst", "pwc", "bosch", "siemens", "oracle", "ntt data", "fujitsu",
             "kpit", "ltts", "l&t tech", "mphasis", "persistent", "hexaware",
             "apexon", "infostretch", "mastech", "exl"]

print("\n--- MNC rows found in current master ---")
found_mncs = {}
for r in range(2, ws.max_row + 1):
    name = str(ws.cell(r, COL_CO).value or "")
    nlow = name.lower()
    for key in MNC_NAMES:
        if key in nlow:
            ems = [ws.cell(r, c).value for c in email_cols if ws.cell(r, c).value]
            ph = ws.cell(r, COL_PH).value
            found_mncs.setdefault(key, []).append((r, name, len(ems), ph))
            break

for k, rows in found_mncs.items():
    for (r, name, ecount, ph) in rows:
        print(f"  [{key if False else k:<12}] row={r:<5} '{name}' emails={ecount} phone={ph}")

print(f"\nTotal MNC-like rows found: {sum(len(v) for v in found_mncs.values())}")
print(f"Unique MNC keys matched: {len(found_mncs)}")
