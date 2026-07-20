"""
For the 780 gap companies (no email/phone/website), guess likely domains
(name.com, name.in, name.co.in, nameindia.com) and probe them directly via
HTTP. If a domain resolves with real content, scrape it for emails/phones.
Fully automated, no API needed, purely additive (never overwrites existing data).
"""
import openpyxl, re, httpx, asyncio, unicodedata
from openpyxl.styles import Font

MASTER = "COMONK_TRUE_MASTER.xlsx"
EMAIL_RE = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
PHONE_RE = re.compile(r'(?:\+91[\s\-]?|0)?[6-9]\d{9}')
GENERIC_DOMAINS = {"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com"}

wb = openpyxl.load_workbook(MASTER)
ws = wb["COMPLETE_MASTER"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CO   = hmap.get("company name", 2)
COL_WEB  = hmap.get("website", 13)
COL_PH   = hmap.get("phone", 12)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]

def clean_slug(name):
    s = unicodedata.normalize('NFKD', str(name).lower())
    s = re.sub(r'\(.*?\)', '', s)  # remove parens content
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

gap_rows = []
for r in range(2, ws.max_row + 1):
    has_email = any(ws.cell(r, c).value for c in email_cols)
    has_web   = ws.cell(r, COL_WEB).value
    has_ph    = ws.cell(r, COL_PH).value
    if not has_email and not has_web and not has_ph:
        name = ws.cell(r, COL_CO).value
        if name:
            gap_rows.append((r, str(name)))

print(f"Gap companies to probe: {len(gap_rows)}")

TLDS = [".com", ".in", ".co.in"]

async def probe_company(client, r, name, sem):
    slug = clean_slug(name)
    if len(slug) < 3:
        return None
    async with sem:
        for tld in TLDS:
            url = f"https://www.{slug}{tld}"
            try:
                resp = await client.get(url, timeout=5, follow_redirects=True,
                    headers={"User-Agent": "Mozilla/5.0"})
                if resp.status_code == 200 and len(resp.text) > 500:
                    text_lower = resp.text.lower()
                    # Skip obvious parked-domain pages
                    if any(p in text_lower for p in ["domain is for sale", "buy this domain",
                                                        "godaddy.com/domainsearch", "sedo.com"]):
                        continue
                    return (r, name, url, resp.text)
            except Exception:
                continue
        return None

async def run_probe():
    sem = asyncio.Semaphore(40)
    limits = httpx.Limits(max_connections=50, max_keepalive_connections=25)
    found = []
    async with httpx.AsyncClient(limits=limits) as client:
        tasks = [probe_company(client, r, name, sem) for r, name in gap_rows]
        done = 0
        for i in range(0, len(tasks), 100):
            batch = tasks[i:i+100]
            results = await asyncio.gather(*batch, return_exceptions=True)
            for res in results:
                if res and not isinstance(res, Exception):
                    found.append(res)
            done += len(batch)
            print(f"  Probed {done}/{len(tasks)} | found {len(found)} live domains so far")
    return found

print("\nPhase 1: Domain probing (this may take a few minutes)...")
found = asyncio.run(run_probe())
print(f"\nPhase 1 done: {len(found)} companies have a live, guessable domain")

# ── Phase 2: Apply website + extract email/phone from already-fetched HTML ────
emails_added = 0; phones_added = 0; webs_added = 0
for (r, name, url, html) in found:
    if not ws.cell(r, COL_WEB).value:
        ws.cell(r, COL_WEB).value = url
        webs_added += 1

    domain = re.sub(r'^https?://(www\.)?', '', url).split('/')[0]
    ems = set()
    for em in EMAIL_RE.findall(html):
        em = em.lower()
        d = em.split("@")[1] if "@" in em else ""
        if d and d not in GENERIC_DOMAINS and domain in d:
            ems.add(em)
    free_slots = [c for c in email_cols if not ws.cell(r, c).value]
    for em, slot in zip(ems, free_slots):
        ws.cell(r, slot).value = em
        ws.cell(r, slot).font = Font(color="0070C0", size=9)
        emails_added += 1

    if not ws.cell(r, COL_PH).value:
        phones = PHONE_RE.findall(html)
        if phones:
            ph = re.sub(r'[\s\-]', '', phones[0])
            if not ph.startswith("+91"):
                ph = "+91" + ph.lstrip("0")
            ws.cell(r, COL_PH).value = ph
            phones_added += 1

print(f"Phase 2 - Applied: +{webs_added} websites, +{emails_added} emails, +{phones_added} phones")

# ── Phase 3: For domains found but no email in homepage, try contact page ────
async def scrape_contact_pages():
    need_contact = [(r, name, url) for (r, name, url, html) in found
                     if not any(ws.cell(r, c).value for c in email_cols)]
    print(f"\nPhase 3: {len(need_contact)} companies need contact-page scrape...")
    if not need_contact:
        return 0
    sem = asyncio.Semaphore(30)
    added = 0

    async def scrape_one(r, name, url, client):
        nonlocal added
        async with sem:
            domain = re.sub(r'^https?://(www\.)?', '', url).split('/')[0]
            for path in ["/contact", "/contact-us", "/contactus"]:
                try:
                    resp = await client.get(f"https://{domain}{path}", timeout=6,
                        follow_redirects=True, headers={"User-Agent": "Mozilla/5.0"})
                    ems = set()
                    for em in EMAIL_RE.findall(resp.text):
                        em = em.lower()
                        d = em.split("@")[1] if "@" in em else ""
                        if d and d not in GENERIC_DOMAINS and domain in d:
                            ems.add(em)
                    if ems:
                        free_slots = [c for c in email_cols if not ws.cell(r, c).value]
                        for em, slot in zip(ems, free_slots):
                            ws.cell(r, slot).value = em
                            ws.cell(r, slot).font = Font(color="0070C0", size=9)
                            added += 1
                        break
                except Exception:
                    continue

    limits = httpx.Limits(max_connections=35, max_keepalive_connections=20)
    async with httpx.AsyncClient(limits=limits) as client:
        for i in range(0, len(need_contact), 30):
            batch = need_contact[i:i+30]
            await asyncio.gather(*[scrape_one(r, n, u, client) for r, n, u in batch],
                                 return_exceptions=True)
    print(f"  Contact-page scrape done: +{added} emails")
    return added

p3 = asyncio.run(scrape_contact_pages())

wb.save(MASTER)

print(f"\n{'='*55}")
print(f"  DOMAIN-PROBE ENRICHMENT COMPLETE")
print(f"{'='*55}")
print(f"  Live domains found  : {len(found)} / {len(gap_rows)}")
print(f"  Websites added      : +{webs_added}")
print(f"  Emails added        : +{emails_added + p3}")
print(f"  Phones added        : +{phones_added}")
print(f"  Saved -> {MASTER}")
