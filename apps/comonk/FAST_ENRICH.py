"""
FAST_ENRICH.py — Upgraded with 429 rate limit fallback
======================================================
Strategy:
- If Google returns 429, immediately switch to DuckDuckGo/Bing
- Dynamic backoff delay when rate limits are hit
- Keep 10 parallel workers, but add a retry/fallback mechanism
"""

import sys, os, re, json, asyncio, random, socket
import openpyxl, httpx
from urllib.parse import quote_plus, urlparse
from datetime import datetime
import concurrent.futures

sys.stdout.reconfigure(encoding='utf-8')

EXCEL     = "COMONK_TRUE_MASTER.xlsx"
SHEET     = "COMPLETE_MASTER"
CKPT      = "fast_enrich_checkpoint.json"
LOG       = "fast_enrich_log.txt"
START     = datetime.now()

C_COMP=2; C_CITY=3; C_CAT=4; C_ROLE=5
C_EMAILS=[6,7,8,9,10,11]
C_PHONE=12; C_WEB=13; C_LI=14
C_ADDR=15; C_CARE=16; C_PRIO=17; C_SRC=18

EMAIL_RE = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,7}\b')
PHONE_RE = re.compile(r'(?:\+91[\-\s]?)?[6-9]\d{9}')
LI_RE    = re.compile(r'linkedin\.com/company/([a-z0-9\-_%\.]+)', re.I)

SKIP_MAIL = {
    "gmail.com","yahoo.com","yahoo.co.in","outlook.com","hotmail.com",
    "rediffmail.com","sentry.io","example.com","googlemail.com",
    "facebook.com","twitter.com","instagram.com","youtube.com",
    "cloudflare.com","google.com","microsoft.com","w3.org","schema.org",
}

HR_PATTERNS = [
    "hr","careers","recruit","jobs","talent","hiring","people","hrd",
    "humanresources","apply","staffing","hr.india","recruitment",
]

UAS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

SEMAPHORE = asyncio.Semaphore(5)  # reduced to 5 to avoid IP blocks
RATE_LIMIT_DELAY = 0.0

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    try:
        with open(LOG, 'a', encoding='utf-8') as f: f.write(line+"\n")
    except: pass

def clean_phone(p):
    if not p: return ""
    d = re.sub(r'[^\d]', '', str(p))
    if d.startswith('91') and len(d)==12: d=d[2:]
    if d.startswith('0') and len(d)==11: d=d[1:]
    if len(d)!=10 or d[0] not in '6789': return ""
    if len(set(d))<=2: return ""
    if d in {"8888888888","9999999999","1234567890","0987654321"}: return ""
    if d[:5]==d[5:]: return ""
    return '+91 '+d[:5]+' '+d[5:]

def domain_of(url):
    if not url: return ""
    try:
        u = str(url).strip()
        if not u.startswith("http"): u="https://"+u
        return urlparse(u).netloc.replace("www.","").strip().lower()
    except: return ""

def valid_email(em):
    if not em or '@' not in em or len(em)>80: return False
    dom = em.split('@')[1].lower()
    return dom not in SKIP_MAIL and not any(x in dom for x in ['example','test','noreply','no-reply','dummy'])

def load_ckpt():
    try:
        with open(CKPT,'r') as f: return set(json.load(f).get("done",[]))
    except: return set()

def save_ckpt(done):
    with open(CKPT,'w') as f: json.dump({"done":list(done),"ts":str(datetime.now())},f)

def extract_emails(html, dom=None):
    found=[]
    for em in EMAIL_RE.findall(html):
        em=em.lower().strip().rstrip('.')
        if valid_email(em) and em not in found:
            if dom and dom in em: found.insert(0,em)
            else: found.append(em)
    return found

def extract_phones(html):
    found=[]
    for m in PHONE_RE.findall(html):
        c=clean_phone(m)
        if c and c not in found: found.append(c)
    return found

def extract_linkedin(html):
    for m in LI_RE.findall(html):
        slug=m.strip('/').split('?')[0]
        if len(slug)>2 and slug not in ('login','company','in','jobs','feed','pub'):
            return f"linkedin.com/company/{slug}"
    return ""

def recalc_priority(ems, ph, web, li):
    if ems and ph: return "1 - Apply Now (Email+Phone) ⭐"
    if ems:        return "2 - Email Available 📧"
    if ph:         return "3 - Phone Only 📞"
    if web:        return "4 - Website Only 🌐"
    if li:         return "5 - LinkedIn Only 🔗"
    return "6 - Research Needed 🔍"

# ── HTTP ──────────────────────────────────────────────────────────────────────
async def fetch(client, url):
    global RATE_LIMIT_DELAY
    try:
        h={"User-Agent":random.choice(UAS),"Accept":"text/html,*/*"}
        if RATE_LIMIT_DELAY > 0:
            await asyncio.sleep(RATE_LIMIT_DELAY)
        
        r=await client.get(url, headers=h, timeout=6, follow_redirects=True)
        if r.status_code == 429:
            RATE_LIMIT_DELAY = min(RATE_LIMIT_DELAY + 1.0, 5.0)  # slow down
            return "429"
        
        # If successfully resolved request
        if r.status_code == 200:
            RATE_LIMIT_DELAY = max(RATE_LIMIT_DELAY - 0.2, 0.0)  # speed back up
            return r.text
        return ""
    except: return ""

async def search(client, q, engine="google"):
    # If engine is google but rate limit is high, fallback to duckduckgo/bing
    actual_engine = engine
    if RATE_LIMIT_DELAY > 2.0 and engine == "google":
        actual_engine = random.choice(["duckduckgo", "bing"])
        
    if actual_engine=="google":
        url=f"https://www.google.com/search?q={quote_plus(q)}&num=6&hl=en"
    elif actual_engine=="bing":
        url=f"https://www.bing.com/search?q={quote_plus(q)}&count=6"
    else:
        url=f"https://html.duckduckgo.com/html/?q={quote_plus(q)}"
        
    html=await fetch(client, url)
    if html == "429" and actual_engine == "google":
        # Google blocked us, retry instantly with Bing
        url=f"https://www.bing.com/search?q={quote_plus(q)}&count=6"
        html=await fetch(client, url)
        
    await asyncio.sleep(random.uniform(0.3, 0.7))
    return html if html != "429" else ""

def extract_domain_from_html(html):
    links = re.findall(r'https?://[a-zA-Z0-9\-\.]+\.[a-zA-Z]{2,6}', html)
    exclude = {
        "google.com", "bing.com", "yahoo.com", "facebook.com", "twitter.com",
        "linkedin.com", "instagram.com", "youtube.com", "naukri.com", "glassdoor.com",
        "justdial.com", "indiamart.com", "ambitionbox.com", "shine.com", "monster.com",
        "wikipedia.org", "pinterest.com", "mapsofindia.com", "duckduckgo.com", "sulekha.com",
        "indeed.com", "cutshort.io", "instahyre.com", "timesjobs.com", "github.com",
        "w3.org", "schema.org", "w3schools.com", "cloudflare.com", "wix.com", "wixpress.com",
        "wordpress.org", "wordpress.com", "blogspot.com", "tumblr.com", "medium.com",
        "github.io", "gitlab.com", "bitbucket.org", "sourceforge.net", "npmtrends.com",
        "npmjs.com", "pypi.org", "packagist.org", "nuget.org", "maven.org",
    }
    for link in links:
        dom = domain_of(link)
        if dom:
            # Check if domain contains any excluded patterns
            if any(ex in dom for ex in exclude):
                continue
            # Strictly block schema/metadata/cloud/search domains
            if any(x in dom for x in ["schema", "xml", "namespace", "xmlns", "metadata", "w3", "live", "office", "microsoft", "onedrive", "sharepoint", "google", "drive", "dropbox", "box.com", "icloud"]):
                continue
            return dom
    return ""

# ── Per-Company Enrichment ────────────────────────────────────────────────────
async def enrich_one(client, name, city, web, dom):
    result = {"emails":[], "phone":"", "linkedin":""}
    nq = name.replace('&','and').replace(',','')

    # Dynamic website finder if domain is missing
    if not dom:
        html = await search(client, f'{nq} official website', "google")
        if html:
            dom = extract_domain_from_html(html)
            if dom:
                log(f"  [🌐] Website found dynamically: {dom} for {name}")

    queries = [
        (f'{nq} {city} HR email contact recruiter', "google"),
        (f'{nq} "@{dom}" OR "hr@" OR "careers@" {city}' if dom else f'{nq} HR email {city} site:naukri.com', "google"),
        (f'{nq} {city} phone number HR contact recruiter', "bing"),
    ]

    for q, eng in queries:
        html = await search(client, q, eng)
        if not html: continue
        for em in extract_emails(html, dom):
            if em not in result["emails"]: result["emails"].append(em)
        for ph in extract_phones(html):
            if not result["phone"]: result["phone"]=ph
        if not result["linkedin"]:
            result["linkedin"] = extract_linkedin(html)

    # Scrape website career page (fallback)
    if dom and len(result["emails"])<2:
        for path in ['/contact','/careers','/about']:
            html=await fetch(client, f"https://{dom}{path}")
            if html and html != "429":
                for em in extract_emails(html, dom):
                    if em not in result["emails"]: result["emails"].append(em)
                for ph in extract_phones(html):
                    if not result["phone"]: result["phone"]=ph
                if result["emails"]: break

    # HR email pattern generation (instant format generation)
    if dom and len(result["emails"])<3:
        for pfx in HR_PATTERNS:
            candidate=f"{pfx}@{dom}"
            if candidate not in result["emails"]:
                result["emails"].append(candidate)
            if len(result["emails"])>=6: break

    return result

# ── Parallel Worker ───────────────────────────────────────────────────────────
async def process_company(client, row_data):
    row, name, city, web, dom, cur_emails, cur_phone, cur_li = row_data
    async with SEMAPHORE:
        try:
            if len(cur_emails)>=3 and cur_phone and cur_li:
                log(f"  [-] Skipped (already done): {name}")
                return row, None
            result = await enrich_one(client, name, city, web, dom)
            found_e = len(result["emails"])
            found_p = "✅" if result["phone"] else "❌"
            found_li = "✅" if result["linkedin"] else "❌"
            log(f"  [+] Enriched: {name[:30]:30s} | E:{found_e} P:{found_p} LI:{found_li}")
            return row, result
        except Exception as e:
            log(f"  [x] Error on {name}: {e}")
            return row, None

# ── Main ──────────────────────────────────────────────────────────────────────
async def main():
    log("\n"+"="*60)
    log("  COMONK FAST ENRICH v2 — Rate-Limit Safe Parallel workers")
    log(f"  Started: {START.strftime('%H:%M:%S')}")
    log(f"  Target : {EXCEL}")
    log("="*60+"\n")

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    total = ws.max_row-1

    done_set = load_ckpt()
    log(f"  Checkpoint: {len(done_set)} already done")
    log(f"  Remaining: {total-len(done_set)}\n")

    prio = {"MNC":0,"IT":1,"AI/ML":2,"FinTech":3,"EdTech":4,"HealthTech":5,"Startup":6,"Pharma":7}
    rows_data = []
    for row in range(2, ws.max_row+1):
        name = str(ws.cell(row,C_COMP).value or "").strip()
        if not name or name in done_set: continue
        cat  = str(ws.cell(row,C_CAT).value or "")
        cat_p = min([v for k,v in prio.items() if k in cat], default=99)
        rows_data.append((cat_p, row, name))
    rows_data.sort(key=lambda x: x[0])

    e_add=p_add=li_add=processed=skipped=0
    BATCH=20

    limits=httpx.Limits(max_connections=10, max_keepalive_connections=8)
    async with httpx.AsyncClient(limits=limits, verify=False, timeout=8) as client:

        for batch_start in range(0, len(rows_data), BATCH):
            batch = rows_data[batch_start:batch_start+BATCH]

            tasks_input = []
            for _, row, name in batch:
                city = str(ws.cell(row,C_CITY).value or "Ahmedabad")
                web  = str(ws.cell(row,C_WEB).value or "")
                dom  = domain_of(web)
                cur_emails = [ws.cell(row,c).value for c in C_EMAILS if ws.cell(row,c).value]
                cur_phone  = ws.cell(row,C_PHONE).value
                cur_li     = ws.cell(row,C_LI).value
                tasks_input.append((row, name, city, web, dom, cur_emails, cur_phone, cur_li))

            results = await asyncio.gather(
                *[process_company(client, t) for t in tasks_input],
                return_exceptions=True
            )

            for (_, row, name), res in zip(batch, results):
                if isinstance(res, Exception) or res is None:
                    continue
                row_idx, data = res
                if data is None:
                    skipped+=1
                    done_set.add(name)
                    continue

                empty_cols=[c for c in C_EMAILS if not ws.cell(row_idx,c).value]
                for em, col in zip(data["emails"][:6], empty_cols):
                    ws.cell(row_idx,col).value=em
                    e_add+=1

                if data["phone"] and not ws.cell(row_idx,C_PHONE).value:
                    cp=clean_phone(data["phone"])
                    if cp:
                        ws.cell(row_idx,C_PHONE).value=cp
                        p_add+=1

                if data["linkedin"] and not ws.cell(row_idx,C_LI).value:
                    ws.cell(row_idx,C_LI).value=data["linkedin"]
                    li_add+=1

                ems=[ws.cell(row_idx,c).value for c in C_EMAILS if ws.cell(row_idx,c).value]
                ph=ws.cell(row_idx,C_PHONE).value
                web_v=ws.cell(row_idx,C_WEB).value
                li_v=ws.cell(row_idx,C_LI).value
                ws.cell(row_idx,C_PRIO).value=recalc_priority(ems,ph,web_v,li_v)
                if not ws.cell(row_idx,C_SRC).value:
                    ws.cell(row_idx,C_SRC).value="Fast Enrich v2"

                done_set.add(name)
                processed+=1

                cat=str(ws.cell(row_idx,C_CAT).value or "")
                found_e=len(data["emails"])
                found_p="✅" if data["phone"] else "❌"
                found_li="✅" if data["linkedin"] else "❌"
                log(f"  [{processed+len(done_set):4d}] {name[:40]:40s} | E:{found_e} P:{found_p} LI:{found_li} [{cat[:8]}]")

            wb.save(EXCEL)
            save_ckpt(done_set)
            elapsed=(datetime.now()-START).seconds//60
            total_done=processed+skipped
            pct=round(total_done/len(rows_data)*100,1) if rows_data else 0
            log(f"\n  --- Batch done | {total_done}/{len(rows_data)} ({pct}%) | +{e_add} emails | +{p_add} phones | +{li_add} LI | Delay: {RATE_LIMIT_DELAY:.1f}s | {elapsed}min ---\n")

    wb.save(EXCEL)
    save_ckpt(done_set)
    log(f"\n  ✅ FAST ENRICH COMPLETE!")

if __name__=="__main__":
    asyncio.run(main())
