import openpyxl, re, zipfile, time
from collections import Counter

F = "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
def stable_load(p, tries=40):
    for _ in range(tries):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, data_only=True)
        except Exception:
            time.sleep(2)
    raise RuntimeError("cannot get stable read")
wb = stable_load(F)
print("="*60)
print("  ANALYSIS:", F)
print("="*60)
print("Sheets:", wb.sheetnames)
print()

def find_header_row(ws, maxscan=5):
    for r in range(1, maxscan+1):
        vals = [str(ws.cell(r,c).value or "").strip().lower() for c in range(1, min(ws.max_column,30)+1)]
        if "company name" in vals or "company" in vals:
            return r
    return 1

for sn in wb.sheetnames:
    ws = wb[sn]
    hr = find_header_row(ws)
    hdr = [ws.cell(hr,c).value for c in range(1, ws.max_column+1)]
    hm = {str(h).strip().lower(): i+1 for i,h in enumerate(hdr) if h}
    print(f"----- {sn} -----")
    print(f"  rows total: {ws.max_row}  | header row: {hr}  | data rows: {ws.max_row-hr}")
    print(f"  columns: {[h for h in hdr if h]}")
    CO = hm.get("company name") or hm.get("company")
    if not CO:
        print("  (no company column detected)\n"); continue
    EM = [hm.get(f"email {i}") for i in range(1,7)] + [hm.get(f"hr email {i}") for i in range(1,16)]
    EM = [c for c in EM if c]
    PH = hm.get("phone") or hm.get("office phone")
    WEB = hm.get("website")
    tot=0; wem=0; wph=0; wweb=0; numeric=0; blanks=0
    email_total=0
    for r in range(hr+1, ws.max_row+1):
        nm = ws.cell(r, CO).value
        if nm is None or str(nm).strip()=="":
            blanks+=1; continue
        tot+=1
        s=str(nm).strip()
        if re.fullmatch(r'[0-9]+', s): numeric+=1
        ec = sum(1 for c in EM if ws.cell(r,c).value and "@" in str(ws.cell(r,c).value))
        email_total += ec
        if ec: wem+=1
        if PH and ws.cell(r,PH).value: wph+=1
        if WEB and ws.cell(r,WEB).value: wweb+=1
    print(f"  companies (non-blank): {tot}")
    if blanks: print(f"  blank rows: {blanks}")
    print(f"  with email: {wem} ({wem*100//tot if tot else 0}%)  | total email cells: {email_total}")
    if PH: print(f"  with phone: {wph} ({wph*100//tot if tot else 0}%)")
    if WEB: print(f"  with website: {wweb} ({wweb*100//tot if tot else 0}%)")
    if numeric: print(f"  numeric-junk names: {numeric}")
    print()
