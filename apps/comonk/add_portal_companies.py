"""
Append companies found from job portals:
Cutshort, Instahyre, Indeed, LinkedIn Jobs, Naukri, Glassdoor,
Jooble, Y Combinator India AI startups
"""

import openpyxl, re, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment

# ── Companies from Job Portals ────────────────────────────────────────────────
# (Company, Category, AI_Roles, Website, Source_Portal)
PORTAL_COMPANIES = [
    # ── Cutshort ─────────────────────────────────────────────────────────
    ("E2M Solutions",             "IT Services","AI/ML Engineer, Data Science, Web Dev",           "e2msolutions.com",       "Cutshort"),
    ("Tops Infosolutions",        "IT Services","AI Integration, Web Dev, ERP",                   "topsinfosolutions.com",  "Cutshort"),
    ("Appsrow Solutions",         "IT Services","AI, Mobile App, React Dev",                      "appsrow.com",            "Cutshort"),
    ("Ascendum Solutions",        "AI / ML",    "AI/ML Engineer, Data Engineer, Cloud AI",         "ascendum.com",           "Cutshort"),
    ("Techwize",                  "AI / ML",    "AI/ML Engineer, Python, Deep Learning",           "techwize.in",            "Cutshort"),
    ("Mangalam Information Technologies","IT Services","AI, Software Dev, Enterprise Tech",        "mangalamgroup.com",      "Cutshort"),
    ("StreamSpace Artificial Intelligence","AI / ML","Generative AI, LLM, AI Streaming",          "streamspaceai.com",      "Cutshort"),
    ("MobileFirst Applications",  "IT Services","AI, Mobile App, Flutter Dev",                    "mobilefirst.in",         "Cutshort"),
    ("Azine Technologies",        "AI / ML",    "OCR AI, Face Recognition, Identity AI, ML",      "azinetech.com",          "Cutshort"),
    ("Emxcel Travel Solutions",   "AI / ML",    "AI/ML for Travel Tech, NLP, Data Science",       "emxcel.com",             "Cutshort"),
    ("Acquire",                   "IT Services","AI Integration, CX Tech, Automation",            "acquire.io",             "Cutshort"),
    # ── Instahyre ────────────────────────────────────────────────────────
    ("Roadzen",                   "AI / ML",    "Deep Learning, Computer Vision, InsurTech AI",   "roadzen.io",             "Instahyre"),
    ("Kanerika Software",         "AI / ML",    "Data Engineering, MLOps, Databricks, Snowflake", "kanerika.com",           "Instahyre"),
    ("Quanteon Solutions",        "AI / ML",    "AI/ML Engineer, Deep Learning",                  "quanteon.in",            "Instahyre"),
    # ── Indeed India ─────────────────────────────────────────────────────
    ("Hk Infosoft",               "IT Services","AI Integration, Mobile App Dev",                 "hkinfosoft.com",         "Indeed"),
    ("Kody Technolab",            "IT Services","AI, Mobile App, Flutter, Kotlin Dev",            "kodytechnolab.com",      "Indeed"),
    ("Gateway Group",             "IT Services","AI, Web Dev, Enterprise Solutions",              "gatewaygroup.co.in",     "Indeed"),
    ("AIVantage Analytics",       "AI / ML",    "AI Analytics, Business Intelligence, ML",        "aivantage.in",           "Indeed"),
    ("Shine Infosoft",            "IT Services","AI Integration, Mobile, Web Dev",                "shineinfosoft.com",      "Indeed"),
    ("Red White Skill Education", "AI / ML",    "AI/ML Training, EdTech AI",                     "redwhiteedu.com",        "Indeed"),
    ("Innvonix Tech Solutions",   "IT Services","AI, Mobile, Custom Software Dev",                "innvonix.com",           "Indeed"),
    # ── LinkedIn Jobs ─────────────────────────────────────────────────────
    ("Analytix Business Solutions","AI / ML",   "AI Analytics, Data Science, ML, BI",             "analytixbs.com",         "LinkedIn"),
    ("atQor",                     "AI / ML",    "AI/ML Engineer, Azure AI, Power Platform",       "atqor.com",              "LinkedIn"),
    ("AI CERTs India",            "AI / ML",    "AI Certification, AI Education, ML Training",    "aicerts.ai",             "LinkedIn"),
    ("DRC Systems",               "AI / ML",    "AI/ML Engineer, Data Science, GenAI",            "drcsystems.com",         "LinkedIn"),
    ("Awaaz AI",                  "AI / ML",    "Voice AI, Vernacular AI, NLP, Conversational AI","awaaz.ai",               "LinkedIn"),
    ("Radix Analytics",           "AI / ML",    "AI Analytics, Data Science, ML Engineering",     "radixanalytics.in",      "LinkedIn"),
    ("Tatvic",                    "AI / ML",    "Marketing AI, Analytics, Google AI Solutions",   "tatvic.com",             "LinkedIn"),
    ("Infineon Technologies",     "AI / ML",    "Automotive AI, Embedded AI, Semiconductor AI",   "infineon.com",           "LinkedIn"),
    ("The Infowarehouse",         "AI / ML",    "Data Warehouse + AI, BI, ML Engineering",        "infowarehouse.in",       "LinkedIn"),
    ("eInfochips (Arrow Company)","AI / ML",    "AI/ML, Embedded AI, IoT AI, Semiconductor AI",   "einfochips.com",         "LinkedIn"),
    ("iGauri Solutions",          "IT Services","AI, SAP, Mobile, Custom Dev",                    "igaurisolutions.com",    "LinkedIn"),
    ("Armakuni",                  "AI / ML",    "Platform Engineering, AI/ML Ops, DevOps + AI",   "armakuni.com",           "LinkedIn"),
    ("Urbanfix",                  "AI / ML",    "PropTech AI, Real Estate AI",                    "urbanfix.in",            "LinkedIn"),
    ("AIMLEAP",                   "AI / ML",    "AI/ML Dev, NLP, Data Science, Automation",       "aimleap.com",            "LinkedIn"),
    ("Shaip (Healthly.AI)",       "AI / ML",    "Healthcare AI Data, ML Training Data, NLP",      "shaip.com",              "LinkedIn"),
    ("Infiria AI",                "AI / ML",    "Generative AI, LLM, AI Product Dev",             "infiria.ai",             "LinkedIn"),
    ("Loopio",                    "IT Services","AI for RFP, Sales AI, ML Automation",            "loopio.com",             "LinkedIn"),
    ("SmartBear",                 "IT Services","AI Testing, QA AI, DevOps AI",                  "smartbear.com",          "LinkedIn"),
    # ── Naukri / Glassdoor / Jooble ─────────────────────────────────────
    ("Mobio Solutions",           "AI / ML",    "AI/ML Engineer, Mobile AI, Data Science",        "mobio.solutions",        "Naukri"),
    ("Digiwagon Technologies",    "IT Services","AI, Digital Marketing, Web Dev",                 "digiwagon.com",          "Naukri"),
    ("Tuvoc Technologies",        "AI / ML",    "AI/ML Engineer, Web Dev, Mobile",                "tuvoc.com",              "Naukri"),
    ("CEPT Research Dev Foundation","AI / ML",  "AI Research, Urban AI, IoT, Smart City",         "crdf.in",                "Naukri"),
    ("Emgage",                    "AI / ML",    "HR AI, Recruitment AI, Workforce AI",            "emgage.com",             "Naukri"),
    ("Wappnet Systems",           "IT Services","AI, Mobile App, Laravel, Flutter Dev",           "wappnet.com",            "Naukri"),
    ("QX Global Group",           "IT Services","AI, Finance AI, BPO + AI, Outsourcing",         "qxglobalgroup.com",      "Naukri"),
    ("RHC India",                 "IT Services","AI Staffing, IT Consulting, HR AI",              "rhcindia.com",           "Naukri"),
    ("Sparks to Ideas",           "IT Services","AI, Mobile App, UI/UX Dev",                     "sparkstoideas.com",      "Naukri"),
    ("Solvix Technologies",       "AI / ML",    "AI/ML Solutions, Data Science, Cloud AI",        "solvix.in",              "Naukri"),
    ("Accumn",                    "AI / ML",    "AI Analytics, FinTech AI, Data Science",         "accumn.in",              "Naukri"),
    ("Techsara Solutions",        "IT Services","AI Integration, Mobile, Web Dev",                "techsara.in",            "Naukri"),
    ("Vrinsoft Technology",       "IT Services","AI, Mobile App, Web Dev",                       "vrinsofttechnology.com", "Naukri"),
    # ── Glassdoor / Gandhinagar ──────────────────────────────────────────
    ("Prompt Softech",            "AI / ML",    "AI/ML Engineer, Data Scientist, NLP",            "promptsoftech.com",      "Glassdoor"),
    ("Spearline",                 "AI / ML",    "Telecom AI, Voice AI, ML Testing",               "spearline.com",          "Glassdoor"),
    ("infoanalytica",             "AI / ML",    "Data Science, AI Analytics, BI",                 "infoanalytica.com",      "Glassdoor"),
    ("AGS Health",                "AI / ML",    "Healthcare AI, Medical Coding AI, RCM AI",       "agshealth.com",          "Glassdoor"),
    ("Smart Working",             "IT Services","AI Staffing, Remote AI, HR Tech",                "smartworking.in",        "Glassdoor"),
    ("Samyak Infotech",           "AI / ML",    "AI/ML, Data Science, Computer Vision",           "samyakinfotech.com",     "Glassdoor"),
    # ── Y Combinator India AI startups ──────────────────────────────────
    ("Bolna AI",                  "AI / ML",    "Voice AI Agents, LLM, Conversational AI",        "bolna.dev",              "YCombinator"),
    ("Peoplebox.ai",              "AI / ML",    "HR AI, Performance AI, OKR AI Platform",         "peoplebox.ai",           "YCombinator"),
    ("BusinessOnBot",             "AI / ML",    "WhatsApp AI, Commerce AI, Chatbot",              "businessonbot.com",      "YCombinator"),
    ("Vahan",                     "AI / ML",    "Blue Collar AI Hiring, HR AI, Recruitment AI",   "vahan.co",               "YCombinator"),
    ("Helloverify",               "AI / ML",    "Background Verification AI, HR AI",              "helloverify.com",        "YCombinator"),
    ("WotNot",                    "AI / ML",    "AI Chatbot Platform, Conversational AI, NLP",    "wotnot.io",              "YCombinator"),
    # ── Additional from web research ─────────────────────────────────────
    ("Hireaidevelopers.io",       "AI / ML",    "AI Developer Hiring, LLM Dev, GenAI",            "hireaidevelopers.io",    "TopDev"),
    ("Whatmaction",               "AI / ML",    "AI/ML Solutions, Automation",                    "whatmaction.com",        "TopDev"),
    ("Innovatics",                "AI / ML",    "AI Analytics, Data Science, ML Engineering",     "innovatics.ai",          "TopDev"),
    ("Pennine Technolab",         "IT Services","AI Integration, Mobile App Dev",                 "pennineindia.com",       "TopDev"),
    ("Rao Information Technology","IT Services","AI, Web Dev, ERP, Custom Dev",                   "raoit.com",              "TopDev"),
    ("GrapesTech Solutions",      "IT Services","AI, Web Dev, Mobile App",                       "grapestechsolutions.com", "TopDev"),
    ("Tech Mitraa",               "AI / ML",    "AI/ML Solutions, Python, Deep Learning",         "techmitraa.com",         "TopDev"),
    ("Verticle Global Tech",      "IT Services","AI Integration, Cloud Dev, Custom Software",     "verticleglobal.com",     "TopDev"),
    ("SNDK Corp",                 "AI / ML",    "AI/ML, Web Dev, Mobile Dev",                    "sndkcorp.com",           "Glassdoor"),
    ("Mobiquity",                 "AI / ML",    "Data Science, AI Analytics, Cloud AI",           "mobiquity.com",          "Glassdoor"),
]

def norm(s):
    if not s: return ""
    s = unicodedata.normalize('NFKD', str(s).lower())
    return re.sub(r'[^a-z0-9]', '', s)

def main():
    wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
    ws = wb["COMPLETE_MASTER"]

    existing = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 2).value
        if n: existing[norm(n)] = r

    fill_new = PatternFill("solid", fgColor="FFF9E6")  # light yellow for portal-found
    added = updated = 0

    for company, category, roles, website_hint, portal in PORTAL_COMPANIES:
        key = norm(company)
        if not key: continue
        website = ("https://www." + website_hint) if not website_hint.startswith("http") else website_hint

        if key in existing:
            row = existing[key]
            if not ws.cell(row, 4).value: ws.cell(row, 4).value = category
            if not ws.cell(row, 5).value: ws.cell(row, 5).value = roles
            if not ws.cell(row, 13).value and website: ws.cell(row, 13).value = website
            cur = ws.cell(row, 18).value or ""
            if portal not in cur:
                ws.cell(row, 18).value = (cur + ", " + portal).strip(", ")
            updated += 1
        else:
            r = ws.max_row + 1
            ws.cell(r, 1, r - 1)
            ws.cell(r, 2, company)
            ws.cell(r, 3, "Ahmedabad")
            ws.cell(r, 4, category)
            ws.cell(r, 5, roles)
            ws.cell(r, 13, website)
            ws.cell(r, 18, portal)
            for c in range(1, 19):
                ws.cell(r, c).fill = fill_new
                ws.cell(r, c).alignment = Alignment(vertical="center")
            ws.cell(r, 2).font = Font(bold=True, color="7D4E00")
            ws.row_dimensions[r].height = 18
            existing[key] = r
            added += 1

    # Renumber
    for i, r in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(r, 1, i)

    total = ws.max_row - 1

    # Rebuild AI_ML_ONLY
    ws2 = wb["AI_ML_ONLY"]
    for r in range(ws2.max_row, 1, -1):
        ws2.delete_rows(r)

    ai_kw = ['ai','ml','data','machine','deep','nlp','computer vision','llm','genai',
             'analytics','automation','intelligence','learning']
    ai_row = 2
    fill_a = PatternFill("solid", fgColor="F5F0FF")
    fill_b = PatternFill("solid", fgColor="FFFFFF")

    for r in range(2, ws.max_row + 1):
        cat  = str(ws.cell(r, 4).value or "").lower()
        role = str(ws.cell(r, 5).value or "").lower()
        if any(k in cat or k in role for k in ai_kw):
            fill = fill_a if ai_row % 2 == 0 else fill_b
            for c in range(1, 19):
                v = ws.cell(r, c).value
                cell2 = ws2.cell(ai_row, c, v)
                cell2.fill = fill
                cell2.alignment = Alignment(vertical="center")
                if c in (6,7,8,9,10,11) and v:
                    cell2.font = Font(color="1a56db", size=9)
            ws2.cell(ai_row, 1, ai_row - 1)
            ws2.row_dimensions[ai_row].height = 18
            ai_row += 1

    ai_count = ws2.max_row - 1

    wb["STATS"].cell(3, 2, total)
    wb["STATS"].cell(6, 2, ai_count)

    wb.save("COMONK_TRUE_MASTER.xlsx")
    print(f"  NEW from portals: +{added}")
    print(f"  Existing updated: {updated}")
    print(f"  Total companies:  {total}")
    print(f"  AI/ML companies:  {ai_count}")
    print(f"  Saved: COMONK_TRUE_MASTER.xlsx")

if __name__ == "__main__":
    main()
