import openpyxl, zipfile, time

F="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
def stable_load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read")

wb=stable_load(F)
print("="*66)
print("  COVERAGE REPORT — Ahmedabad_IT_AIML_FINAL_MASTER.xlsx")
print("="*66)
grand={}
for SHEET in wb.sheetnames:
    if SHEET=="LinkedIn Profiles": continue
    ws=wb[SHEET]
    hrow=None
    for rr in (1,2,3):
        vals=[str(ws.cell(rr,c).value or "").strip().lower() for c in range(1,25)]
        if "company name" in vals or "company" in vals: hrow=rr; break
    if not hrow: continue
    hdr=[ws.cell(hrow,c).value for c in range(1,ws.max_column+1)]
    hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
    CO=hm.get("company name") or hm.get("company")
    PH=hm.get("phone") or hm.get("office phone")
    WEB=hm.get("website")
    ADDR=hm.get("address")
    EM=[hm.get(f"email {i}") for i in range(1,7)]+[hm.get(f"hr email {i}") for i in range(1,16)]
    EM=[c for c in EM if c]
    tot=wem=wph=wweb=wad=wall=w3=0
    total_emails=0
    for row in ws.iter_rows(min_row=hrow+1, values_only=True):
        def g(c): return row[c-1] if c and len(row)>=c else None
        co=g(CO)
        if co is None or str(co).strip()=="": continue
        tot+=1
        ec=sum(1 for c in EM if g(c) and "@" in str(g(c)))
        total_emails+=ec
        e=ec>0; p=bool(g(PH)); w=bool(g(WEB)); a=bool(ADDR and g(ADDR))
        if e: wem+=1
        if p: wph+=1
        if w: wweb+=1
        if a: wad+=1
        if e and p: w3+=1                      # has email AND phone (mailable+callable)
        if e and p and w and a: wall+=1        # complete: all four
    def pct(n): return f"{n*100//tot}%" if tot else "-"
    print(f"\n----- {SHEET} ({tot} companies) -----")
    print(f"  📧 with Email    : {wem:>5}  ({pct(wem)})   [{total_emails} email addresses total]")
    print(f"  📞 with Phone    : {wph:>5}  ({pct(wph)})")
    print(f"  🌐 with Website  : {wweb:>5}  ({pct(wweb)})")
    if ADDR: print(f"  📍 with Address  : {wad:>5}  ({pct(wad)})")
    print(f"  ✅ Email + Phone : {w3:>5}  ({pct(w3)})")
    print(f"  ⭐ ALL 4 complete: {wall:>5}  ({pct(wall)})")
    grand[SHEET]=(tot,wem,wph,wweb,wad,wall,total_emails)

print("\n"+"="*66)
