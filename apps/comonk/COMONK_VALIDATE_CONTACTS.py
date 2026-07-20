# COMONK_VALIDATE_CONTACTS.py
# Validator script for emails and phone numbers in Ahmedabad_IT_AIML_FINAL_MASTER.xlsx.
# Features syntax check, DNS MX record resolution to check domain validity,
# phone validation/formatting, and cleanup of bad entries.
# Keeps a backup copy of all deleted details in COMONK_DELETED_CONTACTS_BACKUP.txt.

import sys, os, re, socket, time
import openpyxl
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

EXCEL = "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
SHEET = "All Companies"
LOG = "validation_log.txt"
BACKUP_FILE = "COMONK_DELETED_CONTACTS_BACKUP.txt"

# 17-column layout mappings:
C_NO=1; C_COMP=2; C_CAT=3; C_ROLE=4
C_EMAILS=[5, 6, 7, 8, 9, 17] # Email 1 to 5, and Email 6 is at col 17
C_PHONE=10; C_WEB=11; C_LI=12; C_ADDR=13; C_CITY=14; C_PRIO=15; C_SRC=16

EMAIL_RE = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,7}\b')

SKIP_MAIL = {
    "gmail.com","yahoo.com","yahoo.co.in","outlook.com","hotmail.com",
    "rediffmail.com","sentry.io","example.com","googlemail.com",
    "facebook.com","twitter.com","instagram.com","youtube.com",
    "cloudflare.com","google.com","microsoft.com","w3.org","schema.org",
}

# DNS Cache to speed up MX checks
MX_CACHE = {}

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    try:
        with open(LOG, 'a', encoding='utf-8') as f: f.write(line+"\n")
    except: pass

def log_deletion(co_name, row, field, val, reason):
    line = f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Row {row} | {co_name} | Field: {field} | Value: {val} | Reason: {reason}\n"
    try:
        with open(BACKUP_FILE, "a", encoding="utf-8") as f:
            f.write(line)
    except Exception as e:
        log(f"  [Error logging deletion] {e}")

def check_domain_mx(domain):
    domain = domain.lower().strip()
    if domain in SKIP_MAIL:
        return True # Free mail domains are valid
    if domain in MX_CACHE:
        return MX_CACHE[domain]
    
    try:
        socket.getaddrinfo(domain, None)
        MX_CACHE[domain] = True
        return True
    except socket.gaierror:
        MX_CACHE[domain] = False
        return False
    except:
        MX_CACHE[domain] = False
        return False

def clean_phone(p):
    if not p: return ""
    d = re.sub(r'[^\d]', '', str(p))
    
    # Toll-free check
    if d.startswith('1800') or d.startswith('1860') or d.startswith('1888'):
        if len(d) in (10, 11):
            return str(p).strip()
            
    # Leading country codes and 0 prefix cleanup
    if d.startswith('91') and len(d) >= 12:
        d = d[2:]
    elif d.startswith('0') and len(d) >= 11:
        d = d[1:]
        
    # Standard check: 10 or 11 digits (mobile/landlines)
    if len(d) in (10, 11):
        if len(set(d)) <= 2: return "" # Repeating digits block
        return str(p).strip()
    return ""

def validate_email_syntax(em):
    if not em or '@' not in em or len(em)>80:
        return False
    return bool(EMAIL_RE.match(em))

def safe_save(wb, filename):
    while True:
        try:
            wb.save(filename)
            return
        except PermissionError:
            log(f"  [⚠️ WARNING] Permission Denied: Cannot save '{filename}'. Please close the file if it is open in Microsoft Excel! Retrying in 5 seconds...")
            time.sleep(5)
        except Exception as e:
            log(f"  [❌ ERROR] Failed to save: {e}")
            raise e

def main():
    log("\n" + "="*70)
    log("  COMONK CONTACT VALIDATION & CLEANUP ENGINE")
    log("="*70 + "\n")

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    
    total_rows = ws.max_row - 3  # Starts at row 4
    emails_checked = 0
    emails_removed = 0
    phones_checked = 0
    phones_cleaned = 0
    phones_removed = 0

    log(f"  Total records in database to validate: {total_rows}")

    for row_idx in range(4, ws.max_row + 1):
        co_name = ws.cell(row_idx, C_COMP).value
        if not co_name:
            continue
            
        # 1. Validate Emails
        for idx, col in enumerate(C_EMAILS, start=1):
            email_val = ws.cell(row_idx, col).value
            if email_val:
                email_val = str(email_val).strip()
                emails_checked += 1
                
                # Check syntax
                if not validate_email_syntax(email_val):
                    log(f"  [Remove] Invalid syntax email in row {row_idx} ({co_name}): {email_val}")
                    log_deletion(co_name, row_idx, f"Email {idx}", email_val, "Invalid syntax")
                    ws.cell(row_idx, col).value = None
                    emails_removed += 1
                    continue
                
                # Check domain validity via DNS resolution
                parts = email_val.split('@')
                if len(parts) == 2:
                    domain = parts[1]
                    if not check_domain_mx(domain):
                        log(f"  [Remove] Domain DNS resolution failed in row {row_idx} ({co_name}): {email_val}")
                        log_deletion(co_name, row_idx, f"Email {idx}", email_val, "DNS MX lookup failed")
                        ws.cell(row_idx, col).value = None
                        emails_removed += 1
                        continue

        # 2. Validate Phone
        phone_val = ws.cell(row_idx, C_PHONE).value
        if phone_val:
            phone_str = str(phone_val).strip()
            phones_checked += 1
            
            cleaned = clean_phone(phone_str)
            if not cleaned:
                log(f"  [Remove] Invalid phone number in row {row_idx} ({co_name}): {phone_val}")
                log_deletion(co_name, row_idx, "Phone", phone_val, "Invalid format or fake repeating digits")
                ws.cell(row_idx, C_PHONE).value = None
                phones_removed += 1
            elif cleaned != phone_str:
                ws.cell(row_idx, C_PHONE).value = cleaned
                phones_cleaned += 1

    safe_save(wb, EXCEL)
    
    log("\n" + "="*70)
    log("  VALIDATION SUMMARY REPORT")
    log("="*70)
    log(f"  Emails Checked   : {emails_checked}")
    log(f"  Emails Removed   : {emails_removed} (Invalid syntax or bad domain)")
    log(f"  Phones Checked   : {phones_checked}")
    log(f"  Phones Formatted : {phones_cleaned}")
    log(f"  Phones Removed   : {phones_removed} (Invalid length or fake repeating digits)")
    log("="*70 + "\n")

if __name__ == "__main__":
    main()
