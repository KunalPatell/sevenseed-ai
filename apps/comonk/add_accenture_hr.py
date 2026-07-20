"""Add Accenture_HR sheet with recruiter contacts found via LinkedIn research."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

MASTER="COMONK_TRUE_MASTER.xlsx"

# (Name, Title, LinkedIn URL, guessed email pattern)
RECRUITERS = [
    ("Parul Narayan",      "HR Talent Acquisition",        "https://in.linkedin.com/in/parul-narayan-b309a2121",   "parul.narayan@accenture.com"),
    ("Sameer Joshi",       "India Talent Acquisition Lead","https://in.linkedin.com/in/sameer-joshi-951754b4",     "sameer.joshi@accenture.com"),
    ("Neha Sharma",        "Talent Acquisition Manager",   "https://in.linkedin.com/in/neha-sharma-2aab90302",     "neha.sharma@accenture.com"),
    ("Pushpinder Singh",   "Talent Acquisition Lead",      "https://in.linkedin.com/in/pushpinder-singh-22b46312", "pushpinder.singh@accenture.com"),
    ("Sushanth N",         "Talent Acquisition Lead",      "https://in.linkedin.com/in/sushanth-n-5187521b9",      "sushanth.n@accenture.com"),
    ("Girish Sharma",      "Director - Human Resources",   "https://in.linkedin.com/in/girish-sharma-6473815",     "girish.sharma@accenture.com"),
    ("Ranjitha Basapally", "HR Recruiter",                 "https://in.linkedin.com/in/ranjitha-basapally-5984b019","ranjitha.basapally@accenture.com"),
    ("Gayathri Sivakumar", "HR Recruiter",                 "https://in.linkedin.com/in/gayathri-sivakumar-953291131","gayathri.sivakumar@accenture.com"),
    ("Abirami Ravichandran","HR Recruiter",                "https://in.linkedin.com/in/abirami-ravichandran-105361265","abirami.ravichandran@accenture.com"),
    ("Ashwini K",          "HR People Advisor",            "https://www.linkedin.com/in/ashwini-k-7a9368b6",       "ashwini.k@accenture.com"),
]

OFFICE = "West Wing, 13th Floor, Venus Stratum, Venus Grounds, Surendra Mangaldas Rd, Nr. Jhansi Ki Rani, Nehrunagar, Satellite, Ahmedabad 380015"
CAREERS = "https://www.accenture.com/in-en/careers/jobsearch"

wb=openpyxl.load_workbook(MASTER)
if "Accenture_HR" in wb.sheetnames: del wb["Accenture_HR"]
ws=wb.create_sheet("Accenture_HR")

# Title rows
ws.merge_cells("A1:F1")
t=ws.cell(1,1,"ACCENTURE — HR / Talent Acquisition Contacts (Ahmedabad / India)")
t.font=Font(bold=True,size=13,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="A100FF")
t.alignment=Alignment(horizontal="center",vertical="center")
ws.row_dimensions[1].height=26
ws.merge_cells("A2:F2")
o=ws.cell(2,1,f"Office: {OFFICE}   |   Apply: {CAREERS}   |   Email pattern: firstname.lastname@accenture.com")
o.font=Font(size=9,italic=True,color="333333"); o.alignment=Alignment(horizontal="center")

# Header
hdr=["#","Name","Title","LinkedIn Profile","Likely Email (verify)","Connect Note"]
W=[4,24,30,52,34,40]
for c,(h,w) in enumerate(zip(hdr,W),1):
    cell=ws.cell(3,c,h); cell.fill=PatternFill("solid",fgColor="2E1A47")
    cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.column_dimensions[chr(64+c)].width=w
ws.freeze_panes="A4"

note="Hi {first}, I'm an AI/ML engineer keen on Accenture Ahmedabad. Could you guide me on open roles / referral? Resume attached. Thanks!"
r=4
for i,(name,title,li,email) in enumerate(RECRUITERS,1):
    first=name.split()[0]
    ws.cell(r,1,i).alignment=Alignment(horizontal="center")
    ws.cell(r,2,name)
    ws.cell(r,3,title)
    c4=ws.cell(r,4,li); c4.font=Font(color="0A66C2",size=9,underline="single"); c4.hyperlink=li
    c5=ws.cell(r,5,email); c5.font=Font(color="1a56db",size=9)
    ws.cell(r,6,note.format(first=first)).font=Font(size=8,italic=True)
    if i%2==0:
        for c in range(1,7): ws.cell(r,c).fill=PatternFill("solid",fgColor="F5F0FF")
    ws.row_dimensions[r].height=16
    r+=1

# Move sheet to position 2 (after COMPLETE_MASTER)
wb.move_sheet("Accenture_HR", -(wb.sheetnames.index("Accenture_HR")-1))
wb.save(MASTER)
print(f"  Accenture_HR sheet created with {len(RECRUITERS)} recruiters")
print(f"  Sheets: {wb.sheetnames}")
