import openpyxl
wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
ws = wb["COMPLETE_MASTER"]

print("Headers:")
for c in range(1, 21):
    print(f"  col {c}: {ws.cell(1, c).value!r}")

print("\nSample rows (2-6):")
for r in range(2, 7):
    print(f"Row {r}:")
    for c in range(1, 6):
        print(f"    col{c}={ws.cell(r,c).value!r}")

# Check AI_ML_ONLY sheet directly
ws_ai = wb["AI_ML_ONLY"]
print(f"\nAI_ML_ONLY sheet rows: {ws_ai.max_row - 1}")
print("Sample from AI_ML_ONLY:")
for r in range(2, 5):
    print(f"  {ws_ai.cell(r,2).value} | cat={ws_ai.cell(r,4).value} | ai={ws_ai.cell(r,5).value}")
