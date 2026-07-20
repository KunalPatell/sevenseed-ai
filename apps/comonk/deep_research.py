"""
Async parallel phone scraper — 20 concurrent requests, finishes in ~3 min.
Strict Indian number validation only.
"""

import openpyxl, re, asyncio, time
import httpx
from html.parser import HTMLParser

EXCEL_PATH   = "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
CONCURRENCY  = 20   # simultaneous requests
TIMEOUT      = 6    # seconds per page

HEADERS_HTTP = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124",
    "Accept": "text/html",
}

PHONE_RE = re.compile(
    r'(?:'
    r'\+91[\s\-]?[6-9]\d{9}'           # +91 mobile
    r'|\+91[\s\-]?79[\s\-]?\d{4}[\s\-]?\d{4}'  # +91 79 Ahmedabad
    r'|079[\s\-]?\d{7,8}'              # 079 landline
    r'|\b[6-9]\d{9}\b'                 # 10-digit mobile
    r'|tel:[\+\d\s\-]{10,15}'         # tel: links
    r')'
)

class TextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.result = []; self._skip = False
    def handle_starttag(self, tag, attrs):
        if tag in ('script','style','noscript'): self._skip = True
    def handle_endtag(self, tag):
        if tag in ('script','style','noscript'): self._skip = False
    def handle_data(self, data):
        if not self._skip: self.result.append(data)
    def get_text(self): return ' '.join(self.result)

def extract_text(html):
    p = TextExtractor()
    try: p.feed(html)
    except: pass
    return p.get_text()

def is_valid_indian(num):
    digits = re.sub(r'[\s\-\(\)\+tel:]', '', num)
    if digits.startswith('91') and len(digits) == 12: digits = digits[2:]
    elif digits.startswith('0') and 10 <= len(digits) <= 11: digits = digits[1:]
    if len(digits) != 10: return False, None
    if digits[0] in '6789': return True, '+91 ' + digits[:5] + ' ' + digits[5:]
    if digits.startswith('79'): return True, '+91 79 ' + digits[2:6] + ' ' + digits[6:]
    return False, None

async def fetch_phones_async(client, base_url):
    pages = [base_url, base_url.rstrip('/') + '/contact',
             base_url.rstrip('/') + '/contact-us']
    for url in pages:
        try:
            r = await client.get(url, timeout=TIMEOUT, follow_redirects=True)
            if r.status_code != 200: continue
            text = extract_text(r.text)
            raw = PHONE_RE.findall(text)
            found = []
            for p in raw:
                ok, cleaned = is_valid_indian(p)
                if ok and cleaned not in found:
                    found.append(cleaned)
            if found:
                return found[0]
        except Exception:
            pass
    return None

async def process_batch(batch, client, sem, results):
    async def one(r, name, website):
        async with sem:
            phone = await fetch_phones_async(client, website)
            results[r] = (name, phone)

    tasks = [one(r, name, website) for r, name, website in batch]
    await asyncio.gather(*tasks)

async def main_async():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Collect missing
    missing = []
    for r in range(4, ws.max_row + 1):
        name    = ws.cell(r, 2).value
        cat     = str(ws.cell(r, 3).value or '').lower()
        website = ws.cell(r, 11).value
        phone   = ws.cell(r, 10).value
        if not name or not website or phone: continue
        prio = 0 if any(k in cat for k in ['ai','ml','data','machine','deep','nlp']) else 1
        missing.append((prio, r, name, website))

    missing.sort()
    total = len(missing)
    print(f"\n  {total} companies missing phones — running {CONCURRENCY} parallel workers\n")

    sem = asyncio.Semaphore(CONCURRENCY)
    results = {}
    updated = 0
    batch_size = 50

    async with httpx.AsyncClient(headers=HEADERS_HTTP) as client:
        for start in range(0, total, batch_size):
            chunk = [(r, name, web) for _, r, name, web in missing[start:start+batch_size]]
            batch_num = start // batch_size + 1
            total_batches = (total + batch_size - 1) // batch_size
            print(f"  Batch {batch_num}/{total_batches} ({len(chunk)} companies)...", flush=True)

            chunk_results = {}
            await process_batch(chunk, client, sem, chunk_results)

            for r, (name, phone) in chunk_results.items():
                short = name[:35] if name else ''
                if phone:
                    ws.cell(r, 10).value = phone
                    updated += 1
                    print(f"    OK   {short:<35} {phone}")
                else:
                    print(f"    --   {short}")

            wb.save(EXCEL_PATH)
            print(f"  >> Saved. Total phones added so far: {updated}\n")

    print(f"\n  DONE: {updated}/{total} phone numbers added.")
    print(f"  Saved to: {EXCEL_PATH}\n")

    # Final count
    total_co = phone_count = email_count = 0
    for r in range(4, ws.max_row+1):
        n = ws.cell(r,2).value
        if not n: continue
        total_co += 1
        if ws.cell(r,10).value: phone_count += 1
        if ws.cell(r,5).value: email_count += 1
    print(f"  === FINAL SHEET STATS ===")
    print(f"  Total companies : {total_co}")
    print(f"  Have email      : {email_count} ({round(email_count/total_co*100)}%)")
    print(f"  Have phone      : {phone_count} ({round(phone_count/total_co*100)}%)")

if __name__ == "__main__":
    asyncio.run(main_async())
