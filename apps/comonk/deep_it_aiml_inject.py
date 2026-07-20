# deep_it_aiml_inject.py - Deep IT/AI-ML injection script
import openpyxl
from datetime import datetime

COMPANIES_DATA = [
    {"name":"DRC Systems India","city":"Gandhinagar","category":"AI-ML","ai_roles":"AI Engineer, ML Engineer, Data Scientist","emails":["careers@drcsystems.com","info@drcsystems.com","ir@drcsystems.com"],"phone":"+91 79 6777 2222","website":"drcsystems.com","address":"24th Floor, GIFT Two Building, Block 56, Road 5C, GIFT CITY, Gandhinagar 382050","careers_url":"https://www.drcsystems.com/careers/"},
    {"name":"AddWeb Solution","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI Engineer, Python Developer","emails":["contact@addwebsolution.com","hr@addwebsolution.com","careers@addwebsolution.com"],"phone":"+91 90237 28518","website":"addwebsolution.com","address":"705, Silicon Tower, Law Garden, Ellisbridge, Ahmedabad 380009","careers_url":"https://addwebsolution.com/careers"},
    {"name":"eSparkBiz Technologies","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI/ML Engineer, LLM Developer","emails":["hr@esparkinfo.com","hiring@esparkinfo.com","career@esparkinfo.com","info@esparkinfo.com"],"phone":"+91 90237 28518","website":"esparkinfo.com","address":"Office No. 1201-1212, The Orion, Gota, Ahmedabad 382481","careers_url":"https://www.esparkinfo.com/careers"},
    {"name":"MindInventory","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI Developer, ML Engineer","emails":["career@mindinventory.com","hr@mindinventory.com","info@mindinventory.com"],"phone":"+91-951-229-3491","website":"mindinventory.com","address":"801, City Center 2, Science City Road, Sola, Ahmedabad 380060","careers_url":"https://www.mindinventory.com/careers.php"},
    {"name":"Innvonix Tech Solutions","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI Agent Developer, LLM Engineer","emails":["info@innvonix.com","hr@innvonix.com","careers@innvonix.com"],"phone":"+91 7096499910","website":"innvonix.com","address":"1105-1108, I Square Corporate Park, Science City Road, Sola, Ahmedabad 380060","careers_url":"https://www.innvonix.com/careers"},
    {"name":"Kody Technolab","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI/ML Lead, AI Engineer","emails":["hiring@kodytechnolab.com","info@kodytechnolab.com","sales@kodytechnolab.com"],"phone":"+91 77177 96599","website":"kodytechnolab.com","address":"2nd Floor, Block-J, Safal Mondeal Retail Park, S.G. Highway, Bodakdev, Ahmedabad 380054","careers_url":"https://kodytechnolab.com/careers/"},
    {"name":"Wappnet Systems","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI Developer, ML Engineer","emails":["hr@wappnet.com","career@wappnet.com","business@wappnet.com","info@wappnet.com"],"phone":"+91 91734 08554","website":"wappnet.com","address":"403, iSQUARE Corporate Park, Science City Road, Sola, Ahmedabad 380060","careers_url":"https://wappnet.com/career/"},
    {"name":"AIS Technolabs","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI Developer, ML Engineer","emails":["hr@aistechnolabs.com","biz@aistechnolabs.com","careers@aistechnolabs.com"],"phone":"+91 75750 06403","website":"aistechnolabs.com","address":"Ahmedabad, Gujarat, India","careers_url":"https://www.aistechnolabs.com/career"},
    {"name":"Rushkar Technology","city":"Ahmedabad","category":"IT Services","ai_roles":"Full Stack Developer, AI Engineer","emails":["hrteam@rushkar.com","management@rushkar.com","hr@rushkar.com"],"phone":"+91 96240 08889","website":"rushkar.com","address":"411, Sarthik Square, SG Highway, Ahmedabad","careers_url":"https://rushkar.com/career"},
    {"name":"Bacancy Technology","city":"Ahmedabad","category":"IT Services","ai_roles":"AI Developer, Full Stack Developer","emails":["solutions@bacancy.com","partnerships@bacancytechnology.com","hr@bacancytechnology.com"],"phone":"+91-079-40037674","website":"bacancytechnology.com","address":"15-16, Times Corporate Park, Thaltej, Ahmedabad 380059","careers_url":"https://www.bacancytechnology.com/careers"},
    {"name":"TatvaSoft","city":"Ahmedabad","category":"IT Services","ai_roles":"Software Developer, AI Engineer","emails":["info@tatvasoft.com","hr@tatvasoft.com","careers@tatvasoft.com"],"phone":"+91 960 142 1472","website":"tatvasoft.com","address":"TatvaSoft House, Rajpath Club Road, Off S.G. Road, Ahmedabad 380054","careers_url":"https://www.tatvasoft.com/career/"},
    {"name":"Simform Solutions","city":"Ahmedabad","category":"IT Services","ai_roles":"AI Developer, Cloud Engineer","emails":["hr@simform.com","careers@simform.com","info@simform.com"],"phone":"+1 321-237-2727","website":"simform.com","address":"Office 601-611, Binori B-Square-2, Ambali-Iskon Road, Ahmedabad 380054","careers_url":"https://www.simform.com/careers/"},
    {"name":"J.P. Morgan GIFT City","city":"Gandhinagar","category":"MNC - Finance","ai_roles":"Data Scientist, AI/ML Engineer","emails":["careers@jpmorgan.com","hr@jpmorgan.com"],"phone":"+91 79 6612 7000","website":"jpmorganchase.com","address":"GIFT City, Gandhinagar, Gujarat 382355","careers_url":"https://www.jpmorganchase.com/about/careers"},
    {"name":"HSBC GIFT City","city":"Gandhinagar","category":"MNC - Finance","ai_roles":"Data Analyst, AI Engineer","emails":["careers@hsbc.com","hr@hsbc.com","india.careers@hsbc.com"],"phone":"+91 79 6191 8000","website":"hsbc.com","address":"GIFT One Tower, GIFT City, Gandhinagar, Gujarat 382355","careers_url":"https://www.hsbc.com/careers/where-we-hire/india"},
    {"name":"Standard Chartered GIFT City","city":"Gandhinagar","category":"MNC - Finance","ai_roles":"Data Scientist, AI Engineer","emails":["careers@sc.com","hr@sc.com","india.careers@sc.com"],"phone":"+91 79 6141 4000","website":"sc.com","address":"GIFT City, Gandhinagar, Gujarat 382355","careers_url":"https://www.sc.com/en/about/careers/"},
    {"name":"Deutsche Bank GIFT City","city":"Gandhinagar","category":"MNC - Finance","ai_roles":"Data Analyst, ML Engineer","emails":["hr@db.com","careers@db.com","india.careers@db.com"],"phone":"+91 79 6141 5000","website":"db.com","address":"GIFT City, Gandhinagar, Gujarat 382355","careers_url":"https://careers.db.com/"},
    {"name":"Axis Bank GIFT City","city":"Gandhinagar","category":"MNC - Finance","ai_roles":"Data Scientist, AI Product Manager","emails":["hr@axisbank.com","careers@axisbank.com"],"phone":"+91 79 7194 5000","website":"axisbank.com","address":"GIFT City, Gandhinagar, Gujarat 382355","careers_url":"https://www.axisbank.com/careers"},
    {"name":"Barclays GIFT City","city":"Gandhinagar","category":"MNC - Finance","ai_roles":"Data Scientist, AI Analyst, Risk Quant","emails":["india.careers@barclays.com","hr@barclays.com"],"phone":"+91 79 6141 6000","website":"barclays.com","address":"GIFT City, Gandhinagar, Gujarat 382355","careers_url":"https://home.barclays/careers/"},
    {"name":"State Bank of India GIFT City","city":"Gandhinagar","category":"MNC - Finance","ai_roles":"Data Analyst, IT Officer","emails":["careers@sbi.co.in","hr@sbi.co.in"],"phone":"1800 11 2211","website":"sbi.co.in","address":"GIFT City Branch, Gandhinagar, Gujarat 382355","careers_url":"https://bank.sbi/careers"},
    {"name":"LTIMindtree","city":"Gandhinagar","category":"MNC - IT","ai_roles":"AI/ML Engineer, Data Scientist","emails":["careers@ltimindtree.com","hr@ltimindtree.com","talent@ltimindtree.com"],"phone":"+91 20 7106 8000","website":"ltimindtree.com","address":"GIFT City, Gandhinagar / Airoli, Navi Mumbai HQ","careers_url":"https://www.ltimindtree.com/careers/"},
    {"name":"Zoho Corporation","city":"Gandhinagar","category":"AI-ML","ai_roles":"AI/ML Engineer, SDE, Product Manager","emails":["careers@zohocorp.com","hr@zohocorp.com"],"phone":"+91 044 6965 1900","website":"zoho.com","address":"Zoho Corporation, GIFT City, Gandhinagar / Chennai HQ","careers_url":"https://careers.zohocorp.com/"},
    {"name":"Softvan","city":"Ahmedabad","category":"IT Services","ai_roles":"Software Developer, AI Engineer","emails":["Sales@softvanlabs.com","cs@softvan.in","hr@softvan.in"],"phone":"+91 6358922311","website":"softvan.in","address":"305, Sigma Legacy, Nr. Vikram Sarabhai Marg, Ambawadi, Ahmedabad 380015","careers_url":"https://softvan.in/career/"},
    {"name":"SOAR Technologies","city":"Ahmedabad","category":"IT Services","ai_roles":"Software Developer, QA Engineer","emails":["hr@soartechnologies.net","info@soartechnologies.net"],"phone":"+91-7948988475","website":"soartechnologies.net","address":"C-405B, Ganesh Meridian, SG Highway, Ahmedabad","careers_url":"https://soartechnologies.net/careers/"},
    {"name":"Vrinsofts Technology","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI/ML Developer, Full Stack Developer","emails":["hr@vrinsofts.com","info@vrinsofts.com","career@vrinsofts.com"],"phone":"+91 98256 37534","website":"vrinsofts.com","address":"Ahmedabad, Gujarat, India","careers_url":"https://vrinsofts.com/career/"},
    {"name":"Vrinsoft Technology","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI/ML Developer, App Developer","emails":["hr@vrinsofts.com","sales@vrinsofts.com"],"phone":"+91 75749 26643","website":"vrinsofts.com","address":"707, Elite Business Park, Shapath Hexa, S G Highway, Ahmedabad 380060","careers_url":"https://www.vrinsofts.com/career/"},
    {"name":"Brilworks","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI Engineer, Data Scientist, Data Engineer","emails":["hr@brilworks.com","sales@brilworks.com"],"phone":"+91 91068 10920","website":"brilworks.com","address":"503, Fortune Business Hub, Science City Road, Sola, Ahmedabad 380060","careers_url":"https://www.brilworks.com/careers/"},
    {"name":"Hidden Brains Infotech","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI/ML Developer, Full Stack Developer","emails":["hr@hiddenbrains.com","biz@hiddenbrains.com"],"phone":"+91-989-892-1444","website":"hiddenbrains.com","address":"301, Sachet-4, Near Prernatirth Derasar, Satellite, Ahmedabad 380015","careers_url":"https://www.hiddenbrains.com/career.html"},
    {"name":"OpenXcell Technolabs","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI Developer, ML Engineer, Mobile Developer","emails":["hr@openxcell.com","sales@openxcell.com","careers@openxcell.com"],"phone":"+91-999-822-2928","website":"openxcell.com","address":"202-203, Baleshwar Avenue, S.G. Highway, Bodakdev, Ahmedabad","careers_url":"https://www.openxcell.com/career/"},
    {"name":"Agile Infoways","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI/ML Engineer, Full Stack Developer","emails":["career@agileinfoways.com","inquiry@agileinfoways.com"],"phone":"+1 470-772-5053","website":"agileinfoways.com","address":"10th Floor, Ashridhar Athens, Shivranjani Cross Rd, Ahmedabad 380015","careers_url":"https://www.agileinfoways.com/career/"},
    {"name":"Uplers","city":"Ahmedabad","category":"IT Services","ai_roles":"Digital Marketer, Developer, AI Engineer","emails":["hello@uplers.com","hr@uplers.com"],"phone":"+91 7940324566","website":"uplers.com","address":"Ahmedabad, Gujarat, India","careers_url":"https://platform.uplers.com"},
    {"name":"Groovy Web","city":"Ahmedabad","category":"IT Services","ai_roles":"Full Stack Developer, AI Developer","emails":["career@groovyweb.co","hr@groovyweb.co"],"phone":"+91 903 357 8483","website":"groovyweb.co","address":"C/513, Siddhi Vinayak Towers, Off S.G. Road, Makarba, Ahmedabad","careers_url":"https://www.groovyweb.co/careers/"},
    {"name":"Citibank GIFT City","city":"Gandhinagar","category":"MNC - Finance","ai_roles":"Data Analyst, AI Engineer, Financial Analyst","emails":["ifscbranch@citi.com","hr@citi.com","careers@citi.com"],"phone":"+91 79 6141 7000","website":"citi.com","address":"GIFT SEZ, GIFT City, Gandhinagar, Gujarat 382355","careers_url":"https://jobs.citi.com/"},
    {"name":"Infy_REC_Helpdesk","city":"Gandhinagar","category":"MNC - IT","ai_roles":"AI/ML Engineer, Data Scientist","emails":["Infy_REC_Helpdesk@infosys.com","careers@infosys.com"],"phone":"+91 80 41179999","website":"infosys.com","address":"Pragya II, Block 15-C1, GIFT SEZ, Gandhinagar 382355","careers_url":"https://www.infosys.com/careers/apply.html"},
    {"name":"Peerbits","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI Engineer, Healthcare Tech Developer","emails":["career@peerbits.com","info@peerbits.com","enquiry@peerbits.com"],"phone":"+91 79 48000686","website":"peerbits.com","address":"409-419, 4th Floor, Solitaire Connect, Nr. Gallops Motors, Makarba, Ahmedabad 380015","careers_url":"https://peerbits.com/career/"},
    {"name":"Radixweb","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI/ML Engineer, LLM Specialist, Data Scientist","emails":["resumes@radixweb.com","hr@radixweb.com","biz@radixweb.in"],"phone":"+91-79-35200685","website":"radixweb.com","address":"Ekyarth, B/H Nirma University, Chharodi, Ahmedabad 382481","careers_url":"https://radixweb.com/career"},
    {"name":"Inventyv Software Services","city":"Ahmedabad","category":"IT Services","ai_roles":"Full Stack Developer, Cloud Engineer, AI Engineer","emails":["info@inventyv.com","hr@inventyv.com"],"phone":"+91-9979982311","website":"inventyv.com","address":"2nd Floor, Office 211, Binori B Square-3, Sindhu Bhavan Road, Bodakdev, Ahmedabad 380054","careers_url":"https://inventyv.com/careers/"},
    {"name":"SPEC INDIA","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI Developer, Software Engineer, Data Analyst","emails":["lead@spec-india.com","hr@spec-india.com","careers@spec-india.com"],"phone":"+91-79-26404031","website":"spec-india.com","address":"SPEC House, Parth Complex, Near Swastik Cross Roads, Navrangpura, Ahmedabad 380009","careers_url":"https://www.spec-india.com/career/"},
    {"name":"Silicon IT Hub","city":"Ahmedabad","category":"AI-ML","ai_roles":"AI/ML Developer, Software Developer","emails":["smmital@siliconithub.com","hr@siliconithub.com"],"phone":"+91 79 26404032","website":"siliconithub.com","address":"B-1, Ground Floor, Safal Profitaire, Prahladnagar, Ahmedabad 380015","careers_url":"https://www.siliconithub.com/careers/"},
]

FAKE_PHONE = "+91 73260 59369"

def normalize(s):
    return str(s).strip().lower()

def main():
    print(f"\n{'='*70}")
    print(f"  DEEP IT/AI-ML INJECT  - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}\n")

    wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
    ws = wb["COMPLETE_MASTER"]

    headers_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    total_cols = len(headers_row)
    hmap = {str(h).strip().lower(): i+1 for i, h in enumerate(headers_row) if h}

    COL_NO=hmap.get("no.",1); COL_CO=hmap.get("company name",2)
    COL_CITY=hmap.get("city",3); COL_CAT=hmap.get("category",4)
    COL_ROLE=hmap.get("target role",5); COL_PH=hmap.get("phone",12)
    COL_WEB=hmap.get("website",13); COL_ADDR=hmap.get("address",15)
    COL_CAR=hmap.get("careers url",16); COL_PRI=hmap.get("priority",17)
    email_cols=[hmap.get(f"email {i}") for i in range(1,7)]
    email_cols=[ec for ec in email_cols if ec and ec<=total_cols]

    updated=0; inserted=0; emails_added=0

    name_to_row={}
    for ridx,row in enumerate(ws.iter_rows(min_row=2,values_only=False),start=2):
        v=normalize(str(row[COL_CO-1].value or ""))
        if v: name_to_row[v]=ridx

    for co in COMPANIES_DATA:
        norm_name=normalize(co["name"])
        eidx=name_to_row.get(norm_name)
        if eidx:
            r=ws[eidx]
            cur_ph=str(r[COL_PH-1].value or "").strip()
            if not cur_ph or cur_ph==FAKE_PHONE: r[COL_PH-1].value=co["phone"]
            ex={str(r[ec-1].value or "").strip().lower() for ec in email_cols if str(r[ec-1].value or "").strip()}
            ne=[e for e in co["emails"] if e.lower() not in ex]
            slots=[ec for ec in email_cols if not str(r[ec-1].value or "").strip()]
            for sl,em in zip(slots,ne): r[sl-1].value=em; emails_added+=1
            if COL_ADDR and COL_ADDR<=total_cols and (not str(r[COL_ADDR-1].value or "").strip()): r[COL_ADDR-1].value=co["address"]
            if COL_CAR and COL_CAR<=total_cols and (not str(r[COL_CAR-1].value or "").strip()): r[COL_CAR-1].value=co["careers_url"]
            updated+=1
            print(f"  [OK] UPDATED: {co['name']}")
        else:
            nr=ws.max_row+1
            ws.cell(nr,COL_NO).value=nr-1
            ws.cell(nr,COL_CO).value=co["name"]
            if COL_CITY<=total_cols: ws.cell(nr,COL_CITY).value=co["city"]
            if COL_CAT<=total_cols: ws.cell(nr,COL_CAT).value=co["category"]
            if COL_ROLE<=total_cols: ws.cell(nr,COL_ROLE).value=co["ai_roles"]
            ws.cell(nr,COL_PH).value=co["phone"]
            if COL_WEB<=total_cols: ws.cell(nr,COL_WEB).value=co["website"]
            if COL_ADDR and COL_ADDR<=total_cols: ws.cell(nr,COL_ADDR).value=co["address"]
            if COL_CAR and COL_CAR<=total_cols: ws.cell(nr,COL_CAR).value=co["careers_url"]
            for ec,em in zip(email_cols,co["emails"]): ws.cell(nr,ec).value=em; emails_added+=1
            if COL_PRI and COL_PRI<=total_cols:
                p="1 - Apply Now (AI/ML) *" if "AI" in co["category"] or "ML" in co["category"] else ("1 - Apply Now (MNC) *" if "MNC" in co["category"] else "2 - High Priority")
                ws.cell(nr,COL_PRI).value=p
            name_to_row[norm_name]=nr
            inserted+=1
            print(f"  [+] INSERTED: {co['name']}")

    wb.save("COMONK_TRUE_MASTER.xlsx")
    print(f"\nDONE: updated={updated} inserted={inserted} emails_added={emails_added}\n")

if __name__=="__main__":
    main()
