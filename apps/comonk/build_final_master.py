"""
FINAL MASTER EXCEL BUILDER
Sources: HR Mail List.xlsx + all_bcc_emails.txt + Mail list.pdf + HR MAIL MAIN_02.pdf
         + DataScience Company List (embedded) + deep web research
"""
import re, pandas as pd, pdfplumber, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ═══════════════════════════════════════════════════════════════════════════
# A.  RESEARCHED COMPANY DATABASE
# ═══════════════════════════════════════════════════════════════════════════
# Key = lowercase domain-first-part (e.g. 'azilen', 'infocusp')
# Values: address, phone, linkedin, website, roles
DB = {
"3rddigital":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/3rddigital","w":"https://www.3rddigital.com","r":"Software Developer"},
"acenurture":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/acenurture","w":"https://www.acenurture.com","r":"IT Recruiter, HR Executive"},
"acespritech":     {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/acespritech","w":"https://www.acespritech.com","r":"Software Developer"},
"actualisation":   {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/actualisation-ai","w":"https://www.actualisation.ai","r":"AI Engineer, ML Researcher"},
"adp":             {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/adp","w":"https://www.adp.com","r":"Payroll Analyst, HR Tech Developer"},
"agileinfoways":   {"a":"303/304 Abhishree Avenue, Near Nehru Nagar Cross Rd, Ambawadi, Ahmedabad 380015","p":"+91 7622081234","l":"linkedin.com/company/agile-infoways-pvt-ltd-","w":"https://www.agileinfoways.com","r":"Data Analyst, AI/ML Developer"},
"aglowiditsolutions":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/aglowid-it-solutions","w":"https://www.aglowiditsolutions.com","r":"Software Developer, Web Developer"},
"aionpixel":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/aionpixel","w":"https://www.aionpixel.com","r":"AI Developer, Computer Vision Engineer"},
"aipxperts":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/aipxperts","w":"https://www.aipxperts.com","r":"AI/ML Engineer, Data Scientist"},
"airtel":          {"a":"Pan India","p":"","l":"linkedin.com/company/bharti-airtel","w":"https://www.airtel.com","r":"Data Analyst, Network Engineer"},
"aistatics":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/aistatics","w":"https://www.aistatics.com","r":"Data Scientist, AI Research Intern"},
"aividtechvision": {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/aividtechvision","w":"https://www.aividtechvision.com","r":"Computer Vision Engineer, AI Developer"},
"alakmalak":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/alakmalak-technologies","w":"https://www.alakmalak.com","r":"Web Developer, Mobile App Developer"},
"alfalabs":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/alfalabs","w":"https://www.alfalabs.in","r":"Software Developer, ML Engineer"},
"allegisgroup":    {"a":"Pan India (US HQ)","p":"","l":"linkedin.com/company/allegis-group","w":"https://www.allegisgroup.com","r":"IT Recruiter, Talent Acquisition"},
"alliancerecruitmentagency":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/alliance-recruitment-agency","w":"https://www.alliancerecruitmentagency.com","r":"HR Recruiter, Talent Acquisition"},
"amnex":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/amnex","w":"https://www.amnex.com","r":"Telecom Software Developer, Data Analyst"},
"analyticsliv":    {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/analyticsliv","w":"https://www.analyticsliv.com","r":"Digital Analytics Consultant"},
"analyticsvidhya": {"a":"Noida / Pan India","p":"","l":"linkedin.com/company/analytics-vidhya","w":"https://www.analyticsvidhya.com","r":"Data Scientist, ML Trainer"},
"anblicks":        {"a":"Ahmedabad / Hyderabad","p":"","l":"linkedin.com/company/anblicks","w":"https://www.anblicks.com","r":"Data Analyst, Cloud Data Engineer"},
"apexon":          {"a":"Shreem Building, Off SG Highway, Near Iscon Crossroad, Jodhpur, Ahmedabad 380015","p":"","l":"linkedin.com/company/apexon","w":"https://www.apexon.com","r":"AI/ML Consultant, Data Analyst, Full Stack Developer"},
"appitsimple":     {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/appitsimple-infotek","w":"https://www.appitsimple.com","r":"Product Manager, Software Developer"},
"appvintech":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/appvin-technologies","w":"https://www.appvintech.com","r":"Mobile App Developer"},
"artoonsolutions":  {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/artoon-solutions","w":"https://www.artoonsolutions.com","r":"Game Developer, Mobile App Developer"},
"ascendion":       {"a":"Pan India (US HQ)","p":"","l":"linkedin.com/company/ascendion","w":"https://www.ascendion.com","r":"Software Engineer, Data Analyst"},
"astrotalk":       {"a":"Noida / Pan India","p":"","l":"linkedin.com/company/astrotalk","w":"https://www.astrotalk.com","r":"Software Engineer, Data Analyst"},
"atqor":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/atqor","w":"https://www.atqor.com","r":"Software Developer, Web Developer"},
"atliq":           {"a":"Ahmedabad / Vadodara, Gujarat","p":"+91 9979738578","l":"linkedin.com/company/atliqtech","w":"https://www.atliq.com","r":"Data Analyst, AI Consultant, BI Developer"},
"atul":            {"a":"Valsad, Gujarat","p":"","l":"linkedin.com/company/atul-ltd","w":"https://www.atul.co.in","r":"Data Analyst (Chemical Manufacturing)"},
"auberginesolutions":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/aubergine-solutions","w":"https://www.auberginesolutions.com","r":"UX Designer, Product Developer"},
"azilen":          {"a":"12th & 13th Floor, B Square-1, Bopal-Ambli Rd, Ahmedabad 380054","p":"+91 79400 93121","l":"linkedin.com/company/azilentechnologies","w":"https://www.azilen.com","r":"AI Engineer, ML Engineer, Full Stack Developer"},
"azilentechnologies":{"a":"12th & 13th Floor, B Square-1, Bopal-Ambli Rd, Ahmedabad 380054","p":"+91 79400 93121","l":"linkedin.com/company/azilentechnologies","w":"https://www.azilen.com","r":"AI Engineer, ML Engineer, Full Stack Developer"},
"bacancy":         {"a":"1207-1208 Times Square Arcade, Thaltej-Shilaj Rd, Ahmedabad 380059","p":"+91 7940037674","l":"linkedin.com/company/bacancy-technology","w":"https://www.bacancytechnology.com","r":"AI/ML Developer, Data Analyst, React/Node Developer"},
"bacancytechnology":{"a":"1207-1208 Times Square Arcade, Thaltej-Shilaj Rd, Ahmedabad 380059","p":"+91 7940037674","l":"linkedin.com/company/bacancy-technology","w":"https://www.bacancytechnology.com","r":"AI/ML Developer, Data Analyst, React/Node Developer"},
"biztechcs":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/biztechcs","w":"https://www.biztechcs.com","r":"Software Developer, Web Developer"},
"bonrix":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/bonrix-software-systems","w":"https://www.bonrix.net","r":"Software Developer, IoT Developer"},
"brainvire":       {"a":"Sheth Corporate Tower, 9th & 10th Floor, Nr. Nagri Hospital, Ellisbridge, Ahmedabad 380009","p":"+91 7941054646","l":"linkedin.com/company/brainvire-infotech-inc","w":"https://www.brainvire.com","r":"Software Engineer (AI/ML), E-commerce Developer"},
"bureau":          {"a":"Ahmedabad / Bengaluru","p":"","l":"linkedin.com/company/bureau-id","w":"https://www.bureau.id","r":"AI Risk Engineer, Data Scientist"},
"bytestechnolab":  {"a":"THE FIRST, B-805, Near Mansi Cross Road, Vastrapur, Ahmedabad 380015","p":"+91 9157320903","l":"linkedin.com/company/bytes-technolab","w":"https://www.bytestechnolab.com","r":"AI/ML Engineer, Data Scientist, Full Stack Developer"},
"capermint":       {"a":"304-305, Shivalik Shilp-1, Iscon Cross Rd, SG Highway, Ahmedabad 380015","p":"+91 81405 47782","l":"linkedin.com/company/caperminttech","w":"https://www.capermint.com","r":"AI App Developer, Game Developer, AR/VR Engineer"},
"caperminttechnologies":{"a":"304-305, Shivalik Shilp-1, Iscon Cross Rd, SG Highway, Ahmedabad 380015","p":"+91 81405 47782","l":"linkedin.com/company/caperminttech","w":"https://www.capermint.com","r":"AI App Developer, Game Developer, AR/VR Engineer"},
"capgemini":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/capgemini","w":"https://www.capgemini.com","r":"AI/ML Consultant, Data Analyst, Data Engineer"},
"citrusbug":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/citrusbug-technolabs","w":"https://www.citrusbug.co","r":"Software Developer, Web Developer"},
"cleardu":         {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/cleardu","w":"https://www.cleardu.com","r":"EdTech Developer, Data Analyst"},
"cmarix":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/cmarix-technolab","w":"https://www.cmarix.com","r":"Mobile App Developer, Web Developer"},
"codezeros":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/codezeros","w":"https://www.codezeros.com","r":"Blockchain Developer, AI Developer"},
"comprinno":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/comprinno-technologies","w":"https://www.comprinno.net","r":"AI/ML Developer, Computer Vision Engineer"},
"concettolabs":    {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/concetto-labs","w":"https://www.concettolabs.com","r":"AI Consultant, Data Analyst"},
"crestdata":       {"a":"CDS House, Nr. Sarkhej-Sanand Circle, SG Rd, Makarba, Ahmedabad 382210","p":"","l":"linkedin.com/company/crest-data-systems","w":"https://www.crestdata.ai","r":"Data Scientist, ML Engineer, Platform Engineer"},
"cygnetinfotech":  {"a":"16-Swastik Society, NR. Amco Bank, Stadium Circle, Navrangpura, Ahmedabad 380009","p":"","l":"linkedin.com/company/cygnet-infotech","w":"https://www.cygnetinfotech.com","r":"Software Engineer (AI/ML), Data Analyst, ERP Consultant"},
"deepcoder":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/deepcoder","w":"https://www.deepcoder.io","r":"AI Developer, Software Engineer"},
"drcsystems":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/drc-systems","w":"https://www.drcsystems.com","r":"Software Developer, ERP Consultant"},
"drivebuddyai":    {"a":"A-1111, World Trade Tower, Off SG Highway, Makarba, Ahmedabad 380051","p":"+91 11 4117 0779","l":"linkedin.com/company/drivebuddyai","w":"https://www.drivebuddyai.co","r":"AI Fleet Safety Engineer, Data Analyst"},
"dxfactor":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/dxfactor","w":"https://www.dxfactor.com","r":"AI Analyst, Data Engineer"},
"ecubix":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/ecubix","w":"https://www.ecubix.com","r":"Software Developer"},
"einfochips":      {"a":"11/A-B, Chandra Colony, Off C.G. Road, Ellisbridge, Ahmedabad 380006","p":"+91 79 26563705","l":"linkedin.com/company/einfochips","w":"https://www.einfochips.com","r":"AI/ML Engineer (Embedded Systems), Data Engineer"},
"elightwalk":      {"a":"611, Shivalik Square, Near Adani CNG Pump, Old Wadaj, Ahmedabad 380013","p":"+91 7600897405","l":"linkedin.com/company/elightwalk-technology-pvt-ltd","w":"https://www.elightwalk.com","r":"E-commerce Developer, Web Developer"},
"elisiontech":     {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/elision-technolab","w":"https://www.elisiontec.com","r":"VoIP Developer, Telecom Software Engineer"},
"elisiontec":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/elision-technolab","w":"https://www.elisiontec.com","r":"VoIP Developer, Telecom Software Engineer"},
"envitics":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/envitics","w":"https://www.envitics.com","r":"Environmental Data Analyst"},
"esparkinfo":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/esparkbiz","w":"https://www.esparkinfo.com","r":"Web Developer, Mobile App Developer"},
"ethicsinfotech":  {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/ethics-infotech","w":"https://www.ethicsinfotech.in","r":"Software Developer, Web Developer"},
"fitterfly":       {"a":"Mumbai / Ahmedabad","p":"","l":"linkedin.com/company/fitterfly","w":"https://www.fitterfly.com","r":"Data Scientist, Health Tech Developer"},
"flodataanalytics": {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/flo-data-analytics","w":"https://www.flodataanalytics.com","r":"Data Analyst, Business Analyst"},
"fxis":            {"a":"4th Floor, Sanskrut Building, Navrangpura, Ahmedabad 380009","p":"+91 9727999595","l":"linkedin.com/company/fxisai","w":"https://www.fxis.ai","r":"AI Solutions Architect, AI Engineer, Data Scientist"},
"fxisai":          {"a":"4th Floor, Sanskrut Building, Navrangpura, Ahmedabad 380009","p":"+91 9727999595","l":"linkedin.com/company/fxisai","w":"https://www.fxis.ai","r":"AI Solutions Architect, AI Engineer, Data Scientist"},
"growexx":         {"a":"Shivalik Abaise, 1105, Prahlad Nagar, Ahmedabad 380015","p":"+91 7940394046","l":"linkedin.com/company/growexx","w":"https://www.growexx.com","r":"AI Consultant, Data Consultant, Full Stack Developer"},
"hhaexchange":     {"a":"Ahmedabad, Gujarat (India Office)","p":"","l":"linkedin.com/company/hhaexchange","w":"https://www.hhaexchange.com","r":"Software Developer, Data Analyst"},
"hiteshi":         {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/hiteshi-technologies","w":"https://www.hiteshi.com","r":"Software Developer, Mobile App Developer"},
"holbox":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/holboxai","w":"https://www.holbox.ai","r":"AI/ML Engineer, Backend Developer, Generative AI"},
"holboxai":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/holboxai","w":"https://www.holbox.ai","r":"AI/ML Engineer, Backend Developer, Generative AI"},
"hyperlinkinfosystem":{"a":"Block C, 106/B Ganesh Meridian, Near Sola Bridge, SG Highway, Ahmedabad 380061","p":"+91 80001 61161","l":"linkedin.com/company/hyperlinkinfosystem","w":"https://www.hyperlinkinfosystem.com","r":"Data Analyst, App Developer (ML), Mobile Developer"},
"ibm":             {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/ibm","w":"https://www.ibm.com","r":"Data Scientist, AI Application Developer, ML Engineer, AI Consultant"},
"indianic":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/indianic-infotech-limited","w":"https://www.indianic.com","r":"Data Analyst, Web Developer"},
"infocusp":        {"a":"301-310, 4th Floor, Gala Hub, Gala Gymkhana Rd, Bopal, Ahmedabad 380058","p":"","l":"linkedin.com/company/infocusp","w":"https://www.infocusp.com","r":"AI/ML Engineer, Data Scientist, Research Engineer"},
"infosys":         {"a":"Ahmedabad / Gandhinagar, Gujarat","p":"","l":"linkedin.com/company/infosys","w":"https://www.infosys.com","r":"AI/ML Engineer, Data Scientist, Data Engineer, Big Data Architect"},
"inexture":        {"a":"Office No. 405, One World West, Near Ambli T-Junction, SP Ring Rd, Bopal, Ahmedabad 380058","p":"+91 6353697824","l":"linkedin.com/company/inexture","w":"https://www.inexture.com","r":"AI Developer, Python Developer, Django Developer"},
"intellyticssolutions":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/intellytics-solutions","w":"https://www.intellyticssolutions.com","r":"Data Analyst, BI Developer"},
"intuz":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/intuz","w":"https://www.intuz.info","r":"Mobile App Developer, IoT Developer"},
"jeavio":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/jeavio","w":"https://www.jeavio.com","r":"Software Engineer, Data Engineer"},
"kodytechnolab":   {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/kody-technolab","w":"https://www.kodytechnolab.com","r":"Mobile App Developer, AI Developer"},
"krishaweb":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/krishaweb","w":"https://www.krishaweb.com","r":"Web Developer, PHP Developer"},
"logicrays":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/logic-rays-technologies","w":"https://www.logicrays.com","r":"Software Developer, Web Developer"},
"magnetoitsolutions":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/magneto-it-solutions","w":"https://www.magnetoitsolutions.com","r":"E-commerce Developer, Data Analyst"},
"manektech":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/manektech","w":"https://www.manektech.com","r":"Software Developer, Web Developer"},
"marutitech":      {"a":"10th Floor, The Ridge, Opp. Novotel, Iscon Char Rasta, Ahmedabad 380060","p":"+91 9879101067","l":"linkedin.com/company/maruti-techlabs-pvt-ltd","w":"https://www.marutitech.com","r":"Deep Learning Engineer, AI Researcher, NLP Engineer"},
"mastek":          {"a":"804/805 President House, Opp. C N Vidyalaya, Nr. Ambawadi Circle, Ahmedabad 380006","p":"+91 22 6722 4200","l":"linkedin.com/company/mastek","w":"https://www.mastek.com","r":"Software Developer, Data Engineer, Cloud Architect"},
"mindinventory":   {"a":"801, City Centre 2, Science City Road, Sola, Ahmedabad 380060","p":"+1 216-609-0691","l":"linkedin.com/company/mindinventory","w":"https://www.mindinventory.com","r":"AI/ML Developer, Mobile App Developer"},
"mobmaxime":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/mobmaxime","w":"https://www.mobmaxime.com","r":"Mobile App Developer"},
"moontechnolabs":  {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/moon-technolabs","w":"https://www.moontechnolabs.com","r":"Mobile App Developer, AI Integration Developer"},
"moweb":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/moweb-technologies","w":"https://www.moweb.com","r":"Mobile App Developer"},
"neuramonks":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/neuramonks","w":"https://www.neuramonks.com","r":"AI/ML Developer, NLP Engineer, Data Scientist"},
"ninjatechnolabs": {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/ninja-technolabs","w":"https://www.ninjatechnolabs.com","r":"Mobile App Developer, Web Developer"},
"nividous":        {"a":"11th Floor, Shivalik Abaise, Prahladnagar, Ahmedabad 380015","p":"+91 79 4008 1681","l":"linkedin.com/company/nividous","w":"https://www.nividous.com","r":"RPA Developer, AI Automation Engineer, IDP Specialist"},
"openxcell":       {"a":"202-203, Baleshwar Avenue, Opp. Rajpath Club, SG Highway, Ahmedabad 380054","p":"+91 9998222929","l":"linkedin.com/company/openxcell","w":"https://www.openxcell.com","r":"AI Developer, ML Engineer, Mobile App Developer"},
"pirimidtech":     {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/pirimidtech","w":"https://www.pirimidtech.com","r":"Software Developer, SaaS Engineer"},
"prydan":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/prydan","w":"https://www.prydan.com","r":"Software Developer, QA Engineer"},
"qmetry":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/qmetry","w":"https://www.qmetry.com","r":"QA Engineer, Test Automation Engineer"},
"qrioustech":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/qrious-tech","w":"https://www.qrioustech.in","r":"Software Developer"},
"radixweb":        {"a":"401, Anandmangal-II, B/H Omkar House, C.G. Road, Navrangpura, Ahmedabad 380009","p":"+91 9909003119","l":"linkedin.com/company/radixweb","w":"https://www.radixweb.com","r":"Data Analyst, BI Specialist, Full Stack Developer"},
"rishabhsoft":     {"a":"Devx, 4th Floor, Binori B Square3, Sindhu Bhavan Rd, Ahmedabad 380054","p":"+91 8511122697","l":"linkedin.com/company/rishabh-software","w":"https://www.rishabhsoft.com","r":"Software Engineer (BI/Analytics), .NET Developer"},
"rysun":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/rysun","w":"https://www.rysun.com","r":"Software Developer, QA Engineer"},
"saras-3d":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/saras-3d","w":"https://www.saras-3d.com","r":"3D Engineer, CAD Designer"},
"sharkstriker":    {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/sharkstriker","w":"https://www.sharkstriker.com","r":"Cybersecurity Engineer, SOC Analyst"},
"silvertouch":     {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/silvertouch-technologies","w":"https://www.silvertouch.com","r":"Software Developer, ERP Consultant"},
"simform":         {"a":"501 Binori BSquare2, Near Double Tree by Hilton, Ambli Rd, Bopal, Ahmedabad 380054","p":"+91 7940070170","l":"linkedin.com/company/simform","w":"https://www.simform.com","r":"Data Engineer, ML Operations Engineer, Cloud Architect"},
"softwebsolutions": {"a":"Block 5 & 6, Garden View, Corporate House, Bodakdev, Ahmedabad 380054","p":"+1 866-345-7638","l":"linkedin.com/company/softwebsolutionsinc","w":"https://www.softwebsolutions.com","r":"AI for IoT, Data Services, Digital Transformation"},
"solulab":         {"a":"812-16, Times Square 1, Opp. Baghban Party Plot, Thaltej, Ahmedabad 380059","p":"+91 9427026888","l":"linkedin.com/company/solulab","w":"https://www.solulab.com","r":"AI Engineer, Blockchain Developer, ML Developer"},
"solutionanalysts":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/solution-analysts-pvt-ltd","w":"https://www.solutionanalysts.com","r":"Business Intelligence Analyst, Data Analyst"},
"spaceogroup":     {"a":"1005, 10th Floor, Abhishree Adroit, Mansi Circle, Ahmedabad 380015","p":"+91 9316757277","l":"linkedin.com/company/space-o-technologies","w":"https://www.spaceotechnologies.com","r":"Mobile App Developer, AI Integration Developer"},
"specindia":       {"a":"SPEC House, Parth Complex, Near Swastik Cross Roads, Navrangpura, Ahmedabad 380009","p":"+91 79 2640 4031","l":"linkedin.com/company/spec-india","w":"https://www.spec-india.com","r":"BI Analyst, Data Analyst, Software Developer"},
"squadtechnologies":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/squad-technologies","w":"https://www.squadtechnologies.com","r":"Software Developer"},
"stalwartitsolution":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/stalwart-it-solution","w":"https://www.stalwartitsolution.com","r":"Software Developer"},
"stellarmind":     {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/stellarmind-ai","w":"https://www.stellarmind.ai","r":"AI Engineer, ML Researcher, NLP Specialist"},
"tatvic":          {"a":"4th Floor, Camps Corner 2, 100 Feet Rd, Prahladnagar, Ahmedabad 380015","p":"+91 9909981217","l":"linkedin.com/company/tatvic","w":"https://www.tatvic.com","r":"Digital Analytics Consultant, Data Analyst, MarTech Expert"},
"tatvasoft":       {"a":"TatvaSoft House, Rajpath Club Rd, Opp. Golf Academy, Ahmedabad 380054","p":"+91 9601421472","l":"linkedin.com/company/tatvasoft","w":"https://www.tatvasoft.com","r":"Software Developer (Data Science), BI Developer"},
"tcs":             {"a":"Gandhinagar / Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/tata-consultancy-services","w":"https://www.tcs.com","r":"AI Consultant, ML Specialist, Data Scientist, Data Engineer"},
"tecblic":         {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/tecblic","w":"https://www.tecblic.com","r":"Software Developer, Mobile App Developer"},
"techkul":         {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/techkul","w":"https://www.techkul.com","r":"E-commerce Developer, Magento Expert"},
"techmahindra":    {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/tech-mahindra","w":"https://www.techmahindra.com","r":"Data Scientist, ML Engineer, AI Architect, Software Engineer"},
"technource":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/technource","w":"https://www.technource.com","r":"Mobile App Developer, Software Developer"},
"teksun":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/teksun-microsys","w":"https://www.teksun.com","r":"IoT Engineer, Embedded Systems Developer"},
"tntra":           {"a":"B/H Rajpath Club, Nr. AUDA Garden, Off SG Road, Ahmedabad","p":"","l":"linkedin.com/company/tntra","w":"https://www.tntra.io","r":"Software Developer, Tech Entrepreneur, Product Manager"},
"trellance":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/trellance","w":"https://www.trellance.com","r":"Data Analyst, BI Developer, Credit Union Tech"},
"tririd":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/tririd","w":"https://www.tririd.com","r":"Software Developer"},
"trootech":        {"a":"605, 6th Floor, B Square, Iscon-Ambali BRTS Rd, Ahmedabad 380058","p":"+91 2717454342","l":"linkedin.com/company/trootech","w":"https://www.trootech.com","r":"Software Developer, Mobile App Developer, AI Solutions"},
"uplers":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/uplers","w":"https://www.uplers.in","r":"Digital Marketer, Web Developer, Remote Talent"},
"valensdatalabs":  {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/valens-data-labs","w":"https://www.valensdatalabs.com","r":"Data Scientist, ML Engineer, Analytics Consultant"},
"varminect":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/varminect","w":"https://www.varminect.com","r":"Software Developer"},
"vrinsoft":        {"a":"707, Elite Business Park, Shapath Hexa, SG Highway, Ahmedabad 380060","p":"+91 7227906117","l":"linkedin.com/company/vrinsoft-technologies-pvt-ltd","w":"https://www.vrinsofts.com","r":"Mobile App Developer, Software Developer, AI Solutions"},
"vrinsofts":       {"a":"707, Elite Business Park, Shapath Hexa, SG Highway, Ahmedabad 380060","p":"+91 7227906117","l":"linkedin.com/company/vrinsoft-technologies-pvt-ltd","w":"https://www.vrinsofts.com","r":"Mobile App Developer, Software Developer, AI Solutions"},
"webcluesinfotech":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/webclues-infotech","w":"https://www.webcluesinfotech.com","r":"Web Developer, Mobile App Developer"},
"webplanex":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/webplanex","w":"https://www.webplanex.com","r":"Web Developer, Digital Marketing"},
"wipro":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/wipro","w":"https://www.wipro.com","r":"Data Scientist, ML Engineer, AI Solution Architect"},
"xceltec":         {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/xceltec","w":"https://www.xceltec.com","r":"Mobile App Developer, Software Developer"},
"yudiz":           {"a":"13th Floor, BSquare2, Iscon-Ambli Road, Ahmedabad 380054","p":"+91 9033975375","l":"linkedin.com/company/yudiz-solutions-ltd","w":"https://www.yudiz.com","r":"AI/ML Developer, Game Developer, AR/VR Engineer"},
"zipaworld":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/zipaworld","w":"https://www.zipaworld.com","r":"Logistics Tech Developer"},
}

EMAIL_RE = re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')

def get_domain(email):
    try: return email.strip().lower().split('@')[1]
    except: return ''

def domain_key(domain):
    return domain.split('.')[0].lower().replace('-','').replace('_','')

def get_details(company_name, domain):
    k = domain_key(domain)
    cn = re.sub(r'[^a-z]', '', company_name.lower())
    for key in [k, cn]:
        if key in DB: return DB[key]
    for key in DB:
        if key in k or key in cn or k in key or cn in key:
            return DB[key]
    return {}

AI_KW = ['ai','ml','data','analyt','neural','brain','cognitiv','vision','nlp','bot',
          'deep','machine','learn','smart','predict','genai','llm','robotic','intelligence',
          'science','mining','automat','insight','model']
AI_DOMAINS = {'fxis','crestdata','holbox','drivebuddyai','aividtechvision','actualisation',
              'aistatics','stellarmind','infocusp','oblivious','datasafeguard','superr',
              'tailinc','datavruti','ddatalabs','bzanalytics','analyticsvidhya','comprinno',
              'humantic','makunaiglobal','panscience','prama','psytech','hunar','voidspace',
              'neuramonks','aipxperts','biobrain','loomylabs','galaxy','toeho','algobrain',
              'growexx','solulab','mindinventory','simform','concettolabs','valensdatalabs',
              'flodataanalytics','tatvic','anblicks','hexadatasolutions','dxfactor',
              'intellyticssolutions','codework','digitalsherpa','crata','decimalpoint'}

def categorize(name, domain):
    d = domain_key(domain)
    n = name.lower()
    if d in AI_DOMAINS: return 'AI / ML'
    for kw in AI_KW:
        if kw in n or kw in d: return 'AI / ML'
    return 'IT Services'

# ═══════════════════════════════════════════════════════════════════════════
# B.  LOAD & MERGE ALL DATA SOURCES
# ═══════════════════════════════════════════════════════════════════════════
company_emails = defaultdict(lambda: {'emails': [], 'name': ''})  # domain → {name, emails}

def add_email(company_name, email, prefer_name=True):
    if not email or not EMAIL_RE.match(email.strip()): return
    d = get_domain(email)
    if not d or any(x in d for x in ['gmail','yahoo','outlook','hotmail']): return
    rec = company_emails[d]
    if prefer_name and company_name and company_name.lower() not in ('nan',''):
        rec['name'] = company_name
    elif not rec['name'] and company_name:
        rec['name'] = company_name
    if email.lower() not in [e.lower() for e in rec['emails']]:
        rec['emails'].append(email.strip())

# ── Source 1: HR Mail List.xlsx ──────────────────────────────────────────
hr_df = pd.read_excel('HR Mail List.xlsx')
for _, row in hr_df.iterrows():
    company = str(row.get('Company Name','') or '').strip()
    if company in ('','nan'): continue
    for col in ['Company Mail','HR Mail','HR Mail2','HR Mail3','HR Mail4','HR Mail5','HR Mail6']:
        v = str(row.get(col,'') or '').strip()
        if '@' in v: add_email(company, v)

# ── Source 2: all_bcc_emails.txt ─────────────────────────────────────────
with open('all_bcc_emails.txt', encoding='utf-8') as f:
    for line in f:
        e = line.strip().rstrip(',')
        if '@' in e: add_email('', e, prefer_name=False)

# ── Source 3+4: Mail list.pdf & HR MAIL MAIN_02.pdf ─────────────────────
def parse_pdf_companies(fname):
    try:
        with pdfplumber.open(fname) as pdf:
            text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        current = ''
        for line in lines:
            emails_in_line = re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', line)
            if emails_in_line:
                for e in emails_in_line:
                    add_email(current, e, prefer_name=bool(current))
            else:
                # Likely a company name line (no @ symbol)
                if len(line) < 60 and not line.startswith('http'):
                    current = line.title()
    except Exception as ex:
        print(f'PDF error {fname}: {ex}')

parse_pdf_companies('Mail list.pdf')
parse_pdf_companies('HR MAIL MAIN_02.pdf')

# Assign name from domain if still empty
for domain, rec in company_emails.items():
    if not rec['name']:
        rec['name'] = domain.split('.')[0].replace('-',' ').replace('_',' ').title()

# ═══════════════════════════════════════════════════════════════════════════
# C.  BUILD FINAL ROWS  (deduplicated by domain)
# ═══════════════════════════════════════════════════════════════════════════
rows = []
for domain, rec in company_emails.items():
    emails = rec['emails']
    name   = rec['name'] or domain.split('.')[0].title()
    det    = get_details(name, domain)
    rows.append({
        'Company Name': name,
        'Category':     categorize(name, domain),
        'Roles':        det.get('r',''),
        'Email 1':      emails[0] if len(emails)>0 else '',
        'Email 2':      emails[1] if len(emails)>1 else '',
        'Email 3':      emails[2] if len(emails)>2 else '',
        'Email 4':      emails[3] if len(emails)>3 else '',
        'Email 5':      emails[4] if len(emails)>4 else '',
        'Phone':        det.get('p',''),
        'Website':      det.get('w') or f'https://www.{domain}',
        'LinkedIn URL': det.get('l',''),
        'Address':      det.get('a',''),
    })

rows.sort(key=lambda x: x['Company Name'].lower())
print(f"Total companies: {len(rows)}")

# ═══════════════════════════════════════════════════════════════════════════
# D.  EXCEL STYLES
# ═══════════════════════════════════════════════════════════════════════════
H_FILL  = PatternFill("solid", fgColor="1F3864")
H_FONT  = Font(bold=True, color="FFFFFF", size=11)
T_FILL  = PatternFill("solid", fgColor="2E75B6")
AI_FILL = PatternFill("solid", fgColor="E8F5E9")
R_FILLS = [PatternFill("solid", fgColor="DDEEFF"), PatternFill("solid", fgColor="FFFFFF")]
LK_FONT = Font(color="0563C1", underline="single", size=10)
BD_FONT = Font(bold=True, size=10)
NM_FONT = Font(size=10)
bs = Side(style='thin', color='B0C4DE')
BRD = Border(left=bs, right=bs, top=bs, bottom=bs)

COLS = [
    ('No',          5 ),('Company Name',28),('Category',   14),('Roles / Skills', 32),
    ('Email 1',     34),('Email 2',     32),('Email 3',    32),('Email 4',        32),
    ('Email 5',     32),('Phone',       18),('Website',    28),('LinkedIn URL',   34),
    ('Address',     46),
]

def write_sheet(ws, data, title=None):
    start = 1
    if title:
        ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
        tc = ws.cell(1,1,title)
        tc.fill = H_FILL
        tc.font = Font(bold=True, color="FFFFFF", size=13)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34
        ws.row_dimensions[2].height = 5
        start = 3

    for ci,(cn,cw) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = cw

    # Header
    ws.row_dimensions[start].height = 30
    for ci,(cn,_) in enumerate(COLS, 1):
        c = ws.cell(start, ci, cn)
        c.fill = H_FILL; c.font = H_FONT; c.border = BRD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data
    ds = start + 1
    for i, r in enumerate(data, 1):
        rn   = ds + i - 1
        isai = r['Category'] == 'AI / ML'
        fill = AI_FILL if isai else R_FILLS[i%2]
        ws.row_dimensions[rn].height = 18
        vals = [i, r['Company Name'], r['Category'], r['Roles'],
                r['Email 1'], r['Email 2'], r['Email 3'], r['Email 4'], r['Email 5'],
                r['Phone'], r['Website'], r['LinkedIn URL'], r['Address']]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(rn, ci, val)
            c.fill = fill; c.border = BRD
            if ci == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = NM_FONT
            elif ci == 13:
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.font = NM_FONT
            elif ci in (5,6,7,8,9,11,12) and val:
                c.font = LK_FONT
                c.alignment = Alignment(vertical="center")
            elif ci == 2:
                c.font = BD_FONT
                c.alignment = Alignment(vertical="center")
            elif ci == 3:
                c.font = Font(bold=True, size=10, color="1A6600" if isai else "1F3864")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.font = NM_FONT
                c.alignment = Alignment(vertical="center", wrap_text=False)

    ws.freeze_panes = ws.cell(ds, 1)
    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(COLS))}{ds+len(data)-1}"

# ═══════════════════════════════════════════════════════════════════════════
# E.  BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# Sheet 1: All companies
ws1 = wb.active
ws1.title = "All Companies"
write_sheet(ws1, rows, "IT & AI/ML Companies — FINAL MASTER LIST (All Sources Merged)")

# Sheet 2: AI/ML only — same columns, same style
ws2 = wb.create_sheet("AI ML Companies")
ai_rows = [r for r in rows if r['Category']=='AI / ML']
write_sheet(ws2, ai_rows, "AI / ML Companies — Detailed Contact List")

wb.save('Ahmedabad_IT_AIML_FINAL_MASTER.xlsx')

ai  = len(ai_rows)
tot = len(rows)
print(f"Saved: Ahmedabad_IT_AIML_FINAL_MASTER.xlsx")
print(f"  Sheet 1 — All companies : {tot}")
print(f"  Sheet 2 — AI/ML only    : {ai}")
print(f"  IT Services             : {tot-ai}")

# ── Quality Report ───────────────────────────────────────────────────────
print("\n" + "="*60 + "\nDATA QUALITY REPORT\n" + "="*60)
bad = []
for r in rows:
    for col in ['Email 1','Email 2','Email 3','Email 4','Email 5']:
        e = r.get(col,'').strip()
        if e and not EMAIL_RE.match(e):
            bad.append(f"  INVALID | {r['Company Name'][:28]:28s} | {e}")
        if e and get_domain(e).endswith('.incom'):
            bad.append(f"  TYPO    | {r['Company Name'][:28]:28s} | {e}")
if bad:
    print(f"[!] {len(bad)} issue(s):"); [print(b) for b in bad]
else:
    print("[OK] All email formats valid.")
phones = sum(1 for r in rows if r['Phone'])
addrs  = sum(1 for r in rows if r['Address'])
links  = sum(1 for r in rows if r['LinkedIn URL'])
print(f"\nCoverage:")
print(f"  Phone numbers  : {phones}/{tot} ({phones*100//tot}%)")
print(f"  Addresses      : {addrs}/{tot}  ({addrs*100//tot}%)")
print(f"  LinkedIn URLs  : {links}/{tot}  ({links*100//tot}%)")
print("="*60)
