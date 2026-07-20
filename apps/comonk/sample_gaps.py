import openpyxl

wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
ws = wb["COMPLETE_MASTER"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CO = hmap.get("company name", 2)
COL_CITY = hmap.get("city", 3)
COL_CAT = hmap.get("category", 4)
COL_WEB = hmap.get("website", 13)
COL_PH = hmap.get("phone", 12)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]

gaps = []
for r in range(2, ws.max_row + 1):
    has_email = any(ws.cell(r, c).value for c in email_cols)
    has_web = ws.cell(r, COL_WEB).value
    has_ph = ws.cell(r, COL_PH).value
    if not has_email and not has_web and not has_ph:
        gaps.append((r, ws.cell(r, COL_CO).value, ws.cell(r, COL_CITY).value, ws.cell(r, COL_CAT).value))

print(f"Total gap rows: {len(gaps)}")
print("\nFirst 30 samples:")
for r, name, city, cat in gaps[:30]:
    print(f"  row={r:<5} {str(name):<45} city={city!r:<15} cat={cat}")

print("\nLast 20 samples:")
for r, name, city, cat in gaps[-20:]:
    print(f"  row={r:<5} {str(name):<45} city={city!r:<15} cat={cat}")
