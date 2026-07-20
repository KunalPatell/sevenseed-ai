import openpyxl
wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
ws_mnc = wb["MNC_HR_Emails"]
counts = {}
for r in range(2, ws_mnc.max_row+1):
    co = ws_mnc.cell(r, 5).value
    em = ws_mnc.cell(r, 3).value
    if co and em and "@" in str(em):
        counts[str(co)] = counts.get(str(co), 0) + 1

MNCS = ["TCS","Infosys","Wipro","HCL Technologies","Tech Mahindra","Capgemini",
        "Accenture","IBM","Cognizant","Deloitte","EY","PwC","Bosch","Siemens",
        "Oracle","NTT Data","Fujitsu","KPIT","LTTS","Mphasis","Persistent",
        "Hexaware","Apexon","Mastech Digital","EXL"]

print("=" * 52)
print("  MNC RECRUITER EMAIL STATUS")
print("=" * 52)
total = 0
ok_count = 0
for m in MNCS:
    c = counts.get(m, 0)
    total += c
    ok = "OK (10+)" if c >= 10 else "NEED +" + str(10 - c)
    if c >= 10:
        ok_count += 1
    print(f"  {m:<22} {c:>3}  {ok}")
print("=" * 52)
print(f"  PASSED: {ok_count} / 25  |  TOTAL CONTACTS: {total}")
print()

ws = wb["COMPLETE_MASTER"]
ph = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 12).value)
em = sum(1 for r in range(2, ws.max_row+1) if any(ws.cell(r, c).value for c in range(6, 12)))
total_rows = ws.max_row - 1
ae = wb["All_Emails"].max_row - 1
ap = wb["All_Phones"].max_row - 1
ai = wb["AI_ML_ONLY"].max_row - 1

print(f"  Companies total     : {total_rows}")
print(f"  With phone          : {ph}  ({ph * 100 // total_rows}%)")
print(f"  Companies with email: {em}  ({em * 100 // total_rows}%)")
print(f"  All_Emails (total)  : {ae}")
print(f"  All_Phones (unique) : {ap}")
print(f"  AI/ML companies     : {ai}")
print(f"  MNC_HR_Emails rows  : {ws_mnc.max_row - 1}")
print(f"  Sheets in workbook  : {len(wb.sheetnames)}")
