"""
DEEP FILL — Fill blank rows in COMONK_TRUE_MASTER.xlsx using:
  1. VCF file (hr_vcards_15-6-26.vcf) — 150 Ahmedabad HR contacts (phone + email)
  2. Verified DB in merge_vcf_update.py — 361 companies (address, phone, linkedin, website, roles)
Company name embedded in VCF contact name (prefix/postfix) + email domain.
"""

import openpyxl, re, unicodedata, os

MASTER = "COMONK_TRUE_MASTER.xlsx"
VCF    = "hr_vcards_15-6-26.vcf"

# ── Column map (COMPLETE_MASTER) ──────────────────────────────────────────────
C_COMP=2; C_CITY=3; C_CAT=4; C_ROLE=5
C_EMAILS=[6,7,8,9,10,11]; C_PHONE=12; C_WEB=13; C_LI=14; C_ADDR=15
C_SRC=18

def norm(s):
    if not s: return ""
    s = unicodedata.normalize('NFKD', str(s).lower())
    return re.sub(r'[^a-z0-9]', '', s)

def clean_phone(p):
    if not p: return ""
    digits = re.sub(r'[^\d+]', '', str(p))
    digits = digits.replace('+', '')
    if digits.startswith('91') and len(digits) == 12: digits = digits[2:]
    elif digits.startswith('0') and len(digits) == 11: digits = digits[1:]
    if len(digits) == 10 and digits[0] in '6789':
        return '+91 ' + digits[:5] + ' ' + digits[5:]
    if len(digits) == 10 and digits.startswith('79'):
        return '+91 79 ' + digits[2:6] + ' ' + digits[6:]
    if len(digits) == 10:
        return '+91 ' + digits[:5] + ' ' + digits[5:]
    return ""

EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
STOP = {"hr","recruiter","recruitment","talent","acquisition","team","careers",
        "career","jobs","job","hiring","people","manager","executive","ms","mr",
        "the","and","pvt","ltd","llp","inc","technologies","technology","solutions",
        "india","ahmedabad","gujarat"}

def company_from_fn(fn):
    """Strip HR/recruiter words from contact name to get company."""
    words = re.split(r'[\s,;]+', fn or "")
    keep = [w for w in words if norm(w) and norm(w) not in STOP]
    return " ".join(keep).strip()

# ── 1. Parse VCF ──────────────────────────────────────────────────────────────
def parse_vcf():
    contacts = []
    cur = {}
    in_photo = False
    for raw in open(VCF, encoding="utf-8", errors="ignore"):
        line = raw.rstrip("\n").rstrip("\r")
        # Skip base64 photo continuation (indented lines)
        if in_photo:
            if line.startswith(" ") or line.startswith("\t"):
                continue
            else:
                in_photo = False
        if line.startswith("PHOTO"):
            in_photo = True
            continue
        if line == "BEGIN:VCARD":
            cur = {"tels": [], "emails": []}
        elif line == "END:VCARD":
            if cur: contacts.append(cur)
            cur = {}
        elif line.startswith("FN:"):
            cur["fn"] = line[3:].strip()
        elif line.startswith("N:"):
            cur["n"] = line[2:].strip()
        elif line.upper().startswith("TEL"):
            num = line.split(":", 1)[-1].strip()
            if num: cur["tels"].append(num)
        elif line.upper().startswith("EMAIL"):
            em = line.split(":", 1)[-1].strip().lower()
            if EMAIL_RE.match(em): cur["emails"].append(em)
    return contacts

# ── 2. Extract verified DB from merge_vcf_update.py ──────────────────────────
def load_db():
    src = open("merge_vcf_update.py", encoding="utf-8").read()
    m = re.search(r'\nDB\s*=\s*\{', src)
    if not m: return {}
    start = m.end() - 1
    depth = 0
    for i in range(start, len(src)):
        if src[i] == '{': depth += 1
        elif src[i] == '}':
            depth -= 1
            if depth == 0:
                end = i + 1; break
    db_literal = src[m.start():end]
    ns = {}
    exec(db_literal, ns)
    return ns.get("DB", {})

# ── 3. Build company -> {phone, emails} from VCF ─────────────────────────────
def build_vcf_map(contacts):
    m = {}
    for c in contacts:
        fn = c.get("fn", "")
        emails = c.get("emails", [])
        tels = [clean_phone(t) for t in c.get("tels", [])]
        tels = [t for t in tels if t]
        # Company key: prefer email domain, else FN
        keys = set()
        for e in emails:
            dom = e.split("@")[1].split(".")[0]
            if dom not in ("gmail","yahoo","outlook","hotmail","rediffmail","gmal","gmai"):
                keys.add(norm(dom))
        cname = company_from_fn(fn)
        if cname: keys.add(norm(cname))
        for k in keys:
            if not k: continue
            if k not in m: m[k] = {"phone": "", "emails": []}
            if tels and not m[k]["phone"]: m[k]["phone"] = tels[0]
            for e in emails:
                if e not in m[k]["emails"]: m[k]["emails"].append(e)
    return m

def main():
    contacts = parse_vcf()
    print(f"  VCF parsed: {len(contacts)} contacts")
    vcf_map = build_vcf_map(contacts)
    print(f"  VCF company keys: {len(vcf_map)}")
    db = load_db()
    print(f"  Verified DB loaded: {len(db)} companies\n")

    wb = openpyxl.load_workbook(MASTER)
    ws = wb["COMPLETE_MASTER"]

    # Index existing rows
    existing = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, C_COMP).value
        if n: existing[norm(n)] = r

    ph_f = addr_f = li_f = web_f = role_f = em_f = 0

    # ── Fill blanks for existing rows using DB + VCF ─────────────────────────
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, C_COMP).value
        if not name: continue
        key = norm(name)

        dbrec = db.get(key)
        vrec  = vcf_map.get(key)

        # Phone
        if not ws.cell(r, C_PHONE).value:
            ph = (vrec or {}).get("phone") or (dbrec or {}).get("p")
            if ph:
                ws.cell(r, C_PHONE).value = clean_phone(ph) or ph
                ph_f += 1
        # Address
        if not ws.cell(r, C_ADDR).value and dbrec and dbrec.get("a"):
            ws.cell(r, C_ADDR).value = dbrec["a"]; addr_f += 1
        # LinkedIn
        if not ws.cell(r, C_LI).value and dbrec and dbrec.get("l"):
            ws.cell(r, C_LI).value = dbrec["l"]; li_f += 1
        # Website
        if not ws.cell(r, C_WEB).value and dbrec and dbrec.get("w"):
            ws.cell(r, C_WEB).value = dbrec["w"]; web_f += 1
        # Roles
        if not ws.cell(r, C_ROLE).value and dbrec and dbrec.get("r"):
            ws.cell(r, C_ROLE).value = dbrec["r"]; role_f += 1
        # Emails from VCF
        if vrec and vrec.get("emails"):
            current = set(filter(None, [ws.cell(r, c).value for c in C_EMAILS]))
            empty = [c for c in C_EMAILS if not ws.cell(r, c).value]
            newem = [e for e in vrec["emails"] if e not in current]
            for e, c in zip(newem, empty):
                ws.cell(r, c).value = e; em_f += 1
        # Mark source
        if vrec:
            cur = ws.cell(r, C_SRC).value or ""
            if "VCF" not in cur:
                ws.cell(r, C_SRC).value = (cur + ", VCF-HR").strip(", ")

    # ── Add NEW companies from VCF not in master ─────────────────────────────
    added = 0
    from openpyxl.styles import Font, PatternFill, Alignment
    fill_new = PatternFill("solid", fgColor="FFFDE6")
    for key, vrec in vcf_map.items():
        if key in existing or not key: continue
        # Find display name from contacts
        disp = key
        dbrec = db.get(key, {})
        r = ws.max_row + 1
        ws.cell(r, 1, r - 1)
        ws.cell(r, C_COMP, disp.title())
        ws.cell(r, C_CITY, "Ahmedabad")
        ws.cell(r, C_CAT, "IT Services")
        if dbrec.get("r"): ws.cell(r, C_ROLE, dbrec["r"])
        for e, c in zip(vrec.get("emails", []), C_EMAILS):
            ws.cell(r, c, e)
        if vrec.get("phone"): ws.cell(r, C_PHONE, vrec["phone"])
        if dbrec.get("w"): ws.cell(r, C_WEB, dbrec["w"])
        if dbrec.get("l"): ws.cell(r, C_LI, dbrec["l"])
        if dbrec.get("a"): ws.cell(r, C_ADDR, dbrec["a"])
        ws.cell(r, C_SRC, "VCF-HR")
        for c in range(1, 19):
            ws.cell(r, c).fill = fill_new
            ws.cell(r, c).alignment = Alignment(vertical="center")
        existing[key] = r
        added += 1

    for i, r in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(r, 1, i)
    total = ws.max_row - 1
    max_col = ws.max_column

    # Rebuild AI_ML_ONLY
    ws2 = wb["AI_ML_ONLY"]
    for r in range(ws2.max_row, 1, -1):
        ws2.delete_rows(r)
    ai_kw = ['ai','ml','data','machine','deep','nlp','computer vision','llm','genai',
             'analytics','automation','intelligence','learning','aiops']
    from openpyxl.styles import Font as F2
    ai_row = 2
    fa = PatternFill("solid", fgColor="F5F0FF"); fb = PatternFill("solid", fgColor="FFFFFF")
    for r in range(2, ws.max_row + 1):
        cat  = str(ws.cell(r, C_CAT).value or "").lower()
        role = str(ws.cell(r, C_ROLE).value or "").lower()
        if any(k in cat or k in role for k in ai_kw):
            fill = fa if ai_row % 2 == 0 else fb
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                cc = ws2.cell(ai_row, c, v)
                cc.fill = fill
                cc.alignment = Alignment(vertical="center")
                if c in C_EMAILS and v: cc.font = F2(color="1a56db", size=9)
            ws2.cell(ai_row, 1, ai_row - 1)
            ai_row += 1
    ai_count = ws2.max_row - 1

    # Coverage
    with_ph = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, C_PHONE).value)
    with_em = sum(1 for r in range(2, ws.max_row+1) if any(ws.cell(r,c).value for c in C_EMAILS))
    with_ad = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, C_ADDR).value)
    with_li = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, C_LI).value)

    wb["STATS"].cell(3, 2, total)
    wb["STATS"].cell(6, 2, ai_count)
    wb.save(MASTER)

    print(f"  === BLANKS FILLED (existing rows) ===")
    print(f"    Phone:    +{ph_f}")
    print(f"    Address:  +{addr_f}")
    print(f"    LinkedIn: +{li_f}")
    print(f"    Website:  +{web_f}")
    print(f"    Roles:    +{role_f}")
    print(f"    Emails:   +{em_f}")
    print(f"  NEW companies from VCF: +{added}")
    print(f"\n  === TOTAL COVERAGE ({total} companies) ===")
    print(f"    Phone:    {with_ph} ({round(with_ph/total*100)}%)")
    print(f"    Email:    {with_em} ({round(with_em/total*100)}%)")
    print(f"    Address:  {with_ad} ({round(with_ad/total*100)}%)")
    print(f"    LinkedIn: {with_li} ({round(with_li/total*100)}%)")
    print(f"    AI/ML:    {ai_count}")

if __name__ == "__main__":
    main()
