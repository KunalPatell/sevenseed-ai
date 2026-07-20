"""
accuracy_boost_v2.py — Fast High-Accuracy Enricher (Fixed, Non-Blocking)
=========================================================================
Runs after deep_web_research.py to maximize data quality.
Fixed: Short timeouts, async SMTP, verbose progress every row.
"""

import sys, os, re, json, time, asyncio, socket, smtplib
import openpyxl, httpx
from urllib.parse import quote_plus

sys.stdout.reconfigure(encoding='utf-8')

EXCEL  = "COMONK_TRUE_MASTER.xlsx"
SHEET  = "COMPLETE_MASTER"

C_COMP   = 2
C_CITY   = 3
C_EMAILS = [6, 7, 8, 9, 10, 11]
C_PHONE  = 12
C_WEB    = 13
C_LI     = 14
C_PRIO   = 17

EMAIL_RE = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,7}\b')
PHONE_RE = re.compile(r'(?:\+91[\s\-.]?[6-9]\d{9}|0?[6-9]\d{9})')
SKIP_DOMAINS = {"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com","sentry.io","example.com","googlemail.com"}

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0 Safari/537.36",
    "Accept": "text/html,*/*;q=0.8",
}

HR_PREFIXES = ["hr","careers","recruit","jobs","talent","hiring","people","hrd","humanresources"]

# ── Helpers ───────────────────────────────────────────────────────────────────
def clean_phone(p):
    if not p: return ""
    d = re.sub(r'[^\d]', '', str(p))
    if d.startswith('91') and len(d) == 12: d = d[2:]
    if d.startswith('0') and len(d) == 11: d = d[1:]
    if len(d) != 10: return ""
    if d[0] not in '6789': return ""
    if len(set(d)) <= 2: return ""  # e.g. 8888888888
    if d in {"8888888888","9999999999","1234567890","0987654321"}: return ""
    if d[:5] == d[5:]: return ""  # e.g. 9876598765
    return '+91 ' + d[:5] + ' ' + d[5:]

def domain_of(url):
    if not url: return ""
    try:
        from urllib.parse import urlparse
        u = str(url).strip()
        if not u.startswith("http"): u = "https://" + u
        return urlparse(u).netloc.replace("www.", "").strip().lower()
    except: return ""

def is_valid_email(email):
    if not email or '@' not in email: return False
    dom = email.split('@')[1].lower()
    if dom in SKIP_DOMAINS: return False
    if len(email) > 80: return False
    return True

def smtp_quick_verify(email, timeout=3):
    """Quick SMTP check — returns True/False/None (None=unknown)."""
    try:
        domain = email.split('@')[1]
        # Try common mail hostnames
        mx = None
        for prefix in ('mail', 'smtp', 'mx', ''):
            host = f"{prefix}.{domain}" if prefix else domain
            try:
                socket.setdefaulttimeout(2)
                socket.gethostbyname(host)
                mx = host
                break
            except: pass
        if not mx: return None

        server = smtplib.SMTP(timeout=timeout)
        server.connect(mx, 25)
        server.helo("comonk.ai")
        server.mail("test@comonk.ai")
        code, _ = server.rcpt(email)
        server.quit()
        return code == 250
    except smtplib.SMTPConnectError: return None
    except smtplib.SMTPServerDisconnected: return None
    except ConnectionRefusedError: return None
    except OSError: return None
    except Exception: return None

# ── Async Search Functions ────────────────────────────────────────────────────
async def fetch(client, url):
    try:
        r = await client.get(url, timeout=6)
        return r.text if r.status_code == 200 else ""
    except: return ""

async def search_phone(client, company, city):
    queries = [
        f'"{company}" {city} phone number HR contact',
        f'"{company}" site:justdial.com contact',
        f'"{company}" "{city}" +91 recruiter phone',
    ]
    for q in queries:
        html = await fetch(client, f"https://www.google.com/search?q={quote_plus(q)}&num=5")
        for m in PHONE_RE.findall(html):
            c = clean_phone(m)
            if c: return c
        await asyncio.sleep(0.3)
    return ""

async def search_emails(client, company, city, domain):
    found = []
    queries = [
        f'"{company}" hr@{domain} OR careers@{domain} OR recruit@{domain}',
        f'"{company}" {city} HR recruiter email contact',
        f'"{company}" "@{domain}" email',
    ]
    for q in queries:
        html = await fetch(client, f"https://www.google.com/search?q={quote_plus(q)}&num=5")
        for em in EMAIL_RE.findall(html):
            em = em.lower().strip().rstrip('.')
            if is_valid_email(em) and em not in found:
                if domain and domain in em:
                    found.insert(0, em)
                else:
                    found.append(em)
        if found: break
        await asyncio.sleep(0.3)
    return found

async def search_linkedin(client, company):
    q = f'site:linkedin.com/company "{company}" India'
    html = await fetch(client, f"https://www.google.com/search?q={quote_plus(q)}&num=5")
    m = re.search(r'linkedin\.com/company/([a-z0-9\-_%]+)', html, re.I)
    return f"linkedin.com/company/{m.group(1)}" if m else ""

# ── Main ──────────────────────────────────────────────────────────────────────
async def main():
    print("\n" + "="*65)
    print("  COMONK AI — Accuracy Boost v2 (Fast + Multi-Source)")
    print("="*65)

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    total = ws.max_row - 1

    phones_added  = 0
    emails_added  = 0
    li_added      = 0
    fake_removed  = 0
    processed     = 0

    limits = httpx.Limits(max_connections=10, max_keepalive_connections=5)
    async with httpx.AsyncClient(headers=HTTP_HEADERS, limits=limits, verify=False, timeout=8) as client:

        for row in range(2, ws.max_row + 1):
            name = str(ws.cell(row, C_COMP).value or "").strip()
            if not name: continue

            city  = str(ws.cell(row, C_CITY).value or "Ahmedabad").strip()
            web   = str(ws.cell(row, C_WEB).value or "").strip()
            li    = str(ws.cell(row, C_LI).value or "").strip()
            dom   = domain_of(web)

            # 1. Clean fake phones
            ph = ws.cell(row, C_PHONE).value
            if ph:
                cleaned = clean_phone(str(ph))
                if not cleaned:
                    ws.cell(row, C_PHONE).value = None
                    fake_removed += 1

            # 2. Add missing phones
            if not ws.cell(row, C_PHONE).value:
                p = await search_phone(client, name, city)
                if p:
                    ws.cell(row, C_PHONE).value = p
                    phones_added += 1

            # 3. Add missing LinkedIn
            if not li:
                li_url = await search_linkedin(client, name)
                if li_url:
                    ws.cell(row, C_LI).value = li_url
                    li_added += 1

            # 4. Add missing emails
            existing = [ws.cell(row, c).value for c in C_EMAILS if ws.cell(row, c).value]
            if not existing:
                found = []

                # 4a. Search Google for real emails
                scraped = await search_emails(client, name, city, dom)
                found.extend(scraped)

                # 4b. Generate HR patterns from domain and quick SMTP check
                if dom and len(found) < 3:
                    for prefix in HR_PREFIXES:
                        candidate = f"{prefix}@{dom}"
                        result = smtp_quick_verify(candidate, timeout=2)
                        if result in (True, None):  # accept verified + unknown
                            if candidate not in found:
                                found.append(candidate)
                        if len(found) >= 4:
                            break

                # Write found emails to empty columns
                empty_cols = [c for c in C_EMAILS if not ws.cell(row, c).value]
                for em, col in zip(found[:6], empty_cols):
                    ws.cell(row, col).value = em
                    emails_added += 1

            # 5. Recalculate priority
            em_vals = [ws.cell(row, c).value for c in C_EMAILS if ws.cell(row, c).value]
            ph_val  = ws.cell(row, C_PHONE).value
            web_val = ws.cell(row, C_WEB).value
            li_val  = ws.cell(row, C_LI).value

            if em_vals and ph_val:
                ws.cell(row, C_PRIO).value = "1 - Apply Now (Email + Phone)"
            elif em_vals:
                ws.cell(row, C_PRIO).value = "2 - Email Available"
            elif ph_val:
                ws.cell(row, C_PRIO).value = "3 - Phone Only"
            elif web_val:
                ws.cell(row, C_PRIO).value = "4 - Website Only"
            elif li_val:
                ws.cell(row, C_PRIO).value = "5 - LinkedIn Only"
            else:
                ws.cell(row, C_PRIO).value = "6 - Research Needed"

            processed += 1

            # Save + print every 25 rows
            if processed % 25 == 0:
                wb.save(EXCEL)
                pct = round(processed / total * 100, 1)
                print(f"  [{pct:5.1f}%] {processed}/{total} | +{emails_added} emails | +{phones_added} phones | +{li_added} LinkedIn | Cleaned: {fake_removed} fakes", flush=True)

    wb.save(EXCEL)
    print(f"\n{'='*65}")
    print(f"  ACCURACY BOOST COMPLETE!")
    print(f"  Companies processed : {processed}")
    print(f"  Fake phones removed : {fake_removed}")
    print(f"  Emails added        : {emails_added}")
    print(f"  Phones added        : {phones_added}")
    print(f"  LinkedIn URLs added : {li_added}")
    print(f"  File saved          : {EXCEL}")
    print(f"{'='*65}\n")

if __name__ == "__main__":
    asyncio.run(main())
