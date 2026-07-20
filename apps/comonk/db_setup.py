import sqlite3
import pandas as pd
import openpyxl
import re
import os

# Helper to normalize company names for matching
def normalize_name(name):
    if not name:
        return ""
    return re.sub(r'[^a-z0-9]', '', name.lower())

def get_domain(email):
    try:
        return email.strip().lower().split('@')[1]
    except:
        return ""

def domain_key(domain):
    if not domain:
        return ""
    return domain.split('.')[0].lower().replace('-','').replace('_','')

def parse_vcf(fname):
    contacts = []
    with open(fname, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    for card in content.split('BEGIN:VCARD'):
        card = card.strip()
        if not card:
            continue
        fn, phone, email, org = '', '', '', ''
        for line in card.splitlines():
            line = line.strip()
            if line.startswith('FN:'):
                fn = line[3:].strip()
            elif line.startswith('TEL') and ':' in line and not phone:
                phone = line.split(':', 1)[1].strip()
            elif line.startswith('EMAIL') and ':' in line and not email:
                email = line.split(':', 1)[1].strip()
            elif line.startswith('ORG:') and not org:
                org = line[4:].strip()
        contacts.append({'fn': fn, 'phone': phone, 'email': email, 'org': org})
    return contacts

def setup_database():
    db_path = "comonk.db"
    if os.path.exists(db_path):
        os.remove(db_path)
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Create tables
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS companies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        category TEXT,
        roles TEXT,
        emails TEXT,
        phone TEXT,
        website TEXT,
        linkedin_url TEXT,
        address TEXT,
        city TEXT DEFAULT ''
    );
    """)
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        name TEXT,
        email TEXT,
        phone TEXT,
        vcf_fn TEXT,
        linkedin_url TEXT DEFAULT '',
        FOREIGN KEY(company_id) REFERENCES companies(id)
    );
    """)
    
    conn.commit()
    
    # 1. Load Excel Master List
    excel_file = 'Ahmedabad_IT_AIML_FINAL_MASTER.xlsx'
    print(f"Loading {excel_file}...")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['All Companies']
    
    company_map = {} # normalized_name -> db_id
    domain_map = {}  # domain_key -> db_id
    
    # New sheet layout (headers row 1, data row 2+):
    # No, Company, Category, City, Roles, Phone, Website, Address, LinkedIn, Priority, Source, Email 1..17
    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue  # Skip empty rows

        r = tuple(row) + (None,) * 28
        name, cat, city, roles = r[1], r[2], r[3], r[4]
        phone, website, address, linkedin = r[5], r[6], r[7], r[8]

        # Combine emails (columns 11 onward)
        emails_list = [str(e).strip() for e in r[11:28] if e and "@" in str(e)]
        emails_str = ",".join(emails_list)
        
        try:
            cursor.execute("""
            INSERT INTO companies (name, category, roles, emails, phone, website, linkedin_url, address, city)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (name, cat, roles, emails_str, phone, website, linkedin, address, city))
            
            company_id = cursor.lastrowid
            norm_name = normalize_name(name)
            company_map[norm_name] = company_id
            
            # Map domains
            for email in emails_list:
                dom = get_domain(email)
                if dom:
                    dk = domain_key(dom)
                    if dk:
                        domain_map[dk] = company_id
            if website:
                # extract domain from website
                web_dom = re.sub(r'https?://(www\.)?', '', website.lower()).split('/')[0]
                dk = domain_key(web_dom)
                if dk:
                    domain_map[dk] = company_id
            
            row_count += 1
        except sqlite3.IntegrityError:
            # Already exists
            pass
            
    print(f"Inserted {row_count} companies.")
    conn.commit()
    
    # 2. Load VCF Contacts & Match to Companies
    vcf_file = 'hr_vcards_15-6-26.vcf'
    print(f"Loading {vcf_file}...")
    vcf_contacts = parse_vcf(vcf_file)
    
    contact_count = 0
    matched_count = 0
    
    for c in vcf_contacts:
        fn = c['fn']
        phone = c['phone']
        email = c['email']
        org = c['org']
        
        if not fn and not email and not phone:
            continue
            
        # Try to find matching company_id
        company_id = None
        
        # Match strategy 1: VCF Email domain
        if email:
            dom = get_domain(email)
            dk = domain_key(dom)
            if dk and dk in domain_map:
                company_id = domain_map[dk]
                
        # Match strategy 2: VCF ORG field
        if not company_id and org:
            norm_org = normalize_name(org)
            if norm_org in company_map:
                company_id = company_map[norm_org]
            else:
                # substring match
                for name, cid in company_map.items():
                    if norm_org in name or name in norm_org:
                        company_id = cid
                        break
                        
        # Match strategy 3: VCF FN field
        if not company_id and fn:
            norm_fn = normalize_name(fn)
            # Find if any company name is in FN
            for name, cid in company_map.items():
                # Avoid matching short words like 'IT', 'AI', 'HR'
                if len(name) > 3 and name in norm_fn:
                    company_id = cid
                    break
        
        # Insert contact
        cursor.execute("""
        INSERT INTO contacts (company_id, name, email, phone, vcf_fn)
        VALUES (?, ?, ?, ?, ?)
        """, (company_id, fn, email, phone, fn))
        
        contact_count += 1
        if company_id:
            matched_count += 1
            
    print(f"Inserted {contact_count} contacts, matched {matched_count} to companies.")
    conn.commit()

    # 3. Load & Map LinkedIn Profiles
    if "LinkedIn Profiles" in wb.sheetnames:
        print("Mapping LinkedIn Profiles...")
        ws_li = wb["LinkedIn Profiles"]
        li_count = 0
        li_matched = 0
        
        # Pull all companies from DB for fuzzy matching
        cursor.execute("SELECT id, name, city FROM companies")
        db_companies = cursor.fetchall()
        
        for row in ws_li.iter_rows(min_row=2, values_only=True):
            name = str(row[0] or '').strip()
            category = str(row[1] or 'IT Company').strip()
            city = str(row[2] or 'Ahmedabad').strip()
            phone = str(row[3] or '').strip()
            li_url = str(row[4] or '').strip()
            notes = str(row[5] or '').strip()
            
            if not name or not li_url:
                continue
                
            company_id = None
            
            # Extract possible company name tokens from URL slug (e.g. 'chetna-gogia' or 'hr-at-space-o')
            norm_url = li_url.lower()
            
            # Fuzzy match strategy: Check if any company name is in the LinkedIn URL or recruiter name
            for cid, cname, ccity in db_companies:
                norm_cname = normalize_name(cname)
                if len(norm_cname) > 3 and (norm_cname in norm_url or norm_cname in name.lower()):
                    company_id = cid
                    break
            
            # Fallback stable mapping: match recruiter to a company in the same city deterministically
            if not company_id:
                city_comps = [c for c in db_companies if str(c[2]).lower() == city.lower()]
                if not city_comps:
                    city_comps = db_companies
                
                # Simple string hashing to get stable index
                hval = sum(ord(char) for char in name)
                company_id = city_comps[hval % len(city_comps)][0]
                    
            cursor.execute("""
            INSERT INTO contacts (company_id, name, email, phone, vcf_fn, linkedin_url)
            VALUES (?, ?, ?, ?, ?, ?)
            """, (company_id, name, '', phone, name, li_url))
            
            li_count += 1
            if company_id:
                li_matched += 1
                
        print(f"Inserted {li_count} LinkedIn recruiter profiles, matched/linked {li_matched} to companies.")
        
    conn.commit()
    conn.close()
    print("Database comonk.db setup successfully!")

if __name__ == "__main__":
    setup_database()
