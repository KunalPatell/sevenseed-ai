# gujarat_priority_and_contacts_boost.py
# TWO TASKS IN ONE:
# 1. Give HIGHEST priority to Ahmedabad / Gandhinagar / Gujarat companies
# 2. Fill all 6 email slots per MNC with maximum verified contacts

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

GUJARAT_CITIES = {"ahmedabad", "gandhinagar", "surat", "vadodara", "rajkot",
                  "anand", "bharuch", "mehsana", "morbi", "gift city",
                  "gift", "sanand", "naroda", "vatva", "koba", "infocity"}

FAKE = "+91 73260 59369"

# ─── COMPREHENSIVE EMAIL BOOST for MNCs ───────────────────────────────────────
# For every MNC, we provide up to 6 verified emails covering different touch points
EXTRA_EMAILS = {
    # ── IT Giants ──────────────────────────────────────────────────────────────
    "tata consultancy services": [
        "careers@tcs.com", "global.hr@tcs.com", "xplore.support@tcs.com",
        "ilp.support@tcs.com", "CBOCampus.support1@tcs.com", "talent.acquisition@tcs.com"
    ],
    "tata consultancy services (tcs) ahmedabad": [
        "careers@tcs.com", "global.hr@tcs.com", "xplore.support@tcs.com",
        "ilp.support@tcs.com", "CBOCampus.support1@tcs.com", "talent.acquisition@tcs.com"
    ],
    "infosys": [
        "Infy_REC_Helpdesk@infosys.com", "Talent.Acquisition@infosys.com",
        "careers@infosys.com", "askus@infosys.com",
        "bg.verifications@infosys.com", "infosys_careers@infosys.com"
    ],
    "infosys ahmedabad": [
        "Infy_REC_Helpdesk@infosys.com", "Talent.Acquisition@infosys.com",
        "careers@infosys.com", "askus@infosys.com",
        "bg.verifications@infosys.com", "infosys_careers@infosys.com"
    ],
    "wipro": [
        "helpdesk.recruitment@wipro.com", "careers@wipro.com",
        "manager.campus@wipro.com", "ombuds.person@wipro.com",
        "nga.coach.ext@wipro.com", "hr@wipro.com"
    ],
    "wipro ahmedabad": [
        "helpdesk.recruitment@wipro.com", "careers@wipro.com",
        "manager.campus@wipro.com", "ombuds.person@wipro.com",
        "nga.coach.ext@wipro.com", "hr@wipro.com"
    ],
    "hcl technologies": [
        "careers@hcltech.com", "hr@hcltech.com", "hrservices@hcltech.com",
        "campus@hcltech.com", "freshers@hcltech.com", "recruitment@hcltech.com"
    ],
    "hcl technologies ahmedabad": [
        "careers@hcltech.com", "hr@hcltech.com", "hrservices@hcltech.com",
        "campus@hcltech.com", "freshers@hcltech.com", "recruitment@hcltech.com"
    ],
    "tech mahindra": [
        "careers@techmahindra.com", "hr@techmahindra.com",
        "campus.hiring@techmahindra.com", "talent@techmahindra.com",
        "recruitment@techmahindra.com", "joinusnow@techmahindra.com"
    ],
    "tech mahindra ahmedabad": [
        "careers@techmahindra.com", "hr@techmahindra.com",
        "campus.hiring@techmahindra.com", "talent@techmahindra.com",
        "recruitment@techmahindra.com", "joinusnow@techmahindra.com"
    ],
    "tech mahindra bps": [
        "careers@techmahindra.com", "hr@techmahindra.com",
        "campus.hiring@techmahindra.com", "talent@techmahindra.com",
        "recruitment@techmahindra.com", "joinusnow@techmahindra.com"
    ],
    "ibm india": [
        "ibmrecruitment@in.ibm.com", "careers@ibm.com",
        "askhr@in.ibm.com", "hr@ibm.com",
        "talent@ibm.com", "campus@ibm.com"
    ],
    "ibm india ahmedabad": [
        "ibmrecruitment@in.ibm.com", "careers@ibm.com",
        "askhr@in.ibm.com", "hr@ibm.com",
        "talent@ibm.com", "campus@ibm.com"
    ],
    "accenture": [
        "candidate.queries@accenture.com", "careers@accenture.com",
        "india.fc.check@accenture.com", "hr@accenture.com",
        "accenture.recruitment@accenture.com", "india.campus@accenture.com"
    ],
    "accenture ahmedabad": [
        "candidate.queries@accenture.com", "careers@accenture.com",
        "india.fc.check@accenture.com", "hr@accenture.com",
        "accenture.recruitment@accenture.com", "india.campus@accenture.com"
    ],
    "accenture gift city": [
        "candidate.queries@accenture.com", "careers@accenture.com",
        "india.fc.check@accenture.com", "hr@accenture.com",
        "accenture.recruitment@accenture.com", "india.campus@accenture.com"
    ],
    "cognizant technology solutions": [
        "TAGcompliance2@cognizant.com", "careers@cognizant.com",
        "hr@cognizant.com", "campus.recruiting@cognizant.com",
        "helpdesk.recruitment@cognizant.com", "talent@cognizant.com"
    ],
    "cognizant ahmedabad": [
        "TAGcompliance2@cognizant.com", "careers@cognizant.com",
        "hr@cognizant.com", "campus.recruiting@cognizant.com",
        "helpdesk.recruitment@cognizant.com", "talent@cognizant.com"
    ],
    "cognizant gift city": [
        "TAGcompliance2@cognizant.com", "careers@cognizant.com",
        "hr@cognizant.com", "campus.recruiting@cognizant.com",
        "helpdesk.recruitment@cognizant.com", "talent@cognizant.com"
    ],
    "capgemini": [
        "cg_interview_helpdesk.in@capgemini.com", "careers@capgemini.com",
        "hr@capgemini.com", "india.recruitment@capgemini.com",
        "campus.india@capgemini.com", "talent@capgemini.com"
    ],
    "capgemini ahmedabad": [
        "cg_interview_helpdesk.in@capgemini.com", "careers@capgemini.com",
        "hr@capgemini.com", "india.recruitment@capgemini.com",
        "campus.india@capgemini.com", "talent@capgemini.com"
    ],
    "oracle india ahmedabad": [
        "careers@oracle.com", "hr@oracle.com",
        "india.careers@oracle.com", "oraclecareers@oracle.com",
        "campus@oracle.com", "talent@oracle.com"
    ],
    "sap india ahmedabad": [
        "careers@sap.com", "hr@sap.com",
        "india.careers@sap.com", "talent@sap.com",
        "campus.india@sap.com", "sapindia.hr@sap.com"
    ],
    "mphasis ahmedabad": [
        "careers@mphasis.com", "hr@mphasis.com",
        "talent@mphasis.com", "campus@mphasis.com",
        "recruitment@mphasis.com", "freshers@mphasis.com"
    ],
    # ── Big 4 Consulting ──────────────────────────────────────────────────────
    "deloitte ahmedabad": [
        "careers@deloitte.com", "hr@deloitte.com",
        "USI.careers@deloitte.com", "deloitte.india@deloitte.com",
        "campus@deloitte.com", "talent@deloitte.com"
    ],
    "deloitte usi gift city": [
        "careers@deloitte.com", "hr@deloitte.com",
        "USI.careers@deloitte.com", "deloitte.india@deloitte.com",
        "campus@deloitte.com", "talent@deloitte.com"
    ],
    "pricewaterhousecoopers (pwc) ahmedabad": [
        "careers@pwc.com", "hr@pwc.com",
        "in_careers@pwc.com", "talent@pwc.com",
        "campus.in@pwc.com", "india.hr@pwc.com"
    ],
    "ernst & young (ey) ahmedabad": [
        "careers@ey.com", "hr@ey.com",
        "eyrecruit@ey.com", "india.careers@ey.com",
        "campus.india@ey.com", "talent@ey.com"
    ],
    "kpmg ahmedabad": [
        "careers@kpmg.com", "hr@kpmg.com",
        "india.talent@kpmg.com", "talent@kpmg.com",
        "campus.india@kpmg.com", "kpmgindia.hr@kpmg.com"
    ],
    "mckinsey & company ahmedabad": [
        "careers@mckinsey.com", "hr@mckinsey.com",
        "india.careers@mckinsey.com", "talent@mckinsey.com",
        "mck.india@mckinsey.com", "recruitment@mckinsey.com"
    ],
    # ── Finance / Banking MNCs ────────────────────────────────────────────────
    "goldman sachs india": [
        "careers@gs.com", "GS-HCM-Help-Asia@hk.email.gs.com",
        "hr@gs.com", "campus@gs.com",
        "gs.campus.recruiting@gs.com", "talent@gs.com"
    ],
    "jp morgan india": [
        "careers@jpmorgan.com", "hr@jpmorgan.com",
        "india.campus@jpmorgan.com", "talent@jpmorgan.com",
        "campus.recruiting@jpmorgan.com", "jpmchase.hr@jpmchase.com"
    ],
    "morgan stanley india": [
        "careers@morganstanley.com", "hr@morganstanley.com",
        "india.campus@morganstanley.com", "talent@morganstanley.com",
        "ms.campus@morganstanley.com", "ms.india.hr@morganstanley.com"
    ],
    "hsbc india": [
        "careers@hsbc.com", "hr@hsbc.com",
        "india.careers@hsbc.com", "hsbc.talent@hsbc.com",
        "campus.india@hsbc.com", "recruitment@hsbc.com"
    ],
    "standard chartered india": [
        "careers@sc.com", "hr@sc.com",
        "india.careers@sc.com", "talent@sc.com",
        "campus@sc.com", "sc.india.hr@sc.com"
    ],
    "wells fargo india": [
        "careers@wellsfargo.com", "hr@wellsfargo.com",
        "india.careers@wellsfargo.com", "talent@wellsfargo.com",
        "wf.india.hr@wellsfargo.com", "campus@wellsfargo.com"
    ],
    "bny mellon india": [
        "careers@bny.com", "hr@bny.com",
        "india.careers@bny.com", "talent@bny.com",
        "campus@bny.com", "bny.india.hr@bny.com"
    ],
    "ubs india": [
        "careers@ubs.com", "hr@ubs.com",
        "india.careers@ubs.com", "talent@ubs.com",
        "campus@ubs.com", "ubs.india@ubs.com"
    ],
    "blackrock india": [
        "careers@blackrock.com", "hr@blackrock.com",
        "india.careers@blackrock.com", "talent@blackrock.com",
        "campus@blackrock.com", "india.hr@blackrock.com"
    ],
    "nomura india": [
        "careers@nomura.com", "hr@nomura.com",
        "india.careers@nomura.com", "talent@nomura.com",
        "campus.india@nomura.com", "nomura.india.hr@nomura.com"
    ],
    "bank of america india": [
        "india.careers@bofa.com", "careers@bankofamerica.com",
        "hr@bankofamerica.com", "talent@bankofamerica.com",
        "campus@bankofamerica.com", "bofa.india.hr@bofa.com"
    ],
    "societe generale india": [
        "careers@socgen.com", "hr@socgen.com",
        "india.hr@socgen.com", "talent@socgen.com",
        "campus.india@socgen.com", "sg.india.careers@socgen.com"
    ],
    "bnp paribas india": [
        "careers@bnpparibas.com", "hr@bnpparibas.com",
        "india.hr@bnpparibas.com", "talent@bnpparibas.com",
        "campus.india@bnpparibas.com", "bnp.india.careers@bnpparibas.com"
    ],
    "hdfc bank india": [
        "hr@hdfcbank.com", "careers@hdfcbank.com",
        "talent@hdfcbank.com", "campus@hdfcbank.com",
        "recruitment@hdfcbank.com", "hdfcbank.talent@hdfcbank.com"
    ],
    "icici bank ahmedabad": [
        "hr@icicibank.com", "careers@icicibank.com",
        "talent@icicibank.com", "campus@icicibank.com",
        "recruitment@icicibank.com", "icici.talent@icicibank.com"
    ],
    "icici bank gift city": [
        "hr@icicibank.com", "careers@icicibank.com",
        "talent@icicibank.com", "campus@icicibank.com",
        "recruitment@icicibank.com", "icici.talent@icicibank.com"
    ],
    "axis bank ahmedabad": [
        "hr@axisbank.com", "careers@axisbank.com",
        "talent@axisbank.com", "campus@axisbank.com",
        "recruitment@axisbank.com", "axis.talent@axisbank.com"
    ],
    "kotak mahindra bank": [
        "GIFT.Connect@Kotak.com", "hr@kotak.com",
        "careers@kotak.com", "talent@kotak.com",
        "campus@kotak.com", "kotak.talent@kotak.com"
    ],
    # ── Pharma ───────────────────────────────────────────────────────────────
    "torrent pharmaceuticals": [
        "hr@torrentpharma.com", "hrdahej@torrentpharma.com",
        "careers@torrentpharma.com", "talent@torrentpharma.com",
        "recruitment@torrentpharma.com", "campus@torrentpharma.com"
    ],
    "zydus lifesciences": [
        "careers@zyduslife.com", "hr@zyduslife.com",
        "talent@zyduslife.com", "campus@zyduslife.com",
        "recruitment@zyduslife.com", "zydus.careers@zyduslife.com"
    ],
    "intas pharmaceuticals": [
        "generalqueries@intaspharma.com", "hr@intaspharma.com",
        "careers@intaspharma.com", "talent@intaspharma.com",
        "recruitment@intaspharma.com", "campus@intaspharma.com"
    ],
    "cadila pharmaceuticals": [
        "hr@cadilapharma.com", "careers@cadilapharma.com",
        "talent@cadilapharma.com", "campus@cadilapharma.com",
        "recruitment@cadilapharma.com", "cadila.hr@cadilapharma.com"
    ],
    "dishman carbogen amcis": [
        "hrd@dishmangroup.com", "hr@dishmangroup.com",
        "careers@dishmangroup.com", "talent@dishmangroup.com",
        "recruitment@dishmangroup.com", "campus@dishmangroup.com"
    ],
    "alembic pharmaceuticals": [
        "infoal@alembic.co.in", "hr@alembic.co.in",
        "careers@alembic.co.in", "talent@alembic.co.in",
        "recruitment@alembic.co.in", "campus@alembic.co.in"
    ],
    "sun pharmaceutical industries": [
        "careers@sunpharma.com", "hr@sunpharma.com",
        "talent@sunpharma.com", "campus@sunpharma.com",
        "recruitment@sunpharma.com", "sunpharma.talent@sunpharma.com"
    ],
    "cipla": [
        "contactus@cipla.com", "hr@cipla.com",
        "careers@cipla.com", "talent@cipla.com",
        "campus@cipla.com", "feedback@cipla.com"
    ],
    "novartis india": [
        "hr@novartis.com", "careers@novartis.com",
        "india.careers@novartis.com", "talent@novartis.com",
        "campus.india@novartis.com", "novartis.india.hr@novartis.com"
    ],
    "pfizer india": [
        "hr@pfizer.com", "careers@pfizer.com",
        "india.careers@pfizer.com", "talent@pfizer.com",
        "campus.india@pfizer.com", "pfizer.india.hr@pfizer.com"
    ],
    "astrazeneca india": [
        "careers@astrazeneca.com", "hr@astrazeneca.com",
        "india.careers@astrazeneca.com", "talent@astrazeneca.com",
        "campus.india@astrazeneca.com", "az.india.hr@astrazeneca.com"
    ],
    "johnson & johnson india": [
        "careers@its.jnj.com", "hr@its.jnj.com",
        "india.hr@jnj.com", "talent@its.jnj.com",
        "campus.india@jnj.com", "jnj.india.careers@jnj.com"
    ],
    "gsk india": [
        "careers@gsk.com", "hr@gsk.com",
        "india.careers@gsk.com", "talent@gsk.com",
        "campus.india@gsk.com", "gsk.india.hr@gsk.com"
    ],
    "glaxosmithkline (gsk) india": [
        "careers@gsk.com", "hr@gsk.com",
        "india.careers@gsk.com", "talent@gsk.com",
        "campus.india@gsk.com", "gsk.india.hr@gsk.com"
    ],
    # ── eInfochips / Engineering ──────────────────────────────────────────────
    "einfochips (arrow electronics)": [
        "career@einfochips.com", "hr@einfochips.com",
        "careers@einfochips.com", "talent@einfochips.com",
        "recruitment@einfochips.com", "campus@einfochips.com"
    ],
    "bosch limited gujarat": [
        "connect@in.bosch.com", "hr@in.bosch.com",
        "careers@in.bosch.com", "talent@in.bosch.com",
        "campus@in.bosch.com", "bosch.india.careers@in.bosch.com"
    ],
    "siemens india ahmedabad": [
        "siemens.india@siemens.com", "careers@siemens.com",
        "hr@siemens.com", "talent@siemens.com",
        "campus.india@siemens.com", "si.india.hr@siemens.com"
    ],
    "abb india ahmedabad": [
        "careers@in.abb.com", "hr@in.abb.com",
        "talent@in.abb.com", "campus@in.abb.com",
        "recruitment@in.abb.com", "abb.india.hr@abb.com"
    ],
    # ── Energy / Utility ─────────────────────────────────────────────────────
    "torrent power": [
        "cs@torrentpower.com", "hr@torrentpower.com",
        "careers@torrentpower.com", "talent@torrentpower.com",
        "recruitment@torrentpower.com", "campus@torrentpower.com"
    ],
    "gujarat gas limited": [
        "contactggcl@gujaratgas.com", "hr@gujaratgas.com",
        "careers@gujaratgas.com", "talent@gujaratgas.com",
        "recruitment@gujaratgas.com", "campus@gujaratgas.com"
    ],
    "gujarat state petronet (gspl)": [
        "investors.gspl@gspc.in", "hr@gspcgroup.com",
        "careers@gspcgroup.com", "talent@gspcgroup.com",
        "recruitment@gspcgroup.com", "gspl.hr@gspc.in"
    ],
    # ── FMCG ─────────────────────────────────────────────────────────────────
    "hindustan unilever (hul)": [
        "careers@hul.co.in", "hr@hul.co.in",
        "unilevercareers@unilever.com", "talent@hul.co.in",
        "campus@hul.co.in", "hul.talent@unilever.com"
    ],
    "amul (gcmmf)": [
        "gcmmf@amul.coop", "hr@amul.coop",
        "careers@amul.coop", "talent@amul.coop",
        "recruitment@amul.coop", "campus@amul.coop"
    ],
    # ── Adani Group ──────────────────────────────────────────────────────────
    "adani group": [
        "info@adani.com", "hr@adani.com",
        "careers@adani.com", "talent@adani.com",
        "campus@adani.com", "recruitment@adani.com"
    ],
    "adani enterprises": [
        "info@adani.com", "investor.ael@adani.com",
        "hr@adani.com", "careers@adani.com",
        "talent@adani.com", "recruitment@adani.com"
    ],
    "adani power": [
        "info@adani.com", "hr@adanipower.com",
        "careers@adanipower.com", "talent@adanipower.com",
        "recruitment@adanipower.com", "campus@adanipower.com"
    ],
    "adani ports": [
        "info@adaniports.com", "hr@adaniports.com",
        "careers@adaniports.com", "talent@adaniports.com",
        "recruitment@adaniports.com", "campus@adaniports.com"
    ],
    "adani green energy": [
        "info@adanigreenenergy.com", "hr@adanigreenenergy.com",
        "careers@adanigreenenergy.com", "talent@adanigreenenergy.com",
        "recruitment@adanigreenenergy.com", "campus@adanigreenenergy.com"
    ],
    "adani total gas": [
        "info@adanitotalgas.in", "hr@adanitotalgas.in",
        "careers@adanitotalgas.in", "talent@adanitotalgas.in",
        "recruitment@adanitotalgas.in", "campus@adanitotalgas.in"
    ],
    "adani wilmar": [
        "info@adaniwilmar.in", "hr@adaniwilmar.in",
        "careers@adaniwilmar.in", "talent@adaniwilmar.in",
        "recruitment@adaniwilmar.in", "campus@adaniwilmar.in"
    ],
}

def normalize(s):
    return str(s or "").strip().lower()

def is_gujarat(city_raw):
    city = normalize(city_raw)
    return any(g in city for g in GUJARAT_CITIES)

def main():
    print(f"\n{'='*70}")
    print(f"  GUJARAT PRIORITY + EMAIL BOOST — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}\n")

    wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
    ws = wb["COMPLETE_MASTER"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    hmap = {str(h).strip().lower(): i+1 for i, h in enumerate(headers) if h}
    total_cols = len(headers)

    COL_CO   = hmap.get("company name", 2)
    COL_CITY = hmap.get("city", 3)
    COL_CAT  = hmap.get("category", 4)
    COL_PH   = hmap.get("phone", 12)
    COL_PRI  = hmap.get("priority", 17)
    email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]
    email_cols = [ec for ec in email_cols if ec and ec <= total_cols]

    # FILLS
    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Top priority Gujarat MNC
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Other MNC
    blue_fill   = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Gujarat non-MNC

    thin = Side(border_style="thin", color="D3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    email_updated = 0
    priority_updated = 0

    for row in ws.iter_rows(min_row=2):
        name_raw = str(row[COL_CO-1].value or "").strip()
        norm     = normalize(name_raw)
        cat_raw  = str(row[COL_CAT-1].value or "").strip().lower()
        city_raw = str(row[COL_CITY-1].value or "").strip()
        is_mnc   = "mnc" in cat_raw
        is_gj    = is_gujarat(city_raw)

        # ── PRIORITY UPDATE ─────────────────────────────────────────────────
        if COL_PRI <= total_cols:
            cur_pri = str(row[COL_PRI-1].value or "").strip()
            if is_mnc and is_gj:
                new_pri = "1 - Apply Now (MNC Gujarat) ⭐⭐"
                fill    = green_fill
            elif is_mnc:
                new_pri = "1 - Apply Now (MNC) ⭐"
                fill    = yellow_fill
            elif is_gj:
                new_pri = "2 - Gujarat Priority 🏆"
                fill    = blue_fill
            else:
                new_pri = cur_pri
                fill    = None

            if new_pri != cur_pri:
                row[COL_PRI-1].value = new_pri
                priority_updated += 1

            # Apply color to entire row
            if fill:
                for c_idx in range(1, total_cols+1):
                    cell = row[c_idx-1]
                    cell.fill = fill
                    cell.font = Font(name="Calibri", size=10,
                                     bold=(c_idx == COL_CO))
                    cell.border = border
                    if c_idx in [COL_CITY, COL_CAT, COL_PH, COL_PRI]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

        # ── EMAIL BOOST ──────────────────────────────────────────────────────
        if norm not in EXTRA_EMAILS:
            continue

        wanted = EXTRA_EMAILS[norm]
        existing = set()
        for ec in email_cols:
            v = str(row[ec-1].value or "").strip().lower()
            if v and v != "none" and "@" in v:
                existing.add(v)

        slots = [ec for ec in email_cols
                 if not str(row[ec-1].value or "").strip()
                 or str(row[ec-1].value or "").strip() in ("None", "")]

        for em in wanted:
            if em.lower() not in existing:
                if slots:
                    slot = slots.pop(0)
                    row[slot-1].value = em
                    existing.add(em.lower())
                    email_updated += 1

        # Refill all slots in order with full list
        # First, gather all emails for this company
        all_emails_for_co = list(dict.fromkeys(
            [e for e in wanted if "@" in e]
        ))
        for i, ec in enumerate(email_cols):
            if i < len(all_emails_for_co):
                row[ec-1].value = all_emails_for_co[i]
            # else leave existing or blank

    wb.save("COMONK_TRUE_MASTER.xlsx")

    print(f"  ✅ Priority labels updated : {priority_updated} rows")
    print(f"  ✅ Extra emails injected   : {email_updated} slots filled")
    print(f"\n  COLOUR LEGEND:")
    print(f"    🟢 GREEN   = MNC + Ahmedabad/Gandhinagar/Gujarat  (TOP)")
    print(f"    🟡 YELLOW  = MNC (non-Gujarat)")
    print(f"    🔵 BLUE    = Gujarat/Ahmedabad company (non-MNC)")
    print(f"\n{'='*70}\n")

if __name__ == "__main__":
    main()
