"""
Comonk AI — HR Email Fetcher
Fetches emails for AI/ML companies using Hunter.io and adds to Excel sheet.
"""

import openpyxl
import httpx
import time
from urllib.parse import urlparse

HUNTER_KEY = "86b42563dd92b15fb26268bdc0f9a697c7609d38"
EXCEL_PATH = "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"

# Top MNC / AI companies to prioritize (fetch first)
PRIORITY_KEYWORDS = [
    "infostretch", "apexon", "tatvasoft", "bosch", "kpit", "l&t", "tata",
    "wipro", "infosys", "cognizant", "accenture", "rapidops", "valuecoders",
    "bacancy", "space-o", "cmarix", "zignuts", "esparkbiz", "tops info",
    "icoderz", "mobisoft", "shreeji", "wdp", "rootsquares", "softvan",
    "syntel", "capgemini", "tech mahindra", "hcl", "mphasis", "hexaware",
    "persistent", "cyient", "zensar", "mastech", "niit", "3i infotech"
]

def extract_domain(url):
    if not url:
        return None
    try:
        url = url.strip()
        if not url.startswith("http"):
            url = "https://" + url
        parsed = urlparse(url)
        domain = parsed.netloc.replace("www.", "").strip()
        return domain if "." in domain else None
    except:
        return None

def hunter_fetch(domain):
    """Fetch emails for a domain using Hunter.io"""
    try:
        r = httpx.get(
            "https://api.hunter.io/v2/domain-search",
            params={"domain": domain, "api_key": HUNTER_KEY, "limit": 5},
            timeout=10
        )
        data = r.json()
        if "data" in data and "emails" in data["data"]:
            emails = [e["value"] for e in data["data"]["emails"]
                      if e.get("value") and ("hr" in e["value"].lower()
                         or "recruit" in e["value"].lower()
                         or "career" in e["value"].lower()
                         or "talent" in e["value"].lower()
                         or "job" in e["value"].lower()
                         or e.get("type") == "generic")]
            # If no HR-specific emails, take all found emails
            if not emails:
                emails = [e["value"] for e in data["data"]["emails"][:3]]
            return emails[:4]  # max 4 new emails
        elif "errors" in data:
            print(f"  Hunter error: {data['errors']}")
        return []
    except Exception as ex:
        print(f"  Request failed: {ex}")
        return []

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Headers are on row 3
    # Col: 1=No, 2=Company, 3=Category, 4=Roles, 5=Email1, 6=Email2, 7=Email3, 8=Email4, 9=Email5, 10=Phone, 11=Website
    COL_NAME     = 2
    COL_CATEGORY = 3
    COL_EMAIL1   = 5
    COL_EMAIL2   = 6
    COL_EMAIL3   = 7
    COL_EMAIL4   = 8
    COL_EMAIL5   = 9
    COL_WEBSITE  = 11

    # Check Hunter quota first
    quota_r = httpx.get(
        "https://api.hunter.io/v2/account",
        params={"api_key": HUNTER_KEY}, timeout=8
    ).json()
    if "data" in quota_r:
        searches = quota_r["data"].get("searches", {})
        used      = searches.get("used", "?")
        available = searches.get("available", "?")
        remaining = int(available) - int(used) if str(available).isdigit() and str(used).isdigit() else available
        print(f"\n Hunter.io quota: {used} used / {available} available — {remaining} remaining\n")
        if isinstance(remaining, int) and remaining < 2:
            print(" QUOTA EXHAUSTED. Wait for monthly reset.")
            return
    else:
        print(" Could not check quota — proceeding anyway.\n")

    # Collect companies to process
    rows_to_process = []
    for r in range(4, ws.max_row + 1):
        name     = ws.cell(r, COL_NAME).value
        category = ws.cell(r, COL_CATEGORY).value or ""
        website  = ws.cell(r, COL_WEBSITE).value
        email1   = ws.cell(r, COL_EMAIL1).value

        if not name or not website:
            continue

        name_lower = str(name).lower()
        cat_lower  = category.lower()
        is_ai      = "ai" in cat_lower or "ml" in cat_lower or "data" in cat_lower or "machine" in cat_lower
        is_priority = any(kw in name_lower for kw in PRIORITY_KEYWORDS)
        has_no_email = not email1

        if is_ai or is_priority or has_no_email:
            rows_to_process.append((r, name, category, website, email1, is_priority))

    # Sort: priority companies first, then AI/ML, then others
    rows_to_process.sort(key=lambda x: (0 if x[5] else 1, x[2]))

    print(f" Found {len(rows_to_process)} companies to process")
    print(f" Processing top 20 (Hunter.io free = 25/month)\n")
    rows_to_process = rows_to_process[:20]

    fetched = 0
    updated = 0

    for r, name, category, website, existing_email, _ in rows_to_process:
        domain = extract_domain(website)
        if not domain:
            print(f"  SKIP {name} - no valid domain")
            continue

        print(f"  >> {name} ({domain})", end=" ... ", flush=True)
        emails = hunter_fetch(domain)
        fetched += 1

        if emails:
            # Get existing emails to avoid duplicates
            existing = set(filter(None, [
                ws.cell(r, c).value for c in [COL_EMAIL1, COL_EMAIL2, COL_EMAIL3, COL_EMAIL4, COL_EMAIL5]
            ]))
            new_emails = [e for e in emails if e not in existing]

            if new_emails:
                # Fill empty email columns
                empty_cols = [c for c in [COL_EMAIL1, COL_EMAIL2, COL_EMAIL3, COL_EMAIL4, COL_EMAIL5]
                              if not ws.cell(r, c).value]
                for i, col in enumerate(empty_cols[:len(new_emails)]):
                    ws.cell(r, col).value = new_emails[i]
                print(f"ADDED {len(new_emails)} email(s): {', '.join(new_emails)}")
                updated += 1
            else:
                print(f"OK - {len(emails)} email(s) already in sheet")
        else:
            print("NOT FOUND")

        time.sleep(1.2)  # Rate limit: be polite to Hunter.io

    wb.save(EXCEL_PATH)
    print(f"\n Done! Fetched from {fetched} companies, updated {updated} rows.")
    print(f" Saved to: {EXCEL_PATH}")

if __name__ == "__main__":
    main()
