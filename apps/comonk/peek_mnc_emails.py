import openpyxl

wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
ws = wb["COMPLETE_MASTER"]

headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i+1 for i, h in enumerate(headers) if h}
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]

for target_row in [131, 304, 140, 65, 74]:  # TCS, Infosys, Wipro, EY, HCL
    name = ws.cell(target_row, 2).value
    print(f"\n{name} (row {target_row}):")
    for c in email_cols:
        v = ws.cell(target_row, c).value
        if v:
            print(f"  {v}")
