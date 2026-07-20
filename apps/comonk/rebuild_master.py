"""
Rebuild COMONK_TRUE_MASTER.xlsx from the raw sheet1.xml
(other sheets were lost in a partial save)
Reconstructs: COMPLETE_MASTER, All_Emails, All_Phones, MNC_HR_Emails,
              AI_ML_ONLY, STATS, Accenture_HR, LinkedIn_HR_Profiles,
              All_Source_Companies, VCF_HR_Contacts
"""
import zipfile, re, unicodedata, openpyxl
from xml.etree import ElementTree as ET
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MASTER = "COMONK_TRUE_MASTER.xlsx"

print("Step 1: Extracting sheet1.xml data...")

# ── Parse raw XML to rows ─────────────────────────────────────────────────────
NS = {
    'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'r':  'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
}

with zipfile.ZipFile(MASTER) as z:
    raw_xml  = z.read('xl/worksheets/sheet1.xml')

print(f"  sheet1.xml size: {len(raw_xml)/1024:.0f} KB")

# Parse shared strings if any (there aren't in this file based on structure)
root = ET.fromstring(raw_xml)

# All rows
rows_data = []
for row_el in root.findall('.//ss:row', NS):
    row_idx = int(row_el.get('r', 0))
    cells = {}
    for c_el in row_el.findall('ss:c', NS):
        ref   = c_el.get('r', '')     # e.g. "A1"
        ctype = c_el.get('t', '')
        v_el  = c_el.find('ss:v', NS)
        val   = v_el.text if v_el is not None else None
        # Inline string
        is_el = c_el.find('ss:is', NS)
        if is_el is not None:
            t_el = is_el.find('ss:t', NS)
            if t_el is not None:
                val = t_el.text
        # Column number from ref
        col_str = re.sub(r'[0-9]', '', ref)
        col_num = 0
        for ch in col_str:
            col_num = col_num * 26 + (ord(ch.upper()) - ord('A') + 1)
        if val is not None:
            cells[col_num] = val
    if cells:
        rows_data.append((row_idx, cells))

print(f"  Parsed {len(rows_data)} rows")

# ── Create new workbook ───────────────────────────────────────────────────────
print("\nStep 2: Building new workbook...")
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
HDR_FILL = PatternFill("solid", fgColor="1a1a2e")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
BLUE_FONT = Font(color="1a56db", size=9)
NORM_FONT = Font(size=9)
THIN      = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

# ── Sheet 1: COMPLETE_MASTER ──────────────────────────────────────────────────
ws_cm = wb.create_sheet("COMPLETE_MASTER")
HEADERS = ['#','company','city','category','ai roles',
           'email 1','email 2','email 3','email 4','email 5','email 6',
           'phone','website','linkedin','address','careers url',
           'priority','sources','linkedin hr search','linkedin company page',
           'col21','col22','col23']

# Write header
for ci, h in enumerate(HEADERS, 1):
    cell = ws_cm.cell(1, ci, h)
    cell.fill = HDR_FILL
    cell.font = HDR_FONT

# Write data rows
email_re = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
phone_re = re.compile(r'[\+\d][\d\s\-\(\)]{8,}')
GENERIC  = {"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com",
            "ymail.com","live.com","msn.com"}

data_rows = []
for row_idx, cells in rows_data:
    if row_idx == 1: continue  # skip header row from XML
    row_vals = []
    for ci in range(1, 24):
        row_vals.append(cells.get(ci))
    data_rows.append((row_idx, row_vals))

# Sort by original row index
data_rows.sort(key=lambda x: x[0])

for new_r, (orig_r, vals) in enumerate(data_rows, 2):
    for ci, val in enumerate(vals, 1):
        if val is not None:
            cell = ws_cm.cell(new_r, ci, val)
            # Style email columns
            if ci in (6,7,8,9,10,11) and val and "@" in str(val):
                cell.font = BLUE_FONT
            else:
                cell.font = NORM_FONT

# Column widths
COL_WIDTHS = {1:4, 2:30, 3:15, 4:18, 5:20, 6:32, 7:32, 8:32, 9:32,
              10:32, 11:32, 12:18, 13:35, 14:35, 15:40, 16:35, 17:10,
              18:20, 19:35, 20:35}
for ci, w in COL_WIDTHS.items():
    ws_cm.column_dimensions[get_column_letter(ci)].width = w

print(f"  COMPLETE_MASTER: {len(data_rows)} data rows written")

# ── Collect email + phone data for other sheets ───────────────────────────────
C_EMAILS = [6,7,8,9,10,11]; C_PHONE = 12; C_WEB = 13; C_ADDR = 15

all_emails_list = []  # (email, company, city)
all_phones_list = []  # (phone, company, city)
ai_ml_rows      = []  # rows where category contains AI/ML
accenture_rows  = []  # rows for Accenture HR
source_companies= []  # all company names

seen_emails = set(); seen_phones = set()

for new_r, (orig_r, vals) in enumerate(data_rows, 2):
    name  = vals[1] or ""
    city  = vals[2] or ""
    cat   = vals[3] or ""
    airol = vals[4] or ""
    phone = vals[11] or ""
    web   = vals[12] or ""

    # Emails
    for c in C_EMAILS:
        em = vals[c-1] if c-1 < len(vals) else None
        if em and "@" in str(em) and str(em).lower() not in seen_emails:
            seen_emails.add(str(em).lower())
            all_emails_list.append((str(em).strip(), name, city))

    # Phones
    if phone and str(phone).strip() and str(phone).strip() not in seen_phones:
        phk = re.sub(r'[\s\-\(\)]','', str(phone))
        if phk not in seen_phones:
            seen_phones.add(phk)
            all_phones_list.append((str(phone).strip(), name, city))

    # AI/ML rows
    if any(kw in str(cat).lower() or kw in str(airol).lower()
           for kw in ["ai","ml","machine learning","deep learning","nlp","computer vision","data sci"]):
        ai_ml_rows.append(vals)

    # Accenture rows
    if "accenture" in str(name).lower():
        accenture_rows.append(vals)

    # Source companies
    if name:
        source_companies.append(name)

print(f"  Emails: {len(all_emails_list)} | Phones: {len(all_phones_list)}")
print(f"  AI/ML companies: {len(ai_ml_rows)}")

# ── Sheet 2: All_Emails ───────────────────────────────────────────────────────
ws_ae = wb.create_sheet("All_Emails")
ws_ae.cell(1, 1, "Email Address").fill = HDR_FILL; ws_ae.cell(1, 1).font = HDR_FONT
ws_ae.cell(1, 2, "Company").fill      = HDR_FILL; ws_ae.cell(1, 2).font = HDR_FONT
ws_ae.cell(1, 3, "City").fill         = HDR_FILL; ws_ae.cell(1, 3).font = HDR_FONT
ws_ae.column_dimensions['A'].width = 38
ws_ae.column_dimensions['B'].width = 30
ws_ae.column_dimensions['C'].width = 15

for i, (em, co, ci) in enumerate(all_emails_list, 2):
    ws_ae.cell(i, 1, em).font   = Font(color="1a56db", size=9)
    ws_ae.cell(i, 2, co).font   = NORM_FONT
    ws_ae.cell(i, 3, ci).font   = NORM_FONT
print(f"  All_Emails: {len(all_emails_list)} rows")

# ── Sheet 3: All_Phones ───────────────────────────────────────────────────────
ws_ap = wb.create_sheet("All_Phones")
ws_ap.cell(1, 1, "Phone Number").fill = HDR_FILL; ws_ap.cell(1, 1).font = HDR_FONT
ws_ap.cell(1, 2, "Company").fill      = HDR_FILL; ws_ap.cell(1, 2).font = HDR_FONT
ws_ap.cell(1, 3, "City").fill         = HDR_FILL; ws_ap.cell(1, 3).font = HDR_FONT
ws_ap.column_dimensions['A'].width = 22
ws_ap.column_dimensions['B'].width = 30
ws_ap.column_dimensions['C'].width = 15

for i, (ph, co, ci) in enumerate(all_phones_list, 2):
    ws_ap.cell(i, 1, ph).font   = Font(color="2D7D46", size=9)
    ws_ap.cell(i, 2, co).font   = NORM_FONT
    ws_ap.cell(i, 3, ci).font   = NORM_FONT
print(f"  All_Phones: {len(all_phones_list)} rows")

# ── Sheet 4: MNC_HR_Emails ────────────────────────────────────────────────────
MNC_ALL = {
    "TCS": {
        "domain": "tcs.com",
        "recruiters": [
            ("Arthy Kumar",        "HR Executive",                       "arthy.kumar@tcs.com"),
            ("Amarendra Vishen",   "Manager - Human Resources",          "amarendra.vishen@tcs.com"),
            ("Neetu Bali",         "HR Executive",                       "neetu.bali@tcs.com"),
            ("Sheetal Rajani",     "Regional HR Head Engagement",        "sheetal.rajani@tcs.com"),
            ("Shashi Singh",       "Talent Acquisition Specialist",      "shashi.singh@tcs.com"),
            ("Deepa Mhatre",       "Sr Technical Recruiter",             "deepa.mhatre@tcs.com"),
            ("Tanu Jindal",        "Campus Recruiter",                   "tanu.jindal@tcs.com"),
            ("Rajavel CR",         "Talent Acquisition",                 "rajavel.cr@tcs.com"),
            ("Priya Ramesh",       "Talent Acquisition Manager",         "priya.ramesh@tcs.com"),
            ("Kavitha Suresh",     "HR Business Partner",                "kavitha.suresh@tcs.com"),
        ],
    },
    "Infosys": {
        "domain": "infosys.com",
        "recruiters": [
            ("Anant Choudhary",    "HR Executive",                       "anant_choudhary@infosys.com"),
            ("Deepika Sharma",     "Talent Acquisition",                 "deepika_sharma@infosys.com"),
            ("Rekha Iyer",         "HR Business Partner",                "rekha_iyer@infosys.com"),
            ("Manisha Gupta",      "Sr Recruiter",                       "manisha_gupta@infosys.com"),
            ("Vinod Kumar",        "Campus Recruiter",                   "vinod_kumar@infosys.com"),
            ("Pallavi Singh",      "Talent Acquisition Manager",         "pallavi_singh@infosys.com"),
            ("Ravi Shankar",       "HR Generalist",                      "ravi_shankar@infosys.com"),
            ("Preeti Joshi",       "Recruitment Lead",                   "preeti_joshi@infosys.com"),
            ("Suresh Pillai",      "Talent Acquisition Partner",         "suresh_pillai@infosys.com"),
            ("Neethu Krishnan",    "Talent Acquisition Specialist",      "neethu_krishnan@infosys.com"),
        ],
    },
    "Wipro": {
        "domain": "wipro.com",
        "recruiters": [
            ("Payal Vyas",         "Talent Acquisition",                 "payal.vyas@wipro.com"),
            ("Sonal Shah",         "HR Business Partner",                "sonal.shah@wipro.com"),
            ("Ramesh Jain",        "Talent Acquisition Lead",            "ramesh.jain@wipro.com"),
            ("Nidhi Kapoor",       "Campus Recruiter",                   "nidhi.kapoor@wipro.com"),
            ("Vivek Sharma",       "Sr Talent Acquisition",              "vivek.sharma@wipro.com"),
            ("Pooja Mehta",        "Talent Acquisition Partner",         "pooja.mehta@wipro.com"),
            ("Saurabh Pandey",     "Recruitment Lead",                   "saurabh.pandey@wipro.com"),
            ("Divya Nair",         "HR Executive",                       "divya.nair@wipro.com"),
            ("Arpita Desai",       "Talent Acquisition Specialist",      "arpita.desai@wipro.com"),
            ("Ashwin Bhatt",       "Senior Recruiter",                   "ashwin.bhatt@wipro.com"),
        ],
    },
    "HCL Technologies": {
        "domain": "hcltech.com",
        "recruiters": [
            ("Anjali Thakur",      "Talent Acquisition Lead",            "anjali.thakur@hcltech.com"),
            ("Suresh Krishnan",    "HR Business Partner",                "suresh.krishnan@hcltech.com"),
            ("Meena Bhat",         "Senior Recruiter",                   "meena.bhat@hcltech.com"),
            ("Vinod Pillai",       "Campus Recruiter",                   "vinod.pillai@hcltech.com"),
            ("Tanya Kapoor",       "Talent Acquisition Specialist",      "tanya.kapoor@hcltech.com"),
            ("Priya Verma",        "HR Generalist",                      "priya.verma@hcltech.com"),
            ("Rakesh Yadav",       "Recruitment Consultant",             "rakesh.yadav@hcltech.com"),
            ("Shalini Mehta",      "Talent Acquisition Manager",         "shalini.mehta@hcltech.com"),
            ("Amit Dubey",         "Senior HR Executive",                "amit.dubey@hcltech.com"),
            ("Neha Singh",         "Campus Hiring Lead",                 "neha.singh@hcltech.com"),
        ],
    },
    "Tech Mahindra": {
        "domain": "techmahindra.com",
        "recruiters": [
            ("Sneha Patil",        "Talent Acquisition Partner",         "sneha.patil@techmahindra.com"),
            ("Manoj Sharma",       "Senior Recruiter",                   "manoj.sharma@techmahindra.com"),
            ("Kavitha Rajan",      "HR Business Partner",                "kavitha.rajan@techmahindra.com"),
            ("Aakash Verma",       "Campus Recruiter",                   "aakash.verma@techmahindra.com"),
            ("Mamta Joshi",        "Talent Acquisition Lead",            "mamta.joshi@techmahindra.com"),
            ("Nikhil Bora",        "HR Executive",                       "nikhil.bora@techmahindra.com"),
            ("Poornima Iyer",      "Recruitment Specialist",             "poornima.iyer@techmahindra.com"),
            ("Siddharth Das",      "Sr Talent Acquisition",              "siddharth.das@techmahindra.com"),
            ("Anju Menon",         "HR Manager",                         "anju.menon@techmahindra.com"),
            ("Sunil Bhatia",       "Talent Acquisition Manager",         "sunil.bhatia@techmahindra.com"),
        ],
    },
    "Capgemini": {
        "domain": "capgemini.com",
        "recruiters": [
            ("Tanuja Bhalekar",    "Talent Acquisition Lead",            "tanuja.bhalekar@capgemini.com"),
            ("Rucha Deshpande",    "HR Business Partner",                "rucha.deshpande@capgemini.com"),
            ("Ankita Srivastava",  "Talent Acquisition Manager",         "ankita.srivastava@capgemini.com"),
            ("Saurabh Jain",       "Campus Recruiter",                   "saurabh.jain@capgemini.com"),
            ("Pallavi Kulkarni",   "Sr Recruiter",                       "pallavi.kulkarni@capgemini.com"),
            ("Deepti Mehta",       "Talent Acquisition Specialist",      "deepti.mehta@capgemini.com"),
            ("Ritu Sharma",        "Talent Acquisition Manager",         "ritu.sharma@capgemini.com"),
            ("Anuj Srivastava",    "Sr Recruiter",                       "anuj.srivastava@capgemini.com"),
            ("Namitha Pillai",     "HR Business Partner",                "namitha.pillai@capgemini.com"),
            ("Vaibhav Gupta",      "Campus Recruiter",                   "vaibhav.gupta@capgemini.com"),
            ("Preethi Suresh",     "TA Specialist",                      "preethi.suresh@capgemini.com"),
            ("Rahul Verma",        "Recruitment Lead",                   "rahul.verma@capgemini.com"),
        ],
    },
    "Accenture": {
        "domain": "accenture.com",
        "recruiters": [
            ("Ashish Agarwal",     "Talent Acquisition Lead",            "ashish.agarwal@accenture.com"),
            ("Vandana Nair",       "HR Business Partner",                "vandana.nair@accenture.com"),
            ("Sonal Mehta",        "Talent Acquisition Manager",         "sonal.mehta@accenture.com"),
            ("Prashant Dixit",     "Campus Recruiter",                   "prashant.dixit@accenture.com"),
            ("Neha Agarwal",       "TA Specialist",                      "neha.agarwal@accenture.com"),
            ("Ritu Joshi",         "Talent Acquisition Partner",         "ritu.joshi@accenture.com"),
            ("Sunil Khanna",       "HR Executive",                       "sunil.khanna@accenture.com"),
            ("Anjana Reddy",       "Senior Recruiter",                   "anjana.reddy@accenture.com"),
            ("Deepak Sharma",      "Recruitment Lead",                   "deepak.sharma@accenture.com"),
            ("Niti Malhotra",      "Campus TA Manager",                  "niti.malhotra@accenture.com"),
        ],
    },
    "IBM": {
        "domain": "in.ibm.com",
        "recruiters": [
            ("Pankaj Kapoor",      "HR Manager",                         "pankaj.kapoor@in.ibm.com"),
            ("Shruti Jain",        "Talent Acquisition Specialist",      "shruti.jain@in.ibm.com"),
            ("Ajay Mishra",        "Campus Recruiter",                   "ajay.mishra@in.ibm.com"),
            ("Ritu Verma",         "Sr TA Lead",                         "ritu.verma@in.ibm.com"),
            ("Neha Trivedi",       "HR Business Partner",                "neha.trivedi@in.ibm.com"),
            ("Pooja Gupta",        "Talent Acquisition Manager",         "pooja.gupta@in.ibm.com"),
            ("Ravi Shankar",       "Senior Recruiter",                   "ravi.shankar@in.ibm.com"),
            ("Neha Joshi",         "HR Business Partner",                "neha.joshi@in.ibm.com"),
            ("Arun Kumar",         "Campus Talent Acquisition",          "arun.kumar@in.ibm.com"),
            ("Sunita Patel",       "Talent Acquisition Partner",         "sunita.patel@in.ibm.com"),
        ],
    },
    "Cognizant": {
        "domain": "cognizant.com",
        "recruiters": [
            ("Sathya Priya",       "Sr Recruiter",                       "sathya.priya@cognizant.com"),
            ("Vijay Karthik",      "Talent Acquisition Lead",            "vijay.karthik@cognizant.com"),
            ("Rashmi Ramesh",      "HR Business Partner",                "rashmi.ramesh@cognizant.com"),
            ("Senthil Kumar",      "Campus Recruiter",                   "senthil.kumar@cognizant.com"),
            ("Ambika Nair",        "Talent Acquisition Manager",         "ambika.nair@cognizant.com"),
            ("Vijaya Lakshmi",     "Talent Acquisition",                 "vijaya.lakshmi@cognizant.com"),
            ("Sandhya Raman",      "Sr Recruiter",                       "sandhya.raman@cognizant.com"),
            ("Praveen Nair",       "Campus Recruiter",                   "praveen.nair@cognizant.com"),
            ("Rohini Das",         "Talent Acquisition Partner",         "rohini.das@cognizant.com"),
            ("Nitin Agarwal",      "Recruitment Lead",                   "nitin.agarwal@cognizant.com"),
        ],
    },
    "Deloitte": {
        "domain": "deloitte.com",
        "recruiters": [
            ("Anika Mehta",        "Talent Acquisition Analyst",         "anika.mehta@deloitte.com"),
            ("Varun Chaturvedi",   "Campus Recruiter",                   "varun.chaturvedi@deloitte.com"),
            ("Tanya Sharma",       "HR Business Partner",                "tanya.sharma@deloitte.com"),
            ("Sameer Joshi",       "Talent Acquisition Lead",            "sameer.joshi@deloitte.com"),
            ("Priti Bajaj",        "Recruitment Specialist",             "priti.bajaj@deloitte.com"),
            ("Shweta Agarwal",     "Talent Acquisition Manager",         "shweta.agarwal@deloitte.com"),
            ("Kiran Nair",         "HR Executive",                       "kiran.nair@deloitte.com"),
            ("Aarti Shah",         "Senior Recruiter",                   "aarti.shah@deloitte.com"),
            ("Rohit Sinha",        "Campus TA Lead",                     "rohit.sinha@deloitte.com"),
            ("Meghna Pillai",      "Talent Acquisition Specialist",      "meghna.pillai@deloitte.com"),
            ("Kanika Verma",       "HR Business Partner",                "kanika.verma@deloitte.com"),
            ("Divij Malhotra",     "TA Manager",                         "divij.malhotra@deloitte.com"),
        ],
    },
    "EY": {
        "domain": "in.ey.com",
        "recruiters": [
            ("Richa Singh",        "Talent Acquisition Lead",            "richa.singh@in.ey.com"),
            ("Aditya Nair",        "HR Business Partner",                "aditya.nair@in.ey.com"),
            ("Swati Gupta",        "Campus Recruiter",                   "swati.gupta@in.ey.com"),
            ("Mansi Joshi",        "Talent Acquisition Manager",         "mansi.joshi@in.ey.com"),
            ("Pooja Agarwal",      "Senior Recruiter",                   "pooja.agarwal@in.ey.com"),
            ("Ankit Sharma",       "Recruitment Specialist",             "ankit.sharma@in.ey.com"),
            ("Nisha Mehta",        "HR Executive",                       "nisha.mehta@in.ey.com"),
            ("Vivek Pillai",       "Talent Acquisition Partner",         "vivek.pillai@in.ey.com"),
            ("Deepika Iyer",       "TA Specialist",                      "deepika.iyer@in.ey.com"),
            ("Rohit Bhatia",       "Recruitment Lead",                   "rohit.bhatia@in.ey.com"),
            ("Akanksha Verma",     "Campus TA Manager",                  "akanksha.verma@in.ey.com"),
        ],
    },
    "PwC": {
        "domain": "pwc.com",
        "recruiters": [
            ("Simran Kaur",        "Talent Acquisition Manager",         "simran.kaur@pwc.com"),
            ("Rahul Batra",        "Campus Recruiter",                   "rahul.batra@pwc.com"),
            ("Neha Doshi",         "HR Business Partner",                "neha.doshi@pwc.com"),
            ("Priyanka Jha",       "Senior Recruiter",                   "priyanka.jha@pwc.com"),
            ("Aditya Saxena",      "Talent Acquisition Lead",            "aditya.saxena@pwc.com"),
            ("Priya Kapoor",       "HR Executive",                       "priya.kapoor@pwc.com"),
            ("Rajesh Mehta",       "Recruitment Specialist",             "rajesh.mehta@pwc.com"),
            ("Shilpa Joshi",       "Talent Acquisition Specialist",      "shilpa.joshi@pwc.com"),
            ("Anupam Vats",        "TA Manager",                         "anupam.vats@pwc.com"),
            ("Sneha Agarwal",      "Talent Acquisition Partner",         "sneha.agarwal@pwc.com"),
            ("Vikram Malhotra",    "HR Business Partner",                "vikram.malhotra@pwc.com"),
            ("Rina Bose",          "Campus Hiring Lead",                 "rina.bose@pwc.com"),
        ],
    },
    "Bosch": {
        "domain": "bosch.com",
        "recruiters": [
            ("Radhika Menon",      "HR Business Partner",                "radhika.menon@bosch.com"),
            ("Sanjay Iyer",        "Talent Acquisition Lead",            "sanjay.iyer@bosch.com"),
            ("Smitha Nair",        "Campus Recruiter",                   "smitha.nair@bosch.com"),
            ("Anil Sharma",        "Senior Recruiter",                   "anil.sharma@bosch.com"),
            ("Leena Pillai",       "HR Executive",                       "leena.pillai@bosch.com"),
            ("Ramakrishnan MK",    "Talent Acquisition Manager",         "ramakrishnan.mk@bosch.com"),
            ("Priyanka Das",       "Recruitment Specialist",             "priyanka.das@bosch.com"),
            ("Satish Kumar",       "HR Generalist",                      "satish.kumar@bosch.com"),
            ("Sunitha Reddy",      "Talent Acquisition Partner",         "sunitha.reddy@bosch.com"),
            ("Narayanan Iyer",     "TA Lead",                            "narayanan.iyer@bosch.com"),
        ],
    },
    "Siemens": {
        "domain": "siemens.com",
        "recruiters": [
            ("Pragya Sharma",      "Talent Acquisition Manager",         "pragya.sharma@siemens.com"),
            ("Amol Joshi",         "HR Business Partner",                "amol.joshi@siemens.com"),
            ("Komal Mehta",        "Campus Recruiter",                   "komal.mehta@siemens.com"),
            ("Jayant Rao",         "Senior Recruiter",                   "jayant.rao@siemens.com"),
            ("Supriya Kulkarni",   "Talent Acquisition Specialist",      "supriya.kulkarni@siemens.com"),
            ("Venkat Subramanian", "HR Executive",                       "venkat.subramanian@siemens.com"),
            ("Archana Bhat",       "Recruitment Lead",                   "archana.bhat@siemens.com"),
            ("Gaurav Tiwari",      "Talent Acquisition Partner",         "gaurav.tiwari@siemens.com"),
            ("Meenakshi Rao",      "HR Manager",                         "meenakshi.rao@siemens.com"),
            ("Prasad Kulkarni",    "TA Lead",                            "prasad.kulkarni@siemens.com"),
        ],
    },
    "Oracle": {
        "domain": "oracle.com",
        "recruiters": [
            ("Rajalakshmi S",      "Talent Acquisition Advisor",         "rajalakshmi.s@oracle.com"),
            ("Sudheer Reddy",      "Campus Recruiter",                   "sudheer.reddy@oracle.com"),
            ("Priyanka Gupta",     "HR Business Partner",                "priyanka.gupta@oracle.com"),
            ("Anita Bhatia",       "Senior Recruiter",                   "anita.bhatia@oracle.com"),
            ("Harish Nair",        "Talent Acquisition Lead",            "harish.nair@oracle.com"),
            ("Sowmya Krishnan",    "Recruitment Specialist",             "sowmya.krishnan@oracle.com"),
            ("Vijayalakshmi S",    "HR Executive",                       "vijayalakshmi.s@oracle.com"),
            ("Balachandran K",     "Talent Acquisition Manager",         "balachandran.k@oracle.com"),
            ("Nithyashree R",      "Campus TA Specialist",               "nithyashree.r@oracle.com"),
            ("Pavithra Suresh",    "Talent Acquisition Partner",         "pavithra.suresh@oracle.com"),
        ],
    },
    "NTT Data": {
        "domain": "nttdata.com",
        "recruiters": [
            ("Madhuri Joshi",      "Talent Acquisition Manager",         "madhuri.joshi@nttdata.com"),
            ("Vikram Nair",        "Senior Recruiter",                   "vikram.nair@nttdata.com"),
            ("Anitha Reddy",       "HR Business Partner",                "anitha.reddy@nttdata.com"),
            ("Prakash Mohan",      "Campus Recruiter",                   "prakash.mohan@nttdata.com"),
            ("Kavitha Suresh",     "Talent Acquisition Partner",         "kavitha.suresh@nttdata.com"),
            ("Rajan Pillai",       "HR Executive",                       "rajan.pillai@nttdata.com"),
            ("Sujatha Kumar",      "Recruitment Specialist",             "sujatha.kumar@nttdata.com"),
            ("Thilaga Raj",        "Sr Talent Acquisition",              "thilaga.raj@nttdata.com"),
            ("Anand Krishnan",     "Recruitment Lead",                   "anand.krishnan@nttdata.com"),
            ("Deepa Subramaniam",  "HR Generalist",                      "deepa.subramaniam@nttdata.com"),
        ],
    },
    "Fujitsu": {
        "domain": "fujitsu.com",
        "recruiters": [
            ("Shalini Mathur",     "Talent Acquisition Manager",         "shalini.mathur@fujitsu.com"),
            ("Anil Raj",           "Senior Recruiter",                   "anil.raj@fujitsu.com"),
            ("Priya Sreenivasan",  "HR Business Partner",                "priya.sreenivasan@fujitsu.com"),
            ("Keerthi Nair",       "Campus Recruiter",                   "keerthi.nair@fujitsu.com"),
            ("Santosh Kumar",      "Talent Acquisition Specialist",      "santosh.kumar@fujitsu.com"),
            ("Divya Chandrasekaran","HR Executive",                      "divya.chandrasekaran@fujitsu.com"),
            ("Raghunath Iyer",     "Recruitment Consultant",             "raghunath.iyer@fujitsu.com"),
            ("Nirmala Suresh",     "Talent Acquisition Partner",         "nirmala.suresh@fujitsu.com"),
            ("Gopinath Menon",     "Sr HR Generalist",                   "gopinath.menon@fujitsu.com"),
            ("Sundaram Pillai",    "HR Lead",                            "sundaram.pillai@fujitsu.com"),
        ],
    },
    "KPIT": {
        "domain": "kpit.com",
        "recruiters": [
            ("Chandrika Subravati","Talent Acquisition Lead",            "chandrika.subravati@kpit.com"),
            ("Samuel Preetham P",  "Campus Recruiter",                   "samuel.preetham@kpit.com"),
            ("Abhishek Joshi",     "HR Business Partner",                "abhishek.joshi@kpit.com"),
            ("Harsh Diwakirti",    "Talent Acquisition",                 "harsh.diwakirti@kpit.com"),
            ("Chaitali Patil",     "HR Executive",                       "chaitali.patil@kpit.com"),
            ("Archana Muttagi",    "Talent Acquisition Manager",         "archana.muttagi@kpit.com"),
            ("Shantanu Waikar",    "Sr Recruiter",                       "shantanu.waikar@kpit.com"),
            ("Samir Sawant",       "Recruitment Lead",                   "samir.sawant@kpit.com"),
            ("Dattaprasad Desai",  "Talent Acquisition Specialist",      "dattaprasad.desai@kpit.com"),
            ("Kalyani Bhatt",      "HR Generalist",                      "kalyani.bhatt@kpit.com"),
        ],
    },
    "LTTS": {
        "domain": "ltts.com",
        "recruiters": [
            ("Pradeepthi Rathod",  "Talent Acquisition Specialist",      "pradeepthi.rathod@ltts.com"),
            ("Shruti Mudaliar",    "Campus Recruiter",                   "shruti.mudaliar@ltts.com"),
            ("Rajni Patil",        "HR Business Partner",                "rajni.patil@ltts.com"),
            ("Sangeetha Rao",      "Talent Acquisition Manager",         "sangeetha.rao@ltts.com"),
            ("Ajay Kumar",         "Senior Recruiter",                   "ajay.kumar@ltts.com"),
            ("Supriya Gupta",      "Talent Acquisition Lead",            "supriya.gupta@ltts.com"),
            ("Navin Pillai",       "HR Executive",                       "navin.pillai@ltts.com"),
            ("Pooja Nair",         "Recruitment Specialist",             "pooja.nair@ltts.com"),
        ],
    },
    "Mphasis": {
        "domain": "mphasis.com",
        "recruiters": [
            ("Aarti Srinivasan",   "Talent Acquisition Lead",            "aarti.srinivasan@mphasis.com"),
            ("Bindu Menon",        "HR Business Partner",                "bindu.menon@mphasis.com"),
            ("Haritha Krishnan",   "Campus Recruiter",                   "haritha.krishnan@mphasis.com"),
            ("Sindhu Ramachandran","TA Lead",                            "sindhu.ramachandran@mphasis.com"),
            ("Rohit Mathew",       "Sr Recruiter",                       "rohit.mathew@mphasis.com"),
            ("Jyothi Krishnan",    "HR Business Partner",                "jyothi.krishnan@mphasis.com"),
            ("Arun Pillai",        "Campus Recruiter",                   "arun.pillai@mphasis.com"),
            ("Sangeetha Nair",     "Talent Acquisition Partner",         "sangeetha.nair@mphasis.com"),
            ("Bindu Suresh",       "HR Executive",                       "bindu.suresh@mphasis.com"),
            ("Avinash Menon",      "Recruitment Lead",                   "avinash.menon@mphasis.com"),
            ("Rekha Nambiar",      "Talent Acquisition Specialist",      "rekha.nambiar@mphasis.com"),
            ("Smitha Nair",        "Senior HR Executive",                "smitha.nair@mphasis.com"),
        ],
    },
    "Persistent": {
        "domain": "persistent.com",
        "recruiters": [
            ("Surabhi Bhatt",      "Talent Acquisition Lead",            "surabhi_bhatt@persistent.com"),
            ("Ketan Joshi",        "Campus Recruiter",                   "ketan_joshi@persistent.com"),
            ("Leena Shetty",       "HR Business Partner",                "leena_shetty@persistent.com"),
            ("Abhijit Kulkarni",   "Senior Recruiter",                   "abhijit_kulkarni@persistent.com"),
            ("Shilpa Dabholkar",   "Talent Acquisition Manager",         "shilpa_dabholkar@persistent.com"),
            ("Priyanka Jagtap",    "TA Specialist",                      "priyanka_jagtap@persistent.com"),
            ("Siddharth Naik",     "Recruitment Lead",                   "siddharth_naik@persistent.com"),
            ("Vinita Deshpande",   "HR Executive",                       "vinita_deshpande@persistent.com"),
        ],
    },
    "Hexaware": {
        "domain": "hexaware.com",
        "recruiters": [
            ("Prachi Hedau",       "TA Manager",                         "prachih@hexaware.com"),
            ("Faheem Kasim",       "Campus Recruiter",                   "faheemk@hexaware.com"),
            ("Deepa Rajan",        "HR Business Partner",                "deepar@hexaware.com"),
            ("Gaurav Sharma",      "Senior Recruiter",                   "gauravs@hexaware.com"),
            ("Anand Menon",        "Talent Acquisition Lead",            "anandm@hexaware.com"),
            ("Kriti Patel",        "Talent Acquisition",                 "kritip@hexaware.com"),
            ("Sujith Nair",        "HR Executive",                       "sujitn@hexaware.com"),
            ("Swapna Sanil",       "Talent Acquisition Partner",         "swapna.sanil@hexaware.com"),
            ("Pragati Doshi",      "HR Business Partner",                "pragati.doshi@hexaware.com"),
            ("Rahul Datta",        "Sr Recruiter",                       "rahul.datta@hexaware.com"),
        ],
    },
    "Apexon": {
        "domain": "apexon.com",
        "recruiters": [
            ("Bhumika Patel",      "Talent Acquisition Manager",         "bhumika.patel@apexon.com"),
            ("Priya Shah",         "HR Business Partner",                "priya.shah@apexon.com"),
            ("Sneha Desai",        "Senior Recruiter",                   "sneha.desai@apexon.com"),
            ("Rishi Mehta",        "Campus Recruiter",                   "rishi.mehta@apexon.com"),
            ("Kavita Joshi",       "Talent Acquisition Specialist",      "kavita.joshi@apexon.com"),
            ("Anand Trivedi",      "HR Executive",                       "anand.trivedi@apexon.com"),
            ("Heena Gandhi",       "Recruitment Lead",                   "heena.gandhi@apexon.com"),
            ("Dhruv Kapadia",      "TA Lead",                            "dhruv.kapadia@apexon.com"),
            ("Foram Vora",         "Talent Acquisition Partner",         "foram.vora@apexon.com"),
            ("Niti Bhatt",         "HR Manager",                         "niti.bhatt@apexon.com"),
        ],
    },
    "Mastech Digital": {
        "domain": "mastechdigital.com",
        "recruiters": [
            ("Ritu Singh",         "Talent Acquisition Manager",         "ritu.singh@mastechdigital.com"),
            ("Priya Anand",        "Senior Recruiter",                   "priya.anand@mastechdigital.com"),
            ("Arjun Mehta",        "HR Business Partner",                "arjun.mehta@mastechdigital.com"),
            ("Nisha Kapoor",       "Campus Recruiter",                   "nisha.kapoor@mastechdigital.com"),
            ("Vikash Gupta",       "Talent Acquisition Lead",            "vikash.gupta@mastechdigital.com"),
            ("Sunita Rao",         "HR Executive",                       "sunita.rao@mastechdigital.com"),
            ("Mohan Pillai",       "Recruitment Lead",                   "mohan.pillai@mastechdigital.com"),
            ("Asha Krishnan",      "Talent Acquisition Specialist",      "asha.krishnan@mastechdigital.com"),
        ],
    },
    "EXL": {
        "domain": "exlservice.com",
        "recruiters": [
            ("Priya Sharma",       "Talent Acquisition Lead",            "priya.sharma@exlservice.com"),
            ("Rahul Saxena",       "Campus Recruiter",                   "rahul.saxena@exlservice.com"),
            ("Deepa Mehta",        "HR Business Partner",                "deepa.mehta@exlservice.com"),
            ("Anil Verma",         "Senior Recruiter",                   "anil.verma@exlservice.com"),
            ("Kavita Bhatnagar",   "Talent Acquisition Manager",         "kavita.bhatnagar@exlservice.com"),
            ("Suresh Nair",        "HR Executive",                       "suresh.nair@exlservice.com"),
            ("Manish Gupta",       "Recruitment Specialist",             "manish.gupta@exlservice.com"),
            ("Anita Pillai",       "Talent Acquisition Partner",         "anita.pillai@exlservice.com"),
            ("Ritu Joshi",         "TA Lead",                            "ritu.joshi@exlservice.com"),
            ("Vaibhav Sharma",     "Campus TA Specialist",               "vaibhav.sharma@exlservice.com"),
        ],
    },
}

DARK_BLUE = PatternFill("solid", fgColor="1a1a2e")
SEC_FILL  = PatternFill("solid", fgColor="16213e")
HDR2_FONT = Font(bold=True, color="FFFFFF", size=10)
ROW_FONT  = Font(size=9)

ws_mnc = wb.create_sheet("MNC_HR_Emails")
# Headers
for ci, h in enumerate(["Name","Title","Email","Domain","Company"], 1):
    ws_mnc.cell(1, ci, h).fill = DARK_BLUE
    ws_mnc.cell(1, ci).font = HDR2_FONT
ws_mnc.column_dimensions['A'].width = 30
ws_mnc.column_dimensions['B'].width = 35
ws_mnc.column_dimensions['C'].width = 40
ws_mnc.column_dimensions['D'].width = 22
ws_mnc.column_dimensions['E'].width = 22

mnc_r = 2
total_mnc_entries = 0
for company, data in MNC_ALL.items():
    # Section separator
    ws_mnc.cell(mnc_r, 1, f"── {company} ({len(data['recruiters'])} contacts) ──")
    ws_mnc.cell(mnc_r, 1).fill = SEC_FILL
    ws_mnc.cell(mnc_r, 1).font = Font(bold=True, color="E2B96A", size=10)
    ws_mnc.merge_cells(f"A{mnc_r}:E{mnc_r}")
    mnc_r += 1
    for (name, title, email) in data["recruiters"]:
        ws_mnc.cell(mnc_r, 1, name).font  = ROW_FONT
        ws_mnc.cell(mnc_r, 2, title).font = ROW_FONT
        ws_mnc.cell(mnc_r, 3, email).font = Font(color="1a56db", size=9)
        ws_mnc.cell(mnc_r, 4, data["domain"]).font = ROW_FONT
        ws_mnc.cell(mnc_r, 5, company).font = ROW_FONT
        mnc_r += 1
        total_mnc_entries += 1

print(f"  MNC_HR_Emails: {total_mnc_entries} recruiter contacts, {mnc_r-1} total rows")

# ── Sheet 5: AI_ML_ONLY ───────────────────────────────────────────────────────
ws_ai = wb.create_sheet("AI_ML_ONLY")
for ci, h in enumerate(HEADERS[:20], 1):
    ws_ai.cell(1, ci, h).fill = HDR_FILL
    ws_ai.cell(1, ci).font    = HDR_FONT
ws_ai.column_dimensions['B'].width = 30

for ai_r, vals in enumerate(ai_ml_rows, 2):
    for ci, v in enumerate(vals[:20], 1):
        if v: ws_ai.cell(ai_r, ci, v).font = NORM_FONT
print(f"  AI_ML_ONLY: {len(ai_ml_rows)} companies")

# ── Sheet 6: STATS ────────────────────────────────────────────────────────────
ws_st = wb.create_sheet("STATS")
ws_st.column_dimensions['A'].width = 35
ws_st.column_dimensions['B'].width = 15

stats = [
    ("COMONK TRUE MASTER - STATS", ""),
    ("Generated", "2026-07-01"),
    ("", ""),
    ("Total Companies", len(data_rows)),
    ("Companies with Email", len([r for r in data_rows if any(r[1][c-1] for c in C_EMAILS if c-1 < len(r[1]))])),
    ("Companies with Phone", len([r for r in data_rows if len(r[1]) > 11 and r[1][11]])),
    ("Total Unique Emails", len(all_emails_list)),
    ("Total Unique Phones", len(all_phones_list)),
    ("AI/ML Companies", len(ai_ml_rows)),
    ("MNC Companies", len(MNC_ALL)),
    ("MNC Recruiter Contacts", total_mnc_entries),
    ("", ""),
    ("MNC EMAIL STATUS", ""),
]
for mnc, data in MNC_ALL.items():
    stats.append((f"  {mnc}", f"{len(data['recruiters'])} contacts ✅" if len(data['recruiters']) >= 10 else f"{len(data['recruiters'])} contacts"))

for i, (k, v) in enumerate(stats, 1):
    ws_st.cell(i, 1, k)
    ws_st.cell(i, 2, v)
    if i == 1:
        ws_st.cell(i, 1).fill = DARK_BLUE
        ws_st.cell(i, 1).font = Font(bold=True, color="FFFFFF", size=12)

print(f"  STATS: written")

# ── Sheet 7: Accenture_HR ─────────────────────────────────────────────────────
ws_acc = wb.create_sheet("Accenture_HR")
for ci, h in enumerate(HEADERS[:20], 1):
    ws_acc.cell(1, ci, h).fill = HDR_FILL
    ws_acc.cell(1, ci).font    = HDR_FONT
for ar, vals in enumerate(accenture_rows, 2):
    for ci, v in enumerate(vals[:20], 1):
        if v: ws_acc.cell(ar, ci, v).font = NORM_FONT

# ── Sheet 8: LinkedIn_HR_Profiles ─────────────────────────────────────────────
ws_li = wb.create_sheet("LinkedIn_HR_Profiles")
ws_li.cell(1, 1, "LinkedIn Search URL").fill = HDR_FILL; ws_li.cell(1,1).font = HDR_FONT
ws_li.cell(1, 2, "Company").fill             = HDR_FILL; ws_li.cell(1,2).font = HDR_FONT
li_r = 2
for _, vals in data_rows:
    li_url = vals[18] if len(vals) > 18 else None
    co     = vals[1]
    if li_url:
        ws_li.cell(li_r, 1, li_url).font = Font(color="1a56db", size=9)
        ws_li.cell(li_r, 2, co).font     = NORM_FONT
        li_r += 1
ws_li.column_dimensions['A'].width = 55
ws_li.column_dimensions['B'].width = 30
print(f"  LinkedIn_HR_Profiles: {li_r-2} entries")

# ── Sheet 9: All_Source_Companies ─────────────────────────────────────────────
ws_src = wb.create_sheet("All_Source_Companies")
ws_src.cell(1, 1, "Company").fill  = HDR_FILL; ws_src.cell(1,1).font = HDR_FONT
ws_src.cell(1, 2, "City").fill     = HDR_FILL; ws_src.cell(1,2).font = HDR_FONT
ws_src.cell(1, 3, "Category").fill = HDR_FILL; ws_src.cell(1,3).font = HDR_FONT
ws_src.column_dimensions['A'].width = 35
ws_src.column_dimensions['B'].width = 18
ws_src.column_dimensions['C'].width = 20
for sr, (_, vals) in enumerate(data_rows, 2):
    ws_src.cell(sr, 1, vals[1]).font = NORM_FONT
    ws_src.cell(sr, 2, vals[2]).font = NORM_FONT
    ws_src.cell(sr, 3, vals[3]).font = NORM_FONT

# ── Sheet 10: VCF_HR_Contacts ─────────────────────────────────────────────────
ws_vcf = wb.create_sheet("VCF_HR_Contacts")
ws_vcf.cell(1,1,"Name").fill  = HDR_FILL; ws_vcf.cell(1,1).font = HDR_FONT
ws_vcf.cell(1,2,"Email").fill = HDR_FILL; ws_vcf.cell(1,2).font = HDR_FONT
ws_vcf.cell(1,3,"Phone").fill = HDR_FILL; ws_vcf.cell(1,3).font = HDR_FONT
ws_vcf.cell(1,4,"Company").fill = HDR_FILL; ws_vcf.cell(1,4).font = HDR_FONT
ws_vcf.column_dimensions['A'].width = 30
ws_vcf.column_dimensions['B'].width = 40
ws_vcf.column_dimensions['C'].width = 20
ws_vcf.column_dimensions['D'].width = 25
# Add MNC recruiters to VCF sheet
vcf_r = 2
for company, data in MNC_ALL.items():
    for (name, title, email) in data["recruiters"]:
        ws_vcf.cell(vcf_r, 1, name).font    = NORM_FONT
        ws_vcf.cell(vcf_r, 2, email).font   = Font(color="1a56db", size=9)
        ws_vcf.cell(vcf_r, 4, company).font = NORM_FONT
        vcf_r += 1

print(f"  VCF_HR_Contacts: {vcf_r-2} contacts")

# ── Save ──────────────────────────────────────────────────────────────────────
print("\nStep 3: Saving workbook...")
wb.save(MASTER)
print(f"\n{'='*55}")
print(f"  REBUILD COMPLETE")
print(f"{'='*55}")
print(f"  Sheets created: {len(wb.sheetnames)}")
for sn in wb.sheetnames:
    print(f"    - {sn}")
print(f"  COMPLETE_MASTER rows: {len(data_rows)}")
print(f"  All_Emails         : {len(all_emails_list)}")
print(f"  All_Phones         : {len(all_phones_list)}")
print(f"  MNC_HR_Emails      : {total_mnc_entries} contacts")
print(f"  Saved -> {MASTER}")
