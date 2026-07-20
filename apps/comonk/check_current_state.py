import openpyxl
from collections import Counter

wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
print("Sheets:", wb.sheetnames)

ws = wb["COMPLETE_MASTER"]
print(f"COMPLETE_MASTER rows: {ws.max_row - 1}")

if "MNC_HR_Emails" in wb.sheetnames:
    ws_mnc = wb["MNC_HR_Emails"]
    print(f"MNC_HR_Emails rows: {ws_mnc.max_row - 1}")
    counts = Counter()
    for r in range(2, ws_mnc.max_row + 1):
        co = ws_mnc.cell(r, 5).value
        em = ws_mnc.cell(r, 3).value
        if co and em and "@" in str(em):
            counts[str(co)] += 1
    print("\nMNC contact counts (from MNC_HR_Emails):")
    for co, cnt in counts.most_common(30):
        print(f"  {cnt:>4}  {co}")
else:
    print("No MNC_HR_Emails sheet found!")

# Check row order for Gujarat priority sort
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("\nHeaders:", headers)
print("\nFirst 10 rows (company, city, priority):")
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CO = hmap.get("company name", 2)
COL_CITY = hmap.get("city", 3)
COL_PRI = hmap.get("priority", 17)
for r in range(2, 12):
    print(f"  {ws.cell(r,COL_CO).value!r:<40} city={ws.cell(r,COL_CITY).value!r} pri={ws.cell(r,COL_PRI).value!r}")
