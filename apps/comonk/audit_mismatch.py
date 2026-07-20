"""
Audit: companies whose emails DON'T match their own company/website domain
(likely wrong emails scraped off their site - e.g. marketing agencies, foreign offices).
Compares current file vs scraper backup to see where backup is cleaner.
"""
import openpyxl, re, zipfile, time, unicodedata
def load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read "+p)
def norm(s): return re.sub(r'[^a-z0-9]','',unicodedata.normalize('NFKD',str(s).lower()))
def dom(url):
    if not url: return None
    u=re.sub(r'^https?://','',str(url).strip().lower()); u=re.sub(r'^www\.','',u)
    return u.split('/')[0].split('?')[0] if '.' in u else None

def analyze(wb, S):
    ws=wb[S]; hrow=3
    hdr=[ws.cell(hrow,c).value for c in range(1,ws.max_column+1)]
    hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
    CO=hm.get("company name",2); WEB=hm.get("website",11)
    EM=[hm.get(f"email {i}") for i in range(1,7)]; EM=[c for c in EM if c]
    total=0; has_email=0; matches=0; nomatch_all=0; nomatch_examples=[]
    for row in ws.iter_rows(min_row=hrow+1, values_only=True):
        def g(c): return row[c-1] if c and len(row)>=c else None
        co=g(CO)
        if not co or not str(co).strip(): continue
        total+=1
        ems=[str(g(c)).strip().lower() for c in EM if g(c) and "@" in str(g(c))]
        if not ems: continue
        has_email+=1
        cnk=norm(co); wdom=dom(g(WEB))
        def m(e):
            d=e.split("@")[1] if "@" in e else ""
            dn=norm(d.split(".")[0])
            # match if email domain relates to company name or website
            return (dn and (dn in cnk or cnk[:6] in dn)) or (wdom and (d==wdom or d in wdom or wdom.split(".")[0] in d))
        if any(m(e) for e in ems): matches+=1
        else:
            nomatch_all+=1
            if len(nomatch_examples)<15: nomatch_examples.append((str(co)[:26], ems[:3]))
    return total, has_email, matches, nomatch_all, nomatch_examples

cur=load("Ahmedabad_IT_AIML_FINAL_MASTER.xlsx")
t,he,mt,nm,ex=analyze(cur,"All Companies")
print("=== All Companies (CURRENT) ===")
print(f"  companies: {t} | with email: {he}")
print(f"  emails match own company/website domain: {mt} ({mt*100//he}%)")
print(f"  NONE of emails match own domain (suspect): {nm} ({nm*100//he}%)")
print("  --- examples of suspect (emails unrelated to company) ---")
for co,ems in ex: print(f"    {co:<26} {ems}")
