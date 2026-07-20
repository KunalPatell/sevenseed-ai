import openpyxl
from collections import Counter

wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
ws = wb["COMPLETE_MASTER"]

headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CAT = hmap.get("category", 4)
COL_WEB = hmap.get("website", 13)
COL_CO  = hmap.get("company name", 2)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]

cats = Counter()
for r in range(2, ws.max_row + 1):
    c = ws.cell(r, COL_CAT).value
    cats[str(c)] += 1

print("Category distribution:")
for cat, cnt in cats.most_common(30):
    print(f"  {cnt:>5}  {cat}")

print(f"\nTotal distinct categories: {len(cats)}")

# Check no-email rows: do they have website?
no_email_no_web = 0
no_email_with_web = 0
for r in range(2, ws.max_row + 1):
    existing = [ws.cell(r, c).value for c in email_cols if ws.cell(r, c).value]
    if not existing:
        web = ws.cell(r, COL_WEB).value
        if web and str(web).strip() and str(web).strip().lower() != "none":
            no_email_with_web += 1
        else:
            no_email_no_web += 1

print(f"\nRows with no email, no website: {no_email_no_web}")
print(f"Rows with no email, HAS website: {no_email_with_web}")

# Sample a few no-email-with-website rows
print("\nSample rows missing email but has website:")
count = 0
for r in range(2, ws.max_row + 1):
    existing = [ws.cell(r, c).value for c in email_cols if ws.cell(r, c).value]
    web = ws.cell(r, COL_WEB).value
    if not existing and web and str(web).strip().lower() != "none":
        print(f"  row={r} name={ws.cell(r,COL_CO).value!r} web={web!r}")
        count += 1
        if count >= 10: break
