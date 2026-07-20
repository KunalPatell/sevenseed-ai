"""
PHASE 2 (runs after the Maps run): deep contact-page scrape for MORE emails.
For every company that has a website (many newly added by the Maps run),
fetch homepage + many contact-type pages + mailto: links, and extract all
emails on the company's own domain. Authorized only (published on their site).
Adds to empty email slots. Checkpoints every 150. Keep Excel CLOSED.
"""
import openpyxl, re, time, zipfile, asyncio, httpx
from urllib.parse import urlparse, urljoin
from openpyxl.styles import Font

F="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"; LOG="scrape_log.txt"; SHEET="All Companies"
EMAIL_RE=re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
MAILTO_RE=re.compile(r'mailto:([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,})', re.I)
GENERIC={"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com","ymail.com","live.com","wixpress.com","sentry.io","example.com","email.com","domain.com","yourdomain.com"}
PATHS=["","/contact","/contact-us","/contactus","/about","/about-us","/aboutus","/team",
       "/careers","/career","/reach-us","/get-in-touch","/support","/enquiry"]
UA="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"

def log(m):
    line=f"[{time.strftime('%H:%M:%S')}] {m}"; print(line,flush=True)
    with open(LOG,"a",encoding="utf-8") as f: f.write(line+"\n")

def stable_load(p,t=60):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read (file open in Excel?)")

def cdom(url):
    try:
        h=urlparse(url if "//" in url else "http://"+url).netloc.lower().replace("www.","")
        return h if "." in h else None
    except: return None

log("=== PHASE 2: deep contact-page email scrape ===")
wb=stable_load(F); ws=wb[SHEET]
hdr=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
CO=hm.get("company name",2); WEB=hm.get("website",11); SRC=hm.get("source",16)
EM=[hm.get(f"email {i}") for i in range(1,7)]; EM=[c for c in EM if c]

targets=[]
for r in range(4, ws.max_row+1):
    nm=ws.cell(r,CO).value; web=ws.cell(r,WEB).value
    if not nm or not web: continue
    free=[c for c in EM if not ws.cell(r,c).value]
    if free and cdom(web):   # has a website + at least one empty email slot
        targets.append((r, str(web), cdom(web)))
log(f"{len(targets)} companies with website + empty email slot")

stats={"email":0,"cos":0}
def checkpoint():
    for _ in range(15):
        try: wb.save(F); return True
        except Exception: time.sleep(2)
    return False

def add(r, emails):
    free=[c for c in EM if not ws.cell(r,c).value]; cur={str(ws.cell(r,c).value or '').lower() for c in EM}; n=0
    for em in sorted(emails):
        if em.lower() not in cur and free:
            slot=free.pop(0); ws.cell(r,slot).value=em; ws.cell(r,slot).font=Font(color="1a56db",size=9); cur.add(em.lower()); n+=1
    if n:
        ws.cell(r,SRC).value="phase2: deep site scrape"; stats["email"]+=n; stats["cos"]+=1
    return n

async def main():
    limits=httpx.Limits(max_connections=40, max_keepalive_connections=20)
    sem=asyncio.Semaphore(25); done=[0]
    async with httpx.AsyncClient(limits=limits, headers={"User-Agent":UA}, follow_redirects=True) as client:
        async def work(r, website, dom):
            async with sem:
                found=set()
                base = website if website.startswith("http") else "https://"+dom
                for p in PATHS:
                    if len(found)>=6: break
                    url = urljoin(base+"/", p.lstrip("/")) if p else base
                    try:
                        resp=await client.get(url, timeout=7)
                        txt=resp.text
                        for em in MAILTO_RE.findall(txt)+EMAIL_RE.findall(txt):
                            em=em.lower().rstrip('.')
                            d=em.split("@")[1] if "@" in em else ""
                            if d and d not in GENERIC and (dom.split('.')[0] in d or d in dom):
                                found.add(em)
                    except Exception: continue
                if found: add(r, found)
                done[0]+=1
                if done[0]%150==0:
                    log(f"  {done[0]}/{len(targets)} | +{stats['email']} emails across {stats['cos']} companies")
                    checkpoint()
        for i in range(0, len(targets), 100):
            await asyncio.gather(*[work(r,w,d) for r,w,d in targets[i:i+100]], return_exceptions=True)
    checkpoint()
    log(f"PHASE 2 DONE. +{stats['email']} emails across {stats['cos']} companies")

asyncio.run(main())
