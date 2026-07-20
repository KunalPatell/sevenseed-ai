"""Full audit of COMONK_TRUE_MASTER.xlsx - check all columns for missing data."""
import openpyxl
import sys

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.max_row < 2:
        continue
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    total = ws.max_row - 1
    print(f"\n=== Sheet: '{sheet_name}' | {total} rows | {ws.max_column} cols ===")

    for c in range(1, ws.max_column + 1):
        h = headers[c-1] or f"Col{c}"
        filled = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, c).value)
        pct = round(filled/total*100) if total else 0
        missing = total - filled
        print(f"  Col{c:02d} {str(h):<28} {filled:>5}/{total} ({pct:>3}%) | missing: {missing}")
