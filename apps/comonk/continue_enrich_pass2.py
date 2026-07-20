"""
Second enrichment pass on remaining gap companies (no email, no phone):
1. Scrape the 55 that already have a website recorded
2. Domain-probe (name.com/.in/.co.in) the 408 that don't
Purely additive, never overwrites existing data.
"""
import openpyxl, re, httpx, asyncio, unicodedata
from openpyxl.styles import Font

MASTER = "COMONK_TRUE_MASTER.xlsx"
EMAIL_RE = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
PHONE_RE = re.compile(r'(?:\+91[\s\-]?|0)?[6-9]\d{9}')
GENERIC_DOMAINS = {"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com"}

wb = openpyxl.load_workbook(MASTER)
ws = wb["COMPLETE_MASTER"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CO   = hmap.get("company name", 2)
COL_WEB  = hmap.get("website", 13)
COL_PH   = hmap.get("phone", 12)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]
email_cols = [c for c in email_cols if c]

def get_domain(url):
    if not url: return None
    url = str(url).strip().lower()
    url = re.sub(r'^https?://(www\.)?', '', url).rstrip('/').split('/')[0].split('?')[0]
    if '.' in url and len(url) > 4 and url not in GENERIC_DOMAINS:
        return url
    return None

def clean_slug(name):
    s = unicodedata.normalize('NFKD', str(name).lower())
    s = re.sub(r'\(.*?\)', '', s)
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

gap_with_web = []
gap_no_web = []
for r in range(2, ws.max_row + 1):
    has_email = any(ws.cell(r, c).value for c in email_cols)
    has_ph = ws.cell(r, COL_PH).value
    if has_email or has_ph:
        continue
    name = ws.cell(r, COL_CO).value
    if not name: continue
    web = ws.cell(r, COL_WEB).value
    if get_domain(web):
        gap_with_web.append((r, str(name), str(web)))
    else:
        gap_no_web.append((r, str(name)))

print(f"Gap with website: {len(gap_with_web)} | Gap needing domain guess: {len(gap_no_web)}")

# ── Phase A: scrape companies that already have a website ────────────────────
async def scrape_existing_websites():
    added_e = 0; added_p = 0
    sem = asyncio.Semaphore(25)

    async def one(r, name, web, client):
        nonlocal added_e, added_p
        async with sem:
            domain = get_domain(web)
            urls = [web] + [f"https://{domain}/contact", f"https://{domain}/contact-us"]
            for url in urls[:2]:
                try:
                    resp = await client.get(url, timeout=7, follow_redirects=True,
                        headers={"User-Agent": "Mozilla/5.0"})
                    text = resp.text
                    ems = set()
                    for em in EMAIL_RE.findall(text):
                        em = em.lower()
                        d = em.split("@")[1] if "@" in em else ""
                        if d and d not in GENERIC_DOMAINS and domain in d:
                            ems.add(em)
                    if ems:
                        free = [c for c in email_cols if not ws.cell(r, c).value]
                        for em, slot in zip(ems, free):
                            ws.cell(r, slot).value = em
                            ws.cell(r, slot).font = Font(color="0070C0", size=9)
                            added_e += 1
                    if not ws.cell(r, COL_PH).value:
                        phones = PHONE_RE.findall(text)
                        if phones:
                            ph = re.sub(r'[\s\-]', '', phones[0])
                            if not ph.startswith("+91"): ph = "+91" + ph.lstrip("0")
                            ws.cell(r, COL_PH).value = ph
                            added_p += 1
                    if ems or ws.cell(r, COL_PH).value:
                        break
                except Exception:
                    continue

    limits = httpx.Limits(max_connections=30, max_keepalive_connections=15)
    async with httpx.AsyncClient(limits=limits) as client:
        await asyncio.gather(*[one(r, n, w, client) for r, n, w in gap_with_web],
                             return_exceptions=True)
    return added_e, added_p

print("\nPhase A: scraping companies with known websites...")
ae, ap_ = asyncio.run(scrape_existing_websites())
print(f"  +{ae} emails, +{ap_} phones")

# ── Phase B: domain-guess the rest ────────────────────────────────────────────
TLDS = [".com", ".in", ".co.in"]

async def probe_and_scrape():
    sem = asyncio.Semaphore(40)
    limits = httpx.Limits(max_connections=50, max_keepalive_connections=25)
    found = []

    async def probe(r, name, client):
        slug = clean_slug(name)
        if len(slug) < 3: return
        async with sem:
            for tld in TLDS:
                url = f"https://www.{slug}{tld}"
                try:
                    resp = await client.get(url, timeout=5, follow_redirects=True,
                        headers={"User-Agent": "Mozilla/5.0"})
                    if resp.status_code == 200 and len(resp.text) > 500:
                        tl = resp.text.lower()
                        if any(p in tl for p in ["domain is for sale", "buy this domain", "godaddy.com/domainsearch"]):
                            continue
                        found.append((r, name, url, resp.text))
                        return
                except Exception:
                    continue

    async with httpx.AsyncClient(limits=limits) as client:
        tasks = [probe(r, n, client) for r, n in gap_no_web]
        for i in range(0, len(tasks), 100):
            batch = tasks[i:i+100]
            await asyncio.gather(*batch, return_exceptions=True)
            print(f"  Probed {min(i+100, len(tasks))}/{len(tasks)} | found {len(found)} domains so far")
    return found

print("\nPhase B: domain-probing companies without a website...")
found = asyncio.run(probe_and_scrape())
print(f"  Found {len(found)} live domains")

webs_added = 0; emails_added = 0; phones_added = 0
for (r, name, url, html) in found:
    if not ws.cell(r, COL_WEB).value:
        ws.cell(r, COL_WEB).value = url
        webs_added += 1
    domain = re.sub(r'^https?://(www\.)?', '', url).split('/')[0]
    ems = set()
    for em in EMAIL_RE.findall(html):
        em = em.lower()
        d = em.split("@")[1] if "@" in em else ""
        if d and d not in GENERIC_DOMAINS and domain in d:
            ems.add(em)
    free = [c for c in email_cols if not ws.cell(r, c).value]
    for em, slot in zip(ems, free):
        ws.cell(r, slot).value = em
        ws.cell(r, slot).font = Font(color="0070C0", size=9)
        emails_added += 1
    if not ws.cell(r, COL_PH).value:
        phones = PHONE_RE.findall(html)
        if phones:
            ph = re.sub(r'[\s\-]', '', phones[0])
            if not ph.startswith("+91"): ph = "+91" + ph.lstrip("0")
            ws.cell(r, COL_PH).value = ph
            phones_added += 1

print(f"  Applied: +{webs_added} websites, +{emails_added} emails, +{phones_added} phones")

# ── Phase C: contact-page scrape for found domains still missing email ───────
async def contact_scrape():
    need = [(r, n, u) for (r, n, u, h) in found if not any(ws.cell(r, c).value for c in email_cols)]
    print(f"\nPhase C: {len(need)} companies need contact-page scrape...")
    if not need: return 0
    added = 0
    sem = asyncio.Semaphore(30)

    async def one(r, name, url, client):
        nonlocal added
        async with sem:
            domain = re.sub(r'^https?://(www\.)?', '', url).split('/')[0]
            for path in ["/contact", "/contact-us", "/contactus"]:
                try:
                    resp = await client.get(f"https://{domain}{path}", timeout=6,
                        follow_redirects=True, headers={"User-Agent": "Mozilla/5.0"})
                    ems = set()
                    for em in EMAIL_RE.findall(resp.text):
                        em = em.lower()
                        d = em.split("@")[1] if "@" in em else ""
                        if d and d not in GENERIC_DOMAINS and domain in d:
                            ems.add(em)
                    if ems:
                        free = [c for c in email_cols if not ws.cell(r, c).value]
                        for em, slot in zip(ems, free):
                            ws.cell(r, slot).value = em
                            ws.cell(r, slot).font = Font(color="0070C0", size=9)
                            added += 1
                        break
                except Exception:
                    continue

    limits = httpx.Limits(max_connections=35, max_keepalive_connections=20)
    async with httpx.AsyncClient(limits=limits) as client:
        for i in range(0, len(need), 30):
            batch = need[i:i+30]
            await asyncio.gather(*[one(r, n, u, client) for r, n, u in batch],
                                 return_exceptions=True)
    print(f"  +{added} emails from contact pages")
    return added

pc = asyncio.run(contact_scrape())

wb.save(MASTER)
print(f"\n{'='*55}")
print(f"  PASS 2 COMPLETE")
print(f"{'='*55}")
print(f"  Phase A (existing websites) : +{ae} emails, +{ap_} phones")
print(f"  Phase B (domain-probe)      : +{webs_added} websites, +{emails_added} emails, +{phones_added} phones")
print(f"  Phase C (contact scrape)    : +{pc} emails")
print(f"  Saved -> {MASTER}")
