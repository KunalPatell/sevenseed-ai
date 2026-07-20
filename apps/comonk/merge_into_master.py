"""
Merge the comprehensive MNC contacts (COMONK_MNC_CONTACTS.xlsx) INTO the main
COMONK_TRUE_MASTER.xlsx so everything is in one file.
- Replaces master's MNC_HR_Emails with the full 1231-contact version (copies values+styling)
- Rebuilds master's All_Emails to include COMPLETE_MASTER emails + all MNC contacts
- Leaves COMPLETE_MASTER / All_Phones / AI_ML_ONLY / STATS untouched
"""
import openpyxl, re, zipfile, time
from openpyxl.styles import Font, PatternFill
from copy import copy

MASTER   = "COMONK_TRUE_MASTER.xlsx"
CONTACTS = "COMONK_MNC_CONTACTS.xlsx"

def stable_load(path, tries=30):
    for _ in range(tries):
        try:
            with zipfile.ZipFile(path):
                pass
            return openpyxl.load_workbook(path)
        except Exception:
            time.sleep(2)
    raise RuntimeError(f"Could not stably read {path}")

wb  = stable_load(MASTER)
src = openpyxl.load_workbook(CONTACTS)
src_mnc = src["MNC_HR_Emails"]

# ── Replace MNC_HR_Emails in master (copy values + styles) ────────────────────
if "MNC_HR_Emails" in wb.sheetnames:
    del wb["MNC_HR_Emails"]
dst = wb.create_sheet("MNC_HR_Emails")

# copy column widths
for col, dim in src_mnc.column_dimensions.items():
    dst.column_dimensions[col].width = dim.width

# copy merged cells
for mc in list(src_mnc.merged_cells.ranges):
    dst.merge_cells(str(mc))

# copy cells with styles
for row in src_mnc.iter_rows():
    for cell in row:
        if cell.value is None and not cell.has_style:
            continue
        nc = dst.cell(row=cell.row, column=cell.column, value=cell.value)
        if cell.has_style:
            nc.font          = copy(cell.font)
            nc.fill          = copy(cell.fill)
            nc.alignment     = copy(cell.alignment)
            nc.border        = copy(cell.border)
            nc.number_format = cell.number_format

mnc_contacts = sum(1 for r in range(2, dst.max_row + 1)
                   if dst.cell(r, 5).value and "@" in str(dst.cell(r, 5).value))
print(f"MNC_HR_Emails merged into master: {dst.max_row-1} rows, {mnc_contacts} contacts")

# ── Rebuild All_Emails in master (COMPLETE_MASTER + MNC contacts) ─────────────
ws = wb["COMPLETE_MASTER"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CO = hmap.get("company name", 2); COL_CITY = hmap.get("city", 3)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]
email_cols = [c for c in email_cols if c]

DARK = PatternFill("solid", fgColor="1a1a2e")
HDRF = Font(bold=True, color="FFFFFF", size=10)

if "All_Emails" in wb.sheetnames:
    del wb["All_Emails"]
ae = wb.create_sheet("All_Emails")
for ci, h in enumerate(["Email Address", "Company", "City / Source"], 1):
    ae.cell(1, ci, h).fill = DARK; ae.cell(1, ci).font = HDRF
for ci, w in enumerate([40, 32, 20], 1):
    ae.column_dimensions[chr(64 + ci)].width = w

seen = set(); rows = []
for rr in range(2, ws.max_row + 1):
    nm = ws.cell(rr, COL_CO).value; cy = ws.cell(rr, COL_CITY).value
    for c in email_cols:
        em = ws.cell(rr, c).value
        if em and "@" in str(em) and str(em).lower() not in seen:
            seen.add(str(em).lower()); rows.append((str(em).strip(), str(nm or ""), str(cy or "")))
# add MNC contacts
for rr in range(2, dst.max_row + 1):
    em = dst.cell(rr, 5).value; co = dst.cell(rr, 1).value
    if em and "@" in str(em) and str(em).lower() not in seen:
        seen.add(str(em).lower()); rows.append((str(em).strip(), str(co or "MNC"), "MNC contact"))
for i, (em, co, cy) in enumerate(rows, 2):
    ae.cell(i, 1, em).font = Font(color="1a56db", size=9)
    ae.cell(i, 2, co).font = Font(size=9)
    ae.cell(i, 3, cy).font = Font(size=9)
print(f"All_Emails rebuilt: {len(rows)} unique emails")

# ── Save (retry if the file is momentarily locked) ────────────────────────────
for attempt in range(20):
    try:
        wb.save(MASTER)
        break
    except Exception:
        time.sleep(2)
else:
    raise RuntimeError("Could not save master (file locked)")

print(f"\nAll-in-one save complete -> {MASTER}")
print("Sheets:", wb.sheetnames)
