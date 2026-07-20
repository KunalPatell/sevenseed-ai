"""
FULL AUDIT — company names + LinkedIn profiles coverage check.
Honest report of what is in master vs what remains in source files.
"""
import openpyxl, re, csv, os, unicodedata

MASTER="COMONK_TRUE_MASTER.xlsx"
def norm(s):
    if not s: return ""
    return re.sub(r'[^a-z0-9]','',unicodedata.normalize('NFKD',str(s).lower()))

# Master company names
wb=openpyxl.load_workbook(MASTER)
ws=wb["COMPLETE_MASTER"]
master_names=set()
for r in range(2,ws.max_row+1):
    n=ws.cell(r,2).value
    if n: master_names.add(norm(n))
print(f"MASTER companies: {len(master_names)}")

# Master LinkedIn profile URLs (any /in/ personal profiles already in workbook)
master_li=set()
for sh in wb.sheetnames:
    for row in wb[sh].iter_rows(values_only=True):
        for v in row:
            if v and "linkedin.com/in/" in str(v).lower():
                master_li.add(str(v).lower().split("linkedin.com/in/")[1].strip("/ ").split("?")[0])

# Source company names (from structured Excels/CSVs)
src_names=set()
def add_name(n):
    k=norm(n)
    if k and len(k)>=3: src_names.add(k)

# CSVs with Company column
for fn,col in [("AIML_Companies_Apply_List.csv","Company"),
               ("Gandhinagar_GIFTCity_Companies.csv","Company"),
               ("NEW_COMPANIES_DISCOVERED.csv","Company")]:
    if os.path.exists(fn):
        for row in csv.DictReader(open(fn,encoding="utf-8-sig",errors="ignore")):
            for key in row:
                if key and key.strip().lower()=="company": add_name(row[key])

# Excels — first text column heuristic
for fn in ["Ahmedabad_AIML_Companies_HR_Contacts.xlsx","Ahmedabad_IT_AIML_Companies_Master.xlsx",
           "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx","AI_Engineer_Job_Targets.xlsx","HR Mail List.xlsx"]:
    if os.path.exists(fn):
        w=openpyxl.load_workbook(fn,read_only=True)
        for s in w.sheetnames:
            for row in w[s].iter_rows(values_only=True):
                cells=[c for c in row if c]
                if len(cells)>=2 and isinstance(row[1] if len(row)>1 else None,str):
                    cand=row[1]
                    if cand and "@" not in str(cand) and not str(cand).startswith("http") and len(str(cand))>3:
                        add_name(cand)
        w.close()

# Company names from PDFs (DataScience list)
if os.path.exists("_extract_datascience_compeny_list_pdf.txt"):
    for line in open("_extract_datascience_compeny_list_pdf.txt",encoding="utf-8"):
        m=re.match(r'^\d+\.\s+(.+)',line.strip())
        if m: add_name(m.group(1))

missing_names=src_names - master_names
print(f"SOURCE company names: {len(src_names)}")
print(f"  Missing from master: {len(missing_names)}")

# LinkedIn HR profiles from sources
src_li=set()
for fn in ["clean_hr_linkedin_list.csv","linkedin_profiles.csv"]:
    if os.path.exists(fn):
        for line in open(fn,encoding="utf-8",errors="ignore"):
            for m in re.findall(r'linkedin\.com/in/([A-Za-z0-9\-_%]+)', line.lower()):
                src_li.add(m.strip("/ "))
if os.path.exists("_extract_hr_mail_list_pdf.txt"):
    for m in re.findall(r'linkedin\.com/in/([A-Za-z0-9\-_%]+)', open("_extract_hr_mail_list_pdf.txt",encoding="utf-8").read().lower()):
        src_li.add(m.strip("/ "))
if os.path.exists("1500+_hr_list.xlsx"):
    w=openpyxl.load_workbook("1500+_hr_list.xlsx",read_only=True)
    for s in w.sheetnames:
        for row in w[s].iter_rows(values_only=True):
            for v in row:
                if v:
                    for m in re.findall(r'linkedin\.com/in/([A-Za-z0-9\-_%]+)', str(v).lower()):
                        src_li.add(m.strip("/ "))
    w.close()

missing_li = src_li - master_li
print(f"\nLinkedIn HR profiles in sources: {len(src_li)}")
print(f"  Already in workbook: {len(src_li & master_li)}")
print(f"  MISSING from workbook: {len(missing_li)}")

print(f"\n=== SAMPLE missing company names ===")
for n in list(missing_names)[:25]: print(f"  {n}")
