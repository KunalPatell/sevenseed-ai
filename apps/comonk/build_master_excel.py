import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from collections import defaultdict

# ═══════════════════════════════════════════════════════════════════════════════
# 1. RESEARCHED COMPANY DETAILS  (address, phone, LinkedIn, website, roles)
# ═══════════════════════════════════════════════════════════════════════════════
COMPANY_DETAILS = {
    # ── IT / AI companies (Ahmedabad) ──────────────────────────────────────
    "azilen":            {"address":"12th & 13th Floor, B Square-1, Bopal-Ambli Rd, Ahmedabad 380054","phone":"+91 79400 93121","linkedin":"linkedin.com/company/azilentechnologies","website":"https://www.azilen.com","roles":"AI Engineer, ML Engineer"},
    "azilentechnologies": {"address":"12th & 13th Floor, B Square-1, Bopal-Ambli Rd, Ahmedabad 380054","phone":"+91 79400 93121","linkedin":"linkedin.com/company/azilentechnologies","website":"https://www.azilen.com","roles":"AI Engineer, ML Engineer"},
    "infocusp":          {"address":"301-310, 4th Floor, Gala Hub, Gala Gymkhana Rd, Bopal, Ahmedabad 380058","phone":"","linkedin":"linkedin.com/company/infocusp","website":"https://www.infocusp.com","roles":"AI/ML Engineer, Data Scientist"},
    "capermint":         {"address":"304-305, Shivalik Shilp-1, Iscon Cross Rd, SG Highway, Ahmedabad 380015","phone":"+91 81405 47782","linkedin":"linkedin.com/company/caperminttech","website":"https://www.capermint.com","roles":"AI App Developer, Game Developer"},
    "caperminttechnologies": {"address":"304-305, Shivalik Shilp-1, Iscon Cross Rd, SG Highway, Ahmedabad 380015","phone":"+91 81405 47782","linkedin":"linkedin.com/company/caperminttech","website":"https://www.capermint.com","roles":"AI App Developer, Game Developer"},
    "vrinsoft":          {"address":"707, Elite Business Park, Shapath Hexa, SG Highway, Ahmedabad 380060","phone":"+91 7227906117","linkedin":"linkedin.com/company/vrinsoft-technologies-pvt-ltd","website":"https://www.vrinsofts.com","roles":"Software Developer, Mobile App Developer"},
    "vrinsofts":         {"address":"707, Elite Business Park, Shapath Hexa, SG Highway, Ahmedabad 380060","phone":"+91 7227906117","linkedin":"linkedin.com/company/vrinsoft-technologies-pvt-ltd","website":"https://www.vrinsofts.com","roles":"Software Developer, Mobile App Developer"},
    "marutitech":        {"address":"10th Floor, The Ridge, Opp. Novotel, Iscon Char Rasta, Ahmedabad 380060","phone":"+91 9879101067","linkedin":"linkedin.com/company/maruti-techlabs-pvt-ltd","website":"https://www.marutitech.com","roles":"Deep Learning Engineer, AI Researcher"},
    "openxcell":         {"address":"202-203, Baleshwar Avenue, Opp. Rajpath Club, SG Highway, Ahmedabad 380054","phone":"+91 9998222929","linkedin":"linkedin.com/company/openxcell","website":"https://www.openxcell.com","roles":"AI Developer, ML Engineer"},
    "trootech":          {"address":"605, 6th Floor, B Square, Iscon-Ambali BRTS Rd, Ahmedabad 380058","phone":"+91 2717454342","linkedin":"linkedin.com/company/trootech","website":"https://www.trootech.com","roles":"Software Developer, AI Solutions"},
    "fxis":              {"address":"4th Floor, Sanskrut Building, Navrangpura, Ahmedabad 380009","phone":"+91 9727999595","linkedin":"linkedin.com/company/fxisai","website":"https://www.fxis.ai","roles":"AI Solutions Architect, AI Engineer"},
    "fxisai":            {"address":"4th Floor, Sanskrut Building, Navrangpura, Ahmedabad 380009","phone":"+91 9727999595","linkedin":"linkedin.com/company/fxisai","website":"https://www.fxis.ai","roles":"AI Solutions Architect, AI Engineer"},
    "crestdata":         {"address":"CDS House, Nr. Sarkhej-Sanand Circle, SG Rd, Makarba, Ahmedabad 382210","phone":"","linkedin":"linkedin.com/company/crest-data-systems","website":"https://www.crestdata.ai","roles":"Data Scientist, ML Engineer"},
    "holbox":            {"address":"Ahmedabad, Gujarat, India","phone":"","linkedin":"linkedin.com/company/holboxai","website":"https://www.holbox.ai","roles":"AI/ML Engineer, Backend Developer"},
    "holboxai":          {"address":"Ahmedabad, Gujarat, India","phone":"","linkedin":"linkedin.com/company/holboxai","website":"https://www.holbox.ai","roles":"AI/ML Engineer, Backend Developer"},
    "einfochips":        {"address":"11/A-B, Chandra Colony, Off C.G. Road, Ellisbridge, Ahmedabad 380006","phone":"+91 79 26563705","linkedin":"linkedin.com/company/einfochips","website":"https://www.einfochips.com","roles":"AI/ML Engineer (Embedded Systems), Data Engineer"},
    "bacancy":           {"address":"1207-1208, Times Square Arcade, Thaltej-Shilaj Rd, Ahmedabad 380059","phone":"+91 7940037674","linkedin":"linkedin.com/company/bacancy-technology","website":"https://www.bacancytechnology.com","roles":"AI/ML Developer, Data Analyst"},
    "bacancytechnology": {"address":"1207-1208, Times Square Arcade, Thaltej-Shilaj Rd, Ahmedabad 380059","phone":"+91 7940037674","linkedin":"linkedin.com/company/bacancy-technology","website":"https://www.bacancytechnology.com","roles":"AI/ML Developer, Data Analyst"},
    "hyperlinkinfosystem":{"address":"Block C, 106/B Ganesh Meridian, Near Sola Bridge, SG Highway, Ahmedabad 380061","phone":"+91 80001 61161","linkedin":"linkedin.com/company/hyperlinkinfosystem","website":"https://www.hyperlinkinfosystem.com","roles":"Data Analyst, App Developer (ML)"},
    "spaceogroup":       {"address":"1005, 10th Floor, Abhishree Adroit, Mansi Circle, Ahmedabad 380015","phone":"+91 9316757277","linkedin":"linkedin.com/company/space-o-technologies","website":"https://www.spaceotechnologies.com","roles":"Mobile App Developer, AI Integration"},
    "spaceotechnologies": {"address":"1005, 10th Floor, Abhishree Adroit, Mansi Circle, Ahmedabad 380015","phone":"+91 9316757277","linkedin":"linkedin.com/company/space-o-technologies","website":"https://www.spaceotechnologies.com","roles":"Mobile App Developer, AI Integration"},
    "bytestechnolab":    {"address":"THE FIRST, B-805, Near Mansi Cross Road, Vastrapur, Ahmedabad 380015","phone":"+91 9157320903","linkedin":"linkedin.com/company/bytes-technolab","website":"https://www.bytestechnolab.com","roles":"AI/ML Engineer, Data Scientist"},
    "tatvasoft":         {"address":"TatvaSoft House, Rajpath Club Rd, Ahmedabad 380054","phone":"+91 9601421472","linkedin":"linkedin.com/company/tatvasoft","website":"https://www.tatvasoft.com","roles":"Software Developer (Data Science), BI Developer"},
    "agileinfoways":     {"address":"303/304, Abhishree Avenue, Near Nehru Nagar Cross Rd, Ambawadi, Ahmedabad 380015","phone":"+91 7622081234","linkedin":"linkedin.com/company/agile-infoways-pvt-ltd-","website":"https://www.agileinfoways.com","roles":"Data Analyst, AI/ML Developer"},
    "radixweb":          {"address":"401, Anandmangal-II, B/H Omkar House, C.G. Road, Navrangpura, Ahmedabad 380009","phone":"+91 9909003119","linkedin":"linkedin.com/company/radixweb","website":"https://www.radixweb.com","roles":"Data Analyst, BI Specialist"},
    "softwebsolutions":  {"address":"Block 5 & 6, Garden View, Corporate House, Bodakdev, Ahmedabad 380054","phone":"+1 866-345-7638","linkedin":"linkedin.com/company/softwebsolutionsinc","website":"https://www.softwebsolutions.com","roles":"AI for IoT, Data Services, Digital Transformation"},
    "atliq":             {"address":"Ahmedabad / Vadodara, Gujarat, India","phone":"+91 9979738578","linkedin":"linkedin.com/company/atliqtech","website":"https://www.atliq.com","roles":"Data Analyst, AI Consultant"},
    "elightwalk":        {"address":"611, Shivalik Square, Near Adani CNG Pump, Old Wadaj, Ahmedabad 380013","phone":"+91 7600897405","linkedin":"linkedin.com/company/elightwalk-technology-pvt-ltd","website":"https://www.elightwalk.com","roles":"E-commerce Developer, SEO"},
    "tntra":             {"address":"B/H Rajpath Club, Nr. AUDA Garden, Off SG Road, Ahmedabad","phone":"","linkedin":"linkedin.com/company/tntra","website":"https://www.tntra.io","roles":"Software Developer, Tech Entrepreneur"},
    "specindia":         {"address":"SPEC House, Parth Complex, Near Swastik Cross Roads, Navrangpura, Ahmedabad 380009","phone":"+91 79 2640 4031","linkedin":"linkedin.com/company/spec-india","website":"https://www.spec-india.com","roles":"BI Analyst, Data Analyst"},
    "inexture":          {"address":"Office No. 405, One World West, Near Ambli T-Junction, SP Ring Rd, Bopal, Ahmedabad 380058","phone":"+91 6353697824","linkedin":"linkedin.com/company/inexture","website":"https://www.inexture.com","roles":"AI Developer, Python Developer"},
    "apexon":            {"address":"Shreem Building, Off SG Highway, Near Iscon Crossroad, Jodhpur, Ahmedabad 380015","phone":"","linkedin":"linkedin.com/company/apexon","website":"https://www.apexon.com","roles":"AI/ML Consultant, Data Analyst"},
    "rishabhsoft":       {"address":"Devx, 4th Floor, Binori B Square3, Sindhu Bhavan Rd, Ahmedabad 380054","phone":"+91 8511122697","linkedin":"linkedin.com/company/rishabh-software","website":"https://www.rishabhsoft.com","roles":"Software Engineer (BI/Analytics)"},
    "drivebuddyai":      {"address":"A-1111, World Trade Tower, Off SG Highway, Makarba, Ahmedabad 380051","phone":"+91 11 4117 0779","linkedin":"linkedin.com/company/drivebuddyai","website":"https://www.drivebuddyai.co","roles":"AI Fleet Safety Engineer, Data Analyst"},
    "drivebuddyaico":    {"address":"A-1111, World Trade Tower, Off SG Highway, Makarba, Ahmedabad 380051","phone":"+91 11 4117 0779","linkedin":"linkedin.com/company/drivebuddyai","website":"https://www.drivebuddyai.co","roles":"AI Fleet Safety Engineer, Data Analyst"},
    # ── DataScience PDF companies ───────────────────────────────────────────
    "infosys":           {"address":"Ahmedabad / Gandhinagar, Gujarat","phone":"","linkedin":"linkedin.com/company/infosys","website":"https://www.infosys.com","roles":"AI/ML Engineer, Data Scientist, Data Engineer, Big Data Architect"},
    "wipro":             {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/wipro","website":"https://www.wipro.com","roles":"Data Scientist, Machine Learning Engineer, AI Solution Architect"},
    "accenture":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/accenture","website":"https://www.accenture.com","roles":"AI Consultant, Data Science Analyst, ML Engineer"},
    "ibm":               {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/ibm","website":"https://www.ibm.com","roles":"Data Scientist, AI Application Developer, ML Engineer, AI Consultant"},
    "capgemini":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/capgemini","website":"https://www.capgemini.com","roles":"AI/ML Consultant, Data Analyst, Data Engineer"},
    "cygnetinfotech":    {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/cygnet-infotech","website":"https://www.cygnetinfotech.com","roles":"Software Engineer (AI/ML), Data Analyst"},
    "simform":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/simform","website":"https://www.simform.com","roles":"Data Engineer, ML Operations Engineer"},
    "indianic":          {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/indianic-infotech-limited","website":"https://www.indianic.com","roles":"Data Analyst"},
    "yudiz":             {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/yudiz-solutions","website":"https://www.yudiz.com","roles":"Software Developer (AI/ML)"},
    "alliancetek":       {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/alliancetek","website":"https://www.alliancetek.com","roles":"Data Analyst"},
    "netclues":          {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/netclues-india","website":"https://www.netclues.com","roles":"Digital Analyst"},
    "mindinventory":     {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/mindinventory","website":"https://www.mindinventory.com","roles":"AI/ML Developer"},
    "bigscal":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/bigscal-technologies-pvt-ltd","website":"https://www.bigscal.com","roles":"Software Engineer (Data focus)"},
    "patoliyainfotech":  {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/patoliya-infotech","website":"https://www.patoliyainfotech.com","roles":"App Developer (ML features)"},
    "quixom":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/quixom-technology","website":"https://www.quixom.com","roles":"Python Developer (ML/Data Science)"},
    "iqlance":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/iqlance-solutions-pvt-ltd","website":"https://www.iqlance.com","roles":"App Developer (AI integration)"},
    "solulab":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/solulab","website":"https://www.solulab.com","roles":"AI Engineer, ML Developer"},
    "algobrain":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/algobrain-ai","website":"https://www.algobrain.ai","roles":"AI/ML Specialist"},
    "avidclantech":      {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/avidclan-technologies","website":"https://www.avidclantech.com","roles":"Software Developer (Data focus)"},
    "concettolabs":      {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/concetto-labs","website":"https://www.concettolabs.com","roles":"AI Consultant, Data Analyst"},
    "compubrain":        {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/compubrain-private-limited","website":"https://www.compubrain.com","roles":"Technology Consultant (Analytics/AI)"},
    "solutionanalysts":  {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/solution-analysts-pvt-ltd","website":"https://www.solutionanalysts.com","roles":"Business Intelligence Analyst, Data Analyst"},
    "magnetoitsolutions": {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/magneto-it-solutions","website":"https://www.magnetoitsolutions.com","roles":"E-commerce Analyst (Data)"},
    "growexx":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/growexx","website":"https://www.growexx.com","roles":"AI Consultant, Data Consultant"},
    "brainvire":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/brainvire-infotech","website":"https://www.brainvire.com","roles":"Software Engineer (AI/ML)"},
    "tcs":               {"address":"Gandhinagar / Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/tata-consultancy-services","website":"https://www.tcs.com","roles":"AI Consultant, ML Specialist, Data Scientist, Data Engineer"},
    "adani":             {"address":"Ahmedabad / Gujarat (Multiple Locations)","phone":"","linkedin":"linkedin.com/company/adani-group","website":"https://www.adani.com","roles":"Data Scientist, AI/ML Engineer"},
    "reliance":          {"address":"Gujarat / Pan India","phone":"","linkedin":"linkedin.com/company/reliance-industries","website":"https://www.ril.com","roles":"Data Scientist, ML Engineer"},
    "techmahindra":      {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/tech-mahindra","website":"https://www.techmahindra.com","roles":"Data Scientist, ML Engineer, AI Architect"},
    "moontechnolabs":    {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/moon-technolabs","website":"https://www.moontechnolabs.com","roles":"Mobile App Developer, AI Integration"},
    "aglowiditsolutions": {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/aglowid-it-solutions","website":"https://www.aglowiditsolutions.com","roles":"Software Developer, Web Developer"},
    "softvan":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/softvan-infotech","website":"https://www.softvan.com","roles":"Software Developer"},
    "stellarmind":       {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/stellarmind-ai","website":"https://www.stellarmind.ai","roles":"AI Engineer, ML Researcher"},
    "appitsimple":       {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/appitsimple-infotek","website":"https://www.appitsimple.com","roles":"Software Developer, Product Manager"},
    "kodytechnolab":     {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/kody-technolab","website":"https://www.kodytechnolab.com","roles":"Mobile App Developer, AI Developer"},
    "ethicsinfotech":    {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/ethics-infotech","website":"https://www.ethicsinfotech.in","roles":"Software Developer, Web Developer"},
    "teksun":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/teksun-microsys","website":"https://www.teksun.com","roles":"IoT Engineer, Embedded Systems"},
    "webcluesinfotech":  {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/webclues-infotech","website":"https://www.webcluesinfotech.com","roles":"Web Developer, Mobile App Developer"},
    "codezeros":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/codezeros","website":"https://www.codezeros.com","roles":"Blockchain Developer, AI Developer"},
    "nividous":          {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/nividous","website":"https://www.nividous.com","roles":"RPA Developer, AI Automation Engineer"},
    "jeavio":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/jeavio","website":"https://www.jeavio.com","roles":"Software Engineer, Data Engineer"},
    "tatvic":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/tatvic","website":"https://www.tatvic.com","roles":"Digital Analytics Consultant, Data Analyst"},
    "scaledge":          {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/scaledge","website":"https://www.scaledge.io","roles":"Software Developer, ML Engineer"},
    "scalege":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/scaledge","website":"https://www.scaledge.io","roles":"Software Developer, ML Engineer"},
    "aionpixel":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/aionpixel","website":"https://www.aionpixel.com","roles":"AI Developer, ML Engineer"},
    "aipxperts":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/aipxperts","website":"https://www.aipxperts.com","roles":"AI/ML Engineer, Data Scientist"},
    "drcsystems":        {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/drc-systems","website":"https://www.drcsystems.com","roles":"Software Developer, ERP Consultant"},
    "amnex":             {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/amnex","website":"https://www.amnex.com","roles":"Telecom Software Developer, Data Analyst"},
    "pirimidtech":       {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/pirimidtech","website":"https://www.pirimidtech.com","roles":"Software Developer, SaaS Engineer"},
    "valensdatalabs":    {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/valens-data-labs","website":"https://www.valensdatalabs.com","roles":"Data Scientist, ML Engineer"},
    "dxfactor":          {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/dxfactor","website":"https://www.dxfactor.com","roles":"AI Analyst, Data Engineer"},
    "rysun":             {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/rysun","website":"https://www.rysun.com","roles":"Software Developer, QA Engineer"},
    "anblicks":          {"address":"Ahmedabad / Hyderabad","phone":"","linkedin":"linkedin.com/company/anblicks","website":"https://www.anblicks.com","roles":"Data Analyst, Cloud Data Engineer"},
    "xceltec":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/xceltec","website":"https://www.xceltec.com","roles":"Mobile App Developer, Software Developer"},
    "biztechcs":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/biztechcs","website":"https://www.biztechcs.com","roles":"Software Developer, Web Developer"},
    "prydan":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/prydan","website":"https://www.prydan.com","roles":"Software Developer, QA Engineer"},
    "hiteshi":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/hiteshi-technologies","website":"https://www.hiteshi.com","roles":"Software Developer, Mobile App Developer"},
    "technource":        {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/technource","website":"https://www.technource.com","roles":"Software Developer, Mobile App Developer"},
    "petpooja":          {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/petpooja","website":"https://www.petpooja.com","roles":"Software Developer, Product Manager"},
    "tatvasoft":         {"address":"TatvaSoft House, Rajpath Club Rd, Ahmedabad 380054","phone":"+91 9601421472","linkedin":"linkedin.com/company/tatvasoft","website":"https://www.tatvasoft.com","roles":"Software Developer (Data Science), BI Developer"},
    "alakmalak":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/alakmalak-technologies","website":"https://www.alakmalak.com","roles":"Web Developer, Mobile App Developer"},
    "manektech":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/manektech","website":"https://www.manektech.com","roles":"Software Developer, Web Developer"},
    "krishaweb":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/krishaweb","website":"https://www.krishaweb.com","roles":"Web Developer, PHP Developer"},
    "ninjatechnolabs":   {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/ninja-technolabs","website":"https://www.ninjatechnolabs.com","roles":"Mobile App Developer, Web Developer"},
    "artoonsolutions":   {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/artoon-solutions","website":"https://www.artoonsolutions.com","roles":"Mobile App Developer, Game Developer"},
    "logicrays":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/logic-rays-technologies","website":"https://www.logicrays.com","roles":"Software Developer, Web Developer"},
    "cmarix":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/cmarix-technolab","website":"https://www.cmarix.com","roles":"Mobile App Developer, Web Developer"},
    "mobmaxime":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/mobmaxime","website":"https://www.mobmaxime.com","roles":"Mobile App Developer"},
    "citrusbug":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/citrusbug-technolabs","website":"https://www.citrusbug.co","roles":"Software Developer, Web Developer"},
    "squadtechnologies":  {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/squad-technologies","website":"https://www.squadtechnologies.com","roles":"Software Developer"},
    "intuz":             {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/intuz","website":"https://www.intuz.info","roles":"Mobile App Developer, IoT Developer"},
    "appvintech":        {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/appvin-technologies","website":"https://www.appvintech.com","roles":"Mobile App Developer"},
    "sharkstriker":      {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/sharkstriker","website":"https://www.sharkstriker.com","roles":"Cybersecurity Engineer"},
    "mastek":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/mastek","website":"https://www.mastek.com","roles":"Software Developer, Data Engineer"},
    "neuramonks":        {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/neuramonks","website":"https://www.neuramonks.com","roles":"AI/ML Developer, Data Scientist"},
    "moweb":             {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/moweb-technologies","website":"https://www.moweb.com","roles":"Mobile App Developer"},
    "conceptinfoway":    {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/concept-infoway","website":"https://www.conceptinfoway.net","roles":"Software Developer, Web Developer"},
    "kloudrac":          {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/kloudrac","website":"https://www.kloudrac.com","roles":"Cloud Engineer, DevOps"},
    "silvertouch":       {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/silvertouch-technologies","website":"https://www.silvertouch.com","roles":"Software Developer, ERP Consultant"},
    "esparkinfo":        {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/esparkbiz","website":"https://www.esparkinfo.com","roles":"Web Developer, Mobile App Developer"},
    "uplers":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/uplers","website":"https://www.uplers.in","roles":"Digital Marketer, Web Developer"},
    "tririd":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/tririd","website":"https://www.tririd.com","roles":"Software Developer"},
    "webplanex":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/webplanex","website":"https://www.webplanex.com","roles":"Web Developer"},
    "qmetry":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/qmetry","website":"https://www.qmetry.com","roles":"QA Engineer, Test Automation Engineer"},
    "aividtechvision":   {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/aividtechvision","website":"https://www.aividtechvision.com","roles":"Computer Vision Engineer, AI Developer"},
    "bonrix":            {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/bonrix-software-systems","website":"https://www.bonrix.net","roles":"Software Developer, IoT Developer"},
    "bureau":            {"address":"Ahmedabad / Bengaluru","phone":"","linkedin":"linkedin.com/company/bureau-id","website":"https://www.bureau.id","roles":"AI Risk Engineer, Data Scientist"},
    "atqor":             {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/atqor","website":"https://www.atqor.com","roles":"Software Developer, Web Developer"},
    "saras-3d":          {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/saras-3d","website":"https://www.saras-3d.com","roles":"3D Engineer, CAD Designer"},
    "auberginesolutions": {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/aubergine-solutions","website":"https://www.auberginesolutions.com","roles":"UX Designer, Product Developer"},
    "oblivious":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/oblivious-ai","website":"https://www.oblivious.ai","roles":"Privacy AI Engineer, Data Scientist"},
    "hhaexchange":       {"address":"Ahmedabad, Gujarat (India Office)","phone":"","linkedin":"linkedin.com/company/hhaexchange","website":"https://www.hhaexchange.com","roles":"Software Developer, Data Analyst"},
    "trellance":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/trellance","website":"https://www.trellance.com","roles":"Data Analyst, BI Developer"},
    "flodataanalytics":  {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/flo-data-analytics","website":"https://www.flodataanalytics.com","roles":"Data Analyst, Business Analyst"},
    "chillitray":        {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/chillitray","website":"https://www.chillitray.com","roles":"Software Developer"},
    "fitterfly":         {"address":"Ahmedabad / Mumbai","phone":"","linkedin":"linkedin.com/company/fitterfly","website":"https://www.fitterfly.com","roles":"Data Scientist, Health Tech Developer"},
    "cleardu":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/cleardu","website":"https://www.cleardu.com","roles":"EdTech Developer, Data Analyst"},
    "intellyticssolutions": {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/intellytics-solutions","website":"https://www.intellyticssolutions.com","roles":"Data Analyst, BI Developer"},
    "thruways":          {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/thruways","website":"https://www.thruways.co","roles":"Software Developer"},
    "deepcoder":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/deepcoder","website":"https://www.deepcoder.io","roles":"AI Developer, Software Engineer"},
    "stalwartitsolution": {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/stalwart-it-solution","website":"https://www.stalwartitsolution.com","roles":"Software Developer"},
    "zipaworld":         {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/zipaworld","website":"https://www.zipaworld.com","roles":"Logistics Tech Developer"},
    "technotery":        {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/technotery","website":"https://www.technotery.com","roles":"Software Developer"},
    "qrioustech":        {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/qrious-tech","website":"https://www.qrioustech.in","roles":"Software Developer"},
    "techkul":           {"address":"Ahmedabad, Gujarat","phone":"","linkedin":"linkedin.com/company/techkul","website":"https://www.techkul.com","roles":"E-commerce Developer, Magento Expert"},
}

def get_domain(email):
    try:
        return email.split('@')[1].lower().strip()
    except:
        return ''

def domain_to_key(domain):
    return domain.split('.')[0].lower()

def lookup_details(company_name, domain):
    cn = company_name.lower().replace(' ', '').replace('.', '').replace('-', '')
    dn = domain_to_key(domain)
    for key in [dn, cn]:
        if key in COMPANY_DETAILS:
            return COMPANY_DETAILS[key]
    for key in COMPANY_DETAILS:
        if key in cn or key in dn or cn in key or dn in key:
            return COMPANY_DETAILS[key]
    return {}

# ═══════════════════════════════════════════════════════════════════════════════
# 2. LOAD EMAIL DATA
# ═══════════════════════════════════════════════════════════════════════════════
hr_df = pd.read_excel('HR Mail List.xlsx')

bcc_emails = []
with open('all_bcc_emails.txt', 'r', encoding='utf-8') as f:
    for line in f:
        email = line.strip().rstrip(',')
        if '@' in email:
            bcc_emails.append(email.lower())

existing_domains = set()
for _, row in hr_df.iterrows():
    for col in ['Company Mail','HR Mail','HR Mail2','HR Mail3','HR Mail4','HR Mail5','HR Mail6']:
        val = str(row.get(col,'') or '')
        if '@' in val:
            existing_domains.add(get_domain(val.lower()))

bcc_by_domain = defaultdict(list)
skip = {'gmail.com','yahoo.com','outlook.com','hotmail.com','pw.live','gmail','yahoo'}
for email in bcc_emails:
    d = get_domain(email)
    if d and not any(s in d for s in ['gmail','yahoo','outlook','hotmail']):
        bcc_by_domain[d].append(email)

extra_companies = []
for domain, emails in bcc_by_domain.items():
    if domain not in existing_domains:
        name = domain.split('.')[0].replace('-',' ').replace('_',' ').title()
        extra_companies.append({'domain': domain, 'emails': emails, 'name': name})

# ═══════════════════════════════════════════════════════════════════════════════
# 3. CATEGORIZE
# ═══════════════════════════════════════════════════════════════════════════════
ai_keywords = ['ai','ml','data','analyt','neural','brain','cognitiv','vision','nlp',
               'bot','deep','machine','learn','smart','predict','genai','llm','robotic',
               'automot','intelligence','science','mining']
aiml_domains = {'fxis','crestdata','holbox','drivebuddyai','aividtechvision','actualisation',
                'aistatics','stellarmind','infocusp','oblivious','datasafeguard','superr',
                'tailinc','datavruti','ddatalabs','bzanalytics','analyticsvidhya','comprinno',
                'humantic','makunaiglobal','panscience','prama','psytech','hunar','voidspace',
                'neuramonks','aipxperts','biobrain','loomylabs','galaxy','toeho','alzieme',
                'gridanalyticsindia','crata','aietheriq','botlabdynamics','sentientgeeks',
                'mandelbulbtech','valensdatalabs','flodataanalytics','dxfactor','intellytics',
                'tatvic','anblicks','hexadatasolutions','rheincs','savvydatacloud','algobrain',
                'growexx','solulab','mindinventory','simform','concettolabs','instinctools'}

def categorize(company_name, domain):
    c = company_name.lower()
    d = domain.lower()
    key = domain_to_key(d)
    if key in aiml_domains:
        return 'AI / ML'
    for kw in ai_keywords:
        if kw in c or kw in d:
            return 'AI / ML'
    if any(x in d for x in ['tech','soft','info','sys','code','dev','digit','web','app','sol',
                              'erp','corp','lab','invent','logic','cloud','byte','ware']):
        return 'IT Services'
    return 'IT / Technology'

# ═══════════════════════════════════════════════════════════════════════════════
# 4. BUILD ROWS
# ═══════════════════════════════════════════════════════════════════════════════
rows = []

for _, row in hr_df.iterrows():
    company = str(row.get('Company Name','') or '').strip()
    if not company or company == 'nan':
        continue
    company_mail = str(row.get('Company Mail','') or '').strip()
    if company_mail == 'nan': company_mail = ''
    hr_mails = []
    for col in ['HR Mail','HR Mail2','HR Mail3','HR Mail4','HR Mail5','HR Mail6']:
        v = str(row.get(col,'') or '').strip()
        if v and v != 'nan' and '@' in v:
            hr_mails.append(v)
    domain = get_domain(company_mail.lower()) if company_mail else ''
    det = lookup_details(company, domain)
    website  = det.get('website') or (('https://www.'+domain) if domain else '')
    linkedin = det.get('linkedin','')
    phone    = det.get('phone','')
    address  = det.get('address','')
    roles    = det.get('roles','')
    category = categorize(company, domain)

    rows.append({
        'Company Name':  company,
        'Category':      category,
        'Roles / Skills': roles,
        'Email 1':       company_mail,
        'Email 2':       hr_mails[0] if len(hr_mails)>0 else '',
        'Email 3':       hr_mails[1] if len(hr_mails)>1 else '',
        'Email 4':       hr_mails[2] if len(hr_mails)>2 else '',
        'Email 5':       hr_mails[3] if len(hr_mails)>3 else '',
        'Phone':         phone,
        'Website':       website,
        'LinkedIn URL':  linkedin,
        'Address':       address,
    })

for ec in extra_companies:
    emails = ec['emails']
    domain = ec['domain']
    det = lookup_details(ec['name'], domain)
    website  = det.get('website') or 'https://www.'+domain
    linkedin = det.get('linkedin','')
    phone    = det.get('phone','')
    address  = det.get('address','')
    roles    = det.get('roles','')
    category = categorize(ec['name'], domain)
    rows.append({
        'Company Name':  ec['name'],
        'Category':      category,
        'Roles / Skills': roles,
        'Email 1':       emails[0] if len(emails)>0 else '',
        'Email 2':       emails[1] if len(emails)>1 else '',
        'Email 3':       emails[2] if len(emails)>2 else '',
        'Email 4':       emails[3] if len(emails)>3 else '',
        'Email 5':       emails[4] if len(emails)>4 else '',
        'Phone':         phone,
        'Website':       website,
        'LinkedIn URL':  linkedin,
        'Address':       address,
    })

seen = set()
final_rows = []
for r in rows:
    key = r['Company Name'].lower().strip()
    if key not in seen:
        seen.add(key)
        final_rows.append(r)
final_rows.sort(key=lambda x: x['Company Name'].lower())
print(f"Total companies: {len(final_rows)}")

# ═══════════════════════════════════════════════════════════════════════════════
# 5. EXCEL STYLE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")
SUBHDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
AI_FILL      = PatternFill("solid", fgColor="E8F5E9")   # light green
ROW_FILLS    = [PatternFill("solid", fgColor="DDEEFF"),
                PatternFill("solid", fgColor="FFFFFF")]
LINK_FONT    = Font(color="0563C1", underline="single", size=10)
BOLD_FONT    = Font(bold=True, size=10)
NORM_FONT    = Font(size=10)
bs = Side(style='thin', color='B0C4DE')
THIN_BORDER  = Border(left=bs, right=bs, top=bs, bottom=bs)

COLUMNS = [
    ('No',            5),
    ('Company Name',  28),
    ('Category',      14),
    ('Roles / Skills',30),
    ('Email 1',       36),
    ('Email 2',       34),
    ('Email 3',       34),
    ('Email 4',       34),
    ('Email 5',       34),
    ('Phone',         18),
    ('Website',       28),
    ('LinkedIn URL',  35),
    ('Address',       45),
]

def write_sheet(ws, data_rows, title=None):
    """Write a fully-styled table to a worksheet."""
    start_row = 1
    if title:
        ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
        tc = ws.cell(row=1, column=1, value=title)
        tc.fill      = HEADER_FILL
        tc.font      = Font(bold=True, color="FFFFFF", size=13)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34
        start_row = 2
        # blank spacer
        ws.row_dimensions[2].height = 6
        start_row = 3

    # Column widths
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Header row
    hdr_row = start_row
    ws.row_dimensions[hdr_row].height = 30
    for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=hdr_row, column=col_idx, value=col_name)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN_BORDER

    # Data rows
    data_start = hdr_row + 1
    for i, r in enumerate(data_rows, start=1):
        row_n = data_start + i - 1
        is_ai = r['Category'] == 'AI / ML'
        fill  = AI_FILL if is_ai else ROW_FILLS[i % 2]
        ws.row_dimensions[row_n].height = 18

        vals = [
            i,
            r['Company Name'],
            r['Category'],
            r['Roles / Skills'],
            r['Email 1'],
            r['Email 2'],
            r['Email 3'],
            r['Email 4'],
            r['Email 5'],
            r['Phone'],
            r['Website'],
            r['LinkedIn URL'],
            r['Address'],
        ]
        for col_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=row_n, column=col_idx, value=val)
            c.fill   = fill
            c.border = THIN_BORDER
            # Alignment
            if col_idx == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 13:  # Address - wrap
                c.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                c.alignment = Alignment(vertical="center", wrap_text=False)
            # Font / hyperlink style
            if col_idx in (5,6,7,8,9) and val:
                c.font = LINK_FONT
            elif col_idx in (11,12) and val:
                c.font = LINK_FONT
            elif col_idx == 2:
                c.font = BOLD_FONT
            elif col_idx == 3:
                c.font = Font(bold=True, size=10,
                              color="1A6600" if is_ai else "1F3864")
            else:
                c.font = NORM_FONT

    # Freeze pane below header, auto-filter on header
    ws.freeze_panes = ws.cell(row=data_start, column=1)
    ws.auto_filter.ref = (
        f"A{hdr_row}:{get_column_letter(len(COLUMNS))}{data_start + len(data_rows) - 1}"
    )

# ═══════════════════════════════════════════════════════════════════════════════
# 6. BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════
wb  = openpyxl.Workbook()

# ── Sheet 1: All companies ───────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "All IT & AIML Companies"
write_sheet(ws1, final_rows, title="IT & AI/ML Companies — Master Contact List (Ahmedabad & Gujarat)")

# ── Sheet 2: AI/ML companies only ────────────────────────────────────────────
ws2 = wb.create_sheet("AI ML Companies")
aiml_rows = [r for r in final_rows if r['Category'] == 'AI / ML']
write_sheet(ws2, aiml_rows, title="AI / ML Companies — Detailed Contact List (Ahmedabad & Gujarat)")

# ═══════════════════════════════════════════════════════════════════════════════
# 7. SAVE
# ═══════════════════════════════════════════════════════════════════════════════
output = 'Ahmedabad_IT_AIML_Companies_Master.xlsx'
wb.save(output)

aiml_count = len(aiml_rows)
it_count   = len(final_rows) - aiml_count
print(f"Saved: {output}")
print(f"  Sheet 1 - All companies : {len(final_rows)}")
print(f"  Sheet 2 - AI/ML only    : {aiml_count}")
print(f"  IT Services             : {it_count}")

# ─── DATA QUALITY REPORT ─────────────────────────────────────────────────────
email_re = re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')
print("\n" + "="*60)
print("DATA QUALITY REPORT")
print("="*60)
issues = []
for r in final_rows:
    for col in ['Email 1','Email 2','Email 3','Email 4','Email 5']:
        e = r.get(col,'').strip()
        if e:
            if not email_re.match(e):
                issues.append(f"INVALID | {r['Company Name']:28s} | {e}")
            if '..' in e:
                issues.append(f"DBL DOT | {r['Company Name']:28s} | {e}")
            if get_domain(e).endswith('.incom'):
                issues.append(f"TYPO    | {r['Company Name']:28s} | {e}")
if issues:
    print(f"[!] {len(issues)} issue(s) found:")
    for iss in issues:
        print("   ", iss)
else:
    print("[OK] All emails look valid.")
print("="*60)
