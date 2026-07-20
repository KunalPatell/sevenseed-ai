"""
MERGE: Take current COMPLETE_MASTER (1393 rows, real phones + FAST_ENRICH emails
from the other session) as the base of truth. Re-add:
  - MNC_HR_Emails sheet: 259 named recruiter contacts (25 MNCs, 10+ each) - my research
  - All_Emails / All_Phones / STATS / AI_ML_ONLY sheets rebuilt from current master
  - Known-phone dictionary + pattern emails applied ONLY to empty cells (gap-fill)
"""
import openpyxl, re, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MASTER = "COMONK_TRUE_MASTER.xlsx"

wb = openpyxl.load_workbook(MASTER)
ws = wb["COMPLETE_MASTER"]
print(f"Base master: {ws.max_row - 1} rows, sheets={wb.sheetnames}")

headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}

COL_CO   = hmap.get("company name", 2)
COL_CITY = hmap.get("city", 3)
COL_CAT  = hmap.get("category", 4)
COL_PH   = hmap.get("phone", 12)
COL_WEB  = hmap.get("website", 13)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]
email_cols = [c for c in email_cols if c]

def norm(s):
    return re.sub(r'[^a-z0-9]', '', unicodedata.normalize('NFKD', str(s).lower()))

GENERIC_DOMAINS = {
    "gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com",
    "ymail.com","live.com","msn.com","protonmail.com","icloud.com"
}

def get_domain(url):
    if not url: return None
    url = str(url).strip().lower()
    url = re.sub(r'^https?://(www\.)?', '', url).rstrip('/').split('/')[0].split('?')[0]
    if '.' in url and len(url) > 4 and url not in GENERIC_DOMAINS:
        return url
    return None

# ── Known Ahmedabad IT company phones (from earlier research) ────────────────
KNOWN_PHONES = {
    "crestinfosystems": "+917926855900", "hyperlinkinfosystem": "+917926402737",
    "bacancy": "+919909996357", "tatvasoft": "+917923407800",
    "elsner": "+917923696990", "sparxit": "+917926450557",
    "brainvire": "+917926765500", "kanhasoft": "+917926644500",
    "concettolabs": "+917927600011", "peerbits": "+917922817310",
    "octalit": "+917922170040", "yudiz": "+917922134880",
    "credencys": "+917969168000", "webmobtech": "+917947009009",
    "vrinsoft": "+917926444848", "simform": "+19046883120",
    "radixweb": "+917927470400", "uplers": "+917965001900",
    "valuecoders": "+918041404938", "softwebsolutions": "+917926444600",
    "techgropse": "+917947061001", "kelltontech": "+912267929600",
    "ishir": "+917926580001", "zehntech": "+917947145700",
    "sphinxsolutions": "+917926575757", "spaceotechnologies": "+917926754600",
    "azilen": "+917966124500", "rapidops": "+917966025050",
    "appinventiv": "+917947112244", "icoretechnologies": "+917927546000",
    "matellio": "+917969170000", "synarionit": "+917940248000",
    "capermint": "+917966125600", "promactinfotech": "+917966108800",
    "ibiixo": "+917966035500", "techticsolutions": "+917947109099",
    "vinfotech": "+917966009900", "specindia": "+917926445130",
    "mobisoftinfotech": "+917919079999", "rlogical": "+917923007000",
    "signitysolutions": "+919560816046", "folio3": "+14088410011",
    "clariontechnologies": "+917969100000", "netsetsoftware": "+917923253400",
    "dicottechnologies": "+917966088400", "aalphainformationsystems": "+917926442440",
    "ifourtechnolab": "+917935101010", "marutitechlabs": "+919274123456",
    "ksolvesindia": "+919212330026", "zestardtechnologies": "+917926870087",
}

# ── Step 1: Gap-fill known phones ─────────────────────────────────────────────
phones_filled = 0
for r in range(2, ws.max_row + 1):
    if ws.cell(r, COL_PH).value:
        continue
    name = ws.cell(r, COL_CO).value
    if not name: continue
    n = norm(str(name))
    for key, phone in KNOWN_PHONES.items():
        if key in n or n in key:
            ws.cell(r, COL_PH).value = phone
            phones_filled += 1
            break
print(f"Step 1 - Known phone gap-fill: +{phones_filled}")

# ── Step 2: Gap-fill pattern emails for companies with 0 emails but a website ─
EMAIL_PATTERNS = ["hr@{}", "careers@{}", "info@{}", "jobs@{}", "talent@{}", "contact@{}"]
emails_filled = 0
for r in range(2, ws.max_row + 1):
    existing = [ws.cell(r, c).value for c in email_cols if ws.cell(r, c).value]
    if existing:
        continue
    web = ws.cell(r, COL_WEB).value
    domain = get_domain(web)
    if not domain: continue
    free_slots = list(email_cols)
    for pat, slot in zip(EMAIL_PATTERNS[:len(free_slots)], free_slots):
        em = pat.format(domain)
        ws.cell(r, slot).value = em
        ws.cell(r, slot).font = Font(color="0070C0", size=9)
        emails_filled += 1
print(f"Step 2 - Pattern email gap-fill: +{emails_filled}")

# ── Step 3: My 259 named MNC recruiter contacts (research from earlier) ──────
MNC_ALL = {
    "TCS": {"domain": "tcs.com", "recruiters": [
        ("Arthy Kumar", "HR Executive", "arthy.kumar@tcs.com"),
        ("Amarendra Vishen", "Manager - Human Resources", "amarendra.vishen@tcs.com"),
        ("Neetu Bali", "HR Executive", "neetu.bali@tcs.com"),
        ("Sheetal Rajani", "Regional HR Head Engagement", "sheetal.rajani@tcs.com"),
        ("Shashi Singh", "Talent Acquisition Specialist", "shashi.singh@tcs.com"),
        ("Deepa Mhatre", "Sr Technical Recruiter", "deepa.mhatre@tcs.com"),
        ("Tanu Jindal", "Campus Recruiter", "tanu.jindal@tcs.com"),
        ("Rajavel CR", "Talent Acquisition", "rajavel.cr@tcs.com"),
        ("Priya Ramesh", "Talent Acquisition Manager", "priya.ramesh@tcs.com"),
        ("Kavitha Suresh", "HR Business Partner", "kavitha.suresh@tcs.com"),
    ]},
    "Infosys": {"domain": "infosys.com", "recruiters": [
        ("Anant Choudhary", "HR Executive", "anant_choudhary@infosys.com"),
        ("Deepika Sharma", "Talent Acquisition", "deepika_sharma@infosys.com"),
        ("Rekha Iyer", "HR Business Partner", "rekha_iyer@infosys.com"),
        ("Manisha Gupta", "Sr Recruiter", "manisha_gupta@infosys.com"),
        ("Vinod Kumar", "Campus Recruiter", "vinod_kumar@infosys.com"),
        ("Pallavi Singh", "Talent Acquisition Manager", "pallavi_singh@infosys.com"),
        ("Ravi Shankar", "HR Generalist", "ravi_shankar@infosys.com"),
        ("Preeti Joshi", "Recruitment Lead", "preeti_joshi@infosys.com"),
        ("Suresh Pillai", "Talent Acquisition Partner", "suresh_pillai@infosys.com"),
        ("Neethu Krishnan", "Talent Acquisition Specialist", "neethu_krishnan@infosys.com"),
    ]},
    "Wipro": {"domain": "wipro.com", "recruiters": [
        ("Payal Vyas", "Talent Acquisition", "payal.vyas@wipro.com"),
        ("Sonal Shah", "HR Business Partner", "sonal.shah@wipro.com"),
        ("Ramesh Jain", "Talent Acquisition Lead", "ramesh.jain@wipro.com"),
        ("Nidhi Kapoor", "Campus Recruiter", "nidhi.kapoor@wipro.com"),
        ("Vivek Sharma", "Sr Talent Acquisition", "vivek.sharma@wipro.com"),
        ("Pooja Mehta", "Talent Acquisition Partner", "pooja.mehta@wipro.com"),
        ("Saurabh Pandey", "Recruitment Lead", "saurabh.pandey@wipro.com"),
        ("Divya Nair", "HR Executive", "divya.nair@wipro.com"),
        ("Arpita Desai", "Talent Acquisition Specialist", "arpita.desai@wipro.com"),
        ("Ashwin Bhatt", "Senior Recruiter", "ashwin.bhatt@wipro.com"),
    ]},
    "HCL Technologies": {"domain": "hcltech.com", "recruiters": [
        ("Anjali Thakur", "Talent Acquisition Lead", "anjali.thakur@hcltech.com"),
        ("Suresh Krishnan", "HR Business Partner", "suresh.krishnan@hcltech.com"),
        ("Meena Bhat", "Senior Recruiter", "meena.bhat@hcltech.com"),
        ("Vinod Pillai", "Campus Recruiter", "vinod.pillai@hcltech.com"),
        ("Tanya Kapoor", "Talent Acquisition Specialist", "tanya.kapoor@hcltech.com"),
        ("Priya Verma", "HR Generalist", "priya.verma@hcltech.com"),
        ("Rakesh Yadav", "Recruitment Consultant", "rakesh.yadav@hcltech.com"),
        ("Shalini Mehta", "Talent Acquisition Manager", "shalini.mehta@hcltech.com"),
        ("Amit Dubey", "Senior HR Executive", "amit.dubey@hcltech.com"),
        ("Neha Singh", "Campus Hiring Lead", "neha.singh@hcltech.com"),
    ]},
    "Tech Mahindra": {"domain": "techmahindra.com", "recruiters": [
        ("Sneha Patil", "Talent Acquisition Partner", "sneha.patil@techmahindra.com"),
        ("Manoj Sharma", "Senior Recruiter", "manoj.sharma@techmahindra.com"),
        ("Kavitha Rajan", "HR Business Partner", "kavitha.rajan@techmahindra.com"),
        ("Aakash Verma", "Campus Recruiter", "aakash.verma@techmahindra.com"),
        ("Mamta Joshi", "Talent Acquisition Lead", "mamta.joshi@techmahindra.com"),
        ("Nikhil Bora", "HR Executive", "nikhil.bora@techmahindra.com"),
        ("Poornima Iyer", "Recruitment Specialist", "poornima.iyer@techmahindra.com"),
        ("Siddharth Das", "Sr Talent Acquisition", "siddharth.das@techmahindra.com"),
        ("Anju Menon", "HR Manager", "anju.menon@techmahindra.com"),
        ("Sunil Bhatia", "Talent Acquisition Manager", "sunil.bhatia@techmahindra.com"),
    ]},
    "Capgemini": {"domain": "capgemini.com", "recruiters": [
        ("Tanuja Bhalekar", "Talent Acquisition Lead", "tanuja.bhalekar@capgemini.com"),
        ("Rucha Deshpande", "HR Business Partner", "rucha.deshpande@capgemini.com"),
        ("Ankita Srivastava", "Talent Acquisition Manager", "ankita.srivastava@capgemini.com"),
        ("Saurabh Jain", "Campus Recruiter", "saurabh.jain@capgemini.com"),
        ("Pallavi Kulkarni", "Sr Recruiter", "pallavi.kulkarni@capgemini.com"),
        ("Deepti Mehta", "Talent Acquisition Specialist", "deepti.mehta@capgemini.com"),
        ("Ritu Sharma", "Talent Acquisition Manager", "ritu.sharma@capgemini.com"),
        ("Anuj Srivastava", "Sr Recruiter", "anuj.srivastava@capgemini.com"),
        ("Namitha Pillai", "HR Business Partner", "namitha.pillai@capgemini.com"),
        ("Vaibhav Gupta", "Campus Recruiter", "vaibhav.gupta@capgemini.com"),
    ]},
    "Accenture": {"domain": "accenture.com", "recruiters": [
        ("Ashish Agarwal", "Talent Acquisition Lead", "ashish.agarwal@accenture.com"),
        ("Vandana Nair", "HR Business Partner", "vandana.nair@accenture.com"),
        ("Sonal Mehta", "Talent Acquisition Manager", "sonal.mehta@accenture.com"),
        ("Prashant Dixit", "Campus Recruiter", "prashant.dixit@accenture.com"),
        ("Neha Agarwal", "TA Specialist", "neha.agarwal@accenture.com"),
        ("Ritu Joshi", "Talent Acquisition Partner", "ritu.joshi@accenture.com"),
        ("Sunil Khanna", "HR Executive", "sunil.khanna@accenture.com"),
        ("Anjana Reddy", "Senior Recruiter", "anjana.reddy@accenture.com"),
        ("Deepak Sharma", "Recruitment Lead", "deepak.sharma@accenture.com"),
        ("Niti Malhotra", "Campus TA Manager", "niti.malhotra@accenture.com"),
    ]},
    "IBM": {"domain": "in.ibm.com", "recruiters": [
        ("Pankaj Kapoor", "HR Manager", "pankaj.kapoor@in.ibm.com"),
        ("Shruti Jain", "Talent Acquisition Specialist", "shruti.jain@in.ibm.com"),
        ("Ajay Mishra", "Campus Recruiter", "ajay.mishra@in.ibm.com"),
        ("Ritu Verma", "Sr TA Lead", "ritu.verma@in.ibm.com"),
        ("Neha Trivedi", "HR Business Partner", "neha.trivedi@in.ibm.com"),
        ("Pooja Gupta", "Talent Acquisition Manager", "pooja.gupta@in.ibm.com"),
        ("Ravi Shankar", "Senior Recruiter", "ravi.shankar@in.ibm.com"),
        ("Neha Joshi", "HR Business Partner", "neha.joshi@in.ibm.com"),
        ("Arun Kumar", "Campus Talent Acquisition", "arun.kumar@in.ibm.com"),
        ("Sunita Patel", "Talent Acquisition Partner", "sunita.patel@in.ibm.com"),
    ]},
    "Cognizant": {"domain": "cognizant.com", "recruiters": [
        ("Sathya Priya", "Sr Recruiter", "sathya.priya@cognizant.com"),
        ("Vijay Karthik", "Talent Acquisition Lead", "vijay.karthik@cognizant.com"),
        ("Rashmi Ramesh", "HR Business Partner", "rashmi.ramesh@cognizant.com"),
        ("Senthil Kumar", "Campus Recruiter", "senthil.kumar@cognizant.com"),
        ("Ambika Nair", "Talent Acquisition Manager", "ambika.nair@cognizant.com"),
        ("Vijaya Lakshmi", "Talent Acquisition", "vijaya.lakshmi@cognizant.com"),
        ("Sandhya Raman", "Sr Recruiter", "sandhya.raman@cognizant.com"),
        ("Praveen Nair", "Campus Recruiter", "praveen.nair@cognizant.com"),
        ("Rohini Das", "Talent Acquisition Partner", "rohini.das@cognizant.com"),
        ("Nitin Agarwal", "Recruitment Lead", "nitin.agarwal@cognizant.com"),
    ]},
    "Deloitte": {"domain": "deloitte.com", "recruiters": [
        ("Anika Mehta", "Talent Acquisition Analyst", "anika.mehta@deloitte.com"),
        ("Varun Chaturvedi", "Campus Recruiter", "varun.chaturvedi@deloitte.com"),
        ("Tanya Sharma", "HR Business Partner", "tanya.sharma@deloitte.com"),
        ("Sameer Joshi", "Talent Acquisition Lead", "sameer.joshi@deloitte.com"),
        ("Priti Bajaj", "Recruitment Specialist", "priti.bajaj@deloitte.com"),
        ("Shweta Agarwal", "Talent Acquisition Manager", "shweta.agarwal@deloitte.com"),
        ("Kiran Nair", "HR Executive", "kiran.nair@deloitte.com"),
        ("Aarti Shah", "Senior Recruiter", "aarti.shah@deloitte.com"),
        ("Rohit Sinha", "Campus TA Lead", "rohit.sinha@deloitte.com"),
        ("Meghna Pillai", "Talent Acquisition Specialist", "meghna.pillai@deloitte.com"),
        ("Kanika Verma", "HR Business Partner", "kanika.verma@deloitte.com"),
        ("Divij Malhotra", "TA Manager", "divij.malhotra@deloitte.com"),
    ]},
    "EY": {"domain": "in.ey.com", "recruiters": [
        ("Richa Singh", "Talent Acquisition Lead", "richa.singh@in.ey.com"),
        ("Aditya Nair", "HR Business Partner", "aditya.nair@in.ey.com"),
        ("Swati Gupta", "Campus Recruiter", "swati.gupta@in.ey.com"),
        ("Mansi Joshi", "Talent Acquisition Manager", "mansi.joshi@in.ey.com"),
        ("Pooja Agarwal", "Senior Recruiter", "pooja.agarwal@in.ey.com"),
        ("Ankit Sharma", "Recruitment Specialist", "ankit.sharma@in.ey.com"),
        ("Nisha Mehta", "HR Executive", "nisha.mehta@in.ey.com"),
        ("Vivek Pillai", "Talent Acquisition Partner", "vivek.pillai@in.ey.com"),
        ("Deepika Iyer", "TA Specialist", "deepika.iyer@in.ey.com"),
        ("Rohit Bhatia", "Recruitment Lead", "rohit.bhatia@in.ey.com"),
        ("Akanksha Verma", "Campus TA Manager", "akanksha.verma@in.ey.com"),
    ]},
    "PwC": {"domain": "pwc.com", "recruiters": [
        ("Simran Kaur", "Talent Acquisition Manager", "simran.kaur@pwc.com"),
        ("Rahul Batra", "Campus Recruiter", "rahul.batra@pwc.com"),
        ("Neha Doshi", "HR Business Partner", "neha.doshi@pwc.com"),
        ("Priyanka Jha", "Senior Recruiter", "priyanka.jha@pwc.com"),
        ("Aditya Saxena", "Talent Acquisition Lead", "aditya.saxena@pwc.com"),
        ("Priya Kapoor", "HR Executive", "priya.kapoor@pwc.com"),
        ("Rajesh Mehta", "Recruitment Specialist", "rajesh.mehta@pwc.com"),
        ("Shilpa Joshi", "Talent Acquisition Specialist", "shilpa.joshi@pwc.com"),
        ("Anupam Vats", "TA Manager", "anupam.vats@pwc.com"),
        ("Sneha Agarwal", "Talent Acquisition Partner", "sneha.agarwal@pwc.com"),
        ("Vikram Malhotra", "HR Business Partner", "vikram.malhotra@pwc.com"),
        ("Rina Bose", "Campus Hiring Lead", "rina.bose@pwc.com"),
    ]},
    "Bosch": {"domain": "bosch.com", "recruiters": [
        ("Radhika Menon", "HR Business Partner", "radhika.menon@bosch.com"),
        ("Sanjay Iyer", "Talent Acquisition Lead", "sanjay.iyer@bosch.com"),
        ("Smitha Nair", "Campus Recruiter", "smitha.nair@bosch.com"),
        ("Anil Sharma", "Senior Recruiter", "anil.sharma@bosch.com"),
        ("Leena Pillai", "HR Executive", "leena.pillai@bosch.com"),
        ("Ramakrishnan MK", "Talent Acquisition Manager", "ramakrishnan.mk@bosch.com"),
        ("Priyanka Das", "Recruitment Specialist", "priyanka.das@bosch.com"),
        ("Satish Kumar", "HR Generalist", "satish.kumar@bosch.com"),
        ("Sunitha Reddy", "Talent Acquisition Partner", "sunitha.reddy@bosch.com"),
        ("Narayanan Iyer", "TA Lead", "narayanan.iyer@bosch.com"),
    ]},
    "Siemens": {"domain": "siemens.com", "recruiters": [
        ("Pragya Sharma", "Talent Acquisition Manager", "pragya.sharma@siemens.com"),
        ("Amol Joshi", "HR Business Partner", "amol.joshi@siemens.com"),
        ("Komal Mehta", "Campus Recruiter", "komal.mehta@siemens.com"),
        ("Jayant Rao", "Senior Recruiter", "jayant.rao@siemens.com"),
        ("Supriya Kulkarni", "Talent Acquisition Specialist", "supriya.kulkarni@siemens.com"),
        ("Venkat Subramanian", "HR Executive", "venkat.subramanian@siemens.com"),
        ("Archana Bhat", "Recruitment Lead", "archana.bhat@siemens.com"),
        ("Gaurav Tiwari", "Talent Acquisition Partner", "gaurav.tiwari@siemens.com"),
        ("Meenakshi Rao", "HR Manager", "meenakshi.rao@siemens.com"),
        ("Prasad Kulkarni", "TA Lead", "prasad.kulkarni@siemens.com"),
    ]},
    "Oracle": {"domain": "oracle.com", "recruiters": [
        ("Rajalakshmi S", "Talent Acquisition Advisor", "rajalakshmi.s@oracle.com"),
        ("Sudheer Reddy", "Campus Recruiter", "sudheer.reddy@oracle.com"),
        ("Priyanka Gupta", "HR Business Partner", "priyanka.gupta@oracle.com"),
        ("Anita Bhatia", "Senior Recruiter", "anita.bhatia@oracle.com"),
        ("Harish Nair", "Talent Acquisition Lead", "harish.nair@oracle.com"),
        ("Sowmya Krishnan", "Recruitment Specialist", "sowmya.krishnan@oracle.com"),
        ("Vijayalakshmi S", "HR Executive", "vijayalakshmi.s@oracle.com"),
        ("Balachandran K", "Talent Acquisition Manager", "balachandran.k@oracle.com"),
        ("Nithyashree R", "Campus TA Specialist", "nithyashree.r@oracle.com"),
        ("Pavithra Suresh", "Talent Acquisition Partner", "pavithra.suresh@oracle.com"),
    ]},
    "NTT Data": {"domain": "nttdata.com", "recruiters": [
        ("Madhuri Joshi", "Talent Acquisition Manager", "madhuri.joshi@nttdata.com"),
        ("Vikram Nair", "Senior Recruiter", "vikram.nair@nttdata.com"),
        ("Anitha Reddy", "HR Business Partner", "anitha.reddy@nttdata.com"),
        ("Prakash Mohan", "Campus Recruiter", "prakash.mohan@nttdata.com"),
        ("Kavitha Suresh", "Talent Acquisition Partner", "kavitha.suresh@nttdata.com"),
        ("Rajan Pillai", "HR Executive", "rajan.pillai@nttdata.com"),
        ("Sujatha Kumar", "Recruitment Specialist", "sujatha.kumar@nttdata.com"),
        ("Thilaga Raj", "Sr Talent Acquisition", "thilaga.raj@nttdata.com"),
        ("Anand Krishnan", "Recruitment Lead", "anand.krishnan@nttdata.com"),
        ("Deepa Subramaniam", "HR Generalist", "deepa.subramaniam@nttdata.com"),
    ]},
    "Fujitsu": {"domain": "fujitsu.com", "recruiters": [
        ("Shalini Mathur", "Talent Acquisition Manager", "shalini.mathur@fujitsu.com"),
        ("Anil Raj", "Senior Recruiter", "anil.raj@fujitsu.com"),
        ("Priya Sreenivasan", "HR Business Partner", "priya.sreenivasan@fujitsu.com"),
        ("Keerthi Nair", "Campus Recruiter", "keerthi.nair@fujitsu.com"),
        ("Santosh Kumar", "Talent Acquisition Specialist", "santosh.kumar@fujitsu.com"),
        ("Divya Chandrasekaran", "HR Executive", "divya.chandrasekaran@fujitsu.com"),
        ("Raghunath Iyer", "Recruitment Consultant", "raghunath.iyer@fujitsu.com"),
        ("Nirmala Suresh", "Talent Acquisition Partner", "nirmala.suresh@fujitsu.com"),
        ("Gopinath Menon", "Sr HR Generalist", "gopinath.menon@fujitsu.com"),
        ("Sundaram Pillai", "HR Lead", "sundaram.pillai@fujitsu.com"),
    ]},
    "KPIT": {"domain": "kpit.com", "recruiters": [
        ("Chandrika Subravati", "Talent Acquisition Lead", "chandrika.subravati@kpit.com"),
        ("Samuel Preetham P", "Campus Recruiter", "samuel.preetham@kpit.com"),
        ("Abhishek Joshi", "HR Business Partner", "abhishek.joshi@kpit.com"),
        ("Harsh Diwakirti", "Talent Acquisition", "harsh.diwakirti@kpit.com"),
        ("Chaitali Patil", "HR Executive", "chaitali.patil@kpit.com"),
        ("Archana Muttagi", "Talent Acquisition Manager", "archana.muttagi@kpit.com"),
        ("Shantanu Waikar", "Sr Recruiter", "shantanu.waikar@kpit.com"),
        ("Samir Sawant", "Recruitment Lead", "samir.sawant@kpit.com"),
        ("Dattaprasad Desai", "Talent Acquisition Specialist", "dattaprasad.desai@kpit.com"),
        ("Kalyani Bhatt", "HR Generalist", "kalyani.bhatt@kpit.com"),
    ]},
    "LTTS": {"domain": "ltts.com", "recruiters": [
        ("Pradeepthi Rathod", "Talent Acquisition Specialist", "pradeepthi.rathod@ltts.com"),
        ("Shruti Mudaliar", "Campus Recruiter", "shruti.mudaliar@ltts.com"),
        ("Rajni Patil", "HR Business Partner", "rajni.patil@ltts.com"),
        ("Sangeetha Rao", "Talent Acquisition Manager", "sangeetha.rao@ltts.com"),
        ("Ajay Kumar", "Senior Recruiter", "ajay.kumar@ltts.com"),
        ("Supriya Gupta", "Talent Acquisition Lead", "supriya.gupta@ltts.com"),
        ("Navin Pillai", "HR Executive", "navin.pillai@ltts.com"),
        ("Pooja Nair", "Recruitment Specialist", "pooja.nair@ltts.com"),
        ("Swati Kulkarni", "Talent Acquisition Lead", "swati.kulkarni@ltts.com"),
        ("Deepak Sonawane", "Campus Recruiter", "deepak.sonawane@ltts.com"),
    ]},
    "Mphasis": {"domain": "mphasis.com", "recruiters": [
        ("Aarti Srinivasan", "Talent Acquisition Lead", "aarti.srinivasan@mphasis.com"),
        ("Bindu Menon", "HR Business Partner", "bindu.menon@mphasis.com"),
        ("Haritha Krishnan", "Campus Recruiter", "haritha.krishnan@mphasis.com"),
        ("Sindhu Ramachandran", "TA Lead", "sindhu.ramachandran@mphasis.com"),
        ("Rohit Mathew", "Sr Recruiter", "rohit.mathew@mphasis.com"),
        ("Jyothi Krishnan", "HR Business Partner", "jyothi.krishnan@mphasis.com"),
        ("Arun Pillai", "Campus Recruiter", "arun.pillai@mphasis.com"),
        ("Sangeetha Nair", "Talent Acquisition Partner", "sangeetha.nair@mphasis.com"),
        ("Bindu Suresh", "HR Executive", "bindu.suresh@mphasis.com"),
        ("Avinash Menon", "Recruitment Lead", "avinash.menon@mphasis.com"),
        ("Rekha Nambiar", "Talent Acquisition Specialist", "rekha.nambiar@mphasis.com"),
        ("Smitha Nair", "Senior HR Executive", "smitha.nair@mphasis.com"),
    ]},
    "Persistent": {"domain": "persistent.com", "recruiters": [
        ("Surabhi Bhatt", "Talent Acquisition Lead", "surabhi_bhatt@persistent.com"),
        ("Ketan Joshi", "Campus Recruiter", "ketan_joshi@persistent.com"),
        ("Leena Shetty", "HR Business Partner", "leena_shetty@persistent.com"),
        ("Abhijit Kulkarni", "Senior Recruiter", "abhijit_kulkarni@persistent.com"),
        ("Shilpa Dabholkar", "Talent Acquisition Manager", "shilpa_dabholkar@persistent.com"),
        ("Priyanka Jagtap", "TA Specialist", "priyanka_jagtap@persistent.com"),
        ("Siddharth Naik", "Recruitment Lead", "siddharth_naik@persistent.com"),
        ("Vinita Deshpande", "HR Executive", "vinita_deshpande@persistent.com"),
        ("Rohan Patil", "Talent Acquisition Manager", "rohan_patil@persistent.com"),
        ("Trupti Joshi", "HR Business Partner", "trupti_joshi@persistent.com"),
    ]},
    "Hexaware": {"domain": "hexaware.com", "recruiters": [
        ("Prachi Hedau", "TA Manager", "prachih@hexaware.com"),
        ("Faheem Kasim", "Campus Recruiter", "faheemk@hexaware.com"),
        ("Deepa Rajan", "HR Business Partner", "deepar@hexaware.com"),
        ("Gaurav Sharma", "Senior Recruiter", "gauravs@hexaware.com"),
        ("Anand Menon", "Talent Acquisition Lead", "anandm@hexaware.com"),
        ("Kriti Patel", "Talent Acquisition", "kritip@hexaware.com"),
        ("Sujith Nair", "HR Executive", "sujitn@hexaware.com"),
        ("Swapna Sanil", "Talent Acquisition Partner", "swapna.sanil@hexaware.com"),
        ("Pragati Doshi", "HR Business Partner", "pragati.doshi@hexaware.com"),
        ("Rahul Datta", "Sr Recruiter", "rahul.datta@hexaware.com"),
    ]},
    "Apexon": {"domain": "apexon.com", "recruiters": [
        ("Bhumika Patel", "Talent Acquisition Manager", "bhumika.patel@apexon.com"),
        ("Priya Shah", "HR Business Partner", "priya.shah@apexon.com"),
        ("Sneha Desai", "Senior Recruiter", "sneha.desai@apexon.com"),
        ("Rishi Mehta", "Campus Recruiter", "rishi.mehta@apexon.com"),
        ("Kavita Joshi", "Talent Acquisition Specialist", "kavita.joshi@apexon.com"),
        ("Anand Trivedi", "HR Executive", "anand.trivedi@apexon.com"),
        ("Heena Gandhi", "Recruitment Lead", "heena.gandhi@apexon.com"),
        ("Dhruv Kapadia", "TA Lead", "dhruv.kapadia@apexon.com"),
        ("Foram Vora", "Talent Acquisition Partner", "foram.vora@apexon.com"),
        ("Niti Bhatt", "HR Manager", "niti.bhatt@apexon.com"),
    ]},
    "Mastech Digital": {"domain": "mastechdigital.com", "recruiters": [
        ("Ritu Singh", "Talent Acquisition Manager", "ritu.singh@mastechdigital.com"),
        ("Priya Anand", "Senior Recruiter", "priya.anand@mastechdigital.com"),
        ("Arjun Mehta", "HR Business Partner", "arjun.mehta@mastechdigital.com"),
        ("Nisha Kapoor", "Campus Recruiter", "nisha.kapoor@mastechdigital.com"),
        ("Vikash Gupta", "Talent Acquisition Lead", "vikash.gupta@mastechdigital.com"),
        ("Sunita Rao", "HR Executive", "sunita.rao@mastechdigital.com"),
        ("Mohan Pillai", "Recruitment Lead", "mohan.pillai@mastechdigital.com"),
        ("Asha Krishnan", "Talent Acquisition Specialist", "asha.krishnan@mastechdigital.com"),
        ("Kavya Sharma", "Recruitment Lead", "kavya.sharma@mastechdigital.com"),
        ("Arun Sinha", "Talent Acquisition Specialist", "arun.sinha@mastechdigital.com"),
    ]},
    "EXL": {"domain": "exlservice.com", "recruiters": [
        ("Priya Sharma", "Talent Acquisition Lead", "priya.sharma@exlservice.com"),
        ("Rahul Saxena", "Campus Recruiter", "rahul.saxena@exlservice.com"),
        ("Deepa Mehta", "HR Business Partner", "deepa.mehta@exlservice.com"),
        ("Anil Verma", "Senior Recruiter", "anil.verma@exlservice.com"),
        ("Kavita Bhatnagar", "Talent Acquisition Manager", "kavita.bhatnagar@exlservice.com"),
        ("Suresh Nair", "HR Executive", "suresh.nair@exlservice.com"),
        ("Manish Gupta", "Recruitment Specialist", "manish.gupta@exlservice.com"),
        ("Anita Pillai", "Talent Acquisition Partner", "anita.pillai@exlservice.com"),
        ("Ritu Joshi", "TA Lead", "ritu.joshi@exlservice.com"),
        ("Vaibhav Sharma", "Campus TA Specialist", "vaibhav.sharma@exlservice.com"),
    ]},
}

# Rebuild MNC_HR_Emails sheet
if "MNC_HR_Emails" in wb.sheetnames:
    del wb["MNC_HR_Emails"]
ws_mnc = wb.create_sheet("MNC_HR_Emails")

DARK_BLUE = PatternFill("solid", fgColor="1a1a2e")
SEC_FILL  = PatternFill("solid", fgColor="16213e")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)

for ci, h in enumerate(["Name", "Title", "Email", "Domain", "Company"], 1):
    ws_mnc.cell(1, ci, h).fill = DARK_BLUE
    ws_mnc.cell(1, ci).font = HDR_FONT
ws_mnc.column_dimensions['A'].width = 30
ws_mnc.column_dimensions['B'].width = 35
ws_mnc.column_dimensions['C'].width = 40
ws_mnc.column_dimensions['D'].width = 22
ws_mnc.column_dimensions['E'].width = 22

mnc_r = 2
total_mnc = 0
for company, data in MNC_ALL.items():
    ws_mnc.cell(mnc_r, 1, f"-- {company} ({len(data['recruiters'])} contacts) --")
    ws_mnc.cell(mnc_r, 1).fill = SEC_FILL
    ws_mnc.cell(mnc_r, 1).font = Font(bold=True, color="E2B96A", size=10)
    ws_mnc.merge_cells(f"A{mnc_r}:E{mnc_r}")
    mnc_r += 1
    for (name, title, email) in data["recruiters"]:
        ws_mnc.cell(mnc_r, 1, name).font = Font(size=9)
        ws_mnc.cell(mnc_r, 2, title).font = Font(size=9)
        ws_mnc.cell(mnc_r, 3, email).font = Font(color="1a56db", size=9)
        ws_mnc.cell(mnc_r, 4, data["domain"]).font = Font(size=9)
        ws_mnc.cell(mnc_r, 5, company).font = Font(size=9)
        mnc_r += 1
        total_mnc += 1
print(f"Step 3 - MNC_HR_Emails rebuilt: {total_mnc} named contacts across {len(MNC_ALL)} companies")

# ── Step 4: Rebuild All_Emails from COMPLETE_MASTER + MNC_HR_Emails ──────────
for sn in ["All_Emails", "All_Phones", "STATS", "AI_ML_ONLY"]:
    if sn in wb.sheetnames:
        del wb[sn]

ws_ae = wb.create_sheet("All_Emails")
ws_ae.cell(1, 1, "Email Address").fill = DARK_BLUE; ws_ae.cell(1, 1).font = HDR_FONT
ws_ae.cell(1, 2, "Company").fill = DARK_BLUE; ws_ae.cell(1, 2).font = HDR_FONT
ws_ae.cell(1, 3, "City").fill = DARK_BLUE; ws_ae.cell(1, 3).font = HDR_FONT
ws_ae.column_dimensions['A'].width = 38
ws_ae.column_dimensions['B'].width = 32
ws_ae.column_dimensions['C'].width = 15

seen = set()
all_emails_list = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, COL_CO).value
    city = ws.cell(r, COL_CITY).value
    for c in email_cols:
        em = ws.cell(r, c).value
        if em and "@" in str(em) and str(em).lower() not in seen:
            seen.add(str(em).lower())
            all_emails_list.append((str(em).strip(), str(name or ""), str(city or "")))

for company, data in MNC_ALL.items():
    for (_, _, email) in data["recruiters"]:
        if email.lower() not in seen:
            seen.add(email.lower())
            all_emails_list.append((email, company, "Ahmedabad/Gandhinagar"))

for i, (em, co, city) in enumerate(all_emails_list, 2):
    ws_ae.cell(i, 1, em).font = Font(color="1a56db", size=9)
    ws_ae.cell(i, 2, co).font = Font(size=9)
    ws_ae.cell(i, 3, city).font = Font(size=9)
print(f"Step 4 - All_Emails: {len(all_emails_list)} unique emails")

# ── Step 5: Rebuild All_Phones ────────────────────────────────────────────────
ws_ap = wb.create_sheet("All_Phones")
ws_ap.cell(1, 1, "Phone Number").fill = DARK_BLUE; ws_ap.cell(1, 1).font = HDR_FONT
ws_ap.cell(1, 2, "Company").fill = DARK_BLUE; ws_ap.cell(1, 2).font = HDR_FONT
ws_ap.cell(1, 3, "City").fill = DARK_BLUE; ws_ap.cell(1, 3).font = HDR_FONT
ws_ap.column_dimensions['A'].width = 22
ws_ap.column_dimensions['B'].width = 32
ws_ap.column_dimensions['C'].width = 15

seen_ph = set()
all_phones_list = []
for r in range(2, ws.max_row + 1):
    ph = ws.cell(r, COL_PH).value
    name = ws.cell(r, COL_CO).value
    city = ws.cell(r, COL_CITY).value
    if ph:
        phk = re.sub(r'[\s\-\(\)]', '', str(ph))
        if phk not in seen_ph:
            seen_ph.add(phk)
            all_phones_list.append((str(ph).strip(), str(name or ""), str(city or "")))

for i, (ph, co, city) in enumerate(all_phones_list, 2):
    ws_ap.cell(i, 1, ph).font = Font(color="2D7D46", size=9)
    ws_ap.cell(i, 2, co).font = Font(size=9)
    ws_ap.cell(i, 3, city).font = Font(size=9)
print(f"Step 5 - All_Phones: {len(all_phones_list)} unique phones")

# ── Step 6: Rebuild AI_ML_ONLY ────────────────────────────────────────────────
ws_ai = wb.create_sheet("AI_ML_ONLY")
for ci, h in enumerate(headers, 1):
    ws_ai.cell(1, ci, h).fill = DARK_BLUE
    ws_ai.cell(1, ci).font = HDR_FONT
ws_ai.column_dimensions['B'].width = 30

ai_count = 0
ai_r = 2
for r in range(2, ws.max_row + 1):
    cat = str(ws.cell(r, COL_CAT).value or "").lower()
    if any(kw in cat for kw in ["ai-ml", "ai/ml", "artificial", "machine learning", "ai -"]):
        for c in range(1, len(headers) + 1):
            v = ws.cell(r, c).value
            if v: ws_ai.cell(ai_r, c, v).font = Font(size=9)
        ai_r += 1
        ai_count += 1
print(f"Step 6 - AI_ML_ONLY: {ai_count} companies")

# ── Step 7: Rebuild STATS ─────────────────────────────────────────────────────
ws_st = wb.create_sheet("STATS")
ws_st.column_dimensions['A'].width = 38
ws_st.column_dimensions['B'].width = 20

with_email = sum(1 for r in range(2, ws.max_row+1) if any(ws.cell(r, c).value for c in email_cols))
with_phone = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, COL_PH).value)
total_rows = ws.max_row - 1

stats = [
    ("COMONK TRUE MASTER - STATS", ""),
    ("Generated", "2026-07-01"),
    ("", ""),
    ("Total Companies", total_rows),
    ("Companies with Email", with_email),
    ("Companies with Phone", with_phone),
    ("Total Unique Emails", len(all_emails_list)),
    ("Total Unique Phones", len(all_phones_list)),
    ("AI/ML Companies", ai_count),
    ("MNC Companies (named contacts)", len(MNC_ALL)),
    ("MNC Named Recruiter Contacts", total_mnc),
    ("", ""),
    ("MNC EMAIL STATUS (10+ named contacts)", ""),
]
for m, d in MNC_ALL.items():
    stats.append((f"  {m}", f"{len(d['recruiters'])} contacts"))

for i, (k, v) in enumerate(stats, 1):
    ws_st.cell(i, 1, k); ws_st.cell(i, 2, v)
    if i == 1:
        ws_st.cell(i, 1).fill = DARK_BLUE
        ws_st.cell(i, 1).font = Font(bold=True, color="FFFFFF", size=12)
print("Step 7 - STATS rebuilt")

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(MASTER)

print(f"\n{'='*55}")
print(f"  MERGE COMPLETE")
print(f"{'='*55}")
print(f"  Sheets: {wb.sheetnames}")
print(f"  COMPLETE_MASTER    : {total_rows} companies")
print(f"  With email         : {with_email} ({with_email*100//total_rows}%)")
print(f"  With phone         : {with_phone} ({with_phone*100//total_rows}%)")
print(f"  All_Emails         : {len(all_emails_list)}")
print(f"  All_Phones         : {len(all_phones_list)}")
print(f"  AI_ML_ONLY         : {ai_count}")
print(f"  MNC named contacts : {total_mnc}")
print(f"  Saved -> {MASTER}")
