"""
DEEP ENRICHMENT: Phone + Email for all companies missing data
1. Apollo.io org enrichment for missing phones/LinkedIn
2. Web scrape contact pages for company emails
"""
import openpyxl, re, httpx, asyncio, time, os
from openpyxl.styles import Font

MASTER  = "COMONK_TRUE_MASTER.xlsx"
API_KEY = os.getenv("APOLLO_API_KEY", "eCItGCG1bp9yfTex_HPlEQ")

EMAIL_RE = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
PHONE_RE = re.compile(r'(?:\+91[\s\-]?)?(?:0)?[6-9]\d{9}')
C_EMAILS = [6, 7, 8, 9, 10, 11]
C_PHONE  = 12
C_WEB    = 13
C_ADDR   = 15

GENERIC_DOMAINS = {
    "gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com",
    "ymail.com","live.com","msn.com","protonmail.com","icloud.com","example.com"
}

wb = openpyxl.load_workbook(MASTER)
ws = wb["COMPLETE_MASTER"]
total_rows = ws.max_row - 1
print(f"Master has {total_rows} companies")

missing_phone, missing_email = [], []
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 2).value
    web  = ws.cell(r, C_WEB).value or ""
    ph   = ws.cell(r, C_PHONE).value
    em   = [ws.cell(r, c).value for c in C_EMAILS if ws.cell(r, c).value]
    if name:
        if not ph: missing_phone.append((r, name, str(web).strip()))
        if not em: missing_email.append((r, name, str(web).strip()))

print(f"Missing phone: {len(missing_phone)} | Missing email: {len(missing_email)}")

def get_domain(url):
    if not url or url in ("", "None"): return None
    url = str(url).strip().lower()
    url = re.sub(r'^https?://(www\.)?', '', url).rstrip('/').split('/')[0].split('?')[0]
    if '.' in url and len(url) > 4 and url not in GENERIC_DOMAINS:
        return url
    return None

# ── Phase 1: Apollo Org Enrichment ───────────────────────────────────────────
async def apollo_enrich_all(companies, limit=300):
    to_do = [(r, name, web) for r, name, web in companies if get_domain(web)][:limit]
    print(f"\n--- Apollo enrichment: {len(to_do)} companies ---")
    phones_added = 0
    async with httpx.AsyncClient(timeout=12) as client:
        for i, (r, name, web) in enumerate(to_do):
            domain = get_domain(web)
            try:
                resp = await client.get(
                    "https://api.apollo.io/v1/organizations/enrich",
                    params={"api_key": API_KEY, "domain": domain}
                )
                if resp.status_code == 200:
                    org = resp.json().get("organization") or {}
                    ph = org.get("phone") or (org.get("primary_phone") or {}).get("number")
                    if ph and not ws.cell(r, C_PHONE).value:
                        ws.cell(r, C_PHONE).value = ph
                        phones_added += 1
                    addr_parts = [org.get("street_address"), org.get("city"),
                                  org.get("state"), org.get("country")]
                    addr = ", ".join(p for p in addr_parts if p)
                    if addr and not ws.cell(r, C_ADDR).value:
                        ws.cell(r, C_ADDR).value = addr
                    li = org.get("linkedin_url")
                    if li and not ws.cell(r, 14).value:
                        ws.cell(r, 14).value = li
            except Exception:
                pass
            if i % 25 == 0 and i > 0:
                print(f"  Apollo {i}/{len(to_do)} | +{phones_added} phones")
            await asyncio.sleep(0.25)
    print(f"  Apollo done: +{phones_added} phones")
    return phones_added

# ── Phase 2: Web scrape emails ────────────────────────────────────────────────
CONTACT_PATHS = ["/contact", "/contact-us", "/contactus", "/about/contact", "/reach-us"]

async def scrape_company_emails(client, url, domain):
    found = set()
    urls = [url] + [f"https://{domain}{p}" for p in CONTACT_PATHS[:3]]
    for u in urls[:3]:
        try:
            resp = await client.get(u, timeout=8, follow_redirects=True,
                headers={"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36"})
            for em in EMAIL_RE.findall(resp.text):
                em = em.lower().rstrip('.')
                dom = em.split("@")[1] if "@" in em else ""
                if dom and dom not in GENERIC_DOMAINS and (domain.split('.')[0] in dom or dom in domain):
                    found.add(em)
            if found: break
        except Exception:
            pass
    return found

async def scrape_emails_all(companies, limit=400):
    to_do = [(r, name, web) for r, name, web in companies if get_domain(web)][:limit]
    print(f"\n--- Email scraping: {len(to_do)} companies ---")
    emails_added = 0; cos_enriched = 0
    sem = asyncio.Semaphore(15)

    async def one(r, name, web, client):
        nonlocal emails_added, cos_enriched
        async with sem:
            domain = get_domain(web)
            if not domain: return
            ems = await scrape_company_emails(client, web, domain)
            if ems:
                existing = {str(ws.cell(r, c).value or "").lower() for c in C_EMAILS}
                new_ems  = [e for e in ems if e not in existing]
                free_slots = [c for c in C_EMAILS if not ws.cell(r, c).value]
                for em, slot in zip(new_ems, free_slots):
                    ws.cell(r, slot).value = em
                    ws.cell(r, slot).font = Font(color="1a56db", size=9)
                    emails_added += 1
                if new_ems: cos_enriched += 1

    limits = httpx.Limits(max_connections=20, max_keepalive_connections=10)
    async with httpx.AsyncClient(limits=limits) as client:
        for i in range(0, len(to_do), 20):
            batch = to_do[i:i+20]
            await asyncio.gather(*[one(r, n, w, client) for r, n, w in batch],
                                 return_exceptions=True)
            if i % 40 == 0 and i > 0:
                print(f"  Email {i}/{len(to_do)} | +{emails_added} emails")
    print(f"  Email scrape: {cos_enriched} companies got emails, +{emails_added} total")
    return emails_added

# ── Phase 3: Phone from website ───────────────────────────────────────────────
async def scrape_phones_all(companies, limit=350):
    to_do = [(r, name, web) for r, name, web in companies if get_domain(web)][:limit]
    print(f"\n--- Phone scraping: {len(to_do)} companies ---")
    phones_added = 0
    sem = asyncio.Semaphore(20)

    async def one(r, name, web, client):
        nonlocal phones_added
        async with sem:
            try:
                resp = await client.get(web, timeout=7, follow_redirects=True,
                    headers={"User-Agent": "Mozilla/5.0"})
                phones = PHONE_RE.findall(resp.text)
                if phones and not ws.cell(r, C_PHONE).value:
                    ph = phones[0].replace(" ", "").replace("-", "")
                    if not ph.startswith("+91"):
                        ph = "+91" + ph.lstrip("0")
                    ws.cell(r, C_PHONE).value = ph
                    phones_added += 1
            except Exception:
                pass

    limits = httpx.Limits(max_connections=25, max_keepalive_connections=15)
    async with httpx.AsyncClient(limits=limits) as client:
        for i in range(0, len(to_do), 25):
            batch = to_do[i:i+25]
            await asyncio.gather(*[one(r, n, w, client) for r, n, w in batch],
                                 return_exceptions=True)
            if i % 50 == 0 and i > 0:
                print(f"  Phone {i}/{len(to_do)} | +{phones_added} phones")
    print(f"  Phone scrape done: +{phones_added} phones")
    return phones_added

async def main():
    t0 = time.time()

    p1 = await apollo_enrich_all(missing_phone, limit=300)

    p2 = await scrape_emails_all(missing_email, limit=400)

    # Re-collect who still has no phone after Apollo
    still_no_phone = [(r, n, w) for r, n, w in missing_phone
                      if not ws.cell(r, C_PHONE).value]
    p3 = await scrape_phones_all(still_no_phone, limit=350)

    wb.save(MASTER)
    t1 = time.time()

    with_phone = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, C_PHONE).value)
    with_email = sum(1 for r in range(2, ws.max_row+1)
                    if any(ws.cell(r, c).value for c in C_EMAILS))
    print(f"\n{'='*55}")
    print(f"  ENRICHMENT COMPLETE ({t1-t0:.0f}s)")
    print(f"{'='*55}")
    print(f"  Apollo phones      : +{p1}")
    print(f"  Scraped emails     : +{p2}")
    print(f"  Scraped phones     : +{p3}")
    print(f"  With phone total   : {with_phone} / {total_rows}")
    print(f"  With email total   : {with_email} / {total_rows}")
    print(f"  Saved -> {MASTER}")

asyncio.run(main())
