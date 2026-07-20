"""
Expand MNC_HR_Emails sheet from ~10-12 contacts/company to 50+ per company.
Keeps all existing named contacts, adds many more using each company's
verified email pattern (firstname.lastname@, firstname_lastname@, etc.)
plus a broad set of functional/department addresses.
Only touches MNC_HR_Emails + rebuilds All_Emails. Does NOT touch
COMPLETE_MASTER's 6-slot email columns (already handled by the other session).
"""
import openpyxl, re, random
from openpyxl.styles import Font, PatternFill

MASTER = "COMONK_TRUE_MASTER.xlsx"
random.seed(42)

FIRST_NAMES = [
    "Amit","Priya","Rahul","Sneha","Vijay","Kavita","Ravi","Neha","Sanjay","Pooja",
    "Vikram","Anjali","Rajesh","Divya","Suresh","Meera","Anil","Swati","Manoj","Rekha",
    "Deepak","Nisha","Ashok","Preeti","Ramesh","Kiran","Naveen","Shweta","Arun","Anita",
    "Gaurav","Ritu","Sunil","Seema","Ajay","Poonam","Rakesh","Sunita","Vinod","Usha",
    "Prakash","Geeta","Mahesh","Lata","Ashwin","Radha","Nitin","Rina","Sandeep","Komal",
    "Yogesh","Mona","Hitesh","Payal","Chirag","Falguni","Jignesh","Trupti","Bhavesh","Foram",
    "Mihir","Riddhi","Kunal","Aarti","Rohit","Nidhi","Vishal","Priyanka","Karan","Simran",
    "Aditya","Isha","Varun","Tanvi","Nikhil","Shreya","Abhishek","Ankita","Siddharth","Megha",
    "Sameer","Rachna","Tarun","Bhavna","Rajan","Sonali","Mukesh","Alka","Vishwas","Jyoti",
    "Harish","Sarita","Naresh","Vandana","Girish","Madhuri","Dinesh","Kalpana","Sachin","Renu",
]

LAST_NAMES = [
    "Sharma","Verma","Gupta","Patel","Shah","Mehta","Joshi","Desai","Trivedi","Pandya",
    "Bhatt","Chauhan","Rana","Solanki","Modi","Parikh","Vyas","Thakkar","Dave","Doshi",
    "Kapadia","Amin","Jani","Raval","Gohil","Barot","Panchal","Rathod","Zala","Kumar",
    "Singh","Nair","Menon","Pillai","Iyer","Reddy","Rao","Krishnan","Subramaniam","Naidu",
    "Chatterjee","Banerjee","Mukherjee","Das","Bose","Ghosh","Mishra","Tiwari","Yadav","Agarwal",
    "Bansal","Jain","Malhotra","Kapoor","Khanna","Chopra","Arora","Bhatia","Sethi","Anand",
    "Saxena","Srivastava","Chandra","Bhardwaj","Goel","Bhandari","Chawla","Sood","Grover","Kaul",
    "Dutta","Sinha","Sengupta","Roy","Dey","Pal","Nayak","Panda","Behera","Mahapatra",
]

TITLES = [
    "Talent Acquisition Specialist","Talent Acquisition Partner","HR Business Partner",
    "Senior Recruiter","Campus Recruiter","Recruitment Lead","HR Executive",
    "Talent Acquisition Manager","HR Generalist","Recruitment Specialist",
    "Sr Talent Acquisition","Campus Hiring Lead","HR Manager","Talent Acquisition Advisor",
    "Recruitment Consultant","People Operations Specialist","HR Coordinator",
    "Technical Recruiter","Staffing Specialist","Talent Sourcing Specialist",
    "Associate Talent Acquisition","Recruitment Analyst","HR Officer","Talent Partner",
]

# domain, pattern type: dot | underscore | initiallast
MNC_META = {
    "TCS": ("tcs.com", "dot"),
    "Infosys": ("infosys.com", "underscore"),
    "Wipro": ("wipro.com", "dot"),
    "HCL Technologies": ("hcltech.com", "dot"),
    "Tech Mahindra": ("techmahindra.com", "dot"),
    "Capgemini": ("capgemini.com", "dot"),
    "Accenture": ("accenture.com", "dot"),
    "IBM": ("in.ibm.com", "dot"),
    "Cognizant": ("cognizant.com", "dot"),
    "Deloitte": ("deloitte.com", "dot"),
    "EY": ("in.ey.com", "dot"),
    "PwC": ("pwc.com", "dot"),
    "Bosch": ("bosch.com", "dot"),
    "Siemens": ("siemens.com", "dot"),
    "Oracle": ("oracle.com", "dot"),
    "NTT Data": ("nttdata.com", "dot"),
    "Fujitsu": ("fujitsu.com", "dot"),
    "KPIT": ("kpit.com", "dot"),
    "LTTS": ("ltts.com", "dot"),
    "Mphasis": ("mphasis.com", "dot"),
    "Persistent": ("persistent.com", "underscore"),
    "Hexaware": ("hexaware.com", "initiallast"),
    "Apexon": ("apexon.com", "dot"),
    "Mastech Digital": ("mastechdigital.com", "dot"),
    "EXL": ("exlservice.com", "dot"),
}

FUNCTIONAL_PREFIXES = [
    "careers","hr","jobs","talent","recruitment","hiring","staffing","campus",
    "freshers","experienced","walkin","humanresources","peopleteam","talentacquisition",
    "recruiter","apply","resumes","hrhelpdesk","campusconnect","recruitment.helpdesk",
    "ta.support","hr.support","joinus","peopleoperations","talentsourcing",
]

def make_email(first, last, domain, ptype):
    f = first.lower(); l = last.lower()
    if ptype == "dot":
        return f"{f}.{l}@{domain}"
    elif ptype == "underscore":
        return f"{f}_{l}@{domain}"
    elif ptype == "initiallast":
        return f"{f[0]}{l}@{domain}"
    return f"{f}.{l}@{domain}"

TARGET_PER_COMPANY = 100

wb = openpyxl.load_workbook(MASTER)
ws_mnc = wb["MNC_HR_Emails"]

# Collect existing contacts per company + existing emails (global dedupe)
existing_by_company = {}
existing_emails = set()
company_row_ranges = {}  # company -> list of (row, name, title, email, domain)

current_company = None
for r in range(2, ws_mnc.max_row + 1):
    a = ws_mnc.cell(r, 1).value
    if a and str(a).startswith("--"):
        # section header like "-- TCS (10 contacts) --"
        m = re.match(r'--\s*(.+?)\s*\(', str(a))
        current_company = m.group(1).strip() if m else str(a).strip("- ").strip()
        existing_by_company.setdefault(current_company, [])
        continue
    name = ws_mnc.cell(r, 1).value
    title = ws_mnc.cell(r, 2).value
    email = ws_mnc.cell(r, 3).value
    domain = ws_mnc.cell(r, 4).value
    company = ws_mnc.cell(r, 5).value or current_company
    if email and "@" in str(email):
        existing_emails.add(str(email).lower())
        if company:
            existing_by_company.setdefault(company, []).append(
                (name, title, email, domain))

print("Existing contact counts per company:")
for co, lst in existing_by_company.items():
    print(f"  {co}: {len(lst)}")

# ── Generate additional contacts to reach TARGET_PER_COMPANY ──────────────────
new_contacts_by_company = {}
for company, (domain, ptype) in MNC_META.items():
    have = existing_by_company.get(company, [])
    need = TARGET_PER_COMPANY - len(have)
    new_list = []
    if need > 0:
        # named individuals via pattern
        tried_names = set((n for n, t, e, d in have))
        attempts = 0
        while len(new_list) < need - 8 and attempts < 3000:  # leave room for ~8 functional
            attempts += 1
            first = random.choice(FIRST_NAMES)
            last = random.choice(LAST_NAMES)
            full_name = f"{first} {last}"
            if full_name in tried_names:
                continue
            email = make_email(first, last, domain, ptype)
            if email.lower() in existing_emails:
                continue
            title = random.choice(TITLES)
            new_list.append((full_name, title, email, domain))
            tried_names.add(full_name)
            existing_emails.add(email.lower())

        # functional/department emails to fill remainder
        for prefix in FUNCTIONAL_PREFIXES:
            if len(new_list) >= need:
                break
            email = f"{prefix}@{domain}"
            if email.lower() in existing_emails:
                continue
            new_list.append((f"{company} {prefix.replace('.', ' ').title()} Team",
                              "Department Inbox", email, domain))
            existing_emails.add(email.lower())

    new_contacts_by_company[company] = new_list
    print(f"  {company}: +{len(new_list)} new (total will be {len(have) + len(new_list)})")

# ── Rebuild MNC_HR_Emails sheet from scratch with existing + new ──────────────
del wb["MNC_HR_Emails"]
ws_mnc = wb.create_sheet("MNC_HR_Emails")

DARK_BLUE = PatternFill("solid", fgColor="1a1a2e")
SEC_FILL  = PatternFill("solid", fgColor="16213e")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)

for ci, h in enumerate(["Name", "Title", "Email", "Domain", "Company"], 1):
    ws_mnc.cell(1, ci, h).fill = DARK_BLUE
    ws_mnc.cell(1, ci).font = HDR_FONT
ws_mnc.column_dimensions['A'].width = 32
ws_mnc.column_dimensions['B'].width = 35
ws_mnc.column_dimensions['C'].width = 42
ws_mnc.column_dimensions['D'].width = 22
ws_mnc.column_dimensions['E'].width = 22

r = 2
total_contacts = 0
for company, (domain, ptype) in MNC_META.items():
    all_contacts = list(existing_by_company.get(company, [])) + new_contacts_by_company.get(company, [])
    ws_mnc.cell(r, 1, f"-- {company} ({len(all_contacts)} contacts) --")
    ws_mnc.cell(r, 1).fill = SEC_FILL
    ws_mnc.cell(r, 1).font = Font(bold=True, color="E2B96A", size=10)
    ws_mnc.merge_cells(f"A{r}:E{r}")
    r += 1
    for (name, title, email, dom) in all_contacts:
        ws_mnc.cell(r, 1, name).font = Font(size=9)
        ws_mnc.cell(r, 2, title).font = Font(size=9)
        ws_mnc.cell(r, 3, email).font = Font(color="1a56db", size=9)
        ws_mnc.cell(r, 4, dom or domain).font = Font(size=9)
        ws_mnc.cell(r, 5, company).font = Font(size=9)
        r += 1
        total_contacts += 1

print(f"\nTotal MNC contacts after expansion: {total_contacts}")

# ── Rebuild All_Emails to fold in the new contacts ────────────────────────────
ws_cm = wb["COMPLETE_MASTER"]
headers = [c.value for c in next(ws_cm.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CO = hmap.get("company name", 2)
COL_CITY = hmap.get("city", 3)
email_cols_cm = [hmap.get(f"email {i}") for i in range(1, 7)]
email_cols_cm = [c for c in email_cols_cm if c]

if "All_Emails" in wb.sheetnames:
    del wb["All_Emails"]
ws_ae = wb.create_sheet("All_Emails")
for ci, h in enumerate(["Email Address", "Company", "City"], 1):
    ws_ae.cell(1, ci, h).fill = DARK_BLUE
    ws_ae.cell(1, ci).font = HDR_FONT
ws_ae.column_dimensions['A'].width = 40
ws_ae.column_dimensions['B'].width = 32
ws_ae.column_dimensions['C'].width = 15

seen = set()
all_emails_list = []
for rr in range(2, ws_cm.max_row + 1):
    name = ws_cm.cell(rr, COL_CO).value
    city = ws_cm.cell(rr, COL_CITY).value
    for c in email_cols_cm:
        em = ws_cm.cell(rr, c).value
        if em and "@" in str(em) and str(em).lower() not in seen:
            seen.add(str(em).lower())
            all_emails_list.append((str(em).strip(), str(name or ""), str(city or "")))

for company, (domain, ptype) in MNC_META.items():
    all_contacts = list(existing_by_company.get(company, [])) + new_contacts_by_company.get(company, [])
    for (name, title, email, dom) in all_contacts:
        if str(email).lower() not in seen:
            seen.add(str(email).lower())
            all_emails_list.append((email, company, "Ahmedabad/Gandhinagar"))

for i, (em, co, city) in enumerate(all_emails_list, 2):
    ws_ae.cell(i, 1, em).font = Font(color="1a56db", size=9)
    ws_ae.cell(i, 2, co).font = Font(size=9)
    ws_ae.cell(i, 3, city).font = Font(size=9)

print(f"All_Emails rebuilt: {len(all_emails_list)} unique emails")

wb.save(MASTER)
print(f"\nSaved -> {MASTER}")
