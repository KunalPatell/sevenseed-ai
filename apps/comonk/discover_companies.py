"""
PHASE 1: Advanced Company Discovery
Finds ALL AI/ML companies in Ahmedabad from multiple sources:
  1. Adzuna Jobs API (multiple queries)
  2. GNews + NewsAPI (news mentions)
  3. Web scraping: Clutch, GoodFirms, TopDevelopers, DesignRush
  4. GESIA / iNDEXTb / LinkedIn company pages
Output: NEW_COMPANIES_DISCOVERED.csv
"""

import httpx, csv, re, asyncio, json, time
from html.parser import HTMLParser
from collections import defaultdict

# ── API keys ─────────────────────────────────────────────────────────────────
ADZUNA_ID  = "f3b116d8"
ADZUNA_KEY = "fa188c839f828013528b020fffd62ae5"
GNEWS_KEY  = "cf82b9056f4e5e8cebe0f215dae2b7bf"
NEWSAPI_KEY = "6bbbb3c56dc5434694d6f6a7b123dbd2"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# ── Text extractor ────────────────────────────────────────────────────────────
class TextExtractor(HTMLParser):
    def __init__(self): super().__init__(); self.result=[]; self._skip=False
    def handle_starttag(self, t, a):
        if t in ('script','style','noscript'): self._skip=True
    def handle_endtag(self, t):
        if t in ('script','style','noscript'): self._skip=False
    def handle_data(self, d):
        if not self._skip and d.strip(): self.result.append(d.strip())
    def get_text(self): return '\n'.join(self.result)

def extract_text(html):
    p=TextExtractor()
    try: p.feed(html)
    except: pass
    return p.get_text()

def clean_co_name(name):
    name = re.sub(r'\s+', ' ', str(name)).strip()
    # Remove trailing generic words
    name = re.sub(r'\s*(Pvt\.?\s*Ltd\.?|Private\s+Limited|Ltd\.?|Inc\.?|LLC|LLP)\s*$', '', name, flags=re.I)
    return name.strip()

# ── 1. ADZUNA — scrape all AI/ML job companies ───────────────────────────────
def fetch_adzuna():
    queries = [
        "artificial intelligence", "machine learning", "deep learning",
        "NLP engineer", "computer vision", "LLM", "generative AI",
        "data scientist", "data engineer", "MLOps", "AI developer",
        "Python AI", "AI researcher", "neural network", "natural language processing",
    ]
    found = {}
    print("\n  [ADZUNA] Fetching from job listings...")
    for q in queries:
        try:
            for page in [1, 2]:
                r = httpx.get(
                    f"https://api.adzuna.com/v1/api/jobs/in/search/{page}",
                    params={
                        "app_id": ADZUNA_ID, "app_key": ADZUNA_KEY,
                        "results_per_page": "50", "what": q, "where": "ahmedabad",
                    }, timeout=12
                )
                data = r.json()
                for job in data.get("results", []):
                    co = job.get("company", {}).get("display_name", "").strip()
                    loc = job.get("location", {}).get("display_name", "")
                    role = job.get("title", "")
                    if co and len(co) > 2:
                        if co not in found:
                            found[co] = {"roles": set(), "city": "Ahmedabad", "source": "Adzuna"}
                        found[co]["roles"].add(role[:60])
                time.sleep(0.3)
        except Exception as e:
            print(f"    Adzuna error [{q}]: {e}")
    print(f"    Found {len(found)} companies from Adzuna")
    return found

# ── 2. GNews — AI companies in Ahmedabad from news ────────────────────────────
def fetch_gnews():
    queries = [
        "AI company Ahmedabad", "ML startup Ahmedabad", "artificial intelligence Ahmedabad",
        "machine learning company Ahmedabad Gujarat", "tech startup Ahmedabad AI",
        "data science company Ahmedabad",
    ]
    found = {}
    print("\n  [GNEWS] Scanning news articles...")
    for q in queries:
        try:
            r = httpx.get("https://gnews.io/api/v4/search",
                params={"q": q, "lang": "en", "max": "10", "token": GNEWS_KEY},
                timeout=10)
            for art in r.json().get("articles", []):
                title = art.get("title", "") + " " + art.get("description", "")
                # Extract company-like names (capitalized words near Ahmedabad)
                matches = re.findall(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,3})\b', title)
                for m in matches:
                    if len(m) > 4 and m not in {"The","This","With","From","India","Gujarat","Ahmedabad"}:
                        if m not in found:
                            found[m] = {"roles": set(), "city": "Ahmedabad", "source": "GNews"}
        except Exception as e:
            print(f"    GNews error [{q}]: {e}")
    print(f"    Found {len(found)} potential names from GNews")
    return found

# ── 3. Clutch.co ─────────────────────────────────────────────────────────────
async def scrape_clutch(client, sem):
    urls = [
        "https://clutch.co/in/artificial-intelligence-companies/ahmedabad",
        "https://clutch.co/in/it-services/ahmedabad",
        "https://clutch.co/in/machine-learning-companies",
        "https://clutch.co/in/analytics/ahmedabad",
    ]
    found = {}
    for url in urls:
        async with sem:
            try:
                r = await client.get(url, timeout=15)
                text = extract_text(r.text)
                # Clutch company names follow patterns
                names = re.findall(r'(?:^|\n)([A-Z][A-Za-z0-9\s&\.\-]{3,45}(?:Technologies?|Solutions?|Systems?|Labs?|AI|Tech|Soft(?:ware)?|Infotech|Digital|Analytics|Data|Innovations?|Consulting|Services?))\b', text, re.M)
                for n in names:
                    n = clean_co_name(n.strip())
                    if n and len(n) > 4:
                        found[n] = {"roles": set(), "city": "Ahmedabad", "source": "Clutch"}
                # Also grab names from title tags
                raw_titles = re.findall(r'<h3[^>]*>([^<]{4,60})</h3>', r.text)
                for t in raw_titles:
                    t = clean_co_name(re.sub(r'<[^>]+>', '', t).strip())
                    if t and len(t) > 4: found[t] = {"roles": set(), "city": "Ahmedabad", "source": "Clutch"}
                await asyncio.sleep(1)
            except Exception as e:
                print(f"    Clutch error: {e}")
    print(f"    [CLUTCH] Found {len(found)} company names")
    return found

# ── 4. GoodFirms ──────────────────────────────────────────────────────────────
async def scrape_goodfirms(client, sem):
    urls = [
        "https://www.goodfirms.co/artificial-intelligence/india/ahmedabad",
        "https://www.goodfirms.co/machine-learning/india/ahmedabad",
        "https://www.goodfirms.co/data-analytics/india/ahmedabad",
        "https://www.goodfirms.co/software-development/india/ahmedabad",
    ]
    found = {}
    for url in urls:
        async with sem:
            try:
                r = await client.get(url, timeout=15)
                # GoodFirms has company names in specific classes
                names = re.findall(r'"company_name"\s*:\s*"([^"]{3,60})"', r.text)
                if not names:
                    names = re.findall(r'<h2[^>]*class="[^"]*company[^"]*"[^>]*>([^<]{3,60})</h2>', r.text, re.I)
                if not names:
                    # Generic extraction from page text
                    text = extract_text(r.text)
                    names = re.findall(r'\b([A-Z][A-Za-z0-9\s]{3,40}(?:Technologies?|Solutions?|AI|Tech|Infotech|Digital|Analytics|Data|Labs?))\b', text)
                for n in names:
                    n = clean_co_name(n.strip())
                    if n and len(n) > 4:
                        found[n] = {"roles": set(), "city": "Ahmedabad", "source": "GoodFirms"}
                await asyncio.sleep(1)
            except Exception as e:
                print(f"    GoodFirms error: {e}")
    print(f"    [GOODFIRMS] Found {len(found)} company names")
    return found

# ── 5. TopDevelopers.co ───────────────────────────────────────────────────────
async def scrape_topdevelopers(client, sem):
    urls = [
        "https://www.topdevelopers.co/directory/artificial-intelligence-companies-in-ahmedabad",
        "https://www.topdevelopers.co/directory/machine-learning-companies-in-india",
        "https://www.topdevelopers.co/directory/data-science-companies-in-ahmedabad",
    ]
    found = {}
    for url in urls:
        async with sem:
            try:
                r = await client.get(url, timeout=15)
                text = r.text
                # TopDevelopers has company names in h3/h2 tags
                names = re.findall(r'<h[23][^>]*>([^<]{3,60})</h[23]>', text)
                for n in names:
                    n = clean_co_name(re.sub(r'<[^>]+>', '', n).strip())
                    if n and len(n) > 4 and len(n) < 60:
                        found[n] = {"roles": set(), "city": "Ahmedabad", "source": "TopDevelopers"}
                await asyncio.sleep(1)
            except Exception as e:
                print(f"    TopDevelopers error: {e}")
    print(f"    [TOPDEVELOPERS] Found {len(found)} names")
    return found

# ── 6. DesignRush ────────────────────────────────────────────────────────────
async def scrape_designrush(client, sem):
    urls = [
        "https://www.designrush.com/agency/artificial-intelligence-companies/india/ahmedabad",
        "https://www.designrush.com/agency/machine-learning/india",
    ]
    found = {}
    for url in urls:
        async with sem:
            try:
                r = await client.get(url, timeout=15)
                names = re.findall(r'"name"\s*:\s*"([^"]{4,60})"', r.text)
                for n in names:
                    n = clean_co_name(n.strip())
                    if n and len(n) > 4:
                        found[n] = {"roles": set(), "city": "Ahmedabad", "source": "DesignRush"}
                await asyncio.sleep(1)
            except Exception as e:
                print(f"    DesignRush error: {e}")
    print(f"    [DESIGNRUSH] Found {len(found)} names")
    return found

# ── 7. Glassdoor / Indeed style job sites ────────────────────────────────────
async def scrape_job_sites(client, sem):
    urls = [
        "https://www.naukri.com/artificial-intelligence-jobs-in-ahmedabad",
        "https://www.naukri.com/machine-learning-jobs-in-ahmedabad",
        "https://www.naukri.com/data-scientist-jobs-in-ahmedabad",
        "https://internshala.com/internships/machine-learning-internship-in-ahmedabad/",
        "https://www.foundit.in/search/artificial-intelligence-jobs-in-ahmedabad",
    ]
    found = {}
    for url in urls:
        async with sem:
            try:
                r = await client.get(url, timeout=15)
                text = extract_text(r.text)
                # Company patterns from job sites
                names = re.findall(r'(?:at|by|@)\s+([A-Z][A-Za-z0-9\s&\.]{3,45}(?:Technologies?|Solutions?|AI|Tech|Infotech|Digital|Analytics|Data|Labs?|Consulting|Services?))', text)
                for n in names:
                    n = clean_co_name(n.strip())
                    if n and 4 < len(n) < 60:
                        found[n] = {"roles": set(), "city": "Ahmedabad", "source": "JobSite"}
                await asyncio.sleep(1.5)
            except Exception as e:
                print(f"    JobSite error: {e}")
    print(f"    [JOB SITES] Found {len(found)} names")
    return found

# ── 8. GESIA + Gujarat government ────────────────────────────────────────────
async def scrape_gesia(client, sem):
    urls = [
        "https://www.gesia.org/member-list/",
        "https://www.gesia.org/it-companies-list/",
        "https://indextb.com/it-ites-companies/",
        "https://www.startupindia.gov.in/content/sih/en/ams-application/startup.html",
    ]
    found = {}
    for url in urls:
        async with sem:
            try:
                r = await client.get(url, timeout=15)
                text = extract_text(r.text)
                names = re.findall(r'\b([A-Z][A-Za-z0-9\s&\.]{3,45}(?:Technologies?|Solutions?|AI|Tech|Infotech|Digital|Analytics|Data|Labs?|Soft(?:ware)?|Systems?))\b', text)
                for n in names:
                    n = clean_co_name(n.strip())
                    if n and 4 < len(n) < 60:
                        found[n] = {"roles": set(), "city": "Ahmedabad", "source": "GESIA"}
                await asyncio.sleep(1)
            except Exception as e:
                print(f"    GESIA error: {e}")
    print(f"    [GESIA/GOV] Found {len(found)} names")
    return found

# ── 9. Google-style search scraping ─────────────────────────────────────────
async def scrape_search(client, sem):
    search_queries = [
        "top AI ML companies Ahmedabad Gujarat 2024 2025",
        "artificial intelligence startups Ahmedabad list",
        "machine learning companies Ahmedabad India",
        "data science firms Ahmedabad Gujarat",
        "AI tech companies GIFT city Gandhinagar",
        "best AI companies to work Ahmedabad 2025",
    ]
    found = {}
    for q in search_queries:
        async with sem:
            try:
                # DuckDuckGo HTML search (no API key needed)
                r = await client.get(
                    "https://html.duckduckgo.com/html/",
                    params={"q": q},
                    timeout=12
                )
                text = extract_text(r.text)
                # Extract company names from snippets
                names = re.findall(
                    r'\b([A-Z][A-Za-z0-9]{2,}(?:\s+[A-Z][A-Za-z0-9]{2,}){0,3})\s+(?:is\s+(?:an?|the)|Technologies?|Solutions?|AI|Tech|Infotech|Digital|Analytics|Labs?)\b',
                    text
                )
                for n in names:
                    n = clean_co_name(n.strip())
                    if n and 4 < len(n) < 60:
                        found[n] = {"roles": set(), "city": "Ahmedabad", "source": "Search"}
                await asyncio.sleep(2)
            except Exception as e:
                print(f"    Search error: {e}")
    print(f"    [WEB SEARCH] Found {len(found)} names")
    return found

# ── Load existing companies for dedup ────────────────────────────────────────
def load_existing():
    import openpyxl, unicodedata
    def norm(s):
        if not s: return ""
        s = unicodedata.normalize('NFKD', str(s).lower())
        return re.sub(r'[^a-z0-9]', '', s)

    existing = set()
    try:
        wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
        ws = wb["COMPLETE_MASTER"]
        for r in range(2, ws.max_row+1):
            n = ws.cell(r, 2).value
            if n: existing.add(norm(n))
        print(f"  Existing companies in master: {len(existing)}")
    except Exception as e:
        print(f"  Could not load master: {e}")
    return existing

# ── Main ──────────────────────────────────────────────────────────────────────
async def main():
    print("\n" + "="*65)
    print("  PHASE 1: COMPANY DISCOVERY — ALL SOURCES")
    print("="*65)

    # Load existing
    existing = load_existing()

    # Sync sources
    print("\n--- Sync Sources ---")
    adzuna_cos  = fetch_adzuna()
    gnews_cos   = fetch_gnews()

    # Async sources
    print("\n--- Async Web Scraping ---")
    sem = asyncio.Semaphore(5)
    async with httpx.AsyncClient(headers=HEADERS, follow_redirects=True) as client:
        results = await asyncio.gather(
            scrape_clutch(client, sem),
            scrape_goodfirms(client, sem),
            scrape_topdevelopers(client, sem),
            scrape_designrush(client, sem),
            scrape_job_sites(client, sem),
            scrape_gesia(client, sem),
            scrape_search(client, sem),
        )

    clutch_cos, goodfirms_cos, topdev_cos, design_cos, job_cos, gesia_cos, search_cos = results

    # Merge all discovered companies
    import unicodedata
    def norm(s):
        if not s: return ""
        s = unicodedata.normalize('NFKD', str(s).lower())
        return re.sub(r'[^a-z0-9]', '', s)

    all_discovered = {}
    for source_dict in [adzuna_cos, gnews_cos, clutch_cos, goodfirms_cos,
                         topdev_cos, design_cos, job_cos, gesia_cos, search_cos]:
        for name, info in source_dict.items():
            key = norm(name)
            if not key or len(key) < 3: continue
            if key in all_discovered:
                all_discovered[key]["sources"].add(info["source"])
                all_discovered[key]["roles"].update(info.get("roles", set()))
                all_discovered[key]["appearances"] += 1
            else:
                all_discovered[key] = {
                    "name": name, "city": info.get("city","Ahmedabad"),
                    "sources": {info["source"]}, "roles": info.get("roles", set()),
                    "appearances": 1
                }

    print(f"\n  Total discovered (all sources): {len(all_discovered)}")

    # Separate: new vs already known
    new_companies = {k: v for k, v in all_discovered.items() if k not in existing}
    known_companies = {k: v for k, v in all_discovered.items() if k in existing}

    print(f"  Already in master: {len(known_companies)}")
    print(f"  NEW companies found: {len(new_companies)}")

    # Sort by appearances (most confirmed first)
    sorted_new = sorted(new_companies.values(), key=lambda x: -x["appearances"])

    # Write to CSV
    out_file = "NEW_COMPANIES_DISCOVERED.csv"
    with open(out_file, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(["Company", "City", "AI Roles", "Sources", "Confidence",
                    "Website", "Email", "Phone", "Address", "LinkedIn"])
        for co in sorted_new:
            confidence = "HIGH" if co["appearances"] >= 3 else ("MEDIUM" if co["appearances"] >= 2 else "LOW")
            w.writerow([
                co["name"], co.get("city","Ahmedabad"),
                "; ".join(list(co["roles"])[:3]),
                ", ".join(sorted(co["sources"])),
                confidence, "", "", "", "", ""
            ])

    print(f"\n  Saved: {out_file}")
    print(f"  HIGH confidence (3+ sources): {sum(1 for c in sorted_new if c['appearances']>=3)}")
    print(f"  MEDIUM (2 sources): {sum(1 for c in sorted_new if c['appearances']==2)}")
    print(f"  LOW (1 source): {sum(1 for c in sorted_new if c['appearances']==1)}")
    print("\n  TOP NEW COMPANIES (multi-source confirmed):")
    for co in sorted_new[:30]:
        print(f"    [{co['appearances']}x] {co['name']} — {', '.join(sorted(co['sources']))}")

if __name__ == "__main__":
    asyncio.run(main())
