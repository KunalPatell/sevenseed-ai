"""
Cross-reference the 909 gap companies (no email, no phone, no website) in
COMONK_TRUE_MASTER against every other source file in the project directory,
pulling in any website/email/phone data already researched elsewhere.
"""
import openpyxl, csv, re, unicodedata, glob

MASTER = "COMONK_TRUE_MASTER.xlsx"

def norm(s):
    return re.sub(r'[^a-z0-9]', '', unicodedata.normalize('NFKD', str(s).lower()))

EMAIL_RE = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
PHONE_RE = re.compile(r'(?:\+91[\s\-]?|0)?[6-9]\d{9}')
WEB_RE   = re.compile(r'https?://[^\s,"\']+|www\.[^\s,"\']+')

wb = openpyxl.load_workbook(MASTER)
ws = wb["COMPLETE_MASTER"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CO = hmap.get("company name", 2)
COL_WEB = hmap.get("website", 13)
COL_PH  = hmap.get("phone", 12)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]

# Build lookup: company norm -> row (for gap companies only)
gap_rows = {}
for r in range(2, ws.max_row + 1):
    existing_email = any(ws.cell(r, c).value for c in email_cols)
    existing_web   = ws.cell(r, COL_WEB).value
    existing_ph    = ws.cell(r, COL_PH).value
    if not existing_email and not existing_web and not existing_ph:
        name = ws.cell(r, COL_CO).value
        if name:
            gap_rows[norm(name)] = r

print(f"Gap companies to fill: {len(gap_rows)}")

# ── Collect data from all other source files ──────────────────────────────────
found_data = {}  # norm_name -> {'email':.., 'phone':.., 'website':..}

def add_found(name, email=None, phone=None, website=None):
    n = norm(name)
    if n not in gap_rows:
        return
    if n not in found_data:
        found_data[n] = {}
    if email and 'email' not in found_data[n]:
        found_data[n]['email'] = email
    if phone and 'phone' not in found_data[n]:
        found_data[n]['phone'] = phone
    if website and 'website' not in found_data[n]:
        found_data[n]['website'] = website

# 1. Scan other XLSX files
xlsx_files = [
    "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx",
    "Ahmedabad_IT_AIML_Companies_Master.xlsx",
    "Ahmedabad_AIML_Companies_HR_Contacts.xlsx",
    "AI_Engineer_Job_Targets.xlsx",
]

for fn in xlsx_files:
    try:
        w = openpyxl.load_workbook(fn, read_only=True, data_only=True)
    except Exception as e:
        print(f"  Skip {fn}: {e}")
        continue
    for sheet in w.sheetnames:
        s = w[sheet]
        rows_iter = s.iter_rows(values_only=True)
        try:
            hdr_row = next(rows_iter)
        except StopIteration:
            continue
        hdr_map = {str(h).strip().lower(): i for i, h in enumerate(hdr_row) if h}
        # Try to find company/name column and email/phone/website columns
        co_idx = None
        for key in ["company name", "company", "name"]:
            if key in hdr_map:
                co_idx = hdr_map[key]; break
        if co_idx is None:
            continue
        email_idxs = [v for k, v in hdr_map.items() if "email" in k]
        phone_idxs = [v for k, v in hdr_map.items() if "phone" in k or "mobile" in k or "contact" in k]
        web_idxs   = [v for k, v in hdr_map.items() if "website" in k or "url" in k and "career" not in k]

        cnt = 0
        for row in rows_iter:
            if co_idx >= len(row): continue
            name = row[co_idx]
            if not name: continue
            n = norm(name)
            if n not in gap_rows: continue
            email = next((row[i] for i in email_idxs if i < len(row) and row[i] and "@" in str(row[i])), None)
            phone = next((row[i] for i in phone_idxs if i < len(row) and row[i]), None)
            web   = next((row[i] for i in web_idxs if i < len(row) and row[i]), None)
            if email or phone or web:
                add_found(name, email, phone, web)
                cnt += 1
    w.close()
    print(f"  {fn}: cross-matched entries scanned")

# 2. Scan CSV files
csv_files = glob.glob("*.csv")
for fn in csv_files:
    try:
        with open(fn, encoding="utf-8", errors="ignore") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                continue
            hdr_map = {h.strip().lower(): h for h in reader.fieldnames if h}
            co_key = None
            for key in ["company name", "company", "name"]:
                if key in hdr_map:
                    co_key = hdr_map[key]; break
            if not co_key:
                continue
            email_keys = [v for k, v in hdr_map.items() if "email" in k]
            phone_keys = [v for k, v in hdr_map.items() if "phone" in k or "mobile" in k]
            web_keys   = [v for k, v in hdr_map.items() if ("website" in k or "url" in k) and "career" not in k]

            for row in reader:
                name = row.get(co_key)
                if not name: continue
                n = norm(name)
                if n not in gap_rows: continue
                email = next((row[k] for k in email_keys if row.get(k) and "@" in row[k]), None)
                phone = next((row[k] for k in phone_keys if row.get(k)), None)
                web   = next((row[k] for k in web_keys if row.get(k)), None)
                if email or phone or web:
                    add_found(name, email, phone, web)
    except Exception as e:
        print(f"  Skip {fn}: {e}")

print(f"\nTotal gap companies with data found elsewhere: {len(found_data)}")

# ── Apply found data to master ────────────────────────────────────────────────
emails_added = 0; phones_added = 0; webs_added = 0
from openpyxl.styles import Font

for n, data in found_data.items():
    r = gap_rows[n]
    if data.get('email'):
        free_slots = [c for c in email_cols if not ws.cell(r, c).value]
        if free_slots:
            ws.cell(r, free_slots[0]).value = str(data['email']).strip()
            ws.cell(r, free_slots[0]).font = Font(color="1a56db", size=9)
            emails_added += 1
    if data.get('phone') and not ws.cell(r, COL_PH).value:
        ws.cell(r, COL_PH).value = str(data['phone']).strip()
        phones_added += 1
    if data.get('website') and not ws.cell(r, COL_WEB).value:
        ws.cell(r, COL_WEB).value = str(data['website']).strip()
        webs_added += 1

print(f"Applied: +{emails_added} emails, +{phones_added} phones, +{webs_added} websites")

wb.save(MASTER)
print(f"Saved -> {MASTER}")
