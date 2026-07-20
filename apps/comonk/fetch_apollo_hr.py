"""
OPTION 2: Apollo.io HR Email Fetcher
Fetches verified HR / Recruiter / Talent Acquisition emails per company.
Free tier: signup at https://www.apollo.io -> Settings -> Integrations -> API -> get key

Usage:
  set APOLLO_API_KEY=your_key   (or put in .env)
  python fetch_apollo_hr.py
"""

import openpyxl, httpx, time, os, re
from urllib.parse import urlparse

EXCEL = "COMONK_TRUE_MASTER.xlsx"

# Read key from env or .env
APOLLO_KEY = os.environ.get("APOLLO_API_KEY", "")
if not APOLLO_KEY and os.path.exists(".env"):
    for line in open(".env", encoding="utf-8"):
        if line.startswith("APOLLO_API_KEY"):
            APOLLO_KEY = line.split("=", 1)[1].strip()

# HR-focused job titles to search
HR_TITLES = ["HR", "Human Resources", "Recruiter", "Talent Acquisition",
             "HR Manager", "Hiring Manager", "Recruitment", "HRBP",
             "Talent", "People Operations"]

def domain_of(url):
    if not url: return None
    url = str(url).strip()
    if not url.startswith("http"): url = "https://" + url
    d = urlparse(url).netloc.replace("www.", "").strip()
    return d if "." in d else None

def apollo_people_search(domain, page=1):
    """Search Apollo for HR people at a company domain."""
    try:
        r = httpx.post(
            "https://api.apollo.io/api/v1/mixed_people/search",
            headers={"Content-Type": "application/json",
                     "Cache-Control": "no-cache",
                     "X-Api-Key": APOLLO_KEY},
            json={
                "q_organization_domains": domain,
                "person_titles": HR_TITLES,
                "page": page,
                "per_page": 5,
            },
            timeout=15
        )
        if r.status_code != 200:
            return [], f"HTTP {r.status_code}"
        data = r.json()
        people = data.get("people", [])
        results = []
        for p in people:
            email = p.get("email")
            name  = p.get("name", "")
            title = p.get("title", "")
            # Apollo sometimes locks email behind reveal; capture what's available
            if email and "email_not_unlocked" not in str(email):
                results.append((name, title, email))
        return results, None
    except Exception as e:
        return [], str(e)

def main():
    if not APOLLO_KEY:
        print("\n  APOLLO_API_KEY nahi mila!")
        print("  Steps:")
        print("   1. https://www.apollo.io pe free signup karo")
        print("   2. Settings -> Integrations -> API -> Connect -> key copy karo")
        print("   3. .env file mein add karo:  APOLLO_API_KEY=your_key_here")
        print("   4. phir: python fetch_apollo_hr.py")
        return

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["AI_ML_ONLY"]   # AI/ML companies first (priority)

    # Find an empty email column slot per row (cols 6-11 are Email 1-6)
    EMAIL_COLS = [6, 7, 8, 9, 10, 11]
    COL_WEBSITE = 13

    # Quota: process companies that have NO email yet, AI/ML first
    targets = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        web  = ws.cell(r, COL_WEBSITE).value
        has_email = any(ws.cell(r, c).value for c in EMAIL_COLS)
        if name and web and not has_email:
            targets.append((r, name, web))

    print(f"\n  {len(targets)} AI/ML companies bina email — Apollo se fetch kar raha hoon\n")

    fetched = updated = 0
    LIMIT = 50  # safety limit per run (Apollo free credits)

    for r, name, web in targets[:LIMIT]:
        domain = domain_of(web)
        if not domain:
            continue
        print(f"  {name[:38]:<38} ({domain})", end=" ... ", flush=True)
        people, err = apollo_people_search(domain)
        fetched += 1

        if err:
            print(f"ERR {err}")
            if "401" in err or "403" in err:
                print("\n  API key galat ya credits khatam. Stop.")
                break
            time.sleep(1); continue

        if people:
            empty = [c for c in EMAIL_COLS if not ws.cell(r, c).value]
            added_here = 0
            for (pname, ptitle, pemail), col in zip(people, empty):
                ws.cell(r, col).value = pemail
                added_here += 1
            print(f"OK +{added_here}: {', '.join(p[2] for p in people[:added_here])}")
            updated += 1
        else:
            print("no HR found")

        time.sleep(1.0)

    wb.save(EXCEL)
    print(f"\n  Done. Queried {fetched}, updated {updated} companies.")
    print(f"  Saved: {EXCEL}")

if __name__ == "__main__":
    main()
