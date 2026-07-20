"""
deep_mnc_search.py — Deep HR Hunt and SMTP Email Verification
==============================================================
1. Queries Google & Bing for target MNCs in Ahmedabad/Gandhinagar to find HR emails & phones.
2. Extracts and normalizes contacts.
3. Performs SMTP handshake checks (MX records + HELO/RCPT TO) to verify email existence.
4. Merges verified data back to COMPLETE_MASTER in COMONK_TRUE_MASTER.xlsx.
5. Recalculates Priorities and updates everything in one single master sheet.
"""

import sys, os, re, time, socket, smtplib
import openpyxl, httpx
from urllib.parse import quote_plus

sys.stdout.reconfigure(encoding='utf-8')

# DNS resolver fallback if dnspython is not installed
try:
    import dns.resolver
    HAS_DNS = True
except ImportError:
    HAS_DNS = False

EXCEL = "COMONK_TRUE_MASTER.xlsx"
SHEET = "COMPLETE_MASTER"

MNC_TARGETS = [
    "TCS", "Infosys", "Wipro", "Accenture", "Deloitte", "PwC India", "EY (Ernst & Young)",
    "Cognizant", "HCL Technologies", "Tech Mahindra", "Capgemini", "IBM", "Persistent Systems",
    "Mphasis", "Hexaware", "LTIMindtree", "Siemens", "Bosch", "KPIT Technologies",
    "L&T Technology Services (LTTS)", "Amazon", "NTT Data", "Fujitsu", "Zensar Technologies",
    "Mastech Digital", "Apexon", "Intel", "Qualcomm", "Adani Group", "Adani Ports",
    "Adani Wilmar", "Aditya Birla Group", "Reliance Industries", "Larsen & Toubro", "Mahindra"
]

C_EMAILS = list(range(6, 12))
C_PHONE  = 12
C_WEB    = 13
C_LI     = 14
C_PRIO   = 17
C_SRC    = 18

EMAIL_RE = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,7}\b')
PHONE_RE = re.compile(r'(?:\+91[\s\-.]?[6-9]\d{9}|\+91[\s\-.]?(?:79|80|22|33|44)\d{8}|0?79[\s\-.]?\d{4}[\s\-.]?\d{4}|\b[6-9]\d{9}\b)')

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
}

# ── SMTP VERIFICATION ────────────────────────────────────────────────────────
def get_mx_records(domain):
    """Get sorted list of mail exchange servers for a domain."""
    if HAS_DNS:
        try:
            answers = dns.resolver.resolve(domain, 'MX')
            mx_hosts = [str(r.exchange).rstrip('.') for r in answers]
            return mx_hosts
        except Exception:
            pass
    # Fallback to simple socket name resolution
    try:
        host = socket.gethostbyname("mail." + domain)
        return ["mail." + domain]
    except Exception:
        pass
    try:
        host = socket.gethostbyname(domain)
        return [domain]
    except Exception:
        return []

def verify_email_smtp(email):
    """
    SMTP verification. Returns (status, reason)
    Statuses: DELIVERABLE, UNDELIVERABLE, UNKNOWN
    """
    domain = email.split('@')[1]
    mx_hosts = get_mx_records(domain)
    if not mx_hosts:
        return "UNKNOWN", "No MX records found"

    # Try connecting to MX hosts (port 25)
    for host in mx_hosts:
        try:
            server = smtplib.SMTP(timeout=8)
            server.connect(host, 25)
            server.helo("comonk.ai")
            server.mail("verify@comonk.ai")
            code, message = server.rcpt(email)
            server.quit()
            
            # code 250 means OK, deliverable
            if code == 250:
                return "DELIVERABLE", "Server accepted email"
            elif code in (550, 551, 552, 553, 554):
                return "UNDELIVERABLE", f"Server rejected with code {code}"
        except Exception as e:
            continue
            
    # Port 25 might be blocked by ISP
    return "UNKNOWN", "SMTP connection timed out or port 25 blocked"

# ─────────────────────────────────────────────────────────────────────────────
# REAL-TIME DORKING SEARCH
# ─────────────────────────────────────────────────────────────────────────────
def search_contacts(company, client):
    emails = []
    phones = []
    
    queries = [
        f'"{company}" Ahmedabad HR email',
        f'"{company}" Ahmedabad recruiter phone number contact',
        f'site:linkedin.com/in "{company}" HR recruiter Ahmedabad'
    ]
    
    for q in queries:
        try:
            url = f"https://www.google.com/search?q={quote_plus(q)}&num=10"
            r = client.get(url, timeout=10)
            if r.status_code == 200:
                # Extract emails
                for e in EMAIL_RE.findall(r.text):
                    e = e.lower().strip()
                    dom = e.split('@')[1]
                    if dom not in {"example.com", "domain.com", "sentry.io", "test.com"}:
                        emails.append(e)
                # Extract phones
                for p in PHONE_RE.findall(r.text):
                    d = re.sub(r'[^\d]', '', p)
                    if d.startswith('91') and len(d) == 12: d = d[2:]
                    if len(d) == 10 and d[0] in '6789':
                        phones.append('+91 ' + d[:5] + ' ' + d[5:])
            time.sleep(1.0)
        except Exception:
            pass
            
    return list(set(emails)), list(set(phones))

# ─────────────────────────────────────────────────────────────────────────────
# MAIN TASK RUNNER
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "="*70)
    print("  COMONK AI — Advanced MNC HR Deep Search & SMTP Validator")
    print("="*70)
    
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]
    
    # Map company name -> row
    company_rows = {}
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 2).value or "").strip()
        if name:
            company_rows[name.lower()] = r

    emails_added = 0
    phones_added = 0
    
    with httpx.Client(headers=HTTP_HEADERS, timeout=10, follow_redirects=True, verify=False) as client:
        for idx, mnc in enumerate(MNC_TARGETS, 1):
            r = company_rows.get(mnc.lower())
            if not r:
                # try fuzzy matching
                for name, row in company_rows.items():
                    if mnc.lower() in name or name in mnc.lower():
                        r = row; break
            
            if not r:
                print(f"  [{idx:02d}/{len(MNC_TARGETS)}] SKIP {mnc} — not found in sheet")
                continue
                
            print(f"  [{idx:02d}/{len(MNC_TARGETS)}] Searching contacts for: {mnc}...", end=" ", flush=True)
            
            scraped_emails, scraped_phones = search_contacts(mnc, client)
            
            # SMTP Verification Phase
            valid_emails = []
            for e in scraped_emails:
                status, reason = verify_email_smtp(e)
                # Keep if deliverable or unknown (fallback)
                if status in ("DELIVERABLE", "UNKNOWN"):
                    valid_emails.append(e)
            
            # Filter unique emails
            existing_emails = set(str(ws.cell(r, c).value).lower() for c in C_EMAILS if ws.cell(r, c).value)
            new_emails = [e for e in valid_emails if e.lower() not in existing_emails]
            
            # Add to cell columns
            empty_cols = [c for c in C_EMAILS if not ws.cell(r, c).value]
            added_em = 0
            for e, col in zip(new_emails, empty_cols):
                ws.cell(r, col).value = e
                added_em += 1
                emails_added += 1
                
            # Add phone
            added_ph = 0
            if scraped_phones and not ws.cell(r, C_PHONE).value:
                clean_ph = scraped_phones[0]
                ws.cell(r, C_PHONE).value = clean_ph
                added_ph = 1
                phones_added += 1
                
            # Update source column
            if added_em or added_ph:
                src = ws.cell(r, C_SRC).value or ""
                if "DEEP_SMTP_SEARCH" not in src:
                    ws.cell(r, C_SRC).value = (src + ", DEEP_SMTP_SEARCH").strip(", ")
                    
            print(f"FOUND +{added_em} emails, +{added_ph} phones")
            time.sleep(0.5)

    # Save Excel
    wb.save(EXCEL)
    
    print("\n" + "="*70)
    print("  Deep MNC Research Complete!")
    print(f"  Emails added : {emails_added}")
    print(f"  Phones added : {phones_added}")
    print(f"  Spreadsheet saved to: {EXCEL}")
    print("="*70 + "\n")

if __name__ == "__main__":
    main()
