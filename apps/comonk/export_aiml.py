# Export ready-to-apply job lists from the master file:
#   1. AIML_Companies_Apply_List.csv          (all AI/ML companies)
#   2. Gandhinagar_GIFTCity_Companies.csv      (Gandhinagar / GIFT City subset)
#   3. AI_Engineer_Job_Targets.xlsx            (formatted, priority-sorted workbook)
import openpyxl, csv
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.load_workbook('Ahmedabad_IT_AIML_FINAL_MASTER.xlsx')
ws = wb['All Companies']

aiml, gnagar = [], []
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[1]:
        continue
    no, name, cat, roles, e1, e2, e3, e4, e5, phone, website, linkedin, address = row
    emails = [e for e in (e1, e2, e3, e4, e5) if e]
    rec = {
        'Company': name, 'Category': cat, 'Roles': roles or '',
        'Email': emails[0] if emails else '', 'All Emails': '; '.join(emails),
        'Phone': phone or '', 'Website': website or '',
        'LinkedIn': linkedin or '', 'Address': address or '',
    }
    if cat == 'AI / ML':
        aiml.append(rec)
    addr_l = (address or '').lower()
    if 'gandhinagar' in addr_l or 'gift city' in addr_l or 'infocity' in addr_l:
        if 'sarkhej' not in addr_l:   # exclude Sarkhej-Gandhinagar Hwy (that's Ahmedabad)
            gnagar.append(rec)

cols = ['Company', 'Category', 'Roles', 'Email', 'All Emails', 'Phone', 'Website', 'LinkedIn', 'Address']

def priority(r):
    if r['Email'] and r['Phone']: return (0, '1 - Apply now (email+phone)')
    if r['Email']:                return (1, '2 - Apply now (email)')
    if r['Phone']:                return (2, '3 - Call / website')
    return (3, '4 - Needs research')

def dump_csv(fname, rows):
    rows = sorted(rows, key=lambda r: r['Company'].lower())
    with open(fname, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader(); w.writerows(rows)
    return len(rows)

n_ai = dump_csv('AIML_Companies_Apply_List.csv', aiml)
n_gn = dump_csv('Gandhinagar_GIFTCity_Companies.csv', gnagar)

# ---- formatted workbook ----
out = openpyxl.Workbook()
hdr_fill = PatternFill('solid', fgColor='1F4E78')
hdr_font = Font(bold=True, color='FFFFFF', size=11)
p_fill = {0: PatternFill('solid', fgColor='C6EFCE'), 1: PatternFill('solid', fgColor='E2EFDA'),
          2: PatternFill('solid', fgColor='FFF2CC'), 3: PatternFill('solid', fgColor='F2F2F2')}
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
xcols = ['Priority', 'Company', 'Roles', 'Email', 'Phone', 'Website', 'LinkedIn', 'Address']
widths = [26, 26, 40, 32, 18, 30, 38, 50]

def make_sheet(title, rows):
    sh = out.create_sheet(title)
    sh.append(xcols)
    for c, w in zip(sh[1], widths):
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(vertical='center')
    for col, w in zip('ABCDEFGH', widths):
        sh.column_dimensions[col].width = w
    rows = sorted(rows, key=lambda r: (priority(r)[0], r['Company'].lower()))
    for r in rows:
        pr_rank, pr_label = priority(r)
        sh.append([pr_label, r['Company'], r['Roles'], r['Email'], r['Phone'],
                   r['Website'], r['LinkedIn'], r['Address']])
        for cell in sh[sh.max_row]:
            cell.fill = p_fill[pr_rank]; cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    sh.freeze_panes = 'A2'
    sh.auto_filter.ref = f'A1:H{sh.max_row}'

out.remove(out.active)
make_sheet('AI-ML Job Targets', aiml)
make_sheet('Gandhinagar-GIFT City', gnagar)
out.save('AI_Engineer_Job_Targets.xlsx')

print(f'AI/ML companies        : {n_ai}  -> AIML_Companies_Apply_List.csv')
print(f'Gandhinagar/GIFT City  : {n_gn}  -> Gandhinagar_GIFTCity_Companies.csv')
print(f'Formatted workbook     : AI_Engineer_Job_Targets.xlsx')
ready = sum(1 for r in aiml if r['Email'])
print(f'AI/ML with a direct email (apply now): {ready}/{n_ai}')
