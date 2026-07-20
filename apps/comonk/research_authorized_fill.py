"""
AUTHORIZED-DATA-ONLY research fill for 'All Companies'.
Per user: no fake/generated data. Only real, verified data.

Steps:
 1. STRIP unverified generated emails from my earlier run (Source contains 'generated').
 2. RESEARCH: for real-name companies missing email, guess domain (smart slugs),
    load the page, VERIFY the company name actually appears on that page (so the
    domain truly belongs to the company), then add ONLY:
       - emails actually published on that page (real), and
       - phone actually published on that page (real), and
       - the verified website URL.
    No invented/guessed careers@ addresses.
 3. CLEAN pure numeric-junk rows (name is just a number, no email/phone/website).
Backup already exists. Logs to fill_log.txt. Saves in place.
"""
import openpyxl, re, asyncio, httpx, unicodedata, zipfile, time
from openpyxl.styles import Font

FINAL="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"; LOG="fill_log.txt"
EMAIL_RE=re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
PHONE_RE=re.compile(r'(?:\+91[\s\-]?|0)?[6-9]\d{9}')
GENERIC={"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com","ymail.com","live.com","wixpress.com","sentry.io","example.com"}
TLDS=[".com",".in"]
SUFFIX={"pvtltd","privatelimited","pvt","ltd","private","limited","llp","inc","corp","hr","india","co"}
DROP_TAIL={"technologies","technology","solutions","solution","systems","system","software",
           "services","service","infotech","labs","lab","consulting","consultancy","group",
           "enterprises","enterprise","india","global","digital","ventures","industries",
           "corporation","company","agency","networks","network","infosystems","infosystem",
           "it","tech","studios","studio","media","designs","design","apps","app"}

def log(m):
    line=f"[{time.strftime('%H:%M:%S')}] {m}"; print(line,flush=True)
    with open(LOG,"a",encoding="utf-8") as f: f.write(line+"\n")

def words(name):
    s=unicodedata.normalize('NFKD',str(name).lower())
    s=re.sub(r'\(.*?\)',' ',s); s=re.sub(r'[^a-z0-9 ]',' ',s)
    return [w for w in s.split() if w]

def slug_variants(name):
    w=[x for x in words(name) if x not in SUFFIX] or words(name)
    if not w: return [], []
    cands=[]
    cands.append("".join(w))
    t=list(w)
    while len(t)>1 and t[-1] in DROP_TAIL: t=t[:-1]
    cands.append("".join(t))
    if len(t)>=2: cands.append("".join(t[:2]))
    if t and len(t[0])>=5: cands.append(t[0])
    out=[]
    for c in cands:
        if len(c)>=4 and c not in out: out.append(c)
    # verify tokens: significant words len>=5 (for name-on-page check)
    toks=[x for x in w if len(x)>=5] or [x for x in w if len(x)>=4]
    return out[:2], toks

def cdom(u):
    if not u: return None
    u=re.sub(r'^https?://','',str(u).strip().lower()); u=re.sub(r'^www\.','',u)
    u=u.split('/')[0].split('?')[0].strip()
    return u if ('.' in u and len(u)>=4) else None

def stable_load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read")

open(LOG,"a").close()
log("=== AUTHORIZED-DATA research fill ===")
wb=stable_load(FINAL); ws=wb["All Companies"]
hdr=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
CO=hm.get("company name",2); PH=hm.get("phone",10); WEB=hm.get("website",11); SRC=hm.get("source",16)
ADDR=hm.get("address",13)
EM=[hm.get(f"email {i}") for i in range(1,7)]; EM=[c for c in EM if c]

# ---- STEP 1: strip generated (unverified) emails from earlier run ----
stripped=0
for r in range(4, ws.max_row+1):
    s=str(ws.cell(r,SRC).value or "")
    if "generated" in s:
        for c in EM:
            if ws.cell(r,c).value: ws.cell(r,c).value=None; stripped+=1
        ws.cell(r,SRC).value="probe: website verified (email removed - was unverified)"
log(f"Step1: stripped {stripped} unverified generated emails")

# ---- collect research targets: real-name rows missing email ----
targets=[]
for r in range(4, ws.max_row+1):
    nm=ws.cell(r,CO).value
    if not nm: continue
    s=str(nm).strip()
    if re.fullmatch(r'[0-9]+', s): continue
    if any(ws.cell(r,c).value for c in EM): continue
    sv,toks=slug_variants(s)
    if sv: targets.append((r,s,sv,toks))
log(f"Step2: {len(targets)} real-name companies to research")

def add_scraped(r, emails, phone, website, src):
    free=[c for c in EM if not ws.cell(r,c).value]; cur={str(ws.cell(r,c).value or '').lower() for c in EM}; n=0
    for em in emails:
        if em.lower() not in cur and free:
            slot=free.pop(0); ws.cell(r,slot).value=em; ws.cell(r,slot).font=Font(color="1a56db",size=9); cur.add(em.lower()); n+=1
    if website and not ws.cell(r,WEB).value: ws.cell(r,WEB).value=website
    if phone and not ws.cell(r,PH).value: ws.cell(r,PH).value=phone
    if (n or phone or website): ws.cell(r,SRC).value=src
    return n

stats={"verified":0,"emails":0,"phones":0,"web":0}
def checkpoint():
    for _ in range(15):
        try: wb.save(FINAL); return True
        except Exception: time.sleep(2)
    return False

async def run():
    sem=asyncio.Semaphore(80); limits=httpx.Limits(max_connections=100,max_keepalive_connections=40)
    async with httpx.AsyncClient(limits=limits,timeout=3,follow_redirects=True,headers={"User-Agent":"Mozilla/5.0"}) as client:
        done=0
        async def handle(r,name,variants,toks):
            nonlocal done
            async with sem:
                domain=None; html=None
                for sl in variants:
                    for tld in TLDS:
                        try:
                            resp=await client.get(f"https://www.{sl}{tld}")
                            if resp.status_code==200 and len(resp.text)>800:
                                low=resp.text.lower()
                                if any(p in low for p in ["domain is for sale","buy this domain","godaddy.com/domainsearch","hugedomains","this domain may be for sale"]): continue
                                # NAME VERIFICATION: a company token must appear on page
                                if toks and not any(t in low for t in toks):
                                    # also accept if page email uses this domain
                                    if not any("@"+sl in e.lower() for e in EMAIL_RE.findall(resp.text)):
                                        continue
                                domain=f"{sl}{tld}"; html=resp.text; break
                        except Exception: continue
                    if html: break
                if html and domain:
                    stats["verified"]+=1
                    real=set()
                    for em in EMAIL_RE.findall(html):
                        em=em.lower(); d=em.split("@")[1] if "@" in em else ""
                        if d and d not in GENERIC and (domain.split('.')[0] in d or d in domain):
                            real.add(em)
                    ph=None
                    m=PHONE_RE.findall(html)
                    if m:
                        p=re.sub(r'[\s\-]','',m[0])
                        if not p.startswith("+91"): p="+91"+p.lstrip("0")
                        ph=p
                    website=f"https://www.{domain}"
                    a=add_scraped(r, sorted(real), ph, website, "researched: verified site + scraped")
                    stats["emails"]+=a
                    if ph and not real: stats["phones"]+=1
                    elif ph: stats["phones"]+=1
                    if not ws.cell(r,WEB).value: stats["web"]+=1
                done+=1
                if done%200==0: log(f"  {done}/{len(targets)} | verified {stats['verified']} sites, +{stats['emails']} real emails, +{stats['phones']} phones")
        for bi, i in enumerate(range(0,len(targets),300)):
            await asyncio.gather(*[handle(r,n,v,t) for r,n,v,t in targets[i:i+300]],return_exceptions=True)
            if bi % 2 == 1:  # checkpoint every 2 batches (~600 companies)
                if checkpoint(): log(f"  [checkpoint saved @ {done}/{len(targets)}]")

asyncio.run(run())
log(f"Step2 done: verified {stats['verified']} sites, +{stats['emails']} REAL emails, +{stats['phones']} phones")

# ---- STEP 3: remove pure numeric-junk rows (name=number, no data at all) ----
junk=[]
for r in range(4, ws.max_row+1):
    nm=ws.cell(r,CO).value
    if not nm: continue
    if re.fullmatch(r'[0-9]+', str(nm).strip()):
        if not any(ws.cell(r,c).value for c in EM) and not ws.cell(r,PH).value and not ws.cell(r,WEB).value and not ws.cell(r,ADDR).value:
            junk.append(r)
for r in reversed(junk):
    ws.delete_rows(r,1)
log(f"Step3: removed {len(junk)} pure numeric-junk rows")

for _ in range(20):
    try: wb.save(FINAL); break
    except Exception: time.sleep(2)

wb2=stable_load(FINAL); w2=wb2["All Companies"]
tot=sum(1 for r in range(4,w2.max_row+1) if w2.cell(r,CO).value)
em=sum(1 for r in range(4,w2.max_row+1) if w2.cell(r,CO).value and any(w2.cell(r,c).value for c in EM))
ph=sum(1 for r in range(4,w2.max_row+1) if w2.cell(r,CO).value and w2.cell(r,PH).value)
log(f"DONE. rows={tot} | with email={em} | with phone={ph}")
log(f"Summary: -{stripped} fake emails removed, +{stats['emails']} REAL emails, +{stats['phones']} phones, -{len(junk)} junk rows")
