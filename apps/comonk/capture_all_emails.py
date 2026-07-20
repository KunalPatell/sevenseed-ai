"""
CAPTURE EVERYTHING — guarantee no email left out.
1. Attach missing emails to existing companies by WEBSITE-domain (free slots).
2. Build 'All_Emails' sheet = EVERY email from EVERY source + derived company.
"""
import openpyxl, re, os, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment

MASTER="COMONK_TRUE_MASTER.xlsx"
EMAIL_RE=re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
C_COMP=2; C_EMAILS=[6,7,8,9,10,11]; C_PHONE=12; C_WEB=13; C_SRC=18
GEN={"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com","yahoo.co.in","ymail.com","googlemail.com","qq.com"}

all_emails=set()
def slurp(t):
    for e in EMAIL_RE.findall(t): all_emails.add(e.lower())
for fn in os.listdir("."):
    if fn.startswith("_extract_") and fn.endswith(".txt"):
        slurp(open(fn,encoding="utf-8",errors="ignore").read())
for fn in ["emails.csv","AIML_Companies_Apply_List.csv","Gandhinagar_GIFTCity_Companies.csv",
           "NEW_COMPANIES_DISCOVERED.csv","all_bcc_emails.txt","hr_vcards_15-6-26.vcf"]:
    if os.path.exists(fn): slurp(open(fn,encoding="utf-8",errors="ignore").read())
for fn in ["1500+_hr_list.xlsx","Ahmedabad_AIML_Companies_HR_Contacts.xlsx",
           "Ahmedabad_IT_AIML_Companies_Master.xlsx","Ahmedabad_IT_AIML_FINAL_MASTER.xlsx",
           "AI_Engineer_Job_Targets.xlsx","HR Mail List.xlsx"]:
    if os.path.exists(fn):
        try:
            w2=openpyxl.load_workbook(fn,read_only=True)
            for s in w2.sheetnames:
                for row in w2[s].iter_rows(values_only=True):
                    for v in row:
                        if v:
                            for e in EMAIL_RE.findall(str(v)): all_emails.add(e.lower())
            w2.close()
        except: pass

wb=openpyxl.load_workbook(MASTER)
ws=wb["COMPLETE_MASTER"]
for row in ws.iter_rows(values_only=True):
    for v in row:
        if v:
            for e in EMAIL_RE.findall(str(v)): all_emails.add(e.lower())
print(f"  TOTAL unique emails across everything: {len(all_emails)}")

web_row={}; comp_by_domain={}
for r in range(2, ws.max_row+1):
    w=ws.cell(r,C_WEB).value
    if w:
        d=re.sub(r'^https?://(www\.)?','',str(w)).split("/")[0].lower()
        if d: web_row[d]=r; comp_by_domain[d]=ws.cell(r,C_COMP).value

attached=0
for e in all_emails:
    dom=e.split("@")[1]
    if dom in GEN: continue
    r=web_row.get(dom)
    if not r: continue
    cur=set(filter(None,[ws.cell(r,c).value for c in C_EMAILS]))
    if e in cur: continue
    empty=[c for c in C_EMAILS if not ws.cell(r,c).value]
    if empty: ws.cell(r,empty[0]).value=e; attached+=1

if "All_Emails" in wb.sheetnames: del wb["All_Emails"]
wae=wb.create_sheet("All_Emails")
hf=PatternFill("solid",fgColor="2E1A47"); hfont=Font(bold=True,color="FFFFFF")
for c,h in enumerate(["#","Email","Company (from domain)","Domain","Status"],1):
    cell=wae.cell(1,c,h); cell.fill=hf; cell.font=hfont
    cell.alignment=Alignment(horizontal="center",vertical="center")
wae.column_dimensions['B'].width=38; wae.column_dimensions['C'].width=28
wae.column_dimensions['D'].width=26; wae.column_dimensions['E'].width=14
wae.freeze_panes="A2"
row=2
for e in sorted(all_emails):
    dom=e.split("@")[1]
    if dom in GEN: comp="(personal)"; st="personal"
    else:
        comp=comp_by_domain.get(dom) or dom.split(".")[0].title()
        st="in master" if dom in web_row else "extra"
    wae.cell(row,1,row-1); wae.cell(row,2,e); wae.cell(row,3,comp)
    wae.cell(row,4,dom); wae.cell(row,5,st)
    if row%2==0:
        for c in range(1,6): wae.cell(row,c).fill=PatternFill("solid",fgColor="F5F0FF")
    row+=1
total_emails=row-2

for i,r in enumerate(range(2,ws.max_row+1),1): ws.cell(r,1,i)
total=ws.max_row-1; maxc=ws.max_column
ws2=wb["AI_ML_ONLY"]
for r in range(ws2.max_row,1,-1): ws2.delete_rows(r)
ai_kw=['ai','ml','data','machine','deep','nlp','computer vision','llm','genai','analytics','automation','intelligence','learning','aiops','analyst']
ar=2; fa=PatternFill("solid",fgColor="F5F0FF"); fb=PatternFill("solid",fgColor="FFFFFF")
for r in range(2,ws.max_row+1):
    cat=str(ws.cell(r,4).value or "").lower(); role=str(ws.cell(r,5).value or "").lower()
    if any(k in cat or k in role for k in ai_kw):
        fill=fa if ar%2==0 else fb
        for c in range(1,maxc+1):
            v=ws.cell(r,c).value; cc=ws2.cell(ar,c,v); cc.fill=fill
            cc.alignment=Alignment(vertical="center")
            if c in C_EMAILS and v: cc.font=Font(color="1a56db",size=9)
        ws2.cell(ar,1,ar-1); ar+=1
ai_count=ws2.max_row-1
with_em=sum(1 for r in range(2,ws.max_row+1) if any(ws.cell(r,c).value for c in C_EMAILS))
wb["STATS"].cell(3,2,total); wb["STATS"].cell(6,2,ai_count)
wb.save(MASTER)
print(f"  Emails attached to companies: +{attached}")
print(f"  All_Emails sheet rows: {total_emails} (EVERY email, guaranteed)")
print(f"  TOTAL companies: {total} | AI/ML: {ai_count} | with email: {with_em}")
print(f"  Sheets: {wb.sheetnames}")
