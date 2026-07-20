import openpyxl

wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
ws = wb["COMPLETE_MASTER"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CO = hmap.get("company name", 2)
COL_WEB = hmap.get("website", 13)
COL_PH = hmap.get("phone", 12)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]

gaps = []
for r in range(2, ws.max_row + 1):
    has_email = any(ws.cell(r, c).value for c in email_cols)
    has_web = ws.cell(r, COL_WEB).value
    has_ph = ws.cell(r, COL_PH).value
    if not has_email and not has_web and not has_ph:
        name = ws.cell(r, COL_CO).value
        if name:
            gaps.append((r, str(name)))

print(f"Remaining full-gap companies: {len(gaps)}")
for r, n in gaps[:15]:
    print(f"  row={r} {n}")
