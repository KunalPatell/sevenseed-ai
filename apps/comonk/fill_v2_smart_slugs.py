"""
Improved fill for 'All Companies' — smarter domain guessing.
Strips corporate suffixes (Pvt Ltd, Technologies, Solutions, India, HR...) and
tries multiple slug variants so real companies (e.g. 'AddWeb Solution Pvt Ltd'
-> addwebsolution.com) are found. Skips numeric-junk names.
Scrapes real emails+phone from live pages; fallback careers@/hr@/info@.
"""
import openpyxl, re, asyncio, httpx, unicodedata, zipfile, time
from openpyxl.styles import Font

FINAL="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"; LOG="fill_log.txt"
EMAIL_RE=re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
PHONE_RE=re.compile(r'(?:\+91[\s\-]?|0)?[6-9]\d{9}')
GENERIC={"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com","ymail.com","live.com"}
TLDS=[".com",".in",".co.in"]
SUFFIX=("pvtltd","privatelimited","pvt","ltd","private","limited","llp","inc","corp",
        "hr","india","indianpvtltd")
DROP_TAIL=("technologies","technology","solutions","solution","systems","system","software",
           "services","service","infotech","labs","lab","consulting","consultancy","group",
           "enterprises","enterprise","india","global","digital","ventures","industries",
           "corporation","company","pvtltd","hr","agency","networks","network","infosystems",
           "infosystem","it","tech")

def log(m):
    line=f"[{time.strftime('%H:%M:%S')}] {m}"; print(line,flush=True)
    with open(LOG,"a",encoding="utf-8") as f: f.write(line+"\n")

def words(name):
    s=unicodedata.normalize('NFKD',str(name).lower())
    s=re.sub(r'\(.*?\)',' ',s); s=re.sub(r'[^a-z0-9 ]',' ',s)
    return [w for w in s.split() if w]

def slug_variants(name):
    w=words(name)
    if not w: return []
    # remove pure suffix tokens
    w2=[x for x in w if x not in SUFFIX]
    if not w2: w2=w
    cands=[]
    full="".join(w2)
    cands.append(full)
    # drop trailing generic descriptor words
    trimmed=list(w2)
    while len(trimmed)>1 and trimmed[-1] in DROP_TAIL:
        trimmed=trimmed[:-1]
    cands.append("".join(trimmed))
    # first two significant words
    if len(trimmed)>=2: cands.append("".join(trimmed[:2]))
    # first word if long enough
    if trimmed and len(trimmed[0])>=5: cands.append(trimmed[0])
    # dedupe, keep len>=4
    out=[]
    for c in cands:
        if len(c)>=4 and c not in out: out.append(c)
    return out[:3]

def cdom(u):
    if not u: return None
    u=re.sub(r'^https?://','',str(u).strip().lower()); u=re.sub(r'^www\.','',u)
    u=u.split('/')[0].split('?')[0].strip()
    return u if ('.' in u and len(u)>=4) else None

def stable_load(p,t=30):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read")

log("=== fill_v2 smart slugs ===")
wb=stable_load(FINAL); ws=wb["All Companies"]
hdr=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
CO=hm.get("company name",2); PH=hm.get("phone",10); WEB=hm.get("website",11); SRC=hm.get("source",16)
EM=[hm.get(f"email {i}") for i in range(1,7)]; EM=[c for c in EM if c]

targets=[]
for r in range(4, ws.max_row+1):
    nm=ws.cell(r,CO).value
    if not nm: continue
    s=str(nm).strip()
    if re.fullmatch(r'[0-9]+', s): continue      # skip numeric junk
    if any(ws.cell(r,c).value for c in EM): continue  # already has email
    if cdom(ws.cell(r,WEB).value): continue      # already has website (handled)
    sv=slug_variants(s)
    if sv: targets.append((r, s, sv))
log(f"{len(targets)} real-name companies to probe (numeric junk skipped)")

def set_emails(r, emails, src):
    free=[c for c in EM if not ws.cell(r,c).value]; cur={str(ws.cell(r,c).value or '').lower() for c in EM}; n=0
    for em in emails:
        if em.lower() not in cur and free:
            slot=free.pop(0); ws.cell(r,slot).value=em; ws.cell(r,slot).font=Font(color="1a56db",size=9); cur.add(em.lower()); n+=1
    if n and not ws.cell(r,SRC).value: ws.cell(r,SRC).value=src
    return n

stats={"web":0,"emails":0,"real":0,"phones":0}
async def run():
    sem=asyncio.Semaphore(60); limits=httpx.Limits(max_connections=70,max_keepalive_connections=30)
    async with httpx.AsyncClient(limits=limits,timeout=4,follow_redirects=True,headers={"User-Agent":"Mozilla/5.0"}) as client:
        done=0
        async def handle(r,name,variants):
            nonlocal done
            async with sem:
                domain=None; html=None
                for sl in variants:
                    for tld in TLDS[:2]:  # .com, .in (fast)
                        try:
                            resp=await client.get(f"https://www.{sl}{tld}")
                            if resp.status_code==200 and len(resp.text)>800:
                                tl=resp.text.lower()
                                if any(p in tl for p in ["domain is for sale","buy this domain","godaddy.com/domainsearch","hugedomains"]): continue
                                domain=f"{sl}{tld}"; html=resp.text; break
                        except Exception: continue
                    if html: break
                if html and domain:
                    if not ws.cell(r,WEB).value: ws.cell(r,WEB).value=f"https://www.{domain}"; stats["web"]+=1
                    real=set()
                    for em in EMAIL_RE.findall(html):
                        em=em.lower(); d=em.split("@")[1] if "@" in em else ""
                        if d and d not in GENERIC and domain.split('.')[0] in d: real.add(em)
                    if real:
                        a=set_emails(r,sorted(real),"probe:scraped-real"); stats["emails"]+=a; stats["real"]+=a
                    else:
                        stats["emails"]+=set_emails(r,[f"careers@{domain}",f"hr@{domain}",f"info@{domain}"],"probe:generated")
                    if not ws.cell(r,PH).value:
                        ph=PHONE_RE.findall(html)
                        if ph:
                            p=re.sub(r'[\s\-]','',ph[0])
                            if not p.startswith("+91"): p="+91"+p.lstrip("0")
                            ws.cell(r,PH).value=p; stats["phones"]+=1
                done+=1
                if done%200==0: log(f"  {done}/{len(targets)} | +{stats['web']} web, +{stats['emails']} em ({stats['real']} real), +{stats['phones']} ph")
        for i in range(0,len(targets),300):
            await asyncio.gather(*[handle(r,n,v) for r,n,v in targets[i:i+300]],return_exceptions=True)

asyncio.run(run())
log(f"Saving... +{stats['web']} web, +{stats['emails']} emails ({stats['real']} real), +{stats['phones']} phones")
for _ in range(20):
    try: wb.save(FINAL); break
    except Exception: time.sleep(2)
wb2=stable_load(FINAL); w2=wb2["All Companies"]
em=sum(1 for r in range(4,w2.max_row+1) if w2.cell(r,CO).value and any(w2.cell(r,c).value for c in EM))
ph=sum(1 for r in range(4,w2.max_row+1) if w2.cell(r,CO).value and w2.cell(r,PH).value)
log(f"DONE. with email={em} | with phone={ph}")
