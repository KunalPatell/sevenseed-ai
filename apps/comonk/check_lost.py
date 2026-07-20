import openpyxl, zipfile, time
def load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read "+p)

def emails_of(wb, S, targets):
    ws=wb[S]
    hrow=3
    hdr=[ws.cell(hrow,c).value for c in range(1,ws.max_column+1)]
    hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
    CO=hm.get("company name",2)
    EM=[hm.get(f"email {i}") for i in range(1,7)]; EM=[c for c in EM if c]
    res={}
    for row in ws.iter_rows(min_row=hrow+1, values_only=True):
        def g(c): return row[c-1] if c and len(row)>=c else None
        co=str(g(CO) or "").strip()
        if co in targets:
            res.setdefault(co,[]).append([str(g(c)).strip() for c in EM if g(c) and "@" in str(g(c))])
    return res

targets={"Adani Enterprises","Axis Bank Ahmedabad","Cadila Pharmaceuticals","Blackrock India","Bayer India"}
cur=load("Ahmedabad_IT_AIML_FINAL_MASTER.xlsx")
bak=load("Ahmedabad_IT_AIML_FINAL_MASTER_SCRAPED_1653.xlsx")
c=emails_of(cur,"All Companies",targets)
b=emails_of(bak,"All Companies",targets)
for t in sorted(targets):
    cn=c.get(t,[[]]); bn=b.get(t,[[]])
    print(f"### {t}")
    print(f"  BACKUP (scraper) emails: {bn}")
    print(f"  CURRENT emails         : {cn}")
    print()
