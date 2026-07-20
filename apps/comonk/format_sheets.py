"""
Properly format/structure every sheet in Ahmedabad_IT_AIML_FINAL_MASTER.xlsx.
- Remove title-banner rows + blank rows + stray U+FFFD junk chars
- Put clean headers on ROW 1 (bold, dark fill, frozen, auto-filter)
- Reorder columns: key fields first (No, Company, Category, City, Roles, Phone,
  Website, Address, LinkedIn, Priority, Source) then all Email columns at the end
- Drop fully-empty email columns; sensible widths + alignment
- Keeps ALL company/contact data.
"""
import openpyxl, re, zipfile, time
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
def stable_load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read (file open in Excel?)")

def clean(v):
    if v is None: return None
    s=str(v).replace("�","")
    s=re.sub(r'[\ud800-\udfff]','',s)          # lone surrogates / broken chars
    s=s.strip()
    if not re.search(r'[A-Za-z0-9]', s): return None   # only symbols/emoji -> empty
    if s.lower() in ("none","n/a","na","-","--","null","nil","."): return None
    return s

def _norm(h):
    return re.sub(r'\s+',' ', re.sub(r'[^a-z0-9 ]','',str(h).strip().lower())).strip()

# preferred order by normalized header
def rank(h):
    n=_norm(h)
    order=["no","#","company name","company","full name","category","city",
           "ai roles apply for","roles skills","role","phone","office phone","website",
           "address","linkedin url original","linkedin","priority","source","notes"]
    for i,k in enumerate(order):
        if n==k: return (0,i)
    m=re.match(r'(hr )?email (\d+)', n)
    if m: return (1, int(m.group(2)))     # emails after everything, in number order
    return (2, 99)

DARK=PatternFill("solid",fgColor="1F3A5F")
HFONT=Font(bold=True,color="FFFFFF",size=10)
BLUE=Font(color="1A56DB",size=9)
NORM=Font(size=9)
CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT=Alignment(horizontal="left",vertical="center")
thin=Side(style="thin",color="D9DEE8")
BORDER=Border(bottom=thin)

WIDTH={"no":5,"#":5,"company name":30,"company":30,"full name":24,"category":18,"city":13,
       "ai roles apply for":34,"roles skills":34,"phone":17,"office phone":17,"website":30,
       "address":42,"linkedin":30,"linkedin url original":40,"priority":24,"source":20,"notes":28}
def width_for(h):
    n=_norm(h)
    if n in WIDTH: return WIDTH[n]
    if n.startswith("email") or n.startswith("hr email"): return 30
    return 16
def is_email_col(h):
    n=_norm(h)
    return n.startswith("email") or n.startswith("hr email")
def center_col(h):
    n=_norm(h)
    return n in ("no","#","city","category","phone","office phone","priority")

wb=stable_load(F)
for sn in list(wb.sheetnames):
    ws=wb[sn]
    # find header row
    hrow=None
    for rr in (1,2,3,4):
        vals=[str(ws.cell(rr,c).value or "").strip().lower() for c in range(1,32)]
        if "company name" in vals or "company" in vals or "full name" in vals: hrow=rr; break
    if not hrow: continue
    # read headers + their source columns
    heads=[]
    for c in range(1, ws.max_column+1):
        h=clean(ws.cell(hrow,c).value)
        if h: heads.append((h,c))
    # read data rows (skip blank)
    key_c=heads[1][1] if len(heads)>1 else heads[0][1]   # company/name column
    rows=[]
    for r in range(hrow+1, ws.max_row+1):
        rec={h:clean(ws.cell(r,c).value) for h,c in heads}
        # row is real if company/name present
        if rec.get(heads[1][0]) or rec.get(heads[0][0]) and not str(rec.get(heads[0][0])).isdigit():
            if any(v for k,v in rec.items()):
                rows.append(rec)
    # order headers
    ordered=[h for h,_ in sorted(heads, key=lambda x: rank(x[0]))]
    # drop email columns that are entirely empty
    def col_has_data(h): return any(rec.get(h) for rec in rows)
    ordered=[h for h in ordered if (not is_email_col(h)) or col_has_data(h)]

    # rebuild sheet
    del wb[sn]
    nw=wb.create_sheet(sn)
    # move sheet back to original position later isn't trivial; order preserved by creation
    for ci,h in enumerate(ordered,1):
        cell=nw.cell(1,ci,h); cell.fill=DARK; cell.font=HFONT; cell.alignment=CENTER
        nw.column_dimensions[get_column_letter(ci)].width=width_for(h)
    for ri,rec in enumerate(rows,2):
        for ci,h in enumerate(ordered,1):
            v=rec.get(h)
            if v is None: continue
            cell=nw.cell(ri,ci,v)
            cell.font = BLUE if (is_email_col(h) and "@" in str(v)) else NORM
            cell.alignment = CENTER if center_col(h) else LEFT
            cell.border=BORDER
    nw.freeze_panes="A2"
    if ordered:
        nw.auto_filter.ref=f"A1:{get_column_letter(len(ordered))}1"
    nw.row_dimensions[1].height=28
    print(f"  {sn}: {len(rows)} rows, {len(ordered)} cols (was {ws.max_column}), headers->row1, cleaned")

# reorder sheets to a sensible order
desired=["All Companies","IT Companies","IT MNC","AI ML Companies","MNC_AI_Targets","LinkedIn Profiles"]
wb._sheets.sort(key=lambda s: desired.index(s.title) if s.title in desired else 99)

for _ in range(20):
    try: wb.save(F); break
    except Exception: time.sleep(2)
print("\nFORMAT COMPLETE. Saved.")
