"""
Merge VCF HR phone numbers into Ahmedabad_IT_AIML_FINAL_MASTER.xlsx
Also adds any new companies/emails found in VCF.
Rebuilds both sheets with updated data.
"""
import re, pandas as pd, pdfplumber, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ═══════════════════════════════════════════════════════════════════════════
# A. RESEARCHED COMPANY DATABASE  (same as build_final_master.py)
# ═══════════════════════════════════════════════════════════════════════════
DB = {
"3rddigital":      {"a":"A-501, Empire Business Hub, Opp. ICICI Bank, Science City Rd, Sola, Ahmedabad","p":"+91 9033463047","l":"linkedin.com/company/3rd-digital","w":"https://www.3rddigital.com","r":"Software Developer"},
"acenurture":      {"a":"Pune / Ahmednagar, Maharashtra (not Ahmedabad)","p":"","l":"linkedin.com/company/acenurture","w":"https://www.acenurture.com","r":"IT Recruiter, HR Executive"},
"acespritech":     {"a":"403, Shalin Complex, Opp. Hotel President, Sector-11, Gandhinagar 382011","p":"+91 9825713947","l":"linkedin.com/company/acespritech-solutions-pvt.-ltd.","w":"https://www.acespritech.com","r":"Software Developer"},
"actualisation":   {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/actualisation-ai","w":"https://www.actualisation.ai","r":"AI Engineer, ML Researcher"},
"adp":             {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/adp","w":"https://www.adp.com","r":"Payroll Analyst, HR Tech Developer"},
"addwebsolution":  {"a":"705, Silicon Tower, Nr Samartheshwar Mahadev Marg, Law Garden, Ellisbridge, Ahmedabad 380009","p":"","l":"linkedin.com/company/addwebsolution","w":"https://www.addwebsolution.com","r":"AI/ML Developer, Laravel/Node Developer, React Developer"},
"agileinfoways":   {"a":"303/304 Abhishree Avenue, Near Nehru Nagar Cross Rd, Ambawadi, Ahmedabad 380015","p":"+91 7622081234","l":"linkedin.com/company/agile-infoways-pvt-ltd-","w":"https://www.agileinfoways.com","r":"Data Analyst, AI/ML Developer"},
"aglowiditsolutions":{"a":"501, City Center, Science City Rd, Sola, Ahmedabad 380060","p":"+91 8487981277","l":"linkedin.com/company/aglowid-it-solutions","w":"https://www.aglowiditsolutions.com","r":"Software Developer, Web Developer"},
"aionpixel":       {"a":"9F2 BCG Tower, Seaport-Airport Rd, Kochi, Kerala 682037 (not Ahmedabad)","p":"+91 4844717111","l":"linkedin.com/company/aionpixel","w":"https://www.aionpixel.com","r":"AI Developer, Computer Vision Engineer"},
"aipxperts":       {"a":"Shilp Epitome, 1003-1004, Bodakdev, Ahmedabad 380054","p":"+91 9099985430","l":"linkedin.com/company/aipxperts","w":"https://www.aipxperts.com","r":"AI/ML Engineer, Data Scientist"},
"airtel":          {"a":"Pan India","p":"","l":"linkedin.com/company/bharti-airtel","w":"https://www.airtel.com","r":"Data Analyst, Network Engineer"},
"aistatics":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/aistatics","w":"https://www.aistatics.com","r":"Data Scientist, AI Research Intern"},
"aividtechvision": {"a":"303 Westface, Zydus Hospital Road, Hebatpur, Thaltej, Ahmedabad 380054","p":"+91 7948993215","l":"linkedin.com/company/aividtechvision","w":"https://www.aividtechvision.com","r":"Computer Vision Engineer, AI Developer"},
"alakmalak":       {"a":"402, Corporate House, Nr Dinesh Hall, Ashram Road, Ahmedabad 380009","p":"","l":"linkedin.com/company/alakmalak-technologies","w":"https://www.alakmalak.com","r":"Web Developer, Mobile App Developer"},
"alfalabs":        {"a":"Mumbai, Maharashtra (not Ahmedabad)","p":"+91 9820555945","l":"linkedin.com/company/alfalabstechnologies","w":"https://www.alfalabstechnologies.com","r":"Software Developer, ML Engineer"},
"allegisgroup":    {"a":"Pan India (US HQ)","p":"","l":"linkedin.com/company/allegis-group","w":"https://www.allegisgroup.com","r":"IT Recruiter, Talent Acquisition"},
"alliancerecruitmentagency":{"a":"Ahmedabad, Gujarat","p":"+91 9334190324","l":"linkedin.com/company/alliance-recruitment-agency","w":"https://www.alliancerecruitmentagency.com","r":"HR Recruiter, Talent Acquisition"},
"amnex":           {"a":"B-1301, Mondeal Heights, Nr Novotel, S.G. Highway, Ahmedabad 380015","p":"+91 9727747451","l":"linkedin.com/company/amnex","w":"https://www.amnex.com","r":"Telecom Software Developer, Data Analyst"},
"analyticsliv":    {"a":"503, 31Five, Corporate Rd, Prahladnagar, Ahmedabad 380015","p":"+91 8320576622","l":"linkedin.com/company/analyticsliv","w":"https://www.analyticsliv.com","r":"Digital Analytics Consultant"},
"analyticsvidhya": {"a":"Noida / Pan India","p":"","l":"linkedin.com/company/analytics-vidhya","w":"https://www.analyticsvidhya.com","r":"Data Scientist, ML Trainer"},
"anblicks":        {"a":"Block-A, 801-806, Navratna Corporate Park, Ambli-Bopal Rd, Ahmedabad 380058","p":"","l":"linkedin.com/company/anblicks","w":"https://www.anblicks.com","r":"Data Analyst, Cloud Data Engineer"},
"apexon":          {"a":"Shreem Building, Off SG Highway, Near Iscon Crossroad, Jodhpur, Ahmedabad 380015","p":"","l":"linkedin.com/company/apexon","w":"https://www.apexon.com","r":"AI/ML Consultant, Data Analyst, Full Stack Developer"},
"appitsimple":     {"a":"D-3 Signature 2, Business Park, Sarkhej-Sanand Cross Rd, Ahmedabad 382210","p":"","l":"linkedin.com/company/appitsimple-infotek","w":"https://www.appitsimple.com","r":"Product Manager, Software Developer"},
"appunik":         {"a":"Ahmedabad, Gujarat","p":"+91 7435047427","l":"linkedin.com/company/appunik","w":"https://www.appunik.com","r":"Mobile App Developer, Software Developer"},
"appvintech":      {"a":"Greater Noida, UP / Netherlands HQ (not Ahmedabad)","p":"","l":"linkedin.com/company/appvin-technologies","w":"https://www.appvintech.com","r":"Mobile App Developer"},
"argusoft":        {"a":"A-66, GIDC Electronics Estate, Sector-25, Gandhinagar 382016","p":"+91 9328573110","l":"linkedin.com/company/argusoft","w":"https://www.argusoft.com","r":"AI Engineer, Data Analyst, IoT/Robotics Engineer"},
"artoonsolutions":  {"a":"Iscon Emporio 301, Nr Star Bazaar, Jodhpur Cross Rd, Satellite, Ahmedabad 380015","p":"+91 8320329068","l":"linkedin.com/company/artoon-solutions","w":"https://www.artoonsolutions.com","r":"Game Developer, Mobile App Developer"},
"ascendion":       {"a":"Pan India (US HQ)","p":"","l":"linkedin.com/company/ascendion","w":"https://www.ascendion.com","r":"Software Engineer, Data Analyst"},
"ashutec":         {"a":"Ahmedabad, Gujarat","p":"+91 6354235108","l":"linkedin.com/company/ashutec","w":"https://www.ashutec.com","r":"Software Developer"},
"astrotalk":       {"a":"Noida / Pan India","p":"","l":"linkedin.com/company/astrotalk","w":"https://www.astrotalk.com","r":"Software Engineer, Data Analyst"},
"atqor":           {"a":"A-311 Navratna Corporate Park, Iskcon-Ambli Rd, Ahmedabad 380058","p":"+91 7069043269","l":"linkedin.com/company/atqor","w":"https://www.atqor.com","r":"Software Developer, Web Developer"},
"atliq":           {"a":"Ahmedabad / Vadodara, Gujarat","p":"+91 9327729553","l":"linkedin.com/company/atliqtech","w":"https://www.atliq.com","r":"Data Analyst, AI Consultant, BI Developer"},
"atul":            {"a":"Valsad, Gujarat","p":"","l":"linkedin.com/company/atul-ltd","w":"https://www.atul.co.in","r":"Data Analyst (Chemical Manufacturing)"},
"auberginesolutions":{"a":"A2/10th Floor, Safal Profitaire, Corporate Rd, Prahlad Nagar, Ahmedabad 380015","p":"+91 7600693463","l":"linkedin.com/company/aubergine-solutions","w":"https://www.aubergine.co","r":"UX Designer, Product Developer"},
"aviasole":        {"a":"Ahmedabad, Gujarat","p":"+91 9104976466","l":"linkedin.com/company/aviasole-technologies","w":"https://www.aviasole.com","r":"Software Developer, Mobile App Developer"},
"axistencolabs":   {"a":"Ahmedabad, Gujarat","p":"+91 9106649361","l":"linkedin.com/company/axis-tencolabs","w":"https://www.axistencolabs.com","r":"Software Developer"},
"avenuesai":       {"a":"28th Floor, GIFT Two Building, Block 56, Road-5C, Zone-5, GIFT City, Gandhinagar 382050","p":"+91 79 67772204","l":"linkedin.com/company/avenuesai","w":"https://www.avenuesai.com","r":"AI Engineer, ML Engineer, Fintech Data Scientist"},
"azilen":          {"a":"12th & 13th Floor, B Square-1, Bopal-Ambli Rd, Ahmedabad 380054","p":"+91 7433001373","l":"linkedin.com/company/azilentechnologies","w":"https://www.azilen.com","r":"AI Engineer, ML Engineer, Full Stack Developer"},
"azilentechnologies":{"a":"12th & 13th Floor, B Square-1, Bopal-Ambli Rd, Ahmedabad 380054","p":"+91 7433001373","l":"linkedin.com/company/azilentechnologies","w":"https://www.azilen.com","r":"AI Engineer, ML Engineer, Full Stack Developer"},
"bacancy":         {"a":"1207-1208 Times Square Arcade, Thaltej-Shilaj Rd, Ahmedabad 380059","p":"+91 7940037674","l":"linkedin.com/company/bacancy-technology","w":"https://www.bacancytechnology.com","r":"AI/ML Developer, Data Analyst, React/Node Developer"},
"bacancytechnology":{"a":"1207-1208 Times Square Arcade, Thaltej-Shilaj Rd, Ahmedabad 380059","p":"+91 7940037674","l":"linkedin.com/company/bacancy-technology","w":"https://www.bacancytechnology.com","r":"AI/ML Developer, Data Analyst, React/Node Developer"},
"biztechcs":       {"a":"C/801, Dev Aurum Commercial, Nr Anandnagar Cross Rds, Prahlad Nagar, Ahmedabad 380015","p":"+91 9106747559","l":"linkedin.com/company/biztechcs","w":"https://www.biztechcs.com","r":"Software Developer, Web Developer"},
"bmcloud":         {"a":"Ahmedabad, Gujarat","p":"+91 7652962632","l":"linkedin.com/company/bm-cloud","w":"https://www.bmcloud.in","r":"Cloud Solutions, Software Developer"},
"bonrix":          {"a":"Ahmedabad, Gujarat","p":"+91 9426045500","l":"linkedin.com/company/bonrix-software-systems","w":"https://www.bonrix.net","r":"Software Developer, IoT Developer"},
"botreetechnologies":{"a":"FF-1,2 JBR Arcade, Science City Rd, Sola, Ahmedabad 380060","p":"","l":"linkedin.com/company/botreetechnologies","w":"https://www.botreetechnologies.com","r":"AI/ML Engineer, Python/Django Developer, Data Scientist"},
"brainvire":       {"a":"Sheth Corporate Tower, 9th & 10th Floor, Nr. Nagri Hospital, Ellisbridge, Ahmedabad 380009","p":"+91 7941054646","l":"linkedin.com/company/brainvire-infotech-inc","w":"https://www.brainvire.com","r":"Software Engineer (AI/ML), E-commerce Developer"},
"brilworks":       {"a":"503, Fortune Business Hub, Science City Rd, near Shell Petrol Pump, Sola, Ahmedabad 380060","p":"+91 9106810920","l":"linkedin.com/company/brilworks","w":"https://www.brilworks.com","r":"AI Engineer, MERN Stack Developer, Mobile App Developer"},
"bureau":          {"a":"Ahmedabad / Bengaluru","p":"","l":"linkedin.com/company/bureau-id","w":"https://www.bureau.id","r":"AI Risk Engineer, Data Scientist"},
"bytestechnolab":  {"a":"THE FIRST, B-805, Near Mansi Cross Road, Vastrapur, Ahmedabad 380015","p":"+91 9157320903","l":"linkedin.com/company/bytes-technolab","w":"https://www.bytestechnolab.com","r":"AI/ML Engineer, Data Scientist, Full Stack Developer"},
"capermint":       {"a":"304-305, Shivalik Shilp-1, Iscon Cross Rd, SG Highway, Ahmedabad 380015","p":"+91 8866687329","l":"linkedin.com/company/caperminttech","w":"https://www.capermint.com","r":"AI App Developer, Game Developer, AR/VR Engineer"},
"caperminttechnologies":{"a":"304-305, Shivalik Shilp-1, Iscon Cross Rd, SG Highway, Ahmedabad 380015","p":"+91 8866687329","l":"linkedin.com/company/caperminttech","w":"https://www.capermint.com","r":"AI App Developer, Game Developer, AR/VR Engineer"},
"capgemini":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/capgemini","w":"https://www.capgemini.com","r":"AI/ML Consultant, Data Analyst, Data Engineer"},
"cipl":            {"a":"Ahmedabad, Gujarat","p":"+91 9978931415","l":"linkedin.com/company/cipl","w":"https://www.cipl.com","r":"Software Developer, IT Services"},
"citrusbug":       {"a":"Ahmedabad, Gujarat","p":"+91 8128442240","l":"linkedin.com/company/citrusbug-technolabs","w":"https://www.citrusbug.co","r":"Software Developer, Web Developer"},
"cleardu":         {"a":"202 Nilkanth Avenue 2, Navjivan Press Rd, Ahmedabad 380014","p":"+91 9898936777","l":"linkedin.com/company/cleardu","w":"https://www.cleardu.com","r":"Fintech Developer, Software Developer, Data Analyst"},
"cmarix":          {"a":"Manav Mandir, Drive-In Rd, Memnagar, Ahmedabad 380051","p":"+91 8000050808","l":"linkedin.com/company/cmarix-technolab","w":"https://www.cmarix.com","r":"Mobile App Developer, Web Developer"},
"cmetric":         {"a":"302, IT Tower-2, Infocity, Gandhinagar 382009","p":"","l":"linkedin.com/company/c-metric-solutions-private-limited","w":"https://www.c-metric.com","r":"Software Developer, IT Solutions, Project Outsourcing"},
"codesclue":       {"a":"915-STC, Near Bopal-Ambli Cross Rd, Ahmedabad 380058","p":"+91 9998977764","l":"linkedin.com/company/codesclue-technologies","w":"https://www.codesclue.com","r":"AI/ML Developer, Data Scientist, Full Stack Developer"},
"codezeros":       {"a":"906-908, Signature 1, Sarkhej-Gandhinagar Hwy, Makarba, Ahmedabad 380051","p":"+91 9638998419","l":"linkedin.com/company/codezeros","w":"https://www.codezeros.com","r":"Blockchain Developer, AI Developer"},
"codeepsilon":     {"a":"602, City Center, Science City Road, Opp. Shukan Mall, Sola, Ahmedabad 380060","p":"","l":"linkedin.com/company/codeepsilon","w":"https://www.codeepsilon.com","r":"Software Developer, Python Developer, Full Stack Developer"},
"comprinno":       {"a":"Bangalore, Karnataka (not Ahmedabad)","p":"","l":"linkedin.com/company/comprinnotech","w":"https://www.comprinno.net","r":"AI/ML Developer, Computer Vision Engineer"},
"concettolabs":    {"a":"City Center, Science City Road, Ahmedabad 380060","p":"+91 9586777575","l":"linkedin.com/company/concetto-labs","w":"https://www.concettolabs.com","r":"AI Consultant, Data Analyst"},
"controlf5":       {"a":"Indore, Madhya Pradesh (not Ahmedabad)","p":"+91 9755524635","l":"linkedin.com/company/control-f5","w":"https://www.controlf5.in","r":"Web Developer, Mobile App Developer"},
"convergesol":     {"a":"Ahmedabad, Gujarat","p":"+91 7778059923","l":"linkedin.com/company/convergesol","w":"https://www.convergesolution.com","r":"Software Developer, IT Solutions"},
"credencys":       {"a":"Ahmedabad, Gujarat","p":"+91 9979871019","l":"linkedin.com/company/credencys-solutions","w":"https://www.credencys.com","r":"Software Developer, Web Developer"},
"cretumadvisory":  {"a":"New Delhi (not Ahmedabad)","p":"","l":"linkedin.com/company/cretum-advisory","w":"https://www.cretumadvisory.com","r":"Audit/CA, Finance Analyst"},
"crestdata":       {"a":"CDS House, Nr. Sarkhej-Sanand Circle, SG Rd, Makarba, Ahmedabad 382210","p":"+91 7940044200","l":"linkedin.com/company/crest-data-systems","w":"https://www.crestdata.ai","r":"Data Scientist, ML Engineer, Platform Engineer"},
"crestinfosystems":{"a":"411-414, Wood Square, L.P. Savani Rd, Adajan, Surat 395009 (not Ahmedabad)","p":"+91 8141333165","l":"linkedin.com/company/crest-infosystems-pvt-ltd","w":"https://www.crestinfosystems.com","r":"Software Developer, Full Stack Developer"},
"cypherox":        {"a":"Ahmedabad, Gujarat","p":"+91 9033199949","l":"linkedin.com/company/cypherox-technologies","w":"https://www.cypherox.com","r":"Software Developer, Mobile App Developer"},
"cygnetinfotech":  {"a":"16-Swastik Society, NR. Amco Bank, Stadium Circle, Navrangpura, Ahmedabad 380009","p":"+91 6358976018","l":"linkedin.com/company/cygnet-infotech","w":"https://www.cygnetinfotech.com","r":"Software Engineer (AI/ML), Data Analyst, ERP Consultant"},
"cybage":           {"a":"7th Floor, Brigade International Financial Centre, Block 14A, Zone-1, GIFT City, Gandhinagar 382355","p":"","l":"linkedin.com/company/cybage-software","w":"https://www.cybage.com","r":"Software Developer, Digital Product Engineer, AI Solutions"},
"datadwip":        {"a":"RE11, Near Vikramnagar, Iscon-Ambli Rd, Ambli, Ahmedabad 380058","p":"+91 9664702180","l":"linkedin.com/company/datadwip","w":"https://www.datadwip.com","r":"AI/ML Data Engineer, Data Annotation Specialist, Web Scraping Engineer"},
"datumquest":      {"a":"Zion Z1, 1109, Sindhu Bhavan Marg, nr Maple County Rd, Bodakdev, Ahmedabad 380054","p":"","l":"linkedin.com/company/datumquest","w":"https://www.datumquest.com","r":"AI Engineer, RPA Developer, ML Engineer, Data Scientist"},
"datavizz":        {"a":"Ahmedabad, Gujarat","p":"+91 7778921793","l":"linkedin.com/company/datavizz","w":"https://www.datavizz.in","r":"Data Analytics, Business Intelligence"},
"dayhawk":         {"a":"Ahmedabad, Gujarat","p":"+91 8120656928","l":"linkedin.com/company/dayhawk","w":"https://www.dayhawk.in","r":"Software Developer, IT Services"},
"deepcoder":       {"a":"504, Jaihind HN Safal, S.G. Highway, Ahmedabad","p":"+91 7016729301","l":"linkedin.com/company/deepcoder","w":"https://www.deepcoder.io","r":"AI Developer, Software Engineer"},
"deftboxsolutions": {"a":"Ahmedabad, Gujarat","p":"+91 7948005672","l":"linkedin.com/company/deftbox-solutions","w":"https://www.deftboxsolutions.com","r":"Software Developer, IT Solutions"},
"digiwagon":        {"a":"1013, Shivalik Shilp, ISKCON Cross Road, SG Highway, Ahmedabad 380054","p":"+91 9727558794","l":"linkedin.com/company/digiwagontechnologies","w":"https://www.digiwagon.com","r":"AI Developer, Digital Engineering, Mobile App Developer"},
"dosepacker":      {"a":"Basement-1, Kala Sagar Mall, Sattadhar Cross Rd, opp Sai Baba Temple, Ghatlodiya, Ahmedabad 380061","p":"","l":"linkedin.com/company/dosepacker","w":"https://www.dosepacker.com","r":"Embedded Software Engineer, Full Stack Developer, Pharma Tech"},
"drcsystems":      {"a":"Ahmedabad (GIFT City), Gujarat","p":"+91 7600353260","l":"linkedin.com/company/drc-systems","w":"https://www.drcsystems.com","r":"Software Developer, ERP Consultant"},
"drivebuddyai":    {"a":"A-1111, World Trade Tower, Off SG Highway, Makarba, Ahmedabad 380051","p":"+91 11 4117 0779","l":"linkedin.com/company/drivebuddyai","w":"https://www.drivebuddyai.co","r":"AI Fleet Safety Engineer, Data Analyst"},
"dxfactor":        {"a":"McLean, Texas, USA (not Ahmedabad)","p":"","l":"linkedin.com/company/dxfactor","w":"https://www.dxfactor.com","r":"AI Analyst, Data Engineer"},
"ecubix":          {"a":"209-212, Ornet Arcade, Opp. AUDA Garden, Vastrapur, Ahmedabad 380015","p":"+91 9512007172","l":"linkedin.com/company/ecubix","w":"https://www.ecubix.com","r":"Software Developer"},
"echoinnovateit":  {"a":"Ahmedabad, Gujarat","p":"+91 8780395644","l":"linkedin.com/company/echo-innovate-it","w":"https://www.echoinnovateit.com","r":"Software Developer, Mobile App Developer"},
"einfochips":      {"a":"11/A-B, Chandra Colony, Off C.G. Road, Ellisbridge, Ahmedabad 380006","p":"+91 79 26563705","l":"linkedin.com/company/einfochips","w":"https://www.einfochips.com","r":"AI/ML Engineer (Embedded Systems), Data Engineer"},
"elightwalk":      {"a":"611, Shivalik Square, Near Adani CNG Pump, Old Wadaj, Ahmedabad 380013","p":"+91 7600897405","l":"linkedin.com/company/elightwalk-technology-pvt-ltd","w":"https://www.elightwalk.com","r":"E-commerce Developer, Web Developer"},
"elisiontech":     {"a":"Ahmedabad, Gujarat","p":"+91 7041649394","l":"linkedin.com/company/elision-technolab","w":"https://www.elisiontec.com","r":"VoIP Developer, Telecom Software Engineer"},
"elisiontec":      {"a":"Ahmedabad, Gujarat","p":"+91 7041649394","l":"linkedin.com/company/elision-technolab","w":"https://www.elisiontec.com","r":"VoIP Developer, Telecom Software Engineer"},
"envitics":        {"a":"Ahmedabad, Gujarat","p":"+91 9974325018","l":"linkedin.com/company/envitics","w":"https://www.envitics.com","r":"Environmental Data Analyst"},
"esparkinfo":      {"a":"1001-1010, City Centre-2, Nr Shukan Mall Cross Rd, Science City Rd, Ahmedabad 380060","p":"+91 9023728510","l":"linkedin.com/company/esparkbiz","w":"https://www.esparkinfo.com","r":"AI/ML Engineer, Web Developer, Mobile App Developer"},
"e2logy":          {"a":"Ahmedabad, Gujarat","p":"+91 7926762385","l":"linkedin.com/company/e2logy","w":"https://www.e2logy.com","r":"Software Developer, Web Developer"},
"e2m":             {"a":"Ahmedabad, Gujarat","p":"+91 8140832518","l":"linkedin.com/company/e2m-solutions","w":"https://www.e2msolutions.com","r":"Content Marketing, Digital Marketing"},
"e2msolutions":    {"a":"Ahmedabad, Gujarat","p":"+91 8140832518","l":"linkedin.com/company/e2m-solutions","w":"https://www.e2msolutions.com","r":"Content Marketing, Digital Marketing"},
"ethicsinfotech":  {"a":"Ahmedabad, Gujarat","p":"+91 9106190473","l":"linkedin.com/company/ethics-infotech","w":"https://www.ethicsinfotech.in","r":"Software Developer, Web Developer"},
"evolutioninfosystem":{"a":"C-1210 Siddhi Vinayak Towers, Nr Kataria Arcade, Makarba, Ahmedabad 380054","p":"","l":"linkedin.com/company/evolution-infosystem","w":"https://www.evolutioninfosystem.com","r":"AI Developer, Web/Software Developer"},
"fitterfly":       {"a":"Mumbai / Ahmedabad","p":"","l":"linkedin.com/company/fitterfly","w":"https://www.fitterfly.com","r":"Data Scientist, Health Tech Developer"},
"flodataanalytics": {"a":"New Delhi (not Ahmedabad)","p":"+91 8048262007","l":"linkedin.com/company/flo-data-analytics","w":"https://www.flodataanalytics.com","r":"Data Analyst, Business Analyst"},
"fusioninformatics":{"a":"Bodakdev, Ahmedabad 380054","p":"+91 6358856146","l":"linkedin.com/company/fusion-informatics","w":"https://www.fusioninformatics.com","r":"AI/ML Engineer, Data Scientist, IoT Developer"},
"fxis":            {"a":"4th Floor, Sanskrut Building, Navrangpura, Ahmedabad 380009","p":"+91 9834135926","l":"linkedin.com/company/fxisai","w":"https://www.fxis.ai","r":"AI Solutions Architect, AI Engineer, Data Scientist"},
"fxisai":          {"a":"4th Floor, Sanskrut Building, Navrangpura, Ahmedabad 380009","p":"+91 9834135926","l":"linkedin.com/company/fxisai","w":"https://www.fxis.ai","r":"AI Solutions Architect, AI Engineer, Data Scientist"},
"gatewaytechnolabs":{"a":"Gateway House, Nr Drive-In Cinema, Thaltej, Ahmedabad 380054","p":"","l":"linkedin.com/company/gateway-group-of-companies","w":"https://www.gatewaytechnolabs.com","r":"Software Engineer, ERP Consultant, AI/ML Developer"},
"girirajdigital":  {"a":"Ahmedabad, Gujarat","p":"+91 9313936238","l":"linkedin.com/company/giriraj-digital","w":"https://www.girirajdigital.com","r":"Digital Marketing, Web Developer"},
"groovyweb":       {"a":"Ahmedabad, Gujarat","p":"+91 9737100736","l":"linkedin.com/company/groovy-web","w":"https://www.groovyweb.in","r":"Web Developer, Mobile App Developer"},
"growexx":         {"a":"Shivalik Abaise, 1105, Prahlad Nagar, Ahmedabad 380015","p":"+91 8866734543","l":"linkedin.com/company/growexx","w":"https://www.growexx.com","r":"AI Consultant, Data Consultant, Full Stack Developer"},
"hhaexchange":     {"a":"Ahmedabad, Gujarat (India Office)","p":"","l":"linkedin.com/company/hhaexchange","w":"https://www.hhaexchange.com","r":"Software Developer, Data Analyst"},
"hcltechnologies": {"a":"GIFT City, Gandhinagar 382426","p":"","l":"linkedin.com/company/hcl-technologies","w":"https://www.hcltech.com","r":"AI Consultant, Data Scientist, Cloud Architect, Digital Transformation"},
"hiteshi":         {"a":"Indore, Madhya Pradesh (not Ahmedabad)","p":"","l":"linkedin.com/company/hiteshi-technologies","w":"https://www.hiteshi.com","r":"Software Developer, Mobile App Developer"},
"holbox":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/holboxai","w":"https://www.holbox.ai","r":"AI/ML Engineer, Backend Developer, Generative AI"},
"holboxai":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/holboxai","w":"https://www.holbox.ai","r":"AI/ML Engineer, Backend Developer, Generative AI"},
"hupptechnologies":{"a":"Ahmedabad, Gujarat","p":"+91 7874515717","l":"linkedin.com/company/hupp-technologies","w":"https://www.hupp.in","r":"Software Developer, Mobile App Developer"},
"hyperlinkinfosystem":{"a":"Block C, 106/B Ganesh Meridian, Near Sola Bridge, SG Highway, Ahmedabad 380061","p":"+91 80001 61161","l":"linkedin.com/company/hyperlinkinfosystem","w":"https://www.hyperlinkinfosystem.com","r":"Data Analyst, App Developer (ML), Mobile Developer"},
"hypeteq":         {"a":"Ahmedabad, Gujarat","p":"+91 9054272806","l":"linkedin.com/company/hypeteq","w":"https://www.hypeteq.com","r":"Software Developer, Mobile App Developer"},
"ibm":             {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/ibm","w":"https://www.ibm.com","r":"Data Scientist, AI Application Developer, ML Engineer, AI Consultant"},
"indianic":        {"a":"2nd Floor, B-201, Devarc Mall, S.G. Highway, Satellite, Ahmedabad 380015","p":"+91 7961916000","l":"linkedin.com/company/indianic-infotech-limited","w":"https://www.indianic.com","r":"Data Analyst, Web Developer"},
"iflair":          {"a":"3rd Floor, Karma Complex, Beside Kaushambi Flats, C.G. Road, Ahmedabad 380007","p":"+91 9558803176","l":"linkedin.com/company/iflair-web-technologies-pvt.-ltd.","w":"https://www.iflair.com","r":"Web Developer, PHP Developer, Software Developer"},
"ifourtechnolab":  {"a":"601, Times Square Arcade, Nr Baghban Party Plot, Thaltej-Shilaj Rd, Thaltej, Ahmedabad 380059","p":"+91 8200979229","l":"linkedin.com/company/ifour-technolab-pvt-ltd","w":"https://www.ifourtechnolab.com","r":".NET Developer, Software Developer, AI Developer"},
"infocusp":        {"a":"301-310, 4th Floor, Gala Hub, Gala Gymkhana Rd, Bopal, Ahmedabad 380058","p":"","l":"linkedin.com/company/infocusp","w":"https://www.infocusp.com","r":"AI/ML Engineer, Data Scientist, Research Engineer"},
"intellytics":     {"a":"E315-320 Ganesh Glory, Jagatpur-Chenpur Rd, Chharodi, Ahmedabad 382481","p":"","l":"linkedin.com/company/intellytics-solutions","w":"https://www.intellyticssolutions.com","r":"Data Scientist, BI Analyst, AI Solutions"},
"infopulsetech":   {"a":"Gujarat, India","p":"+91 7016844972","l":"linkedin.com/company/infopulse-tech","w":"https://www.infopulsetech.com","r":"Software Developer, IT Services"},
"infosys":         {"a":"Ahmedabad / Gandhinagar, Gujarat","p":"","l":"linkedin.com/company/infosys","w":"https://www.infosys.com","r":"AI/ML Engineer, Data Scientist, Data Engineer, Big Data Architect"},
"inexture":        {"a":"Office No. 405, One World West, Near Ambli T-Junction, SP Ring Rd, Bopal, Ahmedabad 380058","p":"+91 9054409426","l":"linkedin.com/company/inexture","w":"https://www.inexture.com","r":"AI Developer, Python Developer, Django Developer"},
"inheritx":        {"a":"8th Floor, Panchdhara Complex, S.G. Highway, Bodakdev, Ahmedabad 380054","p":"+91 9492588586","l":"linkedin.com/company/inheritx-solutions-pvt-ltd","w":"https://www.inheritx.com","r":"Mobile App Developer, Web Developer, AI Integration"},
"infiniumglobal":  {"a":"Ahmedabad, Gujarat","p":"+91 9316104961","l":"linkedin.com/company/infinium-global-technologies","w":"https://www.infiniumglobal.com","r":"Software Developer, IT Services"},
"infilon":         {"a":"1101-1105, Shivalik Shilp 2, Opp. ITC Narmada Hotel, Judges Bunglow Road, Satellite, Ahmedabad 380015","p":"+91 8000230350","l":"linkedin.com/company/infilon-technologies-pvt-ltd","w":"https://www.infilon.com","r":"AI Developer, Web Developer, Mobile App Developer, ERP Consultant"},
"intellyticssolutions":{"a":"Office-915, Ganesh Glory, Jagatpur-Chenpur Rd, Chharodi, Ahmedabad 382481","p":"","l":"linkedin.com/company/intellytics-solutions","w":"https://www.intellyticssolutions.com","r":"Data Analyst, BI Developer"},
"intuz":           {"a":"908-910, Pinnacle Business Park, Corporate Rd, Ahmedabad 380015","p":"","l":"linkedin.com/company/intuz","w":"https://www.intuz.info","r":"Mobile App Developer, IoT Developer"},
"itpathsolution":  {"a":"Ahmedabad, Gujarat","p":"+91 7698608480","l":"linkedin.com/company/it-path-solutions","w":"https://www.itpathsolution.com","r":"Software Developer, Mobile App Developer"},
"j2ml":            {"a":"Ahmedabad, Gujarat","p":"+91 8160460945","l":"linkedin.com/company/j2ml-infotech","w":"https://www.j2ml.com","r":"AI/ML Developer, Data Scientist"},
"jbcodeapp":       {"a":"Ahmedabad, Gujarat","p":"+91 9909918338","l":"linkedin.com/company/jbcodeapp","w":"https://www.jbcodeapp.com","r":"Mobile App Developer, Web Developer"},
"jeavio":          {"a":"709, Sanket Heights, Sun Pharma Rd, Akshar Chowk, Vadodara 390020 (not Ahmedabad)","p":"","l":"linkedin.com/company/jeavio","w":"https://www.jeavio.com","r":"Software Engineer, Data Engineer"},
"latitudetechnolabs":{"a":"A-213 Siddhi Vinayak Tower, Nr Kataria Arcade, Makarba, Ahmedabad 380051","p":"+91 9898332166","l":"linkedin.com/company/latitude-technolabs","w":"https://www.latitudetechnolabs.com","r":"Web Developer, Mobile App Developer"},
"kautilyam":       {"a":"Ahmedabad, Gujarat","p":"+91 7600420470","l":"linkedin.com/company/kautilyam","w":"https://www.kautilyam.com","r":"Software Developer"},
"kenexai":         {"a":"Ahmedabad, Gujarat","p":"+91 9967414744","l":"linkedin.com/company/kenexai","w":"https://www.kenexai.com","r":"AI Solutions, Software Developer"},
"kcsitglobal":     {"a":"Ahmedabad, Gujarat (also Netherlands)","p":"","l":"linkedin.com/company/kcs-krish-compusoft-services","w":"https://www.kcsitglobal.com","r":"Software Engineer, Dynamics/ERP Consultant, IT Consultant"},
"kodytechnolab":   {"a":"2nd Floor, Block-J, Safal Monde, Nr. Rajpath Club, S.G. Highway, Bodakdev, Ahmedabad 380054","p":"+91 9377229944","l":"linkedin.com/company/kody-technolab-pvt-ltd","w":"https://www.kodytechnolab.com","r":"Mobile App Developer, AI Developer"},
"krishaweb":       {"a":"2nd Floor, Kamalkant Complex, Mithakhali, Ahmedabad 380006","p":"+1 3015910989","l":"linkedin.com/company/krishaweb","w":"https://www.krishaweb.com","r":"Web Developer, PHP Developer"},
"logicrays":       {"a":"Avalon Hotel, Off Sindhu Bhavan Marg, Bodakdev, Ahmedabad 380054","p":"+91 9316803688","l":"linkedin.com/company/logicrays","w":"https://www.logicrays.com","r":"Software Developer, Web Developer"},
"magnetoitsolutions":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/magneto-it-solutions","w":"https://www.magnetoitsolutions.com","r":"E-commerce Developer, Data Analyst"},
"manektech":       {"a":"Ahmedabad, Gujarat","p":"+91 8320093859","l":"linkedin.com/company/manektech","w":"https://www.manektech.com","r":"Software Developer, Web Developer"},
"marutitech":      {"a":"10th Floor, The Ridge, Opp. Novotel, Iscon Char Rasta, Ahmedabad 380060","p":"+91 9879101067","l":"linkedin.com/company/maruti-techlabs-pvt-ltd","w":"https://www.marutitech.com","r":"Deep Learning Engineer, AI Researcher, NLP Engineer"},
"mastek":          {"a":"804/805 President House, Opp. C N Vidyalaya, Nr. Ambawadi Circle, Ahmedabad 380006","p":"+91 6358057439","l":"linkedin.com/company/mastek","w":"https://www.mastek.com","r":"Software Developer, Data Engineer, Cloud Architect"},
"metizsoft":       {"a":"A-802, Ganesh Plaza, Nr Navrangpura Post Office, Navrangpura, Ahmedabad 380009","p":"+91 9898964818","l":"linkedin.com/company/metizsoft-solutions","w":"https://www.metizsoft.com","r":"AI/ML Developer, Mobile App Developer, Software Developer"},
"mindinventory":   {"a":"801, City Centre 2, Science City Road, Sola, Ahmedabad 380060","p":"+1 216-609-0691","l":"linkedin.com/company/mindinventory","w":"https://www.mindinventory.com","r":"AI/ML Developer, Mobile App Developer"},
"mobmaxime":       {"a":"308, Lilamani Corporate Heights, New Vadaj, Ahmedabad 380013","p":"","l":"linkedin.com/company/mobmaxime","w":"https://www.mobmaxime.com","r":"Mobile App Developer"},
"moontechnolabs":  {"a":"Ahmedabad, Gujarat","p":"+91 9898255109","l":"linkedin.com/company/moon-technolabs","w":"https://www.moontechnolabs.com","r":"Mobile App Developer, AI Integration Developer"},
"moweb":           {"a":"403, Sarthak House 1, C.G. Road, Navrangpura, Ahmedabad 380009","p":"+91 8469699995","l":"linkedin.com/company/moweb-technologies","w":"https://www.moweb.com","r":"Mobile App Developer"},
"mspitconcepts":   {"a":"Ahmedabad, Gujarat","p":"+91 9106988343","l":"linkedin.com/company/msp-it-concepts","w":"https://www.mspitconcepts.com","r":"Software Developer, IT Services"},
"neuramonks":      {"a":"Ahmedabad, Gujarat","p":"+91 9409525981","l":"linkedin.com/company/neuramonks","w":"https://www.neuramonks.com","r":"AI/ML Developer, NLP Engineer, Data Scientist"},
"nevina":           {"a":"Sun Westbank A-829, Ashram Road, Nr Vallabh Sadan Riverfront, Ahmedabad 380015","p":"","l":"linkedin.com/company/nevinainfotech","w":"https://www.nevinainfotech.com","r":"AI/ML Developer, IoT Developer, AR/VR Engineer, Mobile App Developer"},
"nexgits":         {"a":"E-801, Ganesh Glory 11, Near BSNL Office, S.G. Highway, Jagatpur, Ahmedabad 382470","p":"+91 9265626573","l":"linkedin.com/company/nexgits","w":"https://www.nexgits.com","r":"AI/ML Developer, AR/VR Engineer, Game Developer"},
"nexuslinkservices":{"a":"Ahmedabad, Gujarat","p":"+91 9624411081","l":"linkedin.com/company/nexus-link-services","w":"https://www.nexuslinkservices.com","r":"IT Recruitment, HR Services"},
"nextsavy":        {"a":"Ahmedabad, Gujarat","p":"+91 9106654896","l":"linkedin.com/company/nextsavy-technologies","w":"https://www.nextsavy.com","r":"Software Developer, Web Developer"},
"nichetech":       {"a":"A822, Money Plant High Street, SG Highway, Jagatpur, Ahmedabad 382470","p":"+91 9512180005","l":"linkedin.com/company/nichetechsolutions","w":"https://www.nichetechsolutions.com","r":"AI/ML Developer, Mobile App Developer, Web Developer"},
"nimblechapps":    {"a":"529, Iscon Emporio, Besides Star Bazar, Jodhpur Cross Roads, Satellite, Ahmedabad 380015","p":"+91 9428241327","l":"linkedin.com/company/nimblechapps","w":"https://www.nimblechapps.com","r":"Mobile App Developer, Web Developer, UI/UX Designer"},
"ninjatechnolabs": {"a":"518-528, Yash Arian, Swami Vivekanand Cir, Memnagar, Ahmedabad 380052","p":"+91 9924377804","l":"linkedin.com/company/ninja-technolabs","w":"https://www.ninjatechnolabs.com","r":"Mobile App Developer, Web Developer"},
"nividous":        {"a":"11th Floor, Shivalik Abaise, Prahladnagar, Ahmedabad 380015","p":"+91 79 4008 1681","l":"linkedin.com/company/nividous","w":"https://www.nividous.com","r":"RPA Developer, AI Automation Engineer, IDP Specialist"},
"openxcell":       {"a":"202-203, Baleshwar Avenue, Opp. Rajpath Club, SG Highway, Ahmedabad 380054","p":"+91 9998222929","l":"linkedin.com/company/openxcell","w":"https://www.openxcell.com","r":"AI Developer, ML Engineer, Mobile App Developer"},
"orcaminds":       {"a":"55, 2nd Floor, Samet-2, Khanpur, Ahmedabad 380001","p":"+91 7984433793","l":"linkedin.com/company/orcaminds","w":"https://www.orcaminds.in","r":"AI/ML Engineer, Computer Vision Engineer, NLP Engineer, Data Scientist"},
"pearlinfo":       {"a":"Ahmedabad, Gujarat","p":"+91 8905510009","l":"linkedin.com/company/pearl-info","w":"https://www.pearlinfo.net","r":"Software Developer, IT Services"},
"pirimidtech":     {"a":"A 1001-1002, Ratnakar Nine Square, Opp. ITC Narmada, Ahmedabad 380015","p":"+91 7043536760","l":"linkedin.com/company/pirimidtech","w":"https://www.pirimidtech.com","r":"Software Developer, SaaS Engineer"},
"pragnakalp":      {"a":"D-916, Ganesh Glory 11, Jagatpur Road, Gota, Ahmedabad 382481","p":"+91 9727705677","l":"linkedin.com/company/pragnakalp-techlabs","w":"https://www.pragnakalp.com","r":"Generative AI Engineer, NLP Engineer, Python Developer, Chatbot Developer"},
"prismetric":      {"a":"604, IT Tower-1, Nr. Indroda Circle, InfoCity, Gandhinagar 382007","p":"+91 7283845358","l":"linkedin.com/company/prismetric-technologies","w":"https://www.prismetric.com","r":"AI Engineer, Mobile App Developer, IoT Developer"},
"prydan":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/prydan","w":"https://www.prydan.com","r":"Software Developer, QA Engineer"},
"qdev":            {"a":"Ahmedabad, Gujarat","p":"+91 9328224733","l":"linkedin.com/company/qdev","w":"https://www.qdev.in","r":"Software Developer, Mobile App Developer"},
"qmetry":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/qmetry","w":"https://www.qmetry.com","r":"QA Engineer, Test Automation Engineer"},
"qrioustech":      {"a":"Office 907-909, Block-A, Dev Auram, Anandnagar, Prahlad Nagar, Ahmedabad 380015","p":"+91 6355794606","l":"linkedin.com/company/qrious-tech","w":"https://www.qrioustech.com","r":"Software Developer"},
"quixom":          {"a":"220 Satyam Complex, Opp JBR, Science City Rd, Sola, Ahmedabad 380060","p":"+91 7567751959","l":"linkedin.com/company/quixom-technology","w":"https://www.quixom.com","r":"Python Developer, Web Developer, Mobile App Developer"},
"peerbits":        {"a":"409-419, Solitaire Connect, Nr Gallops Motors, Makarba, Ahmedabad 380015","p":"+91 79 48000686","l":"linkedin.com/company/peerbits","w":"https://www.peerbits.com","r":"Software Developer, Healthcare Platform Engineer, Mobile App Developer"},
"radixweb":        {"a":"401, Anandmangal-II, B/H Omkar House, C.G. Road, Navrangpura, Ahmedabad 380009","p":"+91 7935200685","l":"linkedin.com/company/radixweb","w":"https://www.radixweb.com","r":"Data Analyst, BI Specialist, Full Stack Developer"},
"rapidops":        {"a":"501, Satyamev Eminence, B/S Saptak Bungalows, Science City Rd, Sola, Ahmedabad 380060","p":"+91 9099793087","l":"linkedin.com/company/rapidops","w":"https://www.rapidops.com","r":"AI/ML Engineer, Data Scientist, SaaS Product Engineer"},
"rishabhsoft":     {"a":"Devx, 4th Floor, Binori B Square3, Sindhu Bhavan Rd, Ahmedabad 380054","p":"+91 8511122697","l":"linkedin.com/company/rishabh-software","w":"https://www.rishabhsoft.com","r":"Software Engineer (BI/Analytics), .NET Developer"},
"rushkar":         {"a":"Ahmedabad, Gujarat","p":"+91 9624008889","l":"linkedin.com/company/rushkar-technology","w":"https://www.rushkar.com","r":"Software Developer, Web Developer"},
"rysun":           {"a":"4th Floor, Atal-Kalam Research Park, Opp. GUSEC, Gujarat University, Ahmedabad 380009","p":"+91 7926579333","l":"linkedin.com/company/rysun","w":"https://www.rysun.com","r":"Software Developer, QA Engineer"},
"saras-3d":        {"a":"31, 3rd Floor, Commerce House 4, Prahladnagar, Ahmedabad 380015","p":"+91 7948061514","l":"linkedin.com/company/saras-3d","w":"https://www.saras-3d.com","r":"3D Engineer, CAD Designer"},
"shaligraminfotech":{"a":"Ahmedabad, Gujarat","p":"+91 9909984567","l":"linkedin.com/company/shaligram-infotech","w":"https://www.shaligraminfotech.com","r":"Software Developer, Mobile App Developer"},
"sharkstriker":    {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/sharkstriker","w":"https://www.sharkstriker.com","r":"Cybersecurity Engineer, SOC Analyst"},
"silicon":         {"a":"Ahmedabad, Gujarat","p":"+91 7201023434","l":"linkedin.com/company/silicon-info","w":"https://www.siliconinfo.com","r":"Software Developer, Web Developer"},
"sapphiresolutions":{"a":"C/102-103, Ganesh Meridian, Opp Kargil Petrol Pump, S.G. Highway, Ahmedabad 380060","p":"+91 9429709662","l":"linkedin.com/company/sapphire-software-solutions","w":"https://www.sapphiresolutions.net","r":"Web Developer, Mobile App Developer, Cloud Engineer"},
"sculptsoft":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/sculptsoft","w":"https://www.sculptsoft.com","r":"AI/ML Engineer, Data Scientist, Deep Learning Engineer"},
"shivlab":         {"a":"7th Floor, PV Enclave, Opp Courtyard by Marriott, Off Sindhu Bhavan Rd, Bodakdev, Ahmedabad 380054","p":"+91 9016777787","l":"linkedin.com/company/shiv-technolabs","w":"https://www.shivlab.com","r":"AI Developer, Mobile App Developer, Odoo Developer"},
"sigmasolve":      {"a":"801-803, PV Enclave, ICICI Bank Lane, Sindhu Bhavan Rd, Bodakdev, Ahmedabad 380054","p":"","l":"linkedin.com/company/sigmasolvelimited","w":"https://www.sigmasolve.com","r":"AI Engineer, Data Engineer, BI Analyst"},
"siliconinfo":     {"a":"Ahmedabad, Gujarat","p":"+91 7201023434","l":"linkedin.com/company/silicon-info","w":"https://www.siliconinfo.com","r":"Software Developer, Web Developer"},
"silvertouch":     {"a":"Silver Touch House, Opp. Suryarath Complex, Panchavati Circle, Ahmedabad 380006","p":"+91 7940022770","l":"linkedin.com/company/silver-touch-technologies-ltd","w":"https://www.silvertouch.com","r":"Software Developer, ERP Consultant"},
"simcom":          {"a":"Ahmedabad, Gujarat","p":"+91 9998618213","l":"linkedin.com/company/simcom","w":"https://www.simcom.in","r":"Software Developer, IT Services"},
"simform":         {"a":"501 Binori BSquare2, Near Double Tree by Hilton, Ambli Rd, Bopal, Ahmedabad 380054","p":"+91 7940070170","l":"linkedin.com/company/simform","w":"https://www.simform.com","r":"Data Engineer, ML Operations Engineer, Cloud Architect"},
"sixsigma":        {"a":"Ahmedabad, Gujarat","p":"+91 8690969090","l":"linkedin.com/company/six-sigma-it-solutions","w":"https://www.sixsigmaitsolutions.com","r":"Software Developer, IT Services"},
"skywinds":        {"a":"Ahmedabad, Gujarat","p":"+91 9773124272","l":"linkedin.com/company/skywinds-technologies","w":"https://www.skywinds.tech","r":"Software Developer, Mobile App Developer"},
"softcom":         {"a":"Vadodara (Baroda), Gujarat","p":"+91 9879021944","l":"linkedin.com/company/softcom-it-solutions","w":"https://www.softcomit.com","r":"Software Developer, IT Services"},
"smartsensesolutions":{"a":"Unit 2 & 3, 4th Floor, GIFT One, GIFT City, Gandhinagar 382355","p":"","l":"linkedin.com/company/smartsense-consulting-solutions-pvt-ltd-","w":"https://www.smartsensesolutions.com","r":"IoT Engineer, Software Developer, Data Engineer"},
"softvanlabs":     {"a":"305, Sigma Legacy, IIM Road, Panjrapol Cross-road, Ahmedabad 380015","p":"+91 6358922311","l":"linkedin.com/company/softvanlabs","w":"https://www.softvanlabs.com","r":"ML Engineer, RPA Developer, AI Developer"},
"softwebsolutions": {"a":"Block 5 & 6, Garden View, Corporate House, Bodakdev, Ahmedabad 380054","p":"+1 866-345-7638","l":"linkedin.com/company/softwebsolutionsinc","w":"https://www.softwebsolutions.com","r":"AI for IoT, Data Services, Digital Transformation"},
"solulab":         {"a":"812-16, Times Square 1, Opp. Baghban Party Plot, Thaltej, Ahmedabad 380059","p":"+91 9427026888","l":"linkedin.com/company/solulab","w":"https://www.solulab.com","r":"AI Engineer, Blockchain Developer, ML Developer"},
"solutionanalysts":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/solution-analysts-pvt-ltd","w":"https://www.solutionanalysts.com","r":"Business Intelligence Analyst, Data Analyst"},
"solvative":       {"a":"Ahmedabad, Gujarat","p":"+91 9510735925","l":"linkedin.com/company/solvative","w":"https://www.solvative.com","r":"Software Developer, Digital Solutions"},
"spaceogroup":     {"a":"1005, 10th Floor, Abhishree Adroit, Mansi Circle, Ahmedabad 380015","p":"+91 9316757277","l":"linkedin.com/company/space-o-technologies","w":"https://www.spaceotechnologies.com","r":"Mobile App Developer, AI Integration Developer"},
"specindia":       {"a":"SPEC House, Parth Complex, Near Swastik Cross Roads, Navrangpura, Ahmedabad 380009","p":"+91 79 2640 4031","l":"linkedin.com/company/spec-india","w":"https://www.spec-india.com","r":"BI Analyst, Data Analyst, Software Developer"},
"squadtechnologies":{"a":"B-1005, Shapath Hexa, Nr. Sola Bridge, Ahmedabad 380060","p":"+91 7940398310","l":"linkedin.com/company/squad-technologies","w":"https://www.squadtechnologies.com","r":"Software Developer"},
"stalwartitsolution":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/stalwart-it-solution","w":"https://www.stalwartitsolution.com","r":"Software Developer"},
"stellarmind":     {"a":"A-1307 Privilon, Ambli BRT Rd, Iskcon Cross Rd, Ahmedabad 380059","p":"+91 7863889382","l":"linkedin.com/company/stellarmind-ai","w":"https://www.stellarmind.ai","r":"AI Engineer, ML Researcher, NLP Specialist"},
"spectricssolutions":{"a":"913, The Capital 2, Science City Rd, Ahmedabad 380060","p":"+91 9081431434","l":"linkedin.com/company/spectrics-solutions","w":"https://www.spectricssolutions.com","r":"AI/ML Engineer, Software Developer, SEO"},
"syndell":         {"a":"Ahmedabad, Gujarat","p":"+91 6355614590","l":"linkedin.com/company/syndell-technologies","w":"https://www.syndell.com","r":"Software Developer, Web Developer"},
"thirdrocktechkno":{"a":"103, Sarita Complex, Near KB Dresswala, C.G. Road, Navrangpura, Ahmedabad 380009","p":"+91 8069196722","l":"linkedin.com/company/third-rock-techkno","w":"https://www.thirdrocktechkno.com","r":"AI Developer, Full Stack Developer, MERN Developer"},
"tisindia":        {"a":"Infocity, Gandhinagar 382007 (HQ: Noida, UP)","p":"+91 120 6790400","l":"linkedin.com/company/tis-india","w":"https://www.tisindia.com","r":"Web Developer, Software Developer, Digital Marketing"},
"taglineinfotech": {"a":"D-401, Titanium City Center, 100 Feet Anand Nagar Rd, Ahmedabad 380015","p":"+91 9913808285","l":"linkedin.com/company/tagline-infotech","w":"https://www.taglineinfotech.com","r":"Mobile App Developer, Web Developer, Software Developer"},
"tatvic":          {"a":"4th Floor, Camps Corner 2, 100 Feet Rd, Prahladnagar, Ahmedabad 380015","p":"+91 9909981217","l":"linkedin.com/company/tatvic","w":"https://www.tatvic.com","r":"Digital Analytics Consultant, Data Analyst, MarTech Expert"},
"tatvasoft":       {"a":"TatvaSoft House, Rajpath Club Rd, Opp. Golf Academy, Ahmedabad 380054","p":"+91 9601421472","l":"linkedin.com/company/tatvasoft","w":"https://www.tatvasoft.com","r":"Software Developer (Data Science), BI Developer"},
"techuz":          {"a":"1020, I Square, Nr Shukan Mall Cross Rd, Science City Rd, Sola, Ahmedabad 380060","p":"+91 9724522770","l":"linkedin.com/company/techuz","w":"https://www.techuz.com","r":"Web Developer, Mobile App Developer, Node/React Developer"},
"tcs":             {"a":"Gandhinagar / Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/tata-consultancy-services","w":"https://www.tcs.com","r":"AI Consultant, ML Specialist, Data Scientist, Data Engineer"},
"tecblic":         {"a":"510/511, Shivalik Shilp 2, Opp. ITC Narmada, Judges Bunglow Rd, Ahmedabad 380054","p":"+91 7777907777","l":"linkedin.com/company/tecblic","w":"https://www.tecblic.com","r":"Software Developer, Mobile App Developer"},
"techify":         {"a":"Ahmedabad, Gujarat","p":"+91 7862063131","l":"linkedin.com/company/techify","w":"https://www.techify.in","r":"Software Developer, Mobile App Developer"},
"techkul":         {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/techkul","w":"https://www.techkul.com","r":"E-commerce Developer, Magento Expert"},
"techmahindra":    {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/tech-mahindra","w":"https://www.techmahindra.com","r":"Data Scientist, ML Engineer, AI Architect, Software Engineer"},
"technostacks":    {"a":"10th Floor, Sun Square, Off CG Road, Beside Hotel Regenta, Navrangpura, Ahmedabad 380006","p":"+91 9909712616","l":"linkedin.com/company/technostacks-infotech-pvt-ltd","w":"https://www.technostacks.com","r":"Software Developer, IT Consultant, Digital Transformation"},
"technource":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/technource","w":"https://www.technource.com","r":"Mobile App Developer, Software Developer"},
"technoville":     {"a":"Ahmedabad, Gujarat","p":"+91 8000272722","l":"linkedin.com/company/technoville-consultants","w":"https://www.technovilleconsultants.com","r":"IT Recruitment, HR Consulting"},
"teksun":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/teksun-microsys","w":"https://www.teksun.com","r":"IoT Engineer, Embedded Systems Developer"},
"theonetechnologies":{"a":"Ahmedabad, Gujarat","p":"+91 6358288840","l":"linkedin.com/company/the-one-technologies","w":"https://www.theonetechnologies.com","r":"Software Developer, Web Developer"},
"tntra":           {"a":"B/H Rajpath Club, Nr. AUDA Garden, Off SG Road, Ahmedabad","p":"","l":"linkedin.com/company/tntra","w":"https://www.tntra.io","r":"Software Developer, Tech Entrepreneur, Product Manager"},
"trellance":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/trellance","w":"https://www.trellance.com","r":"Data Analyst, BI Developer, Credit Union Tech"},
"trentiums":       {"a":"Ahmedabad, Gujarat","p":"+91 9157194267","l":"linkedin.com/company/trentiums","w":"https://www.trentiums.com","r":"Software Developer, Enterprise Solutions"},
"tririd":          {"a":"417-418, Tower B, Navratna Corporate Park, Ambli, Ahmedabad 380058","p":"+91 8980010210","l":"linkedin.com/company/tririd","w":"https://www.tririd.com","r":"Software Developer"},
"trootech":        {"a":"605, 6th Floor, B Square, Iscon-Ambali BRTS Rd, Ahmedabad 380058","p":"+91 9033266951","l":"linkedin.com/company/trootech","w":"https://www.trootech.com","r":"Software Developer, Mobile App Developer, AI Solutions"},
"tuvoc":            {"a":"812-816, Times Square 1, Opp. Baghban Party Plot, Thaltej, Ahmedabad 380059","p":"+91 9825096687","l":"linkedin.com/company/tuvoc-technologies","w":"https://www.tuvoc.com","r":"AI/ML Developer, AR/VR Engineer, Blockchain Developer"},
"uffizio":         {"a":"Surat / Ahmedabad, Gujarat","p":"+91 7283855107","l":"linkedin.com/company/uffizio","w":"https://www.uffizio.com","r":"Fleet Management Software, IoT Developer"},
"uplers":          {"a":"Uplers House, Next to Kalasagar Mall, Sattadhar Cross Rd, Ahmedabad 380061","p":"+91 7940324566","l":"linkedin.com/company/uplers","w":"https://www.uplers.com","r":"Digital Marketer, Web Developer, Remote Talent"},
"vagaro":          {"a":"Ahmedabad, Gujarat (India Office)","p":"+91 7965191000","l":"linkedin.com/company/vagaro","w":"https://www.vagaro.com","r":"Software Developer, SaaS Platform"},
"valensdatalabs":  {"a":"Palladium Business Hub, Sarkhej Gandhinagar Hwy, Motera, Ahmedabad 380005","p":"+91 8044562783","l":"linkedin.com/company/valensdatalabs","w":"https://www.valensdatalabs.com","r":"Data Scientist, ML Engineer, Analytics Consultant"},
"varminect":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/varminect","w":"https://www.varminect.com","r":"Software Developer"},
"vihadigitalcommerce":{"a":"Ahmedabad, Gujarat","p":"+91 8866688722","l":"linkedin.com/company/viha-digital-commerce","w":"https://www.vihadigitalcommerce.com","r":"E-commerce Developer, Digital Marketing"},
"volansys":        {"a":"4th Floor, Aurelien Commercial Bldg, Near LJ Campus, S.G. Highway, Ahmedabad 382210","p":"+91 79 40041994","l":"linkedin.com/company/volansys","w":"https://www.volansys.com","r":"Embedded AI Engineer, ML Engineer, IoT Engineer"},
"vrinsoft":        {"a":"707, Elite Business Park, Shapath Hexa, SG Highway, Ahmedabad 380060","p":"+91 7227906117","l":"linkedin.com/company/vrinsoft-technologies-pvt-ltd","w":"https://www.vrinsofts.com","r":"Mobile App Developer, Software Developer, AI Solutions"},
"vrinsofts":       {"a":"707, Elite Business Park, Shapath Hexa, SG Highway, Ahmedabad 380060","p":"+91 7227906117","l":"linkedin.com/company/vrinsoft-technologies-pvt-ltd","w":"https://www.vrinsofts.com","r":"Mobile App Developer, Software Developer, AI Solutions"},
"wappnet":         {"a":"403, iSquare Corporate Park, Before Shukan Mall Cross Rd, Science City Rd, Sola, Ahmedabad 380060","p":"+91 6353604125","l":"linkedin.com/company/wappnet-systems","w":"https://www.wappnet.com","r":"Web Developer, Mobile App Developer, AI Integration"},
"webcluesinfotech":{"a":"Signature 1, Sarkhej-Gandhinagar Hwy, Makarba, Ahmedabad 380051","p":"+91 8141068282","l":"linkedin.com/company/webclues-infotech","w":"https://www.webcluesinfotech.com","r":"Web Developer, Mobile App Developer"},
"webelight":       {"a":"514, Silver Radiance 4, Nr Vivanta, S.G. Highway, Sola, Ahmedabad 380060","p":"+91 7405288281","l":"linkedin.com/company/webelight-solutions","w":"https://www.webelight.com","r":"AI/ML Engineer, Data Scientist, Full Stack Developer"},
"webential":       {"a":"Ahmedabad, Gujarat","p":"+91 7069645064","l":"linkedin.com/company/webential","w":"https://www.webential.com","r":"Web Developer, Digital Solutions"},
"webmobtech":      {"a":"Ahmedabad, Gujarat","p":"+91 7043866892","l":"linkedin.com/company/webmob-technologies","w":"https://www.webmobtech.com","r":"AI/ML Developer, Blockchain Developer, Mobile App Developer"},
"wpwebinfotech":   {"a":"704, City Center 2, Science City Rd, Ahmedabad 380060","p":"+91 7487953665","l":"linkedin.com/company/wpweb-infotech-pvt-ltd","w":"https://www.wpwebinfotech.com","r":"Web Developer, WordPress Developer, Mobile App Developer"},
"weblineindia":    {"a":"401, Arth Complex, Mithakhali 6 Roads, Navrangpura, Ahmedabad 380009","p":"+91 7926420897","l":"linkedin.com/company/weblineindia","w":"https://www.weblineindia.in","r":"Software Developer, Web Developer, Mobile App Developer"},
"webmotion":       {"a":"Ahmedabad, Gujarat","p":"+91 7801916926","l":"linkedin.com/company/webmotion","w":"https://www.webmotion.in","r":"Web Developer, Mobile App Developer"},
"webplanex":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/webplanex","w":"https://www.webplanex.com","r":"Web Developer, Digital Marketing"},
"wipro":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/wipro","w":"https://www.wipro.com","r":"Data Scientist, ML Engineer, AI Solution Architect"},
"worldweb":        {"a":"Ahmedabad, Gujarat","p":"+91 9157723337","l":"linkedin.com/company/world-web-technology","w":"https://www.worldwebtechnology.com","r":"Software Developer, Web Developer"},
"xceltec":         {"a":"Ahmedabad, Gujarat","p":"+91 9879698003","l":"linkedin.com/company/xceltec","w":"https://www.xceltec.com","r":"Mobile App Developer, Software Developer"},
"xenett":          {"a":"Ahmedabad, Gujarat","p":"+91 8780032460","l":"linkedin.com/company/xenett","w":"https://www.xenett.com","r":"Software Developer, SaaS Solutions"},
"zignuts":         {"a":"W210-217, Siddhraj Z Square, Opp The Landmark, Kudasan-Por Rd, Kudasan, Gandhinagar 382421","p":"+91 9427726620","l":"linkedin.com/company/zignuts-technolab","w":"https://www.zignuts.com","r":"Python Developer, AI/ML Engineer, Full Stack Developer"},
"yudiz":           {"a":"13th Floor, BSquare2, Iscon-Ambli Road, Ahmedabad 380054","p":"+91 9033975375","l":"linkedin.com/company/yudiz-solutions-ltd","w":"https://www.yudiz.com","r":"AI/ML Developer, Game Developer, AR/VR Engineer"},
"zipaworld":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/zipaworld","w":"https://www.zipaworld.com","r":"Logistics Tech Developer"},
"zobiweb":         {"a":"Ahmedabad, Gujarat","p":"+91 9510394120","l":"linkedin.com/company/zobi-web-solutions","w":"https://www.zobiweb.com","r":"Web Developer, Mobile App Developer"},
"4cpl":            {"a":"Ahmedabad, Gujarat","p":"+91 9328807088","l":"linkedin.com/company/4cpl","w":"https://www.4cpl.com","r":"IT Staffing, HR Consulting"},
"aditmicrosys":    {"a":"Ahmedabad, Gujarat","p":"+91 9824396414","l":"linkedin.com/company/adit-microsys","w":"https://www.aditmicrosys.com","r":"Software Developer, ERP Solutions"},
"alphaebarcode":   {"a":"Ahmedabad, Gujarat","p":"+91 9879668265","l":"linkedin.com/company/alphae-barcode","w":"https://www.alphaebarcode.com","r":"Barcode Solutions, Software Developer"},
"imperoit":        {"a":"Ahmedabad, Gujarat","p":"+91 9825108987","l":"linkedin.com/company/imperoit","w":"https://www.imperoit.com","r":"Software Developer, IT Solutions"},
"acquaintsofttech":{"a":"Ahmedabad, Gujarat","p":"+91 9327452612","l":"linkedin.com/company/acquaint-softtech","w":"https://www.acquaintsofttech.com","r":"Software Developer, Mobile App Developer"},
"avidclantech":    {"a":"Ahmedabad, Gujarat","p":"+91 9624679717","l":"linkedin.com/company/avidclan-technologies","w":"https://www.avidclantech.com","r":"Software Developer, Mobile App Developer"},
"ambesoft":        {"a":"Ahmedabad, Gujarat","p":"+91 9913341335","l":"linkedin.com/company/ambesoft","w":"https://www.ambesoft.in","r":"Software Developer, IT Services"},
# ── New companies from Mail list (1).pdf + deep research ─────────────────
"petpooja":        {"a":"Unit 81, 8th Floor, Titanium, Corporate Rd, Prahlad Nagar, Ahmedabad 380015","p":"","l":"linkedin.com/company/petpooja","w":"https://www.petpooja.com","r":"Software Developer, Product Manager, DevOps Engineer"},
"glidemtech":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/glidem-tech","w":"https://www.glidemtech.com","r":"Software Developer, Mobile App Developer"},
"magnusminds":     {"a":"503, Mauryansh Elanza, Shyamal Cross Rd, Satellite, Ahmedabad 380015","p":"+91 9714077532","l":"linkedin.com/company/magnusminds","w":"https://www.magnusminds.net","r":"Software Developer, .NET Developer"},
"begenuin":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/begenuin","w":"https://www.begenuin.com","r":"Software Developer, Community Platform"},
"scaledge":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/scaledge","w":"https://www.scaledge.io","r":"SaaS Developer, Fintech Developer"},
"codeverse":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/codeverse","w":"https://www.codeverse.co","r":"Software Developer, EdTech Developer"},
"devstree":        {"a":"B-433-442, Money Plant High Street, Near Ganesh Glory, Jagatpur Rd, S.G. Highway, Ahmedabad 382470","p":"+91 9904045044","l":"linkedin.com/company/devstreeit","w":"https://www.devstree.com","r":"Game Developer, AI/ML Developer, AR/VR Engineer"},
"devitpl":         {"a":"14, Aaryans Corporate Park, Thaltej-Shilaj Rd, Thaltej, Ahmedabad 380059","p":"+91 7926308854","l":"linkedin.com/company/dev-information-technology","w":"https://www.devitpl.com","r":"Software Developer, IT Services"},
"vasyerp":         {"a":"Block A, Neptune Corporate House, Behind Rajpath Club, Bodakdev, Ahmedabad 380054","p":"+91 8140364036","l":"linkedin.com/company/vasyerp","w":"https://www.vasyerp.com","r":"ERP Developer, Software Developer"},
"tridhyain":       {"a":"4th Floor, One World West, Ambli T-Junction, SP Ring Rd, Bopal, Ahmedabad","p":"+91 9638992877","l":"linkedin.com/company/tridhya-tech","w":"https://www.tridhyatech.com","r":"Software Developer, Mobile App Developer"},
"isummation":      {"a":"304, Shapath-3, Sarkhej-Gandhinagar Rd, Bodakdev, Ahmedabad","p":"+91 7926853054","l":"linkedin.com/company/isummation","w":"https://www.isummation.com","r":"Software Developer, IT Solutions"},
"iviewlabs":       {"a":"Adit Centre, Stadium Circle, Navrangpura, Ahmedabad 380009","p":"","l":"linkedin.com/company/iviewlabs","w":"https://www.iviewlabs.com","r":"AI/ML Engineer, Product Developer, Full Stack Developer"},
"synersoft":       {"a":"B-304, Wallstreet 2, Opp. Orient Club, Gujarat College Rd, Ellisbridge, Ahmedabad 380006","p":"+91 8849886408","l":"linkedin.com/company/synersoft","w":"https://www.synersoft.in","r":"ERP Developer, Software Developer"},
"tecnoprism":      {"a":"12th Floor, Vihav Supremus, Gotri Rd, Vadodara 390021 (not Ahmedabad)","p":"+91 6352937566","l":"linkedin.com/company/tecnoprism","w":"https://www.tecnoprism.com","r":"Software Developer, IT Services"},
"iovista":         {"a":"A506-509, The Capital, Science City Rd, Ahmedabad 380060","p":"+91 7990626431","l":"linkedin.com/company/iovista","w":"https://www.iovista.com","r":"E-commerce Developer, Magento Developer"},
"creatpix":        {"a":"253, Iscon Emporio, Nr Jodhpur Cross Rd, Satellite, Ahmedabad 380015","p":"+91 9316914079","l":"linkedin.com/company/creatpix","w":"https://www.creatpix.com","r":"Creative Developer, Software Developer"},
"feeltechsolutions":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/feel-tech-solutions","w":"https://www.feeltechsolutions.com","r":"Software Developer, Web Developer"},
"mnstechnologies":  {"a":"804 Rivera Wave, Kalawad Rd, Rajkot 360005 (not Ahmedabad)","p":"+91 7600015215","l":"linkedin.com/company/mns-technologies","w":"https://www.mnstechnologies.com","r":"Software Developer, IT Services"},
"ecodesoft":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/ecodesoft","w":"https://www.ecodesoft.com","r":"Software Developer, Web Developer"},
"iboontechnologies":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/iboon-technologies","w":"https://www.iboontechnologies.com","r":"Mobile App Developer, Software Developer"},
"tecsense":         {"a":"A-816/817/818, The Capital, Science City Rd, Sola, Ahmedabad 380060","p":"+91 7203014145","l":"linkedin.com/company/tec-sense","w":"https://www.tec-sense.com","r":"Software Developer, IT Solutions"},
"infynno":          {"a":"E-720, Ganesh Glory 11, Jagatpur Rd, Gota, Ahmedabad 382481","p":"","l":"linkedin.com/company/infynno-solutions","w":"https://www.infynno.com","r":"Software Developer, React Developer"},
"kernshell":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/kernshell","w":"https://www.kernshell.com","r":"Software Developer, Embedded Systems"},
"nextgensoft":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/nextgensoft","w":"https://www.nextgensoft.io","r":"Software Developer, IT Services"},
"heritagecyberworld":{"a":"605-606, Solaris Business Hub, Bhuyangdev, Ahmedabad 380013","p":"+91 9054504805","l":"linkedin.com/company/heritage-cyber-world","w":"https://www.heritagecyberworld.com","r":"IT Services, HR Solutions"},
"codiant":          {"a":"Indore, Madhya Pradesh (not Ahmedabad)","p":"+91 7314291704","l":"linkedin.com/company/codiant","w":"https://www.codiant.com","r":"Mobile App Developer, Software Developer"},
"istronix":         {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/istronix","w":"https://www.istronix.co","r":"Software Developer, IT Solutions"},
"lycerix":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/lycerix","w":"https://www.lycerix.com","r":"Software Developer"},
"definelabs":       {"a":"New Delhi / Pune (not Ahmedabad)","p":"+91 1141730344","l":"linkedin.com/company/define-labs","w":"https://www.definelabs.com","r":"Software Developer, Design"},
"rapidise":         {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/rapidise","w":"https://www.rapidise.co","r":"Software Developer, Mobile App"},
"workforce247":     {"a":"CH-1, Inspire Business Park, Shantigram, Nr. Vaishnodevi Circle, Gandhinagar 382421","p":"","l":"linkedin.com/company/workforce-247","w":"https://www.workforce247.com","r":"HR Solutions, Staffing Services"},
"softsages":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/softsages","w":"https://www.softsages.com","r":"Software Developer, Web Developer"},
"chillitray":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/chillitray","w":"https://www.chillitray.com","r":"Software Developer, E-commerce"},
"technotery":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/technotery","w":"https://www.technotery.com","r":"Software Developer"},
"infolabz":         {"a":"405-407, Vraj Avenue, Above Sam's Pizza, Navrangpura, Ahmedabad 380009","p":"+91 8866662662","l":"linkedin.com/company/infolabz","w":"https://www.infolabz.in","r":"Software Developer, IT Training"},
"samyak":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/samyak-computer-training","w":"https://www.samyak.com","r":"IT Training, Software Developer"},
"fx31labs":         {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/fx31-labs","w":"https://www.fx31labs.com","r":"Software Developer, IT Solutions"},
"technoidentity":   {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/techno-identity","w":"https://www.technoidentity.com","r":"Digital Marketing, Software Developer"},
"suesys":           {"a":"Pune, Maharashtra (not Ahmedabad)","p":"","l":"linkedin.com/company/suesys","w":"https://www.suesys.com","r":"Software Developer, IT Services"},
"vedity":           {"a":"C-305, Titanium Square, Thaltej Cross Rd, Ahmedabad 380054","p":"+91 7948002484","l":"linkedin.com/company/vedity","w":"https://www.vedity.com","r":"Software Developer, IT Solutions"},
# ── New AI companies ──────────────────────────────────────────────────────
"neuramoks":        {"a":"Ahmedabad, Gujarat","p":"+91 9537120913","l":"linkedin.com/company/neuramoks","w":"https://www.neuramoks.com","r":"AI/ML Developer, NLP Engineer, Data Scientist"},
"brainyneurals":    {"a":"1109, 11th Floor, The Orion, SG Highway, Gota, Ahmedabad 382481","p":"+91 9875206312","l":"linkedin.com/company/brainy-neurals","w":"https://www.brainyneurals.com","r":"AI/ML Engineer, Deep Learning Developer"},
"datavruti":        {"a":"B-1302, Karmyog Heights, S.V. Desai Marg, Navrangpura, Ahmedabad 380009","p":"","l":"linkedin.com/company/datavruti","w":"https://www.datavruti.com","r":"Data Scientist, AI/ML Engineer"},
"hunar":            {"a":"148, 5th Main Rd, HSR Layout, Bengaluru 560102 (not Ahmedabad)","p":"","l":"linkedin.com/company/hunar-ai","w":"https://www.hunar.ai","r":"AI Talent Platform, Data Scientist"},
"superr":           {"a":"Bengaluru / distributed team (not confirmed Ahmedabad)","p":"","l":"linkedin.com/company/superr-ai","w":"https://www.superr.ai","r":"AI Engineer, Generative AI Developer"},
"datakalp":         {"a":"Bangalore, Karnataka (not Ahmedabad)","p":"","l":"linkedin.com/company/datakalp","w":"https://www.datakalp.com","r":"Data Scientist, AI/ML Engineer, Analytics"},
"datasafeguard":    {"a":"Santa Clara, California, USA (not Ahmedabad)","p":"","l":"linkedin.com/company/datasafeguard","w":"https://www.datasafeguard.ai","r":"AI Privacy Engineer, Data Security Specialist"},
"toeho":            {"a":"Garhwa, Jharkhand (not Ahmedabad)","p":"","l":"linkedin.com/company/toeho-ai","w":"https://www.toeho.ai","r":"AI Solutions Architect, ML Engineer"},
"galaxyai":         {"a":"Cambridge, Massachusetts, USA (not Ahmedabad)","p":"","l":"linkedin.com/company/galaxy-ai","w":"https://www.galaxy.ai","r":"AI Engineer, Generative AI Developer"},
# ── New companies from DataScience Compeny List.pdf ──────────────────────
"bigscal":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/bigscal-technologies-pvt-ltd","w":"https://www.bigscal.com","r":"Software Engineer, Data Science Developer"},
"patoliyainfotech": {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/patoliya-infotech","w":"https://www.patoliyainfotech.com","r":"App Developer, Mobile App Developer"},
"quixom":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/quixom-technology","w":"https://www.quixom.com","r":"Python Developer, AI/ML Projects"},
"bigbrainy":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/bigbrainy-infotech","w":"https://www.bigbrainy.com","r":"Software Developer, Data Skills"},
"iqlance":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/iqlance-solutions-pvt-ltd","w":"https://www.iqlance.com","r":"App Developer, AI Integration"},
"monarchinnovation": {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/monarch-innovation-pvt-ltd","w":"https://www.monarch-innovation.com","r":"Software Developer, Data Analyst"},
"instinctools":     {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/instinctools","w":"https://www.instinctools.com","r":"AI/ML Engineer, Software Developer"},
"workerbees":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/workerbees-technologies","w":"https://www.workerbees.in","r":"Web Developer, Data Analytics"},
"settingsinfotech": {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/settings-infotech","w":"https://www.settingsinfotech.com","r":"Web Developer, App Developer"},
"compubrain":       {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/compubrain-private-limited","w":"https://www.compubrain.com","r":"Technology Consultant, Analytics"},
"alliancetek":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/alliancetek","w":"https://www.alliancetek.com","r":"Data Analyst, Software Developer"},
"netclues":         {"a":"1320, Stratum @ Venus Grounds, Nehru Nagar, Satellite, Ahmedabad 380015","p":"+91 7966637443","l":"linkedin.com/company/netclues-india","w":"https://www.netclues.com","r":"Digital Analyst, Web Developer"},
"uniquesdata":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/uniquesdata","w":"https://www.uniquesdata.com","r":"Data Analyst"},
"shreejidataanalytics":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/shreeji-data-analytics","w":"https://www.shreejidataanalytics.com","r":"Data Analyst, Analytics Consultant"},
"algobrain":        {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/algobrain-ai","w":"https://www.algobrain.ai","r":"AI/ML Specialist, Algorithm Engineer"},
"softvan":          {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/softvan-infotech","w":"https://www.softvan.com","r":"Software Developer, Data Skills"},
"aistechnolabs":    {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/ais-technolabs","w":"https://www.aistechnolabs.com","r":"Data Analyst, App Developer"},
"zeuslearning":     {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/zeus-learning","w":"https://www.zeuslearning.com","r":"Data Analyst, EdTech Developer"},
"iverve":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/i-verve-technologies","w":"https://www.i-verve.com","r":"Software Developer, Data Skills"},
"webkul":           {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/webkul-software","w":"https://www.webkul.com","r":"E-commerce Developer, Data Analyst"},
"clariontech":      {"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/clarion-technologies","w":"https://www.clariontech.com","r":"Software Engineer, BI/Analytics"},
"gatewaytechnolabs":{"a":"Ahmedabad, Gujarat","p":"","l":"linkedin.com/company/gateway-technolabs","w":"https://www.gatewaytechnolabs.com","r":"Software Engineer, Data Developer"},
# ── New companies discovered via VCF + web research (no email domain in sources) ──
"actowiz":         {"a":"101 & 201, Indraprasth Business Park, Beside DAV School, Makarba, Ahmedabad 380051","p":"+91 8401366964","l":"linkedin.com/company/actowiz","w":"https://www.actowizsolutions.com","r":"Web Scraping Developer, Data Engineer, Python Developer"},
"aivhub":          {"a":"Vadodara, Gujarat (Registered: Ahmedabad RoC)","p":"+91 9173611885","l":"linkedin.com/company/aivhub","w":"https://www.aivhub.com","r":"BI Developer, Data Visualization Engineer"},
"atharvasystem":   {"a":"Arista Business Space, Sindhu Bhavan Marg, Bodakdev, Ahmedabad 380054","p":"+91 9558821910","l":"linkedin.com/company/atharvasystem","w":"https://www.atharvasystem.com","r":"AI Software Developer, Mobile App Developer, Full Stack Developer"},
"evitalrx":        {"a":"3rd Floor, 4D Square, Motera, Ahmedabad 380005","p":"+91 9033071022","l":"linkedin.com/company/evitalrx","w":"https://www.evitalrx.in","r":"Software Developer, Cloud Engineer (Pharmacy SaaS)"},
"varianceinfotech":{"a":"Ahmedabad, Gujarat","p":"+91 9870057291","l":"linkedin.com/company/varianceinfo","w":"https://www.varianceinfotech.com","r":"AI/ML Developer, CRM Developer, Mobile App Developer"},
"hrbrainhub":      {"a":"Ahmedabad, Gujarat","p":"+91 7948995210","l":"linkedin.com/company/hrbrainhub","w":"https://www.hrbrainhub.com","r":"HR Consultant, AI Readiness Trainer"},
"isagebrum":       {"a":"501, Heer Aasha Arcade, Opp. Sagar-Sangeet-1, Solagam, Ahmedabad 380060","p":"+91 7990632836","l":"linkedin.com/company/isagebrum-technologies-pvt-ltd","w":"https://www.isagebrum.com","r":"Software Developer, Web Developer, IT Outsourcing"},
# ── New genuine Ahmedabad AI companies found via directory research ──────────
"creolestudios":   {"a":"A-404, Ratnaakar Nine Square, Opp ITC Narmada, Vastrapur, Ahmedabad 380015","p":"+91 7940086120","l":"linkedin.com/company/creole-studios","w":"https://www.creolestudios.com","r":"AI/ML Engineer, Generative AI Developer, Full Stack Developer"},
"codiste":         {"a":"805-807, Ganesh Glory, S.G. Highway, Ahmedabad 382481","p":"+91 9429005987","l":"linkedin.com/company/codiste","w":"https://www.codiste.com","r":"AI/ML Engineer, LLM Developer, Blockchain Developer"},
"nettyfy":         {"a":"717, Fortune Business Hub, Science City Rd, Sola, Ahmedabad 380060","p":"+91 9925051150","l":"linkedin.com/company/nettyfy-technologies","w":"https://www.nettyfy.com","r":"AI/ML Developer, Cloud Engineer, Blockchain Developer"},
"hiddenbrains":    {"a":"301, Sachet-4, Opp. Balaji Garden Restaurant, Satellite, Ahmedabad 380015","p":"+91 9898021433","l":"linkedin.com/company/hiddenbrains-infotech","w":"https://www.hiddenbrains.com","r":"AI/ML Developer, Software Developer, IT Consultant"},
# ── Genuine Ahmedabad presence confirmed (fill in previously blank AI entries) ──
"decimalpoint":    {"a":"D-601, World Trade Center, GIFT City, Gandhinagar 382355","p":"+91 9967066333","l":"linkedin.com/company/decimalpointanalytics","w":"https://decimalpointanalytics.com","r":"Data Analyst, AI/ML Engineer, Financial Analytics"},
"sublimedatasys":  {"a":"210, Iscon Elegance, Prahlad Nagar, Ahmedabad 380015","p":"+91 9825583937","l":"linkedin.com/company/sublime-data-systems","w":"https://www.sublimedatasys.com","r":"Data Engineer, Product Developer"},
"prama":           {"a":"33 Tapovan Colony, Anil Rd, Saraspur, Ahmedabad 380018 (also Phoenix USA)","p":"","l":"linkedin.com/company/prama-services","w":"https://prama.ai","r":"AI/ML Engineer, Data & Cloud Engineer"},
# ── Confirmed real companies but NOT Ahmedabad-based (location corrected) ────
"humantic":        {"a":"Bangalore, Karnataka (not Ahmedabad)","p":"","l":"linkedin.com/company/humantic-ai","w":"https://humantic.ai","r":"AI Engineer, ML Engineer"},
"panscience":      {"a":"New Delhi (not Ahmedabad)","p":"","l":"linkedin.com/company/pan-science-innovation","w":"https://www.panscience.xyz","r":"AI/ML Engineer, Deep Tech Developer"},
"datazip":         {"a":"677, 27th Main, 13th Cross, HSR Layout, Bangalore 560102 (not Ahmedabad)","p":"+91 9403613633","l":"linkedin.com/company/datazipio","w":"https://datazip.io","r":"Data Engineer, AI Engineer"},
"tailinc":         {"a":"Dayton, Ohio, USA (not Ahmedabad)","p":"","l":"linkedin.com/in/tailinc","w":"https://www.tailinc.ai","r":"AI/ML Engineer"},
"oblivious":       {"a":"Dublin, Ireland (not Ahmedabad)","p":"","l":"","w":"","r":"Privacy Engineer, Data Security Engineer"},
"digitalsherpa":   {"a":"Kolkata, West Bengal (not Ahmedabad)","p":"","l":"linkedin.com/company/digitalsherpa-ai","w":"https://www.digitalsherpa.ai","r":"AI Automation Developer"},
"hexadatasolutions":{"a":"Telangana (not Ahmedabad)","p":"","l":"linkedin.com/company/hexa-data-solutions","w":"https://hexadatasolutions.com","r":"Data Analyst, Automation Engineer"},
"crata":           {"a":"Barcelona, Spain (not Ahmedabad)","p":"","l":"","w":"https://www.crata-ai.com","r":"AI Consultant"},
"crataai":         {"a":"Barcelona, Spain (not Ahmedabad)","p":"","l":"","w":"https://www.crata-ai.com","r":"AI Consultant"},
"bzanalytics":     {"a":"Kerala / Dubai UAE (not Ahmedabad)","p":"","l":"linkedin.com/company/bzanalytics","w":"https://www.bzanalytics.ai","r":"Data Analyst"},
"loomylabs":       {"a":"San Francisco / Munich / Delhi (not Ahmedabad)","p":"","l":"","w":"https://loomylabs.ai","r":"AI Consultant, Software Engineer"},
# ── IT Services companies researched in deep-dive round 3 ────────────────────
"lucentinnovation":{"a":"9th Floor-902, Dwarkesh Business Hub, Motera, Ahmedabad 380005","p":"+1 8445823681","l":"linkedin.com/company/lucent-innovation","w":"https://www.lucentinnovation.com","r":"Software Developer, Web Developer"},
"conceptinfoway":  {"a":"801-B Parshwa Tower, S.G. Highway, Bodakdev, Ahmedabad 380054","p":"+91 7926872057","l":"linkedin.com/company/conceptinfoway","w":"https://www.conceptinfoway.net","r":"Software Developer, Web Developer"},
"technocolabs":    {"a":"Indore, Madhya Pradesh (not Ahmedabad)","p":"+91 8319291391","l":"","w":"https://www.technocolabs.com","r":"AI/ML Developer, Data Science Consultant"},
"excellonsoft":    {"a":"Nagpur, Maharashtra (not Ahmedabad)","p":"+91 7126699600","l":"","w":"https://www.excellonsoft.com","r":"Software Developer, ERP Consultant"},
"devkraft":        {"a":"New Delhi (registered); team also reported in Ahmedabad","p":"+91 9711113814","l":"linkedin.com/company/devkraft","w":"https://www.devkraft.com","r":"Software Developer, Product Engineer"},
"innvonix":        {"a":"11th Floor, I Square, Science City Rd, Sola, Ahmedabad 380060","p":"","l":"","w":"https://www.innvonix.com","r":"Software Developer, Web Developer"},
"vyaparapp":       {"a":"Bangalore, Karnataka (not Ahmedabad)","p":"+91 6364444752","l":"","w":"https://vyaparapp.in","r":"Software Developer"},
}

# ── Canonical display names (applied to override domain-derived ugly names) ──
DISPLAY_NAMES = {
    'drcsystems':                 'DRC Systems',
    'drivebuddyai':               'DriveBuddy AI',
    'agileinfoways':              'Agile Infoways',
    'alliancerecruitmentagency':  'Alliance Recruitment Agency',
    'hyperlinkinfosystem':        'Hyperlink InfoSystem',
    'moontechnolabs':             'Moon Technolabs',
    'softwebsolutions':           'Softweb Solutions',
    'spaceogroup':                'SpaceO Technologies',
    'specindia':                  'SPEC INDIA',
    'xceltec':                    'XcelTec',
    'fxis':                       'FXIS AI',
    'fxisai':                     'FXIS AI',
    'skywinds':                   'Skywinds Technologies',
    'comprinno':                  'Comprinno Technologies',
    'esparkinfo':                 'eSpark Info',
    'elisiontec':                 'Elision Technolabs',
    'elisiontech':                'Elision Technolabs',
    'convergesol':                'Converge Solution',
    'convergesolution':           'Converge Solution',
    'ethicsinfotech':             'Ethics Infotech',
    'einfochips':                 'eInfochips',
    'marutitech':                 'Maruti Tech Labs',
    'vrinsoft':                   'Vrinsoft Technologies',
    'vrinsofts':                  'Vrinsoft Technologies',
    'webcluesinfotech':           'WebClues Infotech',
    'thetechnoville':             'The Technoville',
    'neuramoks':                  'Neura Moks',
    'datavruti':                  'Datavruti',
    'datavruit':                  'Datavruti',   # typo variant → merge to same
    'bytestechnolab':             'Bytes Technolab',
    'alphaebarcode':              'Alphae Barcode',
    'nexuslinkservices':          'Nexus Link Services',
    'vihadigitalcommerce':        'Viha Digital Commerce',
    'deftboxsolutions':           'Deftbox Solutions',
    'elightwalk':                 'Elightwalk Technology',
    'siliconinfo':                'Silicon Info',
    'softcom':                    'Softcom IT Solutions',
    'isummation':                 'iSummation Technologies',
    'tridhyain':                  'Tridhya Tech',
    'vasyerp':                    'VasyERP',
    'mnstechnologies':            'MNS Technologies',
    'iboontechnologies':          'iBoon Technologies',
    'infynno':                    'Infynno Solutions',
    'heritagecyberworld':         'Heritage Cyberworld',
    'aglowiditsolutions':         'Aglowid IT Solutions',
    'acquaintsofttech':           'Acquaint Softtech',
    'avidclantech':               'Avidclan Technologies',
    'webclues':                   'WebClues Infotech',
    'weblineindia':               'Webline India',
    'worldweb':                   'World Web Technology',
    'sixsigma':                   'Six Sigma IT Solutions',
    'shaligraminfotech':          'Shaligram Infotech',
    'sharkstriker':               'Sharkstriker',
    'appitsimple':                'AppItSimple Infotek',
    'magnusminds':                'Magnus Minds',
    'tecnoprism':                 'TecnoPrism',
    'glidemtech':                 'Glide M Tech',
    'begenuin':                   'BeGenuin',
    'petpooja':                   'PetPooja',
    'codeverse':                  'Codeverse Technologies',
    'scaledge':                   'ScalEdge',
    'feeltechsolutions':          'Feel Tech Solutions',
    'ecodesoft':                  'Ecodesoftware',
    'bigscal':                    'Bigscal Technologies',
    'patoliyainfotech':           'Patoliya Infotech',
    'quixom':                     'QUIXOM Technology',
    'bigbrainy':                  'Bigbrainy Infotech',
    'iqlance':                    'IQLANCE Solutions',
    'monarchinnovation':          'Monarch Innovation',
    'compubrain':                 'CompuBrain',
    'alliancetek':                'AllianceTek',
    'netclues':                   'Netclues India',
    'uniquesdata':                'Uniquesdata',
    'shreejidataanalytics':       'Shreeji Data Analytics',
    'algobrain':                  'ALGOBRAIN AI',
    'workerbees':                 'WorkerBees Technologies',
    'clariontech':                'Clarion Technologies',
    'gatewaytechnolabs':          'Gateway Technolabs',
    'instinctools':               'Instinctools',
    'iverve':                     'i-Verve Technologies',
    'webkul':                     'Webkul Software',
    'zeuslearning':               'Zeus Learning',
    # AI/ML company canonical names
    'bzanalytics':                'BZAnalytics AI',
    'actualisation':              'Actualisation AI',
    'aipxperts':                  'AIpXperts',
    'aividtechvision':            'AIVID TechVision',
    'analyticsliv':               'AnalyticsLiv',
    'analyticsvidhya':            'Analytics Vidhya',
    'anblicks':                   'Anblicks',
    'badaadata':                  'Bada Data',
    'biobrain':                   'Biobrain AI',
    'brainyneurals':              'Brainy Neurals',
    'codework':                   'Codework AI',
    'crata':                      'Crata AI',
    'crestdata':                  'Crest Data Systems',
    'datakalp':                   'DataKalp',
    'datasafeguard':              'DataSafeguard AI',
    'datavruti':                  'Datavruti',
    'datavruit':                  'Datavruti',
    'datazip':                    'Datazip',
    'ddatalabs':                  'D Data Labs',
    'decimalpoint':               'Decimalpoint Analytics',
    'deepcoder':                  'DeepCoder.io',
    'digitalsherpa':              'DigitalSherpa AI',
    'dxfactor':                   'DXFactor',
    'flodataanalytics':           'Flo Data Analytics',
    'galaxy':                     'Galaxy AI',
    'galaxyai':                   'Galaxy AI',
    'gridanalyticsindia':         'Grid Analytics India',
    'hexadatasolutions':          'Hexa Data Solutions',
    'holbox':                     'Holbox AI',
    'humantic':                   'Humantic AI',
    'hunar':                      'Hunar AI',
    'infocusp':                   'InfoCusp',
    'intellyticssolutions':       'Intellytics Solutions',
    'loomylabs':                  'Loomy Labs',
    'makunaiglobal':              'Makun AI Global',
    'mindbrain':                  'MindBrain',
    'oblivious':                  'Oblivious AI',
    'panscience':                 'PanScience',
    'prama':                      'Prama AI',
    'psytech':                    'Psytech AI',
    'stellarmind':                'Stellarmind AI',
    'sublimedatasys':             'Sublime Data Systems',
    'superr':                     'Superr AI',
    'tailinc':                    'TailInc',
    'toeho':                      'Toeho AI',
    'valensdatalabs':             'Valens Data Labs',
    'voidspace':                  'Voidspace AI',
    'dxfactor':                   'DXFactor Analytics',
    'growexx':                    'Growexx',
    'solulab':                    'SoluLab',
    'mindinventory':              'MindInventory',
    'simform':                    'Simform',
    'concettolabs':               'Concetto Labs',
    'tatvic':                     'Tatvic',
    'neuramonks':                 'Neuramonks',
    # Domain-key variants (hyphen stripped)
    'crataai':                    'Crata AI',
    'holboxai':                   'Holbox AI',
    'bzanalyticsai':              'BZAnalytics AI',
    'actualisationai':            'Actualisation AI',
    'datasafeguardai':            'DataSafeguard AI',
    'stellarmindai':              'Stellarmind AI',
    'superrrai':                  'Superr AI',
    'hunarai':                    'Hunar AI',
    'toehoai':                    'Toeho AI',
    'galaxyai':                   'Galaxy AI',
    # Additional IT company name cleanup
    'rishabhsoft':                'Rishabh Software',
    'openxcell':                  'OpenXcell',
    'cypherox':                   'Cypherox Technologies',
    'atliq':                      'AtliQ Technologies',
    'trootech':                   'Trootech Business Solutions',
    'webential':                  'Webential',
    'solvative':                  'Solvative',
    'mastek':                     'Mastek',
    'bacancy':                    'Bacancy Technology',
    'bacancytechnology':          'Bacancy Technology',
    'tatvasoft':                  'TatvaSoft',
    'inexture':                   'Inexture Solutions',
    'bonrix':                     'Bonrix Software',
    'vagaro':                     'Vagaro India',
    'nividous':                   'Nividous',
    'radixweb':                   'Radixweb',
    'azilen':                     'Azilen Technologies',
    'azilentechnologies':         'Azilen Technologies',
    'capermint':                  'Capermint Technologies',
    'caperminttechnologies':      'Capermint Technologies',
    'manektech':                  'ManekTech',
    'infopulsetech':              'Infopulse Tech',
    'actowiz':                    'Actowiz Solutions',
    'actowizsolutions':           'Actowiz Solutions',
    'aivhub':                     'Aivhub India',
    'atharvasystem':              'Atharva System',
    'evitalrx':                   'eVitalRx',
    'varianceinfotech':           'Variance InfoTech',
    'hrbrainhub':                 'HR Brain Hub Consulting',
    'isagebrum':                  'iSageBrum Technologies',
    'creolestudios':              'Creole Studios',
    'codiste':                    'Codiste',
    'nettyfy':                    'Nettyfy Technologies',
    'hiddenbrains':               'Hidden Brains InfoTech',
    'decimalpoint':               'Decimalpoint Analytics',
    'sublimedatasys':             'Sublime Data Systems',
    'prama':                      'Prama',
    'humantic':                   'Humantic AI',
    'panscience':                 'PanScience Innovations',
    'datazip':                    'Datazip',
    'tailinc':                    'TailInc',
    'oblivious':                  'Oblivious AI',
    'digitalsherpa':              'DigitalSherpa AI',
    'hexadatasolutions':          'Hexa Data Solutions',
    'crata':                      'Crata AI',
    'bzanalytics':                'BZAnalytics',
    'loomylabs':                  'Loomy Labs',
}

# Known typo corrections for specific email addresses
EMAIL_CORRECTIONS = {
    'hr.admin@bitumag.co.incom': 'hr.admin@bitumag.co.in',
}

EMAIL_RE = re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')

def get_domain(email):
    try: return email.strip().lower().split('@')[1]
    except: return ''

def domain_key(domain):
    return domain.split('.')[0].lower().replace('-','').replace('_','')

def get_details(company_name, domain):
    k = domain_key(domain)
    cn = re.sub(r'[^a-z]', '', company_name.lower())
    for key in [k, cn]:
        if key in DB: return DB[key]
    # Partial match — require at least 7 chars to avoid wrong cross-matches
    for key in DB:
        if len(key) >= 7 and (key in k or key in cn):
            return DB[key]
        if len(k) >= 7 and k in key:
            return DB[key]
        if len(cn) >= 7 and cn in key:
            return DB[key]
    return {}

# Short keywords that need word-boundary check (too common in non-AI names)
AI_KW_WORD = ['ai', 'ml', 'bot', 'llm']
# Longer keywords safe to check as substrings
AI_KW_SUB  = ['analyt','neural','cognitiv','vision','nlp','deep','machine',
               'predict','genai','robotic','intelligence','science','mining',
               'automat','insight','model','brain','data','learn','smart']
# Companies incorrectly matched by keywords but clearly NOT AI/ML
EXPLICIT_IT_DOMAINS = {
    'airtel',              # telecom (contains 'ai' at start)
    'brainpayroll',        # payroll software (contains 'brain')
    'airditsoftware',      # general software (contains 'ai')
    'savitarainfotel',     # infotel/telecom ('ai' inside 'rain')
    'srisaioverseas',      # overseas recruitment ('ai' inside)
    'fernwoodporcelain',   # ceramics ('ai' in 'porcelain')
    'saifevetmed',         # veterinary medicine ('ai' in 'Saife')
    'litentertainmentco',  # entertainment ('ai' in 'entertain')
    'retailscan',          # retail tech ('ai' in 'retail')
    'samcomtechnobrains',  # IT services (contains 'brain')
    'wittybrains',         # web development (contains 'brain')
    'omniaautomation',     # business automation (not AI/ML)
    'savvydatacloudconsulting',  # cloud consulting
}
AI_DOMAINS = {'fxis','crestdata','holbox','drivebuddyai','aividtechvision','actualisation',
              'aistatics','stellarmind','infocusp','oblivious','datasafeguard','superr',
              'tailinc','datavruti','ddatalabs','bzanalytics','analyticsvidhya','comprinno',
              'humantic','makunaiglobal','panscience','prama','psytech','hunar','voidspace',
              'neuramonks','aipxperts','biobrain','loomylabs','galaxy','toeho','algobrain',
              'growexx','solulab','mindinventory','simform','concettolabs','valensdatalabs',
              'flodataanalytics','tatvic','anblicks','hexadatasolutions','dxfactor',
              'intellyticssolutions','codework','digitalsherpa','crata','decimalpoint',
              'j2ml','kenexai','datavizz','neuramoks','brainyneurals','datakalp',
              'galaxyai','superrai','datavruti','hunarai','datasafeguardai','toehoai',
              'spectricssolutions','nexgits','orcaminds','codesclue','codezeros',
              'datadwip','datumquest','avenuesai','fusioninformatics',
              'rapidops','webelight','volansys',
              'sigmasolve','brilworks','botreetechnologies',
              'sculptsoft','argusoft','thirdrocktechkno','zignuts',
              'softvanlabs','iviewlabs','webmobtech',
              'pragnakalp','prismetric','infilon','devstree',
              'digiwagon','tuvoc','nevina'}

def categorize(name, domain):
    d = domain_key(domain)
    n = name.lower()
    if d in EXPLICIT_IT_DOMAINS: return 'IT Services'
    if d in AI_DOMAINS: return 'AI / ML'
    for kw in AI_KW_SUB:
        if kw in n or kw in d: return 'AI / ML'
    for kw in AI_KW_WORD:
        if re.search(r'\b' + kw + r'\b', n) or re.search(r'\b' + kw + r'\b', d):
            return 'AI / ML'
    return 'IT Services'

# ═══════════════════════════════════════════════════════════════════════════
# B. LOAD & MERGE ALL DATA SOURCES
# ═══════════════════════════════════════════════════════════════════════════
company_emails = defaultdict(lambda: {'emails': [], 'name': ''})

JUNK_NAMES = {'company name company mail', 'company name email id', 'company name email',
              'company name', 'email id', 'linkedin url', 'nan', ''}

def add_email(company_name, email, prefer_name=True):
    email = EMAIL_CORRECTIONS.get(email.strip().lower(), email.strip())
    if not email or not EMAIL_RE.match(email): return
    d = get_domain(email)
    if not d or any(x in d for x in ['gmail','yahoo','outlook','hotmail','rediff']): return
    rec = company_emails[d]
    cn_clean = (company_name or '').strip().lower()
    is_junk = cn_clean in JUNK_NAMES or cn_clean.startswith('company name')
    if prefer_name and company_name and not is_junk:
        rec['name'] = company_name
    elif not rec['name'] and company_name and not is_junk:
        rec['name'] = company_name
    if email.lower() not in [e.lower() for e in rec['emails']]:
        rec['emails'].append(email.strip())

# ── Source 1: HR Mail List.xlsx ──────────────────────────────────────────
hr_df = pd.read_excel('HR Mail List.xlsx')
for _, row in hr_df.iterrows():
    company = str(row.get('Company Name','') or '').strip()
    if company in ('','nan'): continue
    for col in ['Company Mail','HR Mail','HR Mail2','HR Mail3','HR Mail4','HR Mail5','HR Mail6']:
        v = str(row.get(col,'') or '').strip()
        if '@' in v: add_email(company, v)

# ── Source 2: all_bcc_emails.txt ─────────────────────────────────────────
with open('all_bcc_emails.txt', encoding='utf-8') as f:
    for line in f:
        e = line.strip().rstrip(',')
        if '@' in e: add_email('', e, prefer_name=False)

# ── Source 3+4: Mail list.pdf & HR MAIL MAIN_02.pdf ─────────────────────
# Use emails-only (prefer_name=False) — PDF table merging causes wrong company
# name assignments; HR Mail List.xlsx already provides correct names.
def parse_pdf_companies(fname):
    try:
        with pdfplumber.open(fname) as pdf:
            text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
        for m in re.finditer(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', text):
            add_email('', m.group(), prefer_name=False)
    except Exception as ex:
        print(f'PDF error {fname}: {ex}')

parse_pdf_companies('Mail list.pdf')
parse_pdf_companies('HR MAIL MAIN_02.pdf')

# ── Source 6: Mail list (1).pdf — extract emails only (no company attribution
#   because PDFplumber merges table cells incorrectly for this file)
def parse_pdf_emails_only(fname):
    """Extract all emails from PDF without trusting company name column."""
    try:
        with pdfplumber.open(fname) as pdf:
            text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
    except Exception as ex:
        print(f'PDF error {fname}: {ex}')
        return
    for m in re.finditer(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', text):
        add_email('', m.group(), prefer_name=False)

parse_pdf_emails_only('Mail list (1).pdf')

# ── Source 5 (new): emails.csv ───────────────────────────────────────────
try:
    import csv
    with open('emails.csv', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            e = (row.get('email') or '').strip()
            if '@' in e: add_email('', e, prefer_name=False)
except Exception as ex:
    print(f'emails.csv error: {ex}')

# Assign name from domain if still empty, using DISPLAY_NAMES for canonical names
for domain, rec in company_emails.items():
    if not rec['name']:
        dk = domain_key(domain)
        rec['name'] = DISPLAY_NAMES.get(dk) or domain.split('.')[0].replace('-',' ').replace('_',' ').title()

# ── Source 6: VCF HR Contact Phones + Additional Emails ──────────────────
VCF_EMAILS = [
    # (company_name, email) — non-gmail VCF emails to add
    ("Deftbox Solutions",      "connect@deftboxsolutions.com"),
    ("Vagaro India",           "hrindia@vagaro.com"),
    ("Infopulse Tech",         "rudra.dave@infopulsetech.com"),
    ("Viha Digital Commerce",  "hr@vihadigitalcommerce.com"),
    ("Nexus Link Services",    "jobs@nexuslinkservices.com"),
    ("4C Consulting",          "career1@4cpl.com"),
    ("Alphae Barcode",         "hr@alphaebarcode.com"),
    ("ManekTech",              "tulsi.g@manektech.com"),
    ("Citrusbug Technolabs",   "jobs@citrusbug.co"),
    ("Trentiums Solutions",    "jugalb@trentiums.com"),
    ("Imperoit",               "ritul@imperoit.com"),
    ("Webential",              "hr@webential.com"),
    ("Trootech",               "hr@trootech.com"),
    ("Capermint Technologies", "hr@capermint.com"),
    ("Actowiz Solutions",      "krutis.actowiz@gmail.com"),  # skip gmail
    ("InfoCusp",               "inquiries@infocusp.com"),     # found via web research
    ("Anblicks",               "careers@anblicks.com"),       # found via web research
    ("Kody Technolab",         "hiring@kodytechnolab.com"),   # found via web research
    ("Innvonix",               "purav@innvonix.com"),         # found via web research
    ("iSummation Technologies","info@isummation.com"),        # found via web research
    ("Heritage Cyberworld",    "contact@heritagecyberworld.com"), # found via web research
    ("Saras-3D",               "reachus@saras-3d.com"),       # found via web research
    ("Qrioustech",             "info@qrioustech.com"),        # found via web research
    ("Acespritech",            "info@acespritech.com"),       # found via web research
    ("3rd Digital",            "sagar@3rddigital.com"),       # found via web research
    ("AnalyticsLiv",           "hr@analyticsliv.com"),        # found via web research
    ("Crest Data Systems",     "info@crestdata.ai"),          # found via web research
    ("AIpXperts",              "hr@aipxperts.com"),           # found via web research
    ("Datavruti",              "sales@datavruti.com"),        # found via web research
    ("Cygnet Infotech",        "enquiry@cygnet.one"),         # found via web research
    ("Apexon",                 "info@apexon.com"),            # found via web research
    ("Anblicks",               "marketing@anblicks.com"),     # found via web research (2nd)
    ("Biztechcs",              "career@biztechcs.com"),       # found via web research
    ("Artoonsolutions",        "info@artoonsolutions.com"),   # found via web research
    ("Concettolabs",           "career@concettolabs.com"),    # found via web research
    ("Auberginesolutions",     "hrteam@aubergine.co"),        # found via web research
    ("Alakmalak",              "hello@alakmalak.com"),        # found via web research
    ("DeepCoder.io",           "contact@deepcoder.io"),       # found via web research
    ("Datazip",                "shubham@datazip.io"),         # found via web research (Bangalore)
    ("Aionpixel",              "info@aionpixel.com"),         # found via web research (Kochi)
    ("atQor",                  "info@atqor.com"),             # found via web research
    ("Codiste",                "manager@codiste.com"),        # found via web research
    ("Creole Studios",         "jobs@creolestudios.com"),     # found via web research
    ("Crestinfosystems",       "hr@crestinfosystems.com"),    # found via web research
    ("Devitpl",                "presales@devitpl.com"),       # found via web research
    ("Definelabs",             "anup@definelabs.com"),        # found via web research (Delhi)
    ("Amnex",                  "info@amnex.com"),             # found via web research
    ("Creatpix",               "info@creatpix.com"),          # found via web research
    ("Pirimidtech",            "info@pirimidtech.com"),       # found via web research
    ("Magnus Minds",           "hello@magnusminds.net"),      # found via web research
    ("Netclues India",         "nikhilmehta@netclues.com"),   # found via web research
    ("Intuz.info",             "careers@intuz.info"),         # found via web research
    ("Latitudetechnolabs",     "sales@latitudetechnolabs.com"), # found via web research
    ("MNS Technologies",       "info@mnstechnologies.com"),   # found via web research (Rajkot)
    ("Moweb",                  "info@moweb.com"),             # found via web research
    ("Stellarmind AI",         "info@stellarmind.ai"),        # found via web research
    ("Intellytics Solutions",  "info@intellyticssolutions.com"), # found via web research
    ("Spectrics Solutions",    "info@spectricssolutions.com"),   # NEW Ahmedabad AI/ML firm
    ("Codezeros",              "hello@codezeros.com"),           # found via web research
    ("Cleardu",                "info@cleardu.com"),              # found via web research
    ("Nexgits",                "info@nexgits.com"),              # NEW Ahmedabad AI/ML firm
    ("OrcaMinds",              "info@orcaminds.in"),             # NEW Ahmedabad AI/ML firm
    ("CodesClue Technologies", "business@codesclue.com"),        # NEW Ahmedabad AI/ML firm
    ("DataDwip",               "hello@datadwip.com"),            # NEW Ahmedabad AI data firm
    ("AvenuesAI",              "ir@avenuesai.com"),              # GIFT City Gandhinagar AI/fintech
    ("Fusion Informatics",     "info@fusioninformatics.com"),    # Ahmedabad AI/ML/IoT firm
    ("Rapidops",               "hr@rapidops.com"),               # Ahmedabad AI/ML + SaaS
    ("InheritX Solutions",     "hr@inheritx.com"),               # Ahmedabad IT/AI
    ("WPWeb Infotech",         "sales@wpwebinfotech.com"),       # Ahmedabad web/mobile
    ("AddWeb Solution",        "addweb@addwebsolution.com"),     # Ahmedabad web/AI
    ("Shiv Technolabs",        "contact@shivlab.com"),           # Ahmedabad AI/Odoo/mobile
    ("Tagline Infotech",       "hr@taglineinfotech.com"),        # Ahmedabad mobile/web
    ("Brilworks Software",     "digital@brilworks.com"),         # Ahmedabad AI/cloud
    ("Sigma Solve",            "sales@sigmasolve.com"),          # Ahmedabad AI/data/cloud
    ("Techuz InfoWeb",         "sales@techuz.com"),              # Ahmedabad web/mobile
    ("Nimblechapps",           "contact@nimblechapps.com"),      # Ahmedabad mobile/web
    ("Sapphire Software Solutions", "contact@sapphiresolutions.net"), # Ahmedabad web/app
    ("SculptSoft",             "info@sculptsoft.com"),           # Ahmedabad AI/ML
    ("Argusoft",               "info@argusoft.com"),             # Gandhinagar AI/data/IoT
    ("Third Rock Techkno",     "info@thirdrocktechkno.com"),     # Ahmedabad AI/software
    ("Zignuts Technolab",      "talent@zignuts.com"),            # Gandhinagar Python/AI
    ("Gateway TechnoLabs",     "cs@gatewaytechnolabs.com"),      # Ahmedabad IT/AI (Gateway Group)
    ("Softvan Labs",           "sales@softvanlabs.com"),         # Ahmedabad ML/RPA/AI
    ("iView Labs",             "sales@iviewlabs.com"),           # Ahmedabad AI/ML/product
    ("Wappnet Systems",        "sales@wappnet.com"),             # Ahmedabad web/mobile/AI
    ("WebMob Technologies",    "career@webmobtech.com"),         # Ahmedabad AI/blockchain
    ("iFlair Web Technologies","info@iflair.com"),               # Ahmedabad web/software
    ("iFour Technolab",        "info@ifourtechnolab.com"),       # Ahmedabad .NET/AI dev
    ("KCS IT Global",          "info@kcsitglobal.com"),          # Ahmedabad ERP/IT consulting
    ("Devstree IT Services",   "info@devstree.com"),             # Ahmedabad AI/game/AR/VR
    ("Webline India",          "info@weblineindia.in"),          # Ahmedabad web/software dev
    ("Pragnakalp Techlabs",   "letstalk@pragnakalp.com"),       # Ahmedabad GenAI/NLP/Python
    ("Prismetric Technologies","biz@prismetric.com"),            # Gandhinagar InfoCity AI/mobile
    ("Infilon Technologies",  "contact@infilon.com"),           # Ahmedabad Satellite web/AI
    ("CodeEpsilon",            "info@codeepsilon.com"),          # Ahmedabad Science City Rd
    ("Digiwagon Technologies", "hello@digiwagon.com"),           # Ahmedabad SG Hwy AI/digital
    ("Tuvoc Technologies",     "hr@tuvoc.com"),                  # Ahmedabad Thaltej AI/ML/AR/VR
    ("Nevina Infotech",        "info@nevinainfotech.com"),       # Ahmedabad Ashram Road AI/ML
    ("Technostacks Infotech",  "info@technostacks.com"),         # Ahmedabad CG Road IT consulting
    ("NicheTech Solutions",    "hr@nichetech.in"),               # Ahmedabad SG Hwy AI/ML
    ("C-Metric Solutions",     "contact@c-metric.com"),          # Gandhinagar Infocity IT solutions
    ("Metizsoft Solutions",    "contact@metizsoft.com"),         # Ahmedabad Navrangpura AI/ML
]

for cname, email in VCF_EMAILS:
    add_email(cname, email)

# ── Source 7: Companies found via VCF phone numbers + web research ──────────
# These had only gmail/no email in any source file, so they can't be matched
# via add_email(). Inject them directly using their real domain (found via
# deep research) so they still appear in the output with phone/address/LinkedIn.
PHONE_ONLY_DOMAINS = ['actowizsolutions.com', 'aivhub.com', 'atharvasystem.com',
                      'evitalrx.in', 'varianceinfotech.com', 'hrbrainhub.com',
                      'isagebrum.com',
                      # AI companies found via directory research — no phone/email
                      # found yet, but address/LinkedIn/website confirmed real
                      'creolestudios.com', 'codiste.com', 'nettyfy.com', 'hiddenbrains.com',
                      'datumquest.com', 'tisindia.com',
                      'peerbits.com', 'webelight.com', 'volansys.com',
                      'smartsensesolutions.com', 'botreetechnologies.com',
                      'quixom.com']
for domain in PHONE_ONLY_DOMAINS:
    _ = company_emails[domain]  # touch to create entry with emails=[], name=''

# ── Companies with no verifiable Ahmedabad presence — drop from output ──────
# These domain keys only existed as AI keyword matches with no real DB record;
# web research found no confirmed company info, so they are removed entirely
# rather than shown with fabricated/blank details.
EXCLUDE_KEYS = {'voidspace', 'biobrain', 'mindbrain', 'psytech', 'gridanalyticsindia',
                'ddatalabs', 'codework', 'makunaiglobal', 'badaadata'}
for domain in list(company_emails.keys()):
    if domain_key(domain) in EXCLUDE_KEYS:
        del company_emails[domain]

# ═══════════════════════════════════════════════════════════════════════════
# C. BUILD FINAL ROWS  (deduplicated by domain)
# ═══════════════════════════════════════════════════════════════════════════
rows = []
for domain, rec in company_emails.items():
    emails = rec['emails']
    name   = rec['name'] or domain.split('.')[0].title()
    # Apply canonical display name override based on domain key
    dk = domain_key(domain)
    if dk in DISPLAY_NAMES:
        name = DISPLAY_NAMES[dk]
    det    = get_details(name, domain)
    rows.append({
        'Company Name': name,
        'Category':     categorize(name, domain),
        'Roles':        det.get('r',''),
        'Email 1':      emails[0] if len(emails)>0 else '',
        'Email 2':      emails[1] if len(emails)>1 else '',
        'Email 3':      emails[2] if len(emails)>2 else '',
        'Email 4':      emails[3] if len(emails)>3 else '',
        'Email 5':      emails[4] if len(emails)>4 else '',
        'Phone':        det.get('p',''),
        'Website':      det.get('w') or f'https://www.{domain}',
        'LinkedIn URL': det.get('l',''),
        'Address':      det.get('a',''),
    })

rows.sort(key=lambda x: x['Company Name'].lower())

# ── Deduplicate rows with same normalised company name ───────────────────
from collections import defaultdict as _dd
_groups = _dd(list)
for r in rows:
    nk = re.sub(r'[^a-z0-9]', '', r['Company Name'].lower())
    _groups[nk].append(r)

deduped = []
for nk, grp in _groups.items():
    if len(grp) == 1:
        deduped.append(grp[0])
        continue
    # Merge: combine all emails, pick best phone/address/linkedin
    seen_em = []
    for r in grp:
        for col in ['Email 1','Email 2','Email 3','Email 4','Email 5']:
            e = r[col]
            if e and e.lower() not in [x.lower() for x in seen_em]:
                seen_em.append(e)
    # Canonical name: longest / most spaced variant wins
    best_name = max(grp, key=lambda r: len(r['Company Name']) + r['Company Name'].count(' '))['Company Name']
    best_phone   = next((r['Phone']        for r in grp if r['Phone']),        '')
    best_addr    = next((r['Address']      for r in grp if r['Address']),      '')
    best_link    = next((r['LinkedIn URL'] for r in grp if r['LinkedIn URL']), '')
    best_website = next((r['Website']      for r in grp if r['Website']),      '')
    best_roles   = next((r['Roles']        for r in grp if r['Roles']),        '')
    best_cat     = next((r['Category']     for r in grp if r['Category'] == 'AI / ML'), grp[0]['Category'])
    deduped.append({
        'Company Name': best_name,
        'Category':     best_cat,
        'Roles':        best_roles,
        'Email 1':      seen_em[0] if len(seen_em) > 0 else '',
        'Email 2':      seen_em[1] if len(seen_em) > 1 else '',
        'Email 3':      seen_em[2] if len(seen_em) > 2 else '',
        'Email 4':      seen_em[3] if len(seen_em) > 3 else '',
        'Email 5':      seen_em[4] if len(seen_em) > 4 else '',
        'Phone':        best_phone,
        'Website':      best_website,
        'LinkedIn URL': best_link,
        'Address':      best_addr,
    })

rows = sorted(deduped, key=lambda x: x['Company Name'].lower())
print(f"Total companies after dedup: {len(rows)}")

# ═══════════════════════════════════════════════════════════════════════════
# D. EXCEL STYLES
# ═══════════════════════════════════════════════════════════════════════════
H_FILL  = PatternFill("solid", fgColor="1F3864")
H_FONT  = Font(bold=True, color="FFFFFF", size=11)
T_FILL  = PatternFill("solid", fgColor="2E75B6")
AI_FILL = PatternFill("solid", fgColor="E8F5E9")
R_FILLS = [PatternFill("solid", fgColor="DDEEFF"), PatternFill("solid", fgColor="FFFFFF")]
LK_FONT = Font(color="0563C1", underline="single", size=10)
BD_FONT = Font(bold=True, size=10)
NM_FONT = Font(size=10)
bs = Side(style='thin', color='B0C4DE')
BRD = Border(left=bs, right=bs, top=bs, bottom=bs)

COLS = [
    ('No',          5 ),('Company Name',28),('Category',   14),('Roles / Skills', 32),
    ('Email 1',     34),('Email 2',     32),('Email 3',    32),('Email 4',        32),
    ('Email 5',     32),('Phone',       20),('Website',    28),('LinkedIn URL',   34),
    ('Address',     46),
]

def write_sheet(ws, data, title=None):
    start = 1
    if title:
        ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
        tc = ws.cell(1,1,title)
        tc.fill = H_FILL
        tc.font = Font(bold=True, color="FFFFFF", size=13)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34
        ws.row_dimensions[2].height = 5
        start = 3

    for ci,(cn,cw) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = cw

    # Header
    ws.row_dimensions[start].height = 30
    for ci,(cn,_) in enumerate(COLS, 1):
        c = ws.cell(start, ci, cn)
        c.fill = H_FILL; c.font = H_FONT; c.border = BRD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    ds = start + 1
    for i, r in enumerate(data, 1):
        rn   = ds + i - 1
        isai = r['Category'] == 'AI / ML'
        fill = AI_FILL if isai else R_FILLS[i%2]
        ws.row_dimensions[rn].height = 18
        vals = [i, r['Company Name'], r['Category'], r['Roles'],
                r['Email 1'], r['Email 2'], r['Email 3'], r['Email 4'], r['Email 5'],
                r['Phone'], r['Website'], r['LinkedIn URL'], r['Address']]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(rn, ci, val)
            c.fill = fill; c.border = BRD
            if ci == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = NM_FONT
            elif ci == 13:
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.font = NM_FONT
            elif ci in (5,6,7,8,9,11,12) and val:
                c.font = LK_FONT
                c.alignment = Alignment(vertical="center")
            elif ci == 2:
                c.font = BD_FONT
                c.alignment = Alignment(vertical="center")
            elif ci == 3:
                c.font = Font(bold=True, size=10, color="1A6600" if isai else "1F3864")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.font = NM_FONT
                c.alignment = Alignment(vertical="center", wrap_text=False)

    ws.freeze_panes = ws.cell(ds, 1)
    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(COLS))}{ds+len(data)-1}"

# ═══════════════════════════════════════════════════════════════════════════
# E. BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = "All Companies"
write_sheet(ws1, rows, "IT & AI/ML Companies — FINAL MASTER LIST (All Sources + VCF HR Contacts)")

ws2 = wb.create_sheet("AI ML Companies")
ai_rows = [r for r in rows if r['Category']=='AI / ML']
write_sheet(ws2, ai_rows, "AI / ML Companies — Detailed Contact List")

wb.save('Ahmedabad_IT_AIML_FINAL_MASTER.xlsx')

ai  = len(ai_rows)
tot = len(rows)
print(f"Saved: Ahmedabad_IT_AIML_FINAL_MASTER.xlsx")
print(f"  Sheet 1 - All companies : {tot}")
print(f"  Sheet 2 - AI/ML only    : {ai}")
print(f"  IT Services             : {tot-ai}")

# Phone coverage stats
phones = sum(1 for r in rows if r['Phone'])
addrs  = sum(1 for r in rows if r['Address'])
links  = sum(1 for r in rows if r['LinkedIn URL'])
print(f"\nCoverage:")
print(f"  Phone     : {phones}/{tot} ({100*phones//tot}%)")
print(f"  Address   : {addrs}/{tot}  ({100*addrs//tot}%)")
print(f"  LinkedIn  : {links}/{tot}  ({100*links//tot}%)")

# Quality report
print("\n" + "="*60 + "\nDATA QUALITY REPORT\n" + "="*60)
bad = []
for r in rows:
    for col in ['Email 1','Email 2','Email 3','Email 4','Email 5']:
        e = r.get(col,'').strip()
        if e and not EMAIL_RE.match(e):
            bad.append(f"  INVALID | {r['Company Name'][:28]:28s} | {e}")
        if e and get_domain(e).endswith('.incom'):
            bad.append(f"  TYPO    | {r['Company Name'][:28]:28s} | {e}")
if bad:
    print(f"[!] {len(bad)} issue(s):"); [print(b) for b in bad]
else:
    print("[OK] All email formats valid.")

# Show companies with phones (from VCF)
print("\nSample companies with HR phone numbers:")
ph_rows = [(r['Company Name'], r['Phone']) for r in rows if r['Phone']]
for cn, ph in sorted(ph_rows)[:30]:
    print(f"  {cn[:35]:35s}  {ph}")
print(f"  ... and {len(ph_rows)-30} more" if len(ph_rows) > 30 else "")
