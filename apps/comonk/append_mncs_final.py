"""
Append the 179 MNCs from 'All Companies' into MNC_AI_Targets (same file).
- Keeps the enriched 30 curated MNCs first
- Adds the rest: genuine Gujarat MNCs first, other-state MNCs after (kept, not removed)
- Each row: existing emails + verified + standard functional inboxes (up to 15)
Only touches MNC_AI_Targets.
"""
import openpyxl, re, unicodedata, zipfile, time
from openpyxl.styles import Font, PatternFill, Alignment

FINAL = "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
N_EMAIL = 15
STD = ["careers","hr","recruitment","jobs","talent","campus","freshers","hrhelpdesk","info","hiring"]

def norm(s): return re.sub(r'[^a-z0-9]', '', unicodedata.normalize('NFKD', str(s).lower()))
def clean_domain(url):
    if not url: return None
    u = re.sub(r'^https?://','',str(url).strip().lower()); u=re.sub(r'^www\.','',u)
    u = u.split('/')[0].split('?')[0].strip()
    if '.' not in u or len(u)<4: return None
    p=u.split('.')
    if len(p)>=3 and p[-2] in ("co","com","org","net","gov","ac") and p[-1] in ("in","uk"):
        return ".".join(p[-3:])
    return ".".join(p[-2:])
def edom(e): return str(e).split("@")[1].lower() if e and "@" in str(e) else None

VERIFIED = {
    "tcs":["careers@tcs.com","global.hr@tcs.com","ilp.support@tcs.com","xplore.support@tcs.com"],
    "infosys":["careers@infosys.com","infy_rec_helpdesk@infosys.com","askus@infosys.com"],
    "wipro":["careers@wipro.com","helpdesk.recruitment@wipro.com","ombuds.person@wipro.com"],
    "hcl":["careers@hcltech.com","hrservices@hcltech.com"],
    "capgemini":["cg_interview_helpdesk.in@capgemini.com","talentacquisitioncompliance.in@capgemini.com"],
    "accenture":["candidate.queries@accenture.com","careers@accenture.com"],
    "ibm":["askhr@in.ibm.com","careers@ibm.com"],
    "cognizant":["tagcompliance2@cognizant.com","careers@cognizant.com"],
    "bosch":["connect@in.bosch.com","careers@in.bosch.com"],
}
NAMED = {
    "kpit":["chandrika.subravati@kpit.com","samuel.preetham@kpit.com","abhishek.joshi@kpit.com",
            "harsh.diwakirti@kpit.com","archana.muttagi@kpit.com","shantanu.waikar@kpit.com",
            "samir.sawant@kpit.com","dattaprasad.desai@kpit.com"],
    "ltts":["pradeepthi.rathod@ltts.com","shruti.mudaliar@ltts.com","rajni.patil@ltts.com"],
    "hexaware":["prachih@hexaware.com","faheemk@hexaware.com"],
    "capgemini":["tanuja.bhalekar@capgemini.com","rucha.deshpande@capgemini.com"],
}
GUJ_NAME=("gift","ahmedabad","gandhinagar","gujarat","vadodara","baroda","koba")
GUJ_CO=("adani","torrent","zydus","cadila","intas","nirma","arvind","amul","gcmmf","gspc","gsfc",
        "gnfc","gujaratgas","welspun","meghmani","aiaengineering","elecon","atul","claris","astral",
        "alembic","dishman","eris","gufic","tcs","tataconsultancy","infosys","wipro","hcl",
        "techmahindra","capgemini","accenture","ibm","cognizant","deloitte","ernstyoung","pwc",
        "pricewaterhouse","kpmg","oracle","sap","bosch","siemens","hexaware","ltts","apexon",
        "infostretch","einfochips","persistent","mphasis","atos","birlasoft","exlservice","fujitsu",
        "argusoft","avenues","inubia","icreate","gujaratinformatics")

def stable_load(path, tries=30):
    for _ in range(tries):
        try:
            with zipfile.ZipFile(path): pass
            return openpyxl.load_workbook(path)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read")

def enrich(company, cur, domain):
    out=list(cur); seen={e.lower() for e in out}; nk=norm(company)
    def push(e):
        if e and e.lower() not in seen and len(out)<N_EMAIL: out.append(e); seen.add(e.lower())
    for k,l in VERIFIED.items():
        if k in nk:
            for e in l: push(e)
    for k,l in NAMED.items():
        if k in nk:
            for e in l: push(e)
    if domain:
        for p in STD: push(f"{p}@{domain}")
    return out

wb = stable_load(FINAL)
mt = wb["MNC_AI_Targets"]

# read enriched 30 (col layout: 1=#,2=Co,3=City,4=Roles,5..19=Email1-15,20=Phone,21=Web,22=LI)
existing=[]
for r in range(2, mt.max_row+1):
    co=mt.cell(r,2).value
    if not co: continue
    emails=[mt.cell(r,c).value for c in range(5,5+N_EMAIL) if mt.cell(r,c).value]
    existing.append({"company":str(co).strip(),"city":mt.cell(r,3).value,"roles":mt.cell(r,4).value,
        "emails":[str(e).strip() for e in emails if e and "@" in str(e)],
        "phone":mt.cell(r,5+N_EMAIL).value,"website":mt.cell(r,6+N_EMAIL).value,"linkedin":mt.cell(r,7+N_EMAIL).value})
existing_norms={norm(e["company"]) for e in existing}
print(f"Existing enriched MNCs: {len(existing)}")

# read All Companies MNCs
ac = wb["All Companies"]
hdr=[ac.cell(3,c).value for c in range(1,ac.max_column+1)]
hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
co_c=hm.get("company name",2); cat_c=hm.get("category",3); role_c=hm.get("roles / skills",4)
ph_c=hm.get("phone",10); web_c=hm.get("website",11); city_c=hm.get("city",14)
em_cols=[hm.get(f"email {i}") for i in range(1,7)]; em_cols=[c for c in em_cols if c]

new_guj=[]; new_oth=[]
for r in range(4, ac.max_row+1):
    cat=str(ac.cell(r,cat_c).value or "")
    if "mnc" not in cat.lower(): continue
    nm=ac.cell(r,co_c).value
    if not nm or norm(nm) in existing_norms: continue
    existing_norms.add(norm(nm))
    cur=[str(ac.cell(r,c).value).strip() for c in em_cols if ac.cell(r,c).value and "@" in str(ac.cell(r,c).value)]
    dom=edom(cur[0]) if cur else clean_domain(ac.cell(r,web_c).value)
    emails=enrich(nm,cur,dom)
    nml=str(nm).lower(); nk=norm(nm)
    is_guj=any(k in nml for k in GUJ_NAME) or any(k in nk for k in GUJ_CO)
    rec={"company":str(nm).strip(),"city":ac.cell(r,city_c).value or "Ahmedabad",
         "roles":ac.cell(r,role_c).value or "AI Engineer, ML Engineer, Data Scientist",
         "emails":emails,"phone":ac.cell(r,ph_c).value,"website":ac.cell(r,web_c).value,"linkedin":None}
    (new_guj if is_guj else new_oth).append(rec)

new_guj.sort(key=lambda x:x["company"].lower())
new_oth.sort(key=lambda x:x["company"].lower())
print(f"New MNCs: {len(new_guj)} Gujarat + {len(new_oth)} other-state (kept)")

# rewrite sheet
del wb["MNC_AI_Targets"]
sh=wb.create_sheet("MNC_AI_Targets")
DARK=PatternFill("solid",fgColor="1a1a2e"); HDRF=Font(bold=True,color="FFFFFF",size=10)
GJ=PatternFill("solid",fgColor="E2EFDA")  # light green tint marker via city? keep simple
hdr=["#","Company","City","AI Roles (Apply For)"]+[f"HR Email {i}" for i in range(1,N_EMAIL+1)]+["Office Phone","Website","LinkedIn"]
for ci,h in enumerate(hdr,1):
    sh.cell(1,ci,h).fill=DARK; sh.cell(1,ci).font=HDRF; sh.cell(1,ci).alignment=Alignment(horizontal="center")
sh.column_dimensions['B'].width=32
def wr(r,idx,rec):
    sh.cell(r,1,idx).font=Font(size=9)
    sh.cell(r,2,rec["company"]).font=Font(size=9,bold=True)
    sh.cell(r,3,rec["city"]).font=Font(size=9)
    sh.cell(r,4,rec["roles"]).font=Font(size=9)
    for i,em in enumerate(rec["emails"][:N_EMAIL]): sh.cell(r,5+i,em).font=Font(color="1a56db",size=9)
    sh.cell(r,5+N_EMAIL,rec["phone"]).font=Font(size=9)
    sh.cell(r,6+N_EMAIL,rec["website"]).font=Font(size=9)
    sh.cell(r,7+N_EMAIL,rec["linkedin"]).font=Font(size=9)
r=2; idx=1
for e in existing: wr(r,idx,e); r+=1; idx+=1
for a in new_guj:  wr(r,idx,a); r+=1; idx+=1
for a in new_oth:  wr(r,idx,a); r+=1; idx+=1
total_contacts=sum(min(len(x["emails"]),N_EMAIL) for x in existing+new_guj+new_oth)
print(f"MNC_AI_Targets: {idx-1} MNCs, {total_contacts} email contacts")
print(f"  Order: {len(existing)} curated + {len(new_guj)} Gujarat + {len(new_oth)} other-state")

for _ in range(20):
    try: wb.save(FINAL); break
    except Exception: time.sleep(2)
print(f"Saved -> {FINAL}")
