import openpyxl
from collections import Counter

wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
ws = wb["COMPLETE_MASTER"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CITY = hmap.get("city", 3)
COL_PRI  = hmap.get("priority", 17)

cities = Counter()
for r in range(2, ws.max_row + 1):
    c = ws.cell(r, COL_CITY).value
    cities[str(c)] += 1

print("City distribution:")
for city, cnt in cities.most_common(40):
    print(f"  {cnt:>5}  {city}")

print(f"\nTotal distinct cities: {len(cities)}")

pris = Counter()
for r in range(2, ws.max_row + 1):
    p = ws.cell(r, COL_PRI).value
    pris[str(p)] += 1
print("\nPriority distribution:")
for p, cnt in pris.most_common(20):
    print(f"  {cnt:>5}  {p}")
