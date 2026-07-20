"""
Remove faulty data from all company sheets (keeps everything real).
Clears: w3.org emails, placeholder emails (example.com/domain.com/yourdomain/johnsmith/name@),
image artifacts (@2x/@3x/.webp/.png), fake phones (7326059369, 9240466672), 'N/A' phones.
Then compacts remaining emails leftward. Additive-safe: only clears bad cells.
"""
import openpyxl, re, zipfile, time

F="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
def stable_load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read (file open in Excel?)")

BAD_EMAIL=("w3.org","example.com","@domain.com","yourdomain","johnsmith@","name@example",
           "@2x",".webp",".png","@3x","sentry.io","wixpress.com","noreply@","no-reply@",
           "email@","test@example","user@","abc@","xyz@")
BAD_PHONE_DIGITS={"7326059369","917326059369","09326059369","9240466672","919240466672"}

wb=stable_load(F)
tot_em=tot_ph=0
for SHEET in wb.sheetnames:
    ws=wb[SHEET]
    if SHEET=="LinkedIn Profiles": continue
    # find header row
    hrow=None
    for rr in (1,2,3):
        vals=[str(ws.cell(rr,c).value or "").strip().lower() for c in range(1,25)]
        if "company name" in vals or "company" in vals: hrow=rr; break
    if not hrow: continue
    hdr=[ws.cell(hrow,c).value for c in range(1,ws.max_column+1)]
    hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
    PH=hm.get("phone") or hm.get("office phone")
    EM=[hm.get(f"email {i}") for i in range(1,7)]+[hm.get(f"hr email {i}") for i in range(1,16)]
    EM=[c for c in EM if c]
    s_em=s_ph=0
    for r in range(hrow+1, ws.max_row+1):
        # emails: clear bad, then compact
        vals=[]
        for c in EM:
            v=ws.cell(r,c).value
            if v and "@" in str(v):
                lo=str(v).lower().strip()
                if any(b in lo for b in BAD_EMAIL): s_em+=1; continue
                vals.append(str(v).strip())
            elif v and str(v).strip():
                # non-email junk in an email cell (e.g. artifact)
                s_em+=1
        # rewrite compacted
        for i,c in enumerate(EM):
            ws.cell(r,c).value = vals[i] if i < len(vals) else None
        # phone
        if PH:
            p=ws.cell(r,PH).value
            if p:
                praw=str(p).strip(); pd=re.sub(r'\D','',praw)
                if pd in BAD_PHONE_DIGITS or praw.upper() in ("N/A","NA","-","--","NULL","NONE"):
                    ws.cell(r,PH).value=None; s_ph+=1
    print(f"  {SHEET}: removed {s_em} bad emails, {s_ph} bad phones")
    tot_em+=s_em; tot_ph+=s_ph

for _ in range(20):
    try: wb.save(F); break
    except Exception: time.sleep(2)
print(f"\nFIX COMPLETE: removed {tot_em} fake/placeholder emails, {tot_ph} fake phones. Saved.")
