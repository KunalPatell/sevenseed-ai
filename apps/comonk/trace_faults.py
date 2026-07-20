import openpyxl
from collections import Counter
wb=openpyxl.load_workbook("Ahmedabad_IT_AIML_FINAL_MASTER.xlsx", read_only=True, data_only=True)
for SHEET in ["IT Companies","All Companies","IT MNC"]:
    if SHEET not in wb.sheetnames: continue
    ws=wb[SHEET]
    hdr=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
    hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
    CO=hm.get("company name",2); SRC=hm.get("source",16); PH=hm.get("phone",10)
    EM=[hm.get(f"email {i}") for i in range(1,7)]; EM=[c for c in EM if c]
    w3_src=Counter(); fake_src=Counter(); ex_src=Counter()
    for row in ws.iter_rows(min_row=4, values_only=True):
        def g(c): return row[c-1] if c and len(row)>=c else None
        src=str(g(SRC) or "(blank)")
        if any(g(c) and "w3.org" in str(g(c)) for c in EM): w3_src[src]+=1
        if any(g(c) and "example.com" in str(g(c)) for c in EM): ex_src[src]+=1
        ph=g(PH)
        if ph and "7326059369" in str(ph).replace(" ",""): fake_src[src]+=1
    print("==",SHEET,"==")
    print("  w3.org emails by Source:", dict(w3_src.most_common(4)))
    print("  example.com emails by Source:", dict(ex_src.most_common(4)))
    print("  fake-phone rows by Source:", dict(fake_src.most_common(4)))
    print()
