"""
Safe additive merge: recover scraper progress into the user's edited file.
BASE  = Ahmedabad_IT_AIML_FINAL_MASTER.xlsx      (user's manual edits - authoritative)
SRC   = Ahmedabad_IT_AIML_FINAL_MASTER_SCRAPED_1653.xlsx  (fuller scraper data)
Rule: fill ONLY empty cells in BASE from SRC. Never overwrite anything the user typed.
Sheet: 'All Companies'. Matched by company name (fallback row).
"""
import openpyxl, re, unicodedata, zipfile, time
from openpyxl.styles import Font

BASE="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
SRC ="Ahmedabad_IT_AIML_FINAL_MASTER_SCRAPED_1653.xlsx"
SHEET="All Companies"

def norm(s): return re.sub(r'[^a-z0-9]','',unicodedata.normalize('NFKD',str(s).lower()))
def stable_load(p,ro=False,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, read_only=ro, data_only=False)
        except Exception: time.sleep(2)
    raise RuntimeError(f"cannot read {p}")

wb=stable_load(BASE); ws=wb[SHEET]
hdr=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
CO=hm.get("company name",2); PH=hm.get("phone",10); WEB=hm.get("website",11)
LI=hm.get("linkedin",12); ADDR=hm.get("address",13); SRCc=hm.get("source",16)
EM=[hm.get(f"email {i}") for i in range(1,7)]; EM=[c for c in EM if c]

# build source map
src_wb=stable_load(SRC, ro=True); sws=src_wb[SHEET]
smap={}
for row in sws.iter_rows(min_row=4, values_only=True):
    if len(row)<max(CO,PH,WEB): continue
    nm=row[CO-1]
    if not nm: continue
    k=norm(nm)
    emails=[row[c-1] for c in EM if len(row)>=c and row[c-1] and "@" in str(row[c-1])]
    smap.setdefault(k, {
        "phone":row[PH-1] if len(row)>=PH else None,
        "web":row[WEB-1] if len(row)>=WEB else None,
        "addr":row[ADDR-1] if len(row)>=ADDR else None,
        "li":row[LI-1] if len(row)>=LI else None,
        "emails":emails,
        "src":row[SRCc-1] if len(row)>=SRCc else None,
    })
src_wb.close()
print(f"Source companies indexed: {len(smap)}")

ph=web=addr=li=em=0
for r in range(4, ws.max_row+1):
    nm=ws.cell(r,CO).value
    if not nm: continue
    d=smap.get(norm(nm))
    if not d: continue
    if d["phone"] and not ws.cell(r,PH).value: ws.cell(r,PH).value=d["phone"]; ph+=1
    if d["web"] and not ws.cell(r,WEB).value: ws.cell(r,WEB).value=d["web"]; web+=1
    if d["addr"] and ADDR and not ws.cell(r,ADDR).value: ws.cell(r,ADDR).value=d["addr"]; addr+=1
    if d["li"] and LI and not ws.cell(r,LI).value: ws.cell(r,LI).value=d["li"]; li+=1
    if d["emails"]:
        cur={str(ws.cell(r,c).value or '').lower() for c in EM}
        free=[c for c in EM if not ws.cell(r,c).value]
        for e in d["emails"]:
            if str(e).lower() not in cur and free:
                slot=free.pop(0); ws.cell(r,slot).value=e; ws.cell(r,slot).font=Font(color="1a56db",size=9); cur.add(str(e).lower()); em+=1
    if (d["src"] and not ws.cell(r,SRCc).value): ws.cell(r,SRCc).value=d["src"]

for _ in range(20):
    try: wb.save(BASE); break
    except Exception: time.sleep(2)

# final counts
v=stable_load(BASE, ro=True)[SHEET]
tph=tweb=0
for row in v.iter_rows(min_row=4, values_only=True):
    if len(row)>=WEB and row[PH-1]: tph+=1
    if len(row)>=WEB and row[WEB-1]: tweb+=1
print(f"MERGE COMPLETE (additive, user edits preserved):")
print(f"  filled from scraper -> +{ph} phones, +{web} websites, +{addr} addresses, +{li} linkedin, +{em} emails")
print(f"  All Companies now: phones={tph}, websites={tweb}")
