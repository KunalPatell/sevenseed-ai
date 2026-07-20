"""
IT Companies — Deep Research Batch 2
From: Clutch IT-services, BuiltIn software, TopAppDevelopmentCompanies,
GoodFirms, your-space, aiassistica, Failory, Tracxn, StartupBlink.
Same deep-research approach as AI/ML, now for IT/software/SaaS firms.
"""

import openpyxl, re, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment

# (Company, Category, Roles, Website, City, Source)
IT_COMPANIES_2 = [
    # ── Clutch IT services / developers ──────────────────────────────────
    ("7Span",                   "IT Services","Product Engineering, SaaS, Web, Mobile Dev",        "7span.com",              "Ahmedabad", "Clutch"),
    ("PrimeQA Solutions",       "IT Services","QA, Software Testing, Automation",                 "primeqasolutions.com",   "Ahmedabad", "Clutch"),
    ("Continuum Digital",       "IT Services","Digital Transform, Web, Cloud Dev",                "continuumdigital.in",    "Ahmedabad", "Clutch"),
    ("Evrig Solutions",         "IT Services","eCommerce, Magento, Web Dev",                      "evrig.com",              "Ahmedabad", "Clutch"),
    ("Synoptek",                "IT Services","Managed IT, Cloud, Digital Transform",             "synoptek.com",           "Ahmedabad", "Clutch"),
    ("SharpQuest",              "IT Services","Custom Software, Web, Mobile Dev",                 "sharpquest.com",         "Ahmedabad", "Clutch"),
    ("Kryptoninc",              "IT Services","Web Dev, Mobile App, Custom Software",             "kryptoninc.com",         "Ahmedabad", "Clutch"),
    ("Profusion Systems",       "IT Services","Software Dev, Web, IT Solutions",                 "profusionsystems.com",   "Ahmedabad", "Clutch"),
    ("Logical Wings Infoweb",   "IT Services","Web Dev, Mobile App, Software",                   "logicalwings.com",       "Ahmedabad", "Clutch"),
    ("TRooInbound",             "IT Services","HubSpot, RevOps, Marketing Tech",                 "trooinbound.com",        "Ahmedabad", "Clutch"),
    # ── your-space / aiassistica ─────────────────────────────────────────
    ("Satva Solutions",         "IT Services","Web Dev, .NET, Cloud, FinTech Dev",               "satvasolutions.com",     "Ahmedabad", "your-space"),
    ("Innoventix Solutions",    "IT Services","Web Dev, Mobile App, Custom Software",            "innoventix.in",          "Ahmedabad", "your-space"),
    ("Axone Infotech",          "IT Services","Web Dev, Mobile App, Software",                   "axoneinfotech.com",      "Ahmedabad", "aiassistica"),
    ("Intelligent IT Hub",      "IT Services","Web Dev, Software, IT Solutions",                 "intelligentithub.com",   "Ahmedabad", "aiassistica"),
    ("Silicon IT Hub",          "IT Services","Web Dev, Mobile App, Software Dev",               "siliconithub.com",       "Ahmedabad", "aiassistica"),
    ("Shreeji Software",        "IT Services","Custom Software, Web, ERP",                       "shreejisoftware.com",    "Ahmedabad", "aiassistica"),
    ("Traffic Tail",            "IT Services","Digital Marketing, Web Dev, IT",                  "traffictail.com",        "Ahmedabad", "aiassistica"),
    # ── BuiltIn software companies ───────────────────────────────────────
    ("Cin7",                    "IT Services","Inventory SaaS, Cloud Software",                  "cin7.com",               "Ahmedabad", "BuiltIn"),
    ("SpeedBot",                "AI / ML",    "Algo Trading AI, FinTech, ML",                    "speedbot.tech",          "Ahmedabad", "BuiltIn"),
    ("Zybra",                   "IT Services","Accounting SaaS, FinTech, Cloud",                 "zybra.in",               "Ahmedabad", "BuiltIn"),
    ("Lendingkart",             "AI / ML",    "FinTech, Lending AI, Credit ML",                  "lendingkart.com",        "Ahmedabad", "BuiltIn"),
    ("Petpooja",                "IT Services","Restaurant SaaS, POS Software, Cloud",            "petpooja.com",           "Ahmedabad", "BuiltIn"),
    ("NCrypted Technologies",   "IT Services","Custom Software, SaaS, Web, Mobile",              "ncrypted.com",           "Ahmedabad", "Web-Dir"),
    ("Zybra Accounting",        "IT Services","Accounting Software, FinTech SaaS",               "zybra.in",               "Ahmedabad", "BuiltIn"),
    # ── TopAppDevelopmentCompanies ───────────────────────────────────────
    ("iMOBDEV Technologies",    "IT Services","Mobile App, Web Dev, Game Dev",                   "imobdevtech.com",        "Ahmedabad", "TopAppDev"),
    ("Prolitus Technologies",   "IT Services","Blockchain, ERP, Web Dev",                        "prolitus.com",           "Ahmedabad", "TopAppDev"),
    ("Endive Software",         "IT Services","Web Dev, Mobile App, Custom Software",            "endivesoftware.com",     "Ahmedabad", "TopAppDev"),
    ("Appslure WebSolution",    "IT Services","Mobile App, Web Dev",                            "appslure.com",           "Ahmedabad", "TopAppDev"),
    ("Workholics Infocorp",     "IT Services","Web Dev, Mobile App, Software",                   "workholics.in",          "Ahmedabad", "TopAppDev"),
    ("F5 Buddy",                "IT Services","Web Dev, Mobile App, eCommerce",                  "f5buddy.com",            "Ahmedabad", "TopAppDev"),
    ("Mango IT Solutions",      "IT Services","Web Dev, Mobile App, Custom Software",            "mangoitsolutions.com",   "Ahmedabad", "TopAppDev"),
    ("Consagous Technologies",  "IT Services","Mobile App, Web Dev, Blockchain",                "consagous.co",           "Ahmedabad", "TopAppDev"),
    ("Hyperlink InfoSystem",    "IT Services","Mobile App, Web Dev, AI Integration",            "hyperlinkinfosystem.com","Ahmedabad", "TopAppDev"),
    # ── Failory / StartupBlink / Tracxn — Ahmedabad tech startups ────────
    ("Ignosis",                 "AI / ML",    "FinTech, Account Aggregator AI, Data",            "ignosis.ai",             "Ahmedabad", "Failory"),
    ("Amnex Infotechnologies",  "AI / ML",    "Smart City AI, IoT, GIS, Data Analytics",         "amnex.com",              "Ahmedabad", "Failory"),
    ("Shipturtle",              "IT Services","Marketplace SaaS, eCommerce, Shipping",          "shipturtle.com",         "Ahmedabad", "BuiltIn"),
    ("Gamerji",                 "IT Services","Esports Platform, Gaming Tech, SaaS",             "gamerji.com",            "Ahmedabad", "BuiltIn"),
    ("EduFund",                 "AI / ML",    "FinTech, EdTech, Investment AI",                  "edufund.in",             "Ahmedabad", "BuiltIn"),
    ("Matter",                  "IT Services","EV Tech, Energy Software, IoT",                   "matter.in",              "Ahmedabad", "Failory"),
    ("BluSmart Mobility",       "IT Services","Mobility Tech, EV Fleet Software",                "blusmart.com",           "Ahmedabad", "Failory"),
    ("Medkart Pharmacy",        "IT Services","HealthTech, Pharmacy Software",                   "medkart.in",             "Ahmedabad", "BuiltIn"),
    ("Bummer",                  "IT Services","D2C Tech, eCommerce Platform",                    "bummer.in",              "Ahmedabad", "BuiltIn"),
    # ── Misc verified Ahmedabad IT firms ─────────────────────────────────
    ("Rishabh Software",        "IT Services","Custom Software, Cloud, SAP, Enterprise",         "rishabhsoft.com",        "Ahmedabad", "Web-Dir"),
    ("Grey Chain Design",       "AI / ML",    "AI, Mobile App, Generative AI",                   "greychaindesign.com",    "Ahmedabad", "Web-Dir"),
    ("Webby Central",           "IT Services","Web Dev, Digital Marketing",                      "webbycentral.com",       "Ahmedabad", "Web-Dir"),
]

def norm(s):
    if not s: return ""
    s = unicodedata.normalize('NFKD', str(s).lower())
    return re.sub(r'[^a-z0-9]', '', s)

def main():
    wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
    ws = wb["COMPLETE_MASTER"]

    existing = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 2).value
        if n: existing[norm(n)] = r

    fill_new = PatternFill("solid", fgColor="E6F0FF")
    added = updated = 0

    for company, category, roles, website_hint, city, source in IT_COMPANIES_2:
        key = norm(company)
        if not key: continue
        website = ("https://www." + website_hint) if not website_hint.startswith("http") else website_hint
        if key in existing:
            row = existing[key]
            if not ws.cell(row, 4).value: ws.cell(row, 4).value = category
            if not ws.cell(row, 5).value: ws.cell(row, 5).value = roles
            if not ws.cell(row, 13).value: ws.cell(row, 13).value = website
            cur = ws.cell(row, 18).value or ""
            if source not in cur:
                ws.cell(row, 18).value = (cur + ", " + source).strip(", ")
            updated += 1
        else:
            r = ws.max_row + 1
            ws.cell(r, 1, r - 1)
            ws.cell(r, 2, company); ws.cell(r, 3, city)
            ws.cell(r, 4, category); ws.cell(r, 5, roles)
            ws.cell(r, 13, website); ws.cell(r, 18, source)
            for c in range(1, 19):
                ws.cell(r, c).fill = fill_new
                ws.cell(r, c).alignment = Alignment(vertical="center")
            ws.cell(r, 2).font = Font(bold=True, color="003366")
            ws.row_dimensions[r].height = 18
            existing[key] = r
            added += 1

    for i, r in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(r, 1, i)
    total = ws.max_row - 1

    # Rebuild AI_ML_ONLY
    ws2 = wb["AI_ML_ONLY"]
    for r in range(ws2.max_row, 1, -1):
        ws2.delete_rows(r)
    ai_kw = ['ai','ml','data','machine','deep','nlp','computer vision','llm','genai',
             'analytics','automation','intelligence','learning','aiops']
    ai_row = 2
    fa = PatternFill("solid", fgColor="F5F0FF"); fb = PatternFill("solid", fgColor="FFFFFF")
    max_col = ws.max_column
    for r in range(2, ws.max_row + 1):
        cat  = str(ws.cell(r, 4).value or "").lower()
        role = str(ws.cell(r, 5).value or "").lower()
        if any(k in cat or k in role for k in ai_kw):
            fill = fa if ai_row % 2 == 0 else fb
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                cell2 = ws2.cell(ai_row, c, v)
                cell2.fill = fill
                cell2.alignment = Alignment(vertical="center")
                if c in (6,7,8,9,10,11) and v:
                    cell2.font = Font(color="1a56db", size=9)
            ws2.cell(ai_row, 1, ai_row - 1)
            ws2.row_dimensions[ai_row].height = 18
            ai_row += 1
    ai_count = ws2.max_row - 1

    wb["STATS"].cell(3, 2, total)
    wb["STATS"].cell(6, 2, ai_count)
    wb.save("COMONK_TRUE_MASTER.xlsx")
    print(f"  NEW IT companies: +{added}")
    print(f"  Existing updated: {updated}")
    print(f"  Total companies:  {total}")
    print(f"  AI/ML companies:  {ai_count}")
    print(f"  IT companies:     {total - ai_count}")

if __name__ == "__main__":
    main()
