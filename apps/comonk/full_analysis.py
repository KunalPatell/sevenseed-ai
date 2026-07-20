import openpyxl, re, zipfile, time, unicodedata
from collections import Counter

F="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
def stable_load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read")
def norm(s): return re.sub(r'[^a-z0-9]','',unicodedata.normalize('NFKD',str(s).lower()))
JUNK={"","n/a","na","-","--","none","null","nil","tbd","."}
def hv(v): return v is not None and str(v).strip().lower() not in JUNK
def ismail(v): return v is not None and re.match(r'^[^@\s]+@[^@\s]+\.[A-Za-z]{2,}$',str(v).strip()) is not None

wb=stable_load(F)
def sheet_meta(ws):
    hrow=None
    for rr in (1,2,3,4):
        vals=[str(ws.cell(rr,c).value or "").strip().lower() for c in range(1,26)]
        if "company name" in vals or "company" in vals: hrow=rr; break
    hdr=[ws.cell(hrow,c).value for c in range(1,ws.max_column+1)] if hrow else []
    hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
    return hrow,hm

SHEETS=["All Companies","IT Companies","IT MNC","AI ML Companies","MNC_AI_Targets"]
data={}
for S in SHEETS:
    if S not in wb.sheetnames: continue
    ws=wb[S]; hrow,hm=sheet_meta(ws)
    CO=hm.get("company name") or hm.get("company"); PH=hm.get("phone") or hm.get("office phone")
    WEB=hm.get("website"); ADDR=hm.get("address"); CAT=hm.get("category"); CITY=hm.get("city"); PRI=hm.get("priority")
    EM=[hm.get(f"email {i}") for i in range(1,7)]+[hm.get(f"hr email {i}") for i in range(1,16)]; EM=[c for c in EM if c]
    rows=[]
    for row in ws.iter_rows(min_row=(hrow or 1)+1, values_only=True):
        def g(c): return row[c-1] if c and len(row)>=c else None
        co=g(CO)
        if not hv(co): continue
        rows.append({"co":str(co).strip(),"nk":norm(co),
            "email":any(ismail(g(c)) for c in EM),"nemails":sum(1 for c in EM if ismail(g(c))),
            "phone":hv(g(PH)),"web":hv(g(WEB)),"addr":hv(g(ADDR)) if ADDR else None,
            "cat":str(g(CAT)).strip() if CAT and hv(g(CAT)) else "",
            "city":str(g(CITY)).strip() if CITY and hv(g(CITY)) else "",
            "pri":str(g(PRI)).strip() if PRI and hv(g(PRI)) else ""})
    data[S]=rows

print("="*68)
print("  FULL ANALYSIS — Ahmedabad_IT_AIML_FINAL_MASTER.xlsx")
print("="*68)

# 1. Unique companies overall + overlaps
allkeys={S:set(r["nk"] for r in data[S]) for S in data}
union=set().union(*allkeys.values())
print(f"\n[1] SIZE & UNIQUENESS")
for S in data: print(f"    {S:<18}: {len(data[S]):>5} rows, {len(allkeys[S]):>5} unique names")
print(f"    {'TOTAL UNIQUE':<18}: {len(union):>5} distinct companies across whole file")
# overlap All vs IT
if "All Companies" in allkeys and "IT Companies" in allkeys:
    ov=allkeys["All Companies"]&allkeys["IT Companies"]
    print(f"    Overlap All∩IT Companies: {len(ov)} (IT Companies is {len(ov)*100//max(len(allkeys['IT Companies']),1)}% inside All Companies)")

# 2. Coverage (All Companies = main)
print(f"\n[2] COVERAGE — All Companies ({len(data['All Companies'])} cos)")
A=data["All Companies"]; n=len(A)
for lbl,key in [("Email","email"),("Phone","phone"),("Website","web"),("Address","addr")]:
    c=sum(1 for r in A if r[key]); print(f"    {lbl:<9}: {c:>5} ({c*100//n}%)")
comp=sum(1 for r in A if r["email"] and r["phone"] and r["web"] and r["addr"])
print(f"    COMPLETE : {comp:>5} ({comp*100//n}%)  [email+phone+web+addr]")

# 3. Emails-per-company distribution (All Companies)
print(f"\n[3] EMAILS PER COMPANY — All Companies")
dist=Counter(min(r["nemails"],6) for r in A)
for k in range(0,7):
    lbl=f"{k}+" if k==6 else str(k)
    print(f"    {lbl} emails: {dist.get(k,0):>5}")

# 4. Category breakdown (All Companies)
print(f"\n[4] CATEGORY — All Companies (top 12)")
cats=Counter(r["cat"] or "(none)" for r in A)
for c,ct in cats.most_common(12): print(f"    {ct:>5}  {c[:40]}")

# 5. City breakdown
print(f"\n[5] CITY — All Companies")
cities=Counter(r["city"] or "(none)" for r in A)
for c,ct in cities.most_common(8): print(f"    {ct:>5}  {c[:30]}")

# 6. Priority
print(f"\n[6] PRIORITY — All Companies")
pri=Counter(r["pri"] or "(none)" for r in A)
for c,ct in pri.most_common(8): print(f"    {ct:>5}  {c[:45]}")

print("\n"+"="*68)
