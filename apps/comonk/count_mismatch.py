import openpyxl, re, zipfile, time, unicodedata
def load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read "+p)
def norm(s): return re.sub(r'[^a-z0-9]','',unicodedata.normalize('NFKD',str(s).lower()))
def dom(url):
    if not url: return None
    u=re.sub(r'^https?://','',str(url).strip().lower()); u=re.sub(r'^www\.','',u)
    return u.split('/')[0].split('?')[0] if '.' in u else None
GENERIC={"gmail.com","yahoo.com","yahoo.co.in","hotmail.com","outlook.com","rediffmail.com"}

wb=load("Ahmedabad_IT_AIML_FINAL_MASTER.xlsx")
ws=wb["All Companies"]; hrow=3
hdr=[ws.cell(hrow,c).value for c in range(1,ws.max_column+1)]
hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
CO=hm.get("company name",2); WEB=hm.get("website",11)
EM=[hm.get(f"email {i}") for i in range(1,7)]; EM=[c for c in EM if c]
tot_em=match=generic=mismatch=0; ex=[]
for row in ws.iter_rows(min_row=hrow+1, values_only=True):
    def g(c): return row[c-1] if c and len(row)>=c else None
    co=g(CO)
    if not co or not str(co).strip(): continue
    cnk=norm(co); wdom=dom(g(WEB))
    for c in EM:
        e=g(c)
        if not e or "@" not in str(e): continue
        e=str(e).strip().lower(); d=e.split("@")[1]
        tot_em+=1
        if d in GENERIC: generic+=1; continue
        dn=norm(d.split(".")[0])
        ok=(dn and (dn in cnk or cnk[:6] in dn)) or (wdom and (d==wdom or wdom.split(".")[0] in d or d.split(".")[0] in wdom))
        if ok: match+=1
        else:
            mismatch+=1
            if len(ex)<20: ex.append((str(co)[:24], e))
print(f"All Companies — total email addresses: {tot_em}")
print(f"  match own domain : {match} ({match*100//tot_em}%)")
print(f"  generic (gmail..): {generic} ({generic*100//tot_em}%)")
print(f"  MISMATCH (foreign/wrong domain): {mismatch} ({mismatch*100//tot_em}%)")
print("  --- sample mismatched emails ---")
for co,e in ex: print(f"    {co:<24} {e}")
