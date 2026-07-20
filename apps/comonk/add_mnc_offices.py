"""
ADD MNC OFFICE DATA: addresses + phones + HR emails for Ahmedabad/Gandhinagar
Updates COMPLETE_MASTER rows for MNC companies
"""
import openpyxl, re, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MASTER = "COMONK_TRUE_MASTER.xlsx"

def norm(s):
    if not s: return ""
    return re.sub(r'[^a-z0-9]', '', unicodedata.normalize('NFKD', str(s).lower()))

# ── MNC Office Data: researched from official websites, JustDial, Oracle site ─
MNC_OFFICES = {
    "TCS": {
        "city": "Ahmedabad",
        "address": "Sakar II, Off Ashram Road, Ellis Bridge, Ahmedabad, Gujarat 380006",
        "phone": "+91 79 66071100",
        "website": "https://www.tcs.com",
        "linkedin": "https://www.linkedin.com/company/tata-consultancy-services",
        "email6": "hr@tcs.com",
    },
    "Infosys": {
        "city": "Gandhinagar (GIFT City)",
        "address": "Unit 1, 14th-17th Floor, Pragya II, Block 15-C1, Road 11 Zone 1, GIFT SEZ, Gandhinagar 382355",
        "phone": "+91 80 41179999",
        "website": "https://www.infosys.com",
        "linkedin": "https://www.linkedin.com/company/infosys",
        "email6": "contact@infosys.com",
    },
    "Wipro": {
        "city": "Ahmedabad",
        "address": "807-808 Venus Atlantis, Opp. Safal Pegasus, 100 ft Ring Road, Prahladnagar Satellite, Ahmedabad 380015",
        "phone": "+91 80 28440011",
        "website": "https://www.wipro.com",
        "linkedin": "https://www.linkedin.com/company/wipro",
        "email6": "recruiter@wipro.com",
    },
    "HCL Technologies": {
        "city": "Ahmedabad",
        "address": "Titanium One, 100 ft Road, Prahladnagar, Ahmedabad, Gujarat 380015",
        "phone": "+91 120 4960000",
        "website": "https://www.hcltech.com",
        "linkedin": "https://www.linkedin.com/company/hcl-technologies",
        "email6": "hrsupport@hcltech.com",
    },
    "Tech Mahindra": {
        "city": "Ahmedabad",
        "address": "Solitaire Connect, Makarba, Ahmedabad 380051 / Venus Stratum, Surendra Mangaldas Road, Ahmedabad 380015",
        "phone": "+91 40 66283000",
        "website": "https://www.techmahindra.com",
        "linkedin": "https://www.linkedin.com/company/tech-mahindra",
        "email6": "jobs.amd@techmahindra.com",
    },
    "Capgemini": {
        "city": "Gandhinagar",
        "address": "A-201 & 202, Bldg 1, Mindspace SEZ Koba, Gandhinagar 382009",
        "phone": "+91 79 7122 1000",
        "website": "https://www.capgemini.com/in-en",
        "linkedin": "https://www.linkedin.com/company/capgemini",
        "email6": "cg_interview_helpdesk.in@capgemini.com",
    },
    "Accenture": {
        "city": "Ahmedabad",
        "address": "West Wing, Venus Stratum, Venus Grounds, Nr. Jhansi Ki Rani, Nehrunagar, Ahmedabad 380015",
        "phone": "+91 79 66288888",
        "website": "https://www.accenture.com/in-en",
        "linkedin": "https://www.linkedin.com/company/accenture",
        "email6": "india.careers@accenture.com",
    },
    "IBM": {
        "city": "Ahmedabad",
        "address": "Titanium Unit No. 001, Ground Floor, Near Prahadnagar Garden, Ahmedabad 380015",
        "phone": "+91 79 6190 2400",
        "website": "https://www.ibm.com/in-en",
        "linkedin": "https://www.linkedin.com/company/ibm",
        "email6": "india@ibm.com",
    },
    "Cognizant": {
        "city": "Ahmedabad",
        "address": "4th Floor, Iscon Platinum, Near Iscon Cross Roads, Sarkhej-Gandhinagar Highway, Ahmedabad 380054",
        "phone": "+91 44 46799999",
        "website": "https://www.cognizant.com",
        "linkedin": "https://www.linkedin.com/company/cognizant",
        "email6": "recruitment@cognizant.com",
    },
    "Deloitte": {
        "city": "Ahmedabad",
        "address": "19th Floor, Shapath-V, Besides Crowne Plaza, S.G. Highway, Ahmedabad 380015",
        "phone": "+91 79 66827300",
        "website": "https://www.deloitte.com/in",
        "linkedin": "https://www.linkedin.com/company/deloitte",
        "email6": "india.careers@deloitte.com",
    },
    "EY": {
        "city": "Ahmedabad",
        "address": "22nd Floor, B Wing, Privilon, Ambli-Bopal Road, Ahmedabad 380058",
        "phone": "+91 79 40205000",
        "website": "https://www.ey.com/in",
        "linkedin": "https://www.linkedin.com/company/ernst-young",
        "email6": "careers@in.ey.com",
    },
    "PwC": {
        "city": "Ahmedabad",
        "address": "2nd Floor, Shri Krishna Centre, Mithakhali, Navrangpura, Ahmedabad 380009",
        "phone": "+91 79 30913000",
        "website": "https://www.pwc.in",
        "linkedin": "https://www.linkedin.com/company/pwc",
        "email6": "recruitment@pwc.com",
    },
    "Bosch": {
        "city": "Bangalore (HQ)",
        "address": "Bosch Global Software Technologies, Hosur Road, Koramangala, Bangalore 560095",
        "phone": "+91 80 6657 5757",
        "website": "https://www.bosch-softwaretechnologies.com",
        "linkedin": "https://www.linkedin.com/company/bosch-global-software-technologies",
        "email6": "connect@in.bosch.com",
    },
    "Siemens": {
        "city": "Ahmedabad",
        "address": "Siemens Limited, B-Wing, 2nd Floor, Panchvati, Off C.G. Road, Ahmedabad 380006",
        "phone": "+91 22 33223000",
        "website": "https://www.siemens.com/in",
        "linkedin": "https://www.linkedin.com/company/siemens",
        "email6": "hr.india@siemens.com",
    },
    "Oracle": {
        "city": "Gandhinagar (GIFT City)",
        "address": "GIFT One, Level 23, GIFT City, Block 56A, Gujarat International Finance Tech-City, Gandhinagar 382355",
        "phone": "+91 79 6712 7000",
        "website": "https://www.oracle.com/in",
        "linkedin": "https://www.linkedin.com/company/oracle",
        "email6": "india_info@oracle.com",
    },
    "NTT Data": {
        "city": "Ahmedabad",
        "address": "NTT Data India, RCP Business Park, Mahindra SEZ, Andheri East, Mumbai (Gujarat ops via Ahmedabad)",
        "phone": "+91 22 61426142",
        "website": "https://www.nttdata.com",
        "linkedin": "https://www.linkedin.com/company/ntt-data",
        "email6": "global.careers@nttdata.com",
    },
    "Fujitsu": {
        "city": "Ahmedabad",
        "address": "8th Floor, Unit No. 826, Iconic Shyamal, Shyamal Cross Road, Satellite, Ahmedabad 380015",
        "phone": "+91 80 2841 9990",
        "website": "https://www.fujitsu.com/in",
        "linkedin": "https://www.linkedin.com/company/fujitsu",
        "email6": "india.careers@fujitsu.com",
    },
    "KPIT": {
        "city": "Pune (HQ)",
        "address": "Plot No. 17, Rajiv Gandhi Infotech Park, MIDC-SEZ, Phase-3, Hinjawadi, Pune 411057",
        "phone": "+91 20 6770 6000",
        "website": "https://www.kpit.com",
        "linkedin": "https://www.linkedin.com/company/kpit",
        "email6": "careers@kpit.com",
    },
    "LTTS": {
        "city": "Vadodara (HQ)",
        "address": "L&T Knowledge City, West Block – II, NH No. 8, Village Ankhol, Vadodara, Gujarat 390019",
        "phone": "+91 265 2636511",
        "website": "https://www.ltts.com",
        "linkedin": "https://www.linkedin.com/company/lnttechservices",
        "email6": "info@ltts.com",
    },
    "Mphasis": {
        "city": "Bangalore (HQ)",
        "address": "Bagmane World Technology Centre, Marathahalli-Sarjapur Ring Rd, Bangalore 560037",
        "phone": "+91 80 33520000",
        "website": "https://www.mphasis.com",
        "linkedin": "https://www.linkedin.com/company/mphasis",
        "email6": "hr@mphasis.com",
    },
    "Persistent": {
        "city": "Ahmedabad",
        "address": "Navratna Corporate Park, Block-A, 11th Floor, Ambli-Bhopal Road, Bodakdev, Ahmedabad 380058",
        "phone": "+91 20 67053000",
        "website": "https://www.persistent.com",
        "linkedin": "https://www.linkedin.com/company/persistent-systems",
        "email6": "corpsec@persistent.com",
    },
    "Hexaware": {
        "city": "Gandhinagar (GIFT City)",
        "address": "7th and 8th Floor, Pragya II, Block No 15-C1, Road 11 Zone 1, GIFT City, Gandhinagar 382050",
        "phone": "+91 22 43614444",
        "website": "https://www.hexaware.com",
        "linkedin": "https://www.linkedin.com/company/hexaware-technologies",
        "email6": "hr@hexaware.com",
    },
    "Apexon": {
        "city": "Ahmedabad",
        "address": "Apexon, Iscon Elegance, Near Iscon Breakfree, Satellite, Ahmedabad 380015",
        "phone": "+91 79 26934400",
        "website": "https://www.apexon.com",
        "linkedin": "https://www.linkedin.com/company/apexon",
        "email6": "careers@apexon.com",
    },
    "Mastech Digital": {
        "city": "Pittsburgh (US) / India",
        "address": "1305 Cherrington Parkway, Moon Township, PA 15108, USA; India Operations via Noida/Bangalore",
        "phone": "+1 412 787 2100",
        "website": "https://www.mastechdigital.com",
        "linkedin": "https://www.linkedin.com/company/mastech",
        "email6": "hr@mastechdigital.com",
    },
    "EXL": {
        "city": "Ahmedabad",
        "address": "EXL Service, Iscon Centre, Near Shivranjani, Satellite, Ahmedabad 380015",
        "phone": "+91 120 4770000",
        "website": "https://www.exlservice.com",
        "linkedin": "https://www.linkedin.com/company/exl-service",
        "email6": "india.careers@exlservice.com",
    },
}

print("Loading workbook...")
wb = openpyxl.load_workbook(MASTER)
ws = wb["COMPLETE_MASTER"]

# Read headers to find correct columns
headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
print(f"  Columns ({ws.max_column}): {headers[:20]}")

# Try to detect column positions
def find_col(keywords):
    for i, h in enumerate(headers):
        if any(k in h for k in keywords):
            return i + 1
    return None

C_COMP   = find_col(["company"]) or 2
C_CITY   = find_col(["city", "location"]) or 3
C_EMAILS = [6, 7, 8, 9, 10, 11]
C_PHONE  = find_col(["phone", "mobile", "contact no"]) or 12
C_WEB    = find_col(["website", "web"]) or 13
C_LI_CO  = find_col(["linkedin company", "company page"]) or 14
C_ADDR   = find_col(["address"]) or 15
C_SRC    = find_col(["source"]) or 18

print(f"  Detected: Company={C_COMP}, Phone={C_PHONE}, Website={C_WEB}, Address={C_ADDR}")

# Build lookup: norm(name) -> row
existing = {}
for r in range(2, ws.max_row + 1):
    n = ws.cell(r, C_COMP).value
    if n:
        existing[norm(n)] = r

EMAIL_RE = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')

updated = 0
added_phones = 0
added_addr = 0
added_emails = 0

for company, data in MNC_OFFICES.items():
    # Find row in master
    master_row = None
    search_keys = [
        norm(company),
        norm(company.replace(" ", "")),
        norm(company.split()[0]) if " " in company else norm(company),
        norm(company.replace("Technologies", "").replace("Tech", "").replace("Systems", "")),
    ]
    for k in search_keys:
        if k in existing:
            master_row = existing[k]
            break

    if not master_row:
        # Try partial match
        for k in list(existing.keys()):
            cn = norm(company)
            if cn[:6] in k or k[:6] in cn:
                master_row = existing[k]
                break

    if not master_row:
        print(f"  [SKIP] Not in master: {company}")
        continue

    row_changed = False

    # Add phone
    phone = data.get("phone", "")
    if phone and not ws.cell(master_row, C_PHONE).value:
        ws.cell(master_row, C_PHONE).value = phone
        added_phones += 1; row_changed = True

    # Add website
    web = data.get("website", "")
    if web and not ws.cell(master_row, C_WEB).value:
        ws.cell(master_row, C_WEB).value = web
        row_changed = True

    # Add address
    addr = data.get("address", "")
    if addr and not ws.cell(master_row, C_ADDR).value:
        ws.cell(master_row, C_ADDR).value = addr
        added_addr += 1; row_changed = True

    # Add LinkedIn company page
    li_co = data.get("linkedin", "")
    if li_co and C_LI_CO <= ws.max_column:
        if not ws.cell(master_row, C_LI_CO).value:
            ws.cell(master_row, C_LI_CO).value = li_co
            row_changed = True

    # Add extra email if slot available
    email6 = data.get("email6", "")
    if email6 and EMAIL_RE.match(email6):
        cur_emails = set(str(ws.cell(master_row, c).value or "").lower()
                        for c in C_EMAILS if ws.cell(master_row, c).value)
        if email6.lower() not in cur_emails:
            empty = [c for c in C_EMAILS if not ws.cell(master_row, c).value]
            if empty:
                ws.cell(master_row, empty[0]).value = email6
                ws.cell(master_row, empty[0]).font = Font(color="1a56db", size=9)
                added_emails += 1; row_changed = True

    # Mark city if empty
    city = data.get("city", "")
    if city and not ws.cell(master_row, C_CITY).value:
        ws.cell(master_row, C_CITY).value = city

    if row_changed:
        updated += 1
        print(f"  [OK] Updated: {company}")

# ── Update All_Phones sheet with MNC phones ───────────────────────────────────
if "All_Phones" in wb.sheetnames:
    wp = wb["All_Phones"]
    existing_phones = set()
    for r in range(2, wp.max_row + 1):
        v = wp.cell(r, 2).value
        if v: existing_phones.add(str(v).strip())
    pr = wp.max_row + 1
    for company, data in MNC_OFFICES.items():
        ph = data.get("phone", "")
        if ph and ph not in existing_phones:
            wp.cell(pr, 1, pr - 1)
            wp.cell(pr, 2, ph)
            wp.cell(pr, 3, company)
            wp.cell(pr, 4, "MNC-Office")
            existing_phones.add(ph)
            pr += 1

wb.save(MASTER)
print(f"\n  DONE: {updated} MNC rows updated in COMPLETE_MASTER")
print(f"  Phones added: {added_phones} | Addresses added: {added_addr} | Emails added: {added_emails}")
print(f"  Saved: {MASTER}")
