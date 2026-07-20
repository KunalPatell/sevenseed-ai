import openpyxl, re, zipfile, time
from collections import Counter

F="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
def stable_load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read")

EMAIL_OK=re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')
FAKE_PHONES={"7326059369","+917326059369","917326059369"}
PLACEHOLDER_EM=("example.com","test@","noreply@","no-reply@","yourdomain","domain.com","email@","sentry","wixpress","@2x","@3x")

wb=stable_load(F)

for SHEET in ["All Companies","IT Companies","AI ML Companies","IT MNC"]:
    if SHEET not in wb.sheetnames: continue
    ws=wb[SHEET]
    # header row
    hrow=3
    for rr in (1,2,3):
        vals=[str(ws.cell(rr,c).value or "").strip().lower() for c in range(1,20)]
        if "company name" in vals or "company" in vals: hrow=rr; break
    hdr=[ws.cell(hrow,c).value for c in range(1,ws.max_column+1)]
    hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
    CO=hm.get("company name") or hm.get("company")
    PH=hm.get("phone") or hm.get("office phone")
    EMc=[hm.get(f"email {i}") for i in range(1,7)]+[hm.get(f"hr email {i}") for i in range(1,16)]
    EMc=[c for c in EMc if c]

    rows=list(ws.iter_rows(min_row=hrow+1, values_only=True))
    def g(row,c): return row[c-1] if c and len(row)>=c else None

    bad_email=[]; fake_ph=[]; bad_ph=[]; placeholder=[]; blank_co=0; numeric_co=0
    all_emails=[]; all_phones=[]; co_names=[]
    for row in rows:
        co=g(row,CO)
        if co is None or str(co).strip()=="":
            blank_co+=1; continue
        s=str(co).strip()
        co_names.append(s.lower())
        if re.fullmatch(r'[0-9]+', s): numeric_co+=1
        for c in EMc:
            e=g(row,c)
            if e and str(e).strip():
                e=str(e).strip()
                all_emails.append(e.lower())
                if not EMAIL_OK.match(e): bad_email.append((s,e))
                if any(p in e.lower() for p in PLACEHOLDER_EM): placeholder.append((s,e))
        p=g(row,PH)
        if p and str(p).strip():
            praw=str(p).strip(); pd=re.sub(r'\D','',praw)
            all_phones.append(pd)
            if pd in FAKE_PHONES or praw in FAKE_PHONES: fake_ph.append((s,praw))
            elif not (10<=len(pd)<=13): bad_ph.append((s,praw))

    em_dup=Counter(all_emails); em_repeated={e:n for e,n in em_dup.items() if n>=5}
    ph_dup=Counter(all_phones); ph_repeated={p:n for p,n in ph_dup.items() if n>=5}
    co_dup={c:n for c,n in Counter(co_names).items() if n>1}

    print(f"===== {SHEET} =====")
    print(f"  data rows={len(rows)} | emails={len(all_emails)} | phones={len(all_phones)}")
    print(f"  FAULTS:")
    print(f"    blank company rows: {blank_co}")
    print(f"    numeric junk names: {numeric_co}")
    print(f"    duplicate company names: {len(co_dup)}  e.g. {list(co_dup.items())[:3]}")
    print(f"    invalid email format: {len(bad_email)}  e.g. {bad_email[:3]}")
    print(f"    placeholder/junk emails: {len(placeholder)}  e.g. {placeholder[:3]}")
    print(f"    FAKE phone (7326059369): {len(fake_ph)}  e.g. {fake_ph[:2]}")
    print(f"    invalid phone length: {len(bad_ph)}  e.g. {bad_ph[:3]}")
    print(f"    same email on >=5 companies: {len(em_repeated)}  e.g. {list(em_repeated.items())[:3]}")
    print(f"    same phone on >=5 companies: {len(ph_repeated)}  e.g. {list(ph_repeated.items())[:3]}")
    print()
