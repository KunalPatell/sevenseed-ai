"""
TRUE MASTER BUILDER
====================
Merges ALL data files -> enriches with Hunter.io + Abstract API + website scraping
Output: COMONK_TRUE_MASTER.xlsx with every company's complete profile
"""

import openpyxl, csv, re, asyncio, httpx, time, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from html.parser import HTMLParser
from collections import defaultdict

# ── Config ───────────────────────────────────────────────────────────────────
HUNTER_KEY   = "86b42563dd92b15fb26268bdc0f9a697c7609d38"
ABSTRACT_KEY = "4d5211d783a44a71b7e6072b8c2e015c"
OUT_FILE     = "COMONK_TRUE_MASTER.xlsx"
MAX_EMAILS   = 6

HEADERS_HTTP = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124",
    "Accept": "text/html",
}

# ── Helpers ──────────────────────────────────────────────────────────────────
def norm(s):
    """Normalize company name for matching."""
    if not s: return ""
    s = str(s).lower().strip()
    s = unicodedata.normalize('NFKD', s)
    s = re.sub(r'[^a-z0-9]', '', s)
    return s

def clean_url(url):
    if not url: return ""
    url = str(url).strip()
    if not url.startswith("http"): url = "https://" + url
    return url.rstrip('/')

def domain_from_url(url):
    from urllib.parse import urlparse
    if not url: return None
    url = clean_url(url)
    try:
        d = urlparse(url).netloc.replace("www.", "").strip()
        return d if "." in d else None
    except: return None

def valid_email(e):
    return bool(e and re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', str(e).strip()))

def clean_emails(emails):
    seen = set(); out = []
    for e in emails:
        if not e: continue
        e = str(e).strip().lower()
        for part in re.split(r'[;,\s]+', e):
            if valid_email(part) and part not in seen:
                seen.add(part); out.append(part)
    return out[:MAX_EMAILS]

PHONE_RE = re.compile(
    r'(?:\+91[\s\-]?[6-9]\d{9}'
    r'|\+91[\s\-]?79[\s\-]?\d{4}[\s\-]?\d{4}'
    r'|079[\s\-]?\d{7,8}'
    r'|\b[6-9]\d{9}\b)'
)

def clean_phone(p):
    if not p: return ""
    digits = re.sub(r'[\s\-\(\)\+]', '', str(p))
    if digits.startswith('91') and len(digits) == 12: digits = digits[2:]
    elif digits.startswith('0') and 10 <= len(digits) <= 11: digits = digits[1:]
    if len(digits) == 10 and digits[0] in '6789':
        return '+91 ' + digits[:5] + ' ' + digits[5:]
    if len(digits) == 10 and digits.startswith('79'):
        return '+91 79 ' + digits[2:6] + ' ' + digits[6:]
    return str(p).strip() if str(p).strip() else ""

# ── Data record ──────────────────────────────────────────────────────────────
class Company:
    __slots__ = ['name','category','roles','emails','phone','website',
                 'linkedin','address','careers_url','city','priority','source']
    def __init__(self, name):
        self.name = name
        self.category = ""
        self.roles = ""
        self.emails = []
        self.phone = ""
        self.website = ""
        self.linkedin = ""
        self.address = ""
        self.careers_url = ""
        self.city = "Ahmedabad"
        self.priority = ""
        self.source = set()

    def merge(self, other):
        if not self.category and other.category: self.category = other.category
        if not self.roles and other.roles: self.roles = other.roles
        if other.emails: self.emails = clean_emails(self.emails + other.emails)
        if not self.phone and other.phone: self.phone = other.phone
        if not self.website and other.website: self.website = other.website
        if not self.linkedin and other.linkedin: self.linkedin = other.linkedin
        if (not self.address or len(self.address) < len(other.address)) and other.address:
            self.address = other.address
        if not self.careers_url and other.careers_url: self.careers_url = other.careers_url
        if not self.city and other.city: self.city = other.city
        if not self.priority and other.priority: self.priority = other.priority
        self.source.update(other.source)

# ── Loaders ──────────────────────────────────────────────────────────────────
def load_final_master():
    """Ahmedabad_IT_AIML_FINAL_MASTER.xlsx — 541 companies"""
    wb = openpyxl.load_workbook("Ahmedabad_IT_AIML_FINAL_MASTER.xlsx")
    ws = wb.active
    companies = {}
    for r in range(4, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name: continue
        c = Company(str(name).strip())
        c.category    = str(ws.cell(r, 3).value or "")
        c.roles        = str(ws.cell(r, 4).value or "")
        c.emails       = clean_emails([ws.cell(r, col).value for col in range(5, 10)])
        c.phone        = clean_phone(ws.cell(r, 10).value)
        c.website      = clean_url(str(ws.cell(r, 11).value or ""))
        c.linkedin     = str(ws.cell(r, 12).value or "")
        c.address      = str(ws.cell(r, 13).value or "")
        c.source       = {"FINAL_MASTER"}
        companies[norm(name)] = c
    print(f"  FINAL_MASTER: {len(companies)} companies")
    return companies

def load_older_master(companies):
    """Ahmedabad_IT_AIML_Companies_Master.xlsx — may have extra companies"""
    wb = openpyxl.load_workbook("Ahmedabad_IT_AIML_Companies_Master.xlsx")
    added = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(4, ws.max_row + 1):
            name = ws.cell(r, 2).value
            if not name: continue
            key = norm(name)
            c = Company(str(name).strip())
            c.category = str(ws.cell(r, 3).value or "")
            c.roles    = str(ws.cell(r, 4).value or "")
            c.emails   = clean_emails([ws.cell(r, col).value for col in range(5, 10)])
            c.phone    = clean_phone(ws.cell(r, 10).value)
            c.website  = clean_url(str(ws.cell(r, 11).value or ""))
            c.linkedin = str(ws.cell(r, 12).value or "")
            c.address  = str(ws.cell(r, 13).value or "")
            c.source   = {"OLDER_MASTER"}
            if key in companies:
                companies[key].merge(c)
            else:
                companies[key] = c; added += 1
    print(f"  OLDER_MASTER: +{added} new companies merged")

def load_aiml_hr_contacts(companies):
    """Ahmedabad_AIML_Companies_HR_Contacts.xlsx — 101 AI/ML companies with Careers URLs"""
    wb = openpyxl.load_workbook("Ahmedabad_AIML_Companies_HR_Contacts.xlsx")
    ws = wb.active
    added = 0
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name: continue
        key = norm(name)
        c = Company(str(name).strip())
        c.category    = str(ws.cell(r, 3).value or "AI / ML")
        c.emails      = clean_emails([ws.cell(r, 4).value, ws.cell(r, 5).value])
        c.phone       = clean_phone(ws.cell(r, 6).value)
        c.website     = clean_url(str(ws.cell(r, 7).value or ""))
        c.address     = str(ws.cell(r, 8).value or "")
        c.careers_url = clean_url(str(ws.cell(r, 9).value or ""))
        c.source      = {"AIML_HR_CONTACTS"}
        if key in companies:
            companies[key].merge(c)
        else:
            companies[key] = c; added += 1
    print(f"  AIML_HR_CONTACTS: +{added} new, careers URLs enriched")

def load_aiml_apply_list(companies):
    """AIML_Companies_Apply_List.csv — 77 AI/ML with full addresses"""
    added = 0
    with open("AIML_Companies_Apply_List.csv", encoding='utf-8-sig', errors='ignore') as f:
        for row in csv.DictReader(f):
            name = row.get('Company', '').strip()
            if not name: continue
            key = norm(name)
            c = Company(name)
            c.category = row.get('Category', '')
            c.roles    = row.get('Roles', '')
            c.emails   = clean_emails([row.get('Email',''), row.get('All Emails','')])
            c.phone    = clean_phone(row.get('Phone',''))
            c.website  = clean_url(row.get('Website',''))
            c.linkedin = row.get('LinkedIn','')
            c.address  = row.get('Address','')
            c.source   = {"AIML_APPLY_LIST"}
            if key in companies:
                companies[key].merge(c)
            else:
                companies[key] = c; added += 1
    print(f"  AIML_APPLY_LIST: +{added} new companies, addresses enriched")

def load_job_targets(companies):
    """AI_Engineer_Job_Targets.xlsx — priority + Gandhinagar data"""
    wb = openpyxl.load_workbook("AI_Engineer_Job_Targets.xlsx")
    added = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        city = "Gandhinagar" if "Gandhinagar" in sname or "GIFT" in sname else "Ahmedabad"
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 2).value
            if not name: continue
            key = norm(name)
            c = Company(str(name).strip())
            c.priority = str(ws.cell(r, 1).value or "")
            c.roles    = str(ws.cell(r, 3).value or "")
            c.emails   = clean_emails([ws.cell(r, 4).value])
            c.phone    = clean_phone(ws.cell(r, 5).value)
            c.website  = clean_url(str(ws.cell(r, 6).value or ""))
            c.linkedin = str(ws.cell(r, 7).value or "")
            c.address  = str(ws.cell(r, 8).value or "")
            c.city     = city
            c.source   = {"JOB_TARGETS"}
            if key in companies:
                companies[key].merge(c)
                if not companies[key].priority and c.priority:
                    companies[key].priority = c.priority
                if city == "Gandhinagar":
                    companies[key].city = "Gandhinagar"
            else:
                companies[key] = c; added += 1
    print(f"  JOB_TARGETS: +{added} new, priority + Gandhinagar data enriched")

def load_gandhinagar(companies):
    """Gandhinagar_GIFTCity_Companies.csv"""
    added = 0
    with open("Gandhinagar_GIFTCity_Companies.csv", encoding='utf-8-sig', errors='ignore') as f:
        for row in csv.DictReader(f):
            name = row.get('Company', '').strip()
            if not name: continue
            key = norm(name)
            c = Company(name)
            c.category = row.get('Category','')
            c.roles    = row.get('Roles','')
            c.emails   = clean_emails([row.get('Email',''), row.get('All Emails','')])
            c.phone    = clean_phone(row.get('Phone',''))
            c.website  = clean_url(row.get('Website',''))
            c.linkedin = row.get('LinkedIn','')
            c.address  = row.get('Address','')
            c.city     = "Gandhinagar"
            c.source   = {"GANDHINAGAR_CSV"}
            if key in companies:
                companies[key].merge(c)
                companies[key].city = "Gandhinagar"
            else:
                companies[key] = c; added += 1
    print(f"  GANDHINAGAR_CSV: +{added} new Gandhinagar companies")

def load_hr_mail_list(companies):
    """HR Mail List.xlsx — 325 companies with up to 6 HR emails"""
    wb = openpyxl.load_workbook("HR Mail List.xlsx")
    ws = wb["HR Mail List"]
    enriched = 0
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name: continue
        key = norm(name)
        emails = clean_emails([ws.cell(r, col).value for col in range(2, 9)])
        if key in companies and emails:
            old_count = len(companies[key].emails)
            companies[key].emails = clean_emails(companies[key].emails + emails)
            if len(companies[key].emails) > old_count:
                enriched += 1
                companies[key].source.add("HR_MAIL_LIST")
    print(f"  HR_MAIL_LIST: {enriched} companies got more HR emails")

def load_emails_csv(companies):
    """emails.csv — 674 individual emails, match by domain"""
    domain_map = defaultdict(list)
    with open("emails.csv", encoding='utf-8', errors='ignore') as f:
        for row in csv.DictReader(f):
            email = row.get('email','').strip()
            if valid_email(email):
                domain = email.split('@')[1]
                domain_map[domain].append(email)

    enriched = 0
    for key, co in companies.items():
        domain = domain_from_url(co.website)
        if domain and domain in domain_map:
            old_count = len(co.emails)
            co.emails = clean_emails(co.emails + domain_map[domain])
            if len(co.emails) > old_count:
                enriched += 1
                co.source.add("EMAILS_CSV")
    print(f"  EMAILS_CSV: {enriched} companies got more emails matched by domain")

# ── Website scraper for missing addresses ────────────────────────────────────
class TextExtractor(HTMLParser):
    def __init__(self):
        super().__init__(); self.result = []; self._skip = False
    def handle_starttag(self, t, a):
        if t in ('script','style','noscript'): self._skip = True
    def handle_endtag(self, t):
        if t in ('script','style','noscript'): self._skip = False
    def handle_data(self, d):
        if not self._skip: self.result.append(d)
    def get_text(self): return ' '.join(self.result)

def extract_text(html):
    p = TextExtractor()
    try: p.feed(html)
    except: pass
    return p.get_text()

ADDR_RE = re.compile(
    r'(?:'
    r'\d+[-/,\s]+[A-Za-z][^,\n]{5,50},\s*(?:Ahmedabad|Gandhinagar|Gujarat|GIFT)[^,\n]{0,30}'
    r'|(?:Ahmedabad|Gandhinagar|GIFT\s*City)[^,\n]{0,50}[\-–,]\s*\d{6}'
    r'|[A-Z]\d+[,\s]+[^,\n]{5,60},\s*(?:Ahmedabad|Gandhinagar)[^,\n]{0,30}'
    r')',
    re.IGNORECASE
)

async def fetch_address(client, url, sem):
    async with sem:
        pages = [url, url + '/contact', url + '/contact-us', url + '/about']
        for page in pages:
            try:
                r = await client.get(page, timeout=6, follow_redirects=True)
                if r.status_code != 200: continue
                text = extract_text(r.text)
                matches = ADDR_RE.findall(text)
                if matches:
                    best = max(matches, key=len)
                    return re.sub(r'\s+', ' ', best).strip()
            except: pass
        return None

# ── Hunter.io email fetch ────────────────────────────────────────────────────
async def hunter_email(client, domain, sem):
    async with sem:
        try:
            r = await client.get(
                "https://api.hunter.io/v2/domain-search",
                params={"domain": domain, "api_key": HUNTER_KEY, "limit": 6, "type": "personal"},
                timeout=10
            )
            data = r.json()
            if "data" not in data: return []
            raw = data["data"].get("emails", [])
            hr   = [e["value"] for e in raw if any(k in e.get("value","").lower()
                    for k in ["hr","recruit","career","talent","people","hiring","hrbp"])]
            rest = [e["value"] for e in raw]
            return clean_emails(hr or rest)
        except: return []

# ── Abstract API email validate ──────────────────────────────────────────────
async def validate_emails(client, emails, sem):
    """Validate emails and return only deliverable ones."""
    valid = []
    for email in emails[:3]:  # validate top 3 only (quota)
        async with sem:
            try:
                r = await client.get(
                    "https://emailvalidation.abstractapi.com/v1/",
                    params={"api_key": ABSTRACT_KEY, "email": email},
                    timeout=8
                )
                data = r.json()
                deliverable = data.get("deliverability", "UNKNOWN")
                if deliverable in ("DELIVERABLE", "UNKNOWN"):
                    valid.append(email)
                await asyncio.sleep(0.2)
            except: valid.append(email)
    # Keep remaining emails without validation
    valid += [e for e in emails[3:] if e not in valid]
    return clean_emails(valid)

# ── Main async enrichment ────────────────────────────────────────────────────
async def enrich_all(companies):
    sem_web    = asyncio.Semaphore(25)
    sem_hunter = asyncio.Semaphore(3)
    sem_abs    = asyncio.Semaphore(2)

    # Check Hunter quota
    async with httpx.AsyncClient(headers=HEADERS_HTTP) as client:
        try:
            qr = (await client.get("https://api.hunter.io/v2/account",
                  params={"api_key": HUNTER_KEY}, timeout=8)).json()
            searches = qr.get("data", {}).get("searches", {})
            remaining = searches.get("available", 0) - searches.get("used", 0)
            print(f"  Hunter.io searches remaining: {remaining}")
        except: remaining = 0

    need_addr  = [co for co in companies.values() if not co.address and co.website]
    need_email = [co for co in companies.values()
                  if not co.emails and co.website and remaining > 0][:min(remaining-2, 30)]

    print(f"  Need address: {len(need_addr)}, Need email (Hunter): {len(need_email)}")

    async with httpx.AsyncClient(headers=HEADERS_HTTP, follow_redirects=True) as client:
        # 1. Fetch missing addresses in parallel
        if need_addr:
            print(f"\n  Scraping addresses for {len(need_addr)} companies...")
            addr_tasks = [fetch_address(client, co.website, sem_web) for co in need_addr]
            addr_results = await asyncio.gather(*addr_tasks)
            found = 0
            for co, addr in zip(need_addr, addr_results):
                if addr:
                    co.address = addr
                    co.source.add("WEB_SCRAPE")
                    found += 1
            print(f"  Addresses found: {found}/{len(need_addr)}")

        # 2. Fetch emails via Hunter.io for companies with no emails
        if need_email:
            print(f"\n  Fetching emails via Hunter.io for {len(need_email)} companies...")
            hunter_tasks = [hunter_email(client, domain_from_url(co.website), sem_hunter)
                            for co in need_email]
            hunter_results = await asyncio.gather(*hunter_tasks)
            found = 0
            for co, emails in zip(need_email, hunter_results):
                if emails:
                    co.emails = emails
                    co.source.add("HUNTER")
                    found += 1
                await asyncio.sleep(0.3)
            print(f"  Hunter emails found: {found}/{len(need_email)}")

# ── Write master Excel ───────────────────────────────────────────────────────
def write_master(companies):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Complete Master ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "COMPLETE_MASTER"

    COLS = [
        ("#",          4),  ("Company",       32), ("City",        14),
        ("Category",   16), ("AI Roles",      36), ("Email 1",     32),
        ("Email 2",    32), ("Email 3",        32), ("Email 4",     32),
        ("Email 5",    32), ("Email 6",        32), ("Phone",       18),
        ("Website",    28), ("LinkedIn",       38), ("Address",     48),
        ("Careers URL",28), ("Priority",       22), ("Sources",     22),
    ]

    # Header
    h_fill = PatternFill("solid", fgColor="2E1A47")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    for c, (hdr, w) in enumerate(COLS, 1):
        cell = ws.cell(1, c, hdr)
        cell.fill = h_fill; cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    # Sort: Gandhinagar first then Ahmedabad, priority companies first, then alpha
    def sort_key(co):
        city_rank = 0 if co.city in ("Gandhinagar","GIFT City") else 1
        prio_rank = 0 if co.priority and "Apply now" in co.priority else 1
        return (city_rank, prio_rank, (co.category or "Z"), co.name.lower())

    sorted_cos = sorted(companies.values(), key=sort_key)

    fill_a = PatternFill("solid", fgColor="F5F0FF")
    fill_b = PatternFill("solid", fgColor="FFFFFF")
    fill_g = PatternFill("solid", fgColor="E8F5E9")  # green for Gandhinagar

    ai_kw = ['ai','ml','data','machine','deep','nlp','computer vision','llm','genai']

    for i, co in enumerate(sorted_cos, 1):
        row = i + 1
        is_gandhi = co.city in ("Gandhinagar", "GIFT City")
        is_ai = any(k in (co.category or '').lower() or k in (co.roles or '').lower()
                    for k in ai_kw)
        fill = fill_g if is_gandhi else (fill_a if i % 2 == 0 else fill_b)

        values = [
            i, co.name, co.city or "Ahmedabad", co.category, co.roles,
        ] + (co.emails + [""] * MAX_EMAILS)[:MAX_EMAILS] + [
            co.phone, co.website, co.linkedin, co.address,
            co.careers_url, co.priority, ", ".join(sorted(co.source))
        ]

        for c, v in enumerate(values, 1):
            cell = ws.cell(row, c, v or "")
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (5,14,15)))
            if c == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c in (6,7,8,9,10,11) and v:  # emails blue
                cell.font = Font(color="1a56db", size=9)
            if c == 2 and is_ai:
                cell.font = Font(bold=True, color="4A235A")
            if c == 14 and v:  # LinkedIn
                cell.font = Font(color="0A66C2", size=9)
        ws.row_dimensions[row].height = 18

    # ── Sheet 2: AI/ML Only ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("AI_ML_ONLY")
    ws2.freeze_panes = "A2"
    for c, (hdr, w) in enumerate(COLS, 1):
        cell = ws2.cell(1, c, hdr)
        cell.fill = h_fill; cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[1].height = 26

    ai_only = [co for co in sorted_cos if any(k in (co.category or '').lower()
               or k in (co.roles or '').lower() for k in ai_kw)]
    for i, co in enumerate(ai_only, 1):
        row = i + 1
        is_gandhi = co.city in ("Gandhinagar", "GIFT City")
        fill = fill_g if is_gandhi else (fill_a if i % 2 == 0 else fill_b)
        values = [i, co.name, co.city or "Ahmedabad", co.category, co.roles] + \
                 (co.emails + [""]*MAX_EMAILS)[:MAX_EMAILS] + \
                 [co.phone, co.website, co.linkedin, co.address,
                  co.careers_url, co.priority, ", ".join(sorted(co.source))]
        for c, v in enumerate(values, 1):
            cell = ws2.cell(row, c, v or "")
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            if c in (6,7,8,9,10,11) and v:
                cell.font = Font(color="1a56db", size=9)
        ws2.row_dimensions[row].height = 18

    # ── Sheet 3: Stats ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("STATS")
    total = len(sorted_cos)
    ai_count = len(ai_only)
    gandhi_count = sum(1 for co in sorted_cos if co.city in ("Gandhinagar","GIFT City"))
    with_email = sum(1 for co in sorted_cos if co.emails)
    with_phone = sum(1 for co in sorted_cos if co.phone)
    with_addr  = sum(1 for co in sorted_cos if co.address and len(co.address) > 5)
    with_linked = sum(1 for co in sorted_cos if co.linkedin)
    with_careers = sum(1 for co in sorted_cos if co.careers_url)

    stats = [
        ("COMONK TRUE MASTER — STATS", ""),
        ("", ""),
        ("Total Companies",       total),
        ("  Ahmedabad",           total - gandhi_count),
        ("  Gandhinagar/GIFT City", gandhi_count),
        ("AI/ML Companies",       ai_count),
        ("", ""),
        ("Have Email",            f"{with_email} ({round(with_email/total*100)}%)"),
        ("Have Phone",            f"{with_phone} ({round(with_phone/total*100)}%)"),
        ("Have Address",          f"{with_addr} ({round(with_addr/total*100)}%)"),
        ("Have LinkedIn",         f"{with_linked} ({round(with_linked/total*100)}%)"),
        ("Have Careers URL",      f"{with_careers} ({round(with_careers/total*100)}%)"),
        ("", ""),
        ("Data Sources Merged", ""),
        ("  FINAL_MASTER (541)",              "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"),
        ("  OLDER_MASTER",                   "Ahmedabad_IT_AIML_Companies_Master.xlsx"),
        ("  AIML_HR_CONTACTS (101 + Careers)","Ahmedabad_AIML_Companies_HR_Contacts.xlsx"),
        ("  AIML_APPLY_LIST (77 + addresses)","AIML_Companies_Apply_List.csv"),
        ("  JOB_TARGETS (priority tags)",    "AI_Engineer_Job_Targets.xlsx"),
        ("  GANDHINAGAR (10)",               "Gandhinagar_GIFTCity_Companies.csv"),
        ("  HR_MAIL_LIST (6 emails each)",   "HR Mail List.xlsx"),
        ("  EMAILS_CSV (674 emails)",        "emails.csv"),
        ("  HUNTER.IO API",                  "Email enrichment"),
        ("  WEB_SCRAPE",                     "Address enrichment from company websites"),
    ]
    title_font = Font(bold=True, size=12, color="2E1A47")
    for r, (k, v) in enumerate(stats, 1):
        ws3.cell(r, 1, k)
        ws3.cell(r, 2, v)
        if k and not k.startswith("  ") and k != "":
            ws3.cell(r, 1).font = title_font
    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 45

    wb.save(OUT_FILE)
    print(f"\n  COMPLETE_MASTER: {total} companies")
    print(f"  AI_ML_ONLY: {ai_count} companies")
    print(f"  Gandhinagar/GIFT City: {gandhi_count}")
    print(f"  With email: {with_email} ({round(with_email/total*100)}%)")
    print(f"  With phone: {with_phone} ({round(with_phone/total*100)}%)")
    print(f"  With address: {with_addr} ({round(with_addr/total*100)}%)")
    print(f"  With LinkedIn: {with_linked} ({round(with_linked/total*100)}%)")
    print(f"  With Careers URL: {with_careers}")
    print(f"\n  Saved: {OUT_FILE}")

# ── Entry point ──────────────────────────────────────────────────────────────
def main():
    print("\n" + "="*60)
    print("  COMONK TRUE MASTER BUILDER")
    print("="*60)
    print("\n[1/4] Loading all data sources...")

    companies = load_final_master()
    load_older_master(companies)
    load_aiml_hr_contacts(companies)
    load_aiml_apply_list(companies)
    load_job_targets(companies)
    load_gandhinagar(companies)
    load_hr_mail_list(companies)
    load_emails_csv(companies)

    print(f"\n  Total unique companies after merge: {len(companies)}")

    print("\n[2/4] Enriching with Hunter.io + website scraping...")
    asyncio.run(enrich_all(companies))

    print("\n[3/4] Writing master Excel...")
    write_master(companies)

    print("\n[4/4] Done!")
    print("="*60)

if __name__ == "__main__":
    main()
