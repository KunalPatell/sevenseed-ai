"""
Sort COMPLETE_MASTER so the most valuable Gujarat targets sit at the top.
NO rows are removed. Ranking (top -> bottom):
  Tier 1: MNC + has email        (must-have targets, mailable)
  Tier 2: AI/ML + has email      (career target, mailable)
  Tier 3: IT/other + has email   (mailable)
  Tier 4: MNC + no email         (research later)
  Tier 5: AI/ML + no email
  Tier 6: everything else
Gandhinagar/GIFT City nudged slightly above Ahmedabad within same tier
(both are top-priority Gujarat, but GIFT City = major MNC hub).
Sets a clean Priority label and renumbers the 'No.' column.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

MASTER = "COMONK_TRUE_MASTER.xlsx"
wb = openpyxl.load_workbook(MASTER)
ws = wb["COMPLETE_MASTER"]

headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_NO   = hmap.get("no.", 1)
COL_CO   = hmap.get("company name", 2)
COL_CITY = hmap.get("city", 3)
COL_CAT  = hmap.get("category", 4)
COL_PH   = hmap.get("phone", 12)
COL_PRI  = hmap.get("priority", 17)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]
email_cols = [c for c in email_cols if c]
ncols = len(headers)

# Read all data rows into memory
rows = []
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, ncols + 1)]
    if not vals[COL_CO - 1]:
        continue
    rows.append(vals)

print(f"Read {len(rows)} companies")

def classify(vals):
    cat  = str(vals[COL_CAT - 1] or "").lower()
    city = str(vals[COL_CITY - 1] or "").lower()
    has_email = any(vals[c - 1] for c in email_cols)
    is_mnc   = cat.startswith("mnc") or "mnc" in cat
    is_aiml  = ("ai-ml" in cat or "ai/ml" in cat or "artificial" in cat
                or "machine learning" in cat)
    # Tier
    if is_mnc and has_email:      tier = 1
    elif is_aiml and has_email:   tier = 2
    elif has_email:               tier = 3
    elif is_mnc:                  tier = 4
    elif is_aiml:                 tier = 5
    else:                          tier = 6
    # City nudge: GIFT/Gandhinagar slightly ahead within tier
    city_rank = 0 if ("gandhinagar" in city or "gift" in city) else 1
    label_map = {
        1: "1 - MNC Gujarat (Mailable) TOP",
        2: "2 - AI/ML Gujarat (Mailable)",
        3: "3 - IT Gujarat (Mailable)",
        4: "4 - MNC Gujarat (Research)",
        5: "5 - AI/ML Gujarat (Research)",
        6: "6 - Other Gujarat",
    }
    return tier, city_rank, label_map[tier], has_email

# Sort
def sort_key(vals):
    tier, city_rank, label, _ = classify(vals)
    return (tier, city_rank, str(vals[COL_CO - 1] or "").lower())

rows.sort(key=sort_key)

# Set priority labels + renumber, then write back
from collections import Counter
tier_counts = Counter()
for i, vals in enumerate(rows):
    tier, city_rank, label, has_email = classify(vals)
    vals[COL_NO - 1]  = i + 1
    vals[COL_PRI - 1] = label
    tier_counts[label] += 1

# Clear existing data rows
if ws.max_row >= 2:
    ws.delete_rows(2, ws.max_row - 1)

# Colors per tier
TIER_FILL = {
    "1 - MNC Gujarat (Mailable) TOP": PatternFill("solid", fgColor="C6EFCE"),  # green
    "2 - AI/ML Gujarat (Mailable)":   PatternFill("solid", fgColor="D9EAD3"),  # light green
    "3 - IT Gujarat (Mailable)":      PatternFill("solid", fgColor="FFF2CC"),  # light yellow
    "4 - MNC Gujarat (Research)":     PatternFill("solid", fgColor="FCE4D6"),  # light orange
    "5 - AI/ML Gujarat (Research)":   PatternFill("solid", fgColor="FFF7E6"),
    "6 - Other Gujarat":              None,
}

for vals in rows:
    ws.append(vals)

# Re-apply light styling based on priority tier
for r in range(2, ws.max_row + 1):
    pri = str(ws.cell(r, COL_PRI).value or "")
    fill = TIER_FILL.get(pri)
    for c in range(1, ncols + 1):
        cell = ws.cell(r, c)
        cell.font = Font(size=9)
        if fill:
            cell.fill = fill
        # style emails blue
        if c in email_cols and cell.value and "@" in str(cell.value):
            cell.font = Font(color="1a56db", size=9)

wb.save(MASTER)

print("\n=== Gujarat priority sort complete ===")
for label in sorted(tier_counts):
    print(f"  {tier_counts[label]:>5}  {label}")
print(f"\nTotal: {len(rows)} companies, sorted, renumbered")
print(f"Saved -> {MASTER}")
