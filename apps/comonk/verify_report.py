import openpyxl, re, zipfile, time

F="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
def stable_load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read")

JUNK={"","n/a","na","-","--","none","null","nil","tbd","."}
def has_val(v):
    return v is not None and str(v).strip().lower() not in JUNK
def is_email(v):
    return v is not None and re.match(r'^[^@\s]+@[^@\s]+\.[A-Za-z]{2,}$', str(v).strip()) is not None

wb=stable_load(F)
print("="*70)
print("  VERIFIED COVERAGE REPORT")
print("="*70)
for SHEET in wb.sheetnames:
    if SHEET=="LinkedIn Profiles": continue
    ws=wb[SHEET]
    hrow=None
    for rr in (1,2,3,4):
        vals=[str(ws.cell(rr,c).value or "").strip().lower() for c in range(1,26)]
        if "company name" in vals or "company" in vals: hrow=rr; break
    if not hrow: continue
    hdr=[ws.cell(hrow,c).value for c in range(1,ws.max_column+1)]
    hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
    CO=hm.get("company name") or hm.get("company")
    PH=hm.get("phone") or hm.get("office phone")
    WEB=hm.get("website")
    ADDR=hm.get("address")
    EM=[hm.get(f"email {i}") for i in range(1,7)]+[hm.get(f"hr email {i}") for i in range(1,16)]
    EM=[c for c in EM if c]

    def colname(i): return hdr[i-1] if i else None
    tot=wem=wph=wweb=wad=w_ep=w_all=0
    for row in ws.iter_rows(min_row=hrow+1, values_only=True):
        def g(c): return row[c-1] if c and len(row)>=c else None
        co=g(CO)
        if not has_val(co): continue
        tot+=1
        e=any(is_email(g(c)) for c in EM)
        p=has_val(g(PH))
        w=has_val(g(WEB))
        a=has_val(g(ADDR)) if ADDR else False
        if e:wem+=1
        if p:wph+=1
        if w:wweb+=1
        if a:wad+=1
        if e and p:w_ep+=1
        if e and p and w and (a if ADDR else True):w_all+=1
    def pc(n): return f"{round(n*100/tot)}%" if tot else "-"
    print(f"\n### {SHEET} — {tot} companies")
    print(f"    (header row {hrow}; Company=col{CO}, Phone=col{PH}, Website=col{WEB}, Address=col{ADDR}, Email cols={EM})")
    print(f"    Email   : {wem:>5} ({pc(wem)})")
    print(f"    Phone   : {wph:>5} ({pc(wph)})")
    print(f"    Website : {wweb:>5} ({pc(wweb)})")
    print(f"    Address : {wad:>5} ({pc(wad)})" + ("" if ADDR else "  [no address column]"))
    print(f"    Email+Phone      : {w_ep:>5} ({pc(w_ep)})")
    print(f"    Complete (all)   : {w_all:>5} ({pc(w_all)})")
print("\n"+"="*70)
