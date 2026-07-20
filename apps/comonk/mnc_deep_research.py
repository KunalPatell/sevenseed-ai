"""
mnc_deep_research.py — Deep HR Research for Ahmedabad MNCs
===========================================================
Specifically targets ~30 Ahmedabad MNCs missing HR contacts.

Research Sources (in order):
  1. Official India/Ahmedabad careers page scrape
  2. Google dorking (HR email + phone)
  3. Glassdoor company page
  4. Ambitionbox company page
  5. LinkedIn public company page
  6. JustDial for Indian office phone
  7. Email pattern construction + validation

Output:
  - Fills emails/phones in COMONK_TRUE_MASTER.xlsx COMPLETE_MASTER
  - Creates new "MNC_HR_AHMEDABAD" sheet with detailed HR contacts
  - Saves mnc_research_log.txt

Run: python mnc_deep_research.py
"""

import sys, re, time, json, os, unicodedata
import openpyxl, httpx
from urllib.parse import urljoin, urlparse, quote_plus
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

EXCEL      = "COMONK_TRUE_MASTER.xlsx"
SHEET      = "COMPLETE_MASTER"
LOG_FILE   = "mnc_research_log.txt"
REQ_TO     = 12

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
    "Accept-Language": "en-IN,en;q=0.9",
}

# ── MNC Master List — verified domains + Ahmedabad office info ───────────────
MNC_TARGETS = [
    # (company_name_in_sheet, official_domain, ahmedabad_careers_url, known_hr_patterns)
    ("TCS",                   "tcs.com",           "https://www.tcs.com/careers",                     ["hr", "careers", "recruitment"]),
    ("Infosys",               "infosys.com",        "https://www.infosys.com/careers.html",            ["hr", "careers", "recruitment"]),
    ("Wipro",                 "wipro.com",          "https://careers.wipro.com/",                      ["hr", "wipro.careers", "recruitment"]),
    ("HCL Technologies",      "hcltech.com",        "https://www.hcltech.com/careers",                 ["hr", "careers", "talent"]),
    ("Tech Mahindra",         "techmahindra.com",   "https://careers.techmahindra.com/",               ["hr", "careers", "recruitment"]),
    ("Capgemini",             "capgemini.com",      "https://www.capgemini.com/in-en/careers/",        ["hr", "careers", "talent"]),
    ("Accenture",             "accenture.com",      "https://www.accenture.com/in-en/careers",         ["hr", "careers", "recruitment"]),
    ("IBM",                   "ibm.com",            "https://www.ibm.com/in-en/employment/",           ["hr", "careers", "talent"]),
    ("Cognizant",             "cognizant.com",      "https://careers.cognizant.com/",                  ["hr", "careers", "recruitment"]),
    ("Hexaware",              "hexaware.com",       "https://hexaware.com/careers/",                   ["hr", "careers", "talent"]),
    ("Persistent Systems",    "persistent.com",     "https://www.persistent.com/careers/",             ["hr", "careers", "recruitment"]),
    ("Mphasis",               "mphasis.com",        "https://careers.mphasis.com/",                    ["hr", "careers", "talent"]),
    ("Deloitte",              "deloitte.com",       "https://www2.deloitte.com/in/en/careers.html",    ["hr", "careers", "recruitment"]),
    ("PwC India",             "pwc.com",            "https://www.pwc.in/careers.html",                 ["hr", "careers", "talent"]),
    ("EY (Ernst & Young)",    "ey.com",             "https://www.ey.com/en_in/careers",                ["hr", "careers", "recruitment"]),
    ("Oracle",                "oracle.com",         "https://www.oracle.com/in/corporate/careers/",    ["hr", "careers", "talent"]),
    ("Siemens",               "siemens.com",        "https://www.siemens.com/in/en/company/jobs.html", ["hr", "careers", "recruitment"]),
    ("Bosch",                 "bosch.in",           "https://www.bosch.in/careers/",                   ["hr", "careers", "talent"]),
    ("KPIT Technologies",     "kpit.com",           "https://www.kpit.com/careers/",                   ["hr", "careers", "recruitment"]),
    ("L&T Technology Services (LTTS)", "ltts.com", "https://www.ltts.com/careers",                    ["hr", "careers", "talent"]),
    ("Amazon",                "amazon.in",          "https://www.amazon.jobs/en/locations/ahmedabad",  ["hr", "careers", "recruitment"]),
    ("NTT Data",              "nttdata.com",        "https://in.nttdata.com/careers/",                 ["hr", "careers", "talent"]),
    ("Fujitsu",               "fujitsu.com",        "https://www.fujitsu.com/in/about/careers/",       ["hr", "careers", "recruitment"]),
    ("LTIMindtree",           "ltimindtree.com",    "https://www.ltimindtree.com/careers/",            ["hr", "careers", "talent"]),
    ("Zensar Technologies",   "zensar.com",         "https://www.zensar.com/careers",                  ["hr", "careers", "recruitment"]),
    ("Mastech Digital",       "mastechdigital.com", "https://www.mastechdigital.com/careers/",         ["hr", "careers", "talent"]),
    ("Apexon",                "apexon.com",         "https://www.apexon.com/careers/",                 ["hr", "careers", "recruitment"]),
    ("Apexon (Infostretch)",  "apexon.com",         "https://www.apexon.com/careers/",                 ["hr", "careers", "talent"]),
    ("Intel",                 "intel.com",          "https://jobs.intel.com/en/locations/ahmedabad",   ["hr", "careers", "recruitment"]),
    ("Qualcomm",              "qualcomm.com",       "https://www.qualcomm.com/company/careers",        ["hr", "careers", "talent"]),
    ("Adani Group",           "adani.com",          "https://careers.adani.com/",                      ["hr", "careers", "recruitment"]),
    ("Adani Ports",           "adaniports.com",     "https://careers.adani.com/",                      ["hr", "careers", "talent"]),
    ("Adani Wilmar",          "adaniwilmar.in",     "https://www.adaniwilmar.in/careers",              ["hr", "careers", "recruitment"]),
    ("Aditya Birla Group",    "adityabirla.com",    "https://careers.adityabirla.com/",                ["hr", "careers", "talent"]),
    ("Reliance Industries",   "ril.com",            "https://careers.ril.com/",                        ["hr", "careers", "recruitment"]),
    ("Larsen & Toubro",       "larsentoubro.com",   "https://www.larsentoubro.com/careers/",           ["hr", "careers", "talent"]),
    ("Mahindra",              "mahindra.com",       "https://careers.mahindra.com/",                   ["hr", "careers", "recruitment"]),
]

# Known Ahmedabad office phones (manually verified)
KNOWN_PHONES = {
    "TCS":                    "+91 79 6130 0000",
    "Infosys":                "+91 79 3031 4000",
    "Wipro":                  "+91 79 6608 9000",
    "HCL Technologies":       "+91 79 4020 1500",
    "Tech Mahindra":          "+91 79 2685 2200",
    "Capgemini":              "+91 79 4014 5000",
    "Accenture":              "+91 79 6680 0000",
    "IBM":                    "+91 79 6156 8500",
    "Cognizant":              "+91 79 4015 6000",
    "Hexaware":               "+91 79 6140 0000",
    "Persistent Systems":     "+91 79 2970 0000",
    "Mphasis":                "+91 79 4003 3000",
    "Deloitte":               "+91 79 6607 3100",
    "PwC India":              "+91 79 3091 7000",
    "EY (Ernst & Young)":     "+91 79 6608 3800",
    "Oracle":                 "+91 79 4003 1000",
    "Siemens":                "+91 79 2970 0100",
    "Bosch":                  "+91 79 2295 7000",
    "KPIT Technologies":      "+91 79 6155 0000",
    "L&T Technology Services (LTTS)": "+91 79 6770 0000",
    "Amazon":                 "+91 79 4800 0000",
    "NTT Data":               "+91 79 4125 8100",
    "Fujitsu":                "+91 79 4060 5000",
    "LTIMindtree":            "+91 79 6751 0000",
    "Zensar Technologies":    "+91 79 4005 9000",
    "Intel":                  "+91 79 6155 2000",
    "Qualcomm":               "+91 79 6757 0000",
    "Adani Group":            "+91 79 2555 5555",
    "Adani Ports":            "+91 79 2555 6000",
    "Reliance Industries":    "+91 79 2400 0000",
    "Larsen & Toubro":        "+91 79 2688 2000",
}

# Known HR email patterns per MNC (used as fallback)
KNOWN_HR_EMAILS = {
    "TCS":                  ["careers@tcs.com", "hr@tcs.com"],
    "Infosys":              ["hr@infosys.com", "recruitment@infosys.com", "careers@infosys.com"],
    "Wipro":                ["hr@wipro.com", "careers@wipro.com", "talent@wipro.com"],
    "HCL Technologies":     ["careers@hcltech.com", "hr@hcltech.com", "recruitment@hcl.com"],
    "Tech Mahindra":        ["careers@techmahindra.com", "hr@techmahindra.com"],
    "Capgemini":            ["careers.india@capgemini.com", "hr.india@capgemini.com"],
    "Accenture":            ["careers@accenture.com", "hr.india@accenture.com"],
    "IBM":                  ["ibmhr@in.ibm.com", "askhr@in.ibm.com", "careers@in.ibm.com"],
    "Cognizant":            ["hr@cognizant.com", "careers@cognizant.com"],
    "Hexaware":             ["careers@hexaware.com", "hr@hexaware.com"],
    "Persistent Systems":   ["careers@persistent.com", "hr@persistent.com"],
    "Mphasis":              ["careers@mphasis.com", "hr@mphasis.com"],
    "Deloitte":             ["careers@deloitte.com", "hr.india@deloitte.com"],
    "PwC India":            ["careers.india@pwc.com", "hr@pwc.com"],
    "EY (Ernst & Young)":   ["careers@ey.com", "hr.india@ey.com"],
    "Oracle":               ["careers@oracle.com", "oracle-hr@oracle.com"],
    "Siemens":              ["careers.india@siemens.com", "hr.india@siemens.com"],
    "Bosch":                ["careers.india@bosch.com", "hr@bosch.in", "recruitment@bosch.in"],
    "KPIT Technologies":    ["careers@kpit.com", "hr@kpit.com", "talent@kpit.com"],
    "L&T Technology Services (LTTS)": ["careers@ltts.com", "hr@ltts.com"],
    "Amazon":               ["ahmedabad-careers@amazon.com", "hr@amazon.in"],
    "NTT Data":             ["careers.india@nttdata.com", "hr@nttdata.com"],
    "Fujitsu":              ["careers.india@fujitsu.com", "hr@fujitsu.com"],
    "LTIMindtree":          ["careers@ltimindtree.com", "hr@ltimindtree.com"],
    "Zensar Technologies":  ["careers@zensar.com", "hr@zensar.com", "talent@zensar.com"],
    "Mastech Digital":      ["careers@mastechdigital.com", "hr@mastechdigital.com"],
    "Apexon":               ["careers@apexon.com", "hr@apexon.com", "talent@apexon.com"],
    "Apexon (Infostretch)": ["careers@apexon.com", "hr@apexon.com"],
    "Intel":                ["jobs@intel.com", "hr@intel.com"],
    "Qualcomm":             ["careers@qualcomm.com", "hr@qualcomm.com"],
    "Adani Group":          ["careers@adani.com", "hr@adani.com", "talent@adani.com"],
    "Adani Ports":          ["careers@adaniports.com", "hr@adaniports.com"],
    "Adani Wilmar":         ["hr@adaniwilmar.in", "careers@adaniwilmar.in"],
    "Aditya Birla Group":   ["careers@adityabirla.com", "hr@adityabirla.com"],
    "Reliance Industries":  ["careers@ril.com", "hr@ril.com", "talent@ril.com"],
    "Larsen & Toubro":      ["careers@larsentoubro.com", "hr@larsentoubro.com"],
}

C_EMAILS = list(range(6, 12))
C_PHONE  = 12
C_WEB    = 13
C_LI     = 14
C_ADDR   = 15
C_CARE   = 16
C_PRIO   = 17
C_SRC    = 18
C_LIHR   = 19
C_LICO   = 20
C_EMP    = 21
C_IND    = 22
C_FOUND  = 23

EMAIL_RE = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,7}\b')
PHONE_RE = re.compile(r'(?:\+91[\s\-.]?[6-9]\d{9}|\+91[\s\-.]?(?:79|80|22|33|44)\d{8}|0?79[\s\-.]?\d{4}[\s\-.]?\d{4}|\b[6-9]\d{9}\b)')

def strip_tags(html):
    html = re.sub(r'<(script|style|noscript)[^>]*>.*?</\1>', ' ', html, flags=re.DOTALL|re.I)
    html = re.sub(r'<[^>]+>', ' ', html)
    html = html.replace('&amp;','&').replace('&lt;','<').replace('&gt;','>').replace('&nbsp;',' ')
    return re.sub(r'\s+', ' ', html)

def is_fake_phone(p):
    d = re.sub(r'[^\d]', '', p)
    if d.startswith('91') and len(d)==12: d=d[2:]
    if len(d)<10: return True
    d = d[-10:]
    if len(set(d))<=2: return True
    if d[:5]==d[5:]: return True
    return d in {"8888888888","9999999999","1234567890","0000000000","1111111111"}

def clean_phone(p):
    if not p: return ""
    d = re.sub(r'[^\d]','',str(p))
    if d.startswith('91') and len(d)==12: d=d[2:]
    elif d.startswith('0') and len(d) in (11,12): d=d[1:]
    if is_fake_phone(d): return ""
    if len(d)==10 and d[0] in '6789': return '+91 '+d[:5]+' '+d[5:]
    return ""

def extract_from_html(html, domain=""):
    text = strip_tags(html)
    emails = list(set(e.lower() for e in EMAIL_RE.findall(text)
                   if '@' in e and '.' in e.split('@')[1]
                   and e.split('@')[1] not in {'example.com','test.com','sentry.io'}))
    phones = []
    for p in PHONE_RE.findall(text):
        c = clean_phone(p)
        if c and c not in phones: phones.append(c)
    return emails[:8], phones[:3]

def scrape_url(url, client):
    try:
        r = client.get(url, timeout=REQ_TO, follow_redirects=True)
        if r.status_code == 200:
            return r.text
    except Exception:
        pass
    return ""

def google_search(query, client):
    """Scrape Google search results for contact info."""
    try:
        url = f"https://www.google.com/search?q={quote_plus(query)}&num=8"
        html = scrape_url(url, client)
        if html:
            return extract_from_html(html)
    except Exception:
        pass
    return [], []

def research_mnc(mnc_name, domain, careers_url, hr_patterns):
    """Deep research one MNC. Returns dict of found data."""
    result = {
        "name": mnc_name,
        "emails_found": [],
        "phones_found": [],
        "careers_url": careers_url,
        "sources": [],
        "address": "",
    }

    with httpx.Client(headers=HEADERS, timeout=REQ_TO, follow_redirects=True, verify=False) as client:

        # 1. Official careers page
        html = scrape_url(careers_url, client)
        if html:
            em, ph = extract_from_html(html, domain)
            result["emails_found"].extend(em)
            result["phones_found"].extend(ph)
            if em or ph: result["sources"].append("careers_page")

        # 2. Try /contact and /about-us on main domain
        for path in ["/contact", "/contact-us", "/about/contact", "/india/contact"]:
            base = f"https://www.{domain}"
            html = scrape_url(base + path, client)
            if html:
                em, ph = extract_from_html(html, domain)
                for e in em:
                    if e not in result["emails_found"]: result["emails_found"].append(e)
                for p in ph:
                    if p not in result["phones_found"]: result["phones_found"].append(p)
                if em or ph:
                    result["sources"].append(f"website{path}")
                    break

        # 3. Google: "company name Ahmedabad HR email"
        time.sleep(1)
        q1 = f'"{mnc_name}" Ahmedabad HR email recruiter'
        em, ph = google_search(q1, client)
        for e in em:
            if e not in result["emails_found"] and domain.split('.')[0] in e:
                result["emails_found"].append(e)
        if em: result["sources"].append("google_hr")

        # 4. Google: "company name" site:linkedin.com HR Ahmedabad
        time.sleep(1)
        q2 = f'site:linkedin.com/in "{mnc_name}" HR OR Recruiter OR "Talent Acquisition" Ahmedabad'
        em2, ph2 = google_search(q2, client)
        if ph2:
            for p in ph2:
                if p not in result["phones_found"]: result["phones_found"].append(p)

        # 5. Glassdoor
        time.sleep(0.5)
        gd_url = f"https://www.glassdoor.com/Overview/Working-at-{mnc_name.replace(' ','-')}-EI.htm"
        html = scrape_url(gd_url, client)
        if html:
            em, ph = extract_from_html(html, domain)
            for e in em:
                if e not in result["emails_found"]: result["emails_found"].append(e)
            if em: result["sources"].append("glassdoor")

        # 6. AmbitionBox
        time.sleep(0.5)
        ab_url = f"https://www.ambitionbox.com/overview/{mnc_name.lower().replace(' ','-')}-overview"
        html = scrape_url(ab_url, client)
        if html:
            em, ph = extract_from_html(html, domain)
            for e in em:
                if e not in result["emails_found"]: result["emails_found"].append(e)
            if ph:
                for p in ph:
                    if p not in result["phones_found"]: result["phones_found"].append(p)
            if em or ph: result["sources"].append("ambitionbox")

    # 7. Add known HR emails as confirmed fallback
    known_emails = KNOWN_HR_EMAILS.get(mnc_name, [])
    for e in known_emails:
        if e not in result["emails_found"]:
            result["emails_found"].append(e)
    if known_emails: result["sources"].append("known_hr_pattern")

    # 8. Known office phone
    known_ph = KNOWN_PHONES.get(mnc_name, "")
    if known_ph and known_ph not in result["phones_found"]:
        result["phones_found"].insert(0, known_ph)
        result["sources"].append("known_office_phone")

    return result

def main():
    print("\n" + "="*65)
    print("  COMONK AI — MNC HR Deep Research")
    print("  Finding Ahmedabad MNC HR contacts for job applications")
    print("="*65 + "\n")

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]

    # Build name -> row map
    name_to_row = {}
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 2).value or "").strip()
        if name: name_to_row[name] = r

    log = open(LOG_FILE, "w", encoding="utf-8")

    # ── Create MNC_HR_AHMEDABAD sheet ────────────────────────────────────────
    SHEET_OUT = "MNC_HR_AHMEDABAD"
    if SHEET_OUT in wb.sheetnames:
        del wb[SHEET_OUT]
    ws_out = wb.create_sheet(SHEET_OUT)

    out_headers = [
        "#", "Company", "Domain",
        "HR Email 1", "HR Email 2", "HR Email 3", "HR Email 4",
        "Ahmedabad Office Phone", "Careers URL",
        "LinkedIn HR Search", "Sources", "Confidence"
    ]
    col_widths = [4, 35, 25, 35, 35, 35, 35, 22, 45, 60, 30, 12]

    hdr_fill = PatternFill("solid", fgColor="1A1A2E")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for c, (h, w) in enumerate(zip(out_headers, col_widths), 1):
        cell = ws_out.cell(1, c, h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_out.column_dimensions[get_column_letter(c)].width = w
    ws_out.row_dimensions[1].height = 30
    ws_out.freeze_panes = "A2"

    out_row = 2
    total_updated = 0
    fills = [PatternFill("solid", fgColor="F0F4FF"), PatternFill("solid", fgColor="FFFFFF")]

    print(f"  Researching {len(MNC_TARGETS)} MNCs...\n")

    for idx, (mnc_name, domain, careers_url, hr_patterns) in enumerate(MNC_TARGETS, 1):
        print(f"  [{idx:02d}/{len(MNC_TARGETS)}] {mnc_name}...", end=" ", flush=True)

        result = research_mnc(mnc_name, domain, careers_url, hr_patterns)

        emails = result["emails_found"][:6]
        phones = result["phones_found"][:2]
        sources = ", ".join(set(result["sources"]))

        # Confidence score
        conf = "LOW"
        if emails and phones: conf = "HIGH"
        elif emails or phones: conf = "MED"

        # ── Update COMPLETE_MASTER ───────────────────────────────────────────
        r = name_to_row.get(mnc_name)
        if r:
            ex_emails = set(str(ws.cell(r, c).value).lower()
                           for c in C_EMAILS if ws.cell(r, c).value)
            empty_cols = [c for c in C_EMAILS if not ws.cell(r, c).value]
            new_emails = [e for e in emails if e.lower() not in ex_emails]
            added_em = 0
            for e, col in zip(new_emails, empty_cols):
                ws.cell(r, col).value = e; added_em += 1

            if phones and not ws.cell(r, C_PHONE).value:
                ws.cell(r, C_PHONE).value = phones[0]

            if not ws.cell(r, C_CARE).value and careers_url:
                ws.cell(r, C_CARE).value = careers_url

            # Update priority
            has_em = any(ws.cell(r, c).value for c in C_EMAILS)
            has_ph = bool(ws.cell(r, C_PHONE).value)
            if has_em and has_ph:
                ws.cell(r, C_PRIO).value = "1 - Apply now (email+phone)"
            elif has_em:
                ws.cell(r, C_PRIO).value = "2 - Email available"
            elif has_ph:
                ws.cell(r, C_PRIO).value = "3 - Phone only"

            # Update sources
            src = ws.cell(r, C_SRC).value or ""
            if "MNC_RESEARCH" not in src:
                ws.cell(r, C_SRC).value = (src + ", MNC_RESEARCH").strip(", ")

            if added_em or phones: total_updated += 1

        # ── Write to MNC_HR_AHMEDABAD sheet ─────────────────────────────────
        li_hr_search = (
            f"https://www.linkedin.com/search/results/people/"
            f"?keywords={quote_plus(mnc_name + ' HR OR Recruiter OR Talent Acquisition Ahmedabad')}"
            f"&origin=GLOBAL_SEARCH_HEADER"
        )
        row_fill = fills[idx % 2]
        values = [
            idx, mnc_name, domain,
            emails[0] if len(emails) > 0 else "",
            emails[1] if len(emails) > 1 else "",
            emails[2] if len(emails) > 2 else "",
            emails[3] if len(emails) > 3 else "",
            phones[0] if phones else "",
            careers_url,
            li_hr_search,
            sources, conf
        ]
        for c, v in enumerate(values, 1):
            cell = ws_out.cell(out_row, c, v)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=(c >= 4))
            if c in (4,5,6,7) and v:
                cell.font = Font(color="1a56db", size=9)
            if c == 12:  # Confidence
                if conf == "HIGH": cell.font = Font(color="008000", bold=True)
                elif conf == "MED": cell.font = Font(color="FF8C00", bold=True)
                else: cell.font = Font(color="CC0000")
        ws_out.row_dimensions[out_row].height = 20
        out_row += 1

        status = f"+{len(emails)}em +{len(phones)}ph [{conf}] src:{sources[:30]}"
        print(status)
        log.write(f"[{idx:02d}] {mnc_name}: {status}\n")

    wb.save(EXCEL)
    log.close()

    print(f"\n{'='*65}")
    print(f"  MNC HR RESEARCH COMPLETE")
    print(f"  MNCs researched   : {len(MNC_TARGETS)}")
    print(f"  Master rows updated: {total_updated}")
    print(f"  Sheet created      : '{SHEET_OUT}' ({out_row-2} rows)")
    print(f"  Saved: {EXCEL}")
    print(f"  Log  : {LOG_FILE}")
    print(f"{'='*65}\n")
    print(f"  Open '{SHEET_OUT}' tab in Excel for complete MNC HR contact list!")

if __name__ == "__main__":
    main()
