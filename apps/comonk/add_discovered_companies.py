"""
Add all newly discovered AI/ML companies to COMONK_TRUE_MASTER.xlsx
Sources: Clutch, GoodFirms, TopDevelopers, DesignRush, BuiltIn,
         Tracxn, LinkedIn search, web articles, Adzuna jobs
"""

import openpyxl, re, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── All companies discovered from web research ────────────────────────────────
# Format: (Company, Category, AI_Roles, Website_hint, Source)
DISCOVERED = [
    # ── Clutch.co — AI Companies Ahmedabad ─────────────────────────────
    ("fxis.ai",                     "AI / ML",    "AI Solutions, Generative AI, ML Integration",    "fxis.ai",               "Clutch"),
    ("Simform",                     "IT Services","AI/ML, Cloud, Mobile App, Software Dev",          "simform.com",           "Clutch"),
    ("WPWeb Infotech",              "IT Services","AI Integration, Web Dev, ML Solutions",           "wpwebinfotech.com",     "Clutch"),
    ("Intuz",                       "IT Services","AI Solutions, Mobile App, Cloud Dev",             "intuz.com",             "Clutch"),
    ("MindInventory",               "AI / ML",    "AI/ML, Generative AI, Cloud, Microservices",      "mindinventory.com",     "Clutch"),
    ("eSparkBiz",                   "AI / ML",    "AI/ML Development, Computer Vision, NLP",         "esparkbiz.com",         "Clutch"),
    ("Radixweb",                    "IT Services","AI Integration, Web Dev, .NET, Offshore Dev",     "radixweb.com",          "Clutch"),
    ("Tridhya Tech Limited",        "AI / ML",    "Generative AI, AI Solutions, Enterprise Software","tridhyatech.com",       "Clutch"),
    ("iQlance Solutions",           "IT Services","AI, Mobile App, Web Dev, Custom Software",        "iqlance.com",           "Clutch"),
    ("IIH Global",                  "IT Services","AI/ML, Data Analytics, Web Dev",                  "iihglobal.com",         "Clutch"),
    ("SolGuruz",                    "IT Services","AI Integration, Mobile, Web Dev",                 "solguruz.com",          "Clutch"),
    ("AllianceTek",                 "IT Services","AI Solutions, Custom Software, Mobile Dev",       "alliancetek.com",       "Clutch"),
    ("DevsTree IT Services",        "IT Services","AI/ML, React, Node.js, Custom Dev",              "devstree.com",          "Clutch"),
    ("ScaleTech Solutions",         "IT Services","AI Solutions, Scalable Software, Cloud",          "scaletechsolutions.com","Clutch"),
    ("Crest Coder",                 "IT Services","AI Integration, Mobile, Web Dev",                 "crestcoder.com",        "Clutch"),
    ("Citrusbug Technolabs",        "AI / ML",    "AI/ML Development, Computer Vision, NLP",         "citrusbug.com",         "Clutch"),
    ("Sourceved Technologies",      "IT Services","AI/ML, Mobile App, Web Development",              "sourceved.com",         "Clutch"),
    ("Incipient Infotech",          "IT Services","AI/ML, IoT, Cloud Solutions",                     "incipientinfotech.com", "Clutch"),
    ("Inheritx Solutions",          "IT Services","AI, Mobile App, React Native Dev",                "inheritx.com",          "Clutch"),
    ("Hidden Brains InfoTech",      "IT Services","AI/ML, Mobile App, Web Dev, IT Consulting",       "hiddenbrains.com",      "Clutch"),
    ("Moweb Technologies",          "IT Services","AI, Mobile App Dev, React Native",                "mowebtech.com",         "Clutch"),
    ("Anglara Digital Solutions",   "IT Services","AI, Web Dev, Digital Marketing",                  "anglara.com",           "Clutch"),
    ("Acquaint Softtech",           "IT Services","AI/ML, Laravel, Flutter, Custom Dev",             "acquaintsofttech.com",  "Clutch"),
    ("Krishang Technolab",          "IT Services","AI Solutions, Mobile Dev, Shopify",               "krishangtechnolab.com", "Clutch"),
    ("Linearloop",                  "IT Services","AI Integration, SaaS, Mobile Dev",                "linearloop.io",         "Clutch"),
    ("Innovify",                    "AI / ML",    "AI/ML, Generative AI, Product Dev",               "innovify.com",          "Clutch"),
    ("Prismetric",                  "IT Services","AI, Mobile App, Flutter, React Dev",              "prismetric.com",        "Clutch"),
    ("Codeflash Infotech",          "IT Services","AI Integration, Mobile, Flutter Dev",             "codeflashinfotech.com", "Clutch"),
    ("Bitontree",                   "IT Services","AI, Generative AI, Web Dev",                      "bitontree.com",         "Clutch"),
    ("ENQCODE TECHNOLOGIES",        "IT Services","AI, ML, Mobile App Dev",                          "enqcode.com",           "Clutch"),
    ("SoluLab",                     "AI / ML",    "AI/ML, Blockchain, Generative AI, Web3",          "solulab.com",           "Clutch"),
    ("9series Inc",                 "IT Services","AI, Mobile App, React Native, IoT",               "9series.com",           "Clutch"),
    ("Agile Infoways",              "IT Services","AI/ML, Mobile App, Web Dev, IoT",                 "agileinfoways.com",     "Clutch"),
    ("Creole Studios",              "AI / ML",    "AI/ML, Generative AI, Web3, dApps",               "creolestudios.com",     "Clutch"),
    ("WebClues Infotech",           "AI / ML",    "AI-Based Software, ML, Mobile Dev",               "webcluesinfotech.com",  "Clutch"),
    ("Excellent Webworld",          "IT Services","AI Integration, Mobile, Web Dev",                 "excellentwebworld.com", "Clutch"),
    ("Codiste",                     "AI / ML",    "AI/ML, Blockchain, Generative AI, Web3",          "codiste.com",           "Clutch"),
    ("I Can Infotech",              "IT Services","AI, Mobile, Custom Software Dev",                 "icaninfotech.com",      "Clutch"),
    ("Evince Development",          "IT Services","AI Integration, Flutter, React Dev",              "evincedev.com",         "Clutch"),
    ("Facile Technolab",            "IT Services","AI/ML, PHP, Laravel, Web Dev",                   "faciletechnolab.com",   "Clutch"),
    ("Rishabh Software",            "IT Services","AI/ML, Cloud, SAP, Enterprise Dev",              "rishabhsoft.com",       "Clutch"),
    ("TRooTech Business Solutions", "IT Services","AI, Mobile App, React, Angular Dev",             "trootech.com",          "Clutch"),
    ("Web Mavens",                  "IT Services","AI Integration, Web Dev, WordPress",              "webmavens.org",         "Clutch"),
    ("MultiQoS Technologies",       "IT Services","AI, Mobile App, Flutter, React Dev",             "multiqos.com",          "Clutch"),
    ("Moon Technolabs",             "AI / ML",    "AI/ML, NLP, Computer Vision, Predictive AI",     "moontechnolabs.com",    "Clutch"),
    ("Nettyfy Technologies",        "AI / ML",    "ML Development, AI Platform Integration",         "nettyfy.com",           "Clutch"),
    ("Brilworks Software",          "IT Services","AI, Mobile App, Flutter, Custom Dev",             "brilworks.com",         "Clutch"),
    ("Surekha Technologies",        "IT Services","AI, Odoo ERP, Custom Dev",                       "surekhatechnologies.com","Clutch"),
    ("WebMob Technologies",         "IT Services","AI, Mobile App, React, Angular Dev",             "webmobtech.com",        "Clutch"),
    ("The Intellify",               "AI / ML",    "AI/ML, Deep Learning, NLP, Data Science",        "theintellify.com",      "Clutch"),
    ("Peerbits",                    "IT Services","AI, Mobile App, IoT, Custom Dev",                "peerbits.com",          "Clutch"),
    ("Digiqt Technolabs",           "IT Services","AI, Mobile App, Web Dev",                        "digiqt.com",            "Clutch"),
    ("Third Rock Techkno",          "IT Services","AI, Mobile, Web, Custom Software Dev",           "thirdrocktechkno.com",  "Clutch"),
    ("Alian Software",              "IT Services","AI/ML, Web, Mobile App Dev",                     "aliansoftware.com",     "Clutch"),
    ("KoolMind Technolab",          "AI / ML",    "AI, ML, Computer Vision, NLP",                   "koolmind.in",           "Clutch"),
    ("Saurabh Infosys",             "IT Services","AI, ERP, Mobile, Web Dev",                      "saurabhinfosys.com",    "Clutch"),
    ("IT Path Solutions",           "IT Services","AI Integration, Mobile App, Odoo Dev",           "itpathsolutions.com",   "Clutch"),
    ("Auxano Global Services",      "IT Services","AI, Mobile App, Web Dev, Blockchain",            "auxanoglobalservices.com","Clutch"),
    ("IOTTIVE",                     "AI / ML",    "IoT + AI, Embedded AI, Industrial AI",           "iottive.com",           "Clutch"),
    ("Zealous System",              "IT Services","AI, Mobile App, Custom Software Dev",             "zealousys.com",         "Clutch"),
    ("BOSC Tech Labs",              "IT Services","AI, Ruby on Rails, DevOps, Cloud Dev",           "bosctechlabs.com",      "Clutch"),
    ("Neuramonks",                  "AI / ML",    "AI, Data Science, NLP, ML Development",          "neuramonks.com",        "Clutch"),
    ("Theoddcoders Technologies",   "AI / ML",    "AI Development, Generative AI, LLM",             "theoddcoders.com",      "Clutch"),
    ("Thinkwik",                    "IT Services","AI, Mobile App, Flutter, React Dev",             "thinkwik.com",          "Clutch"),
    ("Lemolite Technologies",       "IT Services","AI Integration, iOS, Android, Web Dev",          "lemolite.com",          "Clutch"),
    ("OpenXcell",                   "IT Services","AI, Mobile App, Web Dev, Blockchain",            "openxcell.com",         "Clutch"),
    ("SoluteLabs",                  "AI / ML",    "AI/ML, Deep Learning, NLP, Product Dev",         "solutelabs.com",        "Clutch"),
    ("ChaitanyaAi",                 "AI / ML",    "Artificial Intelligence, Generative AI",          "chaitanyaai.in",        "Clutch"),
    ("RadheApps",                   "IT Services","AI, Mobile App Development",                     "radheapps.com",         "Clutch"),
    # ── TopDevelopers.co ────────────────────────────────────────────────
    ("ViitorCloud Technologies",    "IT Services","AI, Digital Transform, Cloud, IoT",              "viitorcloud.com",       "TopDev"),
    ("SPEC INDIA",                  "IT Services","AI-driven Software, Enterprise, ERP",            "spec-india.com",        "TopDev"),
    ("Thinkwik Technology",         "IT Services","AI, Mobile, Flutter, Web Dev",                   "thinkwik.com",          "TopDev"),
    ("Concetto Labs",               "IT Services","AI, Mobile App, Node.js, React Dev",             "concettolabs.com",      "TopDev"),
    ("Inferenz",                    "AI / ML",    "ML, Deep Learning, NLP, AI Solutions",            "inferenz.in",           "TopDev"),
    ("Rysun Labs",                  "IT Services","AI, Data Engineering, Cloud Dev",                "rysunlabs.com",         "TopDev"),
    ("INEXTURE Solutions",          "IT Services","AI, Python, Django, Mobile Dev",                 "inexture.com",          "TopDev"),
    ("EncodeDots",                  "IT Services","AI Integration, Mobile, Web Dev",                "encodedots.com",        "TopDev"),
    ("Quixom Technology",           "AI / ML",    "AI/ML, Computer Vision, NLP, Automation",        "quixomtechnology.com",  "TopDev"),
    ("Monarch Innovation",          "IT Services","AI, Mobile App, Custom Software Dev",            "monarchinnovation.com", "TopDev"),
    ("Pranshtech Solutions",        "IT Services","AI, Mobile, Web, Custom Dev",                    "pranshtech.com",        "TopDev"),
    ("Ninja Techno Labs",           "IT Services","AI, Mobile App, Flutter Dev",                    "ninjatechnolabs.com",   "TopDev"),
    ("AQe Digital",                 "IT Services","AI, Web Dev, Digital Marketing",                 "aqedigital.com",        "TopDev"),
    ("Triveni Global Software",     "IT Services","AI, ERP, Custom Software Dev",                   "triveniglobal.com",     "TopDev"),
    ("RaftLabs",                    "IT Services","AI, SaaS, Web, Mobile Dev",                     "raftlabs.co",           "TopDev"),
    ("Matics Analytics",            "AI / ML",    "Data Analytics, AI, Business Intelligence",      "matics.in",             "TopDev"),
    ("CodesClue",                   "AI / ML",    "AI-First Software, ML, Blockchain, Data",        "codesclue.com",         "TopDev"),
    ("IntelloNix",                  "AI / ML",    "AI/ML, NLP, Computer Vision, Deep Learning",    "intellonix.com",        "TopDev"),
    ("AARCHIK SOLUTIONS",           "IT Services","AI, Custom Dev, Web Solutions",                  "aarchik.com",           "TopDev"),
    ("InciPulse",                   "IT Services","AI Integration, Mobile Dev",                     "incipulse.com",         "TopDev"),
    ("Corefragment Technologies",   "IT Services","AI, Web Dev, Custom Software",                   "corefragment.com",      "TopDev"),
    # ── DesignRush ──────────────────────────────────────────────────────
    ("Bytes Technolab",             "IT Services","AI, Mobile App, Web Dev, Blockchain",            "bytestechnolab.com",    "DesignRush"),
    ("Magneto IT Solutions",        "IT Services","AI, eCommerce, Magento, Web Dev",               "magnetoitsolutions.com","DesignRush"),
    ("Probey Services",             "IT Services","AI, Digital Marketing, Web Dev",                 "probeyservices.com",    "DesignRush"),
    ("AddWebSolution",              "IT Services","AI, Mobile App, Laravel, Web Dev",              "addwebsolution.com",    "DesignRush"),
    ("Solution Analysts",           "IT Services","AI, Mobile App, Web Dev, IoT",                  "solutionanalysts.com",  "DesignRush"),
    ("Nadcab Labs",                 "IT Services","AI, Blockchain, Web3, Custom Dev",              "nadcablabs.com",        "DesignRush"),
    ("Techforce Global",            "IT Services","AI, Salesforce, Dev, Cloud",                    "techforceglobal.com",   "DesignRush"),
    ("Rushkar Technology",          "IT Services","AI, Mobile App, Web Dev",                       "rushkar.com",           "DesignRush"),
    ("Avidclan",                    "IT Services","AI, Mobile Dev, UI/UX",                         "avidclan.com",          "DesignRush"),
    ("Solvios Technology",          "AI / ML",    "AI, ML, Data Science, Custom Software",         "solvios.com",           "DesignRush"),
    ("Saeculum Solutions",          "IT Services","AI, Mobile App, Web Dev",                       "saeculumsolutions.com", "DesignRush"),
    ("aPurple",                     "IT Services","AI, Mobile App, Flutter Dev",                   "apurple.com",           "DesignRush"),
    ("WireFuture",                  "IT Services","AI Integration, React, Mobile Dev",             "wirefuture.com",        "DesignRush"),
    ("Softweb Solutions",           "AI / ML",    "AI/ML, IoT, Cloud, Enterprise AI",              "softwebsolutions.com",  "DesignRush"),
    ("CartCoders",                  "IT Services","AI, eCommerce, Shopify, WooCommerce",           "cartcoders.com",        "DesignRush"),
    ("Vrinsoft Technology",         "IT Services","AI, Mobile App, Web Dev",                       "vrinsofttechnology.com","DesignRush"),
    ("Metwaves Technologies",       "AI / ML",    "AI/ML, IoT, Embedded AI, Industrial AI",        "metwave.com",           "DesignRush"),
    ("360 Degree Technosoft",       "IT Services","AI, Mobile App, Web Dev",                       "360degreecloud.com",    "DesignRush"),
    ("Biztech Consulting",          "IT Services","AI, Odoo ERP, Mobile, Web Dev",                 "biztechconsultancy.com","DesignRush"),
    ("Elightwalk Technology",       "IT Services","AI Integration, Mobile, Web Dev",               "elightwalk.com",        "DesignRush"),
    ("Green Apex",                  "IT Services","AI, Mobile App, Web Dev",                       "greenapex.in",          "DesignRush"),
    ("ManekTech",                   "IT Services","AI, Mobile App, Web Dev, IoT",                  "manektech.com",         "DesignRush"),
    ("Techuz",                      "IT Services","AI, Mobile App, Flutter, React Dev",            "techuz.com",            "DesignRush"),
    ("Multiverse Software",         "IT Services","AI, Web Dev, Custom Software",                  "multiversesoftware.com","DesignRush"),
    ("Intelliswift",                "IT Services","AI, Staffing, IT Services",                     "intelliswift.com",      "DesignRush"),
    ("Aviato Consulting",           "AI / ML",    "AI Consulting, Data Science, Analytics",        "aviatoconsulting.com",  "DesignRush"),
    ("Uncanny Consulting",          "IT Services","AI, SAP, ERP, Custom Dev",                     "uncannyconsulting.com", "DesignRush"),
    # ── Tracxn (funded AI startups) ────────────────────────────────────
    ("Indian TTS",                  "AI / ML",    "Text-to-Speech AI, Voice AI, NLP",              "indiantts.com",         "Tracxn"),
    ("drivebuddyAI",                "AI / ML",    "Automotive AI, Driver Safety AI, Computer Vision","drivebuddy.ai",       "Tracxn"),
    ("Metis Intel",                 "AI / ML",    "AI Analytics, Business Intelligence",            "metisintel.com",        "Tracxn"),
    ("Glib",                        "AI / ML",    "Language AI, Communication AI",                  "glib.ai",               "Tracxn"),
    ("Rootle",                      "AI / ML",    "AI Solutions, ML Platform",                      "rootle.ai",             "Tracxn"),
    ("Veloxhire.AI",                "AI / ML",    "AI Hiring, HR AI, Recruitment AI",              "veloxhire.ai",          "Tracxn"),
    ("JetCounter",                  "AI / ML",    "AI Analytics, Counting AI, Computer Vision",    "jetcounter.io",         "Tracxn"),
    ("mxface",                      "AI / ML",    "Face Recognition AI, Biometric AI",             "mxface.ai",             "Tracxn"),
    ("WotNot",                      "AI / ML",    "AI Chatbot, Conversational AI, NLP",            "wotnot.io",             "Tracxn"),
    ("Tusker AI",                   "AI / ML",    "AI Solutions, Enterprise AI",                    "tuskerai.com",          "Tracxn"),
    ("Cygnet Infotech",             "IT Services","AI/ML, RPA, Digital Transform, SAP",            "cygnetinfotech.com",    "Tracxn"),
    # ── LinkedIn Search ─────────────────────────────────────────────────
    ("Brainy Neurals",              "AI / ML",    "AI Development, Healthcare AI, Manufacturing AI","brainyneurals.com",     "LinkedIn"),
    ("Arihant AI",                  "AI / ML",    "AI-powered ERP, Enterprise AI Solutions",        "arihantai.com",         "LinkedIn"),
    ("Dev Information Technology",  "IT Services","AI, IT Services, Consulting, ISO 9001",          "devinfo.net",           "LinkedIn"),
    ("Genesis Artificial Intelligence","AI / ML", "AI/ML Solutions, Enterprise AI",                "genesisai.in",          "LinkedIn"),
    ("Future Development AI (FDAI)","AI / ML",    "XR + AI, Embedded AI, Intelligent Hardware",    "fdai.co.in",            "LinkedIn"),
    ("ImmverseAI",                  "AI / ML",    "Immersive AI, XR + AI, Avatar AI",              "immverseai.com",        "LinkedIn"),
    ("AI Automation Labs",          "AI / ML",    "AI Automation, RPA + AI, Process AI",           "aiautomationlabs.io",   "LinkedIn"),
    # ── BuiltIn / News articles ─────────────────────────────────────────
    ("Prescinto",                   "AI / ML",    "Renewable Energy AI, Solar/Wind AI, Predictive Analytics","prescinto.ai",  "BuiltIn"),
    ("Prepseed",                    "AI / ML",    "EdTech AI, LLM for Education, Adaptive Learning","prepseed.com",         "BuiltIn"),
    ("Yaraa AI",                    "AI / ML",    "Conversational AI, Voice AI, NLP",              "yaraa.io",              "BuiltIn"),
    ("DataDwip",                    "AI / ML",    "Data Scraping, AI/ML Data Intelligence, Annotation","datadwip.com",       "BuiltIn"),
    ("OrcaMinds",                   "AI / ML",    "Generative AI, LLM Integration, Computer Vision, NLP","orcaminds.in",   "Web"),
    ("Infocusp Innovations",        "AI / ML",    "Data-Driven AI, Healthcare AI, eCommerce AI",   "infocusp.com",          "Web"),
    # ── mydigitalcrown / other web ──────────────────────────────────────
    ("Metic.ai",                    "AI / ML",    "AI Solutions, ML Development, NLP",             "metic.ai",              "Web"),
    ("Fusion Informatic",           "IT Services","AI, ERP, Custom Software, Web Dev",             "fusioninformatic.com",  "Web"),
    ("YM Intelligence Tech",        "AI / ML",    "AI/ML Engineering, Data Science",               "ymitech.com",           "Web"),
    ("Kshatra Infotech",            "AI / ML",    "AI-based Web/Mobile App, ML Dev",               "kshatra.com",           "Web"),
    ("Spectrics Solutions",         "AI / ML",    "AI/ML Development, Healthcare AI, Fintech AI",  "spectricssolutions.com","Web"),
    ("Prioxis Technologies",        "AI / ML",    "AI/ML Dev, Deep Learning, Computer Vision",     "prioxis.com",           "Web"),
    ("Yudiz Solutions",             "IT Services","AI, Blockchain, Game Dev, Mobile Dev",          "yudiz.com",             "Web"),
    ("StreamSpace AI",              "AI / ML",    "AI Streaming, Generative AI, LLM",              "streamspaceai.com",     "Adzuna"),
    ("Arihant AI",                  "AI / ML",    "AI ERP, Enterprise AI Solutions",               "arihantai.com",         "Adzuna"),
    ("MeshTek Labs",                "AI / ML",    "AI/ML Research, Deep Learning, NLP",            "meshtek.in",            "Adzuna"),
    ("Growexx",                     "IT Services","AI, Data Science, Digital Marketing",           "growexx.com",           "BuiltIn"),
    ("Uniquesdata",                 "AI / ML",    "Data Science, ML, AI Analytics",                "uniquesdata.com",       "BuiltIn"),
    ("Webelight Solutions",         "IT Services","AI, Mobile App, Web Dev, Blockchain",           "webelight.co.in",       "BuiltIn"),
    ("Analytics8",                  "AI / ML",    "Data Analytics, BI, ML, AI Consulting",         "analytics8.com",        "BuiltIn"),
    ("F(x) Data Labs",              "AI / ML",    "AI, Cloud Consulting, ERP, Data Science",       "fxdatalabs.com",        "Web"),
    ("Viston AI",                   "AI / ML",    "Enterprise AI, AI Strategy, Predictive Analytics","vistonai.com",        "Web"),
    ("NextGenSoft Technologies",    "AI / ML",    "AI, Cloud-native, DevOps, Data Engineering",    "nextgensofttech.com",   "Web"),
    ("DataSlush",                   "AI / ML",    "Full Stack Data & AI, ML, Data Engineering",    "dataslush.com",         "Web"),
    ("Nexgits",                     "AI / ML",    "AI/ML, AR/VR, WebGL, Enterprise AI",            "nexgits.com",           "Web"),
    ("Digital Innovations (DGTL)",  "AI / ML",    "Predictive Analytics, NLP, Computer Vision",    "dgtlinnovations.in",    "Web"),
    ("Technomet Solutions",         "AI / ML",    "AI/ML Development, Gujarat",                     "technomet.in",          "Web"),
    ("Vridhi Infotech",             "IT Services","AI, Web Dev, Digital Marketing",                "vridhiinfotech.com",    "Web"),
]

def norm(s):
    if not s: return ""
    s = unicodedata.normalize('NFKD', str(s).lower())
    return re.sub(r'[^a-z0-9]', '', s)

def main():
    wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
    ws = wb["COMPLETE_MASTER"]

    # Build existing name set
    existing = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 2).value
        if n: existing[norm(n)] = r

    # Column map from master: 1=#, 2=Company, 3=City, 4=Category, 5=AI Roles,
    # 6-11=Emails, 12=Phone, 13=Website, 14=LinkedIn, 15=Address, 16=Careers, 17=Priority, 18=Sources
    added = 0
    updated = 0

    fill_new = PatternFill("solid", fgColor="EAF4FF")

    for company, category, roles, website_hint, source in DISCOVERED:
        key = norm(company)
        if not key: continue

        website = f"https://www.{website_hint}" if website_hint and not website_hint.startswith("http") else website_hint

        if key in existing:
            # Enrich existing row
            row = existing[key]
            if not ws.cell(row, 4).value: ws.cell(row, 4).value = category
            if not ws.cell(row, 5).value: ws.cell(row, 5).value = roles
            if not ws.cell(row, 13).value and website: ws.cell(row, 13).value = website
            # Add source
            cur_src = ws.cell(row, 18).value or ""
            if source not in cur_src:
                ws.cell(row, 18).value = (cur_src + ", " + source).strip(", ")
            updated += 1
        else:
            # Add new row
            next_row = ws.max_row + 1
            ws.cell(next_row, 1, ws.max_row)  # serial number
            ws.cell(next_row, 2, company)
            ws.cell(next_row, 3, "Ahmedabad")
            ws.cell(next_row, 4, category)
            ws.cell(next_row, 5, roles)
            ws.cell(next_row, 13, website)
            ws.cell(next_row, 18, source)

            for c in range(1, 19):
                cell = ws.cell(next_row, c)
                cell.fill = fill_new
                cell.alignment = Alignment(vertical="center", wrap_text=(c in (5, 15)))
                if c == 2:
                    cell.font = Font(bold=True, color="0A3D62")
                if c in (6,7,8,9,10,11) and cell.value:
                    cell.font = Font(color="1a56db", size=9)
            ws.row_dimensions[next_row].height = 18

            existing[key] = next_row
            added += 1

    # Renumber column 1
    for i, r in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(r, 1, i)

    # Update AI_ML_ONLY sheet too
    ws2 = wb["AI_ML_ONLY"]
    ai_kw = ['ai','ml','data','machine','deep','nlp','computer vision','llm','genai','analytics']
    # Rebuild AI_ML_ONLY from COMPLETE_MASTER
    # Clear existing data rows
    for r in range(ws2.max_row, 1, -1):
        ws2.delete_rows(r)

    ai_row = 2
    fill_a = PatternFill("solid", fgColor="F5F0FF")
    fill_b = PatternFill("solid", fgColor="FFFFFF")
    fill_n = PatternFill("solid", fgColor="EAF4FF")
    for r in range(2, ws.max_row + 1):
        cat  = str(ws.cell(r, 4).value or "").lower()
        role = str(ws.cell(r, 5).value or "").lower()
        src  = str(ws.cell(r, 18).value or "")
        if any(k in cat or k in role for k in ai_kw):
            fill = fill_n if any(s in src for s in ["Clutch","TopDev","DesignRush","Tracxn","LinkedIn","BuiltIn","Web"]) else (fill_a if ai_row%2==0 else fill_b)
            for c in range(1, 19):
                v = ws.cell(r, c).value
                cell2 = ws2.cell(ai_row, c, v)
                cell2.fill = fill
                cell2.alignment = Alignment(vertical="center", wrap_text=(c in (5,15)))
                if c in (6,7,8,9,10,11) and v:
                    cell2.font = Font(color="1a56db", size=9)
            ws2.cell(ai_row, 1, ai_row - 1)
            ws2.row_dimensions[ai_row].height = 18
            ai_row += 1

    # Update stats sheet
    ws3 = wb["STATS"]
    total = ws.max_row - 1
    ai_count = ws2.max_row - 1
    ws3.cell(3, 2, total)
    ws3.cell(6, 2, ai_count)

    wb.save("COMONK_TRUE_MASTER.xlsx")
    print(f"\n  NEW companies added:   {added}")
    print(f"  Existing rows updated: {updated}")
    print(f"  Total in master now:   {total}")
    print(f"  AI/ML companies:       {ai_count}")
    print(f"  Saved: COMONK_TRUE_MASTER.xlsx")

if __name__ == "__main__":
    main()
