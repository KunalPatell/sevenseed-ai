"""
FINAL GAP CLOSE — guarantee EVERY source company name is in the workbook.
1. Add legit MNC + IT/tech companies to COMPLETE_MASTER.
2. Dump ALL source company names into 'All_Source_Companies' sheet (reference).
"""
import openpyxl, re, csv, os, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment

MASTER="COMONK_TRUE_MASTER.xlsx"
def norm(s):
    if not s: return ""
    return re.sub(r'[^a-z0-9]','',unicodedata.normalize('NFKD',str(s).lower()))

# tech/IT indicators -> auto-add to main master
TECH_KW = ["tech","soft","infotech","systems","solution","labs","lab","ai","digital",
           "data","analytics","web","app","infosys","cyber","cloud","consulting",
           "studios","studio","innovar","innovation","technolab","computer","infocom",
           "infomedia","ventures","global","networks","robotics","automation"]
# Known MNCs (name fragments) -> always add
MNC = ["cognizant","deloitte","ernst","young","hcl","kpit","l&t","ltts","mphasis",
       "ntt data","oracle","persistent","pwc","pricewaterhouse","fujitsu","sophos",
       "tata electronics","amazon","jabil","skf","milacron","linde","s&p global",
       "dun & bradstreet","arrow","apexon","mastech","gartner","searce","genesys"]
# Non-tech sectors -> keep ONLY in reference sheet (don't pollute master)
NONTECH = ["textile","pharma","steel","forging","chemical","food","jewel","jewell",
           "paper","cotton","yarn","spinning","fashion","fabric","granito","ceramic",
           "agro","agri","seeds","sugar","oil","beverage","brewer","distiller","tea",
           "tobacco","leasing","finance","securit","capital","investment","realty",
           "infra","construct","developer","builder","transformer","gas","petro",
           "alloy","metal","tubes","container","packaging","plastic","polymer","tannery",
           "fert","biotec","biotech","lifescience","lifecare","healthcare","pharmaceutical",
           "wagon","rail","motors","furnish","trading","traders","wool","denim","namkeen",
           "diamond","refiner","minerals","bearings","valve","tools","furnace","wires"]

def classify(name):
    n=name.lower()
    if any(m in n for m in MNC): return "MNC"
    if any(k in n for k in NONTECH): return "skip"
    # fragment / junk: too generic single words
    if norm(name) in {"finance","healthcare","technology","technologies","solutions",
        "industries","company","agreement","agencies","services","software","steel",
        "systems","platform","mobility","energy","power","media","entertainment",
        "infrastructure","infotech","consultancy","corporation","developers","exports",
        "holdings","international","national","global","business","automation","defense",
        "devices","beverages","breweries","gases","pipes","polymers","realty","refractories",
        "credit","credits","close","monday","under","national","communication","instrumentation"}:
        return "junk"
    if any(k in n for k in TECH_KW): return "tech"
    return "other"

# Collect all source company names
src={}
def addn(n):
    if n and isinstance(n,str):
        n=re.sub(r'\s+',' ',n).strip()
        k=norm(n)
        if k and len(k)>=4 and not n.startswith("http") and "@" not in n:
            src.setdefault(k,n)
for fn in ["AIML_Companies_Apply_List.csv","Gandhinagar_GIFTCity_Companies.csv","NEW_COMPANIES_DISCOVERED.csv"]:
    if os.path.exists(fn):
        for row in csv.DictReader(open(fn,encoding="utf-8-sig",errors="ignore")):
            for k in row:
                if k and k.strip().lower() in ("company","company name"): addn(row[k])
def excel_co(fn,col,hr=1):
    if not os.path.exists(fn): return
    w=openpyxl.load_workbook(fn,read_only=True)
    for s in w.sheetnames:
        for i,row in enumerate(w[s].iter_rows(values_only=True),1):
            if i<=hr: continue
            if len(row)>=col and row[col-1]: addn(row[col-1])
    w.close()
excel_co("HR Mail List.xlsx",1,1); excel_co("Ahmedabad_AIML_Companies_HR_Contacts.xlsx",2,1)
excel_co("Ahmedabad_IT_AIML_Companies_Master.xlsx",2,3); excel_co("Ahmedabad_IT_AIML_FINAL_MASTER.xlsx",2,3)
excel_co("AI_Engineer_Job_Targets.xlsx",2,1)
if os.path.exists("_extract_datascience_compeny_list_pdf.txt"):
    for line in open("_extract_datascience_compeny_list_pdf.txt",encoding="utf-8"):
        m=re.match(r'^\d+\.\s+(.+)',line.strip())
        if m: addn(m.group(1))

wb=openpyxl.load_workbook(MASTER)
ws=wb["COMPLETE_MASTER"]
master={norm(ws.cell(r,2).value) for r in range(2,ws.max_row+1) if ws.cell(r,2).value}

# 1. Add legit MNC + tech missing companies to master
added=0
fill=PatternFill("solid",fgColor="EAFFF0")
for k,name in src.items():
    if k in master: continue
    cls=classify(name)
    if cls in ("MNC","tech"):
        r=ws.max_row+1
        cat="AI / ML" if any(x in name.lower() for x in ["ai","data","analytics","ml","intelligence"]) else "IT Services"
        ws.cell(r,1,r-1); ws.cell(r,2,name); ws.cell(r,3,"Ahmedabad")
        ws.cell(r,4,cat); ws.cell(r,18, "MNC-list" if cls=="MNC" else "Discovered")
        for c in range(1,19):
            ws.cell(r,c).fill=fill; ws.cell(r,c).alignment=Alignment(vertical="center")
        master.add(k); added+=1

# 2. All_Source_Companies reference sheet
if "All_Source_Companies" in wb.sheetnames: del wb["All_Source_Companies"]
wsa=wb.create_sheet("All_Source_Companies")
for c,h in enumerate(["#","Company Name (raw from sources)","In Master?","Type"],1):
    cell=wsa.cell(1,c,h); cell.fill=PatternFill("solid",fgColor="2E1A47")
    cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(horizontal="center")
wsa.column_dimensions['B'].width=46; wsa.column_dimensions['C'].width=12; wsa.column_dimensions['D'].width=14
wsa.freeze_panes="A2"
rr=2
for k,name in sorted(src.items(), key=lambda x:x[1].lower()):
    inm = "YES" if k in master else "no"
    wsa.cell(rr,1,rr-1); wsa.cell(rr,2,name); wsa.cell(rr,3,inm); wsa.cell(rr,4,classify(name))
    if rr%2==0:
        for c in range(1,5): wsa.cell(rr,c).fill=PatternFill("solid",fgColor="F5F0FF")
    rr+=1
total_src=rr-2

# Renumber + rebuild AI_ML_ONLY + stats
for i,r in enumerate(range(2,ws.max_row+1),1): ws.cell(r,1,i)
total=ws.max_row-1; maxc=ws.max_column
ws2=wb["AI_ML_ONLY"]
for r in range(ws2.max_row,1,-1): ws2.delete_rows(r)
ai_kw=['ai','ml','data','machine','deep','nlp','computer vision','llm','genai','analytics','automation','intelligence','learning','aiops','analyst']
ar=2; fa=PatternFill("solid",fgColor="F5F0FF"); fb=PatternFill("solid",fgColor="FFFFFF")
EMAILS=[6,7,8,9,10,11]
for r in range(2,ws.max_row+1):
    cat=str(ws.cell(r,4).value or "").lower(); role=str(ws.cell(r,5).value or "").lower()
    if any(k in cat or k in role for k in ai_kw):
        fl=fa if ar%2==0 else fb
        for c in range(1,maxc+1):
            v=ws.cell(r,c).value; cc=ws2.cell(ar,c,v); cc.fill=fl
            cc.alignment=Alignment(vertical="center")
            if c in EMAILS and v: cc.font=Font(color="1a56db",size=9)
        ws2.cell(ar,1,ar-1); ar+=1
ai_count=ws2.max_row-1
wb["STATS"].cell(3,2,total); wb["STATS"].cell(6,2,ai_count)
wb.save(MASTER)
print(f"  Added to master (MNC+tech): +{added}")
print(f"  All_Source_Companies sheet: {total_src} company names (EVERY source name)")
print(f"  In master: {sum(1 for k in src if k in master)} | reference-only: {total_src-sum(1 for k in src if k in master)}")
print(f"\n  TOTAL master companies: {total} | AI/ML: {ai_count}")
print(f"  Sheets: {wb.sheetnames}")
