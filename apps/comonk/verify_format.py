import openpyxl
wb=openpyxl.load_workbook("Ahmedabad_IT_AIML_FINAL_MASTER.xlsx", read_only=True)
print("Sheet order:", wb.sheetnames)
for sn in wb.sheetnames:
    ws=wb[sn]
    hdr=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
    resid=0
    for r in range(1, min(ws.max_row,1000)+1):
        for c in range(1, ws.max_column+1):
            v=ws.cell(r,c).value
            if v and chr(0xFFFD) in str(v): resid+=1
    print(f"\n{sn}: {ws.max_row-1} data rows, {ws.max_column} cols, junk-chars(first1000)={resid}")
    print("  headers:", hdr[:8], ("... +%d email cols" % sum(1 for h in hdr if h and 'mail' in str(h).lower())) if any('mail' in str(h).lower() for h in hdr) else "")
