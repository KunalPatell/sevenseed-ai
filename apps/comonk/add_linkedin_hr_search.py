"""
OPTION 3: LinkedIn HR Finder
Generates ready-to-click LinkedIn people-search URLs for each company
that surface HR / Recruiter / Talent Acquisition profiles.
Adds a 'LinkedIn HR Search' column to COMONK_TRUE_MASTER.xlsx
100% legal — just builds search links you click and connect manually.
"""

import openpyxl
from urllib.parse import quote
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL = "COMONK_TRUE_MASTER.xlsx"

def li_people_search(company):
    """LinkedIn people search: HR/recruiter at this company in Ahmedabad."""
    kw = f'{company} (HR OR Recruiter OR "Talent Acquisition" OR Hiring)'
    return f"https://www.linkedin.com/search/results/people/?keywords={quote(kw)}&origin=GLOBAL_SEARCH_HEADER"

def li_company_people(company):
    """LinkedIn: search company page -> People tab (find employees)."""
    return f"https://www.linkedin.com/search/results/companies/?keywords={quote(company)}"

def process_sheet(ws, hr_col, comp_col=2):
    """Add LinkedIn HR Search + Company Page columns."""
    # Header
    ws.cell(1, hr_col, "LinkedIn HR Search")
    ws.cell(1, hr_col + 1, "LinkedIn Company Page")
    h_fill = PatternFill("solid", fgColor="0A66C2")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    for c in (hr_col, hr_col + 1):
        cell = ws.cell(1, c)
        cell.fill = h_fill; cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = 34

    filled = 0
    for r in range(2, ws.max_row + 1):
        company = ws.cell(r, comp_col).value
        if not company:
            continue
        company = str(company).strip()
        # HR search URL
        c1 = ws.cell(r, hr_col, li_people_search(company))
        c1.font = Font(color="0A66C2", size=9, underline="single")
        c1.hyperlink = li_people_search(company)
        # Company page URL
        c2 = ws.cell(r, hr_col + 1, li_company_people(company))
        c2.font = Font(color="0A66C2", size=9, underline="single")
        c2.hyperlink = li_company_people(company)
        filled += 1
    return filled

def main():
    wb = openpyxl.load_workbook(EXCEL)

    # COMPLETE_MASTER — add after last column (currently 18 = Sources)
    ws = wb["COMPLETE_MASTER"]
    hr_col = ws.max_column + 1
    n1 = process_sheet(ws, hr_col)
    print(f"  COMPLETE_MASTER: {n1} LinkedIn HR search links added (col {get_column_letter(hr_col)})")

    # AI_ML_ONLY — same
    ws2 = wb["AI_ML_ONLY"]
    hr_col2 = ws2.max_column + 1
    n2 = process_sheet(ws2, hr_col2)
    print(f"  AI_ML_ONLY: {n2} LinkedIn HR search links added (col {get_column_letter(hr_col2)})")

    wb.save(EXCEL)
    print(f"\n  Saved: {EXCEL}")
    print("  Har company ke aage ab 2 clickable LinkedIn links hain:")
    print("    1. LinkedIn HR Search   -> us company ke HR/recruiter profiles")
    print("    2. LinkedIn Company Page -> company ka official page -> People tab")

if __name__ == "__main__":
    main()
