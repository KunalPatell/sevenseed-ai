import openpyxl
wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
ws = wb["COMPLETE_MASTER"]

C_EMAILS = list(range(6, 12))
C_PHONE = 12
C_WEB = 13

total = 0
with_email = 0
with_phone = 0
with_both = 0
missing_both = []

for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 2).value
    cat  = str(ws.cell(r, 4).value or "")
    ai   = str(ws.cell(r, 5).value or "")
    if not name:
        continue
    is_ai = any(kw in cat.lower() or kw in ai.lower()
                for kw in ["ai", "ml", "machine learning", "deep learning",
                           "nlp", "computer vision", "data sci"])
    if not is_ai:
        continue
    total += 1
    em = any(ws.cell(r, c).value for c in C_EMAILS)
    ph = ws.cell(r, C_PHONE).value
    web = ws.cell(r, C_WEB).value
    if em: with_email += 1
    if ph: with_phone += 1
    if em and ph: with_both += 1
    if not em and not ph:
        missing_both.append((r, str(name), str(web or "")))

print(f"AI/ML companies total     : {total}")
print(f"  With email              : {with_email} ({with_email*100//total}%)")
print(f"  With phone              : {with_phone} ({with_phone*100//total}%)")
print(f"  With both               : {with_both} ({with_both*100//total}%)")
print(f"  Missing both (email+ph) : {len(missing_both)}")
print(f"  ...of those, with website: {sum(1 for _,_,w in missing_both if w and w != 'None')}")

# Save list for next enrichment pass
with open("_aiml_missing.txt", "w", encoding="utf-8") as f:
    for r, name, web in missing_both:
        f.write(f"{r}\t{name}\t{web}\n")
print("\nSaved missing list to _aiml_missing.txt")
