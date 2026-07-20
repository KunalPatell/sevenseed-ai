"""
Add IT companies (not just AI/ML) from official directories & informative sources:
GESIA (Gujarat IT Association), BuiltIn, GIFT City lists, Gandhinagar directories,
iNDEXTb, ASDM, mygate, Decentro, La Net Team, Cutshort IT directory.
"""

import openpyxl, re, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment

# (Company, Category, Roles, Website, City, Source)
IT_COMPANIES = [
    # ── BuiltIn / mygate — Top IT Ahmedabad ──────────────────────────────
    ("IndiaNIC Infotech",       "IT Services","Mobile App, Web Dev, IoT, Custom Software",        "indianic.com",          "Ahmedabad", "BuiltIn"),
    ("Brainvire Infotech",      "IT Services","AI, eCommerce, Digital Transform, Cloud",          "brainvire.com",         "Ahmedabad", "BuiltIn"),
    ("Cygnet Infotech",         "IT Services","AI/ML, RPA, Tax Tech, Digital Transform",          "cygnet-infotech.com",   "Ahmedabad", "BuiltIn"),
    ("Cybage Software",         "IT Services","AI, Product Engineering, Data Analytics",          "cybage.com",            "Ahmedabad", "BuiltIn"),
    ("Crest Data Systems",      "IT Services","AI/ML, Big Data, DevOps, Cloud, Security",          "crestdatasys.com",      "Ahmedabad", "BuiltIn"),
    ("The Red Eyes",            "IT Services","Web Dev, Mobile App, Digital Marketing",           "theredeyes.com",        "Ahmedabad", "ASDM"),
    ("Silver Touch Technologies","IT Services","AI, ERP, eGov, Enterprise Software",              "silvertouch.com",       "Ahmedabad", "ASDM"),
    # ── Gandhinagar IT directories (rankex / brandveda) ──────────────────
    ("Heptagon Global Services","IT Services","Web Dev, Mobile App, Custom Software",             "heptagonglobal.com",    "Gandhinagar","Rankex"),
    ("Hats Off Solutions",      "IT Services","Web Dev, Mobile App, UI/UX",                       "hatsoffsolutions.com",  "Gandhinagar","Rankex"),
    ("ODD EVEN Infotech",       "IT Services","Web Dev, Mobile App, Software Dev",                "oddeveninfotech.com",   "Gandhinagar","Rankex"),
    ("Quest Infosense",         "IT Services","Software Dev, IT Consulting, Web",                 "questinfosense.com",    "Gandhinagar","Rankex"),
    ("Dreams Technology",       "IT Services","Web Dev, Mobile App, Custom Software",             "dreamstechnologies.com","Gandhinagar","Rankex"),
    ("The Intech Group",        "IT Services","ERP, Microsoft Dynamics, Cloud, Web",             "theintechgroup.com",    "Gandhinagar","Rankex"),
    ("AlpsLogic IT Solutions",  "IT Services","Web Dev, Software, IT Consulting",                "alpslogic.com",         "Gandhinagar","Rankex"),
    ("Xopple Infotech",         "IT Services","Web Dev, Mobile App, eCommerce",                  "xopple.com",            "Gandhinagar","Rankex"),
    ("Accrete Infosolution",    "IT Services","Web Dev, Mobile, Custom Software",                "accreteinfo.com",       "Gandhinagar","Rankex"),
    ("Tech Avidus",             "IT Services","Web Dev, Mobile App, Cloud Dev",                  "techavidus.com",        "Gandhinagar","Rankex"),
    ("eVision IT Solution",     "IT Services","Web Dev, Software, Digital Marketing",            "evisionitsolution.com", "Gandhinagar","Rankex"),
    ("Crucial Software IT Park", "IT Services","Software Dev, IT Park Services",                  "crucialsoftwares.com",  "Gandhinagar","Rankex"),
    ("Silver Touch (Gandhinagar)","IT Services","ERP, eGov, Enterprise Software",                "silvertouch.com",       "Gandhinagar","Rankex"),
    ("Cyfuture",                "IT Services","Cloud, Data Center, AI, IT Services",             "cyfuture.com",          "Gandhinagar","ASDM"),
    ("TIS India",               "IT Services","Web Dev, Software, Digital Marketing",            "tisindia.com",          "Gandhinagar","ASDM"),
    # ── GIFT City companies (Decentro / DRC Systems) ─────────────────────
    ("LTIMindtree",             "IT Services","AI, Digital Transform, Cloud, Enterprise IT",     "ltimindtree.com",       "GIFT City", "Decentro"),
    ("Zensar Technologies",     "IT Services","AI/ML, Digital, Cloud, Enterprise Apps",          "zensar.com",            "GIFT City", "Decentro"),
    ("Infibeam Avenues",        "IT Services","FinTech, Payments AI, eCommerce, CCAvenue",        "infibeam.com",          "GIFT City", "Decentro"),
    ("Decimal Point Analytics", "AI / ML",    "Financial AI, Data Analytics, ML, Research",       "decimalpointanalytics.com","GIFT City","Decentro"),
    ("Zoho",                    "IT Services","SaaS, AI, CRM, Business Software",                 "zoho.com",              "GIFT City", "Decentro"),
    ("Razorpay",                "IT Services","FinTech, Payments AI, ML Fraud Detection",         "razorpay.com",          "GIFT City", "Decentro"),
    ("DevX",                    "IT Services","Coworking + Tech, Startup Incubation",            "devx.work",             "GIFT City", "Decentro"),
    # ── GESIA official member directory ──────────────────────────────────
    ("Skynet Technologies",     "IT Services","Web Dev, Accessibility, eCommerce, AI",            "skynettechnologies.com","Ahmedabad", "GESIA"),
    ("Iplanme Digital Techno",  "IT Services","Digital Solutions, Web, Software Dev",             "iplanme.com",           "Ahmedabad", "GESIA"),
    ("Amarvelly Soft Tech",     "IT Services","Software Dev, IT Solutions",                       "amarvelly.com",         "Ahmedabad", "GESIA"),
    ("Mastek",                  "IT Services","AI, Enterprise Software, Cloud, Digital",          "mastek.com",            "Ahmedabad", "GESIA"),
    ("Shridhar InfoSec Solutions","IT Services","Cybersecurity, InfoSec, IT Audit",              "shridharinfosec.com",   "Ahmedabad", "GESIA"),
    ("Adiance Technologies",    "AI / ML",    "Video AI, CCTV AI, Computer Vision, IoT",          "adiance.com",           "Ahmedabad", "GESIA"),
    ("Iobix Services",          "IT Services","Web Dev, Software, IT Services",                   "iobix.in",              "Ahmedabad", "GESIA"),
    ("Stridely Solutions",      "IT Services","SAP, Microsoft, Cloud, Digital Transform",        "stridelysolutions.com", "Ahmedabad", "GESIA"),
    ("IndieSemiC",              "IT Services","Semiconductor Design, VLSI, Embedded",            "indiesemic.com",        "Ahmedabad", "GESIA"),
    ("Epnovate Technology",     "IT Services","Web Dev, Mobile App, Software",                   "epnovate.com",          "Ahmedabad", "GESIA"),
    ("Seed Data Systems",       "AI / ML",    "Data Systems, Analytics, AI Solutions",            "seeddatasystems.com",   "Ahmedabad", "GESIA"),
    ("Itechops Cloud",          "IT Services","Cloud Services, DevOps, IT Ops",                  "itechops.com",          "Ahmedabad", "GESIA"),
    ("Xtratrust Digisign",      "IT Services","Digital Signature, PKI, Security Tech",            "xtratrust.com",         "Ahmedabad", "GESIA"),
    ("Dhyey Consulting",        "IT Services","ERP, Odoo, Microsoft Dynamics, Cloud",            "dhyey.com",             "Ahmedabad", "GESIA"),
    ("Evince Development",      "IT Services","Web Dev, Mobile App, AI Integration",             "evincedev.com",         "Ahmedabad", "GESIA"),
    ("Stridely",                "IT Services","SAP, Cloud, Digital Transform",                   "stridelysolutions.com", "Ahmedabad", "GESIA"),
    # ── La Net Team / Cutshort IT directory ──────────────────────────────
    ("La Net Team Software",    "IT Services","Web Dev, Software, IT Consulting",                "lanetteam.com",         "Ahmedabad", "LaNetTeam"),
    ("YESQUEST",                "IT Services","Software Dev, IT Solutions",                       "yesquest.com",          "Ahmedabad", "LaNetTeam"),
    ("Theta Technolabs",        "AI / ML",    "AI/ML, iOS, Android, Node.js, AWS Dev",            "thetatechnolabs.com",   "Ahmedabad", "Cutshort-Dir"),
    ("Sublime Data Systems",    "IT Services","Custom Software, Scalable Solutions, Web",         "sublimedatasys.com",    "Ahmedabad", "Cutshort-Dir"),
    ("VAYUZ Technologies",      "AI / ML",    "Web/Mobile Dev, Data Science, AI",                "vayuz.com",             "Ahmedabad", "Cutshort-Dir"),
    # ── Additional well-known Ahmedabad IT firms ─────────────────────────
    ("Motadata (Mindarray)",    "AI / ML",    "AIOps, Network Monitoring AI, IT Ops",             "motadata.com",          "Ahmedabad", "Web-Dir"),
    ("Sigma Infosolutions",     "IT Services","AI, eCommerce, FinTech, Lending Tech",            "sigmainfo.net",         "Ahmedabad", "Web-Dir"),
    ("Elsner Technologies",     "IT Services","Web Dev, eCommerce, Mobile App",                  "elsner.com",            "Ahmedabad", "Web-Dir"),
    ("Webary Infotech",         "IT Services","Web Dev, Mobile App, Software",                   "webary.in",             "Ahmedabad", "Web-Dir"),
    ("Grey Chain",              "AI / ML",    "AI, Mobile App, Generative AI Solutions",          "greychaindesign.com",   "Ahmedabad", "Web-Dir"),
    ("Helios Solutions",        "IT Services","Web Dev, Offshore Dev, IT Outsourcing",           "heliossolutions.in",    "Ahmedabad", "Web-Dir"),
    ("Webs Optimization",       "IT Services","Web Dev, Mobile App, Software",                   "websoptimization.com",  "Ahmedabad", "Web-Dir"),
    ("Azilen Technologies",     "AI / ML",    "Enterprise AI, Product Engineering, ML, HRTech",   "azilen.com",            "Ahmedabad", "GESIA"),
]

def norm(s):
    if not s: return ""
    s = unicodedata.normalize('NFKD', str(s).lower())
    return re.sub(r'[^a-z0-9]', '', s)

def main():
    wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
    ws = wb["COMPLETE_MASTER"]

    existing = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 2).value
        if n: existing[norm(n)] = r

    fill_new = PatternFill("solid", fgColor="FFEEE6")  # light orange for IT-directory sourced
    added = updated = 0

    for company, category, roles, website_hint, city, source in IT_COMPANIES:
        key = norm(company)
        if not key: continue
        website = ("https://www." + website_hint) if not website_hint.startswith("http") else website_hint

        if key in existing:
            row = existing[key]
            if not ws.cell(row, 4).value: ws.cell(row, 4).value = category
            if not ws.cell(row, 5).value: ws.cell(row, 5).value = roles
            if not ws.cell(row, 13).value and website: ws.cell(row, 13).value = website
            # update city if Gandhinagar/GIFT
            if city in ("Gandhinagar", "GIFT City") and ws.cell(row, 3).value == "Ahmedabad":
                ws.cell(row, 3).value = city
            cur = ws.cell(row, 18).value or ""
            if source not in cur:
                ws.cell(row, 18).value = (cur + ", " + source).strip(", ")
            updated += 1
        else:
            r = ws.max_row + 1
            ws.cell(r, 1, r - 1)
            ws.cell(r, 2, company)
            ws.cell(r, 3, city)
            ws.cell(r, 4, category)
            ws.cell(r, 5, roles)
            ws.cell(r, 13, website)
            ws.cell(r, 18, source)
            for c in range(1, 19):
                ws.cell(r, c).fill = fill_new
                ws.cell(r, c).alignment = Alignment(vertical="center")
            ws.cell(r, 2).font = Font(bold=True, color="8B3A00")
            ws.row_dimensions[r].height = 18
            existing[key] = r
            added += 1

    for i, r in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(r, 1, i)

    total = ws.max_row - 1

    # Rebuild AI_ML_ONLY
    ws2 = wb["AI_ML_ONLY"]
    for r in range(ws2.max_row, 1, -1):
        ws2.delete_rows(r)
    ai_kw = ['ai','ml','data','machine','deep','nlp','computer vision','llm','genai',
             'analytics','automation','intelligence','learning','aiops']
    ai_row = 2
    fa = PatternFill("solid", fgColor="F5F0FF")
    fb = PatternFill("solid", fgColor="FFFFFF")
    for r in range(2, ws.max_row + 1):
        cat  = str(ws.cell(r, 4).value or "").lower()
        role = str(ws.cell(r, 5).value or "").lower()
        if any(k in cat or k in role for k in ai_kw):
            fill = fa if ai_row % 2 == 0 else fb
            for c in range(1, 19):
                v = ws.cell(r, c).value
                cell2 = ws2.cell(ai_row, c, v)
                cell2.fill = fill
                cell2.alignment = Alignment(vertical="center")
                if c in (6,7,8,9,10,11) and v:
                    cell2.font = Font(color="1a56db", size=9)
            ws2.cell(ai_row, 1, ai_row - 1)
            ws2.row_dimensions[ai_row].height = 18
            ai_row += 1
    ai_count = ws2.max_row - 1

    # Count by city
    gandhi = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r,3).value in ("Gandhinagar","GIFT City"))

    wb["STATS"].cell(3, 2, total)
    wb["STATS"].cell(6, 2, ai_count)

    wb.save("COMONK_TRUE_MASTER.xlsx")
    print(f"  NEW IT companies: +{added}")
    print(f"  Existing updated: {updated}")
    print(f"  Total companies:  {total}")
    print(f"  AI/ML companies:  {ai_count}")
    print(f"  IT companies:     {total - ai_count}")
    print(f"  Gandhinagar/GIFT: {gandhi}")

if __name__ == "__main__":
    main()
