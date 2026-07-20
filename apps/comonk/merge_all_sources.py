"""
CONSOLIDATE all project data files INTO COMONK_TRUE_MASTER.xlsx (one file).
- Appends NEW companies (not already in master)
- Fills empty email/phone slots for EXISTING companies
- Rebuilds All_Emails / All_Phones
- Additive only: never removes or overwrites existing data
"""
import openpyxl, csv, glob, re, unicodedata, zipfile, time
from openpyxl.styles import Font, PatternFill

MASTER = "COMONK_TRUE_MASTER.xlsx"
SKIP_FILES = {"COMONK_TRUE_MASTER.xlsx", "COMONK_MNC_CONTACTS.xlsx"}
EMAIL_RE = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')

def norm(s): return re.sub(r'[^a-z0-9]', '', unicodedata.normalize('NFKD', str(s).lower()))
def clean_key(h): return str(h).strip().lower().lstrip('﻿') if h else ""

def stable_load(path, tries=30):
    for _ in range(tries):
        try:
            with zipfile.ZipFile(path): pass
            return openpyxl.load_workbook(path)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read master")

wb = stable_load(MASTER)
ws = wb["COMPLETE_MASTER"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
C_NO=hmap.get("no.",1); C_CO=hmap.get("company name",2); C_CITY=hmap.get("city",3)
C_CAT=hmap.get("category",4); C_ROLE=hmap.get("target role",5)
C_PH=hmap.get("phone",12); C_WEB=hmap.get("website",13); C_SRC=hmap.get("source",18)
EMAIL_COLS=[hmap.get(f"email {i}") for i in range(1,7)]; EMAIL_COLS=[c for c in EMAIL_COLS if c]

existing = {}
for r in range(2, ws.max_row + 1):
    nm = ws.cell(r, C_CO).value
    if nm: existing[norm(nm)] = r
print(f"Master starts with {len(existing)} companies")

def find_col(hm, *cands):
    for c in cands:
        if c in hm: return hm[c]
    return None

records = {}
def add_rec(company, city=None, cat=None, roles=None, emails=None, phone=None, web=None):
    if not company: return
    k = norm(company)
    if not k: return
    rec = records.setdefault(k, {"company":str(company).strip(),"city":None,"category":None,
                                  "roles":None,"emails":set(),"phone":None,"website":None})
    if city and not rec["city"]: rec["city"]=str(city).strip()
    if cat and not rec["category"]: rec["category"]=str(cat).strip()
    if roles and not rec["roles"]: rec["roles"]=str(roles).strip()
    if phone and not rec["phone"]: rec["phone"]=str(phone).strip()
    if web and not rec["website"]: rec["website"]=str(web).strip()
    if emails:
        for e in emails:
            for m in EMAIL_RE.findall(str(e)):
                rec["emails"].add(m.lower())

for f in sorted(glob.glob("*.xlsx")):
    if f in SKIP_FILES: continue
    try: w = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception: continue
    for sh in w.sheetnames:
        s = w[sh]; it = s.iter_rows(values_only=True)
        try: h = next(it)
        except StopIteration: continue
        hm = {clean_key(x): i for i, x in enumerate(h) if x}
        ci = find_col(hm,"company name","company","name","organization")
        if ci is None: continue
        cy=find_col(hm,"city","location"); ct=find_col(hm,"category","industry","sector")
        rl=find_col(hm,"target role","roles","role","designation","position")
        ph=find_col(hm,"phone","mobile","contact number","contact"); wb2=find_col(hm,"website","url","web")
        eml=[i for k,i in hm.items() if "email" in k or "mail" in k]
        for row in it:
            if ci>=len(row) or not row[ci]: continue
            ems=[row[i] for i in eml if i<len(row) and row[i]]
            add_rec(row[ci],
                    row[cy] if cy is not None and cy<len(row) else None,
                    row[ct] if ct is not None and ct<len(row) else None,
                    row[rl] if rl is not None and rl<len(row) else None,
                    ems,
                    row[ph] if ph is not None and ph<len(row) else None,
                    row[wb2] if wb2 is not None and wb2<len(row) else None)
    w.close()

for f in sorted(glob.glob("*.csv")):
    try:
        with open(f, encoding="utf-8", errors="ignore") as fh:
            rd = csv.DictReader(fh)
            if not rd.fieldnames: continue
            hm = {clean_key(h): h for h in rd.fieldnames if h}
            ck = None
            for k in ("company name","company","name","organization"):
                if k in hm: ck=hm[k]; break
            if not ck: continue
            cyk=hm.get("city") or hm.get("location")
            ctk=hm.get("category") or hm.get("industry")
            rlk=hm.get("roles") or hm.get("target role") or hm.get("ai roles") or hm.get("role")
            phk=hm.get("phone") or hm.get("mobile")
            emk=[v for k,v in hm.items() if "email" in k or "mail" in k]
            for row in rd:
                comp=row.get(ck)
                if not comp: continue
                ems=[row[k] for k in emk if row.get(k)]
                add_rec(comp, row.get(cyk) if cyk else None, row.get(ctk) if ctk else None,
                        row.get(rlk) if rlk else None, ems, row.get(phk) if phk else None, None)
    except Exception: continue

print(f"Collected {len(records)} distinct company records from source files")

new_added=0; emails_added=0; phones_added=0
maxno=0
for r in range(2, ws.max_row+1):
    v=ws.cell(r,C_NO).value
    if isinstance(v,int): maxno=max(maxno,v)
append_row = ws.max_row + 1
for k, rec in records.items():
    if k in existing:
        r=existing[k]
        cur={str(ws.cell(r,c).value or "").lower() for c in EMAIL_COLS}
        free=[c for c in EMAIL_COLS if not ws.cell(r,c).value]
        for em in rec["emails"]:
            if em not in cur and free:
                slot=free.pop(0); ws.cell(r,slot).value=em
                ws.cell(r,slot).font=Font(color="1a56db",size=9); cur.add(em); emails_added+=1
        if rec["phone"] and not ws.cell(r,C_PH).value:
            ws.cell(r,C_PH).value=rec["phone"]; phones_added+=1
    else:
        maxno+=1
        ws.cell(append_row,C_NO).value=maxno
        ws.cell(append_row,C_CO).value=rec["company"]
        ws.cell(append_row,C_CITY).value=rec["city"] or "Ahmedabad"
        ws.cell(append_row,C_CAT).value=rec["category"] or "IT Services"
        if rec["roles"]: ws.cell(append_row,C_ROLE).value=rec["roles"]
        for em,slot in zip(sorted(rec["emails"]), EMAIL_COLS):
            ws.cell(append_row,slot).value=em; ws.cell(append_row,slot).font=Font(color="1a56db",size=9); emails_added+=1
        if rec["phone"]: ws.cell(append_row,C_PH).value=rec["phone"]; phones_added+=1
        if rec["website"]: ws.cell(append_row,C_WEB).value=rec["website"]
        ws.cell(append_row,C_SRC).value="merged from source files"
        existing[k]=append_row; append_row+=1; new_added+=1

print(f"New companies appended: {new_added}")
print(f"Emails added: {emails_added} | Phones added: {phones_added}")
print(f"COMPLETE_MASTER now: {ws.max_row-1} companies")

DARK=PatternFill("solid",fgColor="1a1a2e"); HDRF=Font(bold=True,color="FFFFFF",size=10)
if "All_Emails" in wb.sheetnames: del wb["All_Emails"]
ae=wb.create_sheet("All_Emails")
for ci,h in enumerate(["Email Address","Company","City / Source"],1):
    ae.cell(1,ci,h).fill=DARK; ae.cell(1,ci).font=HDRF
for ci,wd in enumerate([40,32,20],1): ae.column_dimensions[chr(64+ci)].width=wd
seen=set(); rows=[]
for r in range(2, ws.max_row+1):
    nm=ws.cell(r,C_CO).value; cy=ws.cell(r,C_CITY).value
    for c in EMAIL_COLS:
        em=ws.cell(r,c).value
        if em and "@" in str(em) and str(em).lower() not in seen:
            seen.add(str(em).lower()); rows.append((str(em).strip(),str(nm or ""),str(cy or "")))
if "MNC_HR_Emails" in wb.sheetnames:
    m=wb["MNC_HR_Emails"]
    for r in range(2,m.max_row+1):
        em=m.cell(r,5).value; co=m.cell(r,1).value
        if em and "@" in str(em) and str(em).lower() not in seen:
            seen.add(str(em).lower()); rows.append((str(em).strip(),str(co or "MNC"),"MNC contact"))
for i,(em,co,cy) in enumerate(rows,2):
    ae.cell(i,1,em).font=Font(color="1a56db",size=9); ae.cell(i,2,co).font=Font(size=9); ae.cell(i,3,cy).font=Font(size=9)
if "All_Phones" in wb.sheetnames: del wb["All_Phones"]
ap=wb.create_sheet("All_Phones")
for ci,h in enumerate(["Phone Number","Company","City"],1):
    ap.cell(1,ci,h).fill=DARK; ap.cell(1,ci).font=HDRF
for ci,wd in enumerate([22,32,16],1): ap.column_dimensions[chr(64+ci)].width=wd
seenp=set(); prows=[]
for r in range(2, ws.max_row+1):
    ph=ws.cell(r,C_PH).value; nm=ws.cell(r,C_CO).value; cy=ws.cell(r,C_CITY).value
    if ph:
        pk=re.sub(r'[\s\-\(\)]','',str(ph))
        if pk not in seenp: seenp.add(pk); prows.append((str(ph).strip(),str(nm or ""),str(cy or "")))
for i,(ph,co,cy) in enumerate(prows,2):
    ap.cell(i,1,ph).font=Font(color="2D7D46",size=9); ap.cell(i,2,co).font=Font(size=9); ap.cell(i,3,cy).font=Font(size=9)
print(f"All_Emails: {len(rows)} | All_Phones: {len(prows)}")

for _ in range(20):
    try: wb.save(MASTER); break
    except Exception: time.sleep(2)
print(f"\nConsolidation complete -> {MASTER}")
print("Sheets:", wb.sheetnames)
