"""
VCF CLEANUP v2 — Fix the 132 mangled VCF rows.
Re-match each VCF HR contact to its real company via:
  (a) email domain -> company website
  (b) real-company name as substring of contact name
  (c) real-company name as substring of email string
  (d) phone match
Unmatched personal contacts -> 'VCF_HR_Contacts' sheet (data preserved).
"""

import openpyxl, re, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment

MASTER = "COMONK_TRUE_MASTER.xlsx"
VCF    = "hr_vcards_15-6-26.vcf"
C_COMP=2; C_CITY=3; C_CAT=4; C_ROLE=5
C_EMAILS=[6,7,8,9,10,11]; C_PHONE=12; C_WEB=13; C_LI=14; C_ADDR=15; C_SRC=18

def norm(s):
    if not s: return ""
    s = unicodedata.normalize('NFKD', str(s).lower())
    return re.sub(r'[^a-z0-9]', '', s)

def clean_phone(p):
    if not p: return ""
    d = re.sub(r'[^\d]', '', str(p).replace('+',''))
    if d.startswith('91') and len(d)==12: d=d[2:]
    elif d.startswith('0') and len(d)==11: d=d[1:]
    if len(d)==10 and d[0] in '6789': return '+91 '+d[:5]+' '+d[5:]
    if len(d)==10 and d.startswith('79'): return '+91 79 '+d[2:6]+' '+d[6:]
    if len(d)==10: return '+91 '+d[:5]+' '+d[5:]
    return ""

EMAIL_RE = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')
GENERIC = {"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com",
           "yahoo.co.in","ymail.com","googlemail.com","gmal.com","gmai.com"}

def parse_vcf():
    contacts=[]; cur={}; in_photo=False
    for raw in open(VCF, encoding="utf-8", errors="ignore"):
        line=raw.rstrip("\n").rstrip("\r")
        if in_photo:
            if line.startswith(" ") or line.startswith("\t"): continue
            in_photo=False
        if line.startswith("PHOTO"): in_photo=True; continue
        if line=="BEGIN:VCARD": cur={"tels":[],"emails":[]}
        elif line=="END:VCARD":
            if cur: contacts.append(cur)
            cur={}
        elif line.startswith("FN:"): cur["fn"]=line[3:].strip()
        elif line.upper().startswith("TEL"):
            n=line.split(":",1)[-1].strip()
            if n: cur["tels"].append(n)
        elif line.upper().startswith("EMAIL"):
            e=line.split(":",1)[-1].strip().lower()
            if EMAIL_RE.match(e): cur["emails"].append(e)
    return contacts

def main():
    wb = openpyxl.load_workbook(MASTER)
    ws = wb["COMPLETE_MASTER"]

    # 1. Remove VCF-HR-only rows
    vcf_rows = [r for r in range(2, ws.max_row+1)
                if str(ws.cell(r, C_SRC).value or "").strip() == "VCF-HR"]
    print(f"  VCF-HR-only rows to clean: {len(vcf_rows)}")
    for r in sorted(vcf_rows, reverse=True):
        ws.delete_rows(r)

    # 2. Build real-company index
    real={}; web_to_row={}; phone_to_row={}; real_names=[]
    for r in range(2, ws.max_row+1):
        n = ws.cell(r, C_COMP).value
        if not n: continue
        nn = norm(n); real[nn]=r
        if len(nn)>=5: real_names.append((nn, r))
        w = ws.cell(r, C_WEB).value
        if w:
            d = re.sub(r'^https?://(www\.)?','',str(w)).split("/")[0].lower()
            if d: web_to_row[d]=r
        ph = ws.cell(r, C_PHONE).value
        if ph:
            dd = re.sub(r'[^\d]','',str(ph))[-10:]
            if len(dd)==10: phone_to_row[dd]=r
    real_names.sort(key=lambda x: -len(x[0]))

    def find_match(fn, emails, phone):
        for e in emails:
            dom = e.split("@")[1]
            if dom not in GENERIC and dom in web_to_row:
                return web_to_row[dom]
        if phone:
            dd = re.sub(r'[^\d]','',phone)[-10:]
            if dd in phone_to_row: return phone_to_row[dd]
        nf = norm(fn)
        for nn, r in real_names:
            if len(nn)>=5 and nn in nf: return r
        estr = norm("".join(emails))
        for nn, r in real_names:
            if len(nn)>=6 and nn in estr: return r
        return None

    contacts = parse_vcf()
    matched=0; new_co=0; personal=0

    if "VCF_HR_Contacts" in wb.sheetnames: del wb["VCF_HR_Contacts"]
    wsc = wb.create_sheet("VCF_HR_Contacts")
    hf = PatternFill("solid", fgColor="2E1A47"); hfont = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(["#","Contact Name","Phone","Email","Note"], 1):
        cell=wsc.cell(1,c,h); cell.fill=hf; cell.font=hfont
    wsc.column_dimensions['B'].width=30; wsc.column_dimensions['C'].width=18
    wsc.column_dimensions['D'].width=35; wsc.column_dimensions['E'].width=26
    crow=2

    existing_new = dict(real)
    for ct in contacts:
        fn = ct.get("fn","")
        emails = ct.get("emails",[])
        tels = [clean_phone(t) for t in ct.get("tels",[])]
        tels = [t for t in tels if t]
        phone = tels[0] if tels else ""

        r = find_match(fn, emails, phone)
        if r:
            if phone and not ws.cell(r, C_PHONE).value:
                ws.cell(r, C_PHONE).value = phone
            cur_em = set(filter(None, [ws.cell(r,c).value for c in C_EMAILS]))
            empty = [c for c in C_EMAILS if not ws.cell(r,c).value]
            newe = [e for e in emails if e not in cur_em and e.split("@")[1] not in GENERIC]
            for e,c in zip(newe, empty): ws.cell(r,c).value=e
            src = ws.cell(r, C_SRC).value or ""
            if "VCF-HR" not in src: ws.cell(r, C_SRC).value=(src+", VCF-HR").strip(", ")
            matched += 1
        else:
            cdom = None
            for e in emails:
                dom = e.split("@")[1]
                if dom not in GENERIC: cdom=dom; break
            if cdom:
                cname = cdom.split(".")[0]; key=norm(cname)
                if key in existing_new:
                    rr = existing_new[key]
                    if phone and not ws.cell(rr,C_PHONE).value: ws.cell(rr,C_PHONE).value=phone
                    matched += 1; continue
                rr = ws.max_row + 1
                ws.cell(rr,1,rr-1); ws.cell(rr,C_COMP,cname.title()); ws.cell(rr,C_CITY,"Ahmedabad")
                ws.cell(rr,C_CAT,"IT Services")
                for e,c in zip(emails, C_EMAILS): ws.cell(rr,c,e)
                if phone: ws.cell(rr,C_PHONE,phone)
                ws.cell(rr,C_WEB,"https://www."+cdom); ws.cell(rr,C_SRC,"VCF-HR")
                for c in range(1,19):
                    ws.cell(rr,c).fill = PatternFill("solid", fgColor="E8FFE8")
                    ws.cell(rr,c).alignment = Alignment(vertical="center")
                existing_new[key]=rr; new_co += 1
            else:
                wsc.cell(crow,1,crow-1); wsc.cell(crow,2,fn); wsc.cell(crow,3,phone)
                wsc.cell(crow,4,"; ".join(emails)); wsc.cell(crow,5,"Ahmedabad HR (unmatched)")
                crow += 1; personal += 1

    for i, r in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(r, 1, i)
    total = ws.max_row - 1
    max_col = ws.max_column

    ws2 = wb["AI_ML_ONLY"]
    for r in range(ws2.max_row, 1, -1):
        ws2.delete_rows(r)
    ai_kw = ['ai','ml','data','machine','deep','nlp','computer vision','llm','genai',
             'analytics','automation','intelligence','learning','aiops','analyst']
    ai_row = 2
    fa = PatternFill("solid", fgColor="F5F0FF"); fb = PatternFill("solid", fgColor="FFFFFF")
    for r in range(2, ws.max_row + 1):
        cat=str(ws.cell(r,C_CAT).value or "").lower(); role=str(ws.cell(r,C_ROLE).value or "").lower()
        if any(k in cat or k in role for k in ai_kw):
            fill=fa if ai_row%2==0 else fb
            for c in range(1, max_col+1):
                v=ws.cell(r,c).value; cc=ws2.cell(ai_row,c,v); cc.fill=fill
                cc.alignment=Alignment(vertical="center")
                if c in C_EMAILS and v: cc.font=Font(color="1a56db", size=9)
            ws2.cell(ai_row,1,ai_row-1); ai_row+=1
    ai_count = ws2.max_row - 1

    with_ph=sum(1 for r in range(2,ws.max_row+1) if ws.cell(r,C_PHONE).value)
    with_em=sum(1 for r in range(2,ws.max_row+1) if any(ws.cell(r,c).value for c in C_EMAILS))
    wb["STATS"].cell(3,2,total); wb["STATS"].cell(6,2,ai_count)
    wb.save(MASTER)

    print(f"  Matched to real companies:   {matched}")
    print(f"  New companies (domain email): {new_co}")
    print(f"  Personal -> VCF_HR_Contacts:  {personal}")
    print(f"\n  TOTAL: {total} | AI/ML: {ai_count}")
    print(f"  Phone: {with_ph} ({round(with_ph/total*100)}%) | Email: {with_em} ({round(with_em/total*100)}%)")

if __name__ == "__main__":
    main()
