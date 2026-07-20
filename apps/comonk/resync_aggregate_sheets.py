"""
Resync All_Emails, All_Phones, AI_ML_ONLY, STATS from the current
COMPLETE_MASTER (after domain-probe enrichment). MNC_HR_Emails sheet
is untouched since domain-probing didn't affect it.
"""
import openpyxl, re
from openpyxl.styles import Font, PatternFill

MASTER = "COMONK_TRUE_MASTER.xlsx"
wb = openpyxl.load_workbook(MASTER)
ws = wb["COMPLETE_MASTER"]

headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
hmap = {str(h).strip().lower(): i + 1 for i, h in enumerate(headers) if h}
COL_CO   = hmap.get("company name", 2)
COL_CITY = hmap.get("city", 3)
COL_CAT  = hmap.get("category", 4)
COL_PH   = hmap.get("phone", 12)
email_cols = [hmap.get(f"email {i}") for i in range(1, 7)]

DARK_BLUE = PatternFill("solid", fgColor="1a1a2e")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)

# ── Rebuild All_Emails ────────────────────────────────────────────────────────
if "All_Emails" in wb.sheetnames:
    del wb["All_Emails"]
ws_ae = wb.create_sheet("All_Emails")
for ci, h in enumerate(["Email Address", "Company", "City"], 1):
    ws_ae.cell(1, ci, h).fill = DARK_BLUE
    ws_ae.cell(1, ci).font = HDR_FONT
ws_ae.column_dimensions['A'].width = 38
ws_ae.column_dimensions['B'].width = 32
ws_ae.column_dimensions['C'].width = 15

seen = set()
all_emails_list = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, COL_CO).value
    city = ws.cell(r, COL_CITY).value
    for c in email_cols:
        em = ws.cell(r, c).value
        if em and "@" in str(em) and str(em).lower() not in seen:
            seen.add(str(em).lower())
            all_emails_list.append((str(em).strip(), str(name or ""), str(city or "")))

# Fold in MNC named contacts too
if "MNC_HR_Emails" in wb.sheetnames:
    ws_mnc = wb["MNC_HR_Emails"]
    for r in range(2, ws_mnc.max_row + 1):
        em = ws_mnc.cell(r, 3).value
        co = ws_mnc.cell(r, 5).value
        if em and "@" in str(em) and str(em).lower() not in seen:
            seen.add(str(em).lower())
            all_emails_list.append((str(em).strip(), str(co or "MNC"), "Ahmedabad/Gandhinagar"))

for i, (em, co, city) in enumerate(all_emails_list, 2):
    ws_ae.cell(i, 1, em).font = Font(color="1a56db", size=9)
    ws_ae.cell(i, 2, co).font = Font(size=9)
    ws_ae.cell(i, 3, city).font = Font(size=9)
print(f"All_Emails: {len(all_emails_list)} unique emails")

# ── Rebuild All_Phones ────────────────────────────────────────────────────────
if "All_Phones" in wb.sheetnames:
    del wb["All_Phones"]
ws_ap = wb.create_sheet("All_Phones")
for ci, h in enumerate(["Phone Number", "Company", "City"], 1):
    ws_ap.cell(1, ci, h).fill = DARK_BLUE
    ws_ap.cell(1, ci).font = HDR_FONT
ws_ap.column_dimensions['A'].width = 22
ws_ap.column_dimensions['B'].width = 32
ws_ap.column_dimensions['C'].width = 15

seen_ph = set()
all_phones_list = []
for r in range(2, ws.max_row + 1):
    ph = ws.cell(r, COL_PH).value
    name = ws.cell(r, COL_CO).value
    city = ws.cell(r, COL_CITY).value
    if ph:
        phk = re.sub(r'[\s\-\(\)]', '', str(ph))
        if phk not in seen_ph:
            seen_ph.add(phk)
            all_phones_list.append((str(ph).strip(), str(name or ""), str(city or "")))

for i, (ph, co, city) in enumerate(all_phones_list, 2):
    ws_ap.cell(i, 1, ph).font = Font(color="2D7D46", size=9)
    ws_ap.cell(i, 2, co).font = Font(size=9)
    ws_ap.cell(i, 3, city).font = Font(size=9)
print(f"All_Phones: {len(all_phones_list)} unique phones")

# ── Rebuild AI_ML_ONLY ────────────────────────────────────────────────────────
if "AI_ML_ONLY" in wb.sheetnames:
    del wb["AI_ML_ONLY"]
ws_ai = wb.create_sheet("AI_ML_ONLY")
for ci, h in enumerate(headers, 1):
    ws_ai.cell(1, ci, h).fill = DARK_BLUE
    ws_ai.cell(1, ci).font = HDR_FONT
ws_ai.column_dimensions['B'].width = 30

ai_count = 0
ai_r = 2
for r in range(2, ws.max_row + 1):
    cat = str(ws.cell(r, COL_CAT).value or "").lower()
    if any(kw in cat for kw in ["ai-ml", "ai/ml", "artificial", "machine learning"]):
        for c in range(1, len(headers) + 1):
            v = ws.cell(r, c).value
            if v: ws_ai.cell(ai_r, c, v).font = Font(size=9)
        ai_r += 1
        ai_count += 1
print(f"AI_ML_ONLY: {ai_count} companies")

# ── Rebuild STATS ──────────────────────────────────────────────────────────────
if "STATS" in wb.sheetnames:
    del wb["STATS"]
ws_st = wb.create_sheet("STATS")
ws_st.column_dimensions['A'].width = 38
ws_st.column_dimensions['B'].width = 20

with_email = sum(1 for r in range(2, ws.max_row+1) if any(ws.cell(r, c).value for c in email_cols))
with_phone = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, COL_PH).value)
total_rows = ws.max_row - 1
mnc_contacts = wb["MNC_HR_Emails"].max_row - 1 if "MNC_HR_Emails" in wb.sheetnames else 0

stats = [
    ("COMONK TRUE MASTER - STATS", ""),
    ("Generated", "2026-07-01"),
    ("", ""),
    ("Total Companies", total_rows),
    ("Companies with Email", with_email),
    ("Companies with Phone", with_phone),
    ("Total Unique Emails", len(all_emails_list)),
    ("Total Unique Phones", len(all_phones_list)),
    ("AI/ML Companies", ai_count),
    ("MNC Named Recruiter Contact rows", mnc_contacts),
]
for i, (k, v) in enumerate(stats, 1):
    ws_st.cell(i, 1, k); ws_st.cell(i, 2, v)
    if i == 1:
        ws_st.cell(i, 1).fill = DARK_BLUE
        ws_st.cell(i, 1).font = Font(bold=True, color="FFFFFF", size=12)
print("STATS rebuilt")

wb.save(MASTER)
print(f"\nTotal companies: {total_rows}")
print(f"With email: {with_email} ({with_email*100//total_rows}%)")
print(f"With phone: {with_phone} ({with_phone*100//total_rows}%)")
print(f"Saved -> {MASTER}")
