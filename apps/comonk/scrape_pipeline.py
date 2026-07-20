"""
Authorized contact scraper for 'All Companies' in Ahmedabad_IT_AIML_FINAL_MASTER.xlsx.
Per company missing phone/website, runs a pipeline (authorized data only):
  1. Google Maps (Playwright)  -> verified phone + website + address
  2. DuckDuckGo HTML search     -> official website (fallback)
  3. Contact-page scrape        -> emails + phone published on the site
Checkpoints every 100 companies. IMPORTANT: keep the Excel file CLOSED while running.
"""
import openpyxl, re, time, zipfile, asyncio, httpx, os
from urllib.parse import quote_plus, urlparse
from openpyxl.styles import Font
from playwright.async_api import async_playwright

_last_saved_mtime = [0.0]   # tracks the mtime of our own saves (external-edit guard)

F="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"; LOG="scrape_log.txt"; SHEET="All Companies"
EMAIL_RE=re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
PHONE_RE=re.compile(r'(?:\+91[\s\-]?)?(?:0)?[6-9]\d{9}|0\d{2,4}[\s\-]\d{6,8}')
GENERIC={"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com","ymail.com","live.com","wixpress.com","sentry.io","example.com"}
SKIP_DOM=("justdial","indiamart","linkedin","facebook","instagram","twitter","youtube","naukri",
          "glassdoor","ambitionbox","sulekha","tradeindia","wikipedia","google.","goo.gl",
          "quora","crunchbase","zaubacorp","tofler","yelp","mapsofindia")

def log(m):
    line=f"[{time.strftime('%H:%M:%S')}] {m}"; print(line,flush=True)
    with open(LOG,"a",encoding="utf-8") as f: f.write(line+"\n")

def stable_load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            wb=openpyxl.load_workbook(p)
            _last_saved_mtime[0]=os.path.getmtime(p)
            return wb
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read (is the file open in Excel?)")

def norm_phone(s):
    d=re.sub(r'[^\d+]','',s)
    if d.startswith('+91'): return d
    d=re.sub(r'\D','',s)
    if len(d)==10 and d[0] in '6789': return '+91'+d
    if len(d)==11 and d.startswith('0'): return '+91'+d[1:]
    if len(d)>=10: return '+91'+d[-10:] if d[-10] in '6789' else s.strip()
    return s.strip()

open(LOG,"a").close()
log("=== SCRAPE PIPELINE (Maps + DDG + contact-page) ===")
wb=stable_load(F); ws=wb[SHEET]
hdr=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
CO=hm.get("company name",2); CITY=hm.get("city",14); PH=hm.get("phone",10)
WEB=hm.get("website",11); ADDR=hm.get("address",13); SRC=hm.get("source",16)
EM=[hm.get(f"email {i}") for i in range(1,7)]; EM=[c for c in EM if c]

targets=[]
for r in range(4, ws.max_row+1):
    nm=ws.cell(r,CO).value
    if not nm: continue
    has_ph=ws.cell(r,PH).value; has_web=ws.cell(r,WEB).value
    if not has_ph or not has_web:
        targets.append((r, str(nm).strip(), str(ws.cell(r,CITY).value or "Ahmedabad").strip()))
log(f"{len(targets)} companies missing phone or website")

def clean_domain(url):
    try:
        h=urlparse(url if "//" in url else "http://"+url).netloc.lower()
        h=h.replace("www.","")
        return h if "." in h else None
    except: return None

def good_site(url):
    d=clean_domain(url) or ""
    return d and not any(s in d for s in SKIP_DOM)

stats={"phone":0,"web":0,"email":0,"addr":0,"maps_hit":0,"ddg_hit":0}
CKPT=[0]

ABORTED=[False]
def checkpoint():
    # SAFEGUARD: if the file changed on disk since our last save, someone edited it
    # externally (e.g. in Excel) -> do NOT overwrite their work; abort the run.
    try:
        cur=os.path.getmtime(F)
        if _last_saved_mtime[0] and cur > _last_saved_mtime[0] + 2:
            log("!! EXTERNAL EDIT DETECTED (file changed on disk). Aborting to protect your data. NOT saving.")
            ABORTED[0]=True
            return False
    except Exception: pass
    for _ in range(15):
        try:
            wb.save(F); _last_saved_mtime[0]=os.path.getmtime(F); return True
        except Exception: time.sleep(2)
    return False

async def maps_lookup(page, company, city):
    """Return (phone, website, address) from Google Maps."""
    phone=website=address=None
    try:
        url=f"https://www.google.com/maps/search/{quote_plus(company+' '+city+' Gujarat')}"
        await page.goto(url, wait_until="domcontentloaded", timeout=20000)
        await page.wait_for_timeout(2800)
        # if a results feed, click first result
        try:
            feed=await page.query_selector('a.hfpxzc')
            if feed:
                await feed.click(); await page.wait_for_timeout(2500)
        except Exception: pass
        # website (aria-label = "Website: tatvasoft.com")
        try:
            el=await page.query_selector('[data-item-id="authority"]')
            if el:
                aria=await el.get_attribute('aria-label') or ""
                if ":" in aria:
                    dom=aria.split(":",1)[1].strip().replace("www.","")
                    if "." in dom and good_site(dom): website="https://www."+dom
                if not website:
                    href=await el.get_attribute('href') or ""
                    m=re.search(r'[?&]q=(https?[^&]+)', href)
                    if m:
                        from urllib.parse import unquote
                        u=unquote(m.group(1))
                        if good_site(u): website=u
        except Exception: pass
        # phone (data-item-id = "phone:tel:09601421472")
        try:
            el=await page.query_selector('[data-item-id^="phone:tel:"]')
            if el:
                did=await el.get_attribute('data-item-id')
                if did and "tel:" in did: phone=did.split("tel:")[1]
        except Exception: pass
        # address (aria-label = "Address: ...")
        try:
            el=await page.query_selector('[data-item-id="address"]')
            if el:
                aria=await el.get_attribute('aria-label')
                if aria: address=aria.split(":",1)[1].strip() if ":" in aria else aria
        except Exception: pass
    except Exception: pass
    return phone, website, address

async def ddg_website(client, company, city):
    try:
        r=await client.post("https://html.duckduckgo.com/html/", data={"q":f"{company} {city} official website"}, timeout=10)
        for m in re.finditer(r'href="(https?://[^"]+)"', r.text):
            u=m.group(1)
            if "duckduckgo.com" in u: continue
            if good_site(u): return u
    except Exception: pass
    return None

async def scrape_contact(client, website):
    """Return (emails set, phone) from website + contact page."""
    emails=set(); phone=None
    dom=clean_domain(website)
    if not dom: return emails, phone
    urls=[website]+[f"https://{dom}/contact", f"https://{dom}/contact-us", f"https://{dom}/contactus"]
    for u in urls[:3]:
        try:
            r=await client.get(u, timeout=8, follow_redirects=True, headers={"User-Agent":"Mozilla/5.0"})
            for em in EMAIL_RE.findall(r.text):
                em=em.lower(); d=em.split("@")[1] if "@" in em else ""
                if d and d not in GENERIC and (dom.split('.')[0] in d or d in dom): emails.add(em)
            if not phone:
                pm=PHONE_RE.findall(r.text)
                if pm: phone=norm_phone(pm[0])
            if emails: break
        except Exception: continue
    return emails, phone

def apply(r, phone, website, address, emails, src):
    changed=False
    if phone and not ws.cell(r,PH).value:
        ws.cell(r,PH).value=norm_phone(phone); stats["phone"]+=1; changed=True
    if website and not ws.cell(r,WEB).value:
        ws.cell(r,WEB).value=website if website.startswith("http") else "https://"+website; stats["web"]+=1; changed=True
    if address and ADDR and not ws.cell(r,ADDR).value:
        ws.cell(r,ADDR).value=address; stats["addr"]+=1; changed=True
    if emails:
        free=[c for c in EM if not ws.cell(r,c).value]; cur={str(ws.cell(r,c).value or '').lower() for c in EM}
        for em in sorted(emails):
            if em.lower() not in cur and free:
                slot=free.pop(0); ws.cell(r,slot).value=em; ws.cell(r,slot).font=Font(color="1a56db",size=9); cur.add(em.lower()); stats["email"]+=1; changed=True
    if changed and SRC: ws.cell(r,SRC).value=src

async def main():
    limits=httpx.Limits(max_connections=20)
    async with async_playwright() as pw:
        browser=await pw.chromium.launch(headless=True)
        ctx=await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
        pages=[await ctx.new_page() for _ in range(3)]
        sem=asyncio.Semaphore(3); pidx=[0]
        async with httpx.AsyncClient(limits=limits) as client:
            done=[0]
            async def work(r, company, city):
                async with sem:
                    # pick a page round-robin
                    pg=pages[pidx[0]%len(pages)]; pidx[0]+=1
                    phone, website, address = await maps_lookup(pg, company, city)
                    if phone or website: stats["maps_hit"]+=1
                    emails=set(); cphone=None
                    if website:
                        emails, cphone = await scrape_contact(client, website)
                    apply(r, phone or cphone, website, address, emails, "scraped: maps/ddg/contact")
                    done[0]+=1
                    if done[0]%50==0:
                        log(f"  {done[0]}/{len(targets)} | +{stats['phone']} ph, +{stats['web']} web, +{stats['email']} em, +{stats['addr']} addr (maps {stats['maps_hit']})")
                        if checkpoint(): log(f"  [checkpoint @ {done[0]}]")
            # process sequentially-ish in chunks to bound memory
            for i in range(0, len(targets), 60):
                if ABORTED[0]: break
                await asyncio.gather(*[work(r,n,c) for r,n,c in targets[i:i+60]], return_exceptions=True)
        await browser.close()
    if not ABORTED[0]:
        checkpoint()
        log(f"DONE. +{stats['phone']} phones, +{stats['web']} websites, +{stats['email']} emails, +{stats['addr']} addr")
    else:
        log(f"ABORTED (external edit). Saved up to last safe checkpoint. +{stats['phone']} ph, +{stats['web']} web, +{stats['email']} em so far.")

asyncio.run(main())
