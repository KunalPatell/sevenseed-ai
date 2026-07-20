"""
AUDIT — Compare ALL source files vs COMONK_TRUE_MASTER.xlsx
Reports what emails / phones are captured vs MISSING.
"""
import openpyxl, re, os, unicodedata

MASTER = "COMONK_TRUE_MASTER.xlsx"
EMAIL_RE = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
PHONE_RE = re.compile(r'(?:\+?91[\s\-]?)?[6-9]\d{4}[\s\-]?\d{5}')

def digits10(p):
    d = re.sub(r'[^\d]', '', str(p))[-10:]
    return d if len(d)==10 else None

# 1. Master content
wb = openpyxl.load_workbook(MASTER)
m_emails=set(); m_phones=set(); m_domains=set()
for sheet in wb.sheetnames:
    for row in wb[sheet].iter_rows(values_only=True):
        for v in row:
            if not v: continue
            for e in EMAIL_RE.findall(str(v)):
                m_emails.add(e.lower()); m_domains.add(e.lower().split("@")[1])
            d = digits10(v)
            if d and (d[0] in '6789' or d.startswith('79')): m_phones.add(d)
print(f"MASTER: {len(m_emails)} emails, {len(m_phones)} phones, {len(m_domains)} domains\n")

# 2. All sources
s_emails=set(); s_phones=set()
def slurp(txt):
    for e in EMAIL_RE.findall(txt): s_emails.add(e.lower())
    for p in PHONE_RE.findall(txt):
        d=digits10(p)
        if d: s_phones.add(d)

for fn in os.listdir("."):
    if fn.startswith("_extract_") and fn.endswith(".txt"):
        slurp(open(fn, encoding="utf-8", errors="ignore").read())
for fn in ["emails.csv","AIML_Companies_Apply_List.csv","Gandhinagar_GIFTCity_Companies.csv",
           "NEW_COMPANIES_DISCOVERED.csv","all_bcc_emails.txt","hr_vcards_15-6-26.vcf"]:
    if os.path.exists(fn): slurp(open(fn, encoding="utf-8", errors="ignore").read())
for fn in ["1500+_hr_list.xlsx","Ahmedabad_AIML_Companies_HR_Contacts.xlsx",
           "Ahmedabad_IT_AIML_Companies_Master.xlsx","Ahmedabad_IT_AIML_FINAL_MASTER.xlsx",
           "AI_Engineer_Job_Targets.xlsx","HR Mail List.xlsx"]:
    if os.path.exists(fn):
        try:
            w2 = openpyxl.load_workbook(fn, read_only=True)
            for s in w2.sheetnames:
                for row in w2[s].iter_rows(values_only=True):
                    for v in row:
                        if not v: continue
                        for e in EMAIL_RE.findall(str(v)): s_emails.add(e.lower())
                        d=digits10(v)
                        if d and (d[0] in '6789' or d.startswith('79')): s_phones.add(d)
            w2.close()
        except Exception as ex: print(f"  err {fn}: {ex}")

print(f"ALL SOURCES: {len(s_emails)} emails, {len(s_phones)} phones\n")

miss_e = sorted(s_emails - m_emails)
miss_p = sorted(s_phones - m_phones)
print(f"=== GAP ===")
print(f"  Emails in sources but NOT in master: {len(miss_e)}")
print(f"  Phones in sources but NOT in master: {len(miss_p)}")

open("_missing_emails.txt","w",encoding="utf-8").write("\n".join(miss_e))
open("_missing_phones.txt","w",encoding="utf-8").write("\n".join(miss_p))

gen = {"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com","yahoo.co.in","ymail.com","googlemail.com"}
known = [e for e in miss_e if e.split("@")[1] in m_domains]
newcorp = [e for e in miss_e if e.split("@")[1] not in m_domains and e.split("@")[1] not in gen]
personal = [e for e in miss_e if e.split("@")[1] in gen]
print(f"\n  Missing emails breakdown:")
print(f"    domain already in master (attach): {len(known)}")
print(f"    NEW corporate domain (add company): {len(newcorp)}")
print(f"    personal (gmail/yahoo):            {len(personal)}")
print(f"\n  Sample new corporate-domain emails:")
for e in newcorp[:25]: print(f"    {e}")
