"""
Mine ALL PDFs for company list + emails, add to COMONK_TRUE_MASTER.xlsx
  1. DataScience Compeny List.pdf -> companies (website, linkedin, roles)
  2. Mail list.pdf + HR MAIL MAIN_02.pdf + emails.csv -> emails matched by domain
"""

import openpyxl, re, unicodedata, os, csv
from openpyxl.styles import Font, PatternFill, Alignment

MASTER = "COMONK_TRUE_MASTER.xlsx"
C_COMP=2; C_CITY=3; C_CAT=4; C_ROLE=5
C_EMAILS=[6,7,8,9,10,11]; C_PHONE=12; C_WEB=13; C_LI=14; C_ADDR=15; C_SRC=18

def norm(s):
    if not s: return ""
    s = unicodedata.normalize('NFKD', str(s).lower())
    return re.sub(r'[^a-z0-9]', '', s)

EMAIL_RE = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')

# ── 1. Parse DataScience Company List PDF text ───────────────────────────────
def parse_ds_companies():
    import pdfplumber
    with pdfplumber.open("DataScience Compeny List.pdf") as doc:
        text = "\n".join(pg.extract_text() or "" for pg in doc.pages)
    lines = [l.rstrip() for l in text.splitlines()]
    companies = []
    city = "Ahmedabad"
    cur = None
    GUJ_CITY = {"ahmedabad","gandhinagar","vadodara","anand","halol","vithalapur",
                "limda","valsad","surat","rajkot"}
    for line in lines:
        s = line.strip()
        if not s: continue
        # City header e.g. "Ahmedabad:" / "Gandhinagar:"
        cl = s.lower().rstrip(":")
        if cl in GUJ_CITY and s.endswith(":"):
            city = s.rstrip(":").strip().title()
            continue
        # Numbered company "1. Infosys"
        m = re.match(r'^\d+\.\s+(.*)', s)
        if m:
            if cur: companies.append(cur)
            cur = {"name": m.group(1).strip(), "city": city, "web": "", "li": "", "roles": ""}
            continue
        # bullet lines (○ website / linkedin / roles)
        s2 = s.lstrip("○•- ").strip()
        if cur:
            if re.match(r'^(www\.|https?://|[a-z0-9\-]+\.(com|ai|in|net|org|io|co))', s2, re.I) and "linkedin" not in s2.lower():
                cur["web"] = s2 if s2.startswith("http") else "https://" + s2
            elif "linkedin.com" in s2.lower():
                cur["li"] = s2
            else:
                cur["roles"] = (cur["roles"] + " " + s2).strip() if cur["roles"] else s2
    if cur: companies.append(cur)
    return companies

# ── 2. Collect all emails from PDFs + CSV ────────────────────────────────────
def collect_emails():
    emails = set()
    for fn in os.listdir("."):
        if fn.startswith("_extract_") and fn.endswith(".txt"):
            section = None
            for line in open(fn, encoding="utf-8", errors="ignore"):
                line = line.strip()
                if line == "=== EMAILS ===": section = "e"; continue
                if line.startswith("=== PHONES"): section = None; continue
                if line.startswith("=== FULL"): section = None; continue
                if section == "e" and EMAIL_RE.match(line):
                    emails.add(line.lower())
    # emails.csv
    if os.path.exists("emails.csv"):
        for row in csv.DictReader(open("emails.csv", encoding="utf-8", errors="ignore")):
            e = (row.get("email") or "").strip().lower()
            if EMAIL_RE.match(e): emails.add(e)
    # Build domain -> [emails]
    from collections import defaultdict
    dmap = defaultdict(list)
    GENERIC = {"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com",
               "yahoo.co.in","ymail.com","googlemail.com","protonmail.com"}
    for e in emails:
        dom = e.split("@")[1]
        if dom not in GENERIC:
            dmap[dom].append(e)
    return emails, dmap

def main():
    ds = parse_ds_companies()
    print(f"  DataScience PDF companies parsed: {len(ds)}")
    all_emails, dmap = collect_emails()
    print(f"  Total unique emails from PDFs+CSV: {len(all_emails)}")
    print(f"  Non-generic domains: {len(dmap)}\n")

    wb = openpyxl.load_workbook(MASTER)
    ws = wb["COMPLETE_MASTER"]

    existing = {}
    web_to_row = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, C_COMP).value
        if n: existing[norm(n)] = r
        w = ws.cell(r, C_WEB).value
        if w:
            d = re.sub(r'^https?://(www\.)?', '', str(w)).split("/")[0].lower()
            if d: web_to_row[d] = r

    # ── Add DataScience companies ────────────────────────────────────────────
    added = updated = 0
    fill_new = PatternFill("solid", fgColor="E6FFF7")
    for c in ds:
        key = norm(c["name"])
        if not key: continue
        web = c["web"]
        li  = c["li"]
        roles = c["roles"]
        if key in existing:
            r = existing[key]
            if not ws.cell(r, C_WEB).value and web: ws.cell(r, C_WEB).value = web
            if not ws.cell(r, C_LI).value and li: ws.cell(r, C_LI).value = li
            if not ws.cell(r, C_ROLE).value and roles: ws.cell(r, C_ROLE).value = roles
            cur = ws.cell(r, C_SRC).value or ""
            if "DataSci-PDF" not in cur:
                ws.cell(r, C_SRC).value = (cur + ", DataSci-PDF").strip(", ")
            updated += 1
        else:
            r = ws.max_row + 1
            cat = "AI / ML" if re.search(r'ai|ml|data sci|machine|deep|analyst|analytics', roles, re.I) else "IT Services"
            ws.cell(r, 1, r-1); ws.cell(r, C_COMP, c["name"]); ws.cell(r, C_CITY, c["city"])
            ws.cell(r, C_CAT, cat); ws.cell(r, C_ROLE, roles)
            if web: ws.cell(r, C_WEB, web)
            if li: ws.cell(r, C_LI, li)
            ws.cell(r, C_SRC, "DataSci-PDF")
            for cc in range(1, 19):
                ws.cell(r, cc).fill = fill_new
                ws.cell(r, cc).alignment = Alignment(vertical="center")
            ws.cell(r, C_COMP).font = Font(bold=True, color="006B5E")
            existing[key] = r
            d = re.sub(r'^https?://(www\.)?', '', str(web)).split("/")[0].lower() if web else ""
            if d: web_to_row[d] = r
            added += 1

    # ── Match emails by domain, fill blank email columns ─────────────────────
    em_filled = 0
    for dom, elist in dmap.items():
        r = web_to_row.get(dom)
        if not r: continue
        current = set(filter(None, [ws.cell(r, c).value for c in C_EMAILS]))
        empty = [c for c in C_EMAILS if not ws.cell(r, c).value]
        newem = [e for e in elist if e not in current]
        for e, c in zip(newem, empty):
            ws.cell(r, c).value = e; em_filled += 1
        if newem and empty:
            cur = ws.cell(r, C_SRC).value or ""
            if "PDF-Mail" not in cur:
                ws.cell(r, C_SRC).value = (cur + ", PDF-Mail").strip(", ")

    for i, r in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(r, 1, i)
    total = ws.max_row - 1
    max_col = ws.max_column

    # Rebuild AI_ML_ONLY
    ws2 = wb["AI_ML_ONLY"]
    for r in range(ws2.max_row, 1, -1):
        ws2.delete_rows(r)
    ai_kw = ['ai','ml','data','machine','deep','nlp','computer vision','llm','genai',
             'analytics','automation','intelligence','learning','aiops','analyst']
    ai_row = 2
    fa = PatternFill("solid", fgColor="F5F0FF"); fb = PatternFill("solid", fgColor="FFFFFF")
    for r in range(2, ws.max_row + 1):
        cat  = str(ws.cell(r, C_CAT).value or "").lower()
        role = str(ws.cell(r, C_ROLE).value or "").lower()
        if any(k in cat or k in role for k in ai_kw):
            fill = fa if ai_row % 2 == 0 else fb
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                cc = ws2.cell(ai_row, c, v); cc.fill = fill
                cc.alignment = Alignment(vertical="center")
                if c in C_EMAILS and v: cc.font = Font(color="1a56db", size=9)
            ws2.cell(ai_row, 1, ai_row - 1)
            ai_row += 1
    ai_count = ws2.max_row - 1

    with_em = sum(1 for r in range(2, ws.max_row+1) if any(ws.cell(r,c).value for c in C_EMAILS))
    wb["STATS"].cell(3, 2, total)
    wb["STATS"].cell(6, 2, ai_count)
    wb.save(MASTER)

    print(f"  DataScience companies: +{added} new, {updated} enriched")
    print(f"  PDF emails matched & filled: +{em_filled}")
    print(f"\n  TOTAL: {total} companies | AI/ML: {ai_count} | with email: {with_em} ({round(with_em/total*100)}%)")

if __name__ == "__main__":
    main()
