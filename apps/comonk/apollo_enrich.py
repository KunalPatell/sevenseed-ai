"""
Apollo.io Organization Enrichment (FREE plan)
Fills missing Phone, Address, LinkedIn, + adds Employees/Industry/Founded
for every company in COMONK_TRUE_MASTER.xlsx using Apollo org enrich endpoint.
"""

import openpyxl, httpx, time, os, re
from urllib.parse import urlparse
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL = "COMONK_TRUE_MASTER.xlsx"
KEY = os.environ.get("APOLLO_API_KEY", "")
if not KEY and os.path.exists(".env"):
    for line in open(".env", encoding="utf-8"):
        if line.startswith("APOLLO_API_KEY"):
            KEY = line.split("=", 1)[1].strip()

# Column map in COMPLETE_MASTER:
# 1=#,2=Company,3=City,4=Category,5=Roles,6-11=Email1-6,12=Phone,
# 13=Website,14=LinkedIn,15=Address,16=Careers,17=Priority,18=Sources,
# 19=LinkedIn HR Search,20=LinkedIn Company Page
# NEW: 21=Employees, 22=Industry, 23=Founded
COL_COMPANY=2; COL_CITY=3; COL_PHONE=12; COL_WEBSITE=13
COL_LINKEDIN=14; COL_ADDRESS=15
COL_EMP=21; COL_IND=22; COL_FOUNDED=23

def domain_of(url):
    if not url: return None
    url = str(url).strip()
    if not url.startswith("http"): url = "https://" + url
    d = urlparse(url).netloc.replace("www.", "").strip()
    return d if "." in d else None

def enrich(domain):
    try:
        r = httpx.get("https://api.apollo.io/v1/organizations/enrich",
                      headers={"X-Api-Key": KEY, "Cache-Control": "no-cache"},
                      params={"domain": domain}, timeout=12)
        if r.status_code != 200:
            return None, r.status_code
        return r.json().get("organization", {}), 200
    except Exception as e:
        return None, str(e)

def add_headers(ws):
    """Add Employees / Industry / Founded headers if not present."""
    h_fill = PatternFill("solid", fgColor="2E1A47")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    for col, name, w in [(COL_EMP,"Employees",12),(COL_IND,"Industry",26),(COL_FOUNDED,"Founded",10)]:
        if ws.cell(1, col).value not in ("Employees","Industry","Founded"):
            cell = ws.cell(1, col, name)
            cell.fill = h_fill; cell.font = h_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = w

def main():
    if not KEY:
        print("  APOLLO_API_KEY missing!"); return

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["COMPLETE_MASTER"]
    add_headers(ws)

    # Targets: companies with website but missing phone OR address OR linkedin
    targets = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, COL_COMPANY).value
        web  = ws.cell(r, COL_WEBSITE).value
        if not name or not web: continue
        miss_phone = not ws.cell(r, COL_PHONE).value
        miss_addr  = not ws.cell(r, COL_ADDRESS).value
        miss_li    = not ws.cell(r, COL_LINKEDIN).value
        if miss_phone or miss_addr or miss_li:
            targets.append(r)

    print(f"\n  {len(targets)} companies need enrichment. Apollo org enrich chal raha hai...\n")

    enriched = ph_add = addr_add = li_add = 0
    fail = 0

    for i, r in enumerate(targets, 1):
        name = ws.cell(r, COL_COMPANY).value
        domain = domain_of(ws.cell(r, COL_WEBSITE).value)
        if not domain:
            continue

        org, status = enrich(domain)
        if status != 200:
            fail += 1
            if status in (401, 403):
                print(f"  Stopping — API status {status}")
                break
            if i % 50 == 0:
                print(f"  [{i}/{len(targets)}] ... ({enriched} enriched)")
            continue

        if not org:
            if i % 50 == 0: print(f"  [{i}/{len(targets)}] ... ({enriched} enriched)")
            continue

        got = []
        # Phone
        if not ws.cell(r, COL_PHONE).value:
            ph = org.get("phone") or (org.get("primary_phone") or {}).get("number")
            if ph:
                ws.cell(r, COL_PHONE).value = ph; ph_add += 1; got.append("ph")
        # Address
        if not ws.cell(r, COL_ADDRESS).value:
            addr = org.get("raw_address") or org.get("street_address")
            if addr:
                ws.cell(r, COL_ADDRESS).value = addr; addr_add += 1; got.append("addr")
        # LinkedIn
        if not ws.cell(r, COL_LINKEDIN).value:
            li = org.get("linkedin_url")
            if li:
                ws.cell(r, COL_LINKEDIN).value = li; li_add += 1; got.append("li")
        # Extra: employees, industry, founded
        if org.get("estimated_num_employees"):
            ws.cell(r, COL_EMP).value = org["estimated_num_employees"]
        if org.get("industry"):
            ws.cell(r, COL_IND).value = org["industry"]
        if org.get("founded_year"):
            ws.cell(r, COL_FOUNDED).value = org["founded_year"]
        # City correction if Apollo says Gandhinagar
        ac = (org.get("city") or "").lower()
        if "gandhinagar" in ac and ws.cell(r, COL_CITY).value == "Ahmedabad":
            ws.cell(r, COL_CITY).value = "Gandhinagar"

        if got:
            enriched += 1
            if enriched % 20 == 0:
                print(f"  [{i}/{len(targets)}] {enriched} enriched (ph:{ph_add} addr:{addr_add} li:{li_add})")

        if i % 100 == 0:
            wb.save(EXCEL)
            print(f"  -- checkpoint saved ({i}/{len(targets)}) --")

        time.sleep(0.3)

    wb.save(EXCEL)
    print(f"\n  DONE. Enriched {enriched} companies:")
    print(f"    Phones added:   +{ph_add}")
    print(f"    Addresses added:+{addr_add}")
    print(f"    LinkedIn added: +{li_add}")
    print(f"    Failed lookups: {fail}")

    # Final stats
    total = ws.max_row - 1
    with_ph = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, COL_PHONE).value)
    with_ad = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, COL_ADDRESS).value)
    with_li = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, COL_LINKEDIN).value)
    print(f"\n  === SHEET TOTALS ===")
    print(f"  Phone:    {with_ph}/{total} ({round(with_ph/total*100)}%)")
    print(f"  Address:  {with_ad}/{total} ({round(with_ad/total*100)}%)")
    print(f"  LinkedIn: {with_li}/{total} ({round(with_li/total*100)}%)")

if __name__ == "__main__":
    main()
