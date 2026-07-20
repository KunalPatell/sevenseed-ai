"""
Fetch HR email + phone for top MNC AI companies in Ahmedabad & Gandhinagar.
Saves a dedicated sheet 'MNC_AI_Targets' in Ahmedabad_IT_AIML_FINAL_MASTER.xlsx
"""

import openpyxl
import httpx
import time
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HUNTER_KEY = "86b42563dd92b15fb26268bdc0f9a697c7609d38"
EXCEL_PATH  = "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"

# Top MNCs with AI presence in Ahmedabad / Gandhinagar / GIFT City
# Format: (Company, Domain, City, AI_Roles, Office_Phone, LinkedIn)
MNC_LIST = [
    # ── Global MNCs with Ahmedabad offices ──────────────────────────────
    ("TCS (Tata Consultancy Services)", "tcs.com",           "Ahmedabad", "AI Engineer, Data Scientist, ML Engineer",          "+91 79 6130 0000", "linkedin.com/company/tata-consultancy-services"),
    ("Infosys",                         "infosys.com",        "Ahmedabad", "AI/ML Engineer, Data Engineer, Cloud AI",           "+91 79 3031 4000", "linkedin.com/company/infosys"),
    ("Wipro",                           "wipro.com",          "Ahmedabad", "AI Engineer, NLP Engineer, MLOps",                  "+91 79 6608 9000", "linkedin.com/company/wipro"),
    ("HCL Technologies",                "hcltech.com",        "Ahmedabad", "AI Developer, Data Scientist, Computer Vision",     "+91 79 4020 1500", "linkedin.com/company/hcltech"),
    ("Tech Mahindra",                   "techmahindra.com",   "Ahmedabad", "AI/ML Engineer, Generative AI, LLM Engineer",      "+91 79 2685 2200", "linkedin.com/company/tech-mahindra"),
    ("Capgemini",                       "capgemini.com",      "Ahmedabad", "Data Scientist, AI Architect, ML Platform Eng",     "+91 79 4014 5000", "linkedin.com/company/capgemini"),
    ("Accenture",                       "accenture.com",      "Ahmedabad", "AI Engineer, Applied AI, Data & AI Analyst",        "+91 79 6680 0000", "linkedin.com/company/accenture"),
    ("IBM",                             "ibm.com",            "Ahmedabad", "AI/ML Engineer, Watson AI, Generative AI",         "+91 79 6156 8500", "linkedin.com/company/ibm"),
    ("Hexaware",                        "hexaware.com",       "Ahmedabad", "AI Engineer, Cognitive AI, RPA + AI",              "+91 22 6770 0000", "linkedin.com/company/hexaware"),
    ("Persistent Systems",              "persistent.com",     "Ahmedabad", "AI Engineer, LLM Developer, ML Platform",          "+91 20 6703 0000", "linkedin.com/company/persistent-systems"),
    ("Mphasis",                         "mphasis.com",        "Ahmedabad", "AI/ML Engineer, Data Scientist, AI Ops",           "+91 80 3352 5000", "linkedin.com/company/mphasis"),
    ("L&T Technology Services (LTTS)", "ltts.com",           "Ahmedabad", "AI Engineer, Computer Vision, Embedded AI",        "+91 265 6154 000", "linkedin.com/company/ltts"),
    ("KPIT Technologies",              "kpit.com",           "Ahmedabad", "AI Engineer, Automotive AI, ML for ADAS",          "+91 20 6764 4000", "linkedin.com/company/kpit"),
    ("Cognizant",                       "cognizant.com",      "Ahmedabad", "AI Developer, Data Scientist, ML Engineer",        "+91 44 4209 6000", "linkedin.com/company/cognizant"),
    ("Deloitte",                        "deloitte.com",       "Ahmedabad", "AI Engineer, Data & AI Consultant",               "+91 79 6607 3100", "linkedin.com/company/deloitte"),
    ("PwC India",                       "pwc.com",            "Ahmedabad", "AI/ML Analyst, Data Scientist, AI Consultant",     "+91 79 3091 7000", "linkedin.com/company/pwcindia"),
    ("EY (Ernst & Young)",              "ey.com",             "Ahmedabad", "AI Data Analyst, ML Consultant, Gen AI Eng",       "+91 79 6608 3800", "linkedin.com/company/ernstandyoung"),
    ("Bosch",                           "bosch.com",          "Ahmedabad", "AI Engineer, Deep Learning, Automotive AI",        "+91 80 2296 2296", "linkedin.com/company/bosch"),
    ("Siemens",                         "siemens.com",        "Ahmedabad", "AI Engineer, IIoT + AI, Industrial AI",            "+91 22 3967 7000", "linkedin.com/company/siemens"),
    ("Oracle",                          "oracle.com",         "Ahmedabad", "AI/ML Engineer, Cloud AI, OCI AI Services",       "+91 80 4118 4500", "linkedin.com/company/oracle"),
    ("NTT Data",                        "nttdata.com",        "Ahmedabad", "AI Engineer, Data Science, Intelligent Auto",      "+91 80 4125 8100", "linkedin.com/company/nttdata"),
    ("Fujitsu",                         "fujitsu.com",        "Ahmedabad", "AI Consultant, ML Engineer, Digital Twin AI",     "+91 80 4060 5000", "linkedin.com/company/fujitsu"),
    ("Apexon (Infostretch)",            "apexon.com",         "Ahmedabad", "AI/ML Engineer, Data Engineer, LLM Developer",    "+1 408 927 4026",  "linkedin.com/company/apexon"),
    ("EXL Service",                     "exlservice.com",     "Ahmedabad", "Data Scientist, AI Analyst, ML Engineer",         "+91 79 4005 0500", "linkedin.com/company/exl"),
    ("Mastech Digital",                 "mastechdigital.com", "Ahmedabad", "AI/ML Staffing, Data Scientist, AI Consultant",   "+1 412 384 1300",  "linkedin.com/company/mastech-digital"),
    # ── Gandhinagar / GIFT City MNCs ────────────────────────────────────
    ("Argusoft India",                  "argusoft.com",       "Gandhinagar","AI Engineer, Data Analyst, IoT/Robotics Eng",     "+91 9328573110",   "linkedin.com/company/argusoft"),
    ("AvenuesAI",                       "avenuesai.com",      "GIFT City",  "AI Engineer, ML Engineer, Fintech Data Sci",     "+91 79 6777 2204", "linkedin.com/company/avenuesai"),
    ("iNubia IT Solutions",             "inubia.com",         "Gandhinagar","ML Engineer, AI Developer, Cloud AI",            "+91 79 2322 9100", "linkedin.com/company/inubia"),
    ("Gujarat Informatics Ltd (GIL)",   "gil.gujarat.gov.in", "Gandhinagar","AI Developer, eGov AI, Data Engineer",          "+91 79 2325 6022", ""),
    ("iCreate (GIFT City)",             "icreate.org.in",     "GIFT City",  "AI Engineer, Startup AI, Deep Tech",            "+91 79 2329 0100", "linkedin.com/company/icreate"),
]

def extract_domain(url):
    from urllib.parse import urlparse
    if not url: return url
    url = url.strip()
    if not url.startswith("http"): url = "https://" + url
    return urlparse(url).netloc.replace("www.", "").strip()

def hunter_fetch(domain):
    try:
        r = httpx.get(
            "https://api.hunter.io/v2/domain-search",
            params={"domain": domain, "api_key": HUNTER_KEY, "limit": 10, "type": "personal"},
            timeout=12
        )
        data = r.json()
        if "data" not in data or "emails" not in data["data"]:
            return [], None
        emails_raw = data["data"]["emails"]
        # Prefer HR/recruit/talent emails
        hr_emails = [e["value"] for e in emails_raw
                     if any(k in e.get("value","").lower() for k in
                            ["hr","recruit","career","talent","job","hiring","people","hrbp"])]
        all_emails = [e["value"] for e in emails_raw]
        # Phone from hunter if available
        phone = None
        for e in emails_raw:
            if e.get("phone_number"):
                phone = e["phone_number"]; break
        final_emails = (hr_emails or all_emails)[:5]
        return final_emails, phone
    except Exception as ex:
        print(f"    Hunter error: {ex}")
        return [], None

def main():
    # Check quota
    try:
        qr = httpx.get("https://api.hunter.io/v2/account",
                       params={"api_key": HUNTER_KEY}, timeout=8).json()
        searches = qr.get("data", {}).get("searches", {})
        used = searches.get("used", "?")
        avail = searches.get("available", "?")
        print(f"\n  Hunter.io: {used}/{avail} searches used  ({avail - used if str(avail).isdigit() and str(used).isdigit() else '?'} remaining)\n")
    except:
        print("  (quota check failed — continuing)\n")

    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Remove old sheet if exists
    if "MNC_AI_Targets" in wb.sheetnames:
        del wb["MNC_AI_Targets"]

    ws = wb.create_sheet("MNC_AI_Targets")

    # ── Header styling ────────────────────────────────────────────────
    HEADERS = ["#", "Company", "City", "AI Roles (Apply For)",
               "HR Email 1", "HR Email 2", "HR Email 3",
               "HR Email 4", "HR Email 5",
               "Office Phone", "Website", "LinkedIn"]
    COL_WIDTHS = [4, 30, 14, 42, 32, 32, 32, 32, 32, 18, 30, 40]

    header_fill = PatternFill("solid", fgColor="4A235A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for c, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28

    # ── Fetch & write data ────────────────────────────────────────────
    row = 2
    alt_fill   = PatternFill("solid", fgColor="F3ECF8")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")

    for idx, (company, domain, city, roles, office_phone, linkedin) in enumerate(MNC_LIST, 1):
        print(f"  [{idx:02d}/{len(MNC_LIST)}] {company} ({domain})", end=" ... ", flush=True)

        emails, hunter_phone = hunter_fetch(domain)
        phone = office_phone or hunter_phone or ""

        fill = alt_fill if idx % 2 == 0 else plain_fill
        values = [idx, company, city, roles] + emails + [""] * (5 - len(emails)) + [phone, f"https://{domain}", linkedin]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row, c, v)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if c == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c in (5, 6, 7, 8, 9) and v:
                cell.font = Font(color="1a56db")

        ws.row_dimensions[row].height = 22
        if emails:
            print(f"FOUND {len(emails)} HR emails")
        else:
            print("no emails from Hunter")

        row += 1
        time.sleep(1.1)

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(EXCEL_PATH)
    print(f"\n  Sheet 'MNC_AI_Targets' saved with {len(MNC_LIST)} companies -> {EXCEL_PATH}")

if __name__ == "__main__":
    main()
