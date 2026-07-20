"""
Fill empty rows in 'All Companies' of Ahmedabad_IT_AIML_FINAL_MASTER.xlsx.
For companies missing email/website:
  1. Domain-probe (name.com / .in / .co.in) to find a live official website
  2. Scrape REAL emails + phone from the live page (best quality)
  3. Fallback: generate careers@/hr@/info@ from the confirmed domain
Additive only. Logs progress to fill_log.txt. Saves in place with retry.
"""
import openpyxl, re, asyncio, httpx, unicodedata, zipfile, time

FINAL = "Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
LOG = "fill_log.txt"
EMAIL_RE = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
PHONE_RE = re.compile(r'(?:\+91[\s\-]?|0)?[6-9]\d{9}')
GENERIC = {"gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com","ymail.com","live.com"}
TLDS = [".com", ".in", ".co.in"]

def log(m):
    line = f"[{time.strftime('%H:%M:%S')}] {m}"
    print(line, flush=True)
    with open(LOG, "a", encoding="utf-8") as f: f.write(line+"\n")

def slug(name):
    s = unicodedata.normalize('NFKD', str(name).lower())
    s = re.sub(r'\(.*?\)', '', s); s = re.sub(r'[^a-z0-9]', '', s)
    return s

def cdom(url):
    if not url: return None
    u=re.sub(r'^https?://','',str(url).strip().lower()); u=re.sub(r'^www\.','',u)
    u=u.split('/')[0].split('?')[0].strip()
    return u if ('.' in u and len(u)>=4) else None

def stable_load(path, tries=30):
    for _ in range(tries):
        try:
            with zipfile.ZipFile(path): pass
            return openpyxl.load_workbook(path)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read")

open(LOG,"w").close()
log("Loading workbook...")
wb = stable_load(FINAL)
ws = wb["All Companies"]
hdr=[ws.cell(3,c).value for c in range(1,ws.max_column+1)]
hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
CO=hm.get("company name",2); PH=hm.get("phone",10); WEB=hm.get("website",11); SRC=hm.get("source",16)
EM=[hm.get(f"email {i}") for i in range(1,7)]; EM=[c for c in EM if c]

targets=[]  # (row, name, has_web_domain_or_None)
for r in range(4, ws.max_row+1):
    nm=ws.cell(r,CO).value
    if not nm: continue
    has_em=any(ws.cell(r,c).value for c in EM)
    dom=cdom(ws.cell(r,WEB).value)
    if not has_em:  # needs email
        targets.append((r, str(nm), dom))
log(f"{len(targets)} companies missing email to process")

from openpyxl.styles import Font
def set_emails(r, emails, src):
    free=[c for c in EM if not ws.cell(r,c).value]
    cur={str(ws.cell(r,c).value or '').lower() for c in EM}
    n=0
    for em in emails:
        if em.lower() not in cur and free:
            slot=free.pop(0); ws.cell(r,slot).value=em; ws.cell(r,slot).font=Font(color="1a56db",size=9)
            cur.add(em.lower()); n+=1
    if n and not ws.cell(r,SRC).value: ws.cell(r,SRC).value=src
    return n

stats={"web":0,"emails":0,"phones":0,"real_emails":0}

async def process():
    sem=asyncio.Semaphore(60)
    limits=httpx.Limits(max_connections=70, max_keepalive_connections=30)
    async with httpx.AsyncClient(limits=limits, timeout=4, follow_redirects=True,
                                 headers={"User-Agent":"Mozilla/5.0"}) as client:
        done=0
        async def handle(r, name, dom):
            nonlocal done
            async with sem:
                # find a live domain (use existing or guess)
                domain=dom; html=None
                if domain:
                    try:
                        resp=await client.get(f"https://{domain}")
                        if resp.status_code==200: html=resp.text
                    except Exception: pass
                if html is None:
                    sl=slug(name)
                    if len(sl)>=4:
                        for tld in TLDS:
                            try:
                                resp=await client.get(f"https://www.{sl}{tld}")
                                if resp.status_code==200 and len(resp.text)>500:
                                    tl=resp.text.lower()
                                    if any(p in tl for p in ["domain is for sale","buy this domain","godaddy.com/domainsearch"]):
                                        continue
                                    domain=f"{sl}{tld}"; html=resp.text; break
                            except Exception: continue
                if html and domain:
                    if not ws.cell(r,WEB).value:
                        ws.cell(r,WEB).value=f"https://www.{domain}"; stats["web"]+=1
                    # scrape real emails
                    real=set()
                    for em in EMAIL_RE.findall(html):
                        em=em.lower(); d=em.split("@")[1] if "@" in em else ""
                        if d and d not in GENERIC and domain.split('.')[0] in d:
                            real.add(em)
                    if real:
                        added=set_emails(r, sorted(real), "domain-probe: scraped real")
                        stats["emails"]+=added; stats["real_emails"]+=added
                    else:
                        gen=[f"careers@{domain}",f"hr@{domain}",f"info@{domain}"]
                        stats["emails"]+=set_emails(r, gen, "domain-probe: generated")
                    # phone
                    if not ws.cell(r,PH).value:
                        ph=PHONE_RE.findall(html)
                        if ph:
                            p=re.sub(r'[\s\-]','',ph[0])
                            if not p.startswith("+91"): p="+91"+p.lstrip("0")
                            ws.cell(r,PH).value=p; stats["phones"]+=1
                done+=1
                if done % 250 == 0:
                    log(f"  {done}/{len(targets)} | +{stats['web']} web, +{stats['emails']} emails ({stats['real_emails']} real), +{stats['phones']} phones")

        # process in chunks to bound memory
        for i in range(0, len(targets), 300):
            batch=targets[i:i+300]
            await asyncio.gather(*[handle(r,n,d) for r,n,d in batch], return_exceptions=True)

asyncio.run(process())
log(f"Probing done. Saving... (+{stats['web']} web, +{stats['emails']} emails, +{stats['phones']} phones)")
for _ in range(20):
    try: wb.save(FINAL); break
    except Exception: time.sleep(2)

# final counts
wb2=stable_load(FINAL); w2=wb2["All Companies"]
tot=0; em=0; ph=0
for r in range(4, w2.max_row+1):
    if not w2.cell(r,CO).value: continue
    tot+=1
    if any(w2.cell(r,c).value for c in EM): em+=1
    if w2.cell(r,PH).value: ph+=1
log(f"DONE. Companies={tot} | with email={em} | with phone={ph}")
log(f"Added this run: +{stats['web']} websites, +{stats['emails']} emails ({stats['real_emails']} real scraped), +{stats['phones']} phones")
