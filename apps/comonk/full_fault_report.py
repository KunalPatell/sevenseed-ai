"""
READ-ONLY full fault scan across every sheet. Changes nothing.
Reports: junk names, duplicates, invalid/placeholder emails, wrong-domain emails,
fake/invalid phones, repeated template values, blank rows, structural issues.
"""
import openpyxl, re, zipfile, time, unicodedata
from collections import Counter

F="Ahmedabad_IT_AIML_FINAL_MASTER.xlsx"
def load(p,t=40):
    for _ in range(t):
        try:
            with zipfile.ZipFile(p): pass
            return openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception: time.sleep(2)
    raise RuntimeError("cannot read (file open in Excel?)")
def norm(s): return re.sub(r'[^a-z0-9]','',unicodedata.normalize('NFKD',str(s).lower()))
def dom(url):
    if not url: return None
    u=re.sub(r'^https?://','',str(url).strip().lower()); u=re.sub(r'^www\.','',u)
    return u.split('/')[0].split('?')[0] if '.' in u else None

EMAIL_OK=re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')
GENERIC={"gmail.com","yahoo.com","yahoo.co.in","hotmail.com","outlook.com","rediffmail.com","ymail.com"}
PLACEHOLDER=("example.com","domain.com","yourdomain","johnsmith","name@example","w3.org","@2x",
             ".webp",".png","@3x","sentry.io","wixpress","xxx@","@xxx","@ccc","test@","noreply@",
             "no-reply@","abc@","xyz@","email@example","user@example")
FAKE_PHONE={"7326059369","917326059369","9240466672","919240466672"}
JUNK={"","n/a","na","-","--","none","null","nil","tbd","."}
def hv(v): return v is not None and str(v).strip().lower() not in JUNK

wb=load(F)
print("="*64)
print("  FULL FAULT REPORT (read-only) — nothing changed")
print("="*64)
grand=Counter()
for S in wb.sheetnames:
    if S=="LinkedIn Profiles":
        ws=wb[S]; print(f"\n##### {S} — {ws.max_row-1} URL rows (not scanned: no company data) #####"); continue
    ws=wb[S]
    hrow=None
    for rr in (1,2,3,4):
        vals=[str(ws.cell(rr,c).value or "").strip().lower() for c in range(1,26)]
        if "company name" in vals or "company" in vals: hrow=rr; break
    if not hrow: continue
    hdr=[ws.cell(hrow,c).value for c in range(1,ws.max_column+1)]
    hm={str(h).strip().lower():i+1 for i,h in enumerate(hdr) if h}
    CO=hm.get("company name") or hm.get("company")
    PH=hm.get("phone") or hm.get("office phone")
    WEB=hm.get("website")
    EM=[hm.get(f"email {i}") for i in range(1,7)]+[hm.get(f"hr email {i}") for i in range(1,16)]; EM=[c for c in EM if c]

    blank=numeric=dupe=0
    bad_email=[]; placeholder=[]; wrong_dom=[]; fake_ph=[]; bad_ph=[]
    names=[]; email_ct=Counter(); phone_ct=Counter()
    for row in ws.iter_rows(min_row=hrow+1, values_only=True):
        def g(c): return row[c-1] if c and len(row)>=c else None
        co=g(CO)
        if not hv(co): blank+=1; continue
        s=str(co).strip(); names.append(norm(s))
        if re.fullmatch(r'[0-9]+', s): numeric+=1
        cnk=norm(s); wdom=dom(g(WEB))
        for c in EM:
            e=g(c)
            if not e or "@" not in str(e):
                if e and str(e).strip(): bad_email.append((s,str(e)))  # junk in email col
                continue
            e=str(e).strip(); el=e.lower(); email_ct[el]+=1
            if not EMAIL_OK.match(e): bad_email.append((s,e)); continue
            if any(p in el for p in PLACEHOLDER): placeholder.append((s,e)); continue
            d=el.split("@")[1]
            if d in GENERIC: continue
            dn=norm(d.split(".")[0])
            ok=(dn and (dn in cnk or cnk[:6] in dn)) or (wdom and (d==wdom or wdom.split(".")[0] in d or d.split(".")[0] in wdom))
            if not ok: wrong_dom.append((s,e))
        p=g(PH)
        if hv(p):
            praw=str(p).strip(); pd=re.sub(r'\D','',praw); phone_ct[pd]+=1
            if pd in FAKE_PHONE: fake_ph.append((s,praw))
            elif not (10<=len(pd)<=13): bad_ph.append((s,praw))
    dupe=sum(n-1 for n in Counter(names).values() if n>1)
    rep_em={e:n for e,n in email_ct.items() if n>=5}
    rep_ph={p:n for p,n in phone_ct.items() if n>=5 and p not in FAKE_PHONE}

    print(f"\n##### {S} — {len(names)} companies #####")
    def line(lbl,n,ex=None):
        s=f"  {lbl:<38}: {n:>5}"
        if ex and n: s+="   e.g. "+", ".join(f"{c}:{v}" for c,v in ex[:2])
        print(s)
    line("blank / empty company rows", blank)
    line("numeric junk names", numeric)
    line("duplicate rows (same name)", dupe)
    line("invalid-format / junk emails", len(bad_email), [(c,v[:22]) for c,v in bad_email])
    line("placeholder/fake emails", len(placeholder), [(c,v[:22]) for c,v in placeholder])
    line("wrong-domain emails (unrelated)", len(wrong_dom), [(c,v[:22]) for c,v in wrong_dom])
    line("fake phones (7326.. / 9240..)", len(fake_ph))
    line("invalid-length phones", len(bad_ph), [(c,v[:14]) for c,v in bad_ph])
    line("same email on >=5 companies", len(rep_em), [(e[:20],n) for e,n in list(rep_em.items())])
    line("same phone on >=5 companies", len(rep_ph), [(p[:14],n) for p,n in list(rep_ph.items())])
    for k,v in [("bad_email",len(bad_email)),("placeholder",len(placeholder)),("wrong_dom",len(wrong_dom)),
                ("fake_ph",len(fake_ph)),("bad_ph",len(bad_ph)),("dupe",dupe),("numeric",numeric),("blank",blank)]:
        grand[k]+=v

print("\n"+"="*64)
print("  TOTALS ACROSS ALL SHEETS")
print("="*64)
for k in ["dupe","numeric","blank","bad_email","placeholder","wrong_dom","fake_ph","bad_ph"]:
    print(f"  {k:<16}: {grand[k]}")
