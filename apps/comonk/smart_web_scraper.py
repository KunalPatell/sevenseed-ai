"""
smart_web_scraper.py — HR Email & Phone Collector (100% Free)
=============================================================
Scrapes company websites to find HR emails and phone numbers.
Works on COMONK_TRUE_MASTER.xlsx — fills missing emails & phones.

Features:
- Multi-threaded (5 concurrent) for speed
- Tries homepage + /contact + /contact-us + /about + /careers + /team
- Smart email ranking: prefers hr/recruit/career/talent/jobs emails
- Phone extraction with Indian number normalization
- Checkpoint save every 50 companies (safe to interrupt & resume)
- Resume mode: skips companies already processed

Run: python smart_web_scraper.py
"""

import openpyxl
import re
import time
import os
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin, urlparse
import httpx

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL        = "COMONK_TRUE_MASTER.xlsx"
CHECKPOINT   = "scraper_checkpoint.json"   # tracks which rows are done
BATCH_SIZE   = 50                           # save Excel every N companies
MAX_WORKERS  = 5                            # concurrent threads
REQUEST_TO   = 10                           # seconds timeout per request
MAX_PAGES    = 6                            # pages to try per company

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# Sub-pages to try per company (in priority order)
CONTACT_PATHS = [
    "",             # homepage
    "/contact",
    "/contact-us",
    "/contact-us.html",
    "/contactus",
    "/about",
    "/about-us",
    "/team",
    "/our-team",
    "/careers",
    "/jobs",
    "/work-with-us",
    "/hr",
]

# Email prefixes that indicate HR contact — ranked by quality
HR_PREFIXES = [
    "hr", "hrd", "hrbp", "humanresources",
    "recruit", "recruitment", "recruiter",
    "talent", "talentacquisition",
    "career", "careers",
    "jobs", "job",
    "hiring",
    "people",
    "join", "joinus",
    "staffing",
]

# Skip these generic/noreply emails
SKIP_PREFIXES = {
    "noreply", "no-reply", "nore-ply", "donotreply",
    "support", "help", "admin", "webmaster", "abuse",
    "postmaster", "bounce", "mailer-daemon",
    "privacy", "legal", "billing", "sales",
}

# ── Column map for COMONK_TRUE_MASTER (1-indexed) ────────────────────────────
C_COMP  = 2
C_CITY  = 3
C_EMAIL = [6, 7, 8, 9, 10, 11]   # Email 1-6
C_PHONE = 12
C_WEB   = 13
C_LI    = 14
C_SRC   = 18


# ── Regex patterns ────────────────────────────────────────────────────────────
EMAIL_RE = re.compile(
    r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b'
)
PHONE_RE = re.compile(
    r'(?:\+91[\s\-.]?)?(?:\(?0?[6-9]\d{9}\)?|'
    r'\(?079[\s\-.]?\d{4}[\s\-.]?\d{4}\)?|'
    r'\d{3,5}[\s\-.]?\d{3,5}[\s\-.]?\d{4,5})'
)


# ── Helper: normalize phone ───────────────────────────────────────────────────
def clean_phone(p):
    if not p:
        return ""
    digits = re.sub(r'[^\d]', '', str(p))
    if digits.startswith('91') and len(digits) == 12:
        digits = digits[2:]
    elif digits.startswith('0') and len(digits) == 11:
        digits = digits[1:]
    if len(digits) == 10 and digits[0] in '6789':
        return '+91 ' + digits[:5] + ' ' + digits[5:]
    if len(digits) == 8 and digits.startswith('79'):
        return '+91 79 ' + digits[2:6] + ' ' + digits[6:]
    return ""


# ── Helper: score an email (higher = better HR contact) ──────────────────────
def email_score(email):
    prefix = email.split('@')[0].lower().replace('.', '').replace('-', '').replace('_', '')
    if any(prefix == hp for hp in HR_PREFIXES):
        return 100
    if any(hp in prefix for hp in HR_PREFIXES):
        return 80
    if any(prefix == sp for sp in SKIP_PREFIXES):
        return -1  # mark as skip
    if any(sp in prefix for sp in SKIP_PREFIXES):
        return -1
    return 30  # generic but valid


def is_skip_email(email):
    dom = email.split('@')[1].lower() if '@' in email else ''
    skip_doms = {'example.com', 'test.com', 'yourdomain.com', 'domain.com',
                 'sentry.io', 'googleanalytics.com', 'wixpress.com'}
    if dom in skip_doms:
        return True
    return email_score(email) < 0


# ── Helper: extract emails + phones from HTML text ────────────────────────────
def extract_from_html(html, site_domain):
    emails_raw = EMAIL_RE.findall(html)
    phones_raw = PHONE_RE.findall(html)

    # Filter emails
    valid_emails = []
    seen = set()
    for e in emails_raw:
        e = e.lower().strip('.,;')
        if e in seen:
            continue
        seen.add(e)
        # Must be on company domain or a known HR domain
        edom = e.split('@')[1] if '@' in e else ''
        if edom == site_domain or not is_skip_email(e):
            if '@' in e and '.' in edom and len(edom) > 3:
                valid_emails.append(e)

    # Sort: HR emails first
    valid_emails.sort(key=email_score, reverse=True)
    valid_emails = [e for e in valid_emails if email_score(e) >= 0]

    # Filter phones
    valid_phones = []
    for p in phones_raw:
        cleaned = clean_phone(p)
        if cleaned and cleaned not in valid_phones:
            valid_phones.append(cleaned)

    return valid_emails[:6], valid_phones[:3]


# ── Core: scrape one company ──────────────────────────────────────────────────
def scrape_company(row_data):
    """
    row_data = (row_index, company_name, website_url, has_email, has_phone)
    Returns dict with found emails and phones.
    """
    row_idx, name, website, has_email, has_phone = row_data

    if not website:
        return {"row": row_idx, "emails": [], "phones": [], "status": "no_website"}

    website = website.strip()
    if not website.startswith("http"):
        website = "https://" + website

    parsed = urlparse(website)
    site_domain = parsed.netloc.replace("www.", "").strip()
    base_url = f"{parsed.scheme}://{parsed.netloc}"

    all_emails = []
    all_phones = []

    tried = 0
    with httpx.Client(headers=HEADERS, timeout=REQUEST_TO,
                      follow_redirects=True, verify=False) as client:
        for path in CONTACT_PATHS:
            if tried >= MAX_PAGES:
                break
            # Skip pages we don't need
            if has_email and has_phone:
                break

            url = base_url + path if path else website
            try:
                resp = client.get(url, timeout=REQUEST_TO)
                if resp.status_code not in (200, 201):
                    continue
                html = resp.text
                tried += 1

                emails, phones = extract_from_html(html, site_domain)

                # Add new unique emails
                for e in emails:
                    if e not in all_emails:
                        all_emails.append(e)
                # Add new unique phones
                for p in phones:
                    if p not in all_phones:
                        all_phones.append(p)

                # Stop early if we have both
                if all_emails and all_phones:
                    break

            except httpx.TimeoutException:
                break
            except Exception:
                continue

    return {
        "row": row_idx,
        "emails": all_emails[:6],
        "phones": all_phones[:2],
        "status": "ok" if (all_emails or all_phones) else "not_found",
        "tried_pages": tried,
    }


# ── Load checkpoint ───────────────────────────────────────────────────────────
def load_checkpoint():
    if os.path.exists(CHECKPOINT):
        try:
            with open(CHECKPOINT, "r") as f:
                data = json.load(f)
                return set(data.get("done_rows", []))
        except Exception:
            pass
    return set()


def save_checkpoint(done_rows):
    with open(CHECKPOINT, "w") as f:
        json.dump({"done_rows": list(done_rows)}, f)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("\n=== Comonk HR Web Scraper ===")
    print(f"  Excel: {EXCEL}")
    print(f"  Threads: {MAX_WORKERS}  |  Pages/company: {MAX_PAGES}\n")

    # Load workbook
    wb = openpyxl.load_workbook(EXCEL)
    # Try sheet name
    sheet_name = "COMPLETE_MASTER" if "COMPLETE_MASTER" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    print(f"  Sheet: '{sheet_name}'  |  Rows: {ws.max_row - 1}")

    # Load checkpoint (already processed rows)
    done_rows = load_checkpoint()
    print(f"  Checkpoint: {len(done_rows)} rows already done\n")

    # Build target list
    targets = []
    for r in range(2, ws.max_row + 1):
        if r in done_rows:
            continue
        name    = ws.cell(r, C_COMP).value
        website = ws.cell(r, C_WEB).value
        if not name:
            continue
        has_email = any(ws.cell(r, c).value for c in C_EMAIL)
        has_phone = bool(ws.cell(r, C_PHONE).value)
        # Only process if missing email OR phone
        if has_email and has_phone:
            done_rows.add(r)
            continue
        targets.append((r, str(name), website or "", has_email, has_phone))

    print(f"  Companies to process: {len(targets)}")
    print(f"  (missing email: {sum(1 for t in targets if not t[3])}, "
          f"missing phone: {sum(1 for t in targets if not t[4])})\n")

    if not targets:
        print("  Nothing to do — all companies have email + phone!")
        return

    # Process in batches
    batch_count = 0
    total_em_added = 0
    total_ph_added = 0
    total_found = 0
    processed = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(scrape_company, t): t for t in targets}

        for future in as_completed(futures):
            result = future.result()
            row_idx  = result["row"]
            emails   = result["emails"]
            phones   = result["phones"]
            status   = result["status"]
            row_data = futures[future]
            name     = row_data[1]

            processed += 1
            done_rows.add(row_idx)

            em_added = 0
            ph_added = 0

            if emails:
                # Get existing emails to avoid duplicates
                existing = set(
                    str(ws.cell(row_idx, c).value).lower()
                    for c in C_EMAIL
                    if ws.cell(row_idx, c).value
                )
                empty_cols = [c for c in C_EMAIL if not ws.cell(row_idx, c).value]
                new_emails = [e for e in emails if e.lower() not in existing]
                for e, col in zip(new_emails, empty_cols):
                    ws.cell(row_idx, col).value = e
                    em_added += 1

            if phones and not ws.cell(row_idx, C_PHONE).value:
                ws.cell(row_idx, C_PHONE).value = phones[0]
                ph_added += 1

            if em_added or ph_added:
                total_found += 1
                total_em_added += em_added
                total_ph_added += ph_added
                # Update source
                src = ws.cell(row_idx, C_SRC).value or ""
                if "WEB_SCRAPE" not in src:
                    ws.cell(row_idx, C_SRC).value = (src + ", WEB_SCRAPE").strip(", ")

            batch_count += 1
            symbol = "✓" if (em_added or ph_added) else "·"
            pages = result.get("tried_pages", 0)
            print(
                f"  [{processed:4d}/{len(targets)}] {symbol} {name[:40]:<40} "
                f"| +{em_added}em +{ph_added}ph | pages:{pages}"
            )

            # Checkpoint save every BATCH_SIZE
            if batch_count >= BATCH_SIZE:
                wb.save(EXCEL)
                save_checkpoint(done_rows)
                print(f"\n  --- CHECKPOINT: {total_found} found so far "
                      f"(+{total_em_added} emails, +{total_ph_added} phones) ---\n")
                batch_count = 0

    # Final save
    wb.save(EXCEL)
    save_checkpoint(done_rows)

    # Final stats
    total = ws.max_row - 1
    with_em = sum(1 for r in range(2, ws.max_row + 1)
                  if any(ws.cell(r, c).value for c in C_EMAIL))
    with_ph = sum(1 for r in range(2, ws.max_row + 1)
                  if ws.cell(r, C_PHONE).value)

    print(f"\n{'='*60}")
    print(f"  SCRAPING COMPLETE")
    print(f"  Companies processed: {processed}")
    print(f"  Companies updated:   {total_found}")
    print(f"  Emails added:        +{total_em_added}")
    print(f"  Phones added:        +{total_ph_added}")
    print(f"\n  === FINAL COVERAGE ({total} companies) ===")
    print(f"  Email:  {with_em} ({round(with_em/total*100)}%)")
    print(f"  Phone:  {with_ph} ({round(with_ph/total*100)}%)")
    print(f"  Saved:  {EXCEL}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
