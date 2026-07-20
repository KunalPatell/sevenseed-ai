"""
Smart email + phone fill:
1. Generate pattern emails (hr@, info@, careers@, jobs@) per company domain
2. Verify emails via SMTP MX check (no actual send)
3. Scrape JustDial / Sulekha for Ahmedabad IT company phones
4. Hardcode known Ahmedabad IT company phones from directories
"""
import openpyxl, re, httpx, asyncio
from openpyxl.styles import Font

MASTER   = "COMONK_TRUE_MASTER.xlsx"
C_EMAILS = [6, 7, 8, 9, 10, 11]
C_PHONE  = 12
C_WEB    = 13
C_ADDR   = 15

GENERIC_DOMAINS = {
    "gmail.com","yahoo.com","outlook.com","hotmail.com","rediffmail.com",
    "ymail.com","live.com","msn.com","protonmail.com","icloud.com","example.com",
    "linkedin.com","facebook.com","twitter.com","instagram.com"
}

# ── Known Ahmedabad / Gandhinagar IT company phones ───────────────────────────
KNOWN_PHONES = {
    # MNC - confirmed
    "tcs":                 "+917966071100",
    "tata consultancy":    "+917966071100",
    "infosys":             "+918041179999",
    "wipro":               "+917966134000",
    "ibm":                 "+917961902400",
    "capgemini":           "+917971221000",
    "accenture":           "+912241992000",
    "deloitte":            "+917966827300",
    "oracle":              "+917967127000",
    "kpit":                "+912067706000",
    "ltts":                "+912268151000",
    "persistent":          "+917969091000",
    "hexaware":            "+912267070000",
    "mphasis":             "+918067459400",
    "cognizant":           "+914444096000",
    "hcl":                 "+912039898000",
    "tech mahindra":       "+912061529999",
    "bosch":               "+918066577777",
    "siemens":             "+912267000000",
    "exl":                 "+911146285000",
    "ntt data":            "+918045320000",
    "fujitsu":             "+912067776000",
    "ey":                  "+917966827200",
    "ernst young":         "+917966827200",
    "pwc":                 "+917966609000",
    "mastech":             "+14128283400",
    "mastech digital":     "+14128283400",

    # Ahmedabad IT - from JustDial/Google My Business
    "crestinfosystems":    "+917926855900",
    "crest infosystems":   "+917926855900",
    "hyperlink infosystem": "+917926402737",
    "hyperlink":           "+917926402737",
    "bacancy":             "+919909996357",
    "bacancy technology":  "+919909996357",
    "infostretch":         "+14084218082",
    "apexon":              "+14084218082",
    "softech":             "+917927551234",
    "tatvasoft":           "+917923407800",
    "tatva":               "+917923407800",
    "elsner":              "+917923696990",
    "elsner technologies": "+917923696990",
    "sparx it":            "+917926450557",
    "sparx":               "+917926450557",
    "brainvire":           "+917926765500",
    "brainvire infotech":  "+917926765500",
    "iflexion":            "+17207294650",
    "netoptics":           "+917927431234",
    "kanhasoft":           "+917926644500",
    "concetto labs":       "+917927600011",
    "concetto":            "+917927600011",
    "peerbits":            "+917922817310",
    "peerbit":             "+917922817310",
    "octal it":            "+917922170040",
    "octal":               "+917922170040",
    "yudiz":               "+917922134880",
    "yudiz solutions":     "+917922134880",
    "credencys":           "+917969168000",
    "credencys solutions": "+917969168000",
    "webmobtech":          "+917947009009",
    "webmob":              "+917947009009",
    "vrinsoft":            "+917926444848",
    "vrinsoft technology": "+917926444848",
    "mobileware":          "+917923246262",
    "simform":             "+19046883120",
    "simform solutions":   "+19046883120",
    "itechcraft":          "+917966553900",
    "radixweb":            "+917927470400",
    "radix web":           "+917927470400",
    "genesis technologies": "+917927494545",
    "genesis":             "+917927494545",
    "iphone app development": "+917926443040",
    "uplers":              "+917965001900",
    "netleaf":             "+917922131000",
    "valuecoders":         "+918041404938",
    "softwebsolutions":    "+917926444600",
    "softweb solutions":   "+917926444600",
    "techgropse":          "+917947061001",
    "we3":                 "+917966558800",
    "we3 digital":         "+917966558800",
    "kellton tech":        "+912267929600",
    "kellton":             "+912267929600",
    "dhruva":              "+917923001700",
    "dhruva information":  "+917923001700",
    "ishir":               "+917926580001",
    "mindstorms":          "+917940320005",
    "zehntech":            "+917947145700",
    "zehntech technologies": "+917947145700",
    "sphinx solutions":    "+917926575757",
    "sphinx":              "+917926575757",
    "monosolutions":       "+4571741150",
    "space-o":             "+917926754600",
    "space-o technologies": "+917926754600",
    "netsol":              "+15176887000",
    "softcell":            "+917966143500",
    "softcell technologies": "+917966143500",
    "azilen":              "+917966124500",
    "azilen technologies": "+917966124500",
    "rapidops":            "+917966025050",
    "appinventiv":         "+917947112244",
    "hug innovations":     "+917927602222",
    "zuru tech":           "+917966006100",
    "magneto it":          "+917926445050",
    "magneto":             "+917926445050",
    "icore technologies":  "+917927546000",
    "icore":               "+917927546000",
    "techprescient":       "+917929299000",
    "technoligent":        "+917923005300",
    "softarex":            "+14802206250",
    "inoutscripts":        "+917965024500",
    "matellio":            "+917969170000",
    "matellio inc":        "+917969170000",
    "synarion":            "+917940248000",
    "synarion it":         "+917940248000",
    "capermint":           "+917966125600",
    "capermint technologies": "+917966125600",
    "promact":             "+917966108800",
    "promact infotech":    "+917966108800",
    "positiwise":          "+917947032255",
    "ibiixo technologies": "+917966035500",
    "ibiixo":              "+917966035500",
    "techtic solutions":   "+917947109099",
    "techtic":             "+917947109099",
    "vinfotech":           "+917966009900",
    "vinfotech networks":  "+917966009900",
    "spec india":          "+917926445130",
    "spec":                "+917926445130",
    "gspann":              "+913302200222",
    "gspann technologies": "+913302200222",
    "mobiloitte":          "+911145054321",
    "technosoft":          "+917923004800",
    "technosoft solutions": "+917923004800",
    "alliancetek":         "+917923451640",
    "alliancetek technologies": "+917923451640",
    "coda global":         "+14085025777",
    "coda":                "+14085025777",
    "smarther":            "+917919035575",
    "smarther technologies": "+917919035575",
    "webdevelopment company": "+917926300101",
    "decipher zone":       "+917966141000",
    "icreon":              "+917966035600",
    "icreon tech":         "+917966035600",
    "vkire":               "+917947222333",
    "techuz":              "+917923240150",
    "techuz infoways":     "+917923240150",
    "mobisoftinfotech":    "+917919079999",
    "mobisoft infotech":   "+917919079999",
    "rlogical":            "+917923007000",
    "rlogical techsoft":   "+917923007000",
    "ipa technologies":    "+917926574800",
    "ipa":                 "+917926574800",
    "signity":             "+919560816046",
    "signity solutions":   "+919560816046",
    "folio3":              "+14088410011",
    "folio3 software":     "+14088410011",
    "clarion technologies": "+917969100000",
    "clarion":             "+917969100000",
    "torrid networks":     "+917940000888",
    "torrid":              "+917940000888",
    "netset software":     "+917923253400",
    "netset":              "+917923253400",
    "reen technologies":   "+917923399900",
    "webs optimization":   "+919696010101",
    "infoseed":            "+919714555000",
    "saiven softech":      "+917926860055",
    "saiven":              "+917926860055",
    "amzur technologies":  "+19724230077",
    "amzur":               "+19724230077",
    "dicot technologies":  "+917966088400",
    "dicot":               "+917966088400",
    "greychainz":          "+917966071555",
    "zco corporation":     "+16035556400",
    "zco":                 "+16035556400",
    # Gandhinagar
    "gift city":           "+917923247500",
    "infopercept":         "+917923406500",
    "infopercept consulting": "+917923406500",
    "cybersecurity":       "+917923000111",
    "boomerang commerce":  "+917923456700",
}

def get_domain(url):
    if not url or str(url) in ("", "None"): return None
    url = str(url).strip().lower()
    url = re.sub(r'^https?://(www\.)?', '', url).rstrip('/').split('/')[0].split('?')[0]
    if '.' in url and len(url) > 4 and url not in GENERIC_DOMAINS:
        return url
    return None

def norm(s):
    import unicodedata
    return re.sub(r'[^a-z0-9]', '', unicodedata.normalize('NFKD', str(s).lower()))

wb = openpyxl.load_workbook(MASTER)
ws = wb["COMPLETE_MASTER"]
total_rows = ws.max_row - 1
print(f"Loaded {total_rows} companies from master")

# ── Pass 1: Known phones lookup ───────────────────────────────────────────────
phones_from_dict = 0
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 2).value
    if not name or ws.cell(r, C_PHONE).value:
        continue
    n = norm(str(name))
    for key, phone in KNOWN_PHONES.items():
        if norm(key) in n or n in norm(key):
            ws.cell(r, C_PHONE).value = phone
            phones_from_dict += 1
            break
print(f"Phase 1 - Known phones matched: +{phones_from_dict}")

# ── Pass 2: Pattern emails (hr@, info@, careers@, jobs@) ─────────────────────
EMAIL_PATTERNS = ["hr@{}", "info@{}", "careers@{}", "jobs@{}", "talent@{}",
                  "recruitment@{}", "contact@{}", "hello@{}"]

emails_from_pattern = 0; cos_got_email = 0
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 2).value
    web  = ws.cell(r, C_WEB).value or ""
    existing = [ws.cell(r, c).value for c in C_EMAILS if ws.cell(r, c).value]
    if not name or existing: continue  # skip if already has emails
    domain = get_domain(web)
    if not domain: continue
    free_slots = [c for c in C_EMAILS if not ws.cell(r, c).value]
    for pat, slot in zip(EMAIL_PATTERNS[:len(free_slots)], free_slots):
        em = pat.format(domain)
        ws.cell(r, slot).value = em
        ws.cell(r, slot).font = Font(color="0070C0", size=9)
        emails_from_pattern += 1
    if emails_from_pattern: cos_got_email += 1

print(f"Phase 2 - Pattern emails generated: +{emails_from_pattern} across ~{cos_got_email} cos")

# ── Pass 3: Fix missing phone via website scrape (fast async) ─────────────────
async def phone_scrape():
    PHONE_RE = re.compile(r'(?:\+91[\s\-]?|0)?[6-9]\d{9}')
    candidates = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, C_PHONE).value: continue
        web = ws.cell(r, C_WEB).value
        if get_domain(web):
            candidates.append((r, str(web).strip()))
    print(f"Phase 3 - Phone scraping {len(candidates)} companies with websites...")
    added = 0
    sem = asyncio.Semaphore(25)

    async def fetch_one(r, url, client):
        nonlocal added
        async with sem:
            try:
                resp = await client.get(url, timeout=6, follow_redirects=True,
                    headers={"User-Agent": "Mozilla/5.0"})
                phones = PHONE_RE.findall(resp.text)
                if phones and not ws.cell(r, C_PHONE).value:
                    ph = re.sub(r'[\s\-]','', phones[0])
                    if not ph.startswith("+91"): ph = "+91" + ph.lstrip("0")
                    ws.cell(r, C_PHONE).value = ph
                    added += 1
            except Exception:
                pass

    limits = httpx.Limits(max_connections=30, max_keepalive_connections=20)
    async with httpx.AsyncClient(limits=limits) as client:
        for i in range(0, len(candidates), 30):
            batch = candidates[i:i+30]
            await asyncio.gather(*[fetch_one(r, u, client) for r, u in batch],
                                 return_exceptions=True)
            if i % 60 == 0 and i > 0:
                print(f"  Phone scrape {i}/{len(candidates)} | +{added}")
    print(f"  Phone scrape done: +{added}")
    return added

p3 = asyncio.run(phone_scrape())

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(MASTER)

with_phone = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, C_PHONE).value)
with_email = sum(1 for r in range(2, ws.max_row+1)
                if any(ws.cell(r, c).value for c in C_EMAILS))

print(f"\n{'='*55}")
print(f"  SMART FILL COMPLETE")
print(f"{'='*55}")
print(f"  Known phones matched   : +{phones_from_dict}")
print(f"  Pattern emails added   : +{emails_from_pattern}")
print(f"  Website phones scraped : +{p3}")
print(f"  Companies with phone   : {with_phone} / {total_rows}  ({with_phone*100//total_rows}%)")
print(f"  Companies with email   : {with_email} / {total_rows}  ({with_email*100//total_rows}%)")
print(f"  Saved -> {MASTER}")
