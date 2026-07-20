"""Count existing emails per MNC domain from all our data."""
import openpyxl, re, os
from collections import defaultdict

EMAIL_RE=re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
MNCS = {
 "TCS":"tcs.com","Infosys":"infosys.com","Wipro":"wipro.com","HCL Technologies":"hcltech.com",
 "Tech Mahindra":"techmahindra.com","Capgemini":"capgemini.com","Accenture":"accenture.com",
 "IBM":"ibm.com","Cognizant":"cognizant.com","Deloitte":"deloitte.com","EY":"ey.com",
 "PwC":"pwc.com","Bosch":"bosch.com","Siemens":"siemens.com","Oracle":"oracle.com",
 "NTT Data":"nttdata.com","Fujitsu":"fujitsu.com","KPIT":"kpit.com","LTTS":"ltts.com",
 "Mphasis":"mphasis.com","Persistent":"persistent.com","Hexaware":"hexaware.com",
 "Apexon":"apexon.com","Mastech Digital":"mastechdigital.com","EXL":"exlservice.com",
}
emails=set()
def slurp(t):
    for e in EMAIL_RE.findall(t): emails.add(e.lower())
for fn in os.listdir("."):
    if fn.startswith("_extract_") and fn.endswith(".txt"): slurp(open(fn,encoding="utf-8",errors="ignore").read())
for fn in ["emails.csv","AIML_Companies_Apply_List.csv","Gandhinagar_GIFTCity_Companies.csv","hr_vcards_15-6-26.vcf"]:
    if os.path.exists(fn): slurp(open(fn,encoding="utf-8",errors="ignore").read())
for fn in ["Ahmedabad_IT_AIML_FINAL_MASTER.xlsx","COMONK_TRUE_MASTER.xlsx","HR Mail List.xlsx",
           "Ahmedabad_IT_AIML_Companies_Master.xlsx","Ahmedabad_AIML_Companies_HR_Contacts.xlsx"]:
    if os.path.exists(fn):
        w=openpyxl.load_workbook(fn,read_only=True)
        for s in w.sheetnames:
            for row in w[s].iter_rows(values_only=True):
                for v in row:
                    if v: slurp(str(v))
        w.close()
by_dom=defaultdict(list)
for e in emails: by_dom[e.split("@")[1]].append(e)
print("MNC                domain               have  need")
for name,dom in MNCS.items():
    have=set(by_dom.get(dom,[]))
    for d,el in by_dom.items():
        if d.endswith("."+dom): have.update(el)
    have=sorted(have)
    print(f"  {name:17} {dom:20} {len(have):3}  +{max(0,10-len(have))}")
    for e in have: print(f"       {e}")
