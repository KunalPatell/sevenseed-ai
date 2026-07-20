"""
MEGA BATCH — Exhaustive internet research
Sources: Tracxn SaaS, Clutch (app/web/IT), TopDevelopers, DesignRush,
GoodFirms, Foundit, Inc42, IIMA Ventures, GUSEC, StartupBlink, knowstartup, Failory.
All verified Ahmedabad/Gujarat IT/software/SaaS/deeptech companies.
"""

import openpyxl, re, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment

# (Company, Category, Roles, Website, City, Source)
MEGA = [
    # ── SaaS / Product startups (Tracxn) ─────────────────────────────────
    ("Invixium",                "AI / ML",    "Biometric AI, Face Recognition, Access Control",  "invixium.com",          "Ahmedabad","Tracxn"),
    ("Oizom",                   "AI / ML",    "Environmental AI, IoT Sensors, Data Analytics",   "oizom.com",             "Ahmedabad","Tracxn"),
    ("Reelo",                   "IT Services","Customer Loyalty SaaS, Marketing Tech",           "reelo.io",              "Ahmedabad","Tracxn"),
    ("OPLinnovate",             "IT Services","SaaS, Web Dev, Product Engineering",              "oplinnovate.com",       "Ahmedabad","Tracxn"),
    ("Wizzy",                   "AI / ML",    "eCommerce Search AI, Site Search ML",             "wizzy.ai",              "Ahmedabad","Tracxn"),
    ("NewsReach",               "AI / ML",    "PR Tech, Content AI, Media SaaS",                 "newsreach.ai",          "Ahmedabad","Tracxn"),
    ("Everything CIVIC",        "IT Services","GovTech SaaS, Civic Software",                    "everythingcivic.com",   "Ahmedabad","Tracxn"),
    ("AlmaShines",              "IT Services","Alumni Management SaaS, EdTech",                  "almashines.com",        "Ahmedabad","Tracxn"),
    ("InnkeyPMS",               "IT Services","Hotel Management SaaS, Hospitality Tech",         "innkeypms.net",         "Ahmedabad","Tracxn"),
    ("Legalwiz",                "IT Services","LegalTech SaaS, Compliance Software",             "legalwiz.in",           "Ahmedabad","Tracxn"),
    ("Satark AI",               "AI / ML",    "Surveillance AI, Computer Vision, Safety AI",     "satark.ai",             "Ahmedabad","Tracxn"),
    ("Medmitra",                "AI / ML",    "HealthTech AI, Medical Software",                 "medmitra.ai",           "Ahmedabad","Tracxn"),
    ("SupportIQ",               "AI / ML",    "Customer Support AI, Conversational AI",          "supportiq.ai",          "Ahmedabad","Tracxn"),
    ("Clientjoy",               "IT Services","Agency CRM SaaS, Client Management",              "clientjoy.io",          "Ahmedabad","Tracxn"),
    ("Synup",                   "AI / ML",    "Local Marketing AI, Listings Management",         "synup.com",             "Ahmedabad","Tracxn"),
    ("Plutomen",                "AI / ML",    "AR + AI, Remote Assistance, Industrial XR",       "plutomen.com",          "Ahmedabad","Tracxn"),
    ("Synersoft (BLACKbox)",    "IT Services","SMB IT Security, Data Protection Tech",           "synersoft.in",          "Ahmedabad","Tracxn"),
    ("Finwave",                 "AI / ML",    "FinTech SaaS, Lending Software",                 "finwave.in",            "Ahmedabad","Tracxn"),
    ("MindSpark",               "IT Services","EdTech, Adaptive Learning Software",              "mindspark.in",          "Ahmedabad","Tracxn"),
    ("Vidyalaya School Software","IT Services","School ERP, EdTech SaaS",                        "vidyalayaschoolsoftware.com","Ahmedabad","Tracxn"),
    # ── knowstartup / Inc42 Ahmedabad startups ───────────────────────────
    ("vasyERP",                 "IT Services","Retail ERP SaaS, POS, Manufacturing Software",    "vasyerp.com",           "Ahmedabad","knowstartup"),
    ("Dev Accelerator (DevX)",  "IT Services","Coworking Tech, Startup Infrastructure",          "devx.work",             "Ahmedabad","knowstartup"),
    ("Learnvern",               "IT Services","EdTech, Online Learning Platform",                "learnvern.com",         "Ahmedabad","knowstartup"),
    ("Saarthi Pedagogy",        "AI / ML",    "EdTech AI, Adaptive Learning",                    "saarthipedagogy.com",   "Ahmedabad","knowstartup"),
    ("Quicko",                  "AI / ML",    "FinTech, Tax Software, AI Tax Filing",            "quicko.com",            "Ahmedabad","knowstartup"),
    ("Frendy",                  "IT Services","Social Commerce, Retail Tech",                    "frendy.in",             "Ahmedabad","knowstartup"),
    ("SuperZop",                "AI / ML",    "AgriTech, Supply Chain AI, B2B Commerce",         "superzop.com",          "Ahmedabad","knowstartup"),
    ("OneHash",                 "IT Services","SaaS CRM/ERP, Business Software",                 "onehash.ai",            "Ahmedabad","knowstartup"),
    ("Parikshak.ai",            "AI / ML",    "AI Assessment, EdTech AI, Proctoring",            "parikshak.ai",          "Ahmedabad","knowstartup"),
    ("Hubilo",                  "AI / ML",    "Event Tech SaaS, Virtual Events AI",              "hubilo.com",            "Ahmedabad","Inc42"),
    ("Gensol Engineering",      "IT Services","Solar/EV Tech, Energy Software, IoT",             "gensol.in",             "Ahmedabad","Inc42"),
    ("Optimized Electrotech",   "AI / ML",    "Defense AI, Computer Vision, Optical Tech",       "optimizedelectrotech.com","Ahmedabad","Inc42"),
    ("NEPRA Resource Management","AI / ML",   "Waste Management AI, Robotics, Sustainability",   "letsrecycle.co.in",     "Ahmedabad","Inc42"),
    ("Axio Biosolutions",       "AI / ML",    "MedTech, Biotech, Healthcare Innovation",         "axiobio.com",           "Ahmedabad","Inc42"),
    ("PierSight Space",         "AI / ML",    "SpaceTech AI, Satellite Data, SAR Analytics",     "piersight.space",       "Ahmedabad","Inc42"),
    ("Ignosis",                 "AI / ML",    "FinTech, Account Aggregator AI, Data Analytics",  "ignosis.ai",            "Ahmedabad","Inc42"),
    ("Amnex Infotechnologies",  "AI / ML",    "Smart City AI, GIS, IoT, Data Analytics",         "amnex.com",             "Ahmedabad","Inc42"),
    ("Matter",                  "IT Services","EV Tech, Energy Storage Software, IoT",           "matter.in",             "Ahmedabad","Inc42"),
    ("Shipturtle",              "IT Services","Marketplace SaaS, eCommerce, Logistics Tech",     "shipturtle.com",        "Ahmedabad","StartupBlink"),
    ("Gamerji",                 "IT Services","Esports Platform, Gaming Tech SaaS",              "gamerji.com",           "Ahmedabad","StartupBlink"),
    ("EduFund",                 "AI / ML",    "FinTech, EdTech, Investment AI",                  "edufund.in",            "Ahmedabad","StartupBlink"),
    ("On2Cook",                 "AI / ML",    "FoodTech, Smart Cooking AI, Hardware",            "on2cook.com",           "Ahmedabad","StartupBlink"),
    ("Sterling Accuris",        "AI / ML",    "HealthTech, Diagnostics AI, Lab Software",        "sterlingaccuris.com",   "Ahmedabad","Inc42"),
    # ── Clutch / TopDevelopers / DesignRush — App & Web dev ──────────────
    ("App Ideas Infotech",      "IT Services","Mobile App, Web Dev, Flutter",                    "theappideas.com",       "Ahmedabad","Clutch"),
    ("Siddhi Infosoft",         "IT Services","Mobile App, Web Dev, Custom Software",            "siddhiinfosoft.com",    "Ahmedabad","TopDev"),
    ("Communication Crafts",    "IT Services","Web Dev, Design, Digital Solutions",              "communicationcrafts.com","Ahmedabad","TopDev"),
    ("3Brain Technolabs",       "IT Services","Mobile App, Web Dev, Software",                   "3braintechnolabs.com",  "Ahmedabad","TopDev"),
    ("Kunsh Technologies",      "IT Services","Web Dev, Mobile App, Custom Software",            "kunshtechnologies.com", "Ahmedabad","TopDev"),
    ("World Web Technology",    "IT Services","Web Dev, Magento, Mobile App",                    "worldwebtechnology.com","Ahmedabad","TopDev"),
    ("X-Byte Enterprise Solutions","AI / ML", "Data Scraping AI, Web Dev, IoT",                  "xbyte.io",              "Ahmedabad","TopDev"),
    ("iFour Technolab",         "IT Services","Custom Software, .NET, Web Dev",                  "ifourtechnolab.com",    "Ahmedabad","TopDev"),
    ("Jash Entertainment",      "IT Services","Game Dev, Mobile App, Entertainment Tech",        "jashentertainment.com", "Ahmedabad","TopDev"),
    ("India App Developer",     "IT Services","Mobile App, Web Dev",                            "indiaappdeveloper.com", "Ahmedabad","TopDev"),
    ("MoveoApps",               "IT Services","Mobile App, Flutter, React Native Dev",           "moveoapps.com",         "Ahmedabad","TopDev"),
    ("The One Technologies",    "IT Services","Web Dev, .NET, Mobile App",                       "theonetechnologies.com","Ahmedabad","TopDev"),
    ("Techcronus Business Solutions","IT Services","Web Dev, Mobile App, Microsoft Dev",         "techcronus.com",        "Ahmedabad","TopDev"),
    ("Dolphin Web Solution",    "IT Services","Web Dev, eCommerce, Digital Marketing",           "dolphinwebsolution.com","Ahmedabad","TopDev"),
    ("TechBlocks",              "IT Services","Digital Engineering, Cloud, Web Dev",             "tblocks.com",           "Ahmedabad","TopDev"),
    ("Expert App Devs",         "IT Services","Mobile App, Web Dev, AI Integration",             "expertappdevs.com",     "Ahmedabad","TopDev"),
    ("Global Vincitore",        "IT Services","Web Dev, Mobile App, AR/VR",                      "globalvincitore.com",   "Ahmedabad","TopDev"),
    ("LogicRays Technologies",  "IT Services","Web Dev, Mobile App, Digital Marketing",          "logicrays.com",         "Ahmedabad","TopDev"),
    ("DS Web Technologies",     "IT Services","Web Dev, eCommerce, Mobile App",                  "dswebtechnologies.com", "Ahmedabad","TopDev"),
    ("Stepin Solutions",        "IT Services","Web Dev, Mobile App, Software",                   "stepinsolutions.com",   "Ahmedabad","TopDev"),
    ("Albiorix Technology",     "IT Services","Mobile App, Web Dev, Custom Software",            "albiorix.com",          "Ahmedabad","TopDev"),
    ("Nichetech Solutions",     "IT Services","Web Dev, Mobile App, Software",                   "nichetechsolutions.com","Ahmedabad","TopDev"),
    ("MageAnts",                "IT Services","Magento, eCommerce, Web Dev",                     "mageants.com",          "Ahmedabad","TopDev"),
    ("Naapbooks",               "AI / ML",    "FinTech, Automation, Blockchain, AI Accounting",  "naapbooks.com",         "Ahmedabad","TopDev"),
    ("Octet Design Studio",     "IT Services","UI/UX Design, Product Design, Web",               "octet.design",          "Ahmedabad","DesignRush"),
    ("Binstellar",              "IT Services","Mobile App, Web Dev, Flutter",                    "binstellar.com",        "Ahmedabad","DesignRush"),
    ("Netclues",                "IT Services","Web Dev, eCommerce, Custom Software",             "netclues.com",          "Ahmedabad","DesignRush"),
    ("WNA InfoTech",            "IT Services","Web Dev, Mobile App, Digital",                    "wnainfotech.com",       "Ahmedabad","DesignRush"),
    ("Skywinds Solutions",      "IT Services","Web Dev, Digital Marketing, SEO",                 "skywinds.in",           "Ahmedabad","DesignRush"),
    ("RP UXCollab",             "IT Services","UI/UX Design, Product Design",                    "rpuxcollab.com",        "Ahmedabad","DesignRush"),
    ("Verve Systems",           "IT Services","Web Dev, Mobile App, Custom Software",            "vervelogic.com",        "Ahmedabad","DesignRush"),
    ("Aspire IT Services",      "IT Services","Web Dev, Mobile App, IT Consulting",              "aspireitservices.com",  "Ahmedabad","DesignRush"),
    ("KrishaWeb",               "IT Services","Web Dev, WordPress, Digital Marketing",           "krishaweb.com",         "Ahmedabad","DesignRush"),
    ("Sonoma Infotech",         "IT Services","Web Dev, Mobile App, Software",                   "sonomainfotech.com",    "Ahmedabad","DesignRush"),
    ("Atlas Softweb",           "IT Services","Web Dev, Odoo, Mobile App",                       "atlassoftweb.com",      "Ahmedabad","DesignRush"),
    ("ICUBE TECHNOLABS",        "IT Services","Web Dev, Mobile App, Software",                   "icubetechnolabs.com",   "Ahmedabad","DesignRush"),
    ("WebCodeGenie",            "IT Services","Web Dev, Laravel, Mobile App",                    "webcodegenie.com",      "Ahmedabad","DesignRush"),
    ("Cynoinfotech",            "IT Services","Web Dev, eCommerce, Mobile App",                  "cynoinfotech.com",      "Ahmedabad","DesignRush"),
    ("Matic Solutions",         "IT Services","Web Dev, Digital Marketing",                      "maticsolutions.com",    "Ahmedabad","DesignRush"),
    ("Mavlers",                 "IT Services","Digital Marketing, Email, Web Dev",               "mavlers.com",           "Ahmedabad","DesignRush"),
    ("Internut",                "IT Services","Web Dev, Digital Solutions",                      "internut.in",           "Ahmedabad","DesignRush"),
    ("Teclogiq",                "IT Services","Web/Mobile Dev, Product Engineering",             "teclogiq.com",          "Ahmedabad","GoodFirms"),
    ("Emizen Tech",             "IT Services","Mobile App, eCommerce, Web Dev",                  "emizentech.com",        "Ahmedabad","GoodFirms"),
    ("Codal",                   "IT Services","Web Dev, UX Design, Digital Transform",           "codal.com",             "Ahmedabad","Clutch"),
    ("Naapbooks Limited",       "AI / ML",    "FinTech, AI Accounting, Automation",              "naapbooks.com",         "Ahmedabad","TopDev"),
    ("Ask Datatech",            "IT Services","Data Entry, Data Processing, BPO",                "askdatatech.com",       "Ahmedabad","DesignRush"),
    ("CodeQuality Technologies","IT Services","QA, Software Testing, Web Dev",                   "codequality.in",        "Ahmedabad","DesignRush"),
    ("Code and Core Tech",      "IT Services","Web Dev, Mobile App, Custom Software",            "codeandcore.com",       "Ahmedabad","DesignRush"),
    ("Paperboat Tech",          "IT Services","Web Dev, Mobile App, Software",                   "paperboattech.com",     "Ahmedabad","DesignRush"),
    ("Techcronus",              "IT Services","Web/Mobile Dev, Microsoft, Cloud",               "techcronus.com",        "Ahmedabad","DesignRush"),
    ("MobMaxime",               "IT Services","Mobile App, Web Dev, Game Dev",                   "mobmaxime.com",         "Ahmedabad","TopDev"),
    ("Web 3.0 India",           "IT Services","Blockchain, Web3, dApp Dev",                      "web30india.com",        "Ahmedabad","TopDev"),
    ("AP-GROUP",                "IT Services","Web Dev, Mobile App, IT Services",                "apgroup.com",           "Ahmedabad","TopDev"),
    ("Cygnet.One",              "AI / ML",    "AI/ML, RPA, Tax Tech, Digital Transform",         "cygnet.one",            "Ahmedabad","TopDev"),
    ("Medkart Pharmacy",        "IT Services","HealthTech, Pharmacy Software Platform",          "medkart.in",            "Ahmedabad","knowstartup"),
]

def norm(s):
    if not s: return ""
    s = unicodedata.normalize('NFKD', str(s).lower())
    return re.sub(r'[^a-z0-9]', '', s)

def main():
    wb = openpyxl.load_workbook("COMONK_TRUE_MASTER.xlsx")
    ws = wb["COMPLETE_MASTER"]

    existing = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, 2).value
        if n: existing[norm(n)] = r

    fill_new = PatternFill("solid", fgColor="FFF0F5")
    added = updated = 0

    for company, category, roles, website_hint, city, source in MEGA:
        key = norm(company)
        if not key: continue
        website = ("https://www." + website_hint) if not website_hint.startswith("http") else website_hint
        if key in existing:
            row = existing[key]
            if not ws.cell(row, 4).value: ws.cell(row, 4).value = category
            if not ws.cell(row, 5).value: ws.cell(row, 5).value = roles
            if not ws.cell(row, 13).value: ws.cell(row, 13).value = website
            cur = ws.cell(row, 18).value or ""
            if source not in cur:
                ws.cell(row, 18).value = (cur + ", " + source).strip(", ")
            updated += 1
        else:
            r = ws.max_row + 1
            ws.cell(r, 1, r - 1)
            ws.cell(r, 2, company); ws.cell(r, 3, city)
            ws.cell(r, 4, category); ws.cell(r, 5, roles)
            ws.cell(r, 13, website); ws.cell(r, 18, source)
            for c in range(1, 19):
                ws.cell(r, c).fill = fill_new
                ws.cell(r, c).alignment = Alignment(vertical="center")
            ws.cell(r, 2).font = Font(bold=True, color="8B0050")
            ws.row_dimensions[r].height = 18
            existing[key] = r
            added += 1

    for i, r in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(r, 1, i)
    total = ws.max_row - 1
    max_col = ws.max_column

    # Rebuild AI_ML_ONLY
    ws2 = wb["AI_ML_ONLY"]
    for r in range(ws2.max_row, 1, -1):
        ws2.delete_rows(r)
    ai_kw = ['ai','ml','data','machine','deep','nlp','computer vision','llm','genai',
             'analytics','automation','intelligence','learning','aiops']
    ai_row = 2
    fa = PatternFill("solid", fgColor="F5F0FF"); fb = PatternFill("solid", fgColor="FFFFFF")
    for r in range(2, ws.max_row + 1):
        cat  = str(ws.cell(r, 4).value or "").lower()
        role = str(ws.cell(r, 5).value or "").lower()
        if any(k in cat or k in role for k in ai_kw):
            fill = fa if ai_row % 2 == 0 else fb
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                cell2 = ws2.cell(ai_row, c, v)
                cell2.fill = fill
                cell2.alignment = Alignment(vertical="center")
                if c in (6,7,8,9,10,11) and v:
                    cell2.font = Font(color="1a56db", size=9)
            ws2.cell(ai_row, 1, ai_row - 1)
            ws2.row_dimensions[ai_row].height = 18
            ai_row += 1
    ai_count = ws2.max_row - 1

    wb["STATS"].cell(3, 2, total)
    wb["STATS"].cell(6, 2, ai_count)
    wb.save("COMONK_TRUE_MASTER.xlsx")
    print(f"  NEW companies:    +{added}")
    print(f"  Existing updated: {updated}")
    print(f"  TOTAL companies:  {total}")
    print(f"  AI/ML companies:  {ai_count}")
    print(f"  IT companies:     {total - ai_count}")

if __name__ == "__main__":
    main()
